# -*- coding: utf-8 -*-
"""
시뮬레이터 GUI v3 - 확장된 데이터 수집 + 컬럼 설정 + 불사자 API 연동
- 썸네일 분석 (누끼/텍스트 점수)
- 컬럼 표시/순서 설정
- 더 많은 정보 수집
- 불사자 API로 대표옵션 업데이트
"""
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext, simpledialog
import os
import json
import threading
import subprocess
import math
import webbrowser
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any, Tuple, Optional

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

try:
    from PIL import Image, ImageTk
    import requests
    from io import BytesIO
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False

# 불사자 공통 모듈 (업로더/시뮬레이터 동일하게 사용)
try:
    from bulsaja_common import (
        BulsajaAPIClient, extract_tokens_from_browser,
        filter_bait_options, select_main_option,
        match_thumbnail_to_sku,
        load_bait_keywords, save_bait_keywords,
        load_banned_words, load_excluded_words,
        check_product_safety,
        load_category_risk_settings, save_category_risk_settings,
        DEFAULT_CATEGORY_RISK_SETTINGS,
        MARKET_IDS, DEFAULT_BAIT_KEYWORDS
    )
    BULSAJA_API_AVAILABLE = True
except ImportError:
    BULSAJA_API_AVAILABLE = False

# 엑셀 라이브러리 (상세 형식)
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# ==================== 설정 파일 ====================
CONFIG_FILE = "simulator_gui_v3_config.json"
DEBUG_PORT = 9222


def load_config():
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            pass
    return {}


def save_config(config):
    try:
        with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
        return True
    except:
        return False


# 전체 컬럼 정의 (수집 가능한 모든 데이터)
ALL_COLUMNS = {
    # 기본 정보
    "thumbnail": {"name": "썸네일", "width": 100, "category": "기본", "default": True},
    "option_image": {"name": "옵션이미지", "width": 100, "category": "기본", "default": True},
    "options": {"name": "옵션 선택", "width": 380, "category": "기본", "default": True},
    "product_name": {"name": "상품명", "width": 180, "category": "기본", "default": True},
    "is_safe": {"name": "안전", "width": 50, "category": "기본", "default": True},
    "option_count": {"name": "옵션수", "width": 60, "category": "기본", "default": True},
    "group_name": {"name": "그룹명", "width": 100, "category": "기본", "default": True},

    # 썸네일 분석
    "thumb_score": {"name": "썸네일점수", "width": 80, "category": "썸네일", "default": True},
    "thumb_nukki": {"name": "누끼", "width": 50, "category": "썸네일", "default": False},
    "thumb_text": {"name": "텍스트", "width": 50, "category": "썸네일", "default": False},
    "thumb_action": {"name": "필요작업", "width": 80, "category": "썸네일", "default": True},

    # 가격 정보
    "price_cny": {"name": "위안가", "width": 70, "category": "가격", "default": False},
    "price_krw": {"name": "원화가", "width": 80, "category": "가격", "default": False},
    "sale_price": {"name": "판매가", "width": 80, "category": "가격", "default": True},

    # 옵션 상세
    "total_options": {"name": "전체옵션", "width": 60, "category": "옵션", "default": False},
    "bait_options": {"name": "미끼옵션", "width": 60, "category": "옵션", "default": True},
    "bait_keywords": {"name": "미끼키워드", "width": 120, "category": "옵션", "default": True},
    "option_list": {"name": "옵션명목록", "width": 150, "category": "옵션", "default": True},
    "main_option": {"name": "대표옵션", "width": 100, "category": "옵션", "default": False},

    # 기타
    "product_id": {"name": "불사자ID", "width": 100, "category": "기타", "default": True},
    "unsafe_reason": {"name": "위험키워드", "width": 120, "category": "기타", "default": True},
}

DEFAULT_COLUMN_ORDER = [
    "product_id", "thumbnail", "option_image", "options", "option_list", "product_name",
    "is_safe", "unsafe_reason", "bait_options", "bait_keywords", "sale_price", "option_count", "group_name"
]

# ==================== 수집 설정 (업로더 v1.5와 동일) ====================

# 업로드 조건 (불사자 상태값)
UPLOAD_CONDITIONS = {
    "미업로드(수집완료+수정중+검토완료)": ["0", "1", "2", "수집 완료", "수정중", "검토 완료"],
    "수집완료만": ["0", "수집 완료"],
    "수정중만": ["1", "수정중"],
    "검토완료만": ["2", "검토 완료"],
    "업로드완료(판매중)": ["3", "판매중", "업로드 완료"],
    "전체": None,  # 필터 없음
}

# 상품명 처리 옵션
TITLE_OPTIONS = {
    "원마켓 상품명 그대로 사용": "original",
    "앞4개단어제외 셔플": "shuffle_skip4",
    "앞3개단어제외 셔플": "shuffle_skip3",
    "모든단어 셔플": "shuffle_all",
}

# 옵션 정렬 옵션
OPTION_SORT_OPTIONS = {
    "가격낮은순": "price_asc",
    "주요가격대": "price_main",
    "가격높은순": "price_desc",
}


# ==================== 엑셀 반영 클래스 ====================
class ExcelApplier:
    """엑셀에서 수정한 내용을 불사자에 반영"""

    def __init__(self, api_client, log_callback=None):
        self.api_client = api_client
        self.log = log_callback or print
        self.is_running = False
        self.stats = {
            "total": 0,
            "updated": 0,
            "skipped": 0,
            "failed": 0,
        }

    def read_excel(self, filepath: str) -> List[Dict]:
        """엑셀 파일 읽기 (상세정보 시트 우선)"""
        try:
            wb = load_workbook(filepath, data_only=True)

            # 상세정보 시트 우선, 없으면 첫 번째 시트
            if "상세정보" in wb.sheetnames:
                ws = wb["상세정보"]
            else:
                ws = wb.active

            # 헤더 읽기
            headers = []
            for col in range(1, ws.max_column + 1):
                val = ws.cell(row=1, column=col).value
                headers.append(str(val).strip() if val else f"col_{col}")

            # 데이터 읽기
            data = []
            for row_idx in range(2, ws.max_row + 1):
                row_data = {}
                for col_idx, header in enumerate(headers, 1):
                    val = ws.cell(row=row_idx, column=col_idx).value
                    row_data[header] = val
                # 불사자ID가 있는 행만 추가
                if row_data.get('불사자ID') or row_data.get('id'):
                    data.append(row_data)

            wb.close()
            return data

        except Exception as e:
            self.log(f"❌ 엑셀 읽기 실패: {e}")
            return []

    def parse_selected_option(self, select_value: str, options_text: str) -> Optional[Dict]:
        """
        선택된 옵션 파싱
        select_value: 'A', 'B', 'C' 등
        options_text: 'A. 옵션1(10.5)\nB. 옵션2(15.0)' 형태
        """
        import re
        if not select_value or not options_text:
            return None

        select_value = str(select_value).strip().upper()
        if not select_value:
            return None

        # 옵션 목록 파싱
        lines = options_text.strip().split('\n')
        for idx, line in enumerate(lines):
            line = line.strip()
            if not line:
                continue

            # "A. 옵션명(가격)" 형태 파싱
            if line.startswith(f"{select_value}."):
                option_part = line[2:].strip()
                price_match = re.search(r'\((\d+\.?\d*)\)$', option_part)
                price = float(price_match.group(1)) if price_match else 0
                name = re.sub(r'\(\d+\.?\d*\)$', '', option_part).strip()

                return {
                    'name': name,
                    'price': price,
                    'index': idx,
                    'label': select_value
                }

        return None

    def apply_changes(self, excel_data: List[Dict], options: Dict):
        """엑셀 변경사항을 불사자에 반영"""
        self.is_running = True
        self.stats = {"total": 0, "updated": 0, "skipped": 0, "failed": 0}

        apply_main_option = options.get('apply_main_option', True)
        skip_dangerous = options.get('skip_dangerous', True)

        self.log("")
        self.log("=" * 50)
        self.log("📝 엑셀 반영 시작")
        self.log(f"   총 {len(excel_data)}개 상품")
        self.log(f"   대표옵션 반영: {'O' if apply_main_option else 'X'}")
        self.log(f"   위험상품 스킵: {'O' if skip_dangerous else 'X'}")
        self.log("=" * 50)

        for idx, row in enumerate(excel_data):
            if not self.is_running:
                break

            self.stats['total'] += 1
            product_id = str(row.get('불사자ID') or row.get('id') or '').strip()

            if not product_id:
                self.stats['skipped'] += 1
                continue

            # 안전여부 확인
            safety_value = str(row.get('안전여부', '')).strip().upper()
            is_dangerous = safety_value in ['X', '위험', 'DANGER', 'FALSE', '0']

            if is_dangerous and skip_dangerous:
                self.stats['skipped'] += 1
                continue

            # 대표옵션 변경
            if apply_main_option:
                select_value = row.get('선택', 'A')
                options_text = row.get('최종옵션목록') or row.get('옵션명', '')

                selected = self.parse_selected_option(select_value, options_text)
                if selected and selected['label'] != 'A':
                    # A가 아닌 다른 옵션을 선택한 경우 → 대표옵션 변경
                    try:
                        detail = self.api_client.get_product_detail(product_id)
                        upload_skus = detail.get('uploadSkus', [])

                        if upload_skus and selected['index'] < len(upload_skus):
                            # 모든 옵션 main_product False로
                            for sku in upload_skus:
                                sku['main_product'] = False
                            # 선택된 옵션 main_product True로
                            upload_skus[selected['index']]['main_product'] = True

                            # API 업데이트
                            update_data = {'uploadSkus': upload_skus}
                            success = self.api_client.update_product(product_id, update_data)

                            if success:
                                self.stats['updated'] += 1
                                self.log(f"✅ [{idx+1}] {product_id} → 옵션 {selected['label']} 선택")
                            else:
                                self.stats['failed'] += 1
                                self.log(f"❌ [{idx+1}] {product_id} → 업데이트 실패")
                        else:
                            self.stats['skipped'] += 1
                    except Exception as e:
                        self.stats['failed'] += 1
                        self.log(f"❌ [{idx+1}] {product_id} → 오류: {e}")
                else:
                    self.stats['skipped'] += 1
            else:
                self.stats['skipped'] += 1

            # 10개마다 로그
            if (idx + 1) % 10 == 0:
                self.log(f"   ... {idx+1}/{len(excel_data)} 처리 완료")

        self.log("")
        self.log("=" * 50)
        self.log("📊 반영 완료")
        self.log(f"   총: {self.stats['total']} / 업데이트: {self.stats['updated']}")
        self.log(f"   스킵: {self.stats['skipped']} / 실패: {self.stats['failed']}")
        self.log("=" * 50)


class ColumnSettingsDialog:
    """컬럼 설정 다이얼로그"""

    def __init__(self, parent, current_columns: List[str], column_order: List[str]):
        self.dialog = tk.Toplevel(parent)
        self.dialog.title("컬럼 설정")
        self.dialog.geometry("550x800")
        self.dialog.transient(parent)
        self.dialog.grab_set()

        self.result = None
        self.current_columns = set(current_columns)
        self.column_order = list(column_order)

        self._create_ui()

    def _create_ui(self):
        # 설명
        ttk.Label(self.dialog, text="표시할 컬럼을 선택하고 순서를 조정하세요",
                 font=("맑은 고딕", 10)).pack(pady=10)

        # 메인 프레임
        main_frame = ttk.Frame(self.dialog)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        # 왼쪽: 체크박스 목록
        left_frame = ttk.LabelFrame(main_frame, text="컬럼 선택", padding=5)
        left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))

        # 카테고리별 그룹
        self.checkboxes = {}
        categories = {}
        for col_id, col_info in ALL_COLUMNS.items():
            cat = col_info["category"]
            if cat not in categories:
                categories[cat] = []
            categories[cat].append((col_id, col_info))

        for cat_name, cols in categories.items():
            cat_frame = ttk.LabelFrame(left_frame, text=cat_name, padding=3)
            cat_frame.pack(fill=tk.X, pady=2)

            for col_id, col_info in cols:
                var = tk.BooleanVar(value=col_id in self.current_columns)
                cb = ttk.Checkbutton(cat_frame, text=col_info["name"], variable=var,
                                    command=lambda cid=col_id, v=var: self._on_checkbox_change(cid, v))
                cb.pack(anchor=tk.W)
                self.checkboxes[col_id] = var

        # 오른쪽: 순서 조정
        right_frame = ttk.LabelFrame(main_frame, text="컬럼 순서", padding=5)
        right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

        # 리스트박스
        self.listbox = tk.Listbox(right_frame, height=15, selectmode=tk.SINGLE)
        self.listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        for col_id in self.column_order:
            if col_id in ALL_COLUMNS:
                self.listbox.insert(tk.END, f"{ALL_COLUMNS[col_id]['name']} ({col_id})")

        # 버튼
        btn_frame = ttk.Frame(right_frame)
        btn_frame.pack(side=tk.RIGHT, fill=tk.Y, padx=5)

        ttk.Button(btn_frame, text="▲", width=3, command=self._move_up).pack(pady=2)
        ttk.Button(btn_frame, text="▼", width=3, command=self._move_down).pack(pady=2)

        # 하단 버튼
        bottom_frame = ttk.Frame(self.dialog)
        bottom_frame.pack(fill=tk.X, padx=10, pady=10)

        ttk.Button(bottom_frame, text="기본값", command=self._reset_default).pack(side=tk.LEFT)
        ttk.Button(bottom_frame, text="취소", command=self.dialog.destroy).pack(side=tk.RIGHT, padx=5)
        ttk.Button(bottom_frame, text="적용", command=self._apply).pack(side=tk.RIGHT)

    def _on_checkbox_change(self, col_id, var):
        """체크박스 변경 시 리스트박스 업데이트"""
        if var.get():
            # 체크됨 - 리스트에 추가 (없으면)
            if col_id not in self.column_order:
                self.column_order.append(col_id)
                self.listbox.insert(tk.END, f"{ALL_COLUMNS[col_id]['name']} ({col_id})")
        else:
            # 체크 해제 - 리스트에서 제거
            if col_id in self.column_order:
                idx = self.column_order.index(col_id)
                self.column_order.remove(col_id)
                self.listbox.delete(idx)

    def _move_up(self):
        idx = self.listbox.curselection()
        if idx and idx[0] > 0:
            i = idx[0]
            item = self.listbox.get(i)
            self.listbox.delete(i)
            self.listbox.insert(i - 1, item)
            self.listbox.selection_set(i - 1)
            # 순서 업데이트
            self.column_order[i], self.column_order[i-1] = self.column_order[i-1], self.column_order[i]

    def _move_down(self):
        idx = self.listbox.curselection()
        if idx and idx[0] < self.listbox.size() - 1:
            i = idx[0]
            item = self.listbox.get(i)
            self.listbox.delete(i)
            self.listbox.insert(i + 1, item)
            self.listbox.selection_set(i + 1)
            # 순서 업데이트
            self.column_order[i], self.column_order[i+1] = self.column_order[i+1], self.column_order[i]

    def _reset_default(self):
        # 체크박스 초기화
        for col_id, col_info in ALL_COLUMNS.items():
            self.checkboxes[col_id].set(col_info["default"])

        # 순서 초기화
        self.column_order = list(DEFAULT_COLUMN_ORDER)
        self.listbox.delete(0, tk.END)
        for col_id in self.column_order:
            if col_id in ALL_COLUMNS:
                self.listbox.insert(tk.END, f"{ALL_COLUMNS[col_id]['name']} ({col_id})")

    def _apply(self):
        selected = [col_id for col_id, var in self.checkboxes.items() if var.get()]
        self.result = {
            "columns": selected,
            "order": self.column_order
        }
        self.dialog.destroy()


class SimulatorGUIv3:
    """시뮬레이터 GUI v3 - 탭 구조 (수집|검수|설정)"""

    def __init__(self, root):
        self.root = root
        self.root.title("불사자 시뮬레이터 v3")
        self.root.geometry("1600x1000")
        self.root.minsize(1200, 800)

        self.data = []
        self.selected_options = {}
        self.option_frames = {}
        self.expanded_rows = set()  # 옵션 확장된 행 추적
        self.option_cells = {}  # {row_idx: (cell_frame, item, bg_color)} 옵션 셀 참조

        # === 성능 최적화 ===
        # LRU 캐시 (최대 100개, 오래된 것 자동 삭제)
        self.image_cache = {}
        self.option_image_cache = {}
        self._cache_max_size = 100

        # ThreadPoolExecutor 재사용 (스레드 폭증 방지)
        from concurrent.futures import ThreadPoolExecutor
        self._image_executor = ThreadPoolExecutor(max_workers=8)

        # 키워드 캐시 (파일 I/O 1회만)
        self._bait_keywords_cache = None
        self._banned_words_cache = None

        # 페이지네이션
        self.current_page = 0
        self.page_size = 20  # 한 페이지에 20개

        # 컬럼 설정
        self.visible_columns = [col for col, info in ALL_COLUMNS.items() if info["default"]]
        self.column_order = list(DEFAULT_COLUMN_ORDER)

        # 설정 파일 로드
        self._load_settings()

        self._create_ui()
        # 자동 로드 비활성화 - 사용자가 직접 엑셀 열기 클릭
        # self.root.after(100, self._auto_load_latest)

    def _load_settings(self):
        """설정 파일 로드"""
        settings_file = Path(__file__).parent / "simulator_settings.json"
        if settings_file.exists():
            try:
                with open(settings_file, 'r', encoding='utf-8') as f:
                    settings = json.load(f)
                    self.visible_columns = settings.get("visible_columns", self.visible_columns)
                    self.column_order = settings.get("column_order", self.column_order)
            except:
                pass

    def _save_settings(self):
        """설정 파일 저장"""
        settings_file = Path(__file__).parent / "simulator_settings.json"
        settings = {
            "visible_columns": self.visible_columns,
            "column_order": self.column_order
        }
        with open(settings_file, 'w', encoding='utf-8') as f:
            json.dump(settings, f, ensure_ascii=False, indent=2)

    def _create_ui(self):
        """탭 기반 UI 생성"""
        # 탭 컨테이너
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # 각 탭 프레임 생성
        self.collection_tab = ttk.Frame(self.notebook)
        self.review_tab = ttk.Frame(self.notebook)
        self.settings_tab = ttk.Frame(self.notebook)

        self.notebook.add(self.collection_tab, text="  수집  ")
        self.notebook.add(self.review_tab, text="  검수  ")
        self.notebook.add(self.settings_tab, text="  설정  ")

        # 각 탭 UI 생성
        self._create_collection_tab()
        self._create_review_tab()
        self._create_settings_tab()

        # 기본 탭: 검수
        self.notebook.select(self.review_tab)

    def _create_collection_tab(self):
        """수집 탭 UI - 업로더 v1.5와 동일한 설정"""
        frame = self.collection_tab

        # 스크롤 가능한 프레임
        canvas = tk.Canvas(frame)
        scrollbar = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=canvas.yview)
        scrollable = ttk.Frame(canvas)

        scrollable.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scrollable, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # === 0. API 연결 ===
        conn_frame = ttk.LabelFrame(scrollable, text="🔑 API 연결", padding=5)
        conn_frame.pack(fill=tk.X, padx=10, pady=5)

        conn_row = ttk.Frame(conn_frame)
        conn_row.pack(fill=tk.X, pady=2)

        ttk.Button(conn_row, text="🌐 크롬", command=self._open_debug_chrome, width=8).pack(side=tk.LEFT)
        ttk.Button(conn_row, text="🔑 토큰", command=self._extract_tokens, width=8).pack(side=tk.LEFT, padx=2)
        ttk.Button(conn_row, text="🔗 연결", command=self._connect_api, width=8).pack(side=tk.LEFT, padx=2)

        self.api_status = ttk.Label(conn_row, text="연결 안 됨", foreground="gray")
        self.api_status.pack(side=tk.LEFT, padx=10)

        ttk.Label(conn_row, text="포트:").pack(side=tk.RIGHT)
        self.port_var = tk.StringVar(value="9222")
        ttk.Entry(conn_row, textvariable=self.port_var, width=6).pack(side=tk.RIGHT, padx=2)

        # === 1. 그룹 선택 (시뮬레이터와 동일) ===
        group_frame = ttk.LabelFrame(scrollable, text="📁 마켓그룹 설정", padding=5)
        group_frame.pack(fill=tk.X, padx=10, pady=5)

        row1 = ttk.Frame(group_frame)
        row1.pack(fill=tk.X, pady=2)

        ttk.Label(row1, text="그룹당 최대 상품:").pack(side=tk.LEFT)
        self.max_products_var = tk.StringVar(value="100")
        ttk.Entry(row1, textvariable=self.max_products_var, width=8).pack(side=tk.LEFT, padx=(2, 10))

        ttk.Label(row1, text="작업 그룹:").pack(side=tk.LEFT)
        self.work_groups_var = tk.StringVar(value="1-5")
        ttk.Entry(row1, textvariable=self.work_groups_var, width=15).pack(side=tk.LEFT, padx=2)
        ttk.Label(row1, text="(예: 1-5 또는 1,3,5)", foreground="gray").pack(side=tk.LEFT, padx=5)
        ttk.Button(row1, text="📥 그룹목록", command=self._load_market_groups, width=10).pack(side=tk.RIGHT)

        row1b = ttk.Frame(group_frame)
        row1b.pack(fill=tk.X, pady=2)
        ttk.Label(row1b, text="마켓 그룹 목록 (쉼표 구분, 숫자 맵핑용):").pack(anchor=tk.W)

        self.group_text = scrolledtext.ScrolledText(group_frame, height=2, width=80, font=('Consolas', 9))
        self.group_text.pack(fill=tk.X, expand=True, pady=2)

        ttk.Label(group_frame, text="예: 01_푸로테카,02_스트롬브린 → 작업그룹에서 1, 1-3 등으로 사용",
                  foreground="gray").pack(anchor=tk.W)

        # === 2. 수집 조건 ===
        condition_frame = ttk.LabelFrame(scrollable, text="📋 수집 조건", padding=5)
        condition_frame.pack(fill=tk.X, padx=10, pady=5)

        row2 = ttk.Frame(condition_frame)
        row2.pack(fill=tk.X, pady=2)

        ttk.Label(row2, text="수집조건:").pack(side=tk.LEFT)
        self.upload_condition_var = tk.StringVar(value="미업로드(수집완료+수정중+검토완료)")
        ttk.Combobox(row2, textvariable=self.upload_condition_var, width=35,
                     values=list(UPLOAD_CONDITIONS.keys())).pack(side=tk.LEFT, padx=5)

        ttk.Label(row2, text="수집수:").pack(side=tk.LEFT, padx=(10, 0))
        self.collect_count_var = tk.StringVar(value="9999")
        ttk.Entry(row2, textvariable=self.collect_count_var, width=6).pack(side=tk.LEFT, padx=2)

        # === 3. 옵션 설정 ===
        option_frame = ttk.LabelFrame(scrollable, text="⚙️ 옵션 설정", padding=5)
        option_frame.pack(fill=tk.X, padx=10, pady=5)

        row3 = ttk.Frame(option_frame)
        row3.pack(fill=tk.X, pady=2)

        ttk.Label(row3, text="옵션수:").pack(side=tk.LEFT)
        self.option_count_var = tk.StringVar(value="10")
        ttk.Entry(row3, textvariable=self.option_count_var, width=5).pack(side=tk.LEFT, padx=(2, 10))

        ttk.Label(row3, text="옵션정렬:").pack(side=tk.LEFT)
        self.option_sort_var = tk.StringVar(value="가격낮은순")
        ttk.Combobox(row3, textvariable=self.option_sort_var, width=10,
                     values=list(OPTION_SORT_OPTIONS.keys())).pack(side=tk.LEFT, padx=(2, 10))

        ttk.Label(row3, text="상품명:").pack(side=tk.LEFT)
        self.title_option_var = tk.StringVar(value="앞3개단어제외 셔플")
        ttk.Combobox(row3, textvariable=self.title_option_var, width=18,
                     values=list(TITLE_OPTIONS.keys())).pack(side=tk.LEFT, padx=2)

        row4 = ttk.Frame(option_frame)
        row4.pack(fill=tk.X, pady=2)

        ttk.Label(row4, text="최저가격:").pack(side=tk.LEFT)
        self.min_price_var = tk.StringVar(value="30000")
        ttk.Entry(row4, textvariable=self.min_price_var, width=8).pack(side=tk.LEFT, padx=(2, 10))

        ttk.Label(row4, text="최대가격:").pack(side=tk.LEFT)
        self.max_price_var = tk.StringVar(value="100000000")
        ttk.Entry(row4, textvariable=self.max_price_var, width=10).pack(side=tk.LEFT, padx=2)

        # === 4. 미끼 키워드 설정 ===
        keyword_frame = ttk.LabelFrame(scrollable, text="🚫 미끼 키워드 (학습/수정 가능)", padding=5)
        keyword_frame.pack(fill=tk.X, padx=10, pady=5)

        keyword_row1 = ttk.Frame(keyword_frame)
        keyword_row1.pack(fill=tk.X, pady=2)

        ttk.Label(keyword_row1, text="제외 키워드 (쉼표 구분):").pack(side=tk.LEFT)
        ttk.Button(keyword_row1, text="기본값", command=self._reset_keywords, width=6).pack(side=tk.RIGHT)
        ttk.Button(keyword_row1, text="💾 저장", command=self._save_keywords, width=6).pack(side=tk.RIGHT, padx=2)

        self.keyword_text = scrolledtext.ScrolledText(keyword_frame, height=3, width=80, font=('Consolas', 9))
        self.keyword_text.pack(fill=tk.X, expand=True)
        # 기본 미끼 키워드 로드
        bait_keywords = load_bait_keywords() if BULSAJA_API_AVAILABLE else []
        self.keyword_text.insert("1.0", ','.join(bait_keywords))

        # === 5. 카테고리별 검수 설정 ===
        category_frame = ttk.LabelFrame(scrollable, text="🛡️ 카테고리별 검수 수준", padding=5)
        category_frame.pack(fill=tk.X, padx=10, pady=5)

        cat_row1 = ttk.Frame(category_frame)
        cat_row1.pack(fill=tk.X, pady=2)

        ttk.Label(cat_row1, text="검수 수준:").pack(side=tk.LEFT)

        self.check_level_var = tk.StringVar(value="normal")
        ttk.Radiobutton(cat_row1, text="보통 (프로그램)", variable=self.check_level_var, value="normal").pack(side=tk.LEFT, padx=5)
        ttk.Radiobutton(cat_row1, text="엄격 (AI확인)", variable=self.check_level_var, value="strict").pack(side=tk.LEFT, padx=5)
        ttk.Radiobutton(cat_row1, text="제외", variable=self.check_level_var, value="skip").pack(side=tk.LEFT, padx=5)

        ttk.Button(cat_row1, text="⚙️ 카테고리 설정", command=self._open_category_settings, width=14).pack(side=tk.RIGHT)

        cat_row2 = ttk.Frame(category_frame)
        cat_row2.pack(fill=tk.X, pady=2)

        ttk.Label(cat_row2, text="위험 카테고리 (자동 엄격):").pack(side=tk.LEFT)
        self.risk_categories_var = tk.StringVar(value="패션의류,패션잡화,유아동,의료기기,화장품,시계,가방")
        ttk.Entry(cat_row2, textvariable=self.risk_categories_var, width=60).pack(side=tk.LEFT, padx=2, fill=tk.X, expand=True)

        ttk.Label(category_frame, text="※ 엄격: 위험 키워드 발견 시 Gemini AI로 재확인 | 보통: 안전 컨텍스트로 자동 판단",
                  foreground="gray").pack(anchor=tk.W)

        # === 6. 저장 설정 ===
        save_frame = ttk.LabelFrame(scrollable, text="💾 저장 설정", padding=5)
        save_frame.pack(fill=tk.X, padx=10, pady=5)

        save_row = ttk.Frame(save_frame)
        save_row.pack(fill=tk.X, pady=2)

        ttk.Label(save_row, text="저장 경로:").pack(side=tk.LEFT)
        self.save_path_var = tk.StringVar(value=f"simulation_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")
        ttk.Entry(save_row, textvariable=self.save_path_var, width=50).pack(side=tk.LEFT, padx=2, fill=tk.X, expand=True)
        ttk.Button(save_row, text="찾아보기", command=self._browse_save_path, width=8).pack(side=tk.RIGHT)

        # === 6-1. 직접 API 업로드용 데이터 ===
        api_data_frame = ttk.LabelFrame(scrollable, text="🔗 직접 API 업로드용 (스마트스토어 등)", padding=5)
        api_data_frame.pack(fill=tk.X, padx=10, pady=5)

        api_row = ttk.Frame(api_data_frame)
        api_row.pack(fill=tk.X, pady=2)

        self.fetch_detail_contents_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(api_row, text="상세이미지 (uploadDetailContents)",
                       variable=self.fetch_detail_contents_var).pack(side=tk.LEFT, padx=(0, 15))

        self.fetch_category_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(api_row, text="카테고리 (uploadCategory)",
                       variable=self.fetch_category_var).pack(side=tk.LEFT)

        ttk.Label(api_data_frame, text="※ 체크 시 수집 속도가 느려집니다 (상품당 API 추가 호출)",
                  font=('맑은 고딕', 8), foreground="gray").pack(anchor=tk.W)

        # === 7. 마진 설정 ===
        margin_frame = ttk.LabelFrame(scrollable, text="💰 마진 설정", padding=5)
        margin_frame.pack(fill=tk.X, padx=10, pady=5)

        row5 = ttk.Frame(margin_frame)
        row5.pack(fill=tk.X, pady=2)

        ttk.Label(row5, text="환율(위안):").pack(side=tk.LEFT)
        self.exchange_rate_var = tk.StringVar(value="215")
        ttk.Entry(row5, textvariable=self.exchange_rate_var, width=6).pack(side=tk.LEFT, padx=(2, 10))

        ttk.Label(row5, text="카드수수료(%):").pack(side=tk.LEFT)
        self.card_fee_var = tk.StringVar(value="3.3")
        ttk.Entry(row5, textvariable=self.card_fee_var, width=5).pack(side=tk.LEFT, padx=(2, 10))

        ttk.Label(row5, text="마진(min,max):").pack(side=tk.LEFT)
        self.margin_rate_var = tk.StringVar(value="25,30")
        ttk.Entry(row5, textvariable=self.margin_rate_var, width=8).pack(side=tk.LEFT, padx=(2, 10))

        ttk.Label(row5, text="더하기마진:").pack(side=tk.LEFT)
        self.margin_fixed_var = tk.StringVar(value="15000")
        ttk.Entry(row5, textvariable=self.margin_fixed_var, width=7).pack(side=tk.LEFT, padx=2)

        row6 = ttk.Frame(margin_frame)
        row6.pack(fill=tk.X, pady=2)

        ttk.Label(row6, text="할인율(min,max):").pack(side=tk.LEFT)
        self.discount_rate_var = tk.StringVar(value="20,30")
        ttk.Entry(row6, textvariable=self.discount_rate_var, width=8).pack(side=tk.LEFT, padx=(2, 10))

        ttk.Label(row6, text="가격단위:").pack(side=tk.LEFT)
        self.round_unit_var = tk.StringVar(value="100")
        ttk.Entry(row6, textvariable=self.round_unit_var, width=5).pack(side=tk.LEFT, padx=2)

        # === 6. 버튼 ===
        btn_frame = ttk.Frame(scrollable)
        btn_frame.pack(fill=tk.X, padx=10, pady=10)

        ttk.Button(btn_frame, text="수집 시작", command=self._start_collection).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="수집 중지", command=self._stop_collection).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="엑셀로 저장", command=self._save_collection_to_excel).pack(side=tk.LEFT, padx=20)
        ttk.Button(btn_frame, text="▶ 검수로 이동", command=self._transfer_to_review).pack(side=tk.RIGHT, padx=5)

        # === 7. 진행 상황 ===
        progress_frame = ttk.LabelFrame(scrollable, text="📊 진행 상황", padding=5)
        progress_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        self.collection_status = tk.StringVar(value="대기 중...")
        ttk.Label(progress_frame, textvariable=self.collection_status, font=("맑은 고딕", 10)).pack(anchor=tk.W)

        self.collection_progress = ttk.Progressbar(progress_frame, length=400, mode='determinate')
        self.collection_progress.pack(fill=tk.X, pady=5)

        self.collection_log = tk.Text(progress_frame, height=12, width=80, font=("Consolas", 9))
        self.collection_log.pack(fill=tk.BOTH, expand=True)

    def _create_review_tab(self):
        """검수 탭 UI - 기존 메인 화면"""
        frame = self.review_tab

        # 상단 툴바
        toolbar = ttk.Frame(frame, padding=5)
        toolbar.pack(fill=tk.X)

        ttk.Button(toolbar, text="엑셀 열기", command=self._load_excel).pack(side=tk.LEFT, padx=5)
        ttk.Button(toolbar, text="컬럼 설정", command=self._open_column_settings).pack(side=tk.LEFT, padx=5)
        ttk.Button(toolbar, text="썸네일 로드 (현재페이지)", command=self._load_all_thumbnails).pack(side=tk.LEFT, padx=5)
        ttk.Button(toolbar, text="썸네일 분석 (누끼찾기)", command=self._analyze_thumbnails).pack(side=tk.LEFT, padx=5)
        ttk.Button(toolbar, text="지재권 분석", command=self._analyze_ip_words).pack(side=tk.LEFT, padx=5)

        ttk.Separator(toolbar, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=10)

        ttk.Label(toolbar, text="파일:").pack(side=tk.LEFT)
        self.file_label = ttk.Label(toolbar, text="(없음)", foreground="gray")
        self.file_label.pack(side=tk.LEFT, padx=5)

        ttk.Button(toolbar, text="불사자 업데이트", command=self._update_bulsaja).pack(side=tk.RIGHT, padx=5)
        ttk.Button(toolbar, text="저장", command=self._save_changes).pack(side=tk.RIGHT, padx=5)

        # 썸네일 변경 옵션
        self.update_thumbnail_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(toolbar, text="썸네일도 변경", variable=self.update_thumbnail_var).pack(side=tk.RIGHT, padx=5)

        self.count_label = ttk.Label(toolbar, text="상품: 0개")
        self.count_label.pack(side=tk.RIGHT, padx=20)

        # 페이지네이션 바
        page_bar = ttk.Frame(frame, padding=5)
        page_bar.pack(fill=tk.X)

        ttk.Button(page_bar, text="◀◀ 처음", width=8, command=self._go_first_page).pack(side=tk.LEFT, padx=2)
        ttk.Button(page_bar, text="◀ 이전", width=8, command=self._go_prev_page).pack(side=tk.LEFT, padx=2)

        self.page_label = ttk.Label(page_bar, text="1 / 1 페이지", font=("맑은 고딕", 10, "bold"))
        self.page_label.pack(side=tk.LEFT, padx=20)

        ttk.Button(page_bar, text="다음 ▶", width=8, command=self._go_next_page).pack(side=tk.LEFT, padx=2)
        ttk.Button(page_bar, text="끝 ▶▶", width=8, command=self._go_last_page).pack(side=tk.LEFT, padx=2)

        ttk.Separator(page_bar, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=15)

        ttk.Label(page_bar, text="페이지당:").pack(side=tk.LEFT)
        self.page_size_var = tk.StringVar(value="20")
        page_size_combo = ttk.Combobox(page_bar, textvariable=self.page_size_var, values=["10", "20", "30", "50"], width=5, state="readonly")
        page_size_combo.pack(side=tk.LEFT, padx=5)
        page_size_combo.bind("<<ComboboxSelected>>", self._on_page_size_change)

        self.page_info_label = ttk.Label(page_bar, text="(0 ~ 0 / 총 0개)", foreground="gray")
        self.page_info_label.pack(side=tk.LEFT, padx=10)

        # 메인 영역 (스크롤)
        main_frame = ttk.Frame(frame)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        self.canvas = tk.Canvas(main_frame, bg="white")
        scrollbar_y = ttk.Scrollbar(main_frame, orient=tk.VERTICAL, command=self.canvas.yview)
        scrollbar_x = ttk.Scrollbar(main_frame, orient=tk.HORIZONTAL, command=self.canvas.xview)

        self.scrollable_frame = ttk.Frame(self.canvas)
        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        )

        self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)

        scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)
        scrollbar_x.pack(side=tk.BOTTOM, fill=tk.X)
        self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # 마우스휠 스크롤 핸들러
        def _on_mousewheel(event):
            try:
                self.canvas.yview_scroll(int(-1*(event.delta/120)), "units")
            except:
                pass

        # 마우스휠 바인딩 저장 (자식 위젯에도 적용하기 위해)
        self._mousewheel_handler = _on_mousewheel
        self.canvas.bind("<MouseWheel>", _on_mousewheel)
        self.scrollable_frame.bind("<MouseWheel>", _on_mousewheel)

        # 캔버스 진입/이탈 시 전역 마우스휠 바인딩
        def _bind_mousewheel(event):
            self.root.bind_all("<MouseWheel>", _on_mousewheel)
        def _unbind_mousewheel(event):
            self.root.unbind_all("<MouseWheel>")

        self.canvas.bind("<Enter>", _bind_mousewheel)
        self.canvas.bind("<Leave>", _unbind_mousewheel)

    # ========== 페이지네이션 함수 ==========
    def _get_total_pages(self):
        """총 페이지 수"""
        if not self.data:
            return 1
        return max(1, (len(self.data) + self.page_size - 1) // self.page_size)

    def _go_first_page(self):
        """첫 페이지로"""
        if self.current_page != 0:
            self.current_page = 0
            self._render_data()

    def _go_prev_page(self):
        """이전 페이지"""
        if self.current_page > 0:
            self.current_page -= 1
            self._render_data()

    def _go_next_page(self):
        """다음 페이지"""
        if self.current_page < self._get_total_pages() - 1:
            self.current_page += 1
            self._render_data()

    def _go_last_page(self):
        """마지막 페이지로"""
        last_page = self._get_total_pages() - 1
        if self.current_page != last_page:
            self.current_page = last_page
            self._render_data()

    def _on_page_size_change(self, event=None):
        """페이지 크기 변경"""
        try:
            self.page_size = int(self.page_size_var.get())
        except:
            self.page_size = 20
        self.current_page = 0  # 첫 페이지로
        self._render_data()

    def _update_page_info(self):
        """페이지 정보 업데이트"""
        total = len(self.data)
        total_pages = self._get_total_pages()

        start_idx = self.current_page * self.page_size + 1
        end_idx = min((self.current_page + 1) * self.page_size, total)

        if total == 0:
            start_idx = 0
            end_idx = 0

        self.page_label.config(text=f"{self.current_page + 1} / {total_pages} 페이지")
        self.page_info_label.config(text=f"({start_idx} ~ {end_idx} / 총 {total}개)")

    def _create_settings_tab(self):
        """설정 탭 UI"""
        frame = self.settings_tab

        # 컬럼 설정
        col_frame = ttk.LabelFrame(frame, text="컬럼 설정", padding=10)
        col_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # 왼쪽: 체크박스 목록
        left_frame = ttk.Frame(col_frame)
        left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10))

        # 카테고리별 그룹
        self.settings_checkboxes = {}
        categories = {}
        for col_id, col_info in ALL_COLUMNS.items():
            cat = col_info["category"]
            if cat not in categories:
                categories[cat] = []
            categories[cat].append((col_id, col_info))

        row = 0
        for cat_name, cols in categories.items():
            cat_label = ttk.Label(left_frame, text=f"[{cat_name}]", font=("맑은 고딕", 9, "bold"))
            cat_label.grid(row=row, column=0, sticky=tk.W, pady=(10, 2))
            row += 1

            for col_id, col_info in cols:
                var = tk.BooleanVar(value=col_id in self.visible_columns)
                cb = ttk.Checkbutton(left_frame, text=col_info["name"], variable=var,
                                    command=self._on_settings_column_change)
                cb.grid(row=row, column=0, sticky=tk.W, padx=20)
                self.settings_checkboxes[col_id] = var
                row += 1

        # 오른쪽: 순서 조정
        right_frame = ttk.Frame(col_frame)
        right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

        ttk.Label(right_frame, text="컬럼 순서 (드래그 또는 버튼)", font=("맑은 고딕", 9, "bold")).pack(anchor=tk.W)

        order_frame = ttk.Frame(right_frame)
        order_frame.pack(fill=tk.BOTH, expand=True, pady=5)

        self.order_listbox = tk.Listbox(order_frame, height=15, selectmode=tk.SINGLE, font=("맑은 고딕", 9))
        self.order_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        for col_id in self.column_order:
            if col_id in ALL_COLUMNS:
                self.order_listbox.insert(tk.END, f"{ALL_COLUMNS[col_id]['name']} ({col_id})")

        btn_frame = ttk.Frame(order_frame)
        btn_frame.pack(side=tk.RIGHT, fill=tk.Y, padx=5)

        ttk.Button(btn_frame, text="▲ 위로", width=8, command=self._move_column_up).pack(pady=2)
        ttk.Button(btn_frame, text="▼ 아래로", width=8, command=self._move_column_down).pack(pady=2)
        ttk.Button(btn_frame, text="기본값", width=8, command=self._reset_column_settings).pack(pady=10)

        # 하단 버튼
        bottom_frame = ttk.Frame(frame)
        bottom_frame.pack(fill=tk.X, padx=10, pady=10)

        ttk.Button(bottom_frame, text="설정 저장", command=self._apply_and_save_settings).pack(side=tk.RIGHT, padx=5)

    # ===== API 연결 함수들 =====
    def _open_debug_chrome(self):
        """크롬을 디버그 모드로 열기"""
        import subprocess
        port = self.port_var.get()

        # Windows 크롬 경로
        chrome_paths = [
            r"C:\Program Files\Google\Chrome\Application\chrome.exe",
            r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
        ]

        chrome_path = None
        for path in chrome_paths:
            if os.path.exists(path):
                chrome_path = path
                break

        if not chrome_path:
            messagebox.showerror("오류", "크롬을 찾을 수 없습니다")
            return

        try:
            cmd = f'"{chrome_path}" --remote-debugging-port={port} --user-data-dir="{os.path.expanduser("~")}\\chrome_debug"'
            subprocess.Popen(cmd, shell=True)
            self._log_collection(f"🌐 크롬 디버그 모드 실행 (포트: {port})")
            self._log_collection("   → 불사자 사이트에 로그인하세요")
        except Exception as e:
            messagebox.showerror("오류", f"크롬 실행 실패: {e}")

    def _extract_tokens(self):
        """크롬에서 토큰 추출"""
        if not BULSAJA_API_AVAILABLE:
            messagebox.showerror("오류", "bulsaja_common 모듈이 필요합니다")
            return

        self._log_collection("🔑 토큰 추출 중...")

        def extract():
            success, access_token, refresh_token, error = extract_tokens_from_browser()
            if success:
                self.access_token = access_token
                self.refresh_token = refresh_token
                self._log_collection("✅ 토큰 추출 성공")
                self.root.after(0, lambda: self.api_status.config(text="토큰 추출됨", foreground="orange"))
            else:
                self._log_collection(f"❌ 토큰 추출 실패: {error}")
                self.root.after(0, lambda: self.api_status.config(text="토큰 실패", foreground="red"))

        threading.Thread(target=extract, daemon=True).start()

    def _connect_api(self):
        """API 연결"""
        if not hasattr(self, 'access_token') or not self.access_token:
            messagebox.showwarning("경고", "먼저 토큰을 추출하세요")
            return

        self._log_collection("🔗 API 연결 중...")

        try:
            self.api_client = BulsajaAPIClient(self.access_token, self.refresh_token)
            # 연결 테스트
            success, msg, total = self.api_client.test_connection()
            if success:
                self.api_status.config(text=f"연결됨 (총 {total}개)", foreground="green")
                self._log_collection(f"✅ API 연결 성공 - {msg}")
            else:
                self.api_status.config(text="연결 실패", foreground="red")
                self._log_collection(f"❌ API 연결 실패: {msg}")
        except Exception as e:
            self.api_status.config(text="연결 실패", foreground="red")
            self._log_collection(f"❌ API 연결 실패: {e}")

    # ===== 수집 탭 함수들 =====
    def _load_market_groups(self):
        """마켓 그룹 목록 조회 (시뮬레이터와 동일)"""
        if not hasattr(self, 'api_client') or not self.api_client:
            messagebox.showwarning("경고", "먼저 API에 연결하세요")
            return

        self._log_collection("📥 마켓 그룹 목록 조회 중...")

        try:
            groups = self.api_client.get_market_groups()
            if groups:
                self.group_text.delete("1.0", tk.END)
                self.group_text.insert("1.0", ','.join(groups))
                self._log_collection(f"✅ {len(groups)}개 그룹 로드됨")
            else:
                self._log_collection("⚠️ 그룹 없음 또는 조회 실패")
        except Exception as e:
            self._log_collection(f"❌ 그룹 로드 실패: {e}")

    def _reset_keywords(self):
        """미끼 키워드 기본값으로 초기화"""
        self.keyword_text.delete("1.0", tk.END)
        default_keywords = DEFAULT_BAIT_KEYWORDS if BULSAJA_API_AVAILABLE else []
        self.keyword_text.insert("1.0", ','.join(default_keywords))
        self._log_collection("🔄 미끼 키워드 기본값으로 초기화")

    def _save_keywords(self):
        """미끼 키워드 저장"""
        text = self.keyword_text.get("1.0", tk.END).strip()
        keywords = [k.strip() for k in text.split(',') if k.strip()]
        if BULSAJA_API_AVAILABLE and save_bait_keywords(keywords):
            self._log_collection(f"✅ 미끼 키워드 저장됨 ({len(keywords)}개)")
        else:
            self._log_collection("❌ 미끼 키워드 저장 실패")

    def _open_category_settings(self):
        """카테고리별 검수 설정 다이얼로그"""
        dialog = tk.Toplevel(self.root)
        dialog.title("카테고리별 검수 설정")
        dialog.geometry("600x500")
        dialog.transient(self.root)
        dialog.grab_set()

        main_frame = ttk.Frame(dialog, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        ttk.Label(main_frame, text="카테고리별 검수 수준을 설정합니다. (strict=AI확인, normal=프로그램, skip=제외)",
                  foreground="gray").pack(anchor=tk.W)

        # 현재 설정 로드
        current_settings = load_category_risk_settings() if BULSAJA_API_AVAILABLE else {}

        # 스크롤 가능한 리스트
        list_frame = ttk.Frame(main_frame)
        list_frame.pack(fill=tk.BOTH, expand=True, pady=10)

        canvas = tk.Canvas(list_frame)
        scrollbar_cat = ttk.Scrollbar(list_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar_cat.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar_cat.pack(side="right", fill="y")

        # 카테고리 목록
        categories = [
            "패션의류", "패션잡화", "화장품/미용", "디지털/가전", "가구/인테리어",
            "출산/육아", "식품", "스포츠/레저", "생활/건강", "여가/생활편의",
            "면세점", "도서/음반/DVD"
        ]

        self.category_vars = {}
        for cat in categories:
            row = ttk.Frame(scrollable_frame)
            row.pack(fill=tk.X, pady=2)

            ttk.Label(row, text=cat, width=20).pack(side=tk.LEFT)

            var = tk.StringVar(value=current_settings.get(cat, "normal"))
            self.category_vars[cat] = var

            ttk.Radiobutton(row, text="보통", variable=var, value="normal").pack(side=tk.LEFT, padx=5)
            ttk.Radiobutton(row, text="엄격", variable=var, value="strict").pack(side=tk.LEFT, padx=5)
            ttk.Radiobutton(row, text="제외", variable=var, value="skip").pack(side=tk.LEFT, padx=5)

        # 버튼 프레임
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=tk.X, pady=10)

        def save_cat_settings():
            settings = {cat: var.get() for cat, var in self.category_vars.items()}
            if BULSAJA_API_AVAILABLE:
                save_category_risk_settings(settings)
            self._log_collection("✅ 카테고리 검수 설정 저장됨")
            dialog.destroy()

        ttk.Button(btn_frame, text="저장", command=save_cat_settings).pack(side=tk.RIGHT, padx=5)
        ttk.Button(btn_frame, text="취소", command=dialog.destroy).pack(side=tk.RIGHT, padx=5)

    def _browse_save_path(self):
        """저장 경로 선택"""
        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=self.save_path_var.get()
        )
        if filepath:
            self.save_path_var.set(filepath)

    def _parse_group_mapping(self) -> Dict[str, str]:
        """그룹 매핑 텍스트 파싱 (시뮬레이터와 동일)"""
        import re
        mapping = {}
        text = self.group_text.get("1.0", tk.END).strip()
        if not text:
            return mapping

        groups = [g.strip() for g in text.split(',') if g.strip()]
        prefix_pattern = re.compile(r'^(\d+)[_\-]')

        has_prefix_pattern = any(prefix_pattern.match(g) for g in groups)

        if has_prefix_pattern:
            for group_name in groups:
                match = prefix_pattern.match(group_name)
                if match:
                    num_str = match.group(1)
                    mapping[num_str] = group_name
                    mapping[str(int(num_str))] = group_name
                    mapping[f"{int(num_str):02d}"] = group_name
        else:
            for idx, group_name in enumerate(groups, 1):
                mapping[str(idx)] = group_name
                mapping[f"{idx:02d}"] = group_name

        return mapping

    def _parse_work_range(self, range_str: str) -> List[str]:
        """작업 범위 파싱 (1-5 또는 1,3,5)"""
        result = []
        range_str = range_str.strip()
        if '-' in range_str and ',' not in range_str:
            parts = range_str.split('-')
            if len(parts) == 2:
                try:
                    start = int(parts[0])
                    end = int(parts[1])
                    for i in range(start, end + 1):
                        result.append(str(i))
                except ValueError:
                    pass
        else:
            for item in range_str.split(','):
                item = item.strip()
                if item:
                    result.append(item)
        return result

    def _get_work_group_names(self) -> List[str]:
        """작업 범위에서 실제 그룹명 목록 가져오기"""
        mapping = self._parse_group_mapping()
        range_nums = self._parse_work_range(self.work_groups_var.get())
        group_names = []
        for num in range_nums:
            if num in mapping:
                group_names.append(mapping[num])
            else:
                self._log_collection(f"⚠️ 그룹 번호 {num}에 해당하는 그룹명 없음")
        return group_names

    def _start_collection(self):
        """수집 시작 (시뮬레이터와 동일 - 여러 그룹 지원)"""
        if not hasattr(self, 'api_client') or not self.api_client:
            messagebox.showwarning("경고", "먼저 API에 연결하세요")
            return

        group_names = self._get_work_group_names()
        if not group_names:
            messagebox.showwarning("경고", "작업할 그룹이 없습니다.\n작업범위와 그룹목록을 확인하세요.")
            return

        max_products = int(self.max_products_var.get())

        self._log_collection("")
        self._log_collection("=" * 50)
        self._log_collection("📦 수집 시작")
        self._log_collection(f"   그룹: {', '.join(group_names)}")
        self._log_collection(f"   그룹당 최대 상품: {max_products}개")
        self._log_collection("=" * 50)

        self.collection_status.set(f"수집 중: {len(group_names)}개 그룹")
        self.is_collecting = True

        def collect():
            # 직접 API 업로드용 옵션 (메인 스레드에서 미리 가져옴)
            fetch_detail_contents = self.fetch_detail_contents_var.get()
            fetch_category = self.fetch_category_var.get()

            if fetch_detail_contents or fetch_category:
                self._log_collection(f"🔗 직접 API 옵션: 상세이미지={fetch_detail_contents}, 카테고리={fetch_category}")

            try:
                collected_data = []
                total_groups = len(group_names)

                for group_idx, group_name in enumerate(group_names):
                    if not self.is_collecting:
                        break

                    self._log_collection(f"\n[{group_idx+1}/{total_groups}] '{group_name}' 그룹 처리 중...")

                    # 상품 목록 조회
                    products, total_count = self.api_client.get_products_by_group(group_name, limit=max_products)
                    self._log_collection(f"   📦 {len(products)}개 상품 발견 (전체 {total_count}개)")

                    # 상세 정보 수집
                    for i, prod in enumerate(products[:max_products]):
                        if not self.is_collecting:
                            break

                        progress = ((group_idx * max_products) + i + 1) / (total_groups * max_products) * 100
                        self.root.after(0, lambda v=min(progress, 100): self.collection_progress.config(value=v))
                        self.root.after(0, lambda s=f"수집 중: {group_name} ({i+1}/{len(products)})": self.collection_status.set(s))

                        prod_id = prod.get("ID", "") or prod.get("id", "")
                        try:
                            detail = self.api_client.get_product_detail(prod_id)
                            detail['_group_name'] = group_name  # 그룹명 추가

                            # 직접 API 업로드용 추가 데이터 가져오기
                            if fetch_detail_contents or fetch_category:
                                try:
                                    upload_fields = self.api_client.get_upload_fields(prod_id)
                                    if upload_fields:
                                        if fetch_detail_contents:
                                            detail['uploadDetailContents'] = upload_fields.get('uploadDetailContents', {})
                                        if fetch_category:
                                            detail['uploadCategory'] = upload_fields.get('uploadCategory', {})
                                except Exception as uf_e:
                                    self._log_collection(f"      ⚠️ uploadFields 실패: {uf_e}")

                            collected_data.append(detail)
                            prod_name = prod.get('uploadCommonProductName', '') or prod.get('name', '')
                            extra_info = " [+API]" if (fetch_detail_contents or fetch_category) else ""
                            self._log_collection(f"   [{i+1}/{len(products)}] {prod_name[:25]}...{extra_info}")
                        except Exception as e:
                            self._log_collection(f"   ❌ {prod_id}: {e}")

                self.collected_data = collected_data
                self._log_collection(f"\n✅ 수집 완료: 총 {len(collected_data)}개")
                self.root.after(0, lambda: self.collection_status.set(f"완료: {len(collected_data)}개"))
                self.root.after(0, lambda: self.collection_progress.config(value=100))

            except Exception as e:
                self._log_collection(f"❌ 수집 실패: {e}")
                self.root.after(0, lambda: self.collection_status.set("수집 실패"))

        threading.Thread(target=collect, daemon=True).start()

    def _stop_collection(self):
        """수집 중지"""
        self.is_collecting = False
        self._log_collection("⏹️ 수집 중지 요청")
        self.collection_status.set("중지됨")

    def _save_collection_to_excel(self):
        """수집 데이터를 엑셀로 저장 (시뮬레이터 형식)"""
        if not hasattr(self, 'collected_data') or not self.collected_data:
            messagebox.showwarning("경고", "먼저 수집을 실행하세요")
            return

        if not OPENPYXL_AVAILABLE:
            messagebox.showerror("오류", "openpyxl이 필요합니다: pip install openpyxl")
            return

        # 저장 경로 설정에서 가져오기
        filepath = self.save_path_var.get()
        if not filepath:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M')
            filepath = f"simulation_{timestamp}.xlsx"

        # 파일 선택 다이얼로그 (경로 확인/변경 가능)
        filepath = filedialog.asksaveasfilename(
            title="시뮬레이션 엑셀 저장",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=filepath
        )

        if not filepath:
            return

        self._log_collection(f"💾 엑셀 저장 시작: {filepath}")
        self._log_collection(f"📊 {len(self.collected_data)}개 상품 분석 및 저장 중...")

        def save_task():
            try:
                # 키워드 로드
                banned_words, _ = load_banned_words()
                excluded_words = load_excluded_words()
                bait_keywords = load_bait_keywords()

                # 상품 분석 및 결과 수집
                results = []
                stats = {"total": 0, "safe": 0, "unsafe": 0, "bait_found": 0}

                # 검수 설정 가져오기
                check_level = self.check_level_var.get()
                risk_categories = [c.strip() for c in self.risk_categories_var.get().split(',') if c.strip()]

                for idx, product in enumerate(self.collected_data):
                    stats["total"] += 1

                    # 카테고리 확인하여 검수 레벨 결정
                    product_category = product.get('categoryPath', '') or product.get('category', '') or ''
                    product_check_level = check_level

                    # 위험 카테고리에 해당하면 엄격 검수
                    for risk_cat in risk_categories:
                        if risk_cat and risk_cat.lower() in product_category.lower():
                            product_check_level = 'strict'
                            break

                    result = self._analyze_single_product(product, bait_keywords, excluded_words, product_check_level)
                    result['group_name'] = product.get('_group_name', '')  # 수집 시 저장한 그룹명 사용
                    result['category'] = product_category[:30]  # 카테고리 기록
                    results.append(result)

                    if result['is_safe']:
                        stats["safe"] += 1
                    else:
                        stats["unsafe"] += 1
                    if result['bait_options'] > 0:
                        stats["bait_found"] += 1

                    if (idx + 1) % 10 == 0:
                        self._log_collection(f"  분석 중... {idx+1}/{len(self.collected_data)}")

                # 엑셀 저장
                self._save_results_to_excel(filepath, results, stats)

                self._log_collection(f"✅ 저장 완료: {filepath}")
                self._log_collection(f"   총 {stats['total']}개 / 안전 {stats['safe']}개 / 위험 {stats['unsafe']}개")

                # 저장 후 자동으로 검수탭 전환 + 로드
                def auto_load():
                    messagebox.showinfo("완료", f"엑셀 저장 완료!\n\n{filepath}\n\n검수탭으로 이동합니다.")
                    self._load_excel_file(filepath)
                    self.notebook.select(self.review_tab)

                self.root.after(0, auto_load)

            except Exception as e:
                self._log_collection(f"❌ 저장 실패: {e}")
                self.root.after(0, lambda: messagebox.showerror("오류", f"저장 실패: {e}"))

        threading.Thread(target=save_task, daemon=True).start()

    def _transfer_to_review(self):
        """수집 데이터를 엑셀 저장 없이 바로 검수탭으로 전달"""
        if not hasattr(self, 'collected_data') or not self.collected_data:
            messagebox.showwarning("경고", "먼저 수집을 실행하세요")
            return

        self._log_collection(f"📤 검수탭으로 데이터 전달 중... ({len(self.collected_data)}개)")

        def transfer_task():
            try:
                # 키워드 로드
                banned_words, _ = load_banned_words()
                excluded_words = load_excluded_words()
                bait_keywords = load_bait_keywords()

                # 검수 설정
                check_level = self.check_level_var.get()
                risk_categories = [c.strip() for c in self.risk_categories_var.get().split(',') if c.strip()]

                # 분석 결과 → 검수탭 포맷 변환
                review_data = []
                self.selected_options = {}

                for idx, product in enumerate(self.collected_data):
                    product_category = product.get('categoryPath', '') or product.get('category', '') or ''
                    product_check_level = check_level

                    for risk_cat in risk_categories:
                        if risk_cat and risk_cat.lower() in product_category.lower():
                            product_check_level = 'strict'
                            break

                    result = self._analyze_single_product(product, bait_keywords, excluded_words, product_check_level)
                    result['group_name'] = product.get('_group_name', '')

                    # 검수탭 item 포맷으로 변환
                    item = {
                        "row_idx": idx,
                        "product_name": result.get('name', '')[:30],
                        "product_id": result.get('id', ''),
                        "is_safe": result.get('is_safe', True),
                        "unsafe_reason": result.get('unsafe_reason', ''),
                        "group_name": result.get('group_name', ''),
                        "thumbnail_formula": "",
                        "thumbnail_url": result.get('thumbnail_url', ''),
                        "thumb_score": 0,
                        "thumb_nukki": False,
                        "thumb_text": False,
                        "thumb_action": "-",
                        "option_image_formula": "",
                        "option_image_url": result.get('main_option_image', ''),
                        "total_options": result.get('total_options', 0),
                        "final_options": result.get('final_options', 0),
                        "bait_options": result.get('bait_options', 0),
                        "main_option": result.get('main_option_name', ''),
                        "selected": "A",
                        "option_names": '\n'.join([f"{chr(65+i)}. {opt}" for i, opt in enumerate(result.get('final_option_list', []))]),
                        "cn_option_names": '\n'.join([f"{chr(65+i)}. {opt}" for i, opt in enumerate(result.get('cn_option_list', []))]),
                        "price_cny": result.get('min_price_cny', 0),
                        "price_krw": 0,
                        "sale_price": 0,
                        "option_images": result.get('option_images', {}),
                        "option_count": f"{result.get('final_options', 0)}/{result.get('total_options', 0)}",
                        "all_thumbnails": result.get('all_thumbnails', []),  # 전체 썸네일 (분석용)
                    }

                    # 옵션 파싱
                    item["options"] = self._parse_options(item["option_names"], item["cn_option_names"])
                    review_data.append(item)
                    self.selected_options[idx] = "A"

                    if (idx + 1) % 20 == 0:
                        self._log_collection(f"  변환 중... {idx+1}/{len(self.collected_data)}")

                # 검수탭 데이터 설정
                self.data = review_data

                def update_ui():
                    self.file_label.config(text=f"[메모리] {len(self.data)}개 상품", foreground="blue")
                    self.count_label.config(text=f"상품: {len(self.data)}개")
                    self.current_page = 0  # 첫 페이지로 리셋
                    self._render_data()
                    self.notebook.select(self.review_tab)

                self.root.after(0, update_ui)
                self._log_collection(f"✅ 검수탭 전달 완료: {len(review_data)}개")

            except Exception as e:
                self._log_collection(f"❌ 전달 실패: {e}")
                import traceback
                traceback.print_exc()

        threading.Thread(target=transfer_task, daemon=True).start()

    def _analyze_single_product(self, product: Dict, bait_keywords: list, excluded_words: list, check_level: str = 'normal') -> Dict:
        """단일 상품 분석 (기존 시뮬레이터 로직)

        Args:
            product: 상품 정보
            bait_keywords: 미끼 키워드 목록
            excluded_words: 제외 단어 목록
            check_level: 검수 레벨 (strict/normal/skip)
        """
        product_id = product.get('ID', '') or product.get('id', '')
        product_name = product.get('uploadCommonProductName', '') or product.get('name', '')

        result = {
            'id': product_id,
            'name': product_name,
            'is_safe': True,
            'unsafe_reason': '',
            'unsafe_keywords': [],
            'safe_context': '',  # 안전 컨텍스트로 무시된 키워드
            'check_level': check_level,  # 사용된 검수 레벨
            'ai_judgment': '',  # AI 판단 결과 (strict 모드)
            'total_options': 0,
            'valid_options': 0,
            'final_options': 0,
            'bait_options': 0,
            'bait_option_list': [],
            'main_option_name': '',
            'main_option_method': '',
            'final_option_list': [],
            'cn_option_list': [],
            'thumbnail_url': '',
            'main_option_image': '',
            'min_price_cny': 0,
            'max_price_cny': 0,
            # 직접 API 업로드용 추가 데이터
            'option_images': {},  # {A: url, B: url, C: url, ...}
            'option_prices': {},  # {A: price, B: price, ...}
            'all_thumbnails': [],  # 전체 썸네일 URL 목록
            'all_skus': [],  # 전체 SKU 데이터 (직접 API용)
            'raw_product': None,  # 원본 상품 데이터 전체
            # 직접 API 업로드용 (옵션 선택 시)
            'upload_detail_contents': None,  # 상세이미지 (uploadDetailContents)
            'upload_category': None,  # 카테고리 (uploadCategory)
        }

        try:
            # 1. 상품명 안전 검사 (검수 레벨에 따라 AI 사용 여부 결정)
            safety = check_product_safety(product_name, excluded_words, check_level=check_level)
            result['is_safe'] = safety['is_safe']
            result['unsafe_keywords'] = safety.get('all_found', [])

            # 안전 컨텍스트로 무시된 키워드 기록
            if safety.get('safe_context_found'):
                result['safe_context'] = ', '.join(safety['safe_context_found'][:3])

            # AI 판단 결과 기록 (strict 모드)
            if safety.get('ai_judgment'):
                result['ai_judgment'] = ', '.join(safety['ai_judgment'][:3])

            if not safety['is_safe']:
                categories = []
                cats = safety.get('categories', {})
                if cats.get('adult'):
                    categories.append(f"성인:{','.join(cats['adult'][:2])}")
                if cats.get('medical'):
                    categories.append(f"의료:{','.join(cats['medical'][:2])}")
                if cats.get('child'):
                    categories.append(f"유아:{','.join(cats['child'][:2])}")
                if cats.get('prohibited'):
                    categories.append(f"금지:{','.join(cats['prohibited'][:2])}")
                if cats.get('brand'):
                    categories.append(f"브랜드:{','.join(cats['brand'][:2])}")
                result['unsafe_reason'] = ' / '.join(categories) if categories else '위험키워드감지'

            # 2. 원본 상품 데이터 저장 (직접 API용)
            result['raw_product'] = product

            # 2-1. 직접 API 업로드용 추가 데이터 (수집 시 옵션 선택된 경우)
            if product.get('uploadDetailContents'):
                result['upload_detail_contents'] = product.get('uploadDetailContents')
            if product.get('uploadCategory'):
                result['upload_category'] = product.get('uploadCategory')

            # 3. 썸네일 URL
            thumbnails = product.get('uploadThumbnails', [])
            result['all_thumbnails'] = thumbnails  # 모든 썸네일 저장
            if thumbnails:
                result['thumbnail_url'] = thumbnails[0]

            # 4. SKU 정보
            upload_skus = product.get('uploadSkus', [])
            if not upload_skus:
                upload_skus = product.get('original_skus', [])

            # 전체 SKU 데이터 저장 (직접 API용)
            result['all_skus'] = upload_skus

            result['total_options'] = len(upload_skus)

            if upload_skus:
                # 가격 범위 (None 안전 처리)
                prices = []
                for sku in upload_skus:
                    p = sku.get('_origin_price')
                    if p is not None:
                        try:
                            p = float(p)
                            if p > 0:
                                prices.append(p)
                        except (ValueError, TypeError):
                            pass
                if prices:
                    result['min_price_cny'] = min(prices)
                    result['max_price_cny'] = max(prices)

                # 미끼옵션 필터링
                valid_skus, bait_skus = filter_bait_options(upload_skus, bait_keywords)

                result['valid_options'] = len(valid_skus)
                result['bait_options'] = len(bait_skus)

                # 미끼 옵션 정보 수집
                for bait_sku in bait_skus:
                    option_text_ko = bait_sku.get('text_ko', '') or ''
                    bait_price = bait_sku.get('_origin_price', 0)
                    display_text = option_text_ko[:20] if option_text_ko else ''
                    price_part = f"({bait_price})" if bait_price else ""
                    result['bait_option_list'].append(f"{display_text}{price_part}")

                # 대표옵션 선택: 상품명 매칭 → 첫 번째 옵션 폴백
                if valid_skus:
                    main_sku_idx, main_method = select_main_option(product_name, valid_skus)
                    main_sku = valid_skus[main_sku_idx]
                    result['main_option_name'] = main_sku.get('text_ko', '') or main_sku.get('text', '')
                    result['main_option_method'] = main_method

                    # 대표 옵션 이미지 URL
                    main_option_img = main_sku.get('urlRef', '') or main_sku.get('image', '')
                    if main_option_img:
                        result['main_option_image'] = main_option_img

                    # 옵션 개수 제한 (5개)
                    try:
                        option_count = int(self.option_count_var.get()) if hasattr(self, 'option_count_var') else 5
                    except (ValueError, TypeError):
                        option_count = 5

                    # 가격을 float로 안전하게 변환하는 헬퍼
                    def safe_price(val):
                        if val is None:
                            return 0.0
                        try:
                            return float(val)
                        except (ValueError, TypeError):
                            return 0.0

                    main_sku_price = safe_price(main_sku.get('_origin_price'))

                    if option_count > 0:
                        eligible_skus = [
                            sku for sku in valid_skus
                            if safe_price(sku.get('_origin_price')) >= main_sku_price
                        ]
                        eligible_skus.sort(key=lambda x: safe_price(x.get('_origin_price')))
                        final_skus = eligible_skus[:option_count]
                    else:
                        final_skus = valid_skus

                    result['final_options'] = len(final_skus)

                    # 최종 옵션 목록 + 옵션 이미지/가격 수집
                    labels = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
                    for idx, sku in enumerate(final_skus):
                        opt_name = sku.get('text_ko', '') or sku.get('text', '')
                        opt_cn = sku.get('text', '') or ''
                        opt_price = safe_price(sku.get('_origin_price', 0))
                        opt_image = sku.get('urlRef', '') or sku.get('image', '')

                        result['final_option_list'].append(f"{opt_name[:20]}({opt_price:.1f})")
                        result['cn_option_list'].append(opt_cn[:20])

                        # A, B, C... 라벨로 옵션 이미지/가격 저장
                        label = labels[idx] if idx < len(labels) else str(idx + 1)
                        result['option_images'][label] = opt_image
                        result['option_prices'][label] = opt_price

        except Exception as e:
            result['unsafe_reason'] = f"분석오류: {str(e)[:50]}"

        return result

    def _format_options_abc(self, options: list, max_count: int = 10) -> str:
        """옵션 목록을 A, B, C 형태로 포맷팅"""
        import re
        if not options:
            return ''

        result = []
        labels = 'ABCDEFGHIJ'
        for i, opt in enumerate(options[:max_count]):
            label = labels[i] if i < len(labels) else str(i + 1)
            opt_name = str(opt) if opt else ''
            opt_name = re.sub(r'^[A-Za-z]\.\s*', '', opt_name).strip()
            opt_name = opt_name[:30]
            result.append(f"{label}. {opt_name}")

        return '\n'.join(result)

    def _save_results_to_excel(self, filepath: str, results: list, stats: dict):
        """분석 결과를 엑셀로 저장 (기존 시뮬레이터 형식)"""
        from datetime import datetime
        import json
        wb = Workbook()
        ws = wb.active
        ws.title = "분석결과"

        # 스타일 정의
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        unsafe_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
        safe_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
        select_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        wrap_align = Alignment(vertical="top", wrap_text=True)
        center_align = Alignment(horizontal="center", vertical="center")

        # 헤더
        headers = [
            "썸네일\n이미지", "옵션\n이미지", "상품명", "안전여부", "위험사유",
            "전체옵션", "유효옵션", "최종옵션", "미끼옵션", "미끼옵션목록",
            "대표옵션", "선택방식", "선택", "옵션명", "중국어\n옵션명", "그룹명",
            "불사자ID", "옵션이미지JSON"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border

        # 데이터 입력
        for row_idx, result in enumerate(results, 2):
            col = 1

            # 1. 썸네일 이미지
            thumb_url = result.get('thumbnail_url', '')
            ws.cell(row=row_idx, column=col, value=f'=IMAGE("{thumb_url}")' if thumb_url else '')
            col += 1

            # 2. 옵션 이미지
            option_img = result.get('main_option_image', '')
            ws.cell(row=row_idx, column=col, value=f'=IMAGE("{option_img}")' if option_img else '')
            col += 1

            # 3. 상품명
            ws.cell(row=row_idx, column=col, value=result.get('name', '')[:50])
            col += 1

            # 4. 안전여부
            is_safe = result.get('is_safe', True)
            status_cell = ws.cell(row=row_idx, column=col, value='O' if is_safe else 'X')
            status_cell.alignment = center_align
            status_cell.fill = safe_fill if is_safe else unsafe_fill
            col += 1

            # 5. 위험사유
            ws.cell(row=row_idx, column=col, value=result.get('unsafe_reason', ''))
            col += 1

            # 6-9. 옵션 수량
            ws.cell(row=row_idx, column=col, value=result.get('total_options', 0))
            col += 1
            ws.cell(row=row_idx, column=col, value=result.get('valid_options', 0))
            col += 1
            ws.cell(row=row_idx, column=col, value=result.get('final_options', 0))
            col += 1
            ws.cell(row=row_idx, column=col, value=result.get('bait_options', 0))
            col += 1

            # 10. 미끼옵션목록
            bait_cell = ws.cell(row=row_idx, column=col,
                               value=self._format_options_abc(result.get('bait_option_list', [])[:5]))
            bait_cell.alignment = wrap_align
            col += 1

            # 11. 대표옵션
            ws.cell(row=row_idx, column=col, value=result.get('main_option_name', ''))
            col += 1

            # 12. 선택방식
            ws.cell(row=row_idx, column=col, value=result.get('main_option_method', ''))
            col += 1

            # 13. 선택 (사용자 입력용, 기본값 A)
            select_cell = ws.cell(row=row_idx, column=col, value='A')
            select_cell.alignment = center_align
            select_cell.fill = select_fill
            col += 1

            # 14. 옵션명
            ws.cell(row=row_idx, column=col,
                   value=self._format_options_abc(result.get('final_option_list', []))).alignment = wrap_align
            col += 1

            # 15. 중국어 옵션명
            ws.cell(row=row_idx, column=col,
                   value=self._format_options_abc(result.get('cn_option_list', []))).alignment = wrap_align
            col += 1

            # 16. 그룹명
            ws.cell(row=row_idx, column=col, value=result.get('group_name', ''))
            col += 1

            # 17. 불사자ID
            ws.cell(row=row_idx, column=col, value=result.get('id', ''))
            col += 1

            # 18. 옵션이미지JSON
            opt_images = result.get('option_images', {})
            ws.cell(row=row_idx, column=col, value=json.dumps(opt_images, ensure_ascii=False) if opt_images else '')
            col += 1

            # 테두리 적용
            for c in range(1, col):
                ws.cell(row=row_idx, column=c).border = border

        # 열 너비 조정
        column_widths = [15, 15, 40, 8, 20, 8, 8, 8, 8, 30, 25, 12, 6, 35, 35, 12, 15, 50]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # 행 높이 조정
        for row_idx in range(2, len(results) + 2):
            ws.row_dimensions[row_idx].height = 80
        ws.row_dimensions[1].height = 40

        # 필터 설정
        ws.auto_filter.ref = ws.dimensions

        # === 상세정보 시트 ===
        ws_detail = wb.create_sheet("상세정보")
        detail_headers = [
            "그룹", "불사자ID", "상품명", "안전여부", "위험사유",
            "전체옵션", "유효옵션", "최종옵션", "미끼옵션", "미끼옵션목록",
            "선택", "대표옵션", "최저가(CNY)", "최고가(CNY)", "최종옵션목록", "메인썸네일URL", "옵션이미지URL",
            "옵션이미지JSON", "옵션가격JSON", "전체썸네일", "중국어옵션목록"
        ]
        for col, header in enumerate(detail_headers, 1):
            cell = ws_detail.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border

        for row_idx, result in enumerate(results, 2):
            ws_detail.cell(row=row_idx, column=1, value=result.get('group_name', '')).border = border
            ws_detail.cell(row=row_idx, column=2, value=result.get('id', '')).border = border
            ws_detail.cell(row=row_idx, column=3, value=result.get('name', '')[:50]).border = border

            safe_cell = ws_detail.cell(row=row_idx, column=4, value='안전' if result.get('is_safe') else '위험')
            safe_cell.border = border
            safe_cell.alignment = center_align
            safe_cell.fill = safe_fill if result.get('is_safe') else unsafe_fill

            ws_detail.cell(row=row_idx, column=5, value=result.get('unsafe_reason', '')).border = border
            ws_detail.cell(row=row_idx, column=6, value=result.get('total_options', 0)).border = border
            ws_detail.cell(row=row_idx, column=7, value=result.get('valid_options', 0)).border = border
            ws_detail.cell(row=row_idx, column=8, value=result.get('final_options', 0)).border = border
            ws_detail.cell(row=row_idx, column=9, value=result.get('bait_options', 0)).border = border

            bait_cell = ws_detail.cell(row=row_idx, column=10,
                                       value=self._format_options_abc(result.get('bait_option_list', [])[:5]))
            bait_cell.alignment = wrap_align
            bait_cell.border = border

            select_cell2 = ws_detail.cell(row=row_idx, column=11, value='A')
            select_cell2.alignment = center_align
            select_cell2.fill = select_fill
            select_cell2.border = border

            ws_detail.cell(row=row_idx, column=12, value=result.get('main_option_name', '')).border = border
            ws_detail.cell(row=row_idx, column=13, value=result.get('min_price_cny', 0)).border = border
            ws_detail.cell(row=row_idx, column=14, value=result.get('max_price_cny', 0)).border = border

            final_opt_cell = ws_detail.cell(row=row_idx, column=15,
                                            value=self._format_options_abc(result.get('final_option_list', [])))
            final_opt_cell.alignment = wrap_align
            final_opt_cell.border = border

            ws_detail.cell(row=row_idx, column=16, value=result.get('thumbnail_url', '')).border = border
            ws_detail.cell(row=row_idx, column=17, value=result.get('main_option_image', '')).border = border

            # 추가 데이터 (직접 API 업로드용)
            # 18. 옵션이미지JSON
            opt_images = result.get('option_images', {})
            ws_detail.cell(row=row_idx, column=18, value=json.dumps(opt_images, ensure_ascii=False) if opt_images else '').border = border

            # 19. 옵션가격JSON
            opt_prices = result.get('option_prices', {})
            ws_detail.cell(row=row_idx, column=19, value=json.dumps(opt_prices, ensure_ascii=False) if opt_prices else '').border = border

            # 20. 전체썸네일
            all_thumbs = result.get('all_thumbnails', [])
            ws_detail.cell(row=row_idx, column=20, value='|'.join(all_thumbs) if all_thumbs else '').border = border

            # 21. 중국어옵션목록
            cn_opts = result.get('cn_option_list', [])
            ws_detail.cell(row=row_idx, column=21, value=self._format_options_abc(cn_opts)).border = border

        # 상세시트 열 너비
        detail_widths = [12, 12, 40, 8, 25, 8, 8, 8, 8, 35, 6, 25, 10, 10, 40, 45, 45, 50, 35, 80, 40]
        for i, width in enumerate(detail_widths, 1):
            ws_detail.column_dimensions[get_column_letter(i)].width = width

        # === 통계 시트 ===
        ws_stats = wb.create_sheet("통계")
        stats_data = [
            ["항목", "값"],
            ["전체 상품", stats['total']],
            ["안전 상품", stats['safe']],
            ["위험 상품", stats['unsafe']],
            ["미끼옵션 발견 상품", stats['bait_found']],
            ["분석 일시", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ]
        for row_idx, row_data in enumerate(stats_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                ws_stats.cell(row=row_idx, column=col_idx, value=value)

        # === 원본SKU데이터 시트 (직접 API 업로드용) ===
        ws_raw = wb.create_sheet("원본SKU데이터")
        raw_headers = ["불사자ID", "상품명", "전체SKU_JSON", "전체썸네일_JSON", "원본데이터요약"]
        for col, header in enumerate(raw_headers, 1):
            cell = ws_raw.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border

        for row_idx, result in enumerate(results, 2):
            ws_raw.cell(row=row_idx, column=1, value=result.get('id', '')).border = border
            ws_raw.cell(row=row_idx, column=2, value=result.get('name', '')[:50]).border = border

            # 전체 SKU JSON (직접 API용)
            all_skus = result.get('all_skus', [])
            sku_json = json.dumps(all_skus, ensure_ascii=False) if all_skus else ''
            # 엑셀 셀 크기 제한 (32767자)
            if len(sku_json) > 32000:
                sku_json = sku_json[:32000] + '...[truncated]'
            ws_raw.cell(row=row_idx, column=3, value=sku_json).border = border

            # 전체 썸네일 JSON
            all_thumbs = result.get('all_thumbnails', [])
            ws_raw.cell(row=row_idx, column=4, value=json.dumps(all_thumbs, ensure_ascii=False) if all_thumbs else '').border = border

            # 원본 데이터 요약 (키 목록만)
            raw_product = result.get('raw_product', {})
            if raw_product:
                summary_keys = list(raw_product.keys())[:20]
                ws_raw.cell(row=row_idx, column=5, value=f"keys: {', '.join(summary_keys)}").border = border

        # 원본SKU 시트 열 너비
        raw_widths = [15, 40, 100, 80, 50]
        for i, width in enumerate(raw_widths, 1):
            ws_raw.column_dimensions[get_column_letter(i)].width = width

        wb.save(filepath)

    def _log_collection(self, msg):
        """수집 로그 출력"""
        def update():
            self.collection_log.insert(tk.END, msg + "\n")
            self.collection_log.see(tk.END)
        self.root.after(0, update)

    # ===== 설정 탭 함수들 =====
    def _on_settings_column_change(self):
        """설정 탭에서 컬럼 체크박스 변경 시"""
        pass  # 실시간 미리보기는 추후 구현

    def _move_column_up(self):
        """컬럼 위로 이동"""
        idx = self.order_listbox.curselection()
        if idx and idx[0] > 0:
            i = idx[0]
            item = self.order_listbox.get(i)
            self.order_listbox.delete(i)
            self.order_listbox.insert(i - 1, item)
            self.order_listbox.selection_set(i - 1)
            self.column_order[i], self.column_order[i-1] = self.column_order[i-1], self.column_order[i]

    def _move_column_down(self):
        """컬럼 아래로 이동"""
        idx = self.order_listbox.curselection()
        if idx and idx[0] < self.order_listbox.size() - 1:
            i = idx[0]
            item = self.order_listbox.get(i)
            self.order_listbox.delete(i)
            self.order_listbox.insert(i + 1, item)
            self.order_listbox.selection_set(i + 1)
            self.column_order[i], self.column_order[i+1] = self.column_order[i+1], self.column_order[i]

    def _reset_column_settings(self):
        """컬럼 설정 기본값으로 리셋"""
        # 체크박스 리셋
        for col_id, col_info in ALL_COLUMNS.items():
            if col_id in self.settings_checkboxes:
                self.settings_checkboxes[col_id].set(col_info["default"])

        # 순서 리셋
        self.column_order = list(DEFAULT_COLUMN_ORDER)
        self.order_listbox.delete(0, tk.END)
        for col_id in self.column_order:
            if col_id in ALL_COLUMNS:
                self.order_listbox.insert(tk.END, f"{ALL_COLUMNS[col_id]['name']} ({col_id})")

    def _apply_and_save_settings(self):
        """설정 적용 및 저장"""
        # 선택된 컬럼 수집
        self.visible_columns = [col_id for col_id, var in self.settings_checkboxes.items() if var.get()]
        self._save_settings()
        self._render_data()
        messagebox.showinfo("알림", "설정이 저장되었습니다.")

    def _open_column_settings(self):
        """컬럼 설정 다이얼로그"""
        dialog = ColumnSettingsDialog(self.root, self.visible_columns, self.column_order)
        self.root.wait_window(dialog.dialog)

        if dialog.result:
            self.visible_columns = dialog.result["columns"]
            self.column_order = dialog.result["order"]
            self._save_settings()
            self._render_data()

    def _auto_load_latest(self):
        """최신 시뮬레이션 파일 자동 로드"""
        base_dir = Path(__file__).parent
        simulation_files = list(base_dir.glob("simulation_*.xlsx"))

        if simulation_files:
            simulation_files.sort(key=lambda x: x.stat().st_mtime, reverse=True)
            self._load_excel_file(str(simulation_files[0]))

    def _load_excel(self):
        """파일 선택"""
        try:
            filepath = filedialog.askopenfilename(
                title="시뮬레이션 엑셀 선택",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialdir=str(Path(__file__).parent)
            )
        except Exception:
            # initialdir 오류 시 기본 경로로 재시도
            filepath = filedialog.askopenfilename(
                title="시뮬레이션 엑셀 선택",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )
        if filepath:
            self._load_excel_file(filepath)

    def _load_excel_file(self, filepath):
        """엑셀 파일 로드 - openpyxl로 수식까지 읽기"""
        if not OPENPYXL_AVAILABLE:
            messagebox.showerror("오류", "openpyxl이 필요합니다: pip install openpyxl")
            return

        try:
            # openpyxl로 직접 읽기 (수식 보존)
            from openpyxl import load_workbook
            wb = load_workbook(filepath, data_only=False)  # data_only=False로 수식 읽기

            # 분석결과 시트 선택 (없으면 첫 번째 시트)
            if "분석결과" in wb.sheetnames:
                ws = wb["분석결과"]
            else:
                ws = wb.active

            # 헤더 읽기
            headers = []
            for col in range(1, ws.max_column + 1):
                val = ws.cell(row=1, column=col).value
                headers.append(str(val) if val else f"col_{col}")

            print(f"📊 엑셀 컬럼명: {headers}")

            # 데이터 읽기
            data_rows = []
            for row_idx in range(2, ws.max_row + 1):
                row_data = {}
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    row_data[header] = cell.value
                data_rows.append(row_data)

            # DataFrame으로 변환
            df = pd.DataFrame(data_rows)
            wb.close()

            # 디버그 파일 저장
            with open("excel_debug.txt", "w", encoding="utf-8") as f:
                f.write(f"컬럼명: {headers}\n\n")
                if data_rows:
                    f.write(f"첫 번째 행 데이터:\n")
                    for k, v in data_rows[0].items():
                        f.write(f"  [{k}]: {str(v)[:100]}\n")

        except Exception as e:
            print(f"❌ 엑셀 로드 오류: {e}")
            # 폴백: pandas로 시도
            try:
                df = pd.read_excel(filepath, engine='openpyxl')
            except Exception as e2:
                messagebox.showerror("오류", f"엑셀 로드 실패:\n{e}\n{e2}")
                return

        self._parse_excel_data(df)
        self.file_label.config(text=Path(filepath).name, foreground="black")
        self.count_label.config(text=f"상품: {len(self.data)}개")
        self.current_page = 0  # 첫 페이지로 리셋
        self._render_data()

    def _get_column_value(self, row, *possible_names):
        """여러 가능한 컬럼명을 시도하여 값 가져오기 (줄바꿈 처리 포함)"""
        for name in possible_names:
            # 직접 시도
            if name in row.index:
                val = row.get(name)
                if pd.notna(val) and val != "":
                    return val
            # 줄바꿈을 공백으로 치환해서 시도
            name_no_newline = name.replace("\n", " ")
            if name_no_newline in row.index:
                val = row.get(name_no_newline)
                if pd.notna(val) and val != "":
                    return val
            # 줄바꿈을 제거해서 시도
            name_no_space = name.replace("\n", "")
            if name_no_space in row.index:
                val = row.get(name_no_space)
                if pd.notna(val) and val != "":
                    return val
        return ""

    def _parse_excel_data(self, df):
        """엑셀 데이터 파싱 - 확장된 정보 수집"""
        self.data = []

        # 컬럼명 출력 (디버그용)
        col_list = list(df.columns)
        print(f"📊 엑셀 컬럼명 ({len(col_list)}개): {col_list}")

        # 컬럼명 정규화 맵 생성 (줄바꿈 처리)
        col_map = {}
        for col in df.columns:
            col_str = str(col)
            col_map[col_str] = col
            col_map[col_str.replace("\n", " ")] = col
            col_map[col_str.replace("\n", "")] = col

        for idx, row in df.iterrows():
            # 불사자ID 추출 (여러 컬럼명 시도 - 상세정보 시트와 분석결과 시트 모두 지원)
            product_id = (self._safe_str(self._get_column_value(row, "불사자ID", "상품ID", "id")) or str(idx)).strip()

            # 안전여부 파싱 (O/X 또는 안전/위험 모두 지원)
            safe_val = self._safe_str(self._get_column_value(row, "안전여부"))
            is_safe = safe_val in ["O", "안전", "True", "1", ""]

            # 썸네일 컬럼 (줄바꿈 여러 형태 지원)
            thumb_formula = self._safe_str(self._get_column_value(row,
                "썸네일\n이미지", "썸네일 이미지", "썸네일이미지", "메인썸네일URL"))

            # 옵션이미지 컬럼
            option_img_formula = self._safe_str(self._get_column_value(row,
                "옵션\n이미지", "옵션 이미지", "옵션이미지", "옵션이미지URL"))

            # 중국어 옵션명 컬럼
            cn_options = self._safe_str(self._get_column_value(row,
                "중국어\n옵션명", "중국어 옵션명", "중국어옵션명"))

            # 옵션명 컬럼
            option_names = self._safe_str(self._get_column_value(row,
                "옵션명", "최종옵션목록"))

            item = {
                "row_idx": idx,
                # 기본 정보
                "product_name": self._safe_str(self._get_column_value(row, "상품명"))[:30],
                "product_id": product_id,
                "is_safe": is_safe,
                "unsafe_reason": self._safe_str(self._get_column_value(row, "위험사유")),
                "group_name": self._safe_str(self._get_column_value(row, "그룹명", "그룹")),

                # 썸네일
                "thumbnail_formula": thumb_formula,
                "thumbnail_url": "",

                # 썸네일 분석 결과 (나중에 채움)
                "thumb_score": 0,
                "thumb_nukki": False,
                "thumb_text": False,
                "thumb_action": "-",

                # 옵션 정보
                "option_image_formula": option_img_formula,
                "total_options": self._safe_int(self._get_column_value(row, "전체옵션")),
                "final_options": self._safe_int(self._get_column_value(row, "최종옵션", "유효옵션")),
                "bait_options": self._safe_int(self._get_column_value(row, "미끼옵션")),
                "main_option": self._safe_str(self._get_column_value(row, "대표옵션")),
                "selected": self._safe_str(self._get_column_value(row, "선택")) or "A",
                "option_names": option_names,
                "cn_option_names": cn_options,

                # 가격 정보
                "price_cny": self._safe_float(self._get_column_value(row, "최저가(CNY)", "위안가")),
                "price_krw": self._safe_float(self._get_column_value(row, "원화가")),
                "sale_price": self._safe_float(self._get_column_value(row, "판매가")),
            }

            # URL 추출
            item["thumbnail_url"] = self._extract_image_url(item["thumbnail_formula"])
            item["option_image_url"] = self._extract_image_url(item["option_image_formula"])

            # 옵션이미지JSON 파싱 (직접 API 업로드용)
            import json
            opt_images_json = self._safe_str(self._get_column_value(row, "옵션이미지JSON"))
            if opt_images_json:
                try:
                    item["option_images"] = json.loads(opt_images_json)
                except json.JSONDecodeError:
                    item["option_images"] = {}
            else:
                item["option_images"] = {}

            # 전체썸네일 파싱 (파이프로 구분)
            all_thumbs_str = self._safe_str(self._get_column_value(row, "전체썸네일"))
            if all_thumbs_str:
                item["all_thumbnails"] = [t.strip() for t in all_thumbs_str.split('|') if t.strip()]
            else:
                item["all_thumbnails"] = []

            # 미끼옵션목록 파싱 (줄바꿈으로 구분)
            bait_list_str = self._safe_str(self._get_column_value(row, "미끼옵션목록"))
            if bait_list_str:
                item["bait_option_list"] = [b.strip() for b in bait_list_str.split('\n') if b.strip()]
            else:
                item["bait_option_list"] = []

            # 디버깅 (첫 3개만)
            if idx < 3:
                print(f"[{idx}] 썸네일formula: '{item['thumbnail_formula'][:60]}'" if item['thumbnail_formula'] else f"[{idx}] 썸네일formula: EMPTY")
                print(f"[{idx}] 썸네일URL: '{item['thumbnail_url'][:60]}'" if item['thumbnail_url'] else f"[{idx}] 썸네일URL: EMPTY")
                print(f"[{idx}] 옵션명: '{item['option_names'][:60]}'" if item['option_names'] else f"[{idx}] 옵션명: EMPTY")

            # 옵션 파싱
            item["options"] = self._parse_options(item["option_names"], item["cn_option_names"])

            if idx < 3:
                print(f"[{idx}] 파싱된 옵션 수: {len(item['options'])}")

            # 옵션수 계산
            item["option_count"] = f"{item['final_options']}/{item['total_options']}"

            self.data.append(item)
            self.selected_options[idx] = item["selected"]

    def _safe_str(self, val) -> str:
        """안전한 문자열 변환"""
        if pd.isna(val):
            return ""
        return str(val)

    def _safe_float(self, val) -> float:
        """안전한 숫자 변환"""
        if pd.isna(val):
            return 0.0
        try:
            return float(val)
        except:
            return 0.0

    def _safe_int(self, val) -> int:
        """안전한 정수 변환"""
        if pd.isna(val) or val == "":
            return 0
        try:
            return int(float(val))  # float 경유하여 "5.0" 같은 것도 처리
        except:
            return 0

    def _extract_image_url(self, formula) -> str:
        """=IMAGE("url") 에서 URL 추출"""
        if not formula:
            return ""
        if formula.startswith('=IMAGE("') and formula.endswith('")'):
            return formula[8:-2]
        return formula

    def _parse_options(self, option_names, cn_option_names) -> List[Dict]:
        """옵션명 파싱"""
        options = []
        if not option_names:
            return options

        ko_lines = option_names.strip().split('\n')
        cn_lines = cn_option_names.strip().split('\n') if cn_option_names else []

        for i, line in enumerate(ko_lines):
            line = line.strip()
            if not line:
                continue

            if '. ' in line:
                parts = line.split('. ', 1)
                label = parts[0].strip()
                name = parts[1].strip() if len(parts) > 1 else ""
            else:
                label = chr(ord('A') + i)
                name = line

            cn_name = ""
            if i < len(cn_lines):
                cn_line = cn_lines[i].strip()
                if '. ' in cn_line:
                    cn_name = cn_line.split('. ', 1)[1] if len(cn_line.split('. ', 1)) > 1 else cn_line
                else:
                    cn_name = cn_line

            options.append({
                "label": label,
                "name": name,
                "cn_name": cn_name
            })

        return options

    def _render_data(self):
        """데이터 렌더링 - 현재 페이지만 표시 (페이지네이션)"""
        for widget in self.scrollable_frame.winfo_children():
            widget.destroy()

        self.option_frames = {}
        self.option_cells = {}

        # 페이지 정보 업데이트
        self._update_page_info()

        if not self.data:
            ttk.Label(self.scrollable_frame, text="데이터 없음").pack(pady=50)
            return

        # 표시할 컬럼 (순서대로)
        visible_ordered = [col for col in self.column_order if col in self.visible_columns]
        # 순서에 없는 컬럼 추가
        for col in self.visible_columns:
            if col not in visible_ordered:
                visible_ordered.append(col)

        # 헤더
        self._create_header(visible_ordered)

        # 현재 페이지의 데이터만 표시 (페이지네이션)
        start_idx = self.current_page * self.page_size
        end_idx = start_idx + self.page_size
        page_data = self.data[start_idx:end_idx]

        # 데이터 행 (현재 페이지만) - 배치 업데이트
        for i, item in enumerate(page_data):
            self._create_row(item, visible_ordered)
            # 5행마다 UI 업데이트 (프리징 방지)
            if (i + 1) % 5 == 0:
                self.root.update_idletasks()

    def _create_header(self, columns):
        """헤더 생성 - 픽셀 단위로 정확히 맞춤"""
        header_frame = tk.Frame(self.scrollable_frame, bg="#4472C4")
        header_frame.pack(fill=tk.X, pady=(0, 0))

        for col_id in columns:
            col_info = ALL_COLUMNS.get(col_id, {"name": col_id, "width": 100})
            width = col_info["width"]

            # 헤더 셀 (테두리 포함)
            cell = tk.Frame(header_frame, width=width, height=30, bg="#4472C4",
                           highlightbackground="#2c5282", highlightthickness=1)
            cell.pack(side=tk.LEFT, padx=0, pady=0)
            cell.pack_propagate(False)

            lbl = tk.Label(
                cell,
                text=col_info["name"],
                bg="#4472C4",
                fg="white",
                font=("맑은 고딕", 9, "bold")
            )
            lbl.pack(expand=True)

    def _create_row(self, item, columns):
        """데이터 행 생성 - 세로선으로 구분"""
        row_idx = item["row_idx"]
        bg_color = "#C8E6C9" if item.get("is_safe", True) else "#FFCDD2"
        border_color = "#888888"

        row_frame = tk.Frame(self.scrollable_frame, bg=border_color)
        row_frame.pack(fill=tk.X, pady=0)

        for col_id in columns:
            col_info = ALL_COLUMNS.get(col_id, {"width": 100})
            width = col_info["width"]

            # 셀 프레임 (테두리 효과)
            cell_frame = tk.Frame(row_frame, width=width, height=90, bg=bg_color,
                                 highlightbackground=border_color, highlightthickness=1)
            cell_frame.pack(side=tk.LEFT, padx=0, pady=0)
            cell_frame.pack_propagate(False)

            # 컬럼별 렌더링
            if col_id == "thumbnail":
                self._render_thumbnail(cell_frame, item, bg_color)
            elif col_id == "option_image":
                self._render_option_image(cell_frame, item, bg_color)
            elif col_id == "options":
                self._render_options(cell_frame, item, row_idx, bg_color)
                # 옵션 셀 참조 저장 (확장 시 해당 셀만 업데이트용)
                self.option_cells[row_idx] = (cell_frame, item, bg_color)
            elif col_id == "is_safe":
                self._render_safe(cell_frame, item, bg_color)
            elif col_id == "thumb_score":
                self._render_thumb_score(cell_frame, item, bg_color)
            elif col_id == "thumb_action":
                self._render_thumb_action(cell_frame, item, bg_color)
            elif col_id == "sale_price":
                self._render_price(cell_frame, item.get("sale_price", 0), bg_color)
            elif col_id == "price_cny":
                self._render_price(cell_frame, item.get("price_cny", 0), bg_color, "CNY")
            elif col_id == "price_krw":
                self._render_price(cell_frame, item.get("price_krw", 0), bg_color)
            elif col_id == "product_id":
                self._render_product_id(cell_frame, item, bg_color)
            elif col_id == "bait_options":
                self._render_bait_options(cell_frame, item, bg_color)
            elif col_id == "bait_keywords":
                self._render_bait_keywords(cell_frame, item, bg_color)
            elif col_id == "unsafe_reason":
                self._render_unsafe_reason(cell_frame, item, bg_color)
            elif col_id == "option_list":
                self._render_option_list(cell_frame, item, bg_color)
            else:
                # 일반 텍스트 컬럼
                value = str(item.get(col_id, ""))[:20]
                tk.Label(cell_frame, text=value, bg=bg_color,
                        font=("맑은 고딕", 9), wraplength=width-10).pack(expand=True)

    def _render_product_id(self, frame, item, bg_color):
        """불사자ID 렌더링 - 클릭 시 불사자 옵션탭 열기"""
        product_id = item.get("product_id", "")

        id_label = tk.Label(
            frame,
            text=product_id[:12] + ".." if len(product_id) > 12 else product_id,
            bg=bg_color,
            fg="#2196F3",  # 파란색 링크 스타일
            font=("맑은 고딕", 9, "underline"),
            cursor="hand2"
        )
        id_label.pack(expand=True)

        # 클릭 시 불사자 옵션탭 열기
        if product_id:
            id_label.bind("<Button-1>", lambda e, pid=product_id: self._open_bulsaja_option_tab(pid))

    def _open_bulsaja_option_tab(self, product_id):
        """불사자 상품 상세수정 페이지 열기"""
        url = f"https://www.bulsaja.com/products/manage/list/{product_id}"
        webbrowser.open(url)

    def _render_bait_options(self, frame, item, bg_color):
        """미끼옵션 렌더링 - 클릭 시 상세 보기"""
        bait_count = item.get("bait_options", 0)

        if bait_count > 0:
            # 미끼 있으면 빨간색 + 클릭 가능
            lbl = tk.Label(
                frame,
                text=f"{bait_count}개",
                bg=bg_color,
                fg="#F44336",
                font=("맑은 고딕", 10, "bold"),
                cursor="hand2"
            )
            lbl.pack(expand=True)
            lbl.bind("<Button-1>", lambda e, it=item: self._show_bait_detail(it))
        else:
            # 없으면 그냥 표시
            tk.Label(frame, text="-", bg=bg_color, fg="gray",
                    font=("맑은 고딕", 9)).pack(expand=True)

    def _show_bait_detail(self, item):
        """미끼옵션 상세 보기 팝업"""
        popup = tk.Toplevel(self.root)
        popup.title("미끼옵션 상세")
        popup.geometry("600x450")
        popup.transient(self.root)

        # 상품 정보
        info_frame = ttk.LabelFrame(popup, text="상품 정보", padding=10)
        info_frame.pack(fill=tk.X, padx=10, pady=5)

        product_name = item.get("product_name", "")[:50]
        ttk.Label(info_frame, text=f"상품명: {product_name}").pack(anchor=tk.W)
        ttk.Label(info_frame, text=f"전체옵션: {item.get('total_options', 0)}개 | 유효옵션: {item.get('valid_options', 0)}개 | 미끼옵션: {item.get('bait_options', 0)}개").pack(anchor=tk.W)

        # 미끼옵션 목록
        bait_frame = ttk.LabelFrame(popup, text="미끼옵션 목록", padding=10)
        bait_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        # 트리뷰
        columns = ("option", "reason")
        tree = ttk.Treeview(bait_frame, columns=columns, show="headings", height=10)
        tree.heading("option", text="옵션명")
        tree.heading("reason", text="미끼 판정 사유")
        tree.column("option", width=250)
        tree.column("reason", width=300)

        scrollbar = ttk.Scrollbar(bait_frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # 미끼옵션 목록 표시
        bait_list = item.get("bait_option_list", [])
        bait_keywords = []
        if BULSAJA_API_AVAILABLE:
            bait_keywords = load_bait_keywords()

        for bait_text in bait_list:
            # 매칭된 키워드 찾기
            matched = []
            for kw in bait_keywords:
                if kw.lower() in bait_text.lower():
                    matched.append(kw)

            reason = f"키워드: {', '.join(matched)}" if matched else "가격 기준"
            tree.insert("", tk.END, values=(bait_text, reason, ", ".join(matched)))

        # 오탐 수정 버튼 영역
        fix_frame = ttk.Frame(bait_frame)
        fix_frame.pack(fill=tk.X, pady=5)

        def remove_keyword():
            """선택한 항목의 키워드를 미끼 목록에서 제거 (오탐 수정)"""
            selected = tree.selection()
            if not selected:
                messagebox.showwarning("경고", "제거할 항목을 선택하세요")
                return

            keywords_to_remove = set()
            for sel in selected:
                values = tree.item(sel)['values']
                if len(values) >= 3 and values[2]:
                    for kw in str(values[2]).split(", "):
                        if kw.strip():
                            keywords_to_remove.add(kw.strip())

            if not keywords_to_remove:
                messagebox.showinfo("알림", "제거할 키워드가 없습니다 (가격 기준 판정)")
                return

            if messagebox.askyesno("확인", f"다음 키워드를 미끼 목록에서 제거할까요?\n\n{', '.join(keywords_to_remove)}"):
                current_keywords = load_bait_keywords() if BULSAJA_API_AVAILABLE else []
                new_keywords = [kw for kw in current_keywords if kw not in keywords_to_remove]

                if save_bait_keywords(new_keywords):
                    messagebox.showinfo("완료", f"{len(keywords_to_remove)}개 키워드 제거됨\n\n※ 다시 수집해야 반영됩니다")
                    # 트리뷰에서 삭제
                    for sel in selected:
                        tree.delete(sel)
                else:
                    messagebox.showerror("오류", "저장 실패")

        ttk.Button(fix_frame, text="❌ 선택 키워드 제거 (오탐 수정)", command=remove_keyword).pack(side=tk.LEFT, padx=5)
        ttk.Label(fix_frame, text="← 미끼 아닌데 미끼로 판정된 경우", foreground="gray").pack(side=tk.LEFT)

        # 키워드 설정 영역
        kw_frame = ttk.LabelFrame(popup, text="미끼 키워드 관리", padding=10)
        kw_frame.pack(fill=tk.X, padx=10, pady=5)

        ttk.Label(kw_frame, text="현재 미끼 키워드:").pack(anchor=tk.W)
        kw_text = tk.Text(kw_frame, height=2, width=70)
        kw_text.pack(fill=tk.X, pady=5)
        kw_text.insert("1.0", ", ".join(bait_keywords) if bait_keywords else "(없음)")

        kw_btn_frame = ttk.Frame(kw_frame)
        kw_btn_frame.pack(fill=tk.X)

        def save_keywords():
            new_kw = kw_text.get("1.0", tk.END).strip()
            keywords = [k.strip() for k in new_kw.replace("\n", ",").split(",") if k.strip()]
            if BULSAJA_API_AVAILABLE and save_bait_keywords(keywords):
                messagebox.showinfo("완료", f"{len(keywords)}개 키워드 저장됨\n\n※ 다시 수집해야 반영됩니다")
            else:
                messagebox.showerror("오류", "저장 실패")

        def add_keyword():
            """새 키워드 추가"""
            new_kw = simpledialog.askstring("키워드 추가", "추가할 미끼 키워드:")
            if new_kw and new_kw.strip():
                current = kw_text.get("1.0", tk.END).strip()
                if current and current != "(없음)":
                    kw_text.delete("1.0", tk.END)
                    kw_text.insert("1.0", current + ", " + new_kw.strip())
                else:
                    kw_text.delete("1.0", tk.END)
                    kw_text.insert("1.0", new_kw.strip())

        ttk.Button(kw_btn_frame, text="➕ 키워드 추가", command=add_keyword).pack(side=tk.LEFT, padx=5)
        ttk.Button(kw_btn_frame, text="💾 저장", command=save_keywords).pack(side=tk.LEFT, padx=5)

        # 닫기 버튼
        btn_frame = ttk.Frame(popup, padding=10)
        btn_frame.pack(fill=tk.X)
        ttk.Button(btn_frame, text="닫기", command=popup.destroy).pack(side=tk.RIGHT, padx=5)

    def _render_thumbnail(self, frame, item, bg_color):
        """썸네일 렌더링 - 이미지는 나중에 로드 (성능)"""
        thumb_label = tk.Label(frame, text="[썸네일]", bg=bg_color, font=("맑은 고딕", 8))
        thumb_label.pack(expand=True)

        # 이미지 로딩은 "썸네일 로드" 버튼으로 따로 실행 (500개 이미지 다운로드 너무 느림)
        # 나중에 로드할 수 있도록 참조 저장
        item["_thumb_label"] = thumb_label

    def _render_option_image(self, frame, item, bg_color):
        """옵션 이미지 렌더링 - 이미지는 나중에 로드"""
        opt_img_label = tk.Label(frame, text="[옵션]", bg=bg_color, font=("맑은 고딕", 8))
        opt_img_label.pack(expand=True)

        # 나중에 로드할 수 있도록 참조 저장
        item["_opt_img_label"] = opt_img_label

    def _render_options(self, frame, item, row_idx, bg_color):
        """옵션 선택 영역 렌더링 - 6개까지만 표시"""
        options = item.get("options", [])
        option_images = item.get("option_images", {})  # {"A": url, "B": url, ...}

        # 최대 6개만 표시 (속도 우선)
        display_options = options[:6]

        for i, opt in enumerate(display_options):
            is_selected = (self.selected_options.get(row_idx, "A") == opt["label"])

            opt_frame = tk.Frame(
                frame,
                width=60, height=60,
                bg="#2196F3" if is_selected else "#E0E0E0",
                relief="solid",
                bd=2 if is_selected else 1,
                cursor="hand2"
            )
            opt_frame.pack(side=tk.LEFT, padx=1, pady=1)
            opt_frame.pack_propagate(False)

            opt_frame.bind("<Button-1>", lambda e, r=row_idx, o=opt["label"]: self._on_option_click(r, o))

            lbl_color = "white" if is_selected else "black"
            lbl_bg = "#2196F3" if is_selected else "#E0E0E0"

            # 옵션 이미지 (40x40으로 축소)
            img_label = tk.Label(opt_frame, text="", bg=lbl_bg, width=40, height=40)
            img_label.pack(side=tk.TOP)
            img_label.bind("<Button-1>", lambda e, r=row_idx, o=opt["label"]: self._on_option_click(r, o))

            # 옵션 이미지 비동기 로딩
            opt_label = opt["label"]
            if opt_label in option_images:
                self._load_option_button_image(img_label, option_images[opt_label], lbl_bg)

            # 하단에 라벨만 (간단하게)
            label_widget = tk.Label(opt_frame, text=opt['label'], bg=lbl_bg, fg=lbl_color,
                                   font=("맑은 고딕", 9, "bold"))
            label_widget.pack(side=tk.BOTTOM)
            label_widget.bind("<Button-1>", lambda e, r=row_idx, o=opt["label"]: self._on_option_click(r, o))

            self.option_frames[(row_idx, opt["label"])] = {
                "frame": opt_frame,
                "label": label_widget,
                "name": label_widget,
                "img": img_label
            }

    def _load_option_button_image(self, label, url, bg_color):
        """옵션 버튼용 작은 이미지 비동기 로딩 (최적화)"""
        if not PIL_AVAILABLE or not url:
            return

        # 캐시 확인
        if url in self.option_image_cache:
            try:
                label.config(image=self.option_image_cache[url], text="")
            except tk.TclError:
                pass
            return

        def load():
            try:
                headers = {
                    'User-Agent': 'Mozilla/5.0',
                    'Referer': 'https://www.bulsaja.com/'
                }
                response = requests.get(url, headers=headers, timeout=2)  # 타임아웃 2초로 단축
                if response.status_code != 200:
                    return

                img = Image.open(BytesIO(response.content))
                img.thumbnail((40, 40), Image.Resampling.LANCZOS)
                photo = ImageTk.PhotoImage(img)

                # LRU 캐시 제한 (오래된 것 삭제)
                if len(self.option_image_cache) >= self._cache_max_size:
                    oldest = next(iter(self.option_image_cache))
                    del self.option_image_cache[oldest]
                self.option_image_cache[url] = photo

                # UI 업데이트
                self.root.after(0, lambda: self._update_option_image(label, photo))
            except Exception:
                pass

        # ThreadPoolExecutor 사용 (스레드 폭증 방지)
        self._image_executor.submit(load)

    def _update_option_image(self, label, photo):
        """메인 스레드에서 옵션 이미지 업데이트"""
        try:
            label.config(image=photo, text="")
        except tk.TclError:
            pass  # 위젯이 이미 파괴됨

    def _toggle_options_expand(self, row_idx):
        """옵션 확장/축소 토글 - 해당 셀만 업데이트 (전체 새로고침 X)"""
        if row_idx in self.expanded_rows:
            self.expanded_rows.discard(row_idx)
        else:
            self.expanded_rows.add(row_idx)

        # 해당 행의 옵션 셀만 업데이트 (전체 새로고침 안함!)
        if row_idx in self.option_cells:
            cell_frame, item, bg_color = self.option_cells[row_idx]
            # 기존 내용 삭제
            for widget in cell_frame.winfo_children():
                widget.destroy()
            # 해당 행의 옵션 프레임 정보 제거
            keys_to_remove = [k for k in self.option_frames.keys() if k[0] == row_idx]
            for k in keys_to_remove:
                del self.option_frames[k]
            # 다시 렌더링
            self._render_options(cell_frame, item, row_idx, bg_color)

    def _get_bait_keywords_cached(self):
        """미끼 키워드 캐시 로드 (1회만 파일 I/O)"""
        if self._bait_keywords_cache is None and BULSAJA_API_AVAILABLE:
            self._bait_keywords_cache = load_bait_keywords()
            # 미리 소문자로 변환해서 저장 (매번 lower() 호출 방지)
            self._bait_keywords_lower = {kw.lower(): kw for kw in self._bait_keywords_cache}
        return self._bait_keywords_cache or []

    def _render_bait_keywords(self, frame, item, bg_color):
        """미끼옵션 키워드 렌더링 (최적화)"""
        bait_list = item.get("bait_option_list", [])

        if not bait_list:
            tk.Label(frame, text="-", bg=bg_color, fg="gray",
                    font=("맑은 고딕", 9)).pack(expand=True)
            return

        # 매칭된 키워드 수집 (캐시된 키워드 사용)
        matched_keywords = set()
        bait_keywords = self._get_bait_keywords_cached()
        if bait_keywords:
            for bait_text in bait_list:
                bait_lower = bait_text.lower()
                for kw in bait_keywords:
                    if kw.lower() in bait_lower:
                        matched_keywords.add(kw)
                        break  # 하나 찾으면 다음 옵션으로

        if matched_keywords:
            kw_text = ", ".join(list(matched_keywords)[:3])  # 최대 3개
            if len(matched_keywords) > 3:
                kw_text += f" +{len(matched_keywords)-3}"
        else:
            kw_text = "가격기준"

        lbl = tk.Label(frame, text=kw_text, bg=bg_color, fg="#F44336",
                      font=("맑은 고딕", 8), wraplength=110)
        lbl.pack(expand=True)

    def _render_option_list(self, frame, item, bg_color):
        """옵션명 목록 렌더링 (A. xxx / B. yyy 형태로 세로 표시)"""
        options = item.get("options", [])

        if not options:
            # options가 없으면 option_names에서 파싱
            option_names = item.get("option_names", "")
            if option_names:
                lines = [line.strip() for line in option_names.split('\n') if line.strip()][:6]
                text = '\n'.join(lines)
            else:
                text = "-"
        else:
            # options에서 라벨+이름 조합
            lines = []
            for opt in options[:6]:
                name = opt.get("name", "")[:12]
                lines.append(f"{opt['label']}. {name}")
            text = '\n'.join(lines)

        lbl = tk.Label(frame, text=text, bg=bg_color, fg="black",
                      font=("맑은 고딕", 8), justify=tk.LEFT, anchor="nw")
        lbl.pack(expand=True, fill=tk.BOTH, padx=2, pady=2)

    def _render_unsafe_reason(self, frame, item, bg_color):
        """위험사유 렌더링"""
        is_safe = item.get("is_safe", True)

        if is_safe:
            tk.Label(frame, text="-", bg=bg_color, fg="gray",
                    font=("맑은 고딕", 9)).pack(expand=True)
            return

        unsafe_reason = item.get("unsafe_reason", "")

        # 키워드 간단히 표시 (예: "성인: 바이브 | 의료: 산소" -> "바이브, 산소")
        keywords = []
        for part in unsafe_reason.split("|"):
            part = part.strip()
            if ":" in part:
                _, kw = part.split(":", 1)
                keywords.append(kw.strip()[:6])  # 6자 제한
            elif part:
                keywords.append(part[:6])

        if keywords:
            kw_text = ", ".join(keywords[:3])  # 최대 3개
            if len(keywords) > 3:
                kw_text += f" +{len(keywords)-3}"
            lbl = tk.Label(frame, text=kw_text, bg=bg_color, fg="#F44336",
                          font=("맑은 고딕", 8), wraplength=110)
        else:
            # 키워드 없으면 - 표시 (안전 컬럼에서 이미 X 표시됨)
            lbl = tk.Label(frame, text="-", bg=bg_color, fg="gray",
                          font=("맑은 고딕", 9))

        lbl.pack(expand=True)

    def _render_safe(self, frame, item, bg_color):
        """안전 여부 렌더링 - 클릭하면 토글"""
        is_safe = item.get("is_safe", True)
        safe_text = "O" if is_safe else "X"
        safe_color = "#4CAF50" if is_safe else "#F44336"

        lbl = tk.Label(frame, text=safe_text, bg=bg_color, fg=safe_color,
                      font=("맑은 고딕", 16, "bold"), cursor="hand2")
        lbl.pack(expand=True)

        # 클릭하면 안전/위험 토글
        def toggle_safe(e):
            item["is_safe"] = not item.get("is_safe", True)
            new_safe = item["is_safe"]
            lbl.config(
                text="O" if new_safe else "X",
                fg="#4CAF50" if new_safe else "#F44336"
            )
            # 배경색 변경
            new_bg = "#C8E6C9" if new_safe else "#FFCDD2"
            frame.config(bg=new_bg)
            lbl.config(bg=new_bg)
            # 안전으로 바꾸면 위험사유 제거
            if new_safe:
                item["unsafe_reason"] = ""

        lbl.bind("<Button-1>", toggle_safe)

    def _show_safety_detail(self, item):
        """안전/위험 상세 보기 팝업"""
        popup = tk.Toplevel(self.root)
        popup.title("위험 판정 상세")
        popup.geometry("550x400")
        popup.transient(self.root)

        # 상품 정보
        info_frame = ttk.LabelFrame(popup, text="상품 정보", padding=10)
        info_frame.pack(fill=tk.X, padx=10, pady=5)

        product_name = item.get("product_name", "")
        ttk.Label(info_frame, text=f"상품명: {product_name[:60]}...", wraplength=500).pack(anchor=tk.W)
        ttk.Label(info_frame, text=f"판정: ❌ 위험", foreground="red", font=("맑은 고딕", 10, "bold")).pack(anchor=tk.W)

        # 위험 사유
        reason_frame = ttk.LabelFrame(popup, text="위험 판정 사유", padding=10)
        reason_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        unsafe_reason = item.get("unsafe_reason", "")
        unsafe_keywords = item.get("unsafe_keywords", [])

        # 트리뷰
        columns = ("keyword", "category")
        tree = ttk.Treeview(reason_frame, columns=columns, show="headings", height=8)
        tree.heading("keyword", text="위험 키워드")
        tree.heading("category", text="카테고리")
        tree.column("keyword", width=200)
        tree.column("category", width=150)
        tree.pack(fill=tk.BOTH, expand=True)

        # 위험 키워드 파싱
        if unsafe_reason:
            # unsafe_reason 형식: "성인: 바이브레이터 | 의료: 산소발생기"
            for part in unsafe_reason.split("|"):
                part = part.strip()
                if ":" in part:
                    cat, kw = part.split(":", 1)
                    tree.insert("", tk.END, values=(kw.strip(), cat.strip()))
                elif part:
                    tree.insert("", tk.END, values=(part, "기타"))

        # 오탐 수정 영역
        fix_frame = ttk.LabelFrame(popup, text="오탐 수정", padding=10)
        fix_frame.pack(fill=tk.X, padx=10, pady=5)

        ttk.Label(fix_frame, text="이 상품이 실제로 안전하다면, 키워드를 예외 처리하세요.").pack(anchor=tk.W)

        def add_to_excluded():
            """선택한 키워드를 예외 목록에 추가"""
            selected = tree.selection()
            if not selected:
                messagebox.showwarning("경고", "예외 처리할 키워드를 선택하세요")
                return

            keywords = [tree.item(sel)['values'][0] for sel in selected]

            if messagebox.askyesno("확인", f"다음 키워드를 예외 목록에 추가할까요?\n\n{', '.join(keywords)}\n\n※ 예외 처리하면 다음 검수부터 이 키워드는 무시됩니다"):
                try:
                    from bulsaja_common import load_excluded_words
                    excluded = load_excluded_words()
                    excluded_file = "excluded_words.json"

                    # 기존 데이터 로드
                    import json
                    if os.path.exists(excluded_file):
                        with open(excluded_file, 'r', encoding='utf-8') as f:
                            data = json.load(f)
                    else:
                        data = {'words': []}

                    # 키워드 추가
                    for kw in keywords:
                        if kw not in data['words']:
                            data['words'].append(kw)

                    # 저장
                    with open(excluded_file, 'w', encoding='utf-8') as f:
                        json.dump(data, f, ensure_ascii=False, indent=2)

                    messagebox.showinfo("완료", f"{len(keywords)}개 키워드가 예외 목록에 추가됨\n\n※ 다시 수집해야 반영됩니다")

                    # 트리뷰에서 삭제
                    for sel in selected:
                        tree.delete(sel)

                except Exception as e:
                    messagebox.showerror("오류", f"저장 실패: {e}")

        def mark_as_safe():
            """이 상품을 안전으로 변경"""
            item["is_safe"] = True
            item["unsafe_reason"] = ""
            messagebox.showinfo("완료", "이 상품을 안전으로 변경했습니다.\n\n※ 저장해야 엑셀에 반영됩니다")
            popup.destroy()
            self._render_data()

        btn_frame = ttk.Frame(fix_frame)
        btn_frame.pack(fill=tk.X, pady=5)
        ttk.Button(btn_frame, text="🔓 선택 키워드 예외 처리", command=add_to_excluded).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="✅ 이 상품 안전으로 변경", command=mark_as_safe).pack(side=tk.LEFT, padx=5)

        # 닫기
        ttk.Button(popup, text="닫기", command=popup.destroy).pack(pady=10)

    def _render_thumb_score(self, frame, item, bg_color):
        """썸네일 점수 렌더링"""
        score = item.get("thumb_score", 0)

        if score >= 80:
            color = "#4CAF50"  # 녹색
        elif score >= 50:
            color = "#FF9800"  # 주황
        elif score > 0:
            color = "#F44336"  # 빨강
        else:
            color = "gray"

        tk.Label(frame, text=f"{score}점", bg=bg_color, fg=color,
                font=("맑은 고딕", 12, "bold")).pack(expand=True)

    def _render_thumb_action(self, frame, item, bg_color):
        """필요 작업 렌더링"""
        action = item.get("thumb_action", "-")

        action_colors = {
            "none": ("#4CAF50", "OK"),
            "translate": ("#FF9800", "번역"),
            "nukki": ("#2196F3", "누끼"),
            "both": ("#9C27B0", "둘다"),
            "manual": ("#F44336", "수동"),
            "-": ("gray", "-")
        }

        color, text = action_colors.get(action, ("gray", action))
        tk.Label(frame, text=text, bg=bg_color, fg=color,
                font=("맑은 고딕", 10, "bold")).pack(expand=True)

    def _render_price(self, frame, price, bg_color, prefix=""):
        """가격 렌더링"""
        if price > 0:
            if prefix:
                text = f"{prefix} {price:,.0f}"
            else:
                text = f"{price:,.0f}원"
        else:
            text = "-"

        tk.Label(frame, text=text, bg=bg_color,
                font=("맑은 고딕", 9)).pack(expand=True)

    def _on_option_click(self, row_idx, option_label):
        """옵션 클릭 - 선택 변경 및 옵션이미지 업데이트"""
        old_selected = self.selected_options.get(row_idx, "A")

        if (row_idx, old_selected) in self.option_frames:
            old_widgets = self.option_frames[(row_idx, old_selected)]
            old_widgets["frame"].config(bg="#E0E0E0", bd=1)
            old_widgets["label"].config(bg="#E0E0E0", fg="black")
            old_widgets["name"].config(bg="#E0E0E0", fg="black")
            if "img" in old_widgets:
                old_widgets["img"].config(bg="#E0E0E0")

        if (row_idx, option_label) in self.option_frames:
            new_widgets = self.option_frames[(row_idx, option_label)]
            new_widgets["frame"].config(bg="#2196F3", bd=2)
            new_widgets["label"].config(bg="#2196F3", fg="white")
            new_widgets["name"].config(bg="#2196F3", fg="white")
            if "img" in new_widgets:
                new_widgets["img"].config(bg="#2196F3")

        self.selected_options[row_idx] = option_label

        # 옵션이미지 업데이트 (option_images에서 선택된 옵션의 이미지 URL 가져와서 로드)
        if row_idx < len(self.data):
            item = self.data[row_idx]
            opt_images = item.get("option_images", {})
            new_img_url = opt_images.get(option_label, "")

            if new_img_url and "_opt_img_label" in item:
                # 비동기로 이미지 로드
                self._load_option_image_async(item, new_img_url)

    def _load_option_image_async(self, item, url):
        """옵션 이미지를 비동기로 로드하여 라벨에 표시"""
        import threading

        def load_and_update():
            try:
                response = requests.get(url, timeout=2)
                if response.status_code == 200:
                    img = Image.open(BytesIO(response.content))
                    img = img.resize((75, 75), Image.Resampling.LANCZOS)
                    photo = ImageTk.PhotoImage(img)

                    # UI 업데이트는 메인 스레드에서
                    def update_ui():
                        if "_opt_img_label" in item:
                            label = item["_opt_img_label"]
                            label.config(image=photo, text="")
                            label.image = photo  # 참조 유지

                    self.root.after(0, update_ui)
            except Exception as e:
                print(f"옵션이미지 로드 실패: {e}")

        thread = threading.Thread(target=load_and_update, daemon=True)
        thread.start()

    def _load_all_thumbnails(self):
        """현재 페이지 썸네일 이미지 로드 (병렬 처리로 빠르게)"""
        from concurrent.futures import ThreadPoolExecutor, as_completed

        if not self.data:
            messagebox.showwarning("경고", "먼저 데이터를 로드하세요")
            return

        # 현재 페이지 데이터만
        start_idx = self.current_page * self.page_size
        end_idx = start_idx + self.page_size
        page_data = self.data[start_idx:end_idx]

        if not page_data:
            messagebox.showwarning("경고", "현재 페이지에 데이터가 없습니다")
            return

        # 진행 다이얼로그
        progress = tk.Toplevel(self.root)
        progress.title(f"썸네일 로드 중... (페이지 {self.current_page + 1})")
        progress.geometry("350x120")
        progress.transient(self.root)

        progress_var = tk.StringVar(value="준비 중...")
        ttk.Label(progress, textvariable=progress_var).pack(pady=10)
        pb = ttk.Progressbar(progress, length=300, mode='determinate')
        pb.pack(pady=10)

        cancel_var = tk.BooleanVar(value=False)
        ttk.Button(progress, text="취소", command=lambda: cancel_var.set(True)).pack(pady=5)

        progress.update()

        def load_single_image(task):
            """단일 이미지 로드 (병렬 실행용)"""
            item_idx, img_type, url = task
            if not url:
                return (item_idx, img_type, None, "no_url")
            try:
                headers = {'User-Agent': 'Mozilla/5.0', 'Referer': 'https://www.bulsaja.com/'}
                response = requests.get(url, headers=headers, timeout=2)
                if response.status_code == 200:
                    img = Image.open(BytesIO(response.content))
                    img = img.resize((75, 75), Image.Resampling.LANCZOS)
                    return (item_idx, img_type, img, "ok")
                return (item_idx, img_type, None, "http_error")
            except:
                return (item_idx, img_type, None, "error")

        def load_thread():
            # 현재 페이지의 썸네일 + 옵션이미지 작업 목록 생성
            tasks = []
            for item in page_data:
                i = item.get("row_idx", 0)
                thumb_url = item.get("thumbnail_url", "")
                opt_url = item.get("option_image_url", "")
                if thumb_url:
                    tasks.append((i, "thumb", thumb_url))
                if opt_url:
                    tasks.append((i, "opt", opt_url))

            total = len(tasks)
            if total == 0:
                self.root.after(0, lambda: progress_var.set("로드할 이미지 없음"))
                self.root.after(1500, progress.destroy)
                return

            loaded = 0
            failed = 0
            completed = 0

            # 15개 동시 다운로드
            with ThreadPoolExecutor(max_workers=15) as executor:
                futures = {executor.submit(load_single_image, task): task for task in tasks}

                for future in as_completed(futures):
                    if cancel_var.get():
                        executor.shutdown(wait=False, cancel_futures=True)
                        break

                    completed += 1
                    item_idx, img_type, img, status = future.result()

                    if status == "ok" and img:
                        if img_type == "thumb":
                            label = self.data[item_idx].get("_thumb_label")
                        else:
                            label = self.data[item_idx].get("_opt_img_label")

                        if label:
                            try:
                                photo = ImageTk.PhotoImage(img)
                                def update_label(lbl, p):
                                    try:
                                        lbl.config(image=p, text="")
                                        lbl.image = p
                                    except:
                                        pass
                                self.root.after(0, update_label, label, photo)
                                loaded += 1
                            except:
                                failed += 1
                        else:
                            failed += 1
                    else:
                        if status != "no_url":
                            failed += 1

                    # 진행률 업데이트 (20개마다)
                    if completed % 20 == 0:
                        pct = completed / total * 100
                        self.root.after(0, lambda v=pct, l=loaded, f=failed, c=completed, t=total:
                            (pb.configure(value=v), progress_var.set(f"{c}/{t} ({l}성공/{f}실패)")))

            # 완료
            self.root.after(0, lambda: (
                progress_var.set(f"완료: {loaded}성공 / {failed}실패"),
                pb.configure(value=100)
            ))
            self.root.after(1500, progress.destroy)

        threading.Thread(target=load_thread, daemon=True).start()

    def _analyze_thumbnails(self):
        """
        썸네일 분석 - 전체 썸네일 중 최적 이미지 선택

        점수 기준 (thumbnail_analyzer.py):
        - 누끼 점수: 흰배경 90%+ = 50점, 70%+ = 40점, 밝은배경 = 30점
        - 텍스트 점수: 없음 = +30점, 적음 = +10점, 많음 = -30점
        - 중앙 객체: 있음 = +20점

        총점 높은 썸네일을 대표 이미지로 자동 선택
        """
        if not self.data:
            messagebox.showwarning("경고", "먼저 데이터를 로드하세요")
            return

        try:
            from thumbnail_analyzer import ThumbnailAnalyzer
            has_analyzer = True
        except ImportError:
            has_analyzer = False
            messagebox.showerror("오류", "thumbnail_analyzer.py가 필요합니다.\n\n필요 패키지: pip install opencv-python easyocr pillow")
            return

        # 전체/페이지 선택
        total_count = len(self.data)
        start_idx = self.current_page * self.page_size
        end_idx = min(start_idx + self.page_size, total_count)
        page_count = end_idx - start_idx

        choice_win = tk.Toplevel(self.root)
        choice_win.title("분석 범위 선택")
        choice_win.geometry("350x150")
        choice_win.transient(self.root)
        choice_win.grab_set()

        ttk.Label(choice_win, text="썸네일 분석 범위를 선택하세요", font=("맑은 고딕", 10, "bold")).pack(pady=10)

        result_var = tk.StringVar(value="")

        btn_frame = ttk.Frame(choice_win)
        btn_frame.pack(pady=10)

        def select_page():
            result_var.set("page")
            choice_win.destroy()

        def select_all():
            result_var.set("all")
            choice_win.destroy()

        ttk.Button(btn_frame, text=f"현재 페이지만 ({page_count}개)", command=select_page, width=20).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text=f"전체 분석 ({total_count}개)", command=select_all, width=20).pack(side=tk.LEFT, padx=5)

        ttk.Label(choice_win, text="※ 전체 분석은 시간이 오래 걸릴 수 있습니다", foreground="gray").pack(pady=5)

        choice_win.wait_window()

        if not result_var.get():
            return

        # 분석 대상 결정
        if result_var.get() == "page":
            analyze_data = self.data[start_idx:end_idx]
            analyze_indices = list(range(start_idx, end_idx))
        else:
            analyze_data = self.data
            analyze_indices = list(range(len(self.data)))

        # 진행 다이얼로그
        progress = tk.Toplevel(self.root)
        progress.title("썸네일 분석 중...")
        progress.geometry("450x180")
        progress.transient(self.root)

        progress_var = tk.StringVar(value="분석 준비 중...")
        ttk.Label(progress, textvariable=progress_var, font=("맑은 고딕", 10)).pack(pady=10)

        detail_var = tk.StringVar(value="")
        ttk.Label(progress, textvariable=detail_var, font=("맑은 고딕", 9)).pack(pady=5)

        pb = ttk.Progressbar(progress, length=400, mode='determinate')
        pb.pack(pady=10)

        cancel_var = tk.BooleanVar(value=False)
        ttk.Button(progress, text="취소", command=lambda: cancel_var.set(True)).pack(pady=5)
        progress.update()

        total = len(analyze_data)
        stats = {"analyzed": 0, "changed": 0, "best_scores": []}

        def analyze_thread():
            analyzer = ThumbnailAnalyzer()

            for i, item in enumerate(analyze_data):
                if cancel_var.get():
                    break

                idx = analyze_indices[i]  # 실제 인덱스
                product_name = item.get("product_name", "")[:20]
                all_thumbs = item.get("all_thumbnails", [])
                current_thumb = item.get("thumbnail_url", "")

                # all_thumbnails가 없으면 현재 썸네일만 사용
                if not all_thumbs and current_thumb:
                    all_thumbs = [current_thumb]

                self.root.after(0, lambda cur=i, n=product_name, c=len(all_thumbs): [
                    progress_var.set(f"분석 중... {cur+1}/{total}"),
                    detail_var.set(f"{n}... ({c}개 썸네일)"),
                    pb.config(value=(cur+1)/total*100)
                ])

                if not all_thumbs:
                    continue

                # 모든 썸네일 분석
                best_idx, best_score, action = analyzer.get_best_thumbnail(all_thumbs)

                if best_score:
                    # 분석 결과 저장
                    item["thumb_score"] = best_score.total_score
                    item["thumb_nukki"] = best_score.is_nukki
                    item["thumb_text"] = best_score.has_text
                    item["thumb_action"] = action
                    item["_best_thumb_idx"] = best_idx

                    stats["best_scores"].append(best_score.total_score)
                    stats["analyzed"] += 1

                    # 최적 썸네일이 현재와 다르면 변경
                    best_url = all_thumbs[best_idx]
                    if best_url != current_thumb:
                        item["thumbnail_url"] = best_url
                        item["_thumb_changed"] = True
                        stats["changed"] += 1

                    # 이미지 로드 및 UI 업데이트
                    def load_and_update(i=idx, url=best_url, opt_url=item.get("option_image_url", "")):
                        try:
                            # 썸네일 이미지
                            response = requests.get(url, timeout=2)
                            if response.status_code == 200:
                                img = Image.open(BytesIO(response.content))
                                img = img.resize((75, 75), Image.Resampling.LANCZOS)
                                photo = ImageTk.PhotoImage(img)

                                it = self.data[i]
                                if "_thumb_label" in it:
                                    it["_thumb_label"].config(image=photo, text="")
                                    it["_thumb_label"].image = photo

                            # 옵션 이미지
                            if opt_url:
                                response2 = requests.get(opt_url, timeout=2)
                                if response2.status_code == 200:
                                    img2 = Image.open(BytesIO(response2.content))
                                    img2 = img2.resize((75, 75), Image.Resampling.LANCZOS)
                                    photo2 = ImageTk.PhotoImage(img2)

                                    it = self.data[i]
                                    if "_opt_img_label" in it:
                                        it["_opt_img_label"].config(image=photo2, text="")
                                        it["_opt_img_label"].image = photo2
                        except:
                            pass

                    self.root.after(0, load_and_update)

            def finish():
                progress.destroy()
                self._render_data()

                avg_score = sum(stats["best_scores"]) / len(stats["best_scores"]) if stats["best_scores"] else 0
                msg = f"썸네일 분석 완료!\n\n"
                msg += f"분석: {stats['analyzed']}개 상품\n"
                msg += f"변경: {stats['changed']}개 (더 좋은 썸네일 발견)\n"
                msg += f"평균 점수: {avg_score:.1f}점\n\n"
                msg += "점수 기준:\n"
                msg += "• 50점+ = 완벽 (흰배경 누끼, 텍스트 없음)\n"
                msg += "• 30~50점 = 양호\n"
                msg += "• 30점 미만 = 주의 필요"

                messagebox.showinfo("분석 완료", msg)

            self.root.after(0, finish)

        import threading
        threading.Thread(target=analyze_thread, daemon=True).start()

    def _analyze_ip_words(self):
        """
        지재권 의심 단어 분석

        1. 형태소 분석으로 의심 단어 추출 (일반명사 제외)
        2. AI로 실제 지재권 여부 확인
        3. 확인된 단어 DB에 추가
        """
        if not self.data:
            messagebox.showwarning("경고", "먼저 데이터를 로드하세요")
            return

        # bulsaja_common에서 함수 가져오기
        try:
            from bulsaja_common import (
                analyze_products_for_ip,
                verify_ip_words_with_ai,
                load_ip_words,
                add_ip_words
            )
        except ImportError as e:
            messagebox.showerror("오류", f"bulsaja_common 모듈 로드 실패: {e}")
            return

        # 결과 창
        result_window = tk.Toplevel(self.root)
        result_window.title("지재권 분석")
        result_window.geometry("700x600")
        result_window.transient(self.root)

        # 상단 프레임
        top_frame = ttk.Frame(result_window, padding=10)
        top_frame.pack(fill=tk.X)

        status_var = tk.StringVar(value="분석 준비 중...")
        ttk.Label(top_frame, textvariable=status_var, font=("맑은 고딕", 11, "bold")).pack(anchor=tk.W)

        # 진행바
        pb = ttk.Progressbar(top_frame, length=650, mode='determinate')
        pb.pack(pady=10, fill=tk.X)

        # 결과 영역
        result_frame = ttk.LabelFrame(result_window, text="의심 단어 목록", padding=10)
        result_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        # 트리뷰
        columns = ("word", "count", "type", "status")
        tree = ttk.Treeview(result_frame, columns=columns, show="headings", height=15)
        tree.heading("word", text="단어")
        tree.heading("count", text="출현횟수")
        tree.heading("type", text="유형")
        tree.heading("status", text="상태")

        tree.column("word", width=150)
        tree.column("count", width=80)
        tree.column("type", width=150)
        tree.column("status", width=100)

        scrollbar = ttk.Scrollbar(result_frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)

        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # 버튼 프레임
        btn_frame = ttk.Frame(result_window, padding=10)
        btn_frame.pack(fill=tk.X)

        def verify_with_ai():
            """선택된 단어를 AI로 검증"""
            selected = tree.selection()
            if not selected:
                # 전체 검증
                words = [tree.item(item)['values'][0] for item in tree.get_children()]
            else:
                words = [tree.item(item)['values'][0] for item in selected]

            if not words:
                return

            status_var.set(f"AI 검증 중... ({len(words)}개 단어)")
            result_window.update()

            def verify_thread():
                result = verify_ip_words_with_ai(words)

                def update_ui():
                    for item in tree.get_children():
                        word = tree.item(item)['values'][0]
                        if word in result['ip_confirmed']:
                            tree.set(item, "status", "⚠️ 지재권")
                        elif word in result['ip_safe']:
                            tree.set(item, "status", "✅ 안전")
                        elif word in result['ip_uncertain']:
                            tree.set(item, "status", "❓ 불확실")

                    status_var.set(f"검증 완료: 지재권 {len(result['ip_confirmed'])}개, 안전 {len(result['ip_safe'])}개")

                self.root.after(0, update_ui)

            threading.Thread(target=verify_thread, daemon=True).start()

        def add_to_db():
            """지재권 확인된 단어를 DB에 추가"""
            ip_words = []
            for item in tree.get_children():
                values = tree.item(item)['values']
                if values[3] == "⚠️ 지재권":
                    ip_words.append(values[0])

            if not ip_words:
                messagebox.showinfo("알림", "추가할 지재권 단어가 없습니다.\n먼저 AI 검증을 실행하세요.")
                return

            if add_ip_words(ip_words, 'brands'):
                messagebox.showinfo("완료", f"{len(ip_words)}개 단어가 지재권 DB에 추가되었습니다.")
            else:
                messagebox.showerror("오류", "DB 저장 실패")

        ttk.Button(btn_frame, text="🤖 AI 검증 (선택/전체)", command=verify_with_ai).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="💾 지재권 DB에 추가", command=add_to_db).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="닫기", command=result_window.destroy).pack(side=tk.RIGHT, padx=5)

        # 분석 실행
        def analyze_thread():
            def log(msg):
                self.root.after(0, lambda: status_var.set(msg))

            # 상품 데이터 준비
            products = []
            for item in self.data:
                products.append({
                    'product_name': item.get('product_name', ''),
                    'product_id': item.get('product_id', '')
                })

            # 분석 실행
            result = analyze_products_for_ip(products, log_callback=log)

            # 결과 표시
            def update_tree():
                for word, count in result['suspicious_words'].items():
                    # 이미 DB에 있는지 확인
                    ip_db = load_ip_words()
                    all_ip = ip_db.get('brands', []) + ip_db.get('characters', []) + ip_db.get('trademarks', [])

                    if word in all_ip:
                        status = "⚠️ 지재권(DB)"
                    elif word in ip_db.get('safe_words', []):
                        status = "✅ 안전(DB)"
                    else:
                        status = "❓ 미확인"

                    # 유형 결정
                    word_type = "영어" if word.isascii() else "한글(외래어)"

                    tree.insert("", tk.END, values=(word, count, word_type, status))

                pb['value'] = 100
                status_var.set(f"분석 완료: {len(result['suspicious_words'])}개 의심 단어 발견")

            self.root.after(0, update_tree)

        threading.Thread(target=analyze_thread, daemon=True).start()

    def _load_image(self, url, label, width, height):
        """이미지 로드"""
        try:
            if url in self.image_cache:
                photo = self.image_cache[url]
            else:
                response = requests.get(url, timeout=2)
                img = Image.open(BytesIO(response.content))
                img = img.resize((width, height), Image.Resampling.LANCZOS)
                photo = ImageTk.PhotoImage(img)
                self.image_cache[url] = photo

            label.config(image=photo, text="")
            label.image = photo
        except:
            pass

    def _save_changes(self):
        """변경사항 저장"""
        changes = []
        for row_idx, selected in self.selected_options.items():
            if row_idx < len(self.data):
                original = self.data[row_idx].get("selected", "A")
                if selected != original:
                    changes.append(f"Row {row_idx}: {original} -> {selected}")

        if changes:
            msg = f"변경: {len(changes)}개\n" + "\n".join(changes[:15])
            if len(changes) > 15:
                msg += f"\n... +{len(changes)-15}개"
            messagebox.showinfo("변경사항", msg)
        else:
            messagebox.showinfo("변경사항", "변경 없음")

    def _update_bulsaja(self):
        """불사자 업데이트 - 선택된 옵션을 대표상품으로 변경"""
        if not BULSAJA_API_AVAILABLE:
            messagebox.showerror("오류", "bulsaja_common 모듈을 찾을 수 없습니다.\n\npip install websocket-client 후 재시도")
            return

        if not self.data:
            messagebox.showwarning("경고", "먼저 데이터를 로드하세요")
            return

        # 변경된 항목 수집
        changes = []
        for item in self.data:
            row_idx = item["row_idx"]
            product_id = item.get("product_id", "")
            original_selected = item.get("selected", "A")
            current_selected = self.selected_options.get(row_idx, "A")

            if not product_id:
                continue

            # 선택이 변경된 경우만 추가
            if current_selected != original_selected:
                changes.append({
                    "product_id": product_id,
                    "product_name": item.get("product_name", ""),
                    "old_option": original_selected,
                    "new_option": current_selected,
                    "options": item.get("options", [])
                })

        if not changes:
            messagebox.showinfo("알림", "변경된 옵션이 없습니다.\n\n옵션을 클릭하여 대표상품을 변경하세요.")
            return

        # 확인 메시지
        msg = f"총 {len(changes)}개 상품의 대표옵션을 변경합니다.\n\n"
        for c in changes[:5]:
            msg += f"• {c['product_name'][:15]}... : {c['old_option']} → {c['new_option']}\n"
        if len(changes) > 5:
            msg += f"... +{len(changes) - 5}개"

        if not messagebox.askyesno("불사자 업데이트 확인", msg):
            return

        # 진행 다이얼로그
        progress_dialog = tk.Toplevel(self.root)
        progress_dialog.title("불사자 업데이트 중...")
        progress_dialog.geometry("400x200")
        progress_dialog.transient(self.root)
        progress_dialog.grab_set()

        status_var = tk.StringVar(value="토큰 추출 중...")
        ttk.Label(progress_dialog, textvariable=status_var, font=("맑은 고딕", 10)).pack(pady=20)
        progress_bar = ttk.Progressbar(progress_dialog, length=350, mode='determinate')
        progress_bar.pack(pady=10)

        log_text = tk.Text(progress_dialog, height=5, width=50, font=("Consolas", 9))
        log_text.pack(pady=10, padx=10, fill=tk.BOTH, expand=True)

        def log(msg):
            log_text.insert(tk.END, msg + "\n")
            log_text.see(tk.END)
            progress_dialog.update()

        def run_update():
            # 1. 토큰 추출
            log("🔑 크롬에서 토큰 추출 중...")
            success, access_token, refresh_token, error = extract_tokens_from_browser()

            if not success:
                log(f"❌ 토큰 추출 실패: {error}")
                status_var.set("토큰 추출 실패")
                messagebox.showerror("오류", f"토큰 추출 실패:\n{error}\n\n크롬에서 불사자 사이트에 로그인되어 있어야 합니다.")
                progress_dialog.destroy()
                return

            log(f"✅ 토큰 추출 성공")

            # 2. API 클라이언트 생성
            api_client = BulsajaAPIClient(access_token, refresh_token)

            # 3. 각 상품 업데이트
            success_count = 0
            fail_count = 0
            total = len(changes)

            for i, change in enumerate(changes):
                product_id = change["product_id"]
                new_option_label = change["new_option"]
                options = change["options"]

                status_var.set(f"업데이트 중... ({i+1}/{total})")
                progress_bar['value'] = (i + 1) / total * 100
                progress_dialog.update()

                # 옵션 라벨로 인덱스 찾기 (A=0, B=1, ...)
                try:
                    new_option_idx = ord(new_option_label.upper()) - ord('A')
                except:
                    log(f"⚠️ {product_id}: 잘못된 옵션 라벨 '{new_option_label}'")
                    fail_count += 1
                    continue

                # 4. 상품 상세 정보 조회
                try:
                    detail = api_client.get_product_detail(product_id)
                except Exception as e:
                    log(f"❌ {product_id}: 상세 조회 실패 - {e}")
                    fail_count += 1
                    continue

                upload_skus = detail.get("uploadSkus", [])
                if not upload_skus:
                    log(f"⚠️ {product_id}: SKU 없음")
                    fail_count += 1
                    continue

                # 5. 대표상품 변경 (main_product 플래그)
                # 기존 8. bulsaja_simulator.py의 apply_changes() 로직과 동일하게 구현
                # 모든 옵션 main_product = False → 선택된 인덱스만 True
                target_label = f"{new_option_label.upper()}."  # "A.", "B.", "C." 등

                # 모든 옵션의 main_product를 false로
                for sku in upload_skus:
                    sku['main_product'] = False

                # 선택된 옵션을 main_product = True로 설정
                found_target = False
                if new_option_idx < len(upload_skus):
                    upload_skus[new_option_idx]['main_product'] = True
                    found_target = True
                else:
                    # 인덱스가 범위를 벗어나면 옵션명(text)으로 검색
                    for sku in upload_skus:
                        sku_text = sku.get('text', '') or ''
                        if sku_text.strip().startswith(target_label):
                            sku['main_product'] = True
                            found_target = True
                            break

                if not found_target:
                    log(f"⚠️ {product_id}: 옵션 {new_option_label} 찾지 못함")
                    fail_count += 1
                    continue

                # 6. 썸네일 변경 (옵션 활성화 시)
                update_thumbnail = self.update_thumbnail_var.get()
                upload_thumbnails = detail.get("uploadThumbnails", [])
                upload_sku_props = detail.get("uploadSkuProps", {})

                if update_thumbnail and upload_sku_props:
                    # uploadSkuProps에서 선택된 옵션의 이미지 URL 찾기
                    main_option = upload_sku_props.get("mainOption", {})
                    values = main_option.get("values", [])

                    # 선택된 옵션의 이미지 URL 찾기
                    option_image_url = None
                    for val in values:
                        val_name = val.get("name", "")
                        if val_name.strip().startswith(target_label):
                            option_image_url = val.get("imageUrl", "")
                            break

                    # 인덱스로도 시도 (라벨 매칭 실패 시)
                    if not option_image_url and new_option_idx < len(values):
                        option_image_url = values[new_option_idx].get("imageUrl", "")

                    # 썸네일 배열 업데이트 (옵션 이미지를 첫 번째로)
                    if option_image_url and option_image_url not in upload_thumbnails[:1]:
                        # 기존 배열에서 해당 이미지 제거 후 맨 앞에 추가
                        if option_image_url in upload_thumbnails:
                            upload_thumbnails.remove(option_image_url)
                        upload_thumbnails.insert(0, option_image_url)
                        log(f"🖼️ {product_id}: 썸네일 변경됨")

                # 7. API로 업데이트
                update_data = {"uploadSkus": upload_skus}
                if update_thumbnail and upload_thumbnails:
                    update_data["uploadThumbnails"] = upload_thumbnails

                success_result, msg = api_client.update_product_fields(product_id, update_data)

                if success_result:
                    log(f"✅ {product_id}: 옵션 {new_option_label}로 변경")
                    success_count += 1
                else:
                    log(f"❌ {product_id}: {msg}")
                    fail_count += 1

            # 결과 요약
            status_var.set(f"완료! 성공: {success_count}, 실패: {fail_count}")
            log(f"\n{'='*40}")
            log(f"📊 업데이트 완료: 성공 {success_count}개, 실패 {fail_count}개")

            messagebox.showinfo("완료", f"불사자 업데이트 완료\n\n✅ 성공: {success_count}개\n❌ 실패: {fail_count}개")

            # 원본 선택값 업데이트 (변경 완료된 것으로 반영)
            for change in changes:
                for item in self.data:
                    if item.get("product_id") == change["product_id"]:
                        item["selected"] = change["new_option"]
                        break

        # 스레드로 실행
        threading.Thread(target=run_update, daemon=True).start()


def main():
    root = tk.Tk()
    app = SimulatorGUIv3(root)
    root.mainloop()


if __name__ == "__main__":
    main()
