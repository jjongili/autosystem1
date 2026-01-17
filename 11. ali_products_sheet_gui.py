#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
구글 시트 매출 일일장부 기반 알리 상품 수집 GUI
- 날짜 범위 지정 (여러 월 자동 처리)
- 해외구매처 "알리" 필터링
- 피벗 집계 (상품명 + 옵션명 기준)
"""

import os
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime, timedelta
from collections import defaultdict
from dotenv import load_dotenv
import threading

# Google Sheets
import gspread
from google.oauth2.service_account import Credentials

# ========== 설정 ==========
load_dotenv()

CREDENTIALS_FILE = os.environ.get("SERVICE_ACCOUNT_JSON", r"C:\autosystem\web_system\autosms-466614-951e91617c69.json")
SALES_SHEET_ID = "1MHhu1GdvV1OGS8Wy3NxWOKuqFvgZpqgwn08kG70EDsY"

# ========== 데이터 수집 함수 ==========
def collect_from_sales_sheet(start_date, end_date, progress_callback=None):
    """구글 시트 매출 일일장부에서 알리 상품 수집 및 집계"""
    
    def update_progress(msg):
        if progress_callback:
            progress_callback(msg)
    
    try:
        update_progress("구글 시트 연결 중...")
        
        scope = [
            'https://spreadsheets.google.com/feeds',
            'https://www.googleapis.com/auth/drive'
        ]
        
        creds = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=scope)
        client = gspread.authorize(creds)
        
        sales_sheet = client.open_by_key(SALES_SHEET_ID)
        
        # 날짜 범위에 해당하는 월 목록 생성
        months_to_check = []
        current = start_date.replace(day=1)
        end_month = end_date.replace(day=1)
        
        while current <= end_month:
            months_to_check.append(current.month)
            # 다음 달로
            if current.month == 12:
                current = current.replace(year=current.year + 1, month=1)
            else:
                current = current.replace(month=current.month + 1)
        
        update_progress(f"{len(months_to_check)}개월 데이터 로드 중: {months_to_check}")
        
        # 각 월 시트에서 데이터 수집
        all_rows = []
        headers = None
        
        for month in months_to_check:
            tab_name = f"{month}월"
            try:
                update_progress(f"{tab_name} 시트 로드 중...")
                ws = sales_sheet.worksheet(tab_name)
                tab_data = ws.get_all_values()
                
                if len(tab_data) < 3:
                    continue
                
                if headers is None:
                    headers = tab_data[1]  # 2행이 헤더
                
                all_rows.extend(tab_data[2:])  # 3행부터 데이터
                
            except Exception as e:
                update_progress(f"{tab_name} 로드 실패: {e}")
                continue
        
        if not headers or len(all_rows) == 0:
            return None, "데이터가 없습니다"
        
        update_progress(f"총 {len(all_rows)}건 로드 완료, 필터링 중...")
        
        # 컬럼 인덱스 찾기
        def find_col(names):
            for name in names:
                for idx, h in enumerate(headers):
                    h_clean = h.replace('\n', '').replace('\r', '').replace(' ', '')
                    name_clean = name.replace(' ', '')
                    if name_clean in h_clean:
                        return idx
            return -1
        
        def col_letter_to_idx(letter):
            letter = letter.upper()
            result = 0
            for char in letter:
                result = result * 26 + (ord(char) - ord('A') + 1)
            return result - 1
        
        col_order_date = find_col(["주문일자"])
        col_order_status = find_col(["주문현황"])
        col_seller_code = find_col(["판매자상품코드", "상품코드"])
        col_product_name = find_col(["상품명", "품명"])
        col_option_name = find_col(["옵션명", "옵션"])
        col_quantity = find_col(["수량", "주문수량"])
        col_payment = find_col(["실결제금액(배송비포함)", "실결제금액"])
        col_overseas_seller = find_col(["해외구매처", "구매처"])
        
        # 직접 지정
        if col_order_status < 0: col_order_status = col_letter_to_idx('D')
        if col_order_date < 0: col_order_date = col_letter_to_idx('G')
        if col_seller_code < 0: col_seller_code = col_letter_to_idx('J')
        if col_product_name < 0: col_product_name = col_letter_to_idx('K')
        if col_option_name < 0: col_option_name = col_letter_to_idx('L')
        if col_quantity < 0: col_quantity = col_letter_to_idx('T')
        if col_payment < 0: col_payment = col_letter_to_idx('X')
        if col_overseas_seller < 0: col_overseas_seller = col_letter_to_idx('AK')
        
        # 피벗 집계 (상품명 + 옵션명 기준)
        pivot_data = defaultdict(lambda: {
            "판매자상품코드": "",
            "판매건수": 0,
            "총수량": 0,
            "총금액": 0,
            "정상건수": 0,
            "취소건수": 0
        })
        
        filtered_count = 0
        
        for row in all_rows:
            if len(row) <= max(col_product_name, col_quantity, col_payment, col_overseas_seller):
                continue
            
            # 주문일자 체크
            try:
                date_str = row[col_order_date].strip() if col_order_date < len(row) else ""
                if len(date_str) >= 10:
                    order_date = datetime.strptime(date_str[:10], "%Y-%m-%d").date()
                    if order_date < start_date.date() or order_date > end_date.date():
                        continue
            except:
                continue
            
            # 해외구매처 체크
            overseas_seller = row[col_overseas_seller].strip() if col_overseas_seller < len(row) else ""
            
            if "알리" not in overseas_seller:
                continue
            
            filtered_count += 1
            
            # 데이터 추출
            seller_code = row[col_seller_code].strip() if col_seller_code < len(row) else ""
            product_name = row[col_product_name].strip() if col_product_name < len(row) else ""
            option_name = row[col_option_name].strip() if col_option_name < len(row) else ""
            order_status = row[col_order_status].strip() if col_order_status < len(row) else ""
            
            # 수량 파싱
            quantity = 0
            if col_quantity < len(row):
                try:
                    qty_str = row[col_quantity].replace(",", "").strip()
                    if qty_str:
                        quantity = int(float(qty_str))
                except:
                    pass
            
            # 금액 파싱
            payment = 0
            if col_payment < len(row):
                try:
                    pay_str = row[col_payment].replace(",", "").replace("원", "").replace("₩", "").strip()
                    if pay_str:
                        payment = int(float(pay_str))
                except:
                    pass
            
            # 키 생성 (상품명 + 옵션명)
            key = f"{product_name}|||{option_name}"
            
            # 집계
            pivot_data[key]["판매자상품코드"] = seller_code
            pivot_data[key]["판매건수"] += 1
            pivot_data[key]["총수량"] += quantity
            pivot_data[key]["총금액"] += payment
            
            # 주문 상태별 카운트
            is_cancel = any(x in order_status for x in ["취소", "반품", "환불"])
            if is_cancel:
                pivot_data[key]["취소건수"] += 1
            else:
                pivot_data[key]["정상건수"] += 1
        
        # 결과 리스트로 변환
        result = []
        for key, data in pivot_data.items():
            product_name, option_name = key.split("|||")
            result.append({
                "상품명": product_name,
                "옵션명": option_name,
                "판매자상품코드": data["판매자상품코드"],
                "판매건수": data["판매건수"],
                "총수량": data["총수량"],
                "총금액": data["총금액"],
                "정상건수": data["정상건수"],
                "취소건수": data["취소건수"]
            })
        
        # 판매건수 기준 내림차순 정렬
        result.sort(key=lambda x: x["판매건수"], reverse=True)
        
        update_progress(f"완료! {filtered_count}건 필터링, {len(result)}개 상품 집계")
        
        return result, None
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return None, f"오류: {str(e)}"

# ========== GUI ==========
class AliProductsSheetGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("알리 상품 수집 및 분석 (구글 시트)")
        self.root.geometry("1400x850")
        
        self.data = []
        
        self.create_widgets()
    
    def create_widgets(self):
        # ========== 날짜 선택 프레임 ==========
        date_frame = ttk.LabelFrame(self.root, text="📅 날짜 범위", padding=10)
        date_frame.pack(fill=tk.X, padx=10, pady=5)
        
        # 시작일
        ttk.Label(date_frame, text="시작일:", font=("맑은 고딕", 10, "bold")).pack(side=tk.LEFT, padx=5)
        
        self.start_year = tk.IntVar(value=datetime.now().year)
        self.start_month = tk.IntVar(value=datetime.now().month)
        self.start_day = tk.IntVar(value=1)
        
        ttk.Spinbox(date_frame, from_=2020, to=2030, textvariable=self.start_year, width=8).pack(side=tk.LEFT, padx=2)
        ttk.Label(date_frame, text="년").pack(side=tk.LEFT)
        ttk.Spinbox(date_frame, from_=1, to=12, textvariable=self.start_month, width=5).pack(side=tk.LEFT, padx=2)
        ttk.Label(date_frame, text="월").pack(side=tk.LEFT)
        ttk.Spinbox(date_frame, from_=1, to=31, textvariable=self.start_day, width=5).pack(side=tk.LEFT, padx=2)
        ttk.Label(date_frame, text="일").pack(side=tk.LEFT, padx=10)
        
        # 종료일
        ttk.Label(date_frame, text="종료일:", font=("맑은 고딕", 10, "bold")).pack(side=tk.LEFT, padx=5)
        
        self.end_year = tk.IntVar(value=datetime.now().year)
        self.end_month = tk.IntVar(value=datetime.now().month)
        self.end_day = tk.IntVar(value=datetime.now().day)
        
        ttk.Spinbox(date_frame, from_=2020, to=2030, textvariable=self.end_year, width=8).pack(side=tk.LEFT, padx=2)
        ttk.Label(date_frame, text="년").pack(side=tk.LEFT)
        ttk.Spinbox(date_frame, from_=1, to=12, textvariable=self.end_month, width=5).pack(side=tk.LEFT, padx=2)
        ttk.Label(date_frame, text="월").pack(side=tk.LEFT)
        ttk.Spinbox(date_frame, from_=1, to=31, textvariable=self.end_day, width=5).pack(side=tk.LEFT, padx=2)
        ttk.Label(date_frame, text="일").pack(side=tk.LEFT)
        
        # 빠른 선택
        ttk.Label(date_frame, text="빠른 선택:", font=("맑은 고딕", 9)).pack(side=tk.LEFT, padx=20)
        ttk.Button(date_frame, text="이번달", command=self.set_this_month).pack(side=tk.LEFT, padx=2)
        ttk.Button(date_frame, text="지난달", command=self.set_last_month).pack(side=tk.LEFT, padx=2)
        ttk.Button(date_frame, text="최근 30일", command=self.set_last_30days).pack(side=tk.LEFT, padx=2)
        
        # ========== 컨트롤 프레임 ==========
        control_frame = ttk.Frame(self.root, padding=10)
        control_frame.pack(fill=tk.X)
        
        # 수집 버튼
        self.collect_btn = ttk.Button(
            control_frame,
            text="🔍 데이터 수집",
            command=self.start_collection
        )
        self.collect_btn.pack(side=tk.LEFT, padx=10)
        
        # 엑셀 저장 버튼
        self.save_btn = ttk.Button(
            control_frame,
            text="💾 엑셀 저장",
            command=self.save_to_excel,
            state=tk.DISABLED
        )
        self.save_btn.pack(side=tk.LEFT, padx=5)
        
        # 상태 표시
        self.status_label = ttk.Label(control_frame, text="날짜 범위를 지정한 후 '데이터 수집'을 클릭하세요", foreground="gray")
        self.status_label.pack(side=tk.LEFT, padx=20)
        
        # ========== 통계 프레임 ==========
        stats_frame = ttk.LabelFrame(self.root, text="📊 통계", padding=10)
        stats_frame.pack(fill=tk.X, padx=10, pady=5)
        
        self.stats_text = tk.Text(stats_frame, height=3, font=("맑은 고딕", 9), bg="#f0f0f0")
        self.stats_text.pack(fill=tk.X)
        self.stats_text.insert("1.0", "데이터를 수집하면 통계가 표시됩니다.")
        self.stats_text.config(state=tk.DISABLED)
        
        # ========== 필터 프레임 ==========
        filter_frame = ttk.LabelFrame(self.root, text="🔍 검색 필터", padding=10)
        filter_frame.pack(fill=tk.X, padx=10, pady=5)
        
        ttk.Label(filter_frame, text="상품명/옵션명 검색:", font=("맑은 고딕", 9)).pack(side=tk.LEFT, padx=5)
        
        self.filter_var = tk.StringVar()
        self.filter_var.trace("w", lambda *args: self.apply_filter())
        
        filter_entry = ttk.Entry(filter_frame, textvariable=self.filter_var, width=40)
        filter_entry.pack(side=tk.LEFT, padx=5)
        
        ttk.Button(filter_frame, text="❌ 초기화", command=self.clear_filter).pack(side=tk.LEFT, padx=5)
        
        self.filter_status = ttk.Label(filter_frame, text="", foreground="blue")
        self.filter_status.pack(side=tk.LEFT, padx=10)
        
        # ========== 테이블 프레임 ==========
        table_frame = ttk.LabelFrame(self.root, text="🛒 알리 상품 집계 (판매건수 순)", padding=10)
        table_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        # 스크롤바
        scroll_y = ttk.Scrollbar(table_frame, orient=tk.VERTICAL)
        scroll_y.pack(side=tk.RIGHT, fill=tk.Y)
        
        scroll_x = ttk.Scrollbar(table_frame, orient=tk.HORIZONTAL)
        scroll_x.pack(side=tk.BOTTOM, fill=tk.X)
        
        # Treeview
        columns = ("순위", "상품명", "옵션명", "판매자상품코드", "판매건수", "총수량", "총금액", "정상", "취소")
        
        self.tree = ttk.Treeview(
            table_frame,
            columns=columns,
            show="headings",
            yscrollcommand=scroll_y.set,
            xscrollcommand=scroll_x.set,
            height=20
        )
        
        scroll_y.config(command=self.tree.yview)
        scroll_x.config(command=self.tree.xview)
        
        # 컬럼 설정
        self.tree.heading("순위", text="순위")
        self.tree.heading("상품명", text="상품명 ▲▼", command=lambda: self.sort_column("상품명", False))
        self.tree.heading("옵션명", text="옵션명 ▲▼", command=lambda: self.sort_column("옵션명", False))
        self.tree.heading("판매자상품코드", text="판매자상품코드 ▲▼", command=lambda: self.sort_column("판매자상품코드", False))
        self.tree.heading("판매건수", text="판매건수 ▲▼", command=lambda: self.sort_column("판매건수", False))
        self.tree.heading("총수량", text="총수량 ▲▼", command=lambda: self.sort_column("총수량", False))
        self.tree.heading("총금액", text="총금액 ▲▼", command=lambda: self.sort_column("총금액", False))
        self.tree.heading("정상", text="정상 ▲▼", command=lambda: self.sort_column("정상", False))
        self.tree.heading("취소", text="취소 ▲▼", command=lambda: self.sort_column("취소", False))
        
        self.tree.column("순위", width=50, anchor=tk.CENTER)
        self.tree.column("상품명", width=400, anchor=tk.W)
        self.tree.column("옵션명", width=200, anchor=tk.W)
        self.tree.column("판매자상품코드", width=150, anchor=tk.W)
        self.tree.column("판매건수", width=80, anchor=tk.E)
        self.tree.column("총수량", width=80, anchor=tk.E)
        self.tree.column("총금액", width=120, anchor=tk.E)
        self.tree.column("정상", width=60, anchor=tk.E)
        self.tree.column("취소", width=60, anchor=tk.E)
        
        self.tree.pack(fill=tk.BOTH, expand=True)
        
        # 테이블 스타일
        style = ttk.Style()
        style.configure("Treeview", rowheight=25, font=("맑은 고딕", 9))
        style.configure("Treeview.Heading", font=("맑은 고딕", 9, "bold"))
        
        # 정렬 상태 저장
        self.sort_reverse = {}
    
    def sort_column(self, col_name, reverse):
        """컬럼 클릭 시 정렬"""
        if col_name not in self.sort_reverse:
            self.sort_reverse[col_name] = False
        
        # 토글
        reverse = not self.sort_reverse[col_name]
        self.sort_reverse[col_name] = reverse
        
        # 컬럼 인덱스 매핑
        col_index_map = {
            "상품명": "상품명",
            "옵션명": "옵션명",
            "판매자상품코드": "판매자상품코드",
            "판매건수": "판매건수",
            "총수량": "총수량",
            "총금액": "총금액",
            "정상": "정상건수",
            "취소": "취소건수"
        }
        
        sort_key = col_index_map.get(col_name)
        if not sort_key:
            return
        
        # 데이터 정렬
        try:
            if sort_key in ["판매건수", "총수량", "총금액", "정상건수", "취소건수"]:
                # 숫자 정렬
                self.data.sort(key=lambda x: x[sort_key], reverse=reverse)
            else:
                # 문자열 정렬
                self.data.sort(key=lambda x: x[sort_key], reverse=reverse)
            
            # 테이블 다시 표시
            self.display_data()
            
        except Exception as e:
            print(f"정렬 오류: {e}")
    
    def apply_filter(self):
        """필터 적용"""
        if not hasattr(self, 'data') or not self.data:
            return
        
        filter_text = self.filter_var.get().strip().lower()
        
        if not filter_text:
            # 필터 없으면 전체 표시
            self.display_data()
            self.filter_status.config(text="")
            return
        
        # 필터링된 데이터
        filtered_data = []
        for item in self.data:
            product_name = item["상품명"].lower()
            option_name = item["옵션명"].lower()
            
            if filter_text in product_name or filter_text in option_name:
                filtered_data.append(item)
        
        # 테이블 표시
        self.display_filtered_data(filtered_data)
        self.filter_status.config(text=f"{len(filtered_data)}개 표시 (전체 {len(self.data)}개)")
    
    def clear_filter(self):
        """필터 초기화"""
        self.filter_var.set("")
        self.display_data()
        self.filter_status.config(text="")
    
    def display_filtered_data(self, filtered_data):
        """필터링된 데이터 표시"""
        # 기존 데이터 삭제
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        # 데이터 삽입
        for rank, item in enumerate(filtered_data, 1):
            values = (
                rank,
                item["상품명"][:60] + "..." if len(item["상품명"]) > 60 else item["상품명"],
                item["옵션명"][:30] + "..." if len(item["옵션명"]) > 30 else item["옵션명"],
                item["판매자상품코드"],
                f"{item['판매건수']:,}",
                f"{item['총수량']:,}",
                f"{item['총금액']:,}원",
                f"{item['정상건수']:,}",
                f"{item['취소건수']:,}"
            )
            
            self.tree.insert("", tk.END, values=values)
    
    def set_this_month(self):
        """이번달 설정"""
        now = datetime.now()
        self.start_year.set(now.year)
        self.start_month.set(now.month)
        self.start_day.set(1)
        self.end_year.set(now.year)
        self.end_month.set(now.month)
        self.end_day.set(now.day)
    
    def set_last_month(self):
        """지난달 설정"""
        now = datetime.now()
        first_day = now.replace(day=1)
        last_month = first_day - timedelta(days=1)
        
        self.start_year.set(last_month.year)
        self.start_month.set(last_month.month)
        self.start_day.set(1)
        self.end_year.set(last_month.year)
        self.end_month.set(last_month.month)
        self.end_day.set(last_month.day)
    
    def set_last_30days(self):
        """최근 30일 설정"""
        now = datetime.now()
        start = now - timedelta(days=30)
        
        self.start_year.set(start.year)
        self.start_month.set(start.month)
        self.start_day.set(start.day)
        self.end_year.set(now.year)
        self.end_month.set(now.month)
        self.end_day.set(now.day)
    
    def start_collection(self):
        """데이터 수집 시작"""
        # 날짜 범위
        try:
            start_year = int(self.start_year.get())
            start_month = int(self.start_month.get())
            start_day = int(self.start_day.get())
            end_year = int(self.end_year.get())
            end_month = int(self.end_month.get())
            end_day = int(self.end_day.get())
            
            start_date = datetime(start_year, start_month, start_day)
            end_date = datetime(end_year, end_month, end_day)
        except (ValueError, TypeError) as e:
            messagebox.showerror("오류", f"유효하지 않은 날짜입니다: {e}")
            return
        
        if start_date > end_date:
            messagebox.showerror("오류", "시작일이 종료일보다 늦습니다")
            return
        
        self.collect_btn.config(state=tk.DISABLED)
        self.status_label.config(text=f"📥 데이터 수집 중...", foreground="blue")
        
        # 백그라운드 스레드에서 실행
        thread = threading.Thread(target=self.collect_data, args=(start_date, end_date), daemon=True)
        thread.start()
    
    def collect_data(self, start_date, end_date):
        """데이터 수집 (백그라운드)"""
        
        def progress_callback(msg):
            self.root.after(0, lambda: self.status_label.config(text=msg))
        
        result, error = collect_from_sales_sheet(start_date, end_date, progress_callback)
        
        # UI 업데이트는 메인 스레드에서
        self.root.after(0, lambda: self.on_collection_complete(result, error, start_date, end_date))
    
    def on_collection_complete(self, result, error, start_date, end_date):
        """수집 완료 후 처리"""
        self.collect_btn.config(state=tk.NORMAL)
        
        if error:
            messagebox.showerror("오류", error)
            self.status_label.config(text=f"❌ 오류: {error}", foreground="red")
            return
        
        if not result:
            messagebox.showinfo("알림", "데이터가 없습니다")
            self.status_label.config(text="⚠️ 데이터 없음", foreground="orange")
            return
        
        self.data = result
        self.display_data()
        self.update_stats(start_date, end_date)
        self.save_btn.config(state=tk.NORMAL)
        
        date_range = f"{start_date.strftime('%Y-%m-%d')} ~ {end_date.strftime('%Y-%m-%d')}"
        self.status_label.config(text=f"✅ 수집 완료! ({len(result)}개 상품)", foreground="green")
    
    def display_data(self):
        """데이터 테이블에 표시"""
        # 기존 데이터 삭제
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        # 데이터 삽입
        for rank, item in enumerate(self.data, 1):
            values = (
                rank,
                item["상품명"][:60] + "..." if len(item["상품명"]) > 60 else item["상품명"],
                item["옵션명"][:30] + "..." if len(item["옵션명"]) > 30 else item["옵션명"],
                item["판매자상품코드"],
                f"{item['판매건수']:,}",
                f"{item['총수량']:,}",
                f"{item['총금액']:,}원",
                f"{item['정상건수']:,}",
                f"{item['취소건수']:,}"
            )
            
            self.tree.insert("", tk.END, values=values)
    
    def update_stats(self, start_date, end_date):
        """통계 업데이트"""
        if not self.data:
            return
        
        total_products = len(self.data)
        total_orders = sum(item["판매건수"] for item in self.data)
        total_quantity = sum(item["총수량"] for item in self.data)
        total_amount = sum(item["총금액"] for item in self.data)
        total_normal = sum(item["정상건수"] for item in self.data)
        total_cancel = sum(item["취소건수"] for item in self.data)
        
        date_range = f"{start_date.strftime('%Y-%m-%d')} ~ {end_date.strftime('%Y-%m-%d')}"
        
        stats_text = f"""
📅 기간: {date_range} | 총 {total_products:,}개 상품 | 총 {total_orders:,}건 주문 | 총 {total_quantity:,}개 | 총 {total_amount:,}원
✅ 정상: {total_normal:,}건 ({total_normal/total_orders*100:.1f}%) | ❌ 취소/반품: {total_cancel:,}건 ({total_cancel/total_orders*100:.1f}%)
        """
        
        self.stats_text.config(state=tk.NORMAL)
        self.stats_text.delete("1.0", tk.END)
        self.stats_text.insert("1.0", stats_text.strip())
        self.stats_text.config(state=tk.DISABLED)
    
    def save_to_excel(self):
        """엑셀 파일로 저장"""
        if not self.data:
            messagebox.showwarning("경고", "저장할 데이터가 없습니다")
            return
        
        # 파일명 생성
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_filename = f"알리상품집계_{timestamp}.xlsx"
        
        # 파일 저장 다이얼로그
        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile=default_filename,
            filetypes=[("Excel 파일", "*.xlsx"), ("모든 파일", "*.*")]
        )
        
        if not filepath:
            return
        
        try:
            # openpyxl로 저장
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "알리상품집계"
            
            # 헤더
            headers = ["순위", "상품명", "옵션명", "판매자상품코드", "판매건수", "총수량", "총금액", "정상건수", "취소건수"]
            ws.append(headers)
            
            # 헤더 스타일
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for cell in ws[1]:
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.font = Font(bold=True, size=11, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
            
            # 데이터
            for rank, item in enumerate(self.data, 1):
                ws.append([
                    rank,
                    item["상품명"],
                    item["옵션명"],
                    item["판매자상품코드"],
                    item["판매건수"],
                    item["총수량"],
                    item["총금액"],
                    item["정상건수"],
                    item["취소건수"]
                ])
                
                # 데이터 행 스타일
                row_num = rank + 1
                for col_num in range(1, 10):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.border = thin_border
                    
                    # 숫자 컬럼 오른쪽 정렬
                    if col_num >= 5:
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
            
            # 컬럼 너비 조정
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 60
            ws.column_dimensions['C'].width = 35
            ws.column_dimensions['D'].width = 25
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 15
            ws.column_dimensions['H'].width = 12
            ws.column_dimensions['I'].width = 12
            
            # 자동 필터 활성화 (제대로)
            max_row = len(self.data) + 1
            ws.auto_filter.ref = ws.dimensions
            
            # 창 고정 (헤더 행 고정)
            ws.freeze_panes = ws['A2']
            
            # 저장
            wb.save(filepath)
            
            messagebox.showinfo("저장 완료", f"파일이 저장되었습니다:\n{filepath}\n\n✅ 자동 필터 적용됨\n✅ 헤더 행 고정됨")
            
        except Exception as e:
            messagebox.showerror("저장 실패", f"엑셀 저장 중 오류:\n{str(e)}")

# ========== 메인 실행 ==========
if __name__ == "__main__":
    root = tk.Tk()
    app = AliProductsSheetGUI(root)
    root.mainloop()
