# -*- coding: utf-8 -*-
"""
불사자 시뮬레이터 v1.0

상품 업로드 없이 분석만 수행하여 학습 데이터 수집:
- 브랜드/위험상품 검사
- 미끼옵션 탐지
- 썸네일-옵션 매칭 분석
- 대표옵션 선택 시뮬레이션

결과를 엑셀로 저장하여 검토 및 학습에 활용

by 프코노미
"""

import os
import time
import threading
import json
from datetime import datetime
from typing import List, Dict, Tuple, Optional
from dataclasses import dataclass
import math

import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox, filedialog

# 공통 모듈 임포트
try:
    from bulsaja_common import (
        BulsajaAPIClient, extract_tokens_from_browser,
        load_banned_words, load_excluded_words, load_bait_keywords, save_bait_keywords,
        check_product_safety, filter_bait_options, match_thumbnail_to_sku,
        select_main_option,  # 상품명 기반 대표옵션 선택
        load_category_risk_settings, save_category_risk_settings,  # 카테고리 검수 설정
        DEFAULT_CATEGORY_RISK_SETTINGS,
        MARKET_IDS, DEFAULT_BAIT_KEYWORDS
    )
except ImportError:
    print("⚠️ bulsaja_common.py 모듈을 찾을 수 없습니다. 같은 폴더에 있는지 확인하세요.")
    exit(1)

# 엑셀 라이브러리
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    print("⚠️ openpyxl이 설치되지 않았습니다. pip install openpyxl")
    EXCEL_AVAILABLE = False

# ==================== 설정 ====================
CONFIG_FILE = "bulsaja_simulator_config.json"
DEBUG_PORT = 9222


def load_config():
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            pass
    return {}


def save_config(config):
    try:
        with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
        return True
    except:
        return False


# ==================== 시뮬레이터 클래스 ====================
class BulsajaSimulator:
    def __init__(self, gui):
        self.gui = gui
        self.api_client: Optional[BulsajaAPIClient] = None
        self.is_running = False

        # 키워드 로드
        self.banned_words, _ = load_banned_words()
        self.excluded_words = load_excluded_words()
        self.bait_keywords = load_bait_keywords()

        # 통계
        self.stats = {
            "total": 0,
            "safe": 0,
            "unsafe": 0,
            "bait_found": 0,
            "thumbnail_matched": 0,
            "new_bait_keywords": set(),
        }

        # 결과 데이터
        self.results = []

    def log(self, message):
        if self.gui:
            self.gui.log(message)
        else:
            print(message)

    def init_api_client(self, access_token: str, refresh_token: str) -> Tuple[bool, str, int]:
        self.api_client = BulsajaAPIClient(access_token, refresh_token)
        return self.api_client.test_connection()

    def analyze_product(self, product: Dict, option_count: int = 5, check_level: str = 'normal') -> Dict:
        """단일 상품 분석

        Args:
            product: 상품 정보
            option_count: 옵션 개수 제한
            check_level: 검수 레벨 (strict/normal/skip)
        """
        product_id = product.get('ID', '')
        product_name = product.get('uploadCommonProductName', '')

        result = {
            'id': product_id,
            'name': product_name,
            'is_safe': True,
            'unsafe_reason': '',
            'unsafe_keywords': [],
            'safe_context': '',  # 안전 컨텍스트로 무시된 키워드
            'check_level': check_level,  # 사용된 검수 레벨
            'ai_judgment': '',  # AI 판단 결과 (strict 모드)
            'total_options': 0,
            'valid_options': 0,
            'final_options': 0,  # 옵션 개수 제한 후
            'bait_options': 0,
            'bait_option_list': [],
            'detected_bait_keywords': [],
            'main_option_name': '',
            'main_option_method': '',  # 썸네일매칭 / 최저가
            'final_option_list': [],  # 최종 선택된 옵션들
            'thumbnail_url': '',  # 메인 썸네일
            'main_option_image': '',  # 대표 옵션 이미지
            'min_price_cny': 0,
            'max_price_cny': 0,
        }

        try:
            # 1. 상품명 안전 검사 (검수 레벨에 따라 AI 사용 여부 결정)
            safety = check_product_safety(product_name, self.excluded_words, check_level=check_level)
            result['is_safe'] = safety['is_safe']
            result['unsafe_keywords'] = safety['all_found']

            # 안전 컨텍스트로 무시된 키워드 기록
            if safety.get('safe_context_found'):
                result['safe_context'] = ', '.join(safety['safe_context_found'][:3])

            # AI 판단 결과 기록 (strict 모드)
            if safety.get('ai_judgment'):
                result['ai_judgment'] = ', '.join(safety['ai_judgment'][:3])

            if not safety['is_safe']:
                categories = []
                if safety['categories']['adult']:
                    categories.append(f"성인:{','.join(safety['categories']['adult'][:2])}")
                if safety['categories']['medical']:
                    categories.append(f"의료:{','.join(safety['categories']['medical'][:2])}")
                if safety['categories']['child']:
                    categories.append(f"유아:{','.join(safety['categories']['child'][:2])}")
                if safety['categories']['prohibited']:
                    categories.append(f"금지:{','.join(safety['categories']['prohibited'][:2])}")
                result['unsafe_reason'] = ' / '.join(categories)

            # 2. 상품 상세 정보 조회
            detail = self.api_client.get_product_detail(product_id)

            # 썸네일 URL
            thumbnails = detail.get('uploadThumbnails', [])
            if thumbnails:
                result['thumbnail_url'] = thumbnails[0]

            # SKU 정보
            upload_skus = detail.get('uploadSkus', [])
            if not upload_skus:
                upload_skus = detail.get('original_skus', [])

            result['total_options'] = len(upload_skus)

            if upload_skus:
                # 가격 범위
                prices = [sku.get('_origin_price', 0) for sku in upload_skus if sku.get('_origin_price', 0) > 0]
                if prices:
                    result['min_price_cny'] = min(prices)
                    result['max_price_cny'] = max(prices)

                # 3. 미끼옵션 필터링
                valid_skus, bait_skus = filter_bait_options(upload_skus, self.bait_keywords)

                result['valid_options'] = len(valid_skus)
                result['bait_options'] = len(bait_skus)

                # 미끼 옵션 정보 수집 (옵션명 + 중국어 + 가격)
                for bait_sku in bait_skus:
                    option_text_ko = bait_sku.get('text_ko', '') or ''
                    option_text_cn = bait_sku.get('text', '') or ''
                    bait_price = bait_sku.get('_origin_price', 0)
                    detected_keyword = bait_sku.get('_bait_keyword', '')
                    # 옵션명[중국어](가격) 형식으로 저장
                    display_text = option_text_ko[:20] if option_text_ko else option_text_cn[:20]
                    cn_part = f"[{option_text_cn[:15]}]" if option_text_cn and option_text_cn != option_text_ko else ""
                    price_part = f"({bait_price})" if bait_price else ""
                    result['bait_option_list'].append(f"{display_text}{cn_part}{price_part}")
                    if detected_keyword and detected_keyword not in result['detected_bait_keywords']:
                        result['detected_bait_keywords'].append(detected_keyword)

                # 4. 대표옵션 선택: 상품명 매칭 → 첫 번째 옵션 폴백
                if valid_skus:
                    main_sku_idx, main_method = select_main_option(product_name, valid_skus)
                    main_sku = valid_skus[main_sku_idx]
                    result['main_option_name'] = main_sku.get('text_ko', '') or main_sku.get('text', '')
                    result['main_option_method'] = main_method

                    # 대표 옵션 이미지 URL (urlRef가 불사자 변환 이미지)
                    main_option_img = main_sku.get('urlRef', '') or main_sku.get('image', '') or main_sku.get('img', '')
                    if main_option_img:
                        result['main_option_image'] = main_option_img

                    # 5. 옵션 개수 제한 (대표옵션 = 최저가, 나머지는 그보다 비싼 것)
                    main_sku_price = main_sku.get('_origin_price', 0)

                    if option_count > 0:
                        # 대표옵션 가격 이상인 옵션들만 필터링 후 가격순 정렬
                        eligible_skus = [
                            sku for sku in valid_skus
                            if sku.get('_origin_price', 0) >= main_sku_price
                        ]
                        # 가격 오름차순 정렬
                        eligible_skus.sort(key=lambda x: x.get('_origin_price', 0))
                        # option_count 개수만큼 선택
                        final_skus = eligible_skus[:option_count]
                    else:
                        final_skus = valid_skus

                    result['final_options'] = len(final_skus)

                    # 최종 옵션 목록
                    for sku in final_skus:
                        opt_name = sku.get('text_ko', '') or sku.get('text', '')
                        opt_price = sku.get('_origin_price', 0)
                        result['final_option_list'].append(f"{opt_name[:20]}({opt_price:.1f})")

        except Exception as e:
            result['unsafe_reason'] = f"분석오류: {str(e)[:50]}"

        return result

    def run_simulation(self, group_names: List[str], max_products: int, option_count: int, save_path: str,
                       check_level: str = 'normal', risk_categories: List[str] = None):
        """시뮬레이션 실행

        Args:
            group_names: 그룹명 목록
            max_products: 그룹당 최대 상품 수
            option_count: 옵션 개수 제한
            save_path: 저장 경로
            check_level: 기본 검수 레벨 (strict/normal/skip)
            risk_categories: 엄격 검수 적용할 카테고리 목록
        """
        self.is_running = True
        self.results = []
        self.stats = {
            "total": 0,
            "safe": 0,
            "unsafe": 0,
            "bait_found": 0,
            "thumbnail_matched": 0,
            "new_bait_keywords": set(),
            "ai_checked": 0,  # AI 검증 횟수
        }

        if risk_categories is None:
            risk_categories = []

        level_desc = {'strict': '엄격(AI확인)', 'normal': '보통(자동)', 'skip': '제외'}.get(check_level, check_level)

        self.log("")
        self.log("=" * 50)
        self.log("🔬 시뮬레이션 시작")
        self.log(f"   그룹: {', '.join(group_names)}")
        self.log(f"   최대 상품: {max_products}개/그룹")
        self.log(f"   옵션 개수: {option_count}개")
        self.log(f"   검수 수준: {level_desc}")
        if risk_categories:
            self.log(f"   위험 카테고리: {', '.join(risk_categories[:5])}...")
        self.log(f"   저장 경로: {save_path}")
        self.log("=" * 50)

        try:
            total_products = 0
            for group_idx, group_name in enumerate(group_names):
                if not self.is_running:
                    break

                self.log(f"\n📁 [{group_idx+1}/{len(group_names)}] {group_name} 분석 중...")

                # 상품 목록 조회
                products, total = self.api_client.get_products_by_group(group_name, 0, max_products)

                if not products:
                    self.log(f"   ⚠️ 상품 없음")
                    continue

                self.log(f"   📦 {len(products)}개 상품 발견")

                for prod_idx, product in enumerate(products):
                    if not self.is_running:
                        break

                    # 카테고리 확인하여 검수 레벨 결정
                    product_category = product.get('categoryPath', '') or product.get('category', '') or ''
                    product_check_level = check_level

                    # 위험 카테고리에 해당하면 엄격 검수
                    for risk_cat in risk_categories:
                        if risk_cat and risk_cat.lower() in product_category.lower():
                            product_check_level = 'strict'
                            break

                    # 분석
                    result = self.analyze_product(product, option_count, check_level=product_check_level)
                    result['group_name'] = group_name
                    result['category'] = product_category[:30]  # 카테고리 기록
                    self.results.append(result)

                    # 통계 업데이트
                    self.stats['total'] += 1
                    if result['is_safe']:
                        self.stats['safe'] += 1
                    else:
                        self.stats['unsafe'] += 1
                    if result['bait_options'] > 0:
                        self.stats['bait_found'] += 1
                    if result['main_option_method'] == '썸네일매칭':
                        self.stats['thumbnail_matched'] += 1
                    if result.get('ai_judgment'):
                        self.stats['ai_checked'] += 1

                    # 진행상황 로그 (10개마다)
                    if (prod_idx + 1) % 10 == 0:
                        self.log(f"   ... {prod_idx+1}/{len(products)} 완료")

                    # GUI 업데이트
                    total_products += 1
                    if self.gui:
                        self.gui.update_progress(total_products, len(group_names) * max_products)

                self.log(f"   ✅ {group_name} 완료")

            # 엑셀 저장
            if self.results and EXCEL_AVAILABLE:
                self.save_to_excel(save_path)
                self.log(f"\n📊 엑셀 저장 완료: {save_path}")

        except Exception as e:
            self.log(f"❌ 시뮬레이션 오류: {e}")

        finally:
            # 결과 요약
            self.log("")
            self.log("=" * 50)
            self.log("📊 시뮬레이션 결과")
            self.log(f"   전체 상품: {self.stats['total']}개")
            self.log(f"   안전 상품: {self.stats['safe']}개")
            self.log(f"   위험 상품: {self.stats['unsafe']}개")
            self.log(f"   미끼옵션 발견: {self.stats['bait_found']}개 상품")
            self.log(f"   썸네일 매칭: {self.stats['thumbnail_matched']}개 상품")
            self.log("=" * 50)

            self.is_running = False
            if self.gui:
                self.gui.on_finished()

    def save_to_excel(self, filepath: str):
        """결과를 엑셀로 저장 (깔끔한 형식 - 이미지 수식 포함)"""
        wb = Workbook()
        ws = wb.active
        ws.title = "분석결과"

        # 헤더 스타일
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 헤더 (모든 데이터 포함) - 썸네일/옵션이미지 가깝게 배치
        headers = [
            "썸네일\n이미지", "옵션\n이미지", "상품명", "안전여부", "위험사유", "안전\n컨텍스트", "검수레벨",
            "AI판단", "전체옵션", "유효옵션", "최종옵션", "미끼옵션", "미끼옵션목록",
            "탐지키워드", "대표옵션", "선택방식",
            "선택", "옵션명", "중국어\n옵션명", "그룹명"
        ]
        # 선택 컬럼 스타일 (노란색 배경 - 사용자 입력용)
        select_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border

        # 데이터 스타일
        unsafe_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
        safe_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
        wrap_align = Alignment(vertical="top", wrap_text=True)
        center_align = Alignment(horizontal="center", vertical="center")

        # 데이터 입력 - 썸네일/옵션이미지 가깝게 배치
        for row_idx, result in enumerate(self.results, 2):
            col = 1

            # 1. 썸네일 이미지 (IMAGE 수식)
            thumb_url = result.get('thumbnail_url', '')
            if thumb_url:
                ws.cell(row=row_idx, column=col, value=f'=IMAGE("{thumb_url}")')
            else:
                ws.cell(row=row_idx, column=col, value='')
            col += 1

            # 2. 옵션 이미지 (대표옵션 이미지) - 썸네일 바로 옆
            option_img = result.get('main_option_image', '')
            if option_img:
                ws.cell(row=row_idx, column=col, value=f'=IMAGE("{option_img}")')
            else:
                ws.cell(row=row_idx, column=col, value='')
            col += 1

            # 3. 상품명
            ws.cell(row=row_idx, column=col, value=result.get('name', '')[:50])
            col += 1

            # 4. 안전여부 (O/X)
            is_safe = result.get('is_safe', True)
            status_cell = ws.cell(row=row_idx, column=col, value='O' if is_safe else 'X')
            status_cell.alignment = center_align
            if is_safe:
                status_cell.fill = safe_fill
            else:
                status_cell.fill = unsafe_fill
            col += 1

            # 5. 위험사유
            ws.cell(row=row_idx, column=col, value=result.get('unsafe_reason', ''))
            col += 1

            # 6. 안전컨텍스트
            ws.cell(row=row_idx, column=col, value=result.get('safe_context', ''))
            col += 1

            # 7. 검수레벨
            ws.cell(row=row_idx, column=col, value=result.get('check_level', ''))
            col += 1

            # 8. AI판단
            ws.cell(row=row_idx, column=col, value=result.get('ai_judgment', ''))
            col += 1

            # 9. 전체옵션
            ws.cell(row=row_idx, column=col, value=result.get('total_options', 0))
            col += 1

            # 10. 유효옵션
            ws.cell(row=row_idx, column=col, value=result.get('valid_options', 0))
            col += 1

            # 11. 최종옵션
            ws.cell(row=row_idx, column=col, value=result.get('final_options', 0))
            col += 1

            # 12. 미끼옵션
            ws.cell(row=row_idx, column=col, value=result.get('bait_options', 0))
            col += 1

            # 13. 미끼옵션목록
            bait_cell = ws.cell(row=row_idx, column=col,
                               value=self._format_options_abc(result.get('bait_option_list', [])[:5]))
            bait_cell.alignment = wrap_align
            col += 1

            # 14. 탐지키워드
            ws.cell(row=row_idx, column=col, value=result.get('detected_keywords', ''))
            col += 1

            # 15. 대표옵션
            ws.cell(row=row_idx, column=col, value=result.get('main_option_name', ''))
            col += 1

            # 16. 선택방식 (main_option_method)
            ws.cell(row=row_idx, column=col, value=result.get('main_option_method', ''))
            col += 1

            # 17. 선택 (프로그램 추천 기본값 A, 사용자가 수정 가능)
            select_cell = ws.cell(row=row_idx, column=col, value='A')
            select_cell.alignment = center_align
            select_cell.fill = select_fill
            col += 1

            # 18. 옵션명 (A, B, C 형태로 정리)
            final_options = result.get('final_option_list', [])
            option_text = self._format_options_abc(final_options)
            ws.cell(row=row_idx, column=col, value=option_text).alignment = wrap_align
            col += 1

            # 19. 중국어 옵션명
            cn_options = result.get('cn_option_list', [])
            cn_text = self._format_options_abc(cn_options) if cn_options else ''
            ws.cell(row=row_idx, column=col, value=cn_text).alignment = wrap_align
            col += 1

            # 20. 그룹명
            ws.cell(row=row_idx, column=col, value=result.get('group_name', ''))
            col += 1

            # 테두리 적용
            for c in range(1, col):
                ws.cell(row=row_idx, column=c).border = border

        # 열 너비 조정 - 썸네일/옵션이미지 가깝게 배치된 순서
        ws.column_dimensions['A'].width = 15  # 썸네일
        ws.column_dimensions['B'].width = 15  # 옵션이미지 (썸네일 바로 옆)
        ws.column_dimensions['C'].width = 40  # 상품명
        ws.column_dimensions['D'].width = 8   # 안전여부
        ws.column_dimensions['E'].width = 20  # 위험사유
        ws.column_dimensions['F'].width = 12  # 안전컨텍스트
        ws.column_dimensions['G'].width = 8   # 검수레벨
        ws.column_dimensions['H'].width = 8   # AI판단
        ws.column_dimensions['I'].width = 8   # 전체옵션
        ws.column_dimensions['J'].width = 8   # 유효옵션
        ws.column_dimensions['K'].width = 8   # 최종옵션
        ws.column_dimensions['L'].width = 8   # 미끼옵션
        ws.column_dimensions['M'].width = 30  # 미끼옵션목록
        ws.column_dimensions['N'].width = 15  # 탐지키워드
        ws.column_dimensions['O'].width = 25  # 대표옵션
        ws.column_dimensions['P'].width = 12  # 선택방식
        ws.column_dimensions['Q'].width = 6   # 선택
        ws.column_dimensions['R'].width = 35  # 옵션명
        ws.column_dimensions['S'].width = 35  # 중국어옵션명
        ws.column_dimensions['T'].width = 12  # 그룹명

        # 행 높이 조정 (이미지 표시용)
        for row_idx in range(2, len(self.results) + 2):
            ws.row_dimensions[row_idx].height = 80

        # 헤더 행 높이
        ws.row_dimensions[1].height = 40

        # 필터 설정
        ws.auto_filter.ref = ws.dimensions

        # === 상세 시트 (기존 형식) ===
        ws_detail = wb.create_sheet("상세정보")
        detail_headers = [
            "그룹", "불사자ID", "상품명", "안전여부", "위험사유", "안전컨텍스트",
            "전체옵션", "유효옵션", "최종옵션", "미끼옵션", "미끼옵션목록",
            "선택", "대표옵션", "최저가(CNY)", "최고가(CNY)", "최종옵션목록", "메인썸네일URL", "옵션이미지URL"
        ]
        for col, header in enumerate(detail_headers, 1):
            cell = ws_detail.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border

        for row_idx, result in enumerate(self.results, 2):
            ws_detail.cell(row=row_idx, column=1, value=result.get('group_name', '')).border = border
            ws_detail.cell(row=row_idx, column=2, value=result.get('id', '')).border = border
            ws_detail.cell(row=row_idx, column=3, value=result.get('name', '')[:50]).border = border

            # 안전여부 색상 적용
            safe_cell = ws_detail.cell(row=row_idx, column=4, value='안전' if result.get('is_safe') else '위험')
            safe_cell.border = border
            safe_cell.alignment = center_align
            if result.get('is_safe'):
                safe_cell.fill = safe_fill
            else:
                safe_cell.fill = unsafe_fill

            ws_detail.cell(row=row_idx, column=5, value=result.get('unsafe_reason', '')).border = border
            ws_detail.cell(row=row_idx, column=6, value=result.get('safe_context', '')).border = border
            ws_detail.cell(row=row_idx, column=7, value=result.get('total_options', 0)).border = border
            ws_detail.cell(row=row_idx, column=8, value=result.get('valid_options', 0)).border = border
            ws_detail.cell(row=row_idx, column=9, value=result.get('final_options', 0)).border = border
            ws_detail.cell(row=row_idx, column=10, value=result.get('bait_options', 0)).border = border

            # 미끼옵션목록: A, B, C 형식으로 줄바꿈
            bait_cell = ws_detail.cell(row=row_idx, column=11,
                                       value=self._format_options_abc(result.get('bait_option_list', [])[:5]))
            bait_cell.alignment = wrap_align
            bait_cell.border = border

            # 선택 (프로그램 추천 기본값 A, 사용자가 수정 가능)
            select_cell2 = ws_detail.cell(row=row_idx, column=12, value='A')
            select_cell2.alignment = center_align
            select_cell2.fill = select_fill
            select_cell2.border = border

            ws_detail.cell(row=row_idx, column=13, value=result.get('main_option_name', '')).border = border
            ws_detail.cell(row=row_idx, column=14, value=result.get('min_price_cny', 0)).border = border
            ws_detail.cell(row=row_idx, column=15, value=result.get('max_price_cny', 0)).border = border

            # 최종옵션목록: A, B, C 형식으로 줄바꿈
            final_opt_cell = ws_detail.cell(row=row_idx, column=16,
                                            value=self._format_options_abc(result.get('final_option_list', [])))
            final_opt_cell.alignment = wrap_align
            final_opt_cell.border = border

            ws_detail.cell(row=row_idx, column=17, value=result.get('thumbnail_url', '')).border = border
            ws_detail.cell(row=row_idx, column=18, value=result.get('main_option_image', '')).border = border

        # 상세시트 열 너비 조정
        ws_detail.column_dimensions['A'].width = 12  # 그룹
        ws_detail.column_dimensions['B'].width = 12  # 불사자ID
        ws_detail.column_dimensions['C'].width = 40  # 상품명
        ws_detail.column_dimensions['D'].width = 8   # 안전여부
        ws_detail.column_dimensions['E'].width = 25  # 위험사유
        ws_detail.column_dimensions['F'].width = 15  # 안전컨텍스트
        ws_detail.column_dimensions['K'].width = 35  # 미끼옵션목록
        ws_detail.column_dimensions['L'].width = 6   # 선택
        ws_detail.column_dimensions['M'].width = 25  # 대표옵션
        ws_detail.column_dimensions['P'].width = 40  # 최종옵션목록
        ws_detail.column_dimensions['Q'].width = 45  # 썸네일URL
        ws_detail.column_dimensions['R'].width = 45  # 옵션이미지URL

        # 상세시트 행 높이 (옵션목록 줄바꿈용)
        for row_idx in range(2, len(self.results) + 2):
            ws_detail.row_dimensions[row_idx].height = 60

        # === 통계 시트 ===
        ws_stats = wb.create_sheet("통계")
        stats_data = [
            ["항목", "값"],
            ["전체 상품", self.stats['total']],
            ["안전 상품", self.stats['safe']],
            ["위험 상품", self.stats['unsafe']],
            ["AI 검증 상품", self.stats.get('ai_checked', 0)],
            ["미끼옵션 발견 상품", self.stats['bait_found']],
            ["썸네일 매칭 성공", self.stats['thumbnail_matched']],
            ["분석 일시", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ]
        for row_idx, row_data in enumerate(stats_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                ws_stats.cell(row=row_idx, column=col_idx, value=value)

        # 저장
        wb.save(filepath)

    def _format_options_abc(self, options: list, max_count: int = 10) -> str:
        """옵션 목록을 A, B, C 형태로 포맷팅"""
        import re
        if not options:
            return ''

        result = []
        labels = 'ABCDEFGHIJ'
        for i, opt in enumerate(options[:max_count]):
            label = labels[i] if i < len(labels) else str(i + 1)
            opt_name = str(opt) if opt else ''
            # 기존 라벨 제거 (A. B. C. ... Z. 형태로 시작하면 제거)
            opt_name = re.sub(r'^[A-Za-z]\.\s*', '', opt_name).strip()
            # 30자로 자르기
            opt_name = opt_name[:30]
            result.append(f"{label}. {opt_name}")

        return '\n'.join(result)


# ==================== 엑셀 반영 클래스 ====================
class ExcelApplier:
    """엑셀에서 수정한 내용을 불사자에 반영"""

    def __init__(self, api_client, log_callback=None):
        self.api_client = api_client
        self.log = log_callback or print
        self.is_running = False

        # 통계
        self.stats = {
            "total": 0,
            "updated": 0,
            "skipped": 0,
            "failed": 0,
            "danger_tagged": 0,
        }

    def read_excel(self, filepath: str) -> List[Dict]:
        """엑셀 파일 읽기 (상세정보 시트 우선)"""
        try:
            from openpyxl import load_workbook
            wb = load_workbook(filepath, data_only=True)

            # 상세정보 시트 우선, 없으면 첫 번째 시트
            if "상세정보" in wb.sheetnames:
                ws = wb["상세정보"]
            else:
                ws = wb.active

            # 헤더 읽기
            headers = []
            for col in range(1, ws.max_column + 1):
                val = ws.cell(row=1, column=col).value
                headers.append(str(val).strip() if val else f"col_{col}")

            # 데이터 읽기
            data = []
            for row_idx in range(2, ws.max_row + 1):
                row_data = {}
                for col_idx, header in enumerate(headers, 1):
                    val = ws.cell(row=row_idx, column=col_idx).value
                    row_data[header] = val
                # 불사자ID가 있는 행만 추가
                if row_data.get('불사자ID') or row_data.get('id'):
                    data.append(row_data)

            wb.close()
            return data

        except Exception as e:
            self.log(f"❌ 엑셀 읽기 실패: {e}")
            return []

    def parse_selected_option(self, select_value: str, options_text: str) -> Optional[Dict]:
        """
        선택된 옵션 파싱
        select_value: 'A', 'B', 'C' 등
        options_text: 'A. 옵션1(10.5)\nB. 옵션2(15.0)' 형태

        Returns: {'name': '옵션명', 'price': 10.5, 'index': 0}
        """
        if not select_value or not options_text:
            return None

        select_value = str(select_value).strip().upper()
        if not select_value:
            return None

        # 옵션 목록 파싱
        lines = options_text.strip().split('\n')
        for idx, line in enumerate(lines):
            line = line.strip()
            if not line:
                continue

            # "A. 옵션명(가격)" 형태 파싱
            if line.startswith(f"{select_value}."):
                # 라벨 제거
                option_part = line[2:].strip()

                # 가격 추출 (마지막 괄호 안의 숫자)
                import re
                price_match = re.search(r'\((\d+\.?\d*)\)$', option_part)
                price = float(price_match.group(1)) if price_match else 0

                # 옵션명 (가격 부분 제거)
                name = re.sub(r'\(\d+\.?\d*\)$', '', option_part).strip()

                return {
                    'name': name,
                    'price': price,
                    'index': idx,
                    'label': select_value
                }

        return None

    def apply_changes(self, excel_data: List[Dict], options: Dict):
        """
        엑셀 변경사항을 불사자에 반영

        options:
            - apply_main_option: 대표옵션 변경 반영
            - apply_product_name: 상품명 변경 반영
            - skip_dangerous: 위험상품(X) 스킵
            - tag_dangerous: 위험상품에 태그 추가
            - danger_tag_name: 위험상품 태그명
            - remove_danger_tag: 안전상품에서 위험태그 제거
        """
        self.is_running = True
        self.stats = {"total": 0, "updated": 0, "skipped": 0, "failed": 0, "danger_tagged": 0}

        apply_main_option = options.get('apply_main_option', True)
        apply_product_name = options.get('apply_product_name', False)
        skip_dangerous = options.get('skip_dangerous', True)
        tag_dangerous = options.get('tag_dangerous', False)
        danger_tag_name = options.get('danger_tag_name', '위험상품')
        remove_danger_tag = options.get('remove_danger_tag', False)

        self.log("")
        self.log("=" * 50)
        self.log("📝 엑셀 반영 시작")
        self.log(f"   총 {len(excel_data)}개 상품")
        self.log(f"   대표옵션 반영: {'O' if apply_main_option else 'X'}")
        self.log(f"   상품명 반영: {'O' if apply_product_name else 'X'}")
        self.log(f"   위험상품 스킵: {'O' if skip_dangerous else 'X'}")
        self.log(f"   위험상품 태그: {'O' if tag_dangerous else 'X'}")
        self.log("=" * 50)

        for idx, row in enumerate(excel_data):
            if not self.is_running:
                break

            self.stats['total'] += 1
            product_id = str(row.get('불사자ID') or row.get('id') or '').strip()

            if not product_id:
                self.stats['skipped'] += 1
                continue

            # 안전여부 확인
            safety_value = str(row.get('안전여부', '')).strip().upper()
            is_safe = safety_value in ['O', '안전', 'SAFE', 'TRUE', '1']
            is_dangerous = safety_value in ['X', '위험', 'DANGER', 'FALSE', '0']

            # 위험상품 처리
            if is_dangerous:
                if skip_dangerous and not tag_dangerous:
                    self.stats['skipped'] += 1
                    continue

                if tag_dangerous:
                    # 위험 태그 추가
                    success = self._add_tag(product_id, danger_tag_name)
                    if success:
                        self.stats['danger_tagged'] += 1
                        self.log(f"🏷️ [{idx+1}] {product_id} → 위험태그 추가")

                    if skip_dangerous:
                        self.stats['skipped'] += 1
                        continue

            # 안전상품에서 위험태그 제거
            if is_safe and remove_danger_tag:
                self._remove_tag(product_id, danger_tag_name)

            # 업데이트할 데이터 준비
            update_data = {}

            # 1. 대표옵션 변경
            if apply_main_option:
                select_value = row.get('선택', 'A')
                options_text = row.get('최종옵션목록') or row.get('옵션명', '')

                selected = self.parse_selected_option(select_value, options_text)
                if selected and selected['label'] != 'A':
                    # A가 아닌 다른 옵션을 선택한 경우 → 대표옵션 변경 필요
                    update_data['mainOptionIndex'] = selected['index']
                    update_data['mainOptionName'] = selected['name']

            # 2. 상품명 변경
            if apply_product_name:
                new_name = row.get('상품명', '').strip()
                original_name = row.get('원본상품명', '').strip()

                if new_name and new_name != original_name:
                    update_data['uploadCommonProductName'] = new_name

            # 업데이트 실행
            if update_data:
                success, msg = self.api_client.update_product_fields(product_id, update_data)
                if success:
                    self.stats['updated'] += 1
                    changes = []
                    if 'mainOptionIndex' in update_data:
                        changes.append(f"대표옵션→{update_data.get('mainOptionName', '')[:15]}")
                    if 'uploadCommonProductName' in update_data:
                        changes.append("상품명변경")
                    self.log(f"✅ [{idx+1}] {product_id} → {', '.join(changes)}")
                else:
                    self.stats['failed'] += 1
                    self.log(f"❌ [{idx+1}] {product_id} → {msg[:50]}")
            else:
                self.stats['skipped'] += 1

            # 진행상황 (50개마다)
            if (idx + 1) % 50 == 0:
                self.log(f"   ... {idx+1}/{len(excel_data)} 처리 완료")

        # 결과 요약
        self.log("")
        self.log("=" * 50)
        self.log("📊 반영 결과")
        self.log(f"   전체: {self.stats['total']}개")
        self.log(f"   업데이트: {self.stats['updated']}개")
        self.log(f"   스킵: {self.stats['skipped']}개")
        self.log(f"   실패: {self.stats['failed']}개")
        if tag_dangerous:
            self.log(f"   위험태그: {self.stats['danger_tagged']}개")
        self.log("=" * 50)

        self.is_running = False
        return self.stats

    def _add_tag(self, product_id: str, tag_name: str) -> bool:
        """상품에 태그 추가"""
        try:
            # 현재 상품 정보 조회
            detail = self.api_client.get_product_detail(product_id)
            current_tags = detail.get('tags', []) or []

            # 이미 태그가 있으면 스킵
            if tag_name in current_tags:
                return True

            # 태그 추가
            new_tags = current_tags + [tag_name]
            success, msg = self.api_client.update_product_fields(product_id, {'tags': new_tags})
            return success
        except Exception as e:
            return False

    def _remove_tag(self, product_id: str, tag_name: str) -> bool:
        """상품에서 태그 제거"""
        try:
            detail = self.api_client.get_product_detail(product_id)
            current_tags = detail.get('tags', []) or []

            if tag_name not in current_tags:
                return True

            new_tags = [t for t in current_tags if t != tag_name]
            success, msg = self.api_client.update_product_fields(product_id, {'tags': new_tags})
            return success
        except Exception as e:
            return False


# ==================== GUI 클래스 ====================
class App(tk.Tk):
    def __init__(self):
        super().__init__()

        self.title("불사자 시뮬레이터 v1.1 - 분석 및 반영")
        self.geometry("950x800")
        self.resizable(True, True)

        self.config_data = load_config()
        self.simulator = BulsajaSimulator(self)
        self.excel_applier = None  # API 연결 후 초기화
        self.worker_thread = None

        self.create_widgets()
        self.load_saved_settings()

        self.protocol("WM_DELETE_WINDOW", self.on_close)

    def create_widgets(self):
        # 메인 컨테이너
        main_container = ttk.Frame(self, padding="5")
        main_container.pack(fill=tk.BOTH, expand=True)

        # === 상단: API 연결 (공통) ===
        conn_frame = ttk.LabelFrame(main_container, text="🔑 API 연결", padding="5")
        conn_frame.pack(fill=tk.X, pady=(0, 5))

        row0 = ttk.Frame(conn_frame)
        row0.pack(fill=tk.X, pady=2)
        ttk.Button(row0, text="🌐 크롬", command=self.open_debug_chrome, width=8).pack(side=tk.LEFT)
        ttk.Button(row0, text="🔑 토큰", command=self.extract_tokens, width=8).pack(side=tk.LEFT, padx=2)
        ttk.Button(row0, text="🔗 연결", command=self.connect_api, width=8).pack(side=tk.LEFT, padx=2)
        self.api_status = ttk.Label(row0, text="연결 안 됨", foreground="gray")
        self.api_status.pack(side=tk.LEFT, padx=10)
        ttk.Label(row0, text="포트:").pack(side=tk.RIGHT)
        self.port_var = tk.StringVar(value="9222")
        ttk.Entry(row0, textvariable=self.port_var, width=6).pack(side=tk.RIGHT, padx=2)

        # === 탭 노트북 ===
        self.notebook = ttk.Notebook(main_container)
        self.notebook.pack(fill=tk.BOTH, expand=True)

        # 탭1: 시뮬레이션
        self.tab_simulation = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_simulation, text="🔬 시뮬레이션")
        self.create_simulation_tab()

        # 탭2: 엑셀 반영
        self.tab_apply = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_apply, text="📝 엑셀 반영")
        self.create_apply_tab()

        # === 하단: 로그 (공통) ===
        log_frame = ttk.LabelFrame(main_container, text="📋 로그", padding="5")
        log_frame.pack(fill=tk.BOTH, expand=True, pady=(5, 0))

        self.log_text = scrolledtext.ScrolledText(log_frame, height=12, state='disabled', font=('Consolas', 9))
        self.log_text.pack(fill=tk.BOTH, expand=True)

        # Footer
        footer = ttk.Frame(main_container)
        footer.pack(fill=tk.X, pady=(5, 0))
        ttk.Label(footer, text="v1.1 by 프코노미 | 시뮬레이션 + 엑셀 반영", foreground="gray").pack(side=tk.RIGHT)

    def create_simulation_tab(self):
        """시뮬레이션 탭 생성"""
        main_frame = self.tab_simulation

        # === 1. 시뮬레이션 설정 ===
        sim_frame = ttk.LabelFrame(main_frame, text="🔬 시뮬레이션 설정", padding="5")
        sim_frame.pack(fill=tk.X, pady=(0, 5))

        row1 = ttk.Frame(sim_frame)
        row1.pack(fill=tk.X, pady=2)

        ttk.Label(row1, text="그룹당 최대 상품:").pack(side=tk.LEFT)
        self.max_products_var = tk.StringVar(value="100")
        ttk.Entry(row1, textvariable=self.max_products_var, width=8).pack(side=tk.LEFT, padx=(2, 10))

        ttk.Label(row1, text="옵션 개수:").pack(side=tk.LEFT)
        self.option_count_var = tk.StringVar(value="5")
        ttk.Entry(row1, textvariable=self.option_count_var, width=5).pack(side=tk.LEFT, padx=(2, 10))

        ttk.Label(row1, text="작업 그룹:").pack(side=tk.LEFT)
        self.work_groups_var = tk.StringVar(value="1-5")
        ttk.Entry(row1, textvariable=self.work_groups_var, width=15).pack(side=tk.LEFT, padx=2)
        ttk.Label(row1, text="(예: 1-5 또는 1,3,5)", foreground="gray").pack(side=tk.LEFT, padx=5)

        ttk.Button(row1, text="📥 그룹목록", command=self.load_market_groups, width=10).pack(side=tk.RIGHT)

        # 그룹 목록
        row2 = ttk.Frame(sim_frame)
        row2.pack(fill=tk.X, pady=2)

        ttk.Label(row2, text="마켓 그룹 목록 (쉼표 구분, 숫자 맵핑용):").pack(anchor=tk.W)

        self.group_text = scrolledtext.ScrolledText(sim_frame, height=2, width=80, font=('Consolas', 9))
        self.group_text.pack(fill=tk.X, expand=True, pady=2)

        ttk.Label(sim_frame, text="예: 01_푸로테카,02_스트롬브린 → 작업그룹에서 1, 1-3 등으로 사용",
                  foreground="gray").pack(anchor=tk.W)

        # === 3. 미끼 키워드 설정 ===
        keyword_frame = ttk.LabelFrame(main_frame, text="🚫 미끼 키워드 (학습/수정 가능)", padding="5")
        keyword_frame.pack(fill=tk.X, pady=(0, 5))

        keyword_row1 = ttk.Frame(keyword_frame)
        keyword_row1.pack(fill=tk.X, pady=2)

        ttk.Label(keyword_row1, text="제외 키워드 (쉼표 구분):").pack(side=tk.LEFT)
        ttk.Button(keyword_row1, text="기본값", command=self.reset_keywords, width=6).pack(side=tk.RIGHT)
        ttk.Button(keyword_row1, text="💾 저장", command=self.save_keywords, width=6).pack(side=tk.RIGHT, padx=2)

        self.keyword_text = scrolledtext.ScrolledText(keyword_frame, height=3, width=80, font=('Consolas', 9))
        self.keyword_text.pack(fill=tk.X, expand=True)
        self.keyword_text.insert("1.0", ','.join(self.simulator.bait_keywords))

        # === 4. 카테고리 검수 설정 ===
        category_frame = ttk.LabelFrame(main_frame, text="🛡️ 카테고리별 검수 수준", padding="5")
        category_frame.pack(fill=tk.X, pady=(0, 5))

        cat_row1 = ttk.Frame(category_frame)
        cat_row1.pack(fill=tk.X, pady=2)

        ttk.Label(cat_row1, text="검수 수준:").pack(side=tk.LEFT)

        # 엄격(strict) = AI 확인 필수, 보통(normal) = 프로그램 자동, 제외(skip) = 검수 안함
        self.check_level_var = tk.StringVar(value="normal")
        ttk.Radiobutton(cat_row1, text="보통 (프로그램)", variable=self.check_level_var, value="normal").pack(side=tk.LEFT, padx=5)
        ttk.Radiobutton(cat_row1, text="엄격 (AI확인)", variable=self.check_level_var, value="strict").pack(side=tk.LEFT, padx=5)
        ttk.Radiobutton(cat_row1, text="제외", variable=self.check_level_var, value="skip").pack(side=tk.LEFT, padx=5)

        ttk.Button(cat_row1, text="⚙️ 카테고리 설정", command=self.open_category_settings, width=14).pack(side=tk.RIGHT)

        cat_row2 = ttk.Frame(category_frame)
        cat_row2.pack(fill=tk.X, pady=2)

        ttk.Label(cat_row2, text="위험 카테고리 (자동 엄격):").pack(side=tk.LEFT)
        self.risk_categories_var = tk.StringVar(value="패션의류,패션잡화,유아동,의료기기,화장품,시계,가방")
        ttk.Entry(cat_row2, textvariable=self.risk_categories_var, width=60).pack(side=tk.LEFT, padx=2, fill=tk.X, expand=True)

        ttk.Label(category_frame, text="※ 엄격: 위험 키워드 발견 시 Gemini AI로 재확인 | 보통: 안전 컨텍스트로 자동 판단",
                  foreground="gray").pack(anchor=tk.W)

        # === 5. 저장 경로 ===
        save_frame = ttk.LabelFrame(main_frame, text="💾 저장 설정", padding="5")
        save_frame.pack(fill=tk.X, pady=(0, 5))

        row3 = ttk.Frame(save_frame)
        row3.pack(fill=tk.X, pady=2)

        ttk.Label(row3, text="저장 경로:").pack(side=tk.LEFT)
        self.save_path_var = tk.StringVar(value=f"simulation_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")
        ttk.Entry(row3, textvariable=self.save_path_var, width=50).pack(side=tk.LEFT, padx=2, fill=tk.X, expand=True)
        ttk.Button(row3, text="찾아보기", command=self.browse_save_path, width=8).pack(side=tk.RIGHT)

        # === 진행 상태 ===
        progress_frame = ttk.Frame(main_frame)
        progress_frame.pack(fill=tk.X, pady=(0, 5))

        self.progress_var = tk.StringVar(value="대기 중...")
        ttk.Label(progress_frame, textvariable=self.progress_var).pack(side=tk.LEFT)
        self.progress_bar = ttk.Progressbar(progress_frame, mode='determinate')
        self.progress_bar.pack(side=tk.RIGHT, fill=tk.X, expand=True, padx=(10, 0))

        # === 버튼 ===
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=tk.X, pady=(0, 5))

        self.btn_start = ttk.Button(btn_frame, text="🔬 시뮬레이션 시작", command=self.start_simulation)
        self.btn_start.pack(side=tk.LEFT, padx=(0, 10))

        self.btn_stop = ttk.Button(btn_frame, text="🛑 중지", command=self.stop, state="disabled")
        self.btn_stop.pack(side=tk.LEFT)

        ttk.Button(btn_frame, text="💾 설정 저장", command=self.save_settings).pack(side=tk.RIGHT)

    def create_apply_tab(self):
        """엑셀 반영 탭 생성"""
        main_frame = self.tab_apply

        # === 1. 엑셀 파일 선택 ===
        file_frame = ttk.LabelFrame(main_frame, text="📂 엑셀 파일 선택", padding="5")
        file_frame.pack(fill=tk.X, pady=(0, 5))

        file_row = ttk.Frame(file_frame)
        file_row.pack(fill=tk.X, pady=2)

        ttk.Label(file_row, text="파일:").pack(side=tk.LEFT)
        self.apply_file_var = tk.StringVar()
        ttk.Entry(file_row, textvariable=self.apply_file_var, width=60).pack(side=tk.LEFT, padx=2, fill=tk.X, expand=True)
        ttk.Button(file_row, text="찾아보기", command=self.browse_apply_file, width=10).pack(side=tk.RIGHT)

        # 파일 정보
        self.apply_file_info = ttk.Label(file_frame, text="파일을 선택하세요", foreground="gray")
        self.apply_file_info.pack(anchor=tk.W, pady=2)

        # === 2. 반영 옵션 ===
        option_frame = ttk.LabelFrame(main_frame, text="⚙️ 반영 옵션", padding="5")
        option_frame.pack(fill=tk.X, pady=(0, 5))

        # 체크박스 옵션들
        opt_row1 = ttk.Frame(option_frame)
        opt_row1.pack(fill=tk.X, pady=2)

        self.apply_main_option_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(opt_row1, text="대표옵션 변경 반영", variable=self.apply_main_option_var).pack(side=tk.LEFT, padx=(0, 20))

        self.apply_product_name_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(opt_row1, text="상품명 변경 반영", variable=self.apply_product_name_var).pack(side=tk.LEFT, padx=(0, 20))

        opt_row2 = ttk.Frame(option_frame)
        opt_row2.pack(fill=tk.X, pady=2)

        self.skip_dangerous_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(opt_row2, text="위험상품(X) 스킵", variable=self.skip_dangerous_var).pack(side=tk.LEFT, padx=(0, 20))

        self.tag_dangerous_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(opt_row2, text="위험상품 태그 추가", variable=self.tag_dangerous_var).pack(side=tk.LEFT, padx=(0, 10))

        ttk.Label(opt_row2, text="태그명:").pack(side=tk.LEFT)
        self.danger_tag_var = tk.StringVar(value="위험상품")
        ttk.Entry(opt_row2, textvariable=self.danger_tag_var, width=15).pack(side=tk.LEFT, padx=2)

        opt_row3 = ttk.Frame(option_frame)
        opt_row3.pack(fill=tk.X, pady=2)

        self.remove_danger_tag_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(opt_row3, text="안전상품(O)에서 위험태그 제거", variable=self.remove_danger_tag_var).pack(side=tk.LEFT)

        # 설명
        ttk.Label(option_frame, text="※ 엑셀의 '선택' 컬럼에서 A가 아닌 다른 값(B,C,D...)을 선택한 상품만 대표옵션이 변경됩니다",
                  foreground="gray").pack(anchor=tk.W, pady=(5, 0))

        # === 3. 미리보기 ===
        preview_frame = ttk.LabelFrame(main_frame, text="👁️ 미리보기 (변경될 항목)", padding="5")
        preview_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 5))

        # 미리보기 리스트
        columns = ('불사자ID', '상품명', '선택', '변경내용', '안전여부')
        self.preview_tree = ttk.Treeview(preview_frame, columns=columns, show='headings', height=8)

        for col in columns:
            self.preview_tree.heading(col, text=col)
            width = 80 if col in ['불사자ID', '선택', '안전여부'] else 200
            self.preview_tree.column(col, width=width)

        # 스크롤바
        preview_scroll = ttk.Scrollbar(preview_frame, orient="vertical", command=self.preview_tree.yview)
        self.preview_tree.configure(yscrollcommand=preview_scroll.set)

        self.preview_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        preview_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        # === 4. 버튼 ===
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=tk.X, pady=5)

        ttk.Button(btn_frame, text="🔍 미리보기", command=self.preview_apply).pack(side=tk.LEFT, padx=(0, 10))

        self.btn_apply = ttk.Button(btn_frame, text="📝 반영 실행", command=self.start_apply)
        self.btn_apply.pack(side=tk.LEFT, padx=(0, 10))

        self.btn_apply_stop = ttk.Button(btn_frame, text="🛑 중지", command=self.stop_apply, state="disabled")
        self.btn_apply_stop.pack(side=tk.LEFT)

        # 진행상태
        self.apply_progress_var = tk.StringVar(value="대기 중...")
        ttk.Label(btn_frame, textvariable=self.apply_progress_var).pack(side=tk.RIGHT)

    def browse_apply_file(self):
        """엑셀 반영용 파일 선택"""
        filepath = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialdir="."
        )
        if filepath:
            self.apply_file_var.set(filepath)
            self.load_apply_preview()

    def load_apply_preview(self):
        """선택한 엑셀 파일 미리보기 로드"""
        filepath = self.apply_file_var.get()
        if not filepath or not os.path.exists(filepath):
            return

        try:
            # 임시 ExcelApplier 생성 (API 없이 파일만 읽기)
            temp_applier = ExcelApplier(None, self.log)
            data = temp_applier.read_excel(filepath)

            if not data:
                self.apply_file_info.config(text="⚠️ 데이터를 읽을 수 없습니다", foreground="red")
                return

            # 파일 정보 업데이트
            self.apply_file_info.config(
                text=f"✅ {len(data)}개 상품 로드됨",
                foreground="green"
            )

            # 미리보기 데이터 저장
            self.apply_excel_data = data

            # 미리보기 갱신
            self.preview_apply()

        except Exception as e:
            self.apply_file_info.config(text=f"❌ 오류: {e}", foreground="red")
            self.log(f"❌ 엑셀 로드 오류: {e}")

    def preview_apply(self):
        """변경될 항목 미리보기"""
        # 기존 항목 삭제
        for item in self.preview_tree.get_children():
            self.preview_tree.delete(item)

        if not hasattr(self, 'apply_excel_data') or not self.apply_excel_data:
            self.log("⚠️ 먼저 엑셀 파일을 선택하세요")
            return

        temp_applier = ExcelApplier(None)
        apply_main_option = self.apply_main_option_var.get()
        skip_dangerous = self.skip_dangerous_var.get()
        tag_dangerous = self.tag_dangerous_var.get()

        changes_count = 0
        for row in self.apply_excel_data:
            product_id = str(row.get('불사자ID') or row.get('id') or '').strip()
            product_name = str(row.get('상품명', ''))[:30]
            select_value = str(row.get('선택', 'A')).strip().upper()
            safety_value = str(row.get('안전여부', '')).strip().upper()

            # 안전여부 판정
            is_dangerous = safety_value in ['X', '위험', 'DANGER', 'FALSE', '0']
            safety_display = 'X' if is_dangerous else 'O'

            # 변경 내용 판정
            changes = []

            # 대표옵션 변경 체크
            if apply_main_option and select_value and select_value != 'A':
                options_text = row.get('최종옵션목록') or row.get('옵션명', '')
                selected = temp_applier.parse_selected_option(select_value, options_text)
                if selected:
                    changes.append(f"대표옵션→{select_value}")

            # 위험상품 태그
            if is_dangerous and tag_dangerous:
                changes.append("위험태그추가")

            # 위험상품 스킵
            if is_dangerous and skip_dangerous and not tag_dangerous:
                changes.append("(스킵)")

            change_text = ', '.join(changes) if changes else '-'

            # 변경사항 있는 것만 표시 (또는 전체 표시)
            if changes:
                self.preview_tree.insert('', tk.END, values=(
                    product_id, product_name, select_value, change_text, safety_display
                ))
                changes_count += 1

        self.log(f"📊 미리보기: {changes_count}개 상품 변경 예정")

    def start_apply(self):
        """엑셀 반영 실행"""
        if not self.simulator.api_client:
            messagebox.showwarning("경고", "먼저 API에 연결하세요")
            return

        if not hasattr(self, 'apply_excel_data') or not self.apply_excel_data:
            messagebox.showwarning("경고", "먼저 엑셀 파일을 선택하세요")
            return

        # ExcelApplier 초기화
        self.excel_applier = ExcelApplier(self.simulator.api_client, self.log)

        # 옵션 수집
        options = {
            'apply_main_option': self.apply_main_option_var.get(),
            'apply_product_name': self.apply_product_name_var.get(),
            'skip_dangerous': self.skip_dangerous_var.get(),
            'tag_dangerous': self.tag_dangerous_var.get(),
            'danger_tag_name': self.danger_tag_var.get(),
            'remove_danger_tag': self.remove_danger_tag_var.get(),
        }

        self.btn_apply.config(state="disabled")
        self.btn_apply_stop.config(state="normal")
        self.apply_progress_var.set("반영 중...")

        # 백그라운드 실행
        def task():
            try:
                self.excel_applier.apply_changes(self.apply_excel_data, options)
            finally:
                self.after(0, self.on_apply_finished)

        self.worker_thread = threading.Thread(target=task, daemon=True)
        self.worker_thread.start()

    def stop_apply(self):
        """엑셀 반영 중지"""
        if self.excel_applier:
            self.excel_applier.is_running = False
        self.log("🛑 반영 중지 요청...")

    def on_apply_finished(self):
        """엑셀 반영 완료"""
        self.btn_apply.config(state="normal")
        self.btn_apply_stop.config(state="disabled")
        self.apply_progress_var.set("완료")

    def load_saved_settings(self):
        c = self.config_data
        if "port" in c: self.port_var.set(c["port"])
        if "max_products" in c: self.max_products_var.set(c["max_products"])
        if "option_count" in c: self.option_count_var.set(c["option_count"])
        if "work_groups" in c: self.work_groups_var.set(c["work_groups"])
        if "group_text" in c:
            self.group_text.delete("1.0", tk.END)
            self.group_text.insert("1.0", c["group_text"])
        # save_path는 항상 현재 시간으로 새로 생성 (설정에서 불러오지 않음)
        # 검수 설정
        if "check_level" in c: self.check_level_var.set(c["check_level"])
        if "risk_categories" in c: self.risk_categories_var.set(c["risk_categories"])

    def save_settings(self):
        self.config_data["port"] = self.port_var.get()
        self.config_data["max_products"] = self.max_products_var.get()
        self.config_data["option_count"] = self.option_count_var.get()
        self.config_data["work_groups"] = self.work_groups_var.get()
        self.config_data["group_text"] = self.group_text.get("1.0", tk.END).strip()
        self.config_data["save_path"] = self.save_path_var.get()
        # 검수 설정
        self.config_data["check_level"] = self.check_level_var.get()
        self.config_data["risk_categories"] = self.risk_categories_var.get()
        save_config(self.config_data)
        self.log("✅ 설정 저장됨")

    def log(self, message):
        def _log():
            self.log_text.config(state='normal')
            timestamp = datetime.now().strftime("%H:%M:%S")
            self.log_text.insert(tk.END, f"[{timestamp}] {message}\n")
            self.log_text.see(tk.END)
            self.log_text.config(state='disabled')
        self.after(0, _log)

    def update_progress(self, current, total):
        def _update():
            self.progress_var.set(f"{current}/{total} 상품 분석 중...")
            self.progress_bar['value'] = (current / total) * 100 if total > 0 else 0
        self.after(0, _update)

    def open_debug_chrome(self):
        import subprocess
        port = int(self.port_var.get().strip())
        profile_dir = f"C:\\chrome_debug_profile_{port}"
        chrome_paths = [
            r"C:\Program Files\Google\Chrome\Application\chrome.exe",
            r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
            os.path.expandvars(r"%LOCALAPPDATA%\Google\Chrome\Application\chrome.exe"),
        ]
        chrome_path = None
        for path in chrome_paths:
            if os.path.exists(path):
                chrome_path = path
                break
        if not chrome_path:
            messagebox.showerror("오류", "Chrome을 찾을 수 없습니다")
            return
        url = "https://www.bulsaja.com/products/manage/list/"
        cmd = f'"{chrome_path}" --remote-debugging-port={port} --user-data-dir="{profile_dir}" --remote-allow-origins=* "{url}"'
        try:
            subprocess.Popen(cmd, shell=True)
            self.log(f"🌐 크롬 실행 (포트: {port})")
        except Exception as e:
            self.log(f"❌ 크롬 실행 실패: {e}")

    def extract_tokens(self):
        port = int(self.port_var.get().strip())
        self.log(f"🔍 토큰 추출 중...")

        def task():
            success, access, refresh, err = extract_tokens_from_browser(port)
            if success:
                self.config_data["access_token"] = access
                self.config_data["refresh_token"] = refresh
                self.log("✅ 토큰 추출 성공")
                self.after(500, self.connect_api)
            else:
                self.log(f"❌ 토큰 추출 실패: {err}")

        threading.Thread(target=task, daemon=True).start()

    def connect_api(self):
        access = self.config_data.get("access_token", "")
        refresh = self.config_data.get("refresh_token", "")
        if not access or not refresh:
            messagebox.showwarning("경고", "먼저 토큰을 추출하세요")
            return
        self.log("🔗 API 연결 중...")
        success, msg, total = self.simulator.init_api_client(access, refresh)
        if success:
            self.api_status.config(text=f"✅ 연결됨 ({total}개)", foreground="green")
            self.log(f"✅ {msg}")
        else:
            self.api_status.config(text="❌ 실패", foreground="red")
            self.log(f"❌ 연결 실패: {msg}")

    def load_market_groups(self):
        if not self.simulator.api_client:
            messagebox.showwarning("경고", "먼저 API에 연결하세요")
            return

        self.log("📥 마켓 그룹 목록 조회 중...")

        try:
            groups = self.simulator.api_client.get_market_groups()
            if groups:
                self.group_text.delete("1.0", tk.END)
                self.group_text.insert("1.0", ','.join(groups))
                self.log(f"✅ {len(groups)}개 그룹 로드됨")
            else:
                self.log("⚠️ 그룹 없음 또는 조회 실패")
        except Exception as e:
            self.log(f"❌ 그룹 로드 실패: {e}")

    def reset_keywords(self):
        self.keyword_text.delete("1.0", tk.END)
        self.keyword_text.insert("1.0", ','.join(DEFAULT_BAIT_KEYWORDS))
        self.log("🔄 미끼 키워드 기본값으로 초기화")

    def save_keywords(self):
        text = self.keyword_text.get("1.0", tk.END).strip()
        keywords = [k.strip() for k in text.split(',') if k.strip()]
        if save_bait_keywords(keywords):
            self.simulator.bait_keywords = keywords
            self.log(f"✅ 미끼 키워드 저장됨 ({len(keywords)}개)")
        else:
            self.log("❌ 미끼 키워드 저장 실패")

    def open_category_settings(self):
        """카테고리별 검수 설정 다이얼로그"""
        dialog = tk.Toplevel(self)
        dialog.title("카테고리별 검수 설정")
        dialog.geometry("600x500")
        dialog.transient(self)
        dialog.grab_set()

        main_frame = ttk.Frame(dialog, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # 설명
        ttk.Label(main_frame, text="카테고리별 검수 수준을 설정합니다. (strict=AI확인, normal=프로그램, skip=제외)",
                  foreground="gray").pack(anchor=tk.W)

        # 현재 설정 로드
        current_settings = load_category_risk_settings()

        # 스크롤 가능한 리스트
        list_frame = ttk.Frame(main_frame)
        list_frame.pack(fill=tk.BOTH, expand=True, pady=10)

        canvas = tk.Canvas(list_frame)
        scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # 카테고리별 콤보박스
        self.category_combos = {}
        levels = ['strict', 'normal', 'skip']

        for cat_name in sorted(current_settings.keys()):
            row = ttk.Frame(scrollable_frame)
            row.pack(fill=tk.X, pady=1)

            ttk.Label(row, text=cat_name, width=20).pack(side=tk.LEFT)
            combo_var = tk.StringVar(value=current_settings.get(cat_name, 'normal'))
            combo = ttk.Combobox(row, textvariable=combo_var, values=levels, width=10, state='readonly')
            combo.pack(side=tk.LEFT, padx=5)
            self.category_combos[cat_name] = combo_var

            # 레벨 설명
            level = current_settings.get(cat_name, 'normal')
            level_desc = {'strict': '(AI확인)', 'normal': '(자동)', 'skip': '(제외)'}.get(level, '')
            ttk.Label(row, text=level_desc, foreground="gray", width=10).pack(side=tk.LEFT)

        # 새 카테고리 추가
        add_frame = ttk.Frame(main_frame)
        add_frame.pack(fill=tk.X, pady=5)

        ttk.Label(add_frame, text="카테고리 추가:").pack(side=tk.LEFT)
        self.new_cat_var = tk.StringVar()
        ttk.Entry(add_frame, textvariable=self.new_cat_var, width=20).pack(side=tk.LEFT, padx=2)
        self.new_cat_level_var = tk.StringVar(value='normal')
        ttk.Combobox(add_frame, textvariable=self.new_cat_level_var, values=levels, width=10, state='readonly').pack(side=tk.LEFT, padx=2)

        def add_category():
            cat = self.new_cat_var.get().strip()
            level = self.new_cat_level_var.get()
            if cat and cat not in self.category_combos:
                row = ttk.Frame(scrollable_frame)
                row.pack(fill=tk.X, pady=1)
                ttk.Label(row, text=cat, width=20).pack(side=tk.LEFT)
                combo_var = tk.StringVar(value=level)
                ttk.Combobox(row, textvariable=combo_var, values=levels, width=10, state='readonly').pack(side=tk.LEFT, padx=5)
                self.category_combos[cat] = combo_var
                self.new_cat_var.set("")
                self.log(f"카테고리 추가: {cat} ({level})")

        ttk.Button(add_frame, text="추가", command=add_category, width=8).pack(side=tk.LEFT, padx=5)

        # 버튼
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=tk.X, pady=10)

        def save_and_close():
            new_settings = {}
            for cat_name, combo_var in self.category_combos.items():
                new_settings[cat_name] = combo_var.get()
            if save_category_risk_settings(new_settings):
                self.log(f"✅ 카테고리 검수 설정 저장 ({len(new_settings)}개)")
                # 위험 카테고리 텍스트 업데이트
                risk_cats = [cat for cat, level in new_settings.items() if level == 'strict']
                self.risk_categories_var.set(','.join(risk_cats[:10]))  # 상위 10개만 표시
            else:
                self.log("❌ 카테고리 설정 저장 실패")
            dialog.destroy()

        def reset_to_default():
            for cat_name, combo_var in self.category_combos.items():
                default_level = DEFAULT_CATEGORY_RISK_SETTINGS.get(cat_name, 'normal')
                combo_var.set(default_level)
            self.log("🔄 카테고리 설정 기본값으로 초기화")

        ttk.Button(btn_frame, text="기본값", command=reset_to_default, width=10).pack(side=tk.LEFT)
        ttk.Button(btn_frame, text="취소", command=dialog.destroy, width=10).pack(side=tk.RIGHT, padx=5)
        ttk.Button(btn_frame, text="저장", command=save_and_close, width=10).pack(side=tk.RIGHT)

    def browse_save_path(self):
        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=self.save_path_var.get()
        )
        if filepath:
            self.save_path_var.set(filepath)

    def parse_group_mapping(self) -> Dict[str, str]:
        """그룹 매핑 텍스트 파싱

        그룹명 형식이 '22_리코즈' 처럼 숫자_이름 형태면 → 숫자(22)로 매핑
        그 외에는 순서대로 1,2,3... 매핑
        """
        mapping = {}
        text = self.group_text.get("1.0", tk.END).strip()
        if not text:
            return mapping

        groups = [g.strip() for g in text.split(',') if g.strip()]

        # 숫자 접두사가 있는지 확인 (예: "22_리코즈")
        has_prefix_pattern = False
        import re
        prefix_pattern = re.compile(r'^(\d+)[_\-]')

        for group_name in groups:
            match = prefix_pattern.match(group_name)
            if match:
                has_prefix_pattern = True
                break

        if has_prefix_pattern:
            # 그룹명에 숫자 접두사가 있으면 그 숫자로 매핑
            for group_name in groups:
                match = prefix_pattern.match(group_name)
                if match:
                    num_str = match.group(1)
                    # 01, 1 둘 다 같은 그룹으로 매핑
                    mapping[num_str] = group_name
                    mapping[str(int(num_str))] = group_name  # 앞의 0 제거 버전
                    mapping[f"{int(num_str):02d}"] = group_name  # 2자리 포맷
        else:
            # 숫자 접두사가 없으면 순서대로 매핑
            for idx, group_name in enumerate(groups, 1):
                mapping[str(idx)] = group_name
                mapping[f"{idx:02d}"] = group_name

        return mapping

    def parse_work_range(self, range_str: str) -> List[str]:
        """작업 범위 파싱 (1-20 또는 1,3,5)"""
        result = []
        range_str = range_str.strip()
        if '-' in range_str and ',' not in range_str:
            parts = range_str.split('-')
            if len(parts) == 2:
                try:
                    start = int(parts[0])
                    end = int(parts[1])
                    for i in range(start, end + 1):
                        result.append(str(i))
                except ValueError:
                    pass
        else:
            for item in range_str.split(','):
                item = item.strip()
                if item:
                    result.append(item)
        return result

    def get_group_names_from_range(self) -> List[str]:
        """작업 범위에서 실제 그룹명 목록 가져오기"""
        mapping = self.parse_group_mapping()
        range_nums = self.parse_work_range(self.work_groups_var.get())
        group_names = []
        for num in range_nums:
            if num in mapping:
                group_names.append(mapping[num])
            else:
                self.log(f"⚠️ 그룹 번호 {num}에 해당하는 그룹명 없음")
        return group_names

    def start_simulation(self):
        if not self.simulator.api_client:
            messagebox.showwarning("경고", "먼저 API에 연결하세요")
            return

        group_names = self.get_group_names_from_range()
        if not group_names:
            messagebox.showwarning("경고", "작업할 그룹이 없습니다. 작업범위와 그룹목록을 확인하세요.")
            return

        if not EXCEL_AVAILABLE:
            messagebox.showerror("오류", "openpyxl이 설치되지 않았습니다.\npip install openpyxl")
            return

        # 미끼 키워드 업데이트
        text = self.keyword_text.get("1.0", tk.END).strip()
        self.simulator.bait_keywords = [k.strip() for k in text.split(',') if k.strip()]

        max_products = int(self.max_products_var.get())
        option_count = int(self.option_count_var.get())

        # 파일명 자동 갱신 (실행 시마다 새 타임스탬프 - 날짜_시분)
        new_filename = f"simulation_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        self.save_path_var.set(new_filename)
        save_path = new_filename

        # 검수 설정
        check_level = self.check_level_var.get()
        risk_categories_text = self.risk_categories_var.get()
        risk_categories = [c.strip() for c in risk_categories_text.split(',') if c.strip()]

        self.btn_start.config(state="disabled")
        self.btn_stop.config(state="normal")

        self.worker_thread = threading.Thread(
            target=self.simulator.run_simulation,
            args=(group_names, max_products, option_count, save_path, check_level, risk_categories),
            daemon=True
        )
        self.worker_thread.start()

    def stop(self):
        self.simulator.is_running = False
        self.log("🛑 중지 요청...")

    def on_finished(self):
        def _update():
            self.btn_start.config(state="normal")
            self.btn_stop.config(state="disabled")
            self.progress_var.set("완료")
        self.after(0, _update)

    def on_close(self):
        self.simulator.is_running = False
        self.save_settings()
        self.destroy()


if __name__ == "__main__":
    app = App()
    app.mainloop()
