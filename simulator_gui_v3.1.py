# -*- coding: utf-8 -*-
"""
시뮬레이터 GUI v3.1 - 엑셀 검수 전용 (간소화)
- 업로드 기능 제거
- 엑셀 로드 및 옵션 선택에 집중
- 오류 개선 및 안정화
"""
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import os
import json
from pathlib import Path
from typing import Dict, List, Any

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    print("⚠️ pandas가 필요합니다: pip install pandas openpyxl")

try:
    from PIL import Image, ImageTk
    import requests
    from io import BytesIO
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False


class SimulatorGUIv31:
    """시뮬레이터 GUI v3.1 - 검수 전용"""

    def __init__(self, root):
        self.root = root
        self.root.title("불사자 시뮬레이터 v3.1 - 검수도구")
        self.root.geometry("1400x850")

        self.data = []
        self.selected_options = {}
        self.option_frames = {}
        self.image_cache = {}
        self.current_file = None

        self._create_ui()
        self._auto_load_latest()

    def _create_ui(self):
        # 상단 툴바
        toolbar = ttk.Frame(self.root, padding=5)
        toolbar.pack(fill=tk.X)

        ttk.Label(toolbar, text="불사자 시뮬레이터 v3.1", 
                 font=("맑은 고딕", 12, "bold")).pack(side=tk.LEFT, padx=10)

        ttk.Separator(toolbar, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=10)

        ttk.Button(toolbar, text="📂 엑셀 열기", command=self._load_excel).pack(side=tk.LEFT, padx=5)
        ttk.Button(toolbar, text="💾 선택 저장", command=self._save_excel).pack(side=tk.LEFT, padx=5)
        ttk.Button(toolbar, text="🔄 새로고침", command=self._reload_data).pack(side=tk.LEFT, padx=5)

        ttk.Separator(toolbar, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=10)

        self.file_label = ttk.Label(toolbar, text="(파일 없음)", foreground="gray")
        self.file_label.pack(side=tk.LEFT, padx=5)

        self.count_label = ttk.Label(toolbar, text="상품: 0개")
        self.count_label.pack(side=tk.RIGHT, padx=20)

        # 필터 영역
        filter_frame = ttk.Frame(self.root, padding=5)
        filter_frame.pack(fill=tk.X)

        ttk.Label(filter_frame, text="필터:").pack(side=tk.LEFT)
        
        self.filter_safe = tk.BooleanVar(value=True)
        self.filter_unsafe = tk.BooleanVar(value=True)
        ttk.Checkbutton(filter_frame, text="안전", variable=self.filter_safe, 
                       command=self._apply_filter).pack(side=tk.LEFT, padx=5)
        ttk.Checkbutton(filter_frame, text="위험", variable=self.filter_unsafe,
                       command=self._apply_filter).pack(side=tk.LEFT, padx=5)

        ttk.Separator(filter_frame, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=10)

        ttk.Label(filter_frame, text="그룹:").pack(side=tk.LEFT, padx=5)
        self.group_combo = ttk.Combobox(filter_frame, width=20, state="readonly")
        self.group_combo.pack(side=tk.LEFT, padx=5)
        self.group_combo.bind("<<ComboboxSelected>>", lambda e: self._apply_filter())

        # 메인 영역 (스크롤)
        main_frame = ttk.Frame(self.root)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        self.canvas = tk.Canvas(main_frame, bg="white")
        scrollbar_y = ttk.Scrollbar(main_frame, orient=tk.VERTICAL, command=self.canvas.yview)
        scrollbar_x = ttk.Scrollbar(main_frame, orient=tk.HORIZONTAL, command=self.canvas.xview)

        self.scrollable_frame = ttk.Frame(self.canvas)
        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        )

        self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)

        scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)
        scrollbar_x.pack(side=tk.BOTTOM, fill=tk.X)
        self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        self.canvas.bind_all("<MouseWheel>", lambda e: self.canvas.yview_scroll(int(-1*(e.delta/120)), "units"))

        # 하단 상태바
        status_frame = ttk.Frame(self.root, relief=tk.SUNKEN, padding=2)
        status_frame.pack(fill=tk.X, side=tk.BOTTOM)
        
        self.status_label = ttk.Label(status_frame, text="대기 중...", foreground="gray")
        self.status_label.pack(side=tk.LEFT, padx=5)

    def _auto_load_latest(self):
        """최신 시뮬레이션 파일 자동 로드"""
        if not PANDAS_AVAILABLE:
            return

        base_dir = Path(__file__).parent
        simulation_files = list(base_dir.glob("simulation_*.xlsx"))

        if simulation_files:
            simulation_files.sort(key=lambda x: x.stat().st_mtime, reverse=True)
            self._load_excel_file(str(simulation_files[0]))

    def _load_excel(self):
        """파일 선택"""
        if not PANDAS_AVAILABLE:
            messagebox.showerror("오류", "pandas가 필요합니다: pip install pandas openpyxl")
            return

        filepath = filedialog.askopenfilename(
            title="시뮬레이션 엑셀 선택",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialdir=str(Path(__file__).parent)
        )
        if filepath:
            self._load_excel_file(filepath)

    def _load_excel_file(self, filepath):
        """엑셀 파일 로드"""
        try:
            self.status_label.config(text=f"로딩 중: {Path(filepath).name}...", foreground="blue")
            self.root.update()

            # 엑셀 로드 (상세정보 시트 우선)
            try:
                xls = pd.ExcelFile(filepath, engine='openpyxl')
                if "상세정보" in xls.sheet_names:
                    df = pd.read_excel(filepath, sheet_name="상세정보", engine='openpyxl')
                else:
                    df = pd.read_excel(filepath, engine='openpyxl')
            except Exception as e:
                df = pd.read_excel(filepath)

            self._parse_excel_data(df)
            self.current_file = filepath
            self.file_label.config(text=Path(filepath).name, foreground="black")
            self.count_label.config(text=f"상품: {len(self.data)}개")
            
            # 그룹 목록 업데이트
            groups = sorted(set(item.get("group_name", "") for item in self.data if item.get("group_name")))
            self.group_combo['values'] = ["(전체)"] + groups
            self.group_combo.current(0)
            
            self._render_data()
            self.status_label.config(text=f"로드 완료: {len(self.data)}개 상품", foreground="green")

        except Exception as e:
            messagebox.showerror("오류", f"엑셀 로드 실패:\n{str(e)}")
            self.status_label.config(text=f"로드 실패: {str(e)[:50]}", foreground="red")

    def _parse_excel_data(self, df):
        """엑셀 데이터 파싱"""
        self.data = []

        for idx, row in df.iterrows():
            try:
                item = {
                    "row_idx": idx,
                    # 기본 정보
                    "product_name": self._safe_str(row.get("상품명", ""))[:40],
                    "product_id": self._safe_str(row.get("불사자ID", "") or row.get("상품ID", "")),
                    "is_safe": self._parse_safe_status(row.get("안전여부", "O")),
                    "unsafe_reason": self._safe_str(row.get("위험사유", ""))[:50],
                    "group_name": self._safe_str(row.get("그룹", "") or row.get("그룹명", "")),

                    # 썸네일
                    "thumbnail_url": self._extract_image_url(self._safe_str(row.get("썸네일\n이미지", "") or row.get("메인썸네일URL", ""))),
                    
                    # 옵션 정보
                    "total_options": int(row.get("전체옵션", 0)) if pd.notna(row.get("전체옵션")) else 0,
                    "final_options": int(row.get("최종옵션", 0)) if pd.notna(row.get("최종옵션")) else 0,
                    "bait_options": int(row.get("미끼옵션", 0)) if pd.notna(row.get("미끼옵션")) else 0,
                    "main_option": self._safe_str(row.get("대표옵션", ""))[:30],
                    "selected": self._safe_str(row.get("선택", "A")).strip().upper() or "A",
                    "option_names": self._safe_str(row.get("옵션명", "") or row.get("최종옵션목록", "")),
                }

                # 옵션 파싱
                item["options"] = self._parse_options(item["option_names"])
                item["option_count"] = f"{item['final_options']}/{item['total_options']}"

                self.data.append(item)
                self.selected_options[idx] = item["selected"]

            except Exception as e:
                print(f"⚠️ Row {idx} 파싱 오류: {e}")
                continue

    def _safe_str(self, val) -> str:
        """안전한 문자열 변환"""
        if pd.isna(val):
            return ""
        return str(val).strip()

    def _parse_safe_status(self, val) -> bool:
        """안전여부 파싱"""
        if pd.isna(val):
            return True
        val_str = str(val).strip().upper()
        return val_str in ["O", "안전", "TRUE", "1"]

    def _extract_image_url(self, formula) -> str:
        """=IMAGE("url") 에서 URL 추출"""
        if not formula:
            return ""
        formula = str(formula).strip()
        if formula.startswith('=IMAGE("') and formula.endswith('")'):
            return formula[8:-2]
        if formula.startswith('http'):
            return formula
        return ""

    def _parse_options(self, option_names) -> List[Dict]:
        """옵션명 파싱"""
        options = []
        if not option_names:
            return options

        lines = str(option_names).strip().split('\n')

        for i, line in enumerate(lines):
            line = line.strip()
            if not line:
                continue

            # "A. 옵션명(가격)" 형태 파싱
            if '. ' in line:
                parts = line.split('. ', 1)
                label = parts[0].strip()
                name = parts[1].strip() if len(parts) > 1 else ""
            else:
                label = chr(ord('A') + i) if i < 26 else str(i + 1)
                name = line

            options.append({
                "label": label,
                "name": name
            })

        return options

    def _apply_filter(self):
        """필터 적용하여 재렌더링"""
        self._render_data()

    def _render_data(self):
        """데이터 렌더링 (필터 적용)"""
        for widget in self.scrollable_frame.winfo_children():
            widget.destroy()

        self.option_frames = {}

        if not self.data:
            ttk.Label(self.scrollable_frame, text="데이터 없음", 
                     font=("맑은 고딕", 14)).pack(pady=50)
            return

        # 필터링
        filtered_data = []
        selected_group = self.group_combo.get() if hasattr(self, 'group_combo') else "(전체)"
        
        for item in self.data:
            # 안전 필터
            if item["is_safe"] and not self.filter_safe.get():
                continue
            if not item["is_safe"] and not self.filter_unsafe.get():
                continue
            
            # 그룹 필터
            if selected_group and selected_group != "(전체)":
                if item.get("group_name") != selected_group:
                    continue
            
            filtered_data.append(item)

        if not filtered_data:
            ttk.Label(self.scrollable_frame, text="필터 조건에 맞는 상품이 없습니다", 
                     font=("맑은 고딕", 12), foreground="gray").pack(pady=50)
            return

        # 헤더
        self._create_header()

        # 데이터 행
        for item in filtered_data:
            self._create_row(item)

        self.status_label.config(text=f"표시: {len(filtered_data)}개 / 전체: {len(self.data)}개", foreground="blue")

    def _create_header(self):
        """헤더 생성"""
        header_frame = tk.Frame(self.scrollable_frame, bg="#4472C4")
        header_frame.pack(fill=tk.X, pady=(0, 2))

        headers = [
            ("썸네일", 100),
            ("옵션 선택", 400),
            ("상품명", 250),
            ("안전", 50),
            ("옵션수", 70),
            ("미끼", 50),
            ("그룹", 100),
        ]

        for text, width in headers:
            lbl = tk.Label(
                header_frame,
                text=text,
                width=width // 8,
                bg="#4472C4",
                fg="white",
                font=("맑은 고딕", 9, "bold"),
                pady=5
            )
            lbl.pack(side=tk.LEFT, padx=1)

    def _create_row(self, item):
        """데이터 행 생성"""
        row_idx = item["row_idx"]
        bg_color = "#C8E6C9" if item.get("is_safe", True) else "#FFCDD2"

        row_frame = tk.Frame(self.scrollable_frame, bg=bg_color, relief="solid", bd=1)
        row_frame.pack(fill=tk.X, pady=1)

        # 1. 썸네일
        thumb_frame = tk.Frame(row_frame, width=100, height=90, bg=bg_color)
        thumb_frame.pack(side=tk.LEFT, padx=1, pady=2)
        thumb_frame.pack_propagate(False)
        self._render_thumbnail(thumb_frame, item, bg_color)

        # 2. 옵션 선택
        option_frame = tk.Frame(row_frame, width=400, height=90, bg=bg_color)
        option_frame.pack(side=tk.LEFT, padx=1, pady=2)
        option_frame.pack_propagate(False)
        self._render_options(option_frame, item, row_idx, bg_color)

        # 3. 상품명
        name_frame = tk.Frame(row_frame, width=250, height=90, bg=bg_color)
        name_frame.pack(side=tk.LEFT, padx=1, pady=2)
        name_frame.pack_propagate(False)
        tk.Label(name_frame, text=item["product_name"], bg=bg_color,
                font=("맑은 고딕", 9), wraplength=240, justify=tk.LEFT).pack(expand=True)

        # 4. 안전 여부
        safe_frame = tk.Frame(row_frame, width=50, height=90, bg=bg_color)
        safe_frame.pack(side=tk.LEFT, padx=1, pady=2)
        safe_frame.pack_propagate(False)
        safe_text = "O" if item["is_safe"] else "X"
        safe_color = "#4CAF50" if item["is_safe"] else "#F44336"
        tk.Label(safe_frame, text=safe_text, bg=bg_color, fg=safe_color,
                font=("맑은 고딕", 16, "bold")).pack(expand=True)

        # 5. 옵션수
        count_frame = tk.Frame(row_frame, width=70, height=90, bg=bg_color)
        count_frame.pack(side=tk.LEFT, padx=1, pady=2)
        count_frame.pack_propagate(False)
        tk.Label(count_frame, text=item["option_count"], bg=bg_color,
                font=("맑은 고딕", 9)).pack(expand=True)

        # 6. 미끼옵션
        bait_frame = tk.Frame(row_frame, width=50, height=90, bg=bg_color)
        bait_frame.pack(side=tk.LEFT, padx=1, pady=2)
        bait_frame.pack_propagate(False)
        bait_text = str(item["bait_options"])
        bait_color = "#F44336" if item["bait_options"] > 0 else "#757575"
        tk.Label(bait_frame, text=bait_text, bg=bg_color, fg=bait_color,
                font=("맑은 고딕", 12, "bold")).pack(expand=True)

        # 7. 그룹명
        group_frame = tk.Frame(row_frame, width=100, height=90, bg=bg_color)
        group_frame.pack(side=tk.LEFT, padx=1, pady=2)
        group_frame.pack_propagate(False)
        tk.Label(group_frame, text=item["group_name"], bg=bg_color,
                font=("맑은 고딕", 8), wraplength=90).pack(expand=True)

    def _render_thumbnail(self, frame, item, bg_color):
        """썸네일 렌더링"""
        thumb_label = tk.Label(frame, text="[이미지]", bg=bg_color, font=("맑은 고딕", 8), fg="gray")
        thumb_label.pack(expand=True)

        if PIL_AVAILABLE and item.get("thumbnail_url"):
            try:
                self._load_image(item["thumbnail_url"], thumb_label, 90, 85)
            except:
                pass

    def _render_options(self, frame, item, row_idx, bg_color):
        """옵션 선택 영역 렌더링"""
        options = item.get("options", [])
        max_display = 4

        if not options:
            tk.Label(frame, text="옵션 없음", bg=bg_color, fg="gray",
                    font=("맑은 고딕", 9)).pack(expand=True)
            return

        for i, opt in enumerate(options[:max_display]):
            is_selected = (self.selected_options.get(row_idx, "A") == opt["label"])

            opt_frame = tk.Frame(
                frame,
                width=90, height=80,
                bg="#2196F3" if is_selected else "#E0E0E0",
                relief="solid",
                bd=2 if is_selected else 1,
                cursor="hand2"
            )
            opt_frame.pack(side=tk.LEFT, padx=2, pady=2)
            opt_frame.pack_propagate(False)

            opt_frame.bind("<Button-1>", lambda e, r=row_idx, o=opt["label"]: self._on_option_click(r, o))

            lbl_color = "white" if is_selected else "black"
            lbl_bg = "#2196F3" if is_selected else "#E0E0E0"

            label_widget = tk.Label(opt_frame, text=opt["label"], bg=lbl_bg, fg=lbl_color,
                                   font=("맑은 고딕", 11, "bold"))
            label_widget.pack(pady=2)
            label_widget.bind("<Button-1>", lambda e, r=row_idx, o=opt["label"]: self._on_option_click(r, o))

            name_short = opt["name"][:9] + ".." if len(opt["name"]) > 9 else opt["name"]
            name_widget = tk.Label(opt_frame, text=name_short, bg=lbl_bg, fg=lbl_color,
                                  font=("맑은 고딕", 8), wraplength=80)
            name_widget.pack(pady=1)
            name_widget.bind("<Button-1>", lambda e, r=row_idx, o=opt["label"]: self._on_option_click(r, o))

            self.option_frames[(row_idx, opt["label"])] = {
                "frame": opt_frame,
                "label": label_widget,
                "name": name_widget
            }

        if len(options) > max_display:
            more_label = tk.Label(frame, text=f"+{len(options)-max_display}",
                               bg="#9E9E9E", fg="white", font=("맑은 고딕", 9),
                               width=4)
            more_label.pack(side=tk.LEFT, padx=2, pady=30)

    def _on_option_click(self, row_idx, option_label):
        """옵션 클릭"""
        old_selected = self.selected_options.get(row_idx, "A")

        # 이전 선택 해제
        if (row_idx, old_selected) in self.option_frames:
            old_widgets = self.option_frames[(row_idx, old_selected)]
            old_widgets["frame"].config(bg="#E0E0E0", bd=1)
            old_widgets["label"].config(bg="#E0E0E0", fg="black")
            old_widgets["name"].config(bg="#E0E0E0", fg="black")

        # 새 선택
        if (row_idx, option_label) in self.option_frames:
            new_widgets = self.option_frames[(row_idx, option_label)]
            new_widgets["frame"].config(bg="#2196F3", bd=2)
            new_widgets["label"].config(bg="#2196F3", fg="white")
            new_widgets["name"].config(bg="#2196F3", fg="white")

        self.selected_options[row_idx] = option_label
        self.status_label.config(text=f"선택 변경: Row {row_idx} → {option_label}", foreground="green")

    def _load_image(self, url, label, width, height):
        """이미지 로드"""
        try:
            if url in self.image_cache:
                photo = self.image_cache[url]
            else:
                response = requests.get(url, timeout=5)
                img = Image.open(BytesIO(response.content))
                img = img.resize((width, height), Image.Resampling.LANCZOS)
                photo = ImageTk.PhotoImage(img)
                self.image_cache[url] = photo

            label.config(image=photo, text="")
            label.image = photo
        except:
            pass

    def _save_excel(self):
        """선택 변경사항을 엑셀에 저장"""
        if not self.current_file:
            messagebox.showwarning("경고", "먼저 엑셀 파일을 로드하세요")
            return

        if not PANDAS_AVAILABLE:
            messagebox.showerror("오류", "pandas가 필요합니다")
            return

        try:
            from openpyxl import load_workbook

            wb = load_workbook(self.current_file)
            
            # 상세정보 시트 우선
            if "상세정보" in wb.sheetnames:
                ws = wb["상세정보"]
            else:
                ws = wb.active

            # 선택 컬럼 찾기
            select_col = None
            for col in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col).value
                if header and str(header).strip() == "선택":
                    select_col = col
                    break

            if not select_col:
                messagebox.showerror("오류", "'선택' 컬럼을 찾을 수 없습니다")
                return

            # 변경사항 적용
            changes = 0
            for row_idx, selected in self.selected_options.items():
                # 엑셀은 1-indexed, 헤더가 1행이므로 데이터는 2행부터
                excel_row = row_idx + 2
                if excel_row <= ws.max_row:
                    old_value = ws.cell(row=excel_row, column=select_col).value
                    if str(old_value).strip() != selected:
                        ws.cell(row=excel_row, column=select_col, value=selected)
                        changes += 1

            # 저장
            wb.save(self.current_file)
            wb.close()

            messagebox.showinfo("저장 완료", f"{changes}개 옵션 선택이 저장되었습니다")
            self.status_label.config(text=f"저장 완료: {changes}개 변경", foreground="green")

        except Exception as e:
            messagebox.showerror("저장 실패", f"오류: {str(e)}")
            self.status_label.config(text=f"저장 실패: {str(e)[:50]}", foreground="red")

    def _reload_data(self):
        """현재 파일 다시 로드"""
        if self.current_file:
            self._load_excel_file(self.current_file)
        else:
            messagebox.showwarning("경고", "로드된 파일이 없습니다")


def main():
    root = tk.Tk()
    app = SimulatorGUIv31(root)
    root.mainloop()


if __name__ == "__main__":
    main()
