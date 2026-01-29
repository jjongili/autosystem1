# -*- coding: utf-8 -*-
"""
================================================================================
시뮬레이터 GUI v4 - 불사자 업로드 사전 검증 및 검수 시스템
================================================================================

[핵심 목적]
미끼옵션 제거가 제대로 동작하는지 사전 테스트하고, 수집 시 상품을 검수하여
위험상품 등록을 방지하는 통합 검수 시스템.

[주요 기능]
1. 상품명 검수
   - 중국 구매대행 위험상품 필터링
   - 위험단어 탐지 + 안전 문맥 판단 (위험단어 있어도 주변에 안전단어 있으면 안전)

2. 미끼옵션 제거 검증
   - 저가 미끼옵션 자동 필터링
   - 필터링 결과 검증 (누락/오탐 확인)

3. 대표옵션 ↔ 썸네일 매칭
   - 미끼 제거 후 선택된 대표옵션이 썸네일과 일치하는지
   - 사용자가 육안으로 빠르게 검수

4. 브랜드 가품 탐지 (가장 위험!)
   - 금지단어로 못 거르는 브랜드 단어가 제일 위험
   - 형태소 분석(Kiwi) → 일반명사 아니면 브랜드 의심 → AI 검증 요청
   - 위험 카테고리 상품은 AI 검수 강화

5. 학습 DB화 (지속적 정확도 향상)
   - 시뮬 결과 엑셀 → AI 분석 → DB 축적
   - 위험 분류 → 실제 위험이었는지?
   - 안전 분류 → 놓친 위험단어 있는지?
   - 미끼 키워드 누락/오탐 분석
   - 지속적 학습으로 적중률 향상

[워크플로우]
시뮬레이터 수집/검수 → 엑셀 저장 → 사람 수동 검수 → 불사자 API 업로드
                                              ↓
                           사용자가 변경한 메인썸네일/대표옵션 반영

[공통 모듈 (bulsaja_common.py)]
- filter_bait_options(): 미끼옵션 필터링 (v1.4: 공통키워드 가격분석)
- select_main_option(): 대표옵션 선택 (이미지우선 → 상품명매칭)
- check_product_safety(): 상품 안전성 검사 (위험단어 + 안전문맥)
- analyze_products_for_ip(): 지재권 분석 (Kiwi 형태소 분석)

[기술 스택]
- PyQt6 기반 GUI
- QThreadPool 이미지 병렬 로딩
- 3개 탭: 수집 / 검수 / 설정
================================================================================
"""

import sys
import os
import json
import hashlib
import subprocess
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any, Optional, Tuple
from urllib.request import urlopen, Request
from concurrent.futures import ThreadPoolExecutor

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QGridLayout,
    QLabel, QLineEdit, QPushButton, QComboBox, QSpinBox, QCheckBox, QRadioButton,
    QGroupBox, QTabWidget, QScrollArea, QTextEdit, QProgressBar, QTableWidget,
    QTableWidgetItem, QHeaderView, QMessageBox, QFileDialog, QSplitter, QFrame,
    QButtonGroup, QSizePolicy, QDialog, QDialogButtonBox
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal, QRunnable, QThreadPool, QObject, QSize, QTimer
from PyQt6.QtGui import QPixmap, QImage, QFont, QColor, QPainter

# 불사자 공통 모듈
try:
    from bulsaja_common import (
        BulsajaAPIClient, extract_tokens_from_browser,
        filter_bait_options, select_main_option,
        load_bait_keywords, save_bait_keywords,
        load_banned_words, load_excluded_words, check_product_safety,
        load_category_risk_settings, save_category_risk_settings,
        DEFAULT_CATEGORY_RISK_SETTINGS, get_category_risk_level,
        MARKET_IDS, DEFAULT_BAIT_KEYWORDS,
        analyze_products_for_ip, verify_ip_words_with_ai,  # 지재권 분석
        check_product_name_suspicious, batch_check_product_names,  # 상품명 검수
        load_ai_config, save_ai_config, DEFAULT_AI_CONFIG  # AI 설정
    )
    BULSAJA_API_AVAILABLE = True
except ImportError:
    BULSAJA_API_AVAILABLE = False

# 썸네일 분석 모듈
try:
    from thumbnail_analyzer import ThumbnailAnalyzer, ThumbnailScore
    THUMBNAIL_ANALYZER_AVAILABLE = True
except ImportError:
    THUMBNAIL_ANALYZER_AVAILABLE = False


# 검수 수준 옵션
CHECK_LEVELS = {
    "보통 (자동판단)": "normal",
    "엄격 (AI검수)": "strict",
    "검수제외": "skip",
}

# 엑셀 라이브러리
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font as XLFont, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# ============================================================
# 설정
# ============================================================
CONFIG_FILE = "simulator_gui_v4_config.json"
DEBUG_PORT = 9222

# 수집 조건
UPLOAD_CONDITIONS = {
    "미업로드(수집완료+수정중+검토완료)": ["0", "1", "2", "수집 완료", "수정중", "검토 완료"],
    "수집완료만": ["0", "수집 완료"],
    "수정중만": ["1", "수정중"],
    "검토완료만": ["2", "검토 완료"],
    "업로드완료(판매중)": ["3", "판매중", "업로드 완료"],
    "전체": None,
}

# 옵션 정렬
OPTION_SORT_OPTIONS = {
    "가격낮은순": "price_asc",
    "주요가격대": "price_main",
    "가격높은순": "price_desc",
}

# 상품명 처리
TITLE_OPTIONS = {
    "원마켓 상품명 그대로 사용": "original",
    "앞4개단어제외 셔플": "shuffle_skip4",
    "앞3개단어제외 셔플": "shuffle_skip3",
    "모든단어 셔플": "shuffle_all",
}

# 테이블 컬럼 설정
TABLE_COLUMNS = [
    {"key": "thumbnail", "name": "썸네일", "default": True, "width": 90},
    {"key": "option_image", "name": "옵션이미지", "default": True, "width": 90},
    {"key": "option_select", "name": "옵션선택", "default": True, "width": 400},  # A~F 버튼 (크기 증가)
    {"key": "option_names_list", "name": "옵션명", "default": True, "width": 230},  # 옵션명 목록
    {"key": "option_names_cn", "name": "중국어옵션", "default": False, "width": 230},  # 중국어 옵션명
    {"key": "name", "name": "상품명", "default": True, "width": 0},  # 0 = stretch
    {"key": "danger", "name": "위험", "default": True, "width": 45},
    {"key": "unsafe_reason", "name": "위험사유", "default": True, "width": 100},
    {"key": "options", "name": "옵션", "default": True, "width": 50},
    {"key": "bait", "name": "미끼", "default": True, "width": 45},
    {"key": "price", "name": "가격", "default": True, "width": 70},
    {"key": "price_range", "name": "가격범위(CNY)", "default": False, "width": 100},
    {"key": "category", "name": "카테고리", "default": False, "width": 150},
    {"key": "main_option", "name": "대표옵션", "default": False, "width": 100},
    {"key": "group", "name": "그룹", "default": True, "width": 80},
    {"key": "id", "name": "ID", "default": False, "width": 200},
]


def load_config() -> dict:
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            pass
    return {}


def save_config(config: dict) -> bool:
    try:
        with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
        return True
    except:
        return False


# ============================================================
# 이미지 캐시 (메모리 캐싱)
# ============================================================
class ImageCache:
    """URL+사이즈 해시 기반 메모리 캐시"""

    def __init__(self, max_size: int = 300):
        self._cache: Dict[str, QPixmap] = {}
        self._max_size = max_size

    def get(self, url: str, size: tuple = None) -> Optional[QPixmap]:
        cache_key = f"{url}_{size[0]}x{size[1]}" if size else url
        key = hashlib.md5(cache_key.encode()).hexdigest()
        return self._cache.get(key)

    def put(self, url: str, pixmap: QPixmap, size: tuple = None):
        if len(self._cache) >= self._max_size:
            oldest_key = next(iter(self._cache))
            del self._cache[oldest_key]
        cache_key = f"{url}_{size[0]}x{size[1]}" if size else url
        key = hashlib.md5(cache_key.encode()).hexdigest()
        self._cache[key] = pixmap

    def clear(self):
        self._cache.clear()


image_cache = ImageCache()


# ============================================================
# 이미지 다운로드 워커 (QThreadPool용)
# ============================================================
class WorkerSignals(QObject):
    """워커 시그널"""
    finished = pyqtSignal(str, QPixmap)  # (product_id, pixmap)
    error = pyqtSignal(str, str)  # (product_id, error_msg)


class ImageDownloadWorker(QRunnable):
    """
    백그라운드 이미지 다운로드 워커
    - QThreadPool에서 병렬 실행
    - 다운로드 후 리사이징하여 메모리 최적화
    """

    def __init__(self, product_id: str, url: str, size: QSize = QSize(60, 60)):
        super().__init__()
        self.product_id = product_id
        self.url = url
        self.size = size
        self.signals = WorkerSignals()

    def run(self):
        try:
            # 캐시 확인 (URL + 사이즈로 캐시)
            cache_size = (self.size.width(), self.size.height())
            cached = image_cache.get(self.url, cache_size)
            if cached:
                self.signals.finished.emit(self.product_id, cached)
                return

            # HTTP 요청 (헤더 필수 - CDN 차단 방지)
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
                'Referer': 'https://www.bulsaja.com/'
            }
            req = Request(self.url, headers=headers)
            with urlopen(req, timeout=10) as response:
                data = response.read()

            # QImage 로드
            image = QImage()
            if not image.loadFromData(data):
                raise ValueError("이미지 로드 실패")

            # 리사이징 (SmoothTransformation)
            # 이미지가 작으면 확대, 크면 축소 (영역에 맞게)
            target_w, target_h = self.size.width(), self.size.height()
            img_w, img_h = image.width(), image.height()

            # 비율 유지하면서 영역에 맞는 크기 계산
            scale = min(target_w / img_w, target_h / img_h) if img_w > 0 and img_h > 0 else 1
            new_w = int(img_w * scale)
            new_h = int(img_h * scale)

            scaled = image.scaled(
                new_w, new_h,
                Qt.AspectRatioMode.KeepAspectRatio,
                Qt.TransformationMode.SmoothTransformation
            )

            pixmap = QPixmap.fromImage(scaled)
            image_cache.put(self.url, pixmap, cache_size)
            self.signals.finished.emit(self.product_id, pixmap)

        except Exception as e:
            self.signals.error.emit(self.product_id, str(e))


# ============================================================
# 상품 수집 워커 (QThread)
# ============================================================
class CollectWorker(QThread):
    """상품 수집 백그라운드 워커 + 검수 로직"""
    progress = pyqtSignal(int, int, str)  # current, total, message
    log = pyqtSignal(str)
    finished_signal = pyqtSignal(bool, str, list)  # success, message, data

    def __init__(self, api_client, groups: List[str], settings: dict):
        super().__init__()
        self.api_client = api_client
        self.groups = groups
        self.settings = settings
        self.is_running = True

    def stop(self):
        self.is_running = False

    def run(self):
        try:
            all_products = []
            total_groups = len(self.groups)

            # 검수 설정 로드
            bait_keywords = self.settings.get('bait_keywords', [])
            excluded_words = self.settings.get('excluded_words', set())
            check_level = self.settings.get('check_level', 'normal')
            option_count = self.settings.get('option_count', 10)

            for idx, group_name in enumerate(self.groups):
                if not self.is_running:
                    break

                self.log.emit(f"📁 [{idx+1}/{total_groups}] {group_name} 수집 중...")
                self.progress.emit(idx + 1, total_groups, group_name)

                try:
                    # 상품 조회 (0이면 전체)
                    max_products = self.settings.get('max_products', 0)
                    status_filters = self.settings.get('status_filters')

                    # 0이면 전체 조회 (100000개까지)
                    limit = max_products if max_products > 0 else 100000

                    products, total = self.api_client.get_products_by_group(
                        group_name, 0, limit, status_filters
                    )

                    if products:
                        # 카테고리 수집 옵션
                        fetch_category = self.settings.get('fetch_category', False)

                        if fetch_category:
                            self.log.emit(f"   📂 카테고리 수집 중... ({len(products)}개)")

                        # 각 상품 검수 처리
                        for i, p in enumerate(products):
                            p['group_name'] = group_name

                            # 상품 상세에서 카테고리/옵션 정보 가져오기 (get_product_detail 사용)
                            if fetch_category:
                                try:
                                    prod_id = p.get('ID', '') or p.get('id', '')
                                    product_detail = self.api_client.get_product_detail(prod_id)
                                    if product_detail:
                                        # 카테고리 (항상 덮어쓰기)
                                        p['uploadCategory'] = product_detail.get('uploadCategory', {})
                                        # 썸네일 (항상 덮어쓰기 - 상세 API가 더 정확)
                                        p['uploadThumbnails'] = product_detail.get('uploadThumbnails', [])
                                        # SKU (항상 덮어쓰기 - 옵션 이미지에 필수!)
                                        p['uploadSkus'] = product_detail.get('uploadSkus', [])
                                        # SKU 속성
                                        p['uploadSkuProps'] = product_detail.get('uploadSkuProps', {})
                                        # 가격 정보
                                        p['uploadCommonSalePrice'] = product_detail.get('uploadCommonSalePrice', 0)
                                        # 추가 정보
                                        p['uploadCommonProductName'] = product_detail.get('uploadCommonProductName', p.get('uploadCommonProductName', ''))
                                        p['uploadCommonTags'] = product_detail.get('uploadCommonTags', [])
                                        p['base_price'] = product_detail.get('base_price', {})
                                    # 10개마다 진행 로그
                                    if (i + 1) % 10 == 0:
                                        self.log.emit(f"      {i + 1}/{len(products)} 상세 조회...")
                                except Exception as uf_e:
                                    self.log.emit(f"      ⚠️ {prod_id[:10]}... 상세 조회 실패")

                            self._inspect_product(p, bait_keywords, excluded_words, check_level, option_count)
                        all_products.extend(products)

                        # 안전/위험 카운트
                        safe_count = sum(1 for p in products if p.get('is_safe', True))
                        unsafe_count = len(products) - safe_count
                        self.log.emit(f"   ✅ {len(products)}개 수집 (안전:{safe_count} 위험:{unsafe_count})")
                    else:
                        self.log.emit(f"   ⏭️ 상품 없음")

                except Exception as e:
                    self.log.emit(f"   ❌ 오류: {e}")

            self.finished_signal.emit(True, f"수집 완료: {len(all_products)}개", all_products)

        except Exception as e:
            self.finished_signal.emit(False, str(e), [])

    def _inspect_product(self, product: dict, bait_keywords: list,
                         excluded_words: set, check_level: str, option_count: int):
        """상품 검수 (미끼필터, 대표옵션, 안전검사)"""
        try:
            product_name = product.get('uploadCommonProductName', '')

            # 썸네일 정보 저장
            thumbnails = product.get('uploadThumbnails', []) or []
            product['all_thumbnails'] = thumbnails
            product['thumbnail_url'] = thumbnails[0] if thumbnails else ''

            # SKU 정보 저장
            upload_skus = product.get('uploadSkus', []) or []
            product['all_skus'] = upload_skus

            # 가격 정보
            prices = [float(s.get('_origin_price', 0) or 0) for s in upload_skus if s.get('_origin_price')]
            if prices:
                product['min_price_cny'] = min(prices)
                product['max_price_cny'] = max(prices)
            else:
                product['min_price_cny'] = 0
                product['max_price_cny'] = 0

            # uploadCategory에서 카테고리명 추출 (dict 구조)
            category_name = ''
            upload_cat = product.get('uploadCategory')
            if isinstance(upload_cat, dict):
                # ss_category, esm_category, est_category 등에서 name 추출
                for key in ['ss_category', 'esm_category', 'est_category', 'est_global_category']:
                    cat_obj = upload_cat.get(key)
                    if isinstance(cat_obj, dict) and cat_obj.get('name'):
                        category_name = cat_obj['name']
                        break
            elif isinstance(upload_cat, str):
                category_name = upload_cat

            # 1. 상품명 안전 검사
            if BULSAJA_API_AVAILABLE:
                safety = check_product_safety(
                    product_name, excluded_words,
                    check_level=check_level,
                    category_name=category_name
                )
                product['is_safe'] = safety['is_safe']
                product['unsafe_keywords'] = safety.get('all_found', [])

                # 위험 사유 포맷팅
                if not safety['is_safe']:
                    categories = []
                    cats = safety.get('categories', {})
                    if cats.get('adult'):
                        categories.append(f"성인:{','.join(cats['adult'][:2])}")
                    if cats.get('medical'):
                        categories.append(f"의료:{','.join(cats['medical'][:2])}")
                    if cats.get('child'):
                        categories.append(f"유아:{','.join(cats['child'][:2])}")
                    if cats.get('prohibited'):
                        categories.append(f"금지:{','.join(cats['prohibited'][:2])}")
                    if cats.get('brand'):
                        categories.append(f"브랜드:{','.join(cats['brand'][:2])}")
                    product['unsafe_reason'] = ' / '.join(categories) if categories else '위험키워드감지'
                else:
                    product['unsafe_reason'] = ''
            else:
                product['is_safe'] = True
                product['unsafe_reason'] = ''

            # 1.5 상품명 의심 단어 검수 (패턴 기반, 빠름)
            if BULSAJA_API_AVAILABLE:
                name_check = check_product_name_suspicious(product_name, use_ai=False)
                product['name_check_result'] = name_check

            # 2. SKU 정보 처리
            upload_skus = product.get('uploadSkus', []) or product.get('uploadCommonOptions', []) or []
            product['total_options'] = len(upload_skus)

            # 옵션 관련 필드 기본 초기화
            product['option_images'] = {}
            product['option_prices'] = {}
            product['option_names'] = {}
            product['option_names_cn'] = {}
            product['final_option_list'] = []
            product['valid_options'] = 0
            product['bait_options'] = 0
            product['bait_option_list'] = []
            labels = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
            for idx, sku in enumerate(upload_skus[:26]):  # 최대 26개
                opt_image = sku.get('urlRef', '') or sku.get('optionImage', '') or sku.get('image', '')
                opt_price = float(sku.get('_origin_price', 0) or sku.get('price', 0) or 0)
                opt_name = sku.get('text_ko', '') or sku.get('text', '') or sku.get('optionName', '')
                opt_name_cn = sku.get('_text', '')  # 중국어 원본

                label = labels[idx]
                # 빈 문자열은 저장하지 않음 (폴백 로직이 작동하도록)
                if opt_image:
                    product['option_images'][label] = opt_image
                product['option_prices'][label] = opt_price
                product['option_names'][label] = opt_name
                if opt_name_cn:
                    product['option_names_cn'][label] = opt_name_cn
                # final_option_list에도 추가 (기본값)
                if opt_name:
                    product['final_option_list'].append(f"{opt_name[:20]}({opt_price:.1f})")

            if upload_skus and BULSAJA_API_AVAILABLE:
                # 미끼옵션 필터링
                valid_skus, bait_skus = filter_bait_options(upload_skus, bait_keywords)

                product['valid_options'] = len(valid_skus)
                product['bait_options'] = len(bait_skus)
                product['bait_option_list'] = []

                # 미끼 옵션 정보
                for bait_sku in bait_skus[:5]:
                    opt_name = bait_sku.get('text_ko', '') or bait_sku.get('optionName', '') or ''
                    bait_price = bait_sku.get('_origin_price', 0) or bait_sku.get('price', 0)
                    product['bait_option_list'].append(f"{opt_name[:15]}({bait_price})")

                # 대표옵션 선택 (이미지 있는 최저가)
                if valid_skus:
                    main_idx, main_method = select_main_option(product_name, valid_skus)
                    main_sku = valid_skus[main_idx]

                    product['main_option_name'] = main_sku.get('text_ko', '') or main_sku.get('optionName', '')
                    product['main_option_method'] = main_method
                    product['main_option_image'] = main_sku.get('urlRef', '') or main_sku.get('optionImage', '')

                    # 대표옵션 가격 기준 최종 옵션 목록
                    def safe_price(val):
                        try:
                            return float(val) if val else 0.0
                        except:
                            return 0.0

                    main_price = safe_price(main_sku.get('_origin_price') or main_sku.get('price'))

                    if option_count > 0:
                        eligible = [s for s in valid_skus if safe_price(s.get('_origin_price') or s.get('price')) >= main_price]
                        eligible.sort(key=lambda x: safe_price(x.get('_origin_price') or x.get('price')))
                        final_skus = eligible[:option_count]
                    else:
                        final_skus = valid_skus

                    product['final_options'] = len(final_skus)
                    product['final_option_list'] = []
                    product['option_images'] = {}  # {"A": url, "B": url, ...}
                    product['option_prices'] = {}  # {"A": price, "B": price, ...}
                    product['option_names'] = {}   # {"A": name, "B": name, ...}
                    product['option_names_cn'] = {}  # {"A": 중국어, "B": 중국어, ...}

                    labels = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
                    for idx, sku in enumerate(final_skus):
                        opt_name = sku.get('text_ko', '') or sku.get('text', '') or sku.get('optionName', '')
                        opt_name_cn = sku.get('_text', '')  # 중국어 원본

                        opt_price = safe_price(sku.get('_origin_price') or sku.get('price'))
                        opt_image = sku.get('urlRef', '') or sku.get('optionImage', '') or sku.get('image', '')

                        product['final_option_list'].append(f"{opt_name[:20]}({opt_price:.1f})")

                        # A, B, C... 라벨로 옵션 이미지/가격/이름 저장
                        label = labels[idx] if idx < len(labels) else str(idx + 1)
                        product['option_images'][label] = opt_image
                        product['option_prices'][label] = opt_price
                        product['option_names'][label] = opt_name
                        if opt_name_cn:
                            product['option_names_cn'][label] = opt_name_cn

        except Exception as e:
            product['is_safe'] = True
            product['unsafe_reason'] = f"검수오류: {str(e)[:30]}"


# ============================================================
# 이미지 라벨 (플레이스홀더 포함)
# ============================================================
class ImageLabel(QLabel):
    """플레이스홀더 지원 이미지 라벨"""
    clicked = pyqtSignal(str)  # 클릭 시 이미지 URL 전달

    def __init__(self, size: int = 60, parent=None):
        super().__init__(parent)
        self.setFixedSize(size, size)
        self.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.setCursor(Qt.CursorShape.PointingHandCursor)  # 클릭 가능 커서
        self._image_url = ""
        self.show_placeholder()

    def set_image_url(self, url: str):
        """이미지 URL 저장"""
        self._image_url = url

    def get_image_url(self) -> str:
        return self._image_url

    def mousePressEvent(self, event):
        """클릭 시 시그널 발생"""
        if self._image_url:
            self.clicked.emit(self._image_url)
        super().mousePressEvent(event)

    def show_placeholder(self):
        self.setText("...")
        self.setStyleSheet("""
            QLabel {
                background-color: #E0E0E0;
                border: 1px solid #BDBDBD;
                border-radius: 4px;
                color: #757575;
                font-size: 10px;
            }
        """)

    def show_error(self):
        self.setText("X")
        self.setStyleSheet("""
            QLabel {
                background-color: #FFCDD2;
                border: 1px solid #EF9A9A;
                border-radius: 4px;
                color: #C62828;
            }
        """)

    def set_image(self, pixmap: QPixmap):
        self.setText("")
        self.setPixmap(pixmap)
        self.setStyleSheet("""
            QLabel {
                background-color: white;
                border: 1px solid #BDBDBD;
                border-radius: 4px;
            }
        """)


# ============================================================
# 메인 윈도우
# ============================================================
class SimulatorGUIv4(QMainWindow):
    """시뮬레이터 GUI v4 - PyQt6 최적화 버전"""

    def __init__(self):
        super().__init__()
        self.setWindowTitle("불사자 시뮬레이터 v4 (PyQt6)")
        self.setMinimumSize(1400, 900)

        # 상태
        self.api_client = None
        self.products = []
        self.selected_options = {}
        self.image_labels = {}  # {product_id: ImageLabel}

        # QThreadPool 설정
        self.thread_pool = QThreadPool.globalInstance()
        self.thread_pool.setMaxThreadCount(10)

        # 설정 로드
        self.config = load_config()

        self._build_ui()

    def _build_ui(self):
        """UI 구성"""
        central = QWidget()
        self.setCentralWidget(central)
        main_layout = QVBoxLayout(central)
        main_layout.setContentsMargins(5, 5, 5, 5)

        # 탭 위젯
        self.tabs = QTabWidget()
        self.tabs.setStyleSheet("""
            QTabWidget::pane {
                border: 1px solid #BDBDBD;
                border-radius: 4px;
            }
            QTabBar::tab {
                padding: 8px 20px;
                font-weight: bold;
            }
            QTabBar::tab:selected {
                background-color: #E3F2FD;
            }
        """)

        # 각 탭 생성
        self.collection_tab = QWidget()
        self.review_tab = QWidget()
        self.settings_tab = QWidget()

        self.tabs.addTab(self.collection_tab, "  📥 수집  ")
        self.tabs.addTab(self.review_tab, "  🔍 검수  ")
        self.tabs.addTab(self.settings_tab, "  ⚙️ 설정  ")

        self._build_collection_tab()
        self._build_review_tab()
        self._build_settings_tab()

        main_layout.addWidget(self.tabs)

        # 기본 탭: 검수
        self.tabs.setCurrentIndex(1)

    def _build_collection_tab(self):
        """수집 탭 UI"""
        layout = QVBoxLayout(self.collection_tab)

        # 스크롤 영역
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll_widget = QWidget()
        scroll_layout = QVBoxLayout(scroll_widget)

        # === API 연결 ===
        api_group = QGroupBox("🔑 API 연결")
        api_layout = QVBoxLayout(api_group)

        api_row1 = QHBoxLayout()
        self.chrome_btn = QPushButton("🌐 크롬 열기")
        self.chrome_btn.clicked.connect(self._open_chrome)
        api_row1.addWidget(self.chrome_btn)

        self.token_btn = QPushButton("🔑 토큰 추출")
        self.token_btn.clicked.connect(self._extract_token)
        api_row1.addWidget(self.token_btn)

        self.connect_btn = QPushButton("🔗 연결")
        self.connect_btn.clicked.connect(self._connect_api)
        api_row1.addWidget(self.connect_btn)

        api_row1.addWidget(QLabel("포트:"))
        self.port_spin = QSpinBox()
        self.port_spin.setRange(1000, 65535)
        self.port_spin.setValue(9222)
        self.port_spin.setFixedWidth(80)
        api_row1.addWidget(self.port_spin)

        self.api_status = QLabel("⚫ 미연결")
        self.api_status.setStyleSheet("font-weight: bold;")
        api_row1.addWidget(self.api_status)
        api_row1.addStretch()
        api_layout.addLayout(api_row1)

        api_row2 = QHBoxLayout()
        api_row2.addWidget(QLabel("Access:"))
        self.access_input = QLineEdit()
        self.access_input.setPlaceholderText("토큰 추출 또는 수동 입력")
        api_row2.addWidget(self.access_input)
        api_row2.addWidget(QLabel("Refresh:"))
        self.refresh_input = QLineEdit()
        api_row2.addWidget(self.refresh_input)
        api_layout.addLayout(api_row2)

        scroll_layout.addWidget(api_group)

        # === 그룹 설정 ===
        group_group = QGroupBox("📁 마켓그룹 설정")
        group_layout = QVBoxLayout(group_group)

        group_row1 = QHBoxLayout()
        group_row1.addWidget(QLabel("그룹당 최대:"))
        self.max_products_spin = QSpinBox()
        self.max_products_spin.setRange(0, 100000)  # 0 = 무제한
        self.max_products_spin.setValue(0)  # 기본값 무제한
        self.max_products_spin.setFixedWidth(80)
        self.max_products_spin.setSpecialValueText("전체")  # 0일 때 "전체" 표시
        self.max_products_spin.setToolTip("0 = 전체 (무제한)")
        group_row1.addWidget(self.max_products_spin)

        group_row1.addWidget(QLabel("작업 그룹:"))
        self.work_groups_input = QLineEdit("1-5")
        self.work_groups_input.setFixedWidth(100)
        group_row1.addWidget(self.work_groups_input)
        group_row1.addWidget(QLabel("(예: 1-5 또는 1,3,5)"))

        self.load_groups_btn = QPushButton("📥 그룹 로드")
        self.load_groups_btn.clicked.connect(self._load_groups)
        group_row1.addWidget(self.load_groups_btn)
        group_row1.addStretch()
        group_layout.addLayout(group_row1)

        self.groups_text = QTextEdit()
        self.groups_text.setMaximumHeight(60)
        self.groups_text.setPlaceholderText("그룹 목록 (쉼표 구분)")
        group_layout.addWidget(self.groups_text)

        scroll_layout.addWidget(group_group)

        # === 수집 조건 ===
        condition_group = QGroupBox("📋 수집 조건")
        condition_layout = QHBoxLayout(condition_group)

        condition_layout.addWidget(QLabel("수집조건:"))
        self.condition_combo = QComboBox()
        self.condition_combo.addItems(list(UPLOAD_CONDITIONS.keys()))
        self.condition_combo.setFixedWidth(280)
        condition_layout.addWidget(self.condition_combo)

        condition_layout.addWidget(QLabel("수집수:"))
        self.collect_count_spin = QSpinBox()
        self.collect_count_spin.setRange(1, 99999)
        self.collect_count_spin.setValue(9999)
        self.collect_count_spin.setFixedWidth(80)
        condition_layout.addWidget(self.collect_count_spin)
        condition_layout.addStretch()

        scroll_layout.addWidget(condition_group)

        # === 옵션 설정 ===
        option_group = QGroupBox("⚙️ 옵션 설정")
        option_layout = QHBoxLayout(option_group)

        option_layout.addWidget(QLabel("옵션수:"))
        self.option_count_spin = QSpinBox()
        self.option_count_spin.setRange(1, 100)
        self.option_count_spin.setValue(10)
        self.option_count_spin.setFixedWidth(60)
        option_layout.addWidget(self.option_count_spin)

        option_layout.addWidget(QLabel("표시:"))
        self.option_display_combo = QComboBox()
        self.option_display_combo.addItems(['3개', '4개', '5개', '6개'])
        self.option_display_combo.setCurrentText('6개')
        self.option_display_combo.setFixedWidth(55)
        self.option_display_combo.currentTextChanged.connect(self._on_option_display_changed)
        option_layout.addWidget(self.option_display_combo)

        option_layout.addWidget(QLabel("옵션정렬:"))
        self.option_sort_combo = QComboBox()
        self.option_sort_combo.addItems(list(OPTION_SORT_OPTIONS.keys()))
        option_layout.addWidget(self.option_sort_combo)

        option_layout.addWidget(QLabel("상품명:"))
        self.title_combo = QComboBox()
        self.title_combo.addItems(list(TITLE_OPTIONS.keys()))
        self.title_combo.setCurrentText("앞3개단어제외 셔플")
        option_layout.addWidget(self.title_combo)
        option_layout.addStretch()

        scroll_layout.addWidget(option_group)

        # === 검수 설정 ===
        inspect_group = QGroupBox("🔍 검수 설정")
        inspect_layout = QHBoxLayout(inspect_group)

        inspect_layout.addWidget(QLabel("검수수준:"))
        self.check_level_combo = QComboBox()
        self.check_level_combo.addItems(list(CHECK_LEVELS.keys()))
        self.check_level_combo.setFixedWidth(150)
        self.check_level_combo.setToolTip(
            "보통: 프로그램 자동 판단 (안전 컨텍스트 적용)\n"
            "엄격: AI 확인 필수\n"
            "검수제외: 항상 안전으로 처리"
        )
        inspect_layout.addWidget(self.check_level_combo)

        self.filter_bait_check = QCheckBox("미끼옵션 필터링")
        self.filter_bait_check.setChecked(True)
        inspect_layout.addWidget(self.filter_bait_check)

        self.show_unsafe_only_check = QCheckBox("위험상품만 표시")
        self.show_unsafe_only_check.setChecked(False)
        inspect_layout.addWidget(self.show_unsafe_only_check)

        # 구분선
        sep = QFrame()
        sep.setFrameShape(QFrame.Shape.VLine)
        sep.setStyleSheet("color: #ccc;")
        inspect_layout.addWidget(sep)

        # 카테고리 수집 (직접 API)
        self.fetch_category_check = QCheckBox("카테고리 수집")
        self.fetch_category_check.setChecked(True)
        self.fetch_category_check.setToolTip("uploadCategory 정보 수집 (카테고리별 위험도 판단에 필요)")
        inspect_layout.addWidget(self.fetch_category_check)

        inspect_layout.addStretch()
        scroll_layout.addWidget(inspect_group)

        # === 마진 설정 ===
        margin_group = QGroupBox("💰 마진 설정")
        margin_layout = QHBoxLayout(margin_group)

        margin_layout.addWidget(QLabel("환율:"))
        self.exchange_spin = QSpinBox()
        self.exchange_spin.setRange(100, 500)
        self.exchange_spin.setValue(215)
        self.exchange_spin.setFixedWidth(60)
        margin_layout.addWidget(self.exchange_spin)

        margin_layout.addWidget(QLabel("최저가:"))
        self.min_price_spin = QSpinBox()
        self.min_price_spin.setRange(0, 1000000)
        self.min_price_spin.setValue(30000)
        self.min_price_spin.setFixedWidth(80)
        margin_layout.addWidget(self.min_price_spin)

        margin_layout.addWidget(QLabel("최고가:"))
        self.max_price_spin = QSpinBox()
        self.max_price_spin.setRange(0, 100000000)
        self.max_price_spin.setValue(100000000)
        self.max_price_spin.setFixedWidth(100)
        margin_layout.addWidget(self.max_price_spin)
        margin_layout.addStretch()

        scroll_layout.addWidget(margin_group)

        # === 실행 버튼 ===
        action_group = QGroupBox("🚀 실행")
        action_layout = QHBoxLayout(action_group)

        self.collect_btn = QPushButton("📥 수집 시작")
        self.collect_btn.setFixedHeight(40)
        self.collect_btn.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                font-weight: bold;
                font-size: 14px;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)
        self.collect_btn.clicked.connect(self._start_collect)
        action_layout.addWidget(self.collect_btn)

        self.stop_btn = QPushButton("⏹️ 중지")
        self.stop_btn.setFixedHeight(40)
        self.stop_btn.setEnabled(False)
        self.stop_btn.clicked.connect(self._stop_collect)
        action_layout.addWidget(self.stop_btn)

        scroll_layout.addWidget(action_group)

        # === 진행 상황 ===
        progress_group = QGroupBox("📊 진행 상황")
        progress_layout = QVBoxLayout(progress_group)

        self.progress_bar = QProgressBar()
        progress_layout.addWidget(self.progress_bar)

        self.progress_label = QLabel("대기 중...")
        progress_layout.addWidget(self.progress_label)

        scroll_layout.addWidget(progress_group)

        # === 로그 ===
        log_group = QGroupBox("📋 로그")
        log_layout = QVBoxLayout(log_group)
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setFont(QFont("Consolas", 11))
        self.log_text.setMaximumHeight(200)
        log_layout.addWidget(self.log_text)

        scroll_layout.addWidget(log_group)
        scroll_layout.addStretch()

        scroll.setWidget(scroll_widget)
        layout.addWidget(scroll)

    def _build_review_tab(self):
        """검수 탭 UI - 상품 테이블 + 이미지 그리드"""
        layout = QVBoxLayout(self.review_tab)

        # 상단 컨트롤
        control_layout = QHBoxLayout()

        self.open_excel_btn = QPushButton("📂 엑셀 열기")
        self.open_excel_btn.clicked.connect(self._open_excel)
        control_layout.addWidget(self.open_excel_btn)

        self.save_excel_btn = QPushButton("💾 엑셀 저장")
        self.save_excel_btn.clicked.connect(self._save_excel)
        control_layout.addWidget(self.save_excel_btn)

        self.apply_btn = QPushButton("📤 불사자 반영")
        self.apply_btn.clicked.connect(self._apply_to_bulsaja)
        control_layout.addWidget(self.apply_btn)

        # 구분선
        sep1 = QFrame()
        sep1.setFrameShape(QFrame.Shape.VLine)
        sep1.setStyleSheet("color: #ccc;")
        control_layout.addWidget(sep1)

        # 이미지 로드 버튼
        self.load_images_btn = QPushButton("🖼️ 이미지 로드")
        self.load_images_btn.setToolTip("현재 페이지 썸네일/옵션 이미지 로드")
        self.load_images_btn.clicked.connect(self._load_current_page_images)
        control_layout.addWidget(self.load_images_btn)

        # 컬럼 설정 버튼
        self.column_setting_btn = QPushButton("⚙️ 컬럼설정")
        self.column_setting_btn.clicked.connect(self._show_column_settings)
        control_layout.addWidget(self.column_setting_btn)

        # 구분선2
        sep2 = QFrame()
        sep2.setFrameShape(QFrame.Shape.VLine)
        sep2.setStyleSheet("color: #ccc;")
        control_layout.addWidget(sep2)

        # 썸네일 분석 버튼
        self.thumbnail_analysis_btn = QPushButton("🔍 썸네일자동선택")
        self.thumbnail_analysis_btn.setToolTip("최적 썸네일 자동 선택 (멀티스레드 분석)")
        self.thumbnail_analysis_btn.clicked.connect(self._analyze_thumbnails)
        control_layout.addWidget(self.thumbnail_analysis_btn)

        # 지재권 분석 버튼
        self.ip_analysis_btn = QPushButton("🏷️ 지재권분석")
        self.ip_analysis_btn.setToolTip("상품명에서 지재권 의심 단어 분석")
        self.ip_analysis_btn.clicked.connect(self._analyze_ip)
        control_layout.addWidget(self.ip_analysis_btn)

        # 상품명 검수 버튼 (AI 검증)
        self.name_check_btn = QPushButton("📝 상품명검수")
        self.name_check_btn.setToolTip("상품명에서 의심 단어 AI 검증")
        self.name_check_btn.clicked.connect(self._check_product_names_with_ai)
        control_layout.addWidget(self.name_check_btn)

        control_layout.addStretch()

        self.product_count_label = QLabel("상품: 0개")
        self.product_count_label.setStyleSheet("font-weight: bold;")
        control_layout.addWidget(self.product_count_label)

        layout.addLayout(control_layout)

        # 스플리터 (테이블 + 상세)
        splitter = QSplitter(Qt.Orientation.Horizontal)

        # 왼쪽: 상품 테이블
        table_widget = QWidget()
        table_layout = QVBoxLayout(table_widget)
        table_layout.setContentsMargins(0, 0, 0, 0)

        self.product_table = QTableWidget()
        self.product_table.setColumnCount(len(TABLE_COLUMNS))
        self.product_table.setHorizontalHeaderLabels([c['name'] for c in TABLE_COLUMNS])

        # 컬럼 너비 설정
        header = self.product_table.horizontalHeader()
        header.setSectionsMovable(True)  # 컬럼 드래그로 위치 변경 가능
        header.sectionMoved.connect(self._on_column_moved)  # 위치 변경 시 저장

        for idx, col in enumerate(TABLE_COLUMNS):
            if col['width'] == 0:  # stretch
                header.setSectionResizeMode(idx, QHeaderView.ResizeMode.Stretch)
            else:
                header.setSectionResizeMode(idx, QHeaderView.ResizeMode.Interactive)  # 드래그 이동 가능
                self.product_table.setColumnWidth(idx, col['width'])

        self.product_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.product_table.setRowHeight(0, 90)  # 기본 행 높이 (옵션이미지 크기 증가)
        self.product_table.itemSelectionChanged.connect(self._on_product_selected)
        table_layout.addWidget(self.product_table, 1)  # stretch factor 1: 테이블이 공간 채움

        # 옵션 선택 위젯 저장용
        self.option_widgets = {}  # {row: {label: button_widget}}

        # 페이지네이션
        page_layout = QHBoxLayout()
        self.prev_btn = QPushButton("◀ 이전")
        self.prev_btn.clicked.connect(lambda: self._change_page(-1))
        page_layout.addWidget(self.prev_btn)

        self.page_label = QLabel("1 / 1")
        page_layout.addWidget(self.page_label)

        self.next_btn = QPushButton("다음 ▶")
        self.next_btn.clicked.connect(lambda: self._change_page(1))
        page_layout.addWidget(self.next_btn)

        page_layout.addStretch()

        self.page_size_combo = QComboBox()
        self.page_size_combo.addItems(["20", "50", "100"])
        self.page_size_combo.currentTextChanged.connect(self._on_page_size_changed)
        page_layout.addWidget(QLabel("페이지 크기:"))
        page_layout.addWidget(self.page_size_combo)

        table_layout.addLayout(page_layout)
        splitter.addWidget(table_widget)

        # 오른쪽: 접을 수 있는 상세 정보 사이드바 (최소 폭)
        self.detail_widget = QWidget()
        self.detail_widget.setMinimumWidth(300)  # 최소 너비 축소
        self.detail_widget.setMaximumWidth(320)  # 최대 너비 제한
        detail_layout = QVBoxLayout(self.detail_widget)
        detail_layout.setContentsMargins(3, 3, 3, 3)

        # 헤더 (접기 버튼 포함)
        header_layout = QHBoxLayout()
        detail_label = QLabel("📋 상품 상세")
        detail_label.setStyleSheet("font-weight: bold; font-size: 16px;")
        header_layout.addWidget(detail_label)
        header_layout.addStretch()

        self.detail_close_btn = QPushButton("✕")
        self.detail_close_btn.setFixedSize(25, 25)
        self.detail_close_btn.setStyleSheet("font-weight: bold;")
        self.detail_close_btn.clicked.connect(self._toggle_detail_panel)
        header_layout.addWidget(self.detail_close_btn)
        detail_layout.addLayout(header_layout)

        # 썸네일 크게 (메인) - 280x280 (패널 폭 최소화)
        self.detail_thumbnail = ImageLabel(280)
        self.detail_thumbnail.setFixedSize(280, 280)
        detail_layout.addWidget(self.detail_thumbnail)

        # 썸네일 6장 미리보기 - 42x42 (패널 폭 최소화)
        thumb_preview_layout = QHBoxLayout()
        thumb_preview_layout.setSpacing(2)
        self.detail_thumb_previews = []
        self._selected_thumb_idx = 0  # 현재 선택된 썸네일 인덱스
        for i in range(6):
            thumb_label = ImageLabel(42)
            thumb_label.setFixedSize(42, 42)
            thumb_label.setStyleSheet("border: 1px solid #ccc; background: #f0f0f0;")
            thumb_label.setCursor(Qt.CursorShape.PointingHandCursor)
            # 클릭 시 메인 썸네일에 표시
            thumb_label.mousePressEvent = lambda e, idx=i: self._on_thumb_preview_clicked(idx)
            thumb_preview_layout.addWidget(thumb_label)
            self.detail_thumb_previews.append(thumb_label)
        thumb_preview_layout.addStretch()
        detail_layout.addLayout(thumb_preview_layout)

        # 썸네일 메인 설정 버튼
        thumb_btn_layout = QHBoxLayout()
        self.set_main_thumb_btn = QPushButton("📌 메인으로 설정")
        self.set_main_thumb_btn.setToolTip("선택한 썸네일을 메인(1번)으로 설정")
        self.set_main_thumb_btn.setStyleSheet("padding: 3px 8px; font-size: 11px;")
        self.set_main_thumb_btn.clicked.connect(self._set_selected_thumb_as_main)
        thumb_btn_layout.addWidget(self.set_main_thumb_btn)
        thumb_btn_layout.addStretch()
        detail_layout.addLayout(thumb_btn_layout)

        # 상품 정보
        self.detail_info = QTextEdit()
        self.detail_info.setReadOnly(True)
        detail_layout.addWidget(self.detail_info, 1)

        splitter.addWidget(self.detail_widget)

        # 기본: 상세패널 숨김
        self.detail_widget.hide()
        self._detail_panel_visible = False

        layout.addWidget(splitter, 1)

        # 상세보기 토글 버튼 (테이블 오른쪽 상단에 추가)
        self.detail_toggle_btn = QPushButton("📋 상세")
        self.detail_toggle_btn.setFixedWidth(70)
        self.detail_toggle_btn.setCheckable(True)
        self.detail_toggle_btn.clicked.connect(self._toggle_detail_panel)
        page_layout.addWidget(self.detail_toggle_btn)

        # 페이지 상태
        self.current_page = 0
        self.page_size = 20

        # 저장된 컬럼 설정 불러오기
        self._load_column_settings()

    def _build_settings_tab(self):
        """설정 탭 UI"""
        layout = QVBoxLayout(self.settings_tab)

        # 미끼 키워드
        keyword_group = QGroupBox("🚫 미끼 키워드")
        keyword_layout = QVBoxLayout(keyword_group)

        keyword_layout.addWidget(QLabel("제외할 키워드 (쉼표 구분):"))
        self.keyword_text = QTextEdit()
        self.keyword_text.setMinimumHeight(150)
        self.keyword_text.setMaximumHeight(300)
        if BULSAJA_API_AVAILABLE:
            keywords = load_bait_keywords()
            self.keyword_text.setText(','.join(keywords))
        keyword_layout.addWidget(self.keyword_text)

        keyword_btn_layout = QHBoxLayout()
        save_keyword_btn = QPushButton("💾 저장")
        save_keyword_btn.clicked.connect(self._save_keywords)
        keyword_btn_layout.addWidget(save_keyword_btn)

        reset_keyword_btn = QPushButton("기본값")
        reset_keyword_btn.clicked.connect(self._reset_keywords)
        keyword_btn_layout.addWidget(reset_keyword_btn)
        keyword_btn_layout.addStretch()
        keyword_layout.addLayout(keyword_btn_layout)

        layout.addWidget(keyword_group)

        # 컬럼 설정
        column_group = QGroupBox("📊 표시 컬럼 설정")
        column_layout = QVBoxLayout(column_group)

        # 컬럼별 체크박스
        self._settings_column_cbs = {}
        config = load_config()
        visible_columns = config.get('visible_columns', [c['key'] for c in TABLE_COLUMNS if c['default']])

        for col_info in TABLE_COLUMNS:
            key = col_info['key']
            name = col_info['name']
            cb = QCheckBox(name)
            cb.setChecked(key in visible_columns)
            cb.stateChanged.connect(self._on_column_checkbox_changed)
            self._settings_column_cbs[key] = cb
            column_layout.addWidget(cb)

        layout.addWidget(column_group)

        # 카테고리별 검수 설정
        category_group = QGroupBox("🏷️ 카테고리별 검수 설정")
        category_layout = QVBoxLayout(category_group)
        category_layout.addWidget(QLabel("카테고리별로 검수 수준을 설정합니다."))

        category_btn = QPushButton("⚙️ 카테고리 검수 설정 열기")
        category_btn.clicked.connect(self._show_category_risk_settings)
        category_layout.addWidget(category_btn)

        layout.addWidget(category_group)

        # AI 설정 (Gemini API 키)
        ai_group = QGroupBox("🤖 AI 설정 (Gemini)")
        ai_layout = QVBoxLayout(ai_group)

        # API 키 입력
        api_key_row = QHBoxLayout()
        api_key_row.addWidget(QLabel("API 키:"))
        self.gemini_api_key_input = QLineEdit()
        self.gemini_api_key_input.setEchoMode(QLineEdit.EchoMode.Password)
        self.gemini_api_key_input.setPlaceholderText("Gemini API 키 입력...")
        # 현재 설정 로드
        if BULSAJA_API_AVAILABLE:
            ai_config = load_ai_config()
            current_key = ai_config.get('gemini', {}).get('api_key', '')
            self.gemini_api_key_input.setText(current_key)
        api_key_row.addWidget(self.gemini_api_key_input)

        # 표시/숨기기 버튼
        self.show_api_key_btn = QPushButton("👁")
        self.show_api_key_btn.setFixedWidth(30)
        self.show_api_key_btn.setCheckable(True)
        self.show_api_key_btn.toggled.connect(self._toggle_api_key_visibility)
        api_key_row.addWidget(self.show_api_key_btn)

        ai_layout.addLayout(api_key_row)

        # 모델 선택
        model_row = QHBoxLayout()
        model_row.addWidget(QLabel("모델:"))
        self.gemini_model_combo = QComboBox()
        self.gemini_model_combo.addItems([
            "gemini-2.0-flash",
            "gemini-1.5-flash",
            "gemini-1.5-pro",
        ])
        if BULSAJA_API_AVAILABLE:
            ai_config = load_ai_config()
            current_model = ai_config.get('gemini', {}).get('model', 'gemini-2.0-flash')
            idx = self.gemini_model_combo.findText(current_model)
            if idx >= 0:
                self.gemini_model_combo.setCurrentIndex(idx)
        model_row.addWidget(self.gemini_model_combo)
        model_row.addStretch()
        ai_layout.addLayout(model_row)

        # 저장 버튼
        ai_btn_row = QHBoxLayout()
        save_ai_btn = QPushButton("💾 AI 설정 저장")
        save_ai_btn.clicked.connect(self._save_ai_settings)
        ai_btn_row.addWidget(save_ai_btn)

        test_ai_btn = QPushButton("🧪 연결 테스트")
        test_ai_btn.clicked.connect(self._test_ai_connection)
        ai_btn_row.addWidget(test_ai_btn)

        ai_btn_row.addStretch()
        ai_layout.addLayout(ai_btn_row)

        # 안내 문구
        ai_layout.addWidget(QLabel("💡 Gemini API 키는 Google AI Studio에서 무료로 발급받을 수 있습니다."))

        layout.addWidget(ai_group)

        layout.addStretch()

    def _toggle_api_key_visibility(self, checked: bool):
        """API 키 표시/숨기기"""
        if checked:
            self.gemini_api_key_input.setEchoMode(QLineEdit.EchoMode.Normal)
            self.show_api_key_btn.setText("🙈")
        else:
            self.gemini_api_key_input.setEchoMode(QLineEdit.EchoMode.Password)
            self.show_api_key_btn.setText("👁")

    def _save_ai_settings(self):
        """AI 설정 저장"""
        if not BULSAJA_API_AVAILABLE:
            QMessageBox.warning(self, "오류", "bulsaja_common 모듈을 사용할 수 없습니다.")
            return

        api_key = self.gemini_api_key_input.text().strip()
        model = self.gemini_model_combo.currentText()

        ai_config = load_ai_config()
        ai_config['provider'] = 'gemini'
        ai_config['gemini'] = {
            'api_key': api_key,
            'model': model
        }

        if save_ai_config(ai_config):
            QMessageBox.information(self, "완료", "AI 설정이 저장되었습니다.")
            self._log("✅ AI 설정 저장 완료")
        else:
            QMessageBox.warning(self, "오류", "AI 설정 저장에 실패했습니다.")

    def _test_ai_connection(self):
        """AI 연결 테스트"""
        if not BULSAJA_API_AVAILABLE:
            QMessageBox.warning(self, "오류", "bulsaja_common 모듈을 사용할 수 없습니다.")
            return

        api_key = self.gemini_api_key_input.text().strip()
        if not api_key:
            QMessageBox.warning(self, "오류", "API 키를 입력해주세요.")
            return

        self._log("🧪 AI 연결 테스트 중...")

        try:
            from bulsaja_common import call_ai_api

            # 임시 설정으로 테스트
            test_config = {
                'provider': 'gemini',
                'gemini': {
                    'api_key': api_key,
                    'model': self.gemini_model_combo.currentText()
                }
            }

            success, response, error = call_ai_api("안녕하세요, 테스트입니다. '연결 성공'이라고만 답해주세요.", test_config, timeout=10)

            if success:
                QMessageBox.information(self, "성공", f"AI 연결 성공!\n\n응답: {response[:100]}...")
                self._log("✅ AI 연결 테스트 성공")
            else:
                QMessageBox.warning(self, "실패", f"AI 연결 실패\n\n{error}")
                self._log(f"❌ AI 연결 테스트 실패: {error}")

        except Exception as e:
            QMessageBox.warning(self, "오류", f"테스트 중 오류 발생:\n{e}")
            self._log(f"❌ AI 연결 테스트 오류: {e}")

    # ============================================================
    # API 연결
    # ============================================================
    def _open_chrome(self):
        """크롬 디버그 모드로 열기"""
        port = self.port_spin.value()
        chrome_paths = [
            r"C:\Program Files\Google\Chrome\Application\chrome.exe",
            r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
            os.path.expandvars(r"%LOCALAPPDATA%\Google\Chrome\Application\chrome.exe"),
        ]

        chrome_path = None
        for p in chrome_paths:
            if os.path.exists(p):
                chrome_path = p
                break

        if not chrome_path:
            QMessageBox.warning(self, "오류", "Chrome을 찾을 수 없습니다.")
            return

        user_data = os.path.expandvars(r"%TEMP%\chrome_debug_simulator")
        cmd = [
            chrome_path,
            f"--remote-debugging-port={port}",
            "--remote-allow-origins=*",
            f"--user-data-dir={user_data}",
            "https://www.bulsaja.com"
        ]

        try:
            subprocess.Popen(cmd)
            self._log(f"🌐 크롬 열림 (포트: {port})")
        except Exception as e:
            QMessageBox.warning(self, "오류", f"크롬 실행 실패: {e}")

    def _extract_token(self):
        """토큰 추출"""
        if not BULSAJA_API_AVAILABLE:
            QMessageBox.warning(self, "오류", "bulsaja_common 모듈이 없습니다.")
            return

        port = self.port_spin.value()
        self._log(f"🔑 토큰 추출 중... (포트: {port})")

        success, access, refresh, msg = extract_tokens_from_browser(port)
        if success and access:
            self.access_input.setText(access)
            self.refresh_input.setText(refresh)
            self._log("✅ 토큰 추출 성공")
            self._connect_api()
        else:
            self._log(f"❌ 토큰 추출 실패: {msg}")
            QMessageBox.warning(self, "실패", msg)

    def _connect_api(self):
        """API 연결"""
        if not BULSAJA_API_AVAILABLE:
            QMessageBox.warning(self, "오류", "bulsaja_common 모듈이 없습니다.")
            return

        access = self.access_input.text().strip()
        refresh = self.refresh_input.text().strip()

        if not access or not refresh:
            QMessageBox.warning(self, "경고", "토큰을 입력해주세요.")
            return

        try:
            self.api_client = BulsajaAPIClient(access, refresh)
            groups = self.api_client.get_market_groups()

            self.api_status.setText(f"🟢 연결됨 ({len(groups)}개 그룹)")
            self.api_status.setStyleSheet("font-weight: bold; color: #4CAF50;")
            self._log(f"✅ API 연결 성공! {len(groups)}개 그룹")

            # 그룹 목록 설정
            self.groups_text.setText(', '.join(groups))

        except Exception as e:
            self.api_status.setText("🔴 연결 실패")
            self.api_status.setStyleSheet("font-weight: bold; color: #F44336;")
            self._log(f"❌ 연결 실패: {e}")
            QMessageBox.critical(self, "오류", f"연결 실패:\n{e}")

    def _load_groups(self):
        """그룹 목록 로드"""
        if not self.api_client:
            QMessageBox.warning(self, "경고", "먼저 API에 연결하세요.")
            return

        try:
            groups = self.api_client.get_market_groups()
            self.groups_text.setText(', '.join(groups))
            self._log(f"📥 {len(groups)}개 그룹 로드됨")
        except Exception as e:
            self._log(f"❌ 그룹 로드 실패: {e}")

    # ============================================================
    # 수집
    # ============================================================
    def _start_collect(self):
        """수집 시작"""
        if not self.api_client:
            QMessageBox.warning(self, "경고", "먼저 API에 연결하세요.")
            return

        # 그룹 파싱
        groups_text = self.groups_text.toPlainText().strip()
        if not groups_text:
            QMessageBox.warning(self, "경고", "그룹을 입력하세요.")
            return

        all_groups = [g.strip() for g in groups_text.split(',') if g.strip()]
        work_range = self.work_groups_input.text().strip()

        # 범위 파싱
        selected_groups = []
        try:
            if '-' in work_range:
                start, end = map(int, work_range.split('-'))
                selected_groups = all_groups[start-1:end]
            elif ',' in work_range:
                indices = [int(x.strip()) for x in work_range.split(',')]
                selected_groups = [all_groups[i-1] for i in indices if 0 < i <= len(all_groups)]
            else:
                idx = int(work_range)
                if 0 < idx <= len(all_groups):
                    selected_groups = [all_groups[idx-1]]
        except:
            selected_groups = all_groups

        if not selected_groups:
            QMessageBox.warning(self, "경고", "유효한 그룹이 없습니다.")
            return

        self._log(f"🚀 수집 시작: {len(selected_groups)}개 그룹")

        # 미끼 키워드 로드
        bait_keywords = []
        excluded_words = set()
        if BULSAJA_API_AVAILABLE:
            bait_keywords = load_bait_keywords() if self.filter_bait_check.isChecked() else []
            excluded_words = load_excluded_words()

        # 검수 수준
        check_level_text = self.check_level_combo.currentText()
        check_level = CHECK_LEVELS.get(check_level_text, 'normal')

        # 설정
        settings = {
            'max_products': self.max_products_spin.value(),
            'status_filters': UPLOAD_CONDITIONS.get(self.condition_combo.currentText()),
            'bait_keywords': bait_keywords,
            'excluded_words': excluded_words,
            'check_level': check_level,
            'option_count': self.option_count_spin.value(),
            'fetch_category': self.fetch_category_check.isChecked(),  # 카테고리 수집
        }

        max_p = self.max_products_spin.value()
        max_text = "전체" if max_p == 0 else f"{max_p}개"
        self._log(f"   검수수준: {check_level_text}, 미끼필터: {'O' if bait_keywords else 'X'}, 그룹당: {max_text}")

        # 워커 시작
        self.collect_worker = CollectWorker(self.api_client, selected_groups, settings)
        self.collect_worker.progress.connect(self._on_collect_progress)
        self.collect_worker.log.connect(self._log)
        self.collect_worker.finished_signal.connect(self._on_collect_finished)

        self.collect_btn.setEnabled(False)
        self.stop_btn.setEnabled(True)
        self.progress_bar.setValue(0)

        self.collect_worker.start()

    def _stop_collect(self):
        """수집 중지"""
        if hasattr(self, 'collect_worker'):
            self.collect_worker.stop()
            self._log("⏹️ 중지 요청됨...")

    def _on_collect_progress(self, current: int, total: int, group_name: str):
        """수집 진행 상황"""
        percent = int(current / total * 100) if total > 0 else 0
        self.progress_bar.setValue(percent)
        self.progress_label.setText(f"진행: {current}/{total} - {group_name}")

    def _on_collect_finished(self, success: bool, message: str, products: list):
        """수집 완료"""
        self.collect_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)
        self.progress_bar.setValue(100)
        self.progress_label.setText(message)

        if success and products:
            self.products = products
            self._log(f"✅ {message}")
            self._update_product_table()
            self.tabs.setCurrentIndex(1)  # 검수 탭으로 이동

            # 자동 저장
            self._auto_save_excel()

            # 자동 지재권 분석 (백그라운드)
            self._auto_ip_analysis()
        else:
            self._log(f"⚠️ {message}")

    # ============================================================
    # 검수 탭
    # ============================================================
    def _update_product_table(self):
        """상품 테이블 업데이트 (TABLE_COLUMNS 기반)"""
        self.product_count_label.setText(f"상품: {len(self.products)}개")

        # 페이지 계산
        total_pages = max(1, (len(self.products) + self.page_size - 1) // self.page_size)
        self.current_page = min(self.current_page, total_pages - 1)

        start_idx = self.current_page * self.page_size
        end_idx = min(start_idx + self.page_size, len(self.products))
        page_products = self.products[start_idx:end_idx]

        self.page_label.setText(f"{self.current_page + 1} / {total_pages}")

        # 위험상품만 표시 옵션
        show_unsafe_only = self.show_unsafe_only_check.isChecked() if hasattr(self, 'show_unsafe_only_check') else False
        if show_unsafe_only:
            page_products = [p for p in page_products if not p.get('is_safe', True)]

        # 테이블 채우기
        self.product_table.setRowCount(len(page_products))
        self.option_widgets = {}  # 옵션 위젯 초기화

        # 행 높이 계산 (옵션 표시 개수에 따라)
        display_count = int(self.option_display_combo.currentText().replace('개', '')) if hasattr(self, 'option_display_combo') else 6
        row_height = max(90, 13 * display_count + 20)  # 옵션당 13px + 여백 (옵션이미지 크기 증가)

        for row, product in enumerate(page_products):
            product_id = product.get('ID', '')
            self.product_table.setRowHeight(row, row_height)

            # 컬럼별 렌더링
            for col_idx, col_info in enumerate(TABLE_COLUMNS):
                key = col_info['key']

                if key == 'thumbnail':
                    # 썸네일 (중앙 정렬 컨테이너) - 크기 증가
                    thumb_container = QWidget()
                    thumb_layout = QHBoxLayout(thumb_container)
                    thumb_layout.setContentsMargins(0, 0, 0, 0)
                    thumb_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
                    thumb_label = ImageLabel(80)
                    thumb_layout.addWidget(thumb_label)
                    self.image_labels[f"{product_id}_thumb"] = thumb_label
                    self.product_table.setCellWidget(row, col_idx, thumb_container)
                    # 썸네일 URL 저장
                    thumb_url = product.get('thumbnail_url', '') or product.get('uploadCommonThumbnail', '')
                    if not thumb_url:
                        thumbnails = product.get('uploadThumbnails', []) or product.get('all_thumbnails', [])
                        if thumbnails:
                            thumb_url = thumbnails[0]
                    product['_thumb_url'] = thumb_url
                    # 클릭 시 상세패널에 이미지 표시
                    thumb_label.set_image_url(thumb_url)
                    thumb_label.clicked.connect(self._on_image_clicked)

                elif key == 'option_image':
                    # 대표 옵션 이미지 (중앙 정렬 컨테이너) - 크기 증가
                    opt_container = QWidget()
                    opt_layout = QHBoxLayout(opt_container)
                    opt_layout.setContentsMargins(0, 0, 0, 0)
                    opt_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
                    opt_img_label = ImageLabel(80)
                    opt_layout.addWidget(opt_img_label)
                    self.image_labels[f"{product_id}_opt"] = opt_img_label
                    self.product_table.setCellWidget(row, col_idx, opt_container)
                    # 옵션 이미지 URL 저장
                    opt_img_url = product.get('main_option_image', '')
                    if not opt_img_url:
                        option_images = product.get('option_images', {})
                        if option_images and isinstance(option_images, dict):
                            opt_img_url = option_images.get('A', '') or (list(option_images.values())[0] if option_images.values() else '')
                    # 폴백: uploadSkus 또는 all_thumbnails에서 가져오기
                    if not opt_img_url:
                        upload_skus = product.get('uploadSkus', [])
                        if upload_skus and isinstance(upload_skus, list) and len(upload_skus) > 0:
                            opt_img_url = upload_skus[0].get('urlRef', '')
                    if not opt_img_url:
                        all_thumbs = product.get('all_thumbnails', []) or product.get('uploadThumbnails', [])
                        if all_thumbs and len(all_thumbs) > 1:
                            opt_img_url = all_thumbs[1]  # 두 번째 썸네일
                        elif all_thumbs:
                            opt_img_url = all_thumbs[0]
                    product['_opt_img_url'] = opt_img_url
                    # 클릭 시 상세패널에 이미지 표시
                    opt_img_label.set_image_url(opt_img_url)
                    opt_img_label.clicked.connect(self._on_image_clicked)

                elif key == 'option_select':
                    # 옵션 선택 버튼들 (A~F)
                    option_widget = self._create_option_buttons(row, product)
                    self.product_table.setCellWidget(row, col_idx, option_widget)

                elif key == 'option_names_list':
                    # 옵션명 목록 (작은 폰트, 줄간격 최소화)
                    option_names = product.get('option_names', {})
                    display_count = int(self.option_display_combo.currentText().replace('개', '')) if hasattr(self, 'option_display_combo') else 6
                    if option_names and isinstance(option_names, dict):
                        names_lines = [f"{k}:{v}" for k, v in list(option_names.items())[:display_count] if v]
                        names_text = '\n'.join(names_lines)
                    else:
                        final_list = product.get('final_option_list', [])
                        if final_list:
                            labels = 'ABCDEF'
                            names_lines = [f"{labels[i]}:{opt.split('(')[0]}" for i, opt in enumerate(final_list[:display_count])]
                            names_text = '\n'.join(names_lines)
                        else:
                            names_text = ''
                    names_label = QLabel(names_text)
                    names_label.setStyleSheet("font-size: 11px; padding: 2px; line-height: 1.2;")
                    names_label.setToolTip(names_text.replace('\n', '\n'))
                    self.product_table.setCellWidget(row, col_idx, names_label)

                elif key == 'option_names_cn':
                    # 중국어 옵션명 목록
                    option_names_cn = product.get('option_names_cn', {})
                    display_count = int(self.option_display_combo.currentText().replace('개', '')) if hasattr(self, 'option_display_combo') else 6
                    if option_names_cn and isinstance(option_names_cn, dict):
                        cn_lines = [f"{k}:{v}" for k, v in list(option_names_cn.items())[:display_count] if v]
                        cn_text = '\n'.join(cn_lines)
                    else:
                        cn_text = ''
                    cn_label = QLabel(cn_text)
                    cn_label.setStyleSheet("font-size: 11px; padding: 2px;")
                    cn_label.setToolTip(cn_text)
                    self.product_table.setCellWidget(row, col_idx, cn_label)

                elif key == 'name':
                    name = product.get('uploadCommonProductName', '')
                    # 상품명 검수 결과 확인
                    name_check = product.get('name_check_result', {})
                    suspicious_words = name_check.get('suspicious_words', [])

                    name_label = QLabel()
                    name_label.setWordWrap(True)

                    if suspicious_words:
                        # 의심 단어가 있으면 HTML로 강조 표시
                        highlighted = name_check.get('highlighted_name', name)
                        name_label.setText(highlighted)
                        name_label.setTextFormat(Qt.TextFormat.RichText)
                        # 배경색 약간 노란색
                        name_label.setStyleSheet("font-size: 11px; padding: 2px; background-color: #FFFDE7;")
                        # 툴팁에 의심 단어 목록 표시
                        suspicious_list = ', '.join([w['word'] for w in suspicious_words])
                        name_label.setToolTip(f"⚠️ 의심 단어: {suspicious_list}\n\n{name}")
                    else:
                        name_label.setText(name)
                        name_label.setStyleSheet("font-size: 11px; padding: 2px;")
                        name_label.setToolTip(name)

                    self.product_table.setCellWidget(row, col_idx, name_label)

                elif key == 'danger':
                    # 위험 컬럼 (체크=위험, 미체크=안전)
                    is_safe = product.get('is_safe', True)
                    is_danger = not is_safe
                    danger_widget = QWidget()
                    danger_layout = QHBoxLayout(danger_widget)
                    danger_layout.setContentsMargins(0, 0, 0, 0)
                    danger_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
                    danger_check = QCheckBox()
                    danger_check.setChecked(is_danger)
                    danger_check.stateChanged.connect(lambda state, p=product: self._on_danger_toggle(p, state))
                    danger_layout.addWidget(danger_check)
                    if is_danger:
                        danger_widget.setStyleSheet("background-color: #FFCDD2;")
                    self.product_table.setCellWidget(row, col_idx, danger_widget)

                elif key == 'unsafe_reason':
                    reason = product.get('unsafe_reason', '')
                    reason_item = QTableWidgetItem(reason[:25] if reason else '')
                    if reason:
                        reason_item.setBackground(QColor("#FFCDD2"))
                    self.product_table.setItem(row, col_idx, reason_item)

                elif key == 'options':
                    total_opts = product.get('total_options', 0)
                    valid_opts = product.get('valid_options', total_opts)
                    opt_text = f"{valid_opts}/{total_opts}" if valid_opts != total_opts else str(total_opts)
                    opt_item = QTableWidgetItem(opt_text)
                    opt_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    self.product_table.setItem(row, col_idx, opt_item)

                elif key == 'bait':
                    bait_count = product.get('bait_options', 0)
                    bait_item = QTableWidgetItem(str(bait_count) if bait_count > 0 else '')
                    bait_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    if bait_count > 0:
                        bait_item.setBackground(QColor("#FFF9C4"))
                    self.product_table.setItem(row, col_idx, bait_item)

                elif key == 'price':
                    # 가격 (여러 필드에서 시도)
                    price = product.get('uploadCommonSalePrice', 0) or product.get('salePrice', 0) or product.get('price', 0)
                    if not price:
                        # CNY 가격으로 대체 표시
                        min_cny = product.get('min_price_cny', 0)
                        if min_cny:
                            price_text = f"¥{min_cny:.0f}"
                        else:
                            price_text = ''
                    else:
                        price_text = f"{int(price):,}"
                    price_item = QTableWidgetItem(price_text)
                    price_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                    self.product_table.setItem(row, col_idx, price_item)

                elif key == 'price_range':
                    min_p = product.get('min_price_cny', 0)
                    max_p = product.get('max_price_cny', 0)
                    range_text = f"{min_p:.0f}~{max_p:.0f}" if min_p and max_p else ''
                    self.product_table.setItem(row, col_idx, QTableWidgetItem(range_text))

                elif key == 'category':
                    cat = ''
                    upload_cat = product.get('uploadCategory')
                    if isinstance(upload_cat, dict):
                        for k in ['ss_category', 'esm_category', 'est_category']:
                            cat_obj = upload_cat.get(k)
                            if isinstance(cat_obj, dict) and cat_obj.get('name'):
                                cat = cat_obj['name'].split('>')[-1]  # 마지막 카테고리만
                                break
                    self.product_table.setItem(row, col_idx, QTableWidgetItem(cat[:15]))

                elif key == 'main_option':
                    # 선택된 옵션 레이블 (A, B, C, D...) + 옵션명
                    selected_label = product.get('_selected_option', 'A')
                    main_opt_name = product.get('main_option_name', '')[:12]
                    display_text = f"{selected_label}: {main_opt_name}" if main_opt_name else selected_label
                    item = QTableWidgetItem(display_text)
                    item.setToolTip(product.get('main_option_name', ''))
                    self.product_table.setItem(row, col_idx, item)

                elif key == 'group':
                    group = product.get('group_name', '')
                    self.product_table.setItem(row, col_idx, QTableWidgetItem(group))

                elif key == 'id':
                    self.product_table.setItem(row, col_idx, QTableWidgetItem(product_id))

    def _create_option_buttons(self, row: int, product: dict) -> QWidget:
        """옵션 선택 버튼 위젯 생성 (A~F)"""
        widget = QWidget()
        layout = QHBoxLayout(widget)
        layout.setContentsMargins(2, 2, 2, 2)
        layout.setSpacing(2)

        option_images = product.get('option_images', {})
        option_prices = product.get('option_prices', {})
        product_id = product.get('ID', '')

        # option_images가 비어있으면 uploadSkus에서 가져오기 (1차 폴백)
        if not option_images:
            upload_skus = product.get('uploadSkus', [])
            if upload_skus and isinstance(upload_skus, list):
                labels_fb = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
                option_images = {}
                for i, sku in enumerate(upload_skus[:6]):
                    url_ref = sku.get('urlRef', '')
                    if url_ref:
                        option_images[labels_fb[i]] = url_ref
                if option_images:
                    product['option_images'] = option_images

        # 여전히 비어있으면 all_thumbnails에서 생성 (2차 폴백)
        if not option_images:
            all_thumbs = product.get('all_thumbnails', []) or product.get('uploadThumbnails', []) or []
            if all_thumbs:
                labels_fb = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
                option_images = {labels_fb[i]: thumb for i, thumb in enumerate(all_thumbs[:6])}
                product['option_images'] = option_images

        # 선택된 옵션 (기본 A)
        selected = product.get('_selected_option', 'A')

        # 표시할 옵션 개수 (설정에서 가져오기)
        display_count = int(self.option_display_combo.currentText().replace('개', '')) if hasattr(self, 'option_display_combo') else 6
        labels = 'ABCDEF'[:display_count]
        for label in labels:
            btn_frame = QFrame()
            btn_frame.setFixedSize(65, 85)  # 크기 증가

            is_selected = (label == selected)
            if is_selected:
                btn_frame.setStyleSheet("background-color: #2196F3; border: 2px solid #1976D2; border-radius: 3px;")
            else:
                btn_frame.setStyleSheet("background-color: #E0E0E0; border: 1px solid #BDBDBD; border-radius: 3px;")

            btn_layout = QVBoxLayout(btn_frame)
            btn_layout.setContentsMargins(2, 2, 2, 2)
            btn_layout.setSpacing(1)

            # 옵션 이미지 (55x55 - 크기 증가)
            img_label = ImageLabel(55)
            img_label.setStyleSheet("border: none;")
            btn_layout.addWidget(img_label, alignment=Qt.AlignmentFlag.AlignCenter)

            # 저장 (이미지 로드용) - 폴백 로직 추가
            img_url = option_images.get(label, '')
            # 폴백 1: uploadSkus에서 직접 찾기
            if not img_url:
                upload_skus = product.get('uploadSkus', [])
                label_idx = ord(label) - ord('A')
                if upload_skus and label_idx < len(upload_skus):
                    img_url = upload_skus[label_idx].get('urlRef', '') or upload_skus[label_idx].get('optionImage', '')
            # 폴백 2: 대표옵션이미지 사용 (A일 경우)
            if not img_url and label == 'A':
                img_url = product.get('main_option_image', '') or product.get('대표옵션이미지', '')
            # 폴백 3: 썸네일 사용
            if not img_url:
                all_thumbs = product.get('all_thumbnails', []) or product.get('uploadThumbnails', [])
                label_idx = ord(label) - ord('A')
                if all_thumbs and label_idx < len(all_thumbs):
                    img_url = all_thumbs[label_idx]
            if img_url:
                self.image_labels[f"{product_id}_opt_{label}"] = img_label
                product[f'_opt_{label}_url'] = img_url
                option_images[label] = img_url  # 업데이트
                # 클릭 시 상세패널에 이미지 표시
                img_label.set_image_url(img_url)
                img_label.clicked.connect(self._on_image_clicked)

            # 라벨 (A, B, C...)
            text_label = QLabel(label)
            text_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            text_color = "white" if is_selected else "black"
            text_label.setStyleSheet(f"color: {text_color}; font-weight: bold; font-size: 11px; border: none;")
            btn_layout.addWidget(text_label)

            # 클릭 이벤트 추가 (대표옵션 변경)
            btn_frame.setCursor(Qt.CursorShape.PointingHandCursor)
            btn_frame.mousePressEvent = lambda event, r=row, l=label, p=product: self._on_option_select(r, l, p)

            layout.addWidget(btn_frame)

            # 옵션 위젯 저장
            if row not in self.option_widgets:
                self.option_widgets[row] = {}
            self.option_widgets[row][label] = {
                'frame': btn_frame,
                'img_label': img_label,
                'text_label': text_label
            }

        layout.addStretch()
        return widget

    def _on_option_select(self, row: int, label: str, product: dict):
        """옵션 버튼 클릭 시 대표옵션 변경"""
        # 이전 선택 해제
        old_selected = product.get('_selected_option', 'A')

        # 새 선택 저장
        product['_selected_option'] = label

        # 대표옵션 정보 업데이트
        option_names = product.get('option_names', {})
        option_names_cn = product.get('option_names_cn', {})
        option_images = product.get('option_images', {})
        option_prices = product.get('option_prices', {})

        product['main_option_name'] = option_names.get(label, '')
        product['main_option_name_cn'] = option_names_cn.get(label, '')
        product['main_option_image'] = option_images.get(label, '')
        product['main_option_price'] = option_prices.get(label, 0)

        # UI 스타일 업데이트
        if row in self.option_widgets:
            for opt_label, widgets in self.option_widgets[row].items():
                frame = widgets.get('frame')
                text_label = widgets.get('text_label')
                if opt_label == label:
                    # 선택된 버튼 강조
                    frame.setStyleSheet("background-color: #2196F3; border: 2px solid #1976D2; border-radius: 3px;")
                    text_label.setStyleSheet("color: white; font-weight: bold; font-size: 11px; border: none;")
                else:
                    # 비선택 버튼
                    frame.setStyleSheet("background-color: #E0E0E0; border: 1px solid #BDBDBD; border-radius: 3px;")
                    text_label.setStyleSheet("color: black; font-weight: bold; font-size: 11px; border: none;")

        # 대표옵션 이미지 업데이트 (옵션이미지 칼럼)
        product_id = product.get('ID', '')
        opt_key = f"{product_id}_opt"
        new_img_url = option_images.get(label, '')
        if new_img_url and opt_key in self.image_labels:
            product['_opt_img_url'] = new_img_url
            self._load_image_async(opt_key, new_img_url)

        # 대표옵션 컬럼 텍스트 업데이트
        for col_idx, col_info in enumerate(TABLE_COLUMNS):
            if col_info.get('key') == 'main_option':
                main_opt_name = product.get('main_option_name', '')[:12]
                display_text = f"{label}: {main_opt_name}" if main_opt_name else label
                item = QTableWidgetItem(display_text)
                item.setToolTip(product.get('main_option_name', ''))
                self.product_table.setItem(row, col_idx, item)
                break

        self._log(f"   옵션 변경: {product.get('uploadCommonProductName', '')[:20]}... → {label}")

    def _on_danger_toggle(self, product: dict, state: int):
        """위험여부 체크박스 토글 (체크=위험, 미체크=안전)"""
        is_danger = (state == 2)  # Qt.CheckState.Checked = 2
        product['is_safe'] = not is_danger
        if not is_danger:  # 안전해지면 사유 삭제
            product['unsafe_reason'] = ''
        self._log(f"   위험여부 변경: {product.get('uploadCommonProductName', '')[:20]}... → {'위험' if is_danger else '안전'}")

    def _load_image_async(self, product_id: str, url: str):
        """비동기 이미지 로드"""
        worker = ImageDownloadWorker(product_id, url, QSize(50, 50))
        worker.signals.finished.connect(self._on_image_loaded)
        worker.signals.error.connect(self._on_image_error)
        self.thread_pool.start(worker)

    def _on_image_loaded(self, product_id: str, pixmap: QPixmap):
        """이미지 로드 완료"""
        if product_id in self.image_labels:
            self.image_labels[product_id].set_image(pixmap)

    def _on_image_error(self, product_id: str, error: str):
        """이미지 로드 에러"""
        if product_id in self.image_labels:
            self.image_labels[product_id].show_error()

    def _change_page(self, delta: int):
        """페이지 변경"""
        total_pages = max(1, (len(self.products) + self.page_size - 1) // self.page_size)
        new_page = self.current_page + delta

        if 0 <= new_page < total_pages:
            self.current_page = new_page
            self._update_product_table()

    def _on_page_size_changed(self, size_text: str):
        """페이지 크기 변경"""
        self.page_size = int(size_text)
        self.current_page = 0
        self._update_product_table()

    def _on_product_selected(self):
        """상품 선택 시"""
        rows = self.product_table.selectedItems()
        if not rows:
            return

        row = self.product_table.currentRow()
        start_idx = self.current_page * self.page_size
        product_idx = start_idx + row

        if 0 <= product_idx < len(self.products):
            product = self.products[product_idx]
            # 상세패널 자동 열기
            if not self._detail_panel_visible:
                self._toggle_detail_panel()
            self._show_product_detail(product)

    def _toggle_detail_panel(self):
        """상세 패널 토글"""
        self._detail_panel_visible = not self._detail_panel_visible

        if self._detail_panel_visible:
            self.detail_widget.show()
            self.detail_toggle_btn.setChecked(True)
            self.detail_toggle_btn.setText("📋 닫기")
        else:
            self.detail_widget.hide()
            self.detail_toggle_btn.setChecked(False)
            self.detail_toggle_btn.setText("📋 상세")

    def _on_option_display_changed(self, text: str):
        """옵션 표시 개수 변경 시 테이블 새로고침"""
        if self.products:
            self._update_product_table()

    def _on_image_clicked(self, image_url: str):
        """이미지 클릭 시 상세 패널에 큰 이미지 표시"""
        if image_url:
            worker = ImageDownloadWorker("clicked", image_url, QSize(280, 280))
            worker.signals.finished.connect(
                lambda pid, pix: self.detail_thumbnail.set_image(pix)
            )
            self.thread_pool.start(worker)
            # URL도 저장해서 상세패널 현재 이미지 URL 추적
            self.detail_thumbnail.set_image_url(image_url)

    def _show_product_detail(self, product: dict):
        """상품 상세 표시"""
        self._detail_current_product = product  # 현재 상품 저장 (옵션 선택용)
        self._selected_thumb_idx = 0  # 선택 인덱스 초기화

        # 전체 썸네일 목록 (최대 6장)
        all_thumbs = product.get('uploadThumbnails', [])
        if not all_thumbs:
            # 폴백: uploadCommonThumbnail
            thumb_url = product.get('uploadCommonThumbnail', '')
            all_thumbs = [thumb_url] if thumb_url else []
        self._detail_thumb_urls = all_thumbs[:6]  # 클릭용 저장

        # 메인 썸네일 (첫 번째)
        if all_thumbs:
            thumb_url = all_thumbs[0]
            worker = ImageDownloadWorker("detail", thumb_url, QSize(280, 280))
            worker.signals.finished.connect(
                lambda pid, pix: self.detail_thumbnail.set_image(pix)
            )
            self.thread_pool.start(worker)
            self.detail_thumbnail.set_image_url(thumb_url)

        # 6장 미리보기 로드
        for i, preview_label in enumerate(self.detail_thumb_previews):
            if i < len(all_thumbs):
                url = all_thumbs[i]
                # 선택된 썸네일 테두리 강조
                if i == 0:
                    preview_label.setStyleSheet("border: 2px solid #007bff; background: #f0f0f0;")
                else:
                    preview_label.setStyleSheet("border: 1px solid #ccc; background: #f0f0f0;")
                # 이미지 로드
                worker = ImageDownloadWorker(f"thumb_preview_{i}", url, QSize(42, 42))
                worker.signals.finished.connect(
                    lambda pid, pix, label=preview_label: label.set_image(pix)
                )
                self.thread_pool.start(worker)
            else:
                # 빈 슬롯
                preview_label.clear()
                preview_label.setStyleSheet("border: 1px solid #ccc; background: #f0f0f0;")

        # 상세 정보 표시
        self._update_detail_info(product)

    def _on_thumb_preview_clicked(self, idx: int):
        """썸네일 미리보기 클릭 시 메인에 표시"""
        if not hasattr(self, '_detail_thumb_urls') or idx >= len(self._detail_thumb_urls):
            return

        self._selected_thumb_idx = idx  # 선택 인덱스 저장
        url = self._detail_thumb_urls[idx]

        # 메인 썸네일에 표시
        worker = ImageDownloadWorker("detail_click", url, QSize(280, 280))
        worker.signals.finished.connect(
            lambda pid, pix: self.detail_thumbnail.set_image(pix)
        )
        self.thread_pool.start(worker)
        self.detail_thumbnail.set_image_url(url)

        # 선택 테두리 업데이트
        for i, preview_label in enumerate(self.detail_thumb_previews):
            if i == idx:
                preview_label.setStyleSheet("border: 2px solid #007bff; background: #f0f0f0;")
            elif i < len(self._detail_thumb_urls):
                preview_label.setStyleSheet("border: 1px solid #ccc; background: #f0f0f0;")

    def _set_selected_thumb_as_main(self):
        """선택한 썸네일을 메인(1번)으로 설정"""
        if not hasattr(self, '_detail_current_product') or not self._detail_current_product:
            QMessageBox.warning(self, "오류", "상품이 선택되지 않았습니다.")
            return

        if not hasattr(self, '_detail_thumb_urls') or not self._detail_thumb_urls:
            QMessageBox.warning(self, "오류", "썸네일이 없습니다.")
            return

        idx = getattr(self, '_selected_thumb_idx', 0)
        if idx == 0:
            QMessageBox.information(self, "알림", "이미 첫 번째 썸네일입니다.")
            return

        if idx >= len(self._detail_thumb_urls):
            return

        # 썸네일 순서 변경
        thumbnails = self._detail_thumb_urls.copy()
        selected_thumb = thumbnails.pop(idx)
        thumbnails.insert(0, selected_thumb)

        # 상품 데이터 업데이트
        product = self._detail_current_product
        product['uploadThumbnails'] = thumbnails
        product['all_thumbnails'] = thumbnails
        product['uploadCommonThumbnail'] = thumbnails[0]
        product['thumbnail_url'] = thumbnails[0]

        # 내부 데이터 업데이트
        self._detail_thumb_urls = thumbnails
        self._selected_thumb_idx = 0

        # 썸네일 미리보기 새로고침
        self._refresh_detail_thumbnails()

        # 테이블 새로고침 + 이미지 즉시 로드
        self._update_product_table()
        self._load_current_page_images()

        # 자동 저장
        self._auto_save_excel()

        self._log(f"✅ 썸네일 변경: {product.get('uploadCommonProductName', '')[:25]}... #{idx+1}→#1")

    def _refresh_detail_thumbnails(self):
        """상세 패널 썸네일 미리보기 새로고침"""
        if not hasattr(self, '_detail_thumb_urls'):
            return

        # 메인 썸네일 먼저 새로고침
        if self._detail_thumb_urls:
            url = self._detail_thumb_urls[0]
            self.detail_thumbnail.set_image_url(url)
            worker = ImageDownloadWorker("detail_main_refresh", url, QSize(280, 280))
            worker.signals.finished.connect(
                lambda pid, pix, lbl=self.detail_thumbnail: lbl.set_image(pix) if pix else None
            )
            self.thread_pool.start(worker)

        # 미리보기 썸네일 새로고침
        for i, preview_label in enumerate(self.detail_thumb_previews):
            if i < len(self._detail_thumb_urls):
                url = self._detail_thumb_urls[i]
                preview_label.show()
                # 선택 테두리 (새 첫번째 = 파란색)
                if i == 0:
                    preview_label.setStyleSheet("border: 2px solid #007bff; background: #f0f0f0;")
                else:
                    preview_label.setStyleSheet("border: 1px solid #ccc; background: #f0f0f0;")
                # 이미지 로드 (클로저 문제 방지를 위해 idx 저장)
                def load_thumbnail(label, image_url, idx):
                    w = ImageDownloadWorker(f"thumb_refresh_{idx}", image_url, QSize(42, 42))
                    w.signals.finished.connect(
                        lambda pid, pix, lbl=label: lbl.set_image(pix) if pix else None
                    )
                    self.thread_pool.start(w)
                load_thumbnail(preview_label, url, i)
            else:
                preview_label.hide()

    def _update_detail_info(self, product: dict):
        """상세 정보 텍스트 업데이트"""
        is_safe = product.get('is_safe', True)
        safe_status = "✅ 안전" if is_safe else "⚠️ 위험"
        unsafe_reason = product.get('unsafe_reason', '')
        main_option = product.get('main_option_name', '')
        main_method = product.get('main_option_method', '')
        bait_count = product.get('bait_options', 0)
        bait_list = product.get('bait_option_list', [])
        final_options = product.get('final_option_list', [])

        # 카테고리명 추출 (dict 구조 처리)
        category_display = ''
        upload_cat = product.get('uploadCategory')
        if isinstance(upload_cat, dict):
            for key in ['ss_category', 'esm_category', 'est_category', 'est_global_category']:
                cat_obj = upload_cat.get(key)
                if isinstance(cat_obj, dict) and cat_obj.get('name'):
                    category_display = cat_obj['name']
                    break
        elif isinstance(upload_cat, str):
            category_display = upload_cat

        # 추가 정보 수집
        tags = product.get('uploadCommonTags', [])
        tags_str = ', '.join(tags[:5]) if tags else '없음'
        min_price = product.get('min_price_cny', 0)
        max_price = product.get('max_price_cny', 0)
        option_names = product.get('option_names', {})
        option_prices = product.get('option_prices', {})
        selected_opt = product.get('_selected_option', 'A')

        # 옵션 상세 리스트
        option_details = []
        for label in 'ABCDEF':
            name = option_names.get(label, '')
            price = option_prices.get(label, 0)
            if name:
                mark = '→' if label == selected_opt else ' '
                option_details.append(f"{mark}{label}: {name[:20]} ({price:.1f}위안)")

        info = f"""상품명: {product.get('uploadCommonProductName', '')}

[검수결과] {safe_status}
{f'위험사유: {unsafe_reason}' if unsafe_reason else ''}

[기본정보]
그룹: {product.get('group_name', '')}
ID: {product.get('ID', '')}
가격: {product.get('uploadCommonSalePrice', 0):,}원
가격범위: {min_price:.1f} ~ {max_price:.1f} CNY
카테고리: {category_display}
태그: {tags_str}

[옵션통계]
전체: {product.get('total_options', 0)}개 / 유효: {product.get('valid_options', 0)}개 / 미끼: {bait_count}개
{f'미끼목록: {", ".join(bait_list[:3])}' if bait_list else ''}

[대표옵션] {selected_opt}
{main_option} ({main_method})

[전체옵션] ({len(option_details)}개)
{chr(10).join(option_details) if option_details else '옵션 없음'}

[최종옵션목록]
{chr(10).join(f"• {opt}" for opt in final_options[:10]) if final_options else '없음'}
"""
        self.detail_info.setText(info)

    # ============================================================
    # 엑셀
    # ============================================================
    def _open_excel(self):
        """엑셀 열기"""
        if not OPENPYXL_AVAILABLE:
            QMessageBox.warning(self, "오류", "openpyxl이 설치되지 않았습니다.")
            return

        filepath, _ = QFileDialog.getOpenFileName(
            self, "엑셀 열기", "", "Excel Files (*.xlsx)"
        )
        if not filepath:
            return

        try:
            wb = load_workbook(filepath)
            ws = wb.active

            # 데이터 읽기
            headers = [cell.value for cell in ws[1]]
            self.products = []

            # 한글 → 영문 필드 매핑
            field_mapping = {
                '불사자ID': 'ID',
                '상품명': 'uploadCommonProductName',
                '그룹': 'group_name',
                '판매가': 'uploadCommonSalePrice',
                '카테고리': 'category_name',
                '안전': 'is_safe',
                '위험사유': 'unsafe_reason',
                '전체옵션': 'total_options',
                '유효옵션': 'valid_options',
                '최종옵션': 'final_options',
                '미끼옵션': 'bait_options',
                '대표옵션': 'main_option_name',
                '대표옵션방식': 'main_option_method',
                '최소가(CNY)': 'min_price_cny',
                '최대가(CNY)': 'max_price_cny',
                '썸네일URL': 'thumbnail_url',
                '대표옵션이미지': 'main_option_image',
                '옵션이미지JSON': 'option_images_json',
                '옵션가격JSON': 'option_prices_json',
                '옵션명JSON': 'option_names_json',
                '중국어옵션명JSON': 'option_names_cn_json',
                '전체썸네일': 'all_thumbnails_str',
                '최종옵션목록': 'final_option_list_str',
                '미끼옵션목록': 'bait_option_list_str',
                'uploadSkusJSON': 'upload_skus_json',
                'uploadThumbnailsJSON': 'upload_thumbs_json',
                'uploadCategoryJSON': 'upload_category_json',
            }

            for row in ws.iter_rows(min_row=2, values_only=True):
                product = dict(zip(headers, row))
                if product.get('ID') or product.get('불사자ID'):
                    # 필드명 정규화 (한글 → 영문)
                    for kr_key, en_key in field_mapping.items():
                        if kr_key in product and product[kr_key]:
                            product[en_key] = product[kr_key]

                    # 안전 필드 변환 (문자열 → bool)
                    if product.get('is_safe') == '안전':
                        product['is_safe'] = True
                    elif product.get('is_safe') == '위험':
                        product['is_safe'] = False

                    # 로우데이터 파싱 (uploadSkusJSON, uploadThumbnailsJSON)
                    skus_json = product.get('upload_skus_json', '') or product.get('uploadSkusJSON', '')
                    if skus_json and isinstance(skus_json, str):
                        try:
                            product['uploadSkus'] = json.loads(skus_json)
                        except:
                            pass

                    thumbs_json = product.get('upload_thumbs_json', '') or product.get('uploadThumbnailsJSON', '')
                    if thumbs_json and isinstance(thumbs_json, str):
                        try:
                            product['uploadThumbnails'] = json.loads(thumbs_json)
                            product['all_thumbnails'] = product['uploadThumbnails']
                        except:
                            pass

                    # uploadCategory 파싱
                    cat_json = product.get('upload_category_json', '') or product.get('uploadCategoryJSON', '')
                    if cat_json and isinstance(cat_json, str):
                        try:
                            product['uploadCategory'] = json.loads(cat_json)
                        except:
                            pass

                    # 옵션 가격/이름/중국어 JSON 파싱
                    prices_json = product.get('option_prices_json', '') or product.get('옵션가격JSON', '')
                    if prices_json and isinstance(prices_json, str):
                        try:
                            product['option_prices'] = json.loads(prices_json)
                        except:
                            pass

                    names_json = product.get('option_names_json', '') or product.get('옵션명JSON', '')
                    if names_json and isinstance(names_json, str):
                        try:
                            product['option_names'] = json.loads(names_json)
                        except:
                            pass

                    names_cn_json = product.get('option_names_cn_json', '') or product.get('중국어옵션명JSON', '')
                    if names_cn_json and isinstance(names_cn_json, str):
                        try:
                            product['option_names_cn'] = json.loads(names_cn_json)
                        except:
                            pass

                    # 전체썸네일 파싱 (|로 구분된 문자열 → 리스트) - 폴백
                    if not product.get('uploadThumbnails'):
                        all_thumbs_str = product.get('all_thumbnails_str', '')
                        if all_thumbs_str and isinstance(all_thumbs_str, str):
                            product['all_thumbnails'] = all_thumbs_str.split('|')
                            product['uploadThumbnails'] = product['all_thumbnails']

                    # 옵션이미지 JSON 파싱
                    opt_json = product.get('option_images_json', '') or product.get('옵션이미지JSON', '')
                    if opt_json and isinstance(opt_json, str):
                        try:
                            data = json.loads(opt_json)
                            # option_images와 option_names 분리
                            if isinstance(data, dict):
                                if 'images' in data:
                                    # 새 형식: {"images": {...}, "names": {...}}
                                    product['option_images'] = data.get('images', {})
                                    product['option_names'] = data.get('names', {})
                                else:
                                    # 기존 형식: {"A": "url", "B": "url", ...}
                                    product['option_images'] = data
                        except Exception as e:
                            print(f"[DEBUG] option_images_json 파싱 실패: {e}")

                    # 최종옵션목록에서 option_names 추출 (폴백)
                    final_list_str = product.get('최종옵션목록', '')
                    if final_list_str and isinstance(final_list_str, str):
                        final_list = [x.strip() for x in final_list_str.split('\n') if x.strip()]
                        labels = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'

                        # option_names 없으면 생성
                        if not product.get('option_names'):
                            product['option_names'] = {}
                        for i, opt in enumerate(final_list[:26]):
                            opt_name = opt.split('(')[0].strip() if '(' in opt else opt.strip()
                            if labels[i] not in product['option_names']:
                                product['option_names'][labels[i]] = opt_name

                    # uploadSkus에서 option_images 생성 (가장 신뢰할 수 있는 소스)
                    upload_skus = product.get('uploadSkus', [])
                    if upload_skus and isinstance(upload_skus, list):
                        labels = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
                        product['option_images'] = {}
                        product['option_names'] = product.get('option_names', {})
                        product['option_names_cn'] = {}  # 중국어 원본
                        for idx, sku in enumerate(upload_skus[:26]):
                            url_ref = sku.get('urlRef', '')
                            opt_name = sku.get('text', '') or sku.get('text_ko', '')
                            opt_name_cn = sku.get('_text', '')  # 중국어 원본
                            label = labels[idx]
                            if url_ref:
                                product['option_images'][label] = url_ref
                            if opt_name and label not in product['option_names']:
                                product['option_names'][label] = opt_name
                            if opt_name_cn:
                                product['option_names_cn'][label] = opt_name_cn

                    # option_images 여전히 없으면 대표옵션이미지 사용 (최후 폴백)
                    if not product.get('option_images'):
                        product['option_images'] = {}
                        main_opt_img = product.get('main_option_image', '') or product.get('대표옵션이미지', '')
                        if main_opt_img:
                            product['option_images']['A'] = main_opt_img

                    self.products.append(product)

            wb.close()
            self._log(f"📂 엑셀 로드: {len(self.products)}개 상품")

            # 디버그: 첫 상품 옵션 정보 확인
            if self.products:
                p = self.products[0]
                skus = p.get('uploadSkus', [])
                imgs = p.get('option_images', {})
                self._log(f"   [DEBUG] uploadSkus: {len(skus)}개, option_images: {list(imgs.keys())}")
                if skus and len(skus) > 0:
                    self._log(f"   [DEBUG] 첫 SKU urlRef: {skus[0].get('urlRef', 'NONE')[:50]}...")
                if imgs:
                    first_key = list(imgs.keys())[0]
                    self._log(f"   [DEBUG] 첫 옵션이미지 {first_key}: {imgs[first_key][:50] if imgs[first_key] else 'EMPTY'}...")

            self._update_product_table()

        except Exception as e:
            self._log(f"❌ 엑셀 로드 실패: {e}")
            QMessageBox.critical(self, "오류", f"엑셀 로드 실패:\n{e}")

    def _save_excel(self):
        """엑셀 저장 (검수 결과 포함)"""
        if not OPENPYXL_AVAILABLE:
            QMessageBox.warning(self, "오류", "openpyxl이 설치되지 않았습니다.")
            return

        if not self.products:
            QMessageBox.warning(self, "경고", "저장할 상품이 없습니다.")
            return

        filepath, _ = QFileDialog.getSaveFileName(
            self, "엑셀 저장",
            f"simulation_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            "Excel Files (*.xlsx)"
        )
        if not filepath:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "상품목록"

            # 헤더 (검수 결과 포함)
            headers = ['ID', '상품명', '그룹', '가격', '카테고리', '안전', '위험사유',
                       '전체옵션', '유효옵션', '미끼옵션', '대표옵션', '썸네일URL']
            ws.append(headers)

            # 데이터
            for p in self.products:
                # 카테고리명 추출
                category_name = ''
                upload_cat = p.get('uploadCategory')
                if isinstance(upload_cat, dict):
                    for key in ['ss_category', 'esm_category', 'est_category']:
                        cat_obj = upload_cat.get(key)
                        if isinstance(cat_obj, dict) and cat_obj.get('name'):
                            category_name = cat_obj['name']
                            break
                elif isinstance(upload_cat, str):
                    category_name = upload_cat

                ws.append([
                    p.get('ID', ''),
                    p.get('uploadCommonProductName', ''),
                    p.get('group_name', ''),
                    p.get('uploadCommonSalePrice', 0),
                    category_name,
                    '안전' if p.get('is_safe', True) else '위험',
                    p.get('unsafe_reason', ''),
                    p.get('total_options', 0),
                    p.get('valid_options', 0),
                    p.get('bait_options', 0),
                    p.get('main_option_name', ''),
                    p.get('uploadCommonThumbnail', ''),
                ])

            wb.save(filepath)
            self._log(f"💾 엑셀 저장: {filepath}")
            QMessageBox.information(self, "저장 완료", f"저장됨: {filepath}")

        except Exception as e:
            self._log(f"❌ 저장 실패: {e}")
            QMessageBox.critical(self, "오류", f"저장 실패:\n{e}")

    def _auto_save_excel(self):
        """수집 완료 시 자동 저장 (v3 형식 - 모든 정보 포함)"""
        if not OPENPYXL_AVAILABLE or not self.products:
            return

        try:
            # 자동 파일명 생성 (현재 작업 디렉토리에 저장)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"simulation_{timestamp}.xlsx"
            filepath = os.path.abspath(filename)

            wb = Workbook()
            ws = wb.active
            ws.title = "수집결과"

            # 확장된 헤더 (v3 형식 + 로우데이터) - 최대한 많은 정보
            headers = [
                'ID', '상품명', '그룹', '판매가', '카테고리',
                '안전', '위험사유', '전체옵션', '유효옵션', '최종옵션', '미끼옵션',
                '대표옵션', '대표옵션방식', '최소가(CNY)', '최대가(CNY)',
                '썸네일URL', '대표옵션이미지',
                '옵션이미지JSON', '옵션가격JSON', '옵션명JSON', '중국어옵션명JSON',
                '전체썸네일', '최종옵션목록', '미끼옵션목록',
                'uploadSkusJSON', 'uploadThumbnailsJSON', 'uploadCategoryJSON'  # 로우데이터
            ]
            ws.append(headers)

            # 데이터
            for p in self.products:
                # 카테고리명 추출
                category_name = ''
                upload_cat = p.get('uploadCategory')
                if isinstance(upload_cat, dict):
                    for key in ['ss_category', 'esm_category', 'est_category']:
                        cat_obj = upload_cat.get(key)
                        if isinstance(cat_obj, dict) and cat_obj.get('name'):
                            category_name = cat_obj['name']
                            break
                elif isinstance(upload_cat, str):
                    category_name = upload_cat

                # 옵션 이미지/가격/이름 JSON (개별 저장)
                option_images = p.get('option_images', {})
                option_prices = p.get('option_prices', {})
                option_names = p.get('option_names', {})
                option_names_cn = p.get('option_names_cn', {})

                option_images_json = json.dumps(option_images, ensure_ascii=False) if option_images else ''
                option_prices_json = json.dumps(option_prices, ensure_ascii=False) if option_prices else ''
                option_names_json = json.dumps(option_names, ensure_ascii=False) if option_names else ''
                option_names_cn_json = json.dumps(option_names_cn, ensure_ascii=False) if option_names_cn else ''

                # 전체 썸네일
                all_thumbs = p.get('all_thumbnails', []) or p.get('uploadThumbnails', [])
                all_thumbs_str = '|'.join(all_thumbs) if all_thumbs else ''

                # 최종 옵션 목록
                final_opts = p.get('final_option_list', [])
                final_opts_str = '\n'.join(final_opts) if final_opts else ''

                # 미끼 옵션 목록
                bait_opts = p.get('bait_option_list', [])
                bait_opts_str = ', '.join(bait_opts) if bait_opts else ''

                # 로우데이터 JSON
                upload_skus = p.get('uploadSkus', [])
                upload_skus_json = json.dumps(upload_skus, ensure_ascii=False) if upload_skus else ''
                upload_thumbs = p.get('uploadThumbnails', [])
                upload_thumbs_json = json.dumps(upload_thumbs, ensure_ascii=False) if upload_thumbs else ''
                upload_category = p.get('uploadCategory', {})
                upload_category_json = json.dumps(upload_category, ensure_ascii=False) if upload_category else ''

                ws.append([
                    p.get('ID', ''),
                    p.get('uploadCommonProductName', ''),
                    p.get('group_name', ''),
                    p.get('uploadCommonSalePrice', 0),
                    category_name,
                    '안전' if p.get('is_safe', True) else '위험',
                    p.get('unsafe_reason', ''),
                    p.get('total_options', 0),
                    p.get('valid_options', 0),
                    p.get('final_options', 0),  # 최종옵션
                    p.get('bait_options', 0),
                    p.get('main_option_name', ''),
                    p.get('main_option_method', ''),
                    p.get('min_price_cny', 0),
                    p.get('max_price_cny', 0),
                    p.get('thumbnail_url', ''),
                    p.get('main_option_image', ''),
                    option_images_json,     # 옵션이미지JSON
                    option_prices_json,     # 옵션가격JSON
                    option_names_json,      # 옵션명JSON
                    option_names_cn_json,   # 중국어옵션명JSON
                    all_thumbs_str,
                    final_opts_str,
                    bait_opts_str,
                    upload_skus_json,       # uploadSkusJSON
                    upload_thumbs_json,     # uploadThumbnailsJSON
                    upload_category_json,   # uploadCategoryJSON
                ])

            wb.save(filepath)
            self._log(f"💾 자동 저장 완료!")
            self._log(f"   📁 경로: {filepath}")
            self._log(f"   📊 상품: {len(self.products)}개")

        except Exception as e:
            self._log(f"⚠️ 자동 저장 실패: {e}")

    def _apply_to_bulsaja(self):
        """불사자에 반영"""
        QMessageBox.information(self, "알림", "추후 구현 예정")

    # ============================================================
    # 설정
    # ============================================================
    def _save_keywords(self):
        """키워드 저장"""
        if BULSAJA_API_AVAILABLE:
            keywords = [k.strip() for k in self.keyword_text.toPlainText().split(',') if k.strip()]
            save_bait_keywords(keywords)
            self._log("💾 미끼 키워드 저장됨")
            QMessageBox.information(self, "저장", "미끼 키워드가 저장되었습니다.")

    def _reset_keywords(self):
        """키워드 초기화"""
        if BULSAJA_API_AVAILABLE:
            self.keyword_text.setText(','.join(DEFAULT_BAIT_KEYWORDS))
            self._log("🔄 미끼 키워드 초기화됨")

    def _on_column_checkbox_changed(self):
        """설정 탭 컬럼 체크박스 변경 시"""
        if not hasattr(self, '_settings_column_cbs'):
            return

        # 테이블 컬럼 표시/숨김 적용
        for idx, col_info in enumerate(TABLE_COLUMNS):
            key = col_info['key']
            cb = self._settings_column_cbs.get(key)
            if cb:
                self.product_table.setColumnHidden(idx, not cb.isChecked())

        # 설정 저장
        config = load_config()
        config['visible_columns'] = [k for k, cb in self._settings_column_cbs.items() if cb.isChecked()]
        save_config(config)

    def _show_category_risk_settings(self):
        """카테고리별 검수 설정 다이얼로그"""
        if not BULSAJA_API_AVAILABLE:
            QMessageBox.warning(self, "오류", "bulsaja_common 모듈이 필요합니다.")
            return

        dialog = QDialog(self)
        dialog.setWindowTitle("🏷️ 카테고리별 검수 설정")
        dialog.setFixedSize(450, 500)

        layout = QVBoxLayout(dialog)

        # 설명
        layout.addWidget(QLabel("카테고리별 검수 수준을 설정합니다. (strict=AI확인, normal=프로그램, skip=제외)"))

        # 스크롤 영역
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll_widget = QWidget()
        scroll_layout = QVBoxLayout(scroll_widget)

        # 카테고리 목록 (대분류)
        main_categories = [
            '패션의류', '패션잡화', '화장품/미용', '디지털/가전',
            '가구/인테리어', '출산/육아', '식품', '스포츠/레저',
            '생활/건강', '여가/생활편의', '면세점', '도서/음반/DVD',
            '캠핑', '낚시', '골프'
        ]

        # 현재 설정 로드
        current_settings = load_category_risk_settings()

        self._category_radios = {}

        for cat in main_categories:
            row_widget = QWidget()
            row_layout = QHBoxLayout(row_widget)
            row_layout.setContentsMargins(0, 2, 0, 2)

            # 카테고리명
            cat_label = QLabel(cat)
            cat_label.setFixedWidth(100)
            row_layout.addWidget(cat_label)

            # 라디오 버튼 그룹
            btn_group = QButtonGroup(row_widget)
            current_level = current_settings.get(cat, 'normal')

            rb_normal = QRadioButton("보통")
            rb_strict = QRadioButton("엄격")
            rb_skip = QRadioButton("제외")

            btn_group.addButton(rb_normal, 0)
            btn_group.addButton(rb_strict, 1)
            btn_group.addButton(rb_skip, 2)

            if current_level == 'strict':
                rb_strict.setChecked(True)
            elif current_level == 'skip':
                rb_skip.setChecked(True)
            else:
                rb_normal.setChecked(True)

            row_layout.addWidget(rb_normal)
            row_layout.addWidget(rb_strict)
            row_layout.addWidget(rb_skip)
            row_layout.addStretch()

            self._category_radios[cat] = btn_group
            scroll_layout.addWidget(row_widget)

        scroll_layout.addStretch()
        scroll.setWidget(scroll_widget)
        layout.addWidget(scroll)

        # 버튼
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()

        cancel_btn = QPushButton("취소")
        cancel_btn.clicked.connect(dialog.reject)
        btn_layout.addWidget(cancel_btn)

        save_btn = QPushButton("저장")
        save_btn.clicked.connect(lambda: self._save_category_risk_settings(dialog))
        btn_layout.addWidget(save_btn)

        layout.addLayout(btn_layout)

        dialog.exec()

    def _save_category_risk_settings(self, dialog):
        """카테고리 검수 설정 저장"""
        if not BULSAJA_API_AVAILABLE:
            return

        settings = load_category_risk_settings()

        for cat, btn_group in self._category_radios.items():
            checked_id = btn_group.checkedId()
            if checked_id == 0:
                settings[cat] = 'normal'
            elif checked_id == 1:
                settings[cat] = 'strict'
            elif checked_id == 2:
                settings[cat] = 'skip'

        save_category_risk_settings(settings)
        self._log("💾 카테고리 검수 설정 저장됨")
        QMessageBox.information(self, "저장", "카테고리 검수 설정이 저장되었습니다.")
        dialog.accept()

    # ============================================================
    # 이미지 로드
    # ============================================================
    def _load_current_page_images(self):
        """현재 페이지 이미지 로드 (썸네일 + 옵션이미지)"""
        if not self.products:
            return

        # 현재 페이지 상품들
        start_idx = self.current_page * self.page_size
        end_idx = min(start_idx + self.page_size, len(self.products))
        page_products = self.products[start_idx:end_idx]

        total_images = 0
        self._log(f"🖼️ 이미지 로드 중... ({len(page_products)}개 상품)")

        for product in page_products:
            product_id = product.get('ID', '')

            # 1. 썸네일 로드
            thumb_url = product.get('_thumb_url', '')
            thumb_key = f"{product_id}_thumb"
            if thumb_url and thumb_key in self.image_labels:
                self._load_image_async(thumb_key, thumb_url)
                total_images += 1

            # 2. 대표 옵션 이미지 로드
            opt_img_url = product.get('_opt_img_url', '')
            opt_key = f"{product_id}_opt"
            if opt_img_url and opt_key in self.image_labels:
                self._load_image_async(opt_key, opt_img_url)
                total_images += 1

            # 3. A~F 옵션 버튼 이미지 로드
            for label in 'ABCDEF':
                opt_url = product.get(f'_opt_{label}_url', '')
                opt_label_key = f"{product_id}_opt_{label}"
                if opt_url and opt_label_key in self.image_labels:
                    self._load_image_async(opt_label_key, opt_url)
                    total_images += 1

        self._log(f"✅ 이미지 로드 요청: {total_images}개")

    # ============================================================
    # 컬럼 설정
    # ============================================================
    def _show_column_settings(self):
        """컬럼 설정 다이얼로그"""
        dialog = QDialog(self)
        dialog.setWindowTitle("⚙️ 표시 컬럼 설정")
        dialog.setFixedSize(300, 400)

        layout = QVBoxLayout(dialog)

        # 설명
        layout.addWidget(QLabel("표시할 컬럼을 선택하세요:"))

        # 컬럼 체크박스들
        self._column_checkboxes = {}
        for col_info in TABLE_COLUMNS:
            key = col_info['key']
            name = col_info['name']

            cb = QCheckBox(name)
            # 현재 컬럼 표시 상태 확인
            col_idx = [c['key'] for c in TABLE_COLUMNS].index(key)
            is_hidden = self.product_table.isColumnHidden(col_idx)
            cb.setChecked(not is_hidden)

            self._column_checkboxes[key] = cb
            layout.addWidget(cb)

        layout.addStretch()

        # 버튼
        btn_layout = QHBoxLayout()

        apply_btn = QPushButton("적용")
        apply_btn.clicked.connect(lambda: self._apply_column_settings(dialog))
        btn_layout.addWidget(apply_btn)

        cancel_btn = QPushButton("취소")
        cancel_btn.clicked.connect(dialog.reject)
        btn_layout.addWidget(cancel_btn)

        layout.addLayout(btn_layout)

        dialog.exec()

    def _apply_column_settings(self, dialog):
        """컬럼 설정 적용"""
        for idx, col_info in enumerate(TABLE_COLUMNS):
            key = col_info['key']
            cb = self._column_checkboxes.get(key)
            if cb:
                self.product_table.setColumnHidden(idx, not cb.isChecked())

        # 설정 저장
        config = load_config()
        config['visible_columns'] = [k for k, cb in self._column_checkboxes.items() if cb.isChecked()]
        save_config(config)

        self._log("✅ 컬럼 설정 적용됨")
        dialog.accept()

    def _on_column_moved(self, logical_idx: int, old_visual: int, new_visual: int):
        """컬럼 위치 변경 시 저장"""
        header = self.product_table.horizontalHeader()
        column_order = [header.logicalIndex(i) for i in range(header.count())]
        config = load_config()
        config['column_order'] = column_order
        save_config(config)

    def _load_column_settings(self):
        """저장된 컬럼 설정 불러오기"""
        config = load_config()
        visible_columns = config.get('visible_columns')
        column_order = config.get('column_order')

        # 컬럼 표시/숨김
        if visible_columns:
            for idx, col_info in enumerate(TABLE_COLUMNS):
                key = col_info['key']
                should_show = key in visible_columns
                self.product_table.setColumnHidden(idx, not should_show)

        # 컬럼 순서 복원
        if column_order and len(column_order) == len(TABLE_COLUMNS):
            header = self.product_table.horizontalHeader()
            for visual_idx, logical_idx in enumerate(column_order):
                current_visual = header.visualIndex(logical_idx)
                if current_visual != visual_idx:
                    header.moveSection(current_visual, visual_idx)

    # ============================================================
    # 유틸
    # ============================================================
    def _log(self, message: str):
        """로그 출력"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.append(f"[{timestamp}] {message}")

    # ============================================================
    # 썸네일 분석
    # ============================================================
    def _analyze_thumbnails(self):
        """선택 상품 썸네일 분석 → 자동으로 최고 점수 썸네일을 메인으로 일괄 적용"""
        if not THUMBNAIL_ANALYZER_AVAILABLE:
            QMessageBox.warning(self, "오류", "thumbnail_analyzer 모듈이 필요합니다.\npip install opencv-python pillow easyocr")
            return

        # 선택된 상품만 (필수)
        rows = self.product_table.selectedItems()
        if not rows:
            QMessageBox.warning(self, "경고", "분석할 상품을 선택해주세요.")
            return

        selected_rows = list(set(item.row() for item in rows))
        start_idx = self.current_page * self.page_size
        products_to_analyze = [self.products[start_idx + r] for r in selected_rows if start_idx + r < len(self.products)]

        if not products_to_analyze:
            QMessageBox.warning(self, "경고", "분석할 상품이 없습니다.")
            return

        self._log(f"🔍 썸네일 분석 시작: {len(products_to_analyze)}개 상품 (멀티스레드)")
        self.thumbnail_analysis_btn.setEnabled(False)
        self.thumbnail_analysis_btn.setText("분석중...")
        QApplication.processEvents()

        # 멀티스레드 분석 함수
        def analyze_single_product(product):
            """단일 상품 썸네일 분석"""
            try:
                analyzer = ThumbnailAnalyzer()
                thumbnails = product.get('all_thumbnails', []) or product.get('uploadThumbnails', []) or []

                if not thumbnails or len(thumbnails) < 2:
                    return None

                best_idx, best_score, action = analyzer.get_best_thumbnail(thumbnails)
                return {
                    'product': product,
                    'best_idx': best_idx,
                    'best_score': best_score,
                    'thumbnails': thumbnails
                }
            except Exception as e:
                return None

        # 멀티스레드로 병렬 분석 (최대 4개 동시)
        changed_count = 0
        analyzed_count = 0

        with ThreadPoolExecutor(max_workers=4) as executor:
            futures = {executor.submit(analyze_single_product, p): p for p in products_to_analyze}

            for future in futures:
                result = future.result()
                if result:
                    analyzed_count += 1
                    product = result['product']
                    best_idx = result['best_idx']
                    best_score = result['best_score']
                    thumbnails = result['thumbnails']

                    # 최고 점수 썸네일을 메인으로 자동 설정
                    if best_idx != 0 and best_idx < len(thumbnails):
                        # 썸네일 순서 변경 (best를 맨 앞으로)
                        new_thumbnails = [thumbnails[best_idx]] + [t for i, t in enumerate(thumbnails) if i != best_idx]
                        product['uploadThumbnails'] = new_thumbnails
                        product['all_thumbnails'] = new_thumbnails
                        product['uploadCommonThumbnail'] = new_thumbnails[0]
                        product['thumbnail_url'] = new_thumbnails[0]
                        changed_count += 1

                        self._log(f"  ✅ {product.get('uploadCommonProductName', '')[:25]}... #{best_idx+1}→#1 (점수:{best_score.total_score if best_score else 0})")
                    else:
                        self._log(f"  ⏭️ {product.get('uploadCommonProductName', '')[:25]}... 이미 최적")

                    # 분석 결과 저장
                    product['_thumbnail_analysis'] = {
                        'total_score': best_score.total_score if best_score else 0,
                        'is_nukki': best_score.is_nukki if best_score else False,
                        'recommendation': best_score.recommendation if best_score else 'unknown'
                    }

        self.thumbnail_analysis_btn.setEnabled(True)
        self.thumbnail_analysis_btn.setText("🔍 썸네일자동선택")

        self._log(f"✅ 분석 완료: {analyzed_count}개 분석, {changed_count}개 변경")

        # 테이블 새로고침 (변경된 썸네일 반영)
        if changed_count > 0:
            self._update_product_table()
            self._auto_save_excel()  # 자동 저장
            QMessageBox.information(self, "썸네일 분석 완료",
                f"분석: {analyzed_count}개\n변경: {changed_count}개\n\n메인 썸네일이 자동 변경되었습니다.")

    def _show_thumbnail_selector(self, products: list):
        """썸네일 수동 선택 다이얼로그 (복수 상품 지원)"""
        if not products:
            return

        dialog = QDialog(self)
        dialog.setWindowTitle(f"🖼️ 썸네일 선택 ({len(products)}개 상품)")
        dialog.setMinimumSize(900, 700)

        main_layout = QVBoxLayout(dialog)

        # 상태 표시 라벨
        status_label = QLabel("썸네일을 클릭하여 메인으로 선택하세요. 분석 실행 시 점수가 표시됩니다.")
        status_label.setStyleSheet("color: #666; padding: 5px;")
        main_layout.addWidget(status_label)

        # 스크롤 영역
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll_widget = QWidget()
        scroll_layout = QVBoxLayout(scroll_widget)

        # 상품별 썸네일 선택 데이터 저장
        self._thumb_selector_data = {}

        for product in products:
            product_id = product.get('_id', '')
            product_name = product.get('uploadCommonProductName', '')[:40]
            thumbnails = product.get('all_thumbnails', []) or product.get('uploadThumbnails', []) or []

            if not thumbnails:
                continue

            # 상품별 그룹 박스
            group_box = QGroupBox(f"📦 {product_name}... ({len(thumbnails)}장)")
            group_box.setStyleSheet("""
                QGroupBox {
                    font-weight: bold;
                    border: 1px solid #ddd;
                    border-radius: 5px;
                    margin-top: 10px;
                    padding: 10px;
                }
                QGroupBox::title {
                    subcontrol-origin: margin;
                    left: 10px;
                    padding: 0 5px;
                }
            """)
            group_layout = QVBoxLayout(group_box)

            # 썸네일 그리드 (가로로 나열)
            thumb_grid = QHBoxLayout()
            thumb_grid.setSpacing(10)

            # 썸네일 라벨 저장용
            thumb_labels = []

            for idx, thumb_url in enumerate(thumbnails[:10]):  # 최대 10장
                # 썸네일 컨테이너
                thumb_container = QWidget()
                thumb_container.setFixedSize(110, 150)
                thumb_v_layout = QVBoxLayout(thumb_container)
                thumb_v_layout.setContentsMargins(2, 2, 2, 2)
                thumb_v_layout.setSpacing(3)

                # 썸네일 이미지 라벨
                thumb_label = QLabel()
                thumb_label.setFixedSize(100, 100)
                thumb_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
                thumb_label.setStyleSheet("""
                    QLabel {
                        border: 3px solid #ddd;
                        border-radius: 5px;
                        background-color: #f5f5f5;
                    }
                """)
                thumb_label.setProperty("thumb_idx", idx)
                thumb_label.setProperty("product_id", product_id)
                thumb_label.setCursor(Qt.CursorShape.PointingHandCursor)

                # 클릭 이벤트
                def make_click_handler(pid, tidx, labels):
                    def handler(event):
                        self._on_thumbnail_clicked(pid, tidx, labels)
                    return handler
                thumb_label.mousePressEvent = make_click_handler(product_id, idx, thumb_labels)

                thumb_labels.append(thumb_label)
                thumb_v_layout.addWidget(thumb_label)

                # 썸네일 번호 + 점수 라벨
                info_label = QLabel(f"#{idx+1}")
                info_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
                info_label.setStyleSheet("font-size: 10px; color: #666;")
                info_label.setProperty("info_label", True)
                thumb_v_layout.addWidget(info_label)

                # 이미지 비동기 로드
                if hasattr(self, 'image_loader') and self.image_loader:
                    self.image_loader.load_image(
                        thumb_url,
                        lambda pid=product_id, pix=None, lbl=thumb_label: lbl.setPixmap(pix.scaled(100, 100, Qt.AspectRatioMode.KeepAspectRatio)) if pix else None
                    )
                else:
                    # 동기 로드 시도
                    try:
                        import requests
                        response = requests.get(thumb_url, timeout=5)
                        pix = QPixmap()
                        pix.loadFromData(response.content)
                        thumb_label.setPixmap(pix.scaled(100, 100, Qt.AspectRatioMode.KeepAspectRatio))
                    except:
                        thumb_label.setText("로딩중...")

                thumb_grid.addWidget(thumb_container)

            thumb_grid.addStretch()
            group_layout.addLayout(thumb_grid)

            # 선택 정보 저장
            self._thumb_selector_data[product_id] = {
                'product': product,
                'thumbnails': thumbnails,
                'thumb_labels': thumb_labels,
                'selected_idx': 0,  # 기본: 첫 번째 선택
                'scores': None  # 분석 후 채워짐
            }

            # 첫 번째 썸네일 기본 선택 표시
            if thumb_labels:
                thumb_labels[0].setStyleSheet("""
                    QLabel {
                        border: 3px solid #4CAF50;
                        border-radius: 5px;
                        background-color: #E8F5E9;
                    }
                """)

            scroll_layout.addWidget(group_box)

        scroll.setWidget(scroll_widget)
        main_layout.addWidget(scroll)

        # 하단 버튼
        btn_layout = QHBoxLayout()

        # 분석 버튼
        analyze_btn = QPushButton("🔍 분석 실행 (점수 확인)")
        analyze_btn.setStyleSheet("background-color: #2196F3; color: white; padding: 8px 15px;")
        analyze_btn.clicked.connect(lambda: self._run_thumbnail_analysis(dialog, status_label))
        btn_layout.addWidget(analyze_btn)

        # 자동 선택 버튼
        auto_select_btn = QPushButton("⭐ 최고점수 자동선택")
        auto_select_btn.setStyleSheet("background-color: #FF9800; color: white; padding: 8px 15px;")
        auto_select_btn.clicked.connect(lambda: self._auto_select_best_thumbnails(dialog))
        btn_layout.addWidget(auto_select_btn)

        btn_layout.addStretch()

        # 적용 버튼 (로컬만)
        apply_local_btn = QPushButton("✅ 로컬 적용")
        apply_local_btn.setStyleSheet("background-color: #4CAF50; color: white; padding: 8px 15px;")
        apply_local_btn.clicked.connect(lambda: self._apply_thumbnail_selection(dialog, local_only=True))
        btn_layout.addWidget(apply_local_btn)

        # 불사자 반영 버튼
        apply_api_btn = QPushButton("📤 불사자 반영")
        apply_api_btn.setStyleSheet("background-color: #9C27B0; color: white; padding: 8px 15px;")
        apply_api_btn.setEnabled(self.api_client is not None)
        apply_api_btn.clicked.connect(lambda: self._apply_thumbnail_selection(dialog, local_only=False))
        btn_layout.addWidget(apply_api_btn)

        # 닫기
        close_btn = QPushButton("닫기")
        close_btn.clicked.connect(dialog.reject)
        btn_layout.addWidget(close_btn)

        main_layout.addLayout(btn_layout)

        dialog.exec()

    def _on_thumbnail_clicked(self, product_id: str, thumb_idx: int, thumb_labels: list):
        """썸네일 클릭 시 선택 처리"""
        if product_id not in self._thumb_selector_data:
            return

        # 모든 라벨 선택 해제
        for lbl in thumb_labels:
            lbl.setStyleSheet("""
                QLabel {
                    border: 3px solid #ddd;
                    border-radius: 5px;
                    background-color: #f5f5f5;
                }
            """)

        # 클릭한 라벨 선택 표시
        if thumb_idx < len(thumb_labels):
            thumb_labels[thumb_idx].setStyleSheet("""
                QLabel {
                    border: 3px solid #4CAF50;
                    border-radius: 5px;
                    background-color: #E8F5E9;
                }
            """)

        # 선택 인덱스 저장
        self._thumb_selector_data[product_id]['selected_idx'] = thumb_idx

    def _run_thumbnail_analysis(self, dialog, status_label):
        """썸네일 분석 실행 (점수 계산)"""
        status_label.setText("분석 중... 잠시 기다려주세요.")
        status_label.setStyleSheet("color: #2196F3; padding: 5px; font-weight: bold;")
        QApplication.processEvents()

        analyzer = ThumbnailAnalyzer()
        analyzed_count = 0

        for product_id, data in self._thumb_selector_data.items():
            thumbnails = data['thumbnails']
            thumb_labels = data['thumb_labels']

            if len(thumbnails) < 2:
                continue

            try:
                # 전체 썸네일 분석
                results = analyzer.analyze_thumbnails(thumbnails)
                data['scores'] = {r.index: r for r in results}
                analyzed_count += 1

                # 라벨 업데이트 (점수 표시)
                for idx, lbl in enumerate(thumb_labels):
                    parent = lbl.parentWidget()
                    if parent:
                        for child in parent.children():
                            if isinstance(child, QLabel) and child.property("info_label"):
                                score_info = data['scores'].get(idx)
                                if score_info:
                                    score_text = f"#{idx+1} ({score_info.total_score}점)"
                                    if score_info.is_nukki:
                                        score_text += " 🔲"
                                    if score_info.has_text:
                                        score_text += " 📝"
                                    child.setText(score_text)

                                    # 최고 점수에 표시
                                    if results and results[0].index == idx:
                                        child.setStyleSheet("font-size: 10px; color: #4CAF50; font-weight: bold;")
                                    else:
                                        child.setStyleSheet("font-size: 10px; color: #666;")
                                break

            except Exception as e:
                self._log(f"분석 오류 ({product_id}): {e}")

        status_label.setText(f"✅ 분석 완료: {analyzed_count}개 상품. 썸네일을 클릭하여 선택하세요.")
        status_label.setStyleSheet("color: #4CAF50; padding: 5px; font-weight: bold;")

    def _auto_select_best_thumbnails(self, dialog):
        """모든 상품에서 최고점수 썸네일 자동 선택"""
        if not hasattr(self, '_thumb_selector_data'):
            QMessageBox.warning(self, "알림", "먼저 '분석 실행'을 해주세요.")
            return

        changed = 0
        for product_id, data in self._thumb_selector_data.items():
            scores = data.get('scores')
            if not scores:
                continue

            # 최고 점수 찾기
            best_score = max(scores.values(), key=lambda x: x.total_score)
            best_idx = best_score.index

            if best_idx != data['selected_idx']:
                # 선택 변경
                self._on_thumbnail_clicked(product_id, best_idx, data['thumb_labels'])
                changed += 1

        if changed > 0:
            QMessageBox.information(self, "자동 선택", f"{changed}개 상품의 메인 썸네일이 최고 점수로 변경되었습니다.")
        else:
            QMessageBox.information(self, "자동 선택", "이미 모든 상품이 최적 썸네일로 선택되어 있습니다.")

    def _apply_thumbnail_selection(self, dialog, local_only: bool = True):
        """선택된 썸네일 적용"""
        if not hasattr(self, '_thumb_selector_data'):
            return

        changed_count = 0
        api_success = 0
        api_fail = 0

        for product_id, data in self._thumb_selector_data.items():
            product = data['product']
            thumbnails = data['thumbnails']
            selected_idx = data['selected_idx']

            if selected_idx == 0:
                # 이미 첫 번째가 선택됨 → 변경 필요 없음
                continue

            # 썸네일 순서 변경
            new_thumbnails = thumbnails.copy()
            selected_thumb = new_thumbnails.pop(selected_idx)
            new_thumbnails.insert(0, selected_thumb)

            # 로컬 데이터 업데이트
            product['uploadThumbnails'] = new_thumbnails
            product['all_thumbnails'] = new_thumbnails
            product['uploadCommonThumbnail'] = new_thumbnails[0]
            product['thumbnail_url'] = new_thumbnails[0]
            changed_count += 1

            # API 반영
            if not local_only and self.api_client:
                try:
                    update_data = {'uploadThumbnails': new_thumbnails}
                    success, msg = self.api_client.update_product_fields(product_id, update_data)
                    if success:
                        api_success += 1
                        self._log(f"  ✅ {product.get('uploadCommonProductName', '')[:25]}: #{selected_idx+1} → #1")
                    else:
                        api_fail += 1
                        self._log(f"  ❌ {product.get('uploadCommonProductName', '')[:25]}: {msg}")
                except Exception as e:
                    api_fail += 1
                    self._log(f"  ❌ {product.get('uploadCommonProductName', '')[:25]}: {e}")

        # 결과 메시지
        if local_only:
            if changed_count > 0:
                self._update_product_table()
                QMessageBox.information(self, "적용 완료", f"로컬 변경: {changed_count}개\n\n테이블이 새로고침되었습니다.")
            else:
                QMessageBox.information(self, "알림", "변경된 항목이 없습니다.")
        else:
            self._update_product_table()
            QMessageBox.information(self, "반영 완료",
                f"로컬 변경: {changed_count}개\n불사자 성공: {api_success}개\n불사자 실패: {api_fail}개")

        dialog.accept()

    def _show_thumbnail_analysis_result(self, results: list):
        """썸네일 분석 결과 다이얼로그"""
        self._thumb_analysis_results = results  # 반영용 저장

        dialog = QDialog(self)
        dialog.setWindowTitle("🔍 썸네일 분석 결과")
        dialog.setFixedSize(850, 550)

        layout = QVBoxLayout(dialog)

        # 요약
        nukki_count = sum(1 for r in results if r['score'] and r['score'].is_nukki)
        text_count = sum(1 for r in results if r['score'] and r['score'].has_text)
        best_count = sum(1 for r in results if r['score'] and r['score'].recommendation == 'best')
        change_count = sum(1 for r in results if r.get('need_change', False))

        summary = QLabel(f"""
📊 분석 요약 ({len(results)}개 상품)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
✅ 누끼 배경: {nukki_count}개 ({nukki_count*100//len(results) if results else 0}%)
📝 텍스트 있음: {text_count}개 ({text_count*100//len(results) if results else 0}%)
⭐ 즉시사용가능(best): {best_count}개
🔄 메인 변경 필요: {change_count}개
        """)
        summary.setStyleSheet("font-size: 12px; padding: 10px; background-color: #E3F2FD; border-radius: 5px;")
        layout.addWidget(summary)

        # 상세 테이블
        table = QTableWidget()
        table.setColumnCount(9)
        table.setHorizontalHeaderLabels(['상품명', '썸네일', '최적', '변경', '누끼', '점수', '텍스트', '총점', '추천'])
        table.setRowCount(len(results))

        header = table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        for i in range(1, 9):
            header.setSectionResizeMode(i, QHeaderView.ResizeMode.Fixed)
            table.setColumnWidth(i, 55)

        for row, r in enumerate(results):
            score = r.get('score')
            if not score:
                continue

            table.setItem(row, 0, QTableWidgetItem(r['product_name']))

            # 썸네일 개수
            thumb_item = QTableWidgetItem(str(r.get('thumbnail_count', 0)))
            thumb_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            table.setItem(row, 1, thumb_item)

            # 최적 썸네일 번호
            best_idx = r.get('best_index', 0)
            best_item = QTableWidgetItem(f"#{best_idx + 1}")
            best_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            table.setItem(row, 2, best_item)

            # 변경 필요 여부
            need_change = r.get('need_change', False)
            change_item = QTableWidgetItem("🔄" if need_change else "-")
            change_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            if need_change:
                change_item.setBackground(QColor("#BBDEFB"))
            table.setItem(row, 3, change_item)

            # 누끼
            nukki_item = QTableWidgetItem("✅" if score.is_nukki else "❌")
            nukki_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            if score.is_nukki:
                nukki_item.setBackground(QColor("#C8E6C9"))
            else:
                nukki_item.setBackground(QColor("#FFCDD2"))
            table.setItem(row, 4, nukki_item)

            table.setItem(row, 5, QTableWidgetItem(str(score.nukki_score)))

            # 텍스트
            text_item = QTableWidgetItem(str(score.text_count) if score.has_text else "-")
            text_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            if score.has_text:
                text_item.setBackground(QColor("#FFF9C4"))
            table.setItem(row, 6, text_item)

            table.setItem(row, 7, QTableWidgetItem(str(score.total_score)))

            rec_text = {
                'best': '⭐최고',
                'needs_nukki': '🔲누끼',
                'needs_translate': '📝번역',
                'needs_both': '⚠️둘다',
                'poor': '❌재촬영',
                'error': '⚠️오류'
            }.get(score.recommendation, score.recommendation)
            table.setItem(row, 8, QTableWidgetItem(rec_text))

        layout.addWidget(table)

        # 버튼
        btn_layout = QHBoxLayout()

        apply_btn = QPushButton(f"📤 불사자 반영 ({change_count}개)")
        apply_btn.setEnabled(change_count > 0 and self.api_client is not None)
        apply_btn.setStyleSheet("background-color: #4CAF50; color: white; font-weight: bold; padding: 8px;")
        apply_btn.clicked.connect(lambda: self._apply_thumbnail_changes(dialog))
        btn_layout.addWidget(apply_btn)

        btn_layout.addStretch()

        close_btn = QPushButton("닫기")
        close_btn.clicked.connect(dialog.accept)
        btn_layout.addWidget(close_btn)

        layout.addLayout(btn_layout)

        dialog.exec()

    def _apply_thumbnail_changes(self, dialog):
        """분석된 최적 썸네일을 불사자에 반영"""
        if not self.api_client:
            QMessageBox.warning(self, "오류", "API 연결이 필요합니다.")
            return

        results = getattr(self, '_thumb_analysis_results', [])
        change_items = [r for r in results if r.get('need_change', False)]

        if not change_items:
            QMessageBox.information(self, "알림", "변경할 상품이 없습니다.")
            return

        # 확인
        reply = QMessageBox.question(
            self, "확인",
            f"{len(change_items)}개 상품의 메인 썸네일을 변경하시겠습니까?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if reply != QMessageBox.StandardButton.Yes:
            return

        self._log(f"📤 썸네일 반영 시작: {len(change_items)}개")

        success_count = 0
        fail_count = 0

        for item in change_items:
            product_id = item['product_id']
            best_idx = item['best_index']
            thumbnails = item.get('thumbnails', [])

            if not thumbnails or best_idx >= len(thumbnails):
                self._log(f"  ❌ {product_id}: 썸네일 데이터 없음")
                fail_count += 1
                continue

            # 썸네일 순서 재정렬 (best를 맨 앞으로)
            new_thumbnails = thumbnails.copy()
            best_thumb = new_thumbnails.pop(best_idx)
            new_thumbnails.insert(0, best_thumb)

            # API 업데이트
            try:
                update_data = {
                    'uploadThumbnails': new_thumbnails
                }
                success, msg = self.api_client.update_product_fields(product_id, update_data)

                if success:
                    self._log(f"  ✅ {item['product_name'][:20]}: #{best_idx+1} → #1")
                    success_count += 1
                else:
                    self._log(f"  ❌ {item['product_name'][:20]}: {msg}")
                    fail_count += 1

            except Exception as e:
                self._log(f"  ❌ {item['product_name'][:20]}: {e}")
                fail_count += 1

        self._log(f"✅ 썸네일 반영 완료: 성공 {success_count}개, 실패 {fail_count}개")
        QMessageBox.information(self, "완료", f"성공: {success_count}개\n실패: {fail_count}개")

        dialog.accept()

    # ============================================================
    # 지재권 분석
    # ============================================================
    def _auto_ip_analysis(self):
        """수집 완료 후 자동 지재권 분석 (형태소 분석 → AI 검증)"""
        if not BULSAJA_API_AVAILABLE or not self.products:
            return

        try:
            self._log(f"🏷️ 자동 지재권 분석 중...")

            # 1단계: 형태소 분석으로 의심 단어 추출
            result = analyze_products_for_ip(self.products)

            if result and result.get('products_with_issues', 0) > 0:
                suspicious_words = list(result.get('suspicious_words', {}).keys())
                self._log(f"  📋 형태소 분석: {len(suspicious_words)}종 의심 단어 발견")

                # 2단계: Gemini API 키 확인 후 AI 검증
                ai_config = load_ai_config()
                api_key = ai_config.get('gemini', {}).get('api_key', '')

                ip_confirmed_words = set()

                if api_key and suspicious_words:
                    self._log(f"  🤖 Gemini AI 검증 중... ({len(suspicious_words[:50])}개 단어)")
                    try:
                        verified = verify_ip_words_with_ai(suspicious_words[:50], log_callback=self._log)
                        if verified:
                            ip_confirmed_words = set(verified.get('ip_confirmed', []))
                            ip_safe = verified.get('ip_safe', [])
                            ip_uncertain = verified.get('ip_uncertain', [])
                            self._log(f"  ✅ AI 검증 완료: 🔴IP확정 {len(ip_confirmed_words)}개, 🟢안전 {len(ip_safe)}개, 🟡불확실 {len(ip_uncertain)}개")
                    except Exception as e:
                        self._log(f"  ⚠️ AI 검증 실패: {e}")
                        # AI 실패 시 형태소 분석 결과만 사용
                        ip_confirmed_words = set(suspicious_words)
                elif not api_key:
                    self._log(f"  ⚠️ Gemini API 키 미설정 - 형태소 분석 결과만 사용")
                    ip_confirmed_words = set(suspicious_words)

                # 3단계: IP 확정 단어가 포함된 상품에 위험 마킹
                ip_count = 0
                for product in self.products:
                    product_name = product.get('product_name', '') or product.get('name', '') or product.get('uploadCommonProductName', '')
                    # 상품명에 IP 확정 단어가 포함되어 있는지 확인
                    for ip_word in ip_confirmed_words:
                        if ip_word in product_name:
                            product['_ip_warning'] = True
                            product['_ip_reason'] = f"지재권 의심: {ip_word}"
                            # 안전 여부도 위험으로 변경
                            if product.get('_is_safe') != 'X':
                                product['_is_safe'] = 'X'
                                product['_unsafe_reason'] = f"IP침해:{ip_word}"
                            ip_count += 1
                            break

                if ip_count > 0:
                    self._log(f"  ⚠️ 지재권 위험 상품: {ip_count}개 (위험 분류됨)")
                else:
                    self._log(f"  ✅ 지재권 확정 상품 없음")
            else:
                self._log(f"  ✅ 지재권 의심 단어 없음")

        except Exception as e:
            self._log(f"  ⚠️ 지재권 분석 스킵: {e}")

    def _check_product_names_with_ai(self):
        """상품명 의심 단어 AI 검증"""
        if not BULSAJA_API_AVAILABLE:
            QMessageBox.warning(self, "오류", "bulsaja_common 모듈이 필요합니다.")
            return

        if not self.products:
            QMessageBox.warning(self, "경고", "검수할 상품이 없습니다.\n먼저 상품을 수집하세요.")
            return

        # AI 설정 확인
        ai_config = load_ai_config()
        api_key = ai_config.get('gemini', {}).get('api_key', '')
        if not api_key:
            reply = QMessageBox.question(
                self, "API 키 필요",
                "Gemini API 키가 설정되지 않았습니다.\n\n패턴 기반 검수만 실행할까요?\n(AI 검증 없이 빠르게 진행)",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            if reply == QMessageBox.StandardButton.No:
                return
            use_ai = False
        else:
            reply = QMessageBox.question(
                self, "AI 검증",
                f"상품 {len(self.products)}개의 상품명을 검수합니다.\n\nAI 검증을 사용할까요?\n(AI 검증 시 더 정확하지만 시간이 걸립니다)",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            use_ai = (reply == QMessageBox.StandardButton.Yes)

        self._log(f"📝 상품명 검수 시작: {len(self.products)}개 상품 (AI: {'O' if use_ai else 'X'})")

        self.name_check_btn.setEnabled(False)
        self.name_check_btn.setText("검수중...")
        QApplication.processEvents()

        try:
            # 일괄 검수
            batch_check_product_names(self.products, use_ai=use_ai, log_callback=self._log)

            # 테이블 새로고침 (검수 결과 반영)
            self._update_product_table()

            # 통계
            suspicious_count = sum(1 for p in self.products if p.get('name_check_result', {}).get('suspicious_words'))

            self.name_check_btn.setEnabled(True)
            self.name_check_btn.setText("📝 상품명검수")

            self._log(f"✅ 상품명 검수 완료: {suspicious_count}/{len(self.products)}개 의심")

            if suspicious_count > 0:
                # 의심 상품 목록 표시
                self._show_name_check_result(suspicious_count)
            else:
                QMessageBox.information(self, "검수 완료", "의심 단어가 발견되지 않았습니다.")

        except Exception as e:
            self._log(f"❌ 검수 오류: {e}")
            import traceback
            traceback.print_exc()
            self.name_check_btn.setEnabled(True)
            self.name_check_btn.setText("📝 상품명검수")
            QMessageBox.warning(self, "오류", f"검수 중 오류 발생:\n{e}")

    def _show_name_check_result(self, suspicious_count: int):
        """상품명 검수 결과 표시"""
        dialog = QDialog(self)
        dialog.setWindowTitle("📝 상품명 검수 결과")
        dialog.setFixedSize(800, 500)

        layout = QVBoxLayout(dialog)

        # 요약
        summary = QLabel(f"""
📊 검수 결과
━━━━━━━━━━━━━━━━━━━━━━━━━━
📦 검수 상품: {len(self.products)}개
⚠️ 의심 상품: {suspicious_count}개
        """)
        summary.setStyleSheet("font-size: 12px; padding: 10px; background-color: #FFF3E0; border-radius: 5px;")
        layout.addWidget(summary)

        # 의심 상품 목록
        table = QTableWidget()
        table.setColumnCount(3)
        table.setHorizontalHeaderLabels(['상품명', '의심 단어', '카테고리'])
        table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        table.setColumnWidth(1, 150)
        table.setColumnWidth(2, 100)

        # 의심 상품만 필터
        suspicious_products = [p for p in self.products if p.get('name_check_result', {}).get('suspicious_words')]
        table.setRowCount(len(suspicious_products))

        for row, product in enumerate(suspicious_products):
            name = product.get('uploadCommonProductName', '')[:50]
            check_result = product.get('name_check_result', {})
            words = check_result.get('suspicious_words', [])

            word_str = ', '.join([w['word'] for w in words])
            category_str = ', '.join(set([w.get('category', '') for w in words]))

            table.setItem(row, 0, QTableWidgetItem(name))

            word_item = QTableWidgetItem(word_str)
            word_item.setForeground(QColor('red'))
            word_item.setBackground(QColor('#FFFDE7'))
            table.setItem(row, 1, word_item)

            table.setItem(row, 2, QTableWidgetItem(category_str))

        layout.addWidget(table)

        # 안내
        info = QLabel("💡 의심 단어가 포함된 상품은 테이블에서 노란 배경으로 표시됩니다.")
        layout.addWidget(info)

        # 닫기 버튼
        close_btn = QPushButton("닫기")
        close_btn.clicked.connect(dialog.close)
        layout.addWidget(close_btn)

        dialog.exec()

    def _analyze_ip(self):
        """상품명 지재권 분석"""
        if not BULSAJA_API_AVAILABLE:
            QMessageBox.warning(self, "오류", "bulsaja_common 모듈이 필요합니다.")
            return

        if not self.products:
            QMessageBox.warning(self, "경고", "분석할 상품이 없습니다.\n먼저 상품을 수집하세요.")
            return

        self._log(f"🏷️ 지재권 분석 시작: {len(self.products)}개 상품")

        # 동기 방식으로 변경 (UI 응답성 유지를 위해 processEvents 사용)
        self.ip_analysis_btn.setEnabled(False)
        self.ip_analysis_btn.setText("분석중...")
        QApplication.processEvents()

        try:
            result = analyze_products_for_ip(self.products, log_callback=self._log)

            self.ip_analysis_btn.setEnabled(True)
            self.ip_analysis_btn.setText("🏷️ 지재권분석")

            if result and result.get('products_with_issues', 0) > 0:
                self._log(f"✅ 지재권 분석 완료: {result['products_with_issues']}개 의심 상품")
                self._show_ip_analysis_result(result)
            elif result:
                self._log(f"✅ 분석 완료: 의심 단어 없음")
                QMessageBox.information(self, "분석 완료", "지재권 의심 단어가 발견되지 않았습니다.")
            else:
                self._log("⚠️ 분석 결과 없음")

        except Exception as e:
            self._log(f"❌ 분석 오류: {e}")
            import traceback
            traceback.print_exc()
            self.ip_analysis_btn.setEnabled(True)
            self.ip_analysis_btn.setText("🏷️ 지재권분석")
            QMessageBox.warning(self, "오류", f"분석 중 오류 발생:\n{e}")

    def _show_ip_analysis_result(self, result: dict):
        """지재권 분석 결과 다이얼로그"""
        self._ip_analysis_result = result  # AI 검증용 저장

        dialog = QDialog(self)
        dialog.setWindowTitle("🏷️ 지재권 분석 결과")
        dialog.setFixedSize(900, 650)

        layout = QVBoxLayout(dialog)

        # 요약
        self._ip_summary_label = QLabel(f"""
📊 분석 요약 (KoNLPy 형태소 분석)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
📦 분석 상품: {result['total_analyzed']}개
⚠️ 의심 상품: {result['products_with_issues']}개
🏷️ 의심 단어: {len(result['suspicious_words'])}종류
        """)
        self._ip_summary_label.setStyleSheet("font-size: 12px; padding: 10px; background-color: #FFF3E0; border-radius: 5px;")
        layout.addWidget(self._ip_summary_label)

        # 탭 (단어 / 상품 / AI검증)
        self._ip_tabs = QTabWidget()

        # 탭1: 의심 단어 목록
        word_tab = QWidget()
        word_layout = QVBoxLayout(word_tab)

        self._ip_word_table = QTableWidget()
        self._ip_word_table.setColumnCount(3)
        self._ip_word_table.setHorizontalHeaderLabels(['의심 단어', '출현 횟수', 'AI 검증'])
        self._ip_word_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self._ip_word_table.setColumnWidth(1, 80)
        self._ip_word_table.setColumnWidth(2, 100)

        words = list(result['suspicious_words'].items())
        self._ip_word_table.setRowCount(len(words))

        for row, (word, count) in enumerate(words):
            self._ip_word_table.setItem(row, 0, QTableWidgetItem(word))
            count_item = QTableWidgetItem(str(count))
            count_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self._ip_word_table.setItem(row, 1, count_item)
            # AI 검증 결과 (초기값: 미검증)
            ai_item = QTableWidgetItem("⏳ 미검증")
            ai_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self._ip_word_table.setItem(row, 2, ai_item)

        word_layout.addWidget(self._ip_word_table)
        self._ip_tabs.addTab(word_tab, f"📝 의심 단어 ({len(words)})")

        # 탭2: 의심 상품 목록
        product_tab = QWidget()
        product_layout = QVBoxLayout(product_tab)

        product_table = QTableWidget()
        product_table.setColumnCount(3)
        product_table.setHorizontalHeaderLabels(['상품명', 'ID', '의심 단어'])
        product_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        product_table.setColumnWidth(1, 150)
        product_table.setColumnWidth(2, 200)

        ip_products = result['products_with_ip']
        product_table.setRowCount(len(ip_products))

        for row, p in enumerate(ip_products):
            product_table.setItem(row, 0, QTableWidgetItem(p['product_name'][:50]))
            product_table.setItem(row, 1, QTableWidgetItem(p['product_id'][:20]))

            suspicious_words = [s['word'] for s in p['suspicious']]
            product_table.setItem(row, 2, QTableWidgetItem(', '.join(suspicious_words[:5])))

        product_layout.addWidget(product_table)
        self._ip_tabs.addTab(product_tab, f"📦 의심 상품 ({len(ip_products)})")

        layout.addWidget(self._ip_tabs)

        # 버튼
        btn_layout = QHBoxLayout()

        # AI 검증 버튼
        ai_verify_btn = QPushButton(f"🤖 AI 검증 ({len(words)}개 단어)")
        ai_verify_btn.setStyleSheet("background-color: #2196F3; color: white; font-weight: bold; padding: 8px;")
        ai_verify_btn.clicked.connect(lambda: self._verify_ip_with_ai(dialog))
        btn_layout.addWidget(ai_verify_btn)

        btn_layout.addStretch()

        close_btn = QPushButton("닫기")
        close_btn.clicked.connect(dialog.accept)
        btn_layout.addWidget(close_btn)

        layout.addLayout(btn_layout)

        dialog.exec()

    def _verify_ip_with_ai(self, dialog):
        """의심 단어를 AI로 검증"""
        result = getattr(self, '_ip_analysis_result', None)
        if not result:
            return

        words = list(result['suspicious_words'].keys())
        if not words:
            QMessageBox.information(self, "알림", "검증할 단어가 없습니다.")
            return

        self._log(f"🤖 AI 지재권 검증 시작: {len(words)}개 단어")

        # 비동기 검증
        def run_verify():
            return verify_ip_words_with_ai(words[:50], log_callback=self._log)  # 최대 50개

        def on_complete(verified):
            if not verified:
                self._log("⚠️ AI 검증 실패")
                return

            ip_confirmed = verified.get('ip_confirmed', [])
            ip_safe = verified.get('ip_safe', [])
            ip_uncertain = verified.get('ip_uncertain', [])

            self._log(f"✅ AI 검증 완료")
            self._log(f"   🔴 지재권 확정: {len(ip_confirmed)}개")
            self._log(f"   🟢 안전 단어: {len(ip_safe)}개")
            self._log(f"   🟡 불확실: {len(ip_uncertain)}개")

            # 테이블 업데이트
            for row in range(self._ip_word_table.rowCount()):
                word_item = self._ip_word_table.item(row, 0)
                if not word_item:
                    continue
                word = word_item.text()

                ai_item = self._ip_word_table.item(row, 2)
                if word in ip_confirmed:
                    ai_item.setText("🔴 지재권")
                    ai_item.setBackground(QColor("#FFCDD2"))
                elif word in ip_safe:
                    ai_item.setText("🟢 안전")
                    ai_item.setBackground(QColor("#C8E6C9"))
                elif word in ip_uncertain:
                    ai_item.setText("🟡 불확실")
                    ai_item.setBackground(QColor("#FFF9C4"))
                else:
                    ai_item.setText("⚪ 미응답")

            # 요약 업데이트
            self._ip_summary_label.setText(f"""
📊 분석 요약 (KoNLPy + AI 검증 완료)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
📦 분석 상품: {result['total_analyzed']}개
⚠️ 의심 상품: {result['products_with_issues']}개
🏷️ 의심 단어: {len(result['suspicious_words'])}종류
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
🔴 지재권 확정: {len(ip_confirmed)}개
🟢 안전 단어: {len(ip_safe)}개
🟡 불확실: {len(ip_uncertain)}개
            """)
            self._ip_summary_label.setStyleSheet("font-size: 12px; padding: 10px; background-color: #E8F5E9; border-radius: 5px;")

        executor = ThreadPoolExecutor(max_workers=1)
        future = executor.submit(run_verify)

        def check_complete():
            if future.done():
                try:
                    verified = future.result()
                    on_complete(verified)
                except Exception as e:
                    self._log(f"❌ AI 검증 오류: {e}")
            else:
                QTimer.singleShot(500, check_complete)

        QTimer.singleShot(100, check_complete)

# ============================================================
# 메인
# ============================================================
def main():
    app = QApplication(sys.argv)
    app.setStyle("Fusion")

    window = SimulatorGUIv4()
    window.show()

    sys.exit(app.exec())


if __name__ == "__main__":
    main()
