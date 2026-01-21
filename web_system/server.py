#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
êµ¬ë§¤ëŒ€í–‰ í†µí•©ê´€ë¦¬ ì‹œìŠ¤í…œ - ë°±ì—”ë“œ ì„œë²„
- ë‹´ë‹¹ì ë¡œê·¸ì¸ (ì„¸ì…˜ ê¸°ë°˜)
- ê³„ì • ê´€ë¦¬ (êµ¬ê¸€ ì‹œíŠ¸ ì—°ë™)
- SMS í†µí•© (3ê°œ í° í”„ë¡œí•„)
- ì‹¤ì‹œê°„ WebSocket
- ìŠ¤ì¼€ì¤„ëŸ¬ (APScheduler)

ì‹¤í–‰: python server.py
ì ‘ì†: http://localhost:8000 ë˜ëŠ” http://ì„œë²„IP:8000
"""

import os
import re
import sys
import json
import time
import asyncio
import hashlib
import threading

import secrets
import requests
from pathlib import Path
from datetime import datetime, timedelta
from typing import Optional, List, Dict, Any
from dataclasses import dataclass, asdict
from contextlib import asynccontextmanager

from fastapi import FastAPI, HTTPException, Depends, Request, WebSocket, WebSocketDisconnect, BackgroundTasks, File, UploadFile, Form
from fastapi.staticfiles import StaticFiles
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse, StreamingResponse
import io
import pandas as pd
from fastapi.templating import Jinja2Templates
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
import uvicorn

# APScheduler
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.cron import CronTrigger
from modules.delivery_check import DeliveryChecker
from modules.ali_tracking import AliTrackingCollector
from modules.daily_sync import DailyJournalSyncer
from apscheduler.triggers.interval import IntervalTrigger
from apscheduler.jobstores.memory import MemoryJobStore

# Google Sheets
import gspread
from google.oauth2.service_account import Credentials
from dotenv import load_dotenv

# Playwright
from playwright.async_api import async_playwright, Page, BrowserContext

# ========== ì„¤ì • ==========
APP_DIR = Path(__file__).resolve().parent

# .env íŒŒì¼ ë¡œë“œ (í˜„ì¬ í´ë” ë˜ëŠ” ìƒìœ„ í´ë”)
env_path = APP_DIR / ".env"
if not env_path.exists():
    env_path = APP_DIR.parent / ".env"
load_dotenv(env_path)
print(f"[ENV] .env ë¡œë“œ: {env_path} (ì¡´ì¬: {env_path.exists()})")

# Playwright ë¸Œë¼ìš°ì € ê²½ë¡œ ì„¤ì • (sms_gui.pyì™€ ë™ì¼ - ë‹¤ë¥¸ PCì—ì„œë„ ì‘ë™í•˜ë„ë¡)
PLAYWRIGHT_BROWSER_DIR = os.environ.get(
    "PLAYWRIGHT_BROWSERS_PATH",
    str(APP_DIR / "pw_browsers")
)
os.environ["PLAYWRIGHT_BROWSERS_PATH"] = PLAYWRIGHT_BROWSER_DIR
print(f"[PLAYWRIGHT] Playwright ë¸Œë¼ìš°ì € ê²½ë¡œ: {PLAYWRIGHT_BROWSER_DIR}")

# Playwright (SMS ë¸Œë¼ìš°ì €) - async_playwright ì „ìš©
# (sync_playwright import ì œê±°ë¨)

    
# êµ¬ê¸€ ì‹œíŠ¸ ì„¤ì •
def get_rel_path(env_key, default_name):
    env_val = os.environ.get(env_key)
    if env_val:
        # 1. ì ˆëŒ€ ê²½ë¡œì¸ ê²½ìš° ê·¸ëŒ€ë¡œ í™•ì¸
        if os.path.isabs(env_val) and os.path.exists(env_val):
            return env_val
        # 2. APP_DIR ê¸°ì¤€ ìƒëŒ€ ê²½ë¡œ í™•ì¸
        path_in_app = APP_DIR / env_val
        if path_in_app.exists():
            return str(path_in_app)
        # 3. í”„ë¡œì íŠ¸ ë£¨íŠ¸(APP_DIR.parent) ê¸°ì¤€ ìƒëŒ€ ê²½ë¡œ í™•ì¸
        path_in_root = APP_DIR.parent / env_val
        if path_in_root.exists():
            return str(path_in_root)
            
    # ë¡œì»¬ ë””ë ‰í† ë¦¬ì—ì„œ ê¸°ë³¸ ì´ë¦„ìœ¼ë¡œ ì°¾ê¸°
    local_path = APP_DIR / default_name
    if local_path.exists():
        return str(local_path)
    
    # ìƒìœ„ ë””ë ‰í† ë¦¬(í”„ë¡œì íŠ¸ ë£¨íŠ¸)ì—ì„œ ê¸°ë³¸ ì´ë¦„ìœ¼ë¡œ ì°¾ê¸°
    parent_path = APP_DIR.parent / default_name
    if parent_path.exists():
        return str(parent_path)
        
    return str(local_path)

CREDENTIALS_FILE = get_rel_path("SERVICE_ACCOUNT_JSON", "autosms-466614-951e91617c69.json")
SPREADSHEET_KEY = os.environ.get("SPREADSHEET_KEY", "1r-ROJ7ksv6qOtOTXbkrprxu17EQmbO-n1J1pm_N5Hh8")
MARKETING_SPREADSHEET_KEY = os.environ.get("MARKETING_SPREADSHEET_KEY", "14l6Y7y7bHcn6LRGlfQ0QOKHNGWI5tFqPWRioxE8aoTo")

# ë“±ë¡ê°¯ìˆ˜ ì „ìš© ì¸ì¦ íŒŒì¼
COUNT_CREDENTIALS_FILE = get_rel_path("COUNT_CREDENTIALS_FILE", "auto-smartstore-update-61c3a948c45c.json")

# ì‹œíŠ¸ íƒ­ ì´ë¦„
ACCOUNTS_TAB = "ê³„ì •ëª©ë¡"  # í”Œë«í¼ ê³„ì •
USERS_TAB = "ë‹´ë‹¹ì"       # ë‹´ë‹¹ì ë¡œê·¸ì¸ ì •ë³´
SMS_LOG_TAB = "SMSë¡œê·¸"    # SMS ë°œì†¡ ê¸°ë¡ (ì„ íƒ)
WORK_LOG_SHEET = "ì‘ì—…ë¡œê·¸"  # ì‘ì—… ê¸°ë¡ (ìº˜ë¦°ë”ìš©)

# í° í”„ë¡œí•„
PHONE_PROFILES = ["8295", "8217", "4682"]

# í”„ë¡œí•„ ID â†’ ì‹¤ì œ ë””ë ‰í† ë¦¬ ë§¤í•‘ (QR ì¸ì¦ ì˜ëª»ëœ ê²½ìš° ì—¬ê¸°ì„œ êµì²´)
# í˜„ì¬: ì •ìƒ (ë””ë ‰í† ë¦¬ ì´ë¦„ êµì²´ ì™„ë£Œë¨)
PROFILE_DIR_MAPPING = {
    "8295": "8295",
    "8217": "8217",
    "4682": "4682",
}

# PC ì‹ë³„ì (ë¹ˆ ë¬¸ìì—´ - í”„ë¡œí•„ ì§ì ‘ ì‚¬ìš©, ë‹¤ë¥¸ PCë¡œ ë³µì‚¬ ê°€ëŠ¥)
import socket
SERVER_ID = ""  # pw_sessions/8295 ì§ì ‘ ì‚¬ìš©

# ì‹œíŠ¸ ì»¬ëŸ¼ ë§¤í•‘ (ì‹¤ì œ ì‹œíŠ¸ í—¤ë”ì™€ ì¼ì¹˜)
SHEET_COLUMNS = [
    "í”Œë«í¼",
    "ì•„ì´ë””",
    "íŒ¨ìŠ¤ì›Œë“œ",
    "ìŠ¤í† ì–´ëª…",
    "ì‚¬ì—…ìë²ˆí˜¸",
    "ìŠ¤ë§ˆíŠ¸ìŠ¤í† ì–´ API ì—°ë™ìš© íŒë§¤ìID",
    "ìŠ¤ë§ˆíŠ¸ìŠ¤í† ì–´ ì• í”Œë¦¬ì¼€ì´ì…˜ ID",
    "ìŠ¤ë§ˆíŠ¸ìŠ¤í† ì–´ ì• í”Œë¦¬ì¼€ì´ì…˜ ì‹œí¬ë¦¿",
    "ì¿ íŒ¡ ì—…ì²´ì½”ë“œ",
    "ì¿ íŒ¡ Access Key",
    "ì¿ íŒ¡ Secret Key",
    "11ë²ˆê°€ API KEY",
    "ESMí†µí•©ê³„ì •",  # ì§€ë§ˆì¼“/ì˜¥ì…˜ì´ ì—°ê²°ëœ ESMí†µí•© ê³„ì • ID
    "ESMí†µí•©ë¹„ë°€ë²ˆí˜¸"  # ESMí†µí•© ê³„ì • ë¹„ë°€ë²ˆí˜¸
]

# ìŠ¤í† ì–´ëª… í—¤ë” ë³„ì¹­ ëª©ë¡ (í†µì¼)
STORE_NAME_ALIASES = [
    "ìŠ¤í† ì–´ëª…", "store_name", "ì‡¼í•‘ëª° ë³„ì¹­", "ì‡¼í•‘ëª°ë³„ì¹­", 
    "ê³„ì •ëª…", "ê³„ì •", "ë³„ì¹­", "ì‚¬ì—…ì", "ì´ë¦„"
]

# ì„¸ì…˜ ì„¤ì •
SESSION_EXPIRE_HOURS = 8
SECRET_KEY = os.environ.get("SECRET_KEY", secrets.token_hex(32))

# ì„œë²„ ì„¤ì •
HOST = os.environ.get("HOST", "0.0.0.0")
PORT = int(os.environ.get("PORT", 8000))
# PORT = 8001

# ========== ìŠ¤ì¼€ì¤„ëŸ¬ ì„¤ì • ==========
scheduler = AsyncIOScheduler(
    jobstores={'default': MemoryJobStore()},
    timezone='Asia/Seoul'
)

# ìŠ¤ì¼€ì¤„ ì €ì¥ (íŒŒì¼ ê¸°ë°˜)
SCHEDULES_FILE = APP_DIR / "schedules.json"

def load_schedules() -> List[Dict]:
    """ì €ì¥ëœ ìŠ¤ì¼€ì¤„ ë¡œë“œ"""
    if SCHEDULES_FILE.exists():
        try:
            with open(SCHEDULES_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            pass
    return []

def save_schedules(schedules: List[Dict]):
    """ìŠ¤ì¼€ì¤„ ì €ì¥"""
    with open(SCHEDULES_FILE, 'w', encoding='utf-8') as f:
        json.dump(schedules, f, ensure_ascii=False, indent=2)

# ìŠ¤ì¼€ì¤„ ì‘ì—… ì‹¤í–‰ í•¨ìˆ˜
async def execute_scheduled_task(schedule_id: str, platform: str, task: str, stores: List[str], options: Dict):
    """ìŠ¤ì¼€ì¤„ëœ ì‘ì—… ì‹¤í–‰"""
    print(f"[ìŠ¤ì¼€ì¤„ëŸ¬] ì‘ì—… ì‹œì‘: {schedule_id} - {platform}/{task}")

    # ì‘ì—… ë¡œê·¸ ê¸°ë¡
    store_count = len(stores) if stores else 0
    log_work(f"ìŠ¤ì¼€ì¤„-{task}", platform, store_count, f"ìŠ¤ì¼€ì¤„: {schedule_id}", "ì˜ˆì•½")
    
    try:
        if platform == "ìŠ¤ë§ˆíŠ¸ìŠ¤í† ì–´":
            # smartstore_allinone.py subprocess ì‹¤í–‰ (í™˜ê²½ë³€ìˆ˜ë¡œ ì‘ì—…/ìŠ¤í† ì–´ ì „ë‹¬)
            env = os.environ.copy()
            env["SERVICE_ACCOUNT_JSON"] = os.environ.get("SERVICE_ACCOUNT_JSON", "")
            env["SPREADSHEET_KEY"] = os.environ.get("SPREADSHEET_KEY", "")
            env["DELETE_SOURCE_SPREADSHEET_KEY"] = SALES_SHEET_ID
            env["PARALLEL_STORES"] = "true"
            env["PARALLEL_WORKERS"] = "4"
            env["PYTHONIOENCODING"] = "utf-8"
            env["AIO_TASK"] = task  # ì‘ì—…ëª… ì „ë‹¬ (ëŒ€ìƒ ìŠ¤í† ì–´ëŠ” ì‹œíŠ¸ì˜ 'í™œì„±í™”' ì»¬ëŸ¼ ì°¸ì¡°)
            
            module_path = os.path.join(os.path.dirname(__file__), "modules", "smartstore_allinone.py")
            log_file = os.path.join(os.path.dirname(__file__), "logs", f"schedule_{schedule_id}.log")
            os.makedirs(os.path.dirname(log_file), exist_ok=True)
            
            print(f"[ìŠ¤ì¼€ì¤„ëŸ¬] smartstore_allinone.py ì‹¤í–‰: {task}")
            
            process = subprocess.Popen(
                [sys.executable, module_path],
                stdout=open(log_file, "w", encoding="utf-8"),
                stderr=subprocess.STDOUT,
                env=env,
                cwd=os.path.dirname(__file__)
            )
            
            # í”„ë¡œì„¸ìŠ¤ ì™„ë£Œ ëŒ€ê¸° (ìµœëŒ€ 30ë¶„)
            try:
                process.wait(timeout=1800)
                print(f"[ìŠ¤ì¼€ì¤„ëŸ¬] í”„ë¡œì„¸ìŠ¤ ì™„ë£Œ: exit code {process.returncode}")
            except subprocess.TimeoutExpired:
                process.kill()
                print(f"[ìŠ¤ì¼€ì¤„ëŸ¬] í”„ë¡œì„¸ìŠ¤ íƒ€ì„ì•„ì›ƒ (30ë¶„)")
        
        elif platform == "11ë²ˆê°€":
            # 11ë²ˆê°€ ì‘ì—…
            print(f"[ìŠ¤ì¼€ì¤„ëŸ¬] 11ë²ˆê°€ {task} - ì•„ì§ ë¯¸êµ¬í˜„")
        
        # ì‹¤í–‰ ê²°ê³¼ ê¸°ë¡
        schedules = load_schedules()
        for s in schedules:
            if s.get('id') == schedule_id:
                s['last_run'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                s['run_count'] = s.get('run_count', 0) + 1
                break
        save_schedules(schedules)
        
        print(f"[ìŠ¤ì¼€ì¤„ëŸ¬] ì‘ì—… ì™„ë£Œ: {schedule_id}")
    except Exception as e:
        print(f"[ìŠ¤ì¼€ì¤„ëŸ¬] ì‘ì—… ì˜¤ë¥˜: {schedule_id} - {e}")
        import traceback
        traceback.print_exc()

# ë¶ˆì‚¬ì ì‹œìŠ¤í…œ ì„¤ì •
BULSAJA_SYSTEMS = {
    1: {
        "name": "ë°˜ëŒ€ëŸ‰í”„ë¦¬ë¯¸ì—„ 1",
        "folder": "C:\\ìë™í™”ì‹œìŠ¤í…œ",
        "json": "civic-kayak-410304-03d60ceb535f.json",
        "sheet_key": "19glTugSCcouvFQWALuguAO0jeJt1gdnhI-zbDhMds_Y"
    },
    2: {
        "name": "ë°˜ëŒ€ëŸ‰í”„ë¦¬ë¯¸ì—„ 2",
        "folder": "C:\\ìë™í™”ì‹œìŠ¤í…œ2",
        "json": "ornate-chemist-466108-p8-0ac0d011d2cd.json",
        "sheet_key": "1lMY0g2P2TKFTI23-zqGO48oCLPCTGqRK2KJcYfD4pf8"
    }
}

BULSAJA_TAB_NAME = "njb_ìƒí’ˆê´€ë¦¬"
BULSAJA_SHEET_KEY = "19glTugSCcouvFQWALuguAO0jeJt1gdnhI-zbDhMds_Y"
BULSAJA_GROUP_SELECTOR_CELL = "C15"  # ê¸°ë³¸ê°’, ëª¨ë“œì— ë”°ë¼ ë³€ê²½
BULSAJA_GROUP_TOKEN_FMT = "{n}ë²ˆ ë§ˆì¼“ê·¸ë£¹"

# í”Œë«í¼ ì„¤ì •
PLATFORM_CONFIG = {
    "ìŠ¤ë§ˆíŠ¸ìŠ¤í† ì–´": {
        "login_url": "https://accounts.commerce.naver.com/login?url=https%3A%2F%2Fsell.smartstore.naver.com%2F%23%2Flogin-callback",
        "color": "#03C75A"
    },
    "ì¿ íŒ¡": {
        "login_url": "https://xauth.coupang.com/auth/realms/seller/protocol/openid-connect/auth?response_type=code&client_id=wing&redirect_uri=https%3A%2F%2Fwing.coupang.com%2Fsso%2Flogin?returnUrl%3D%252F&state=78ad277c-bf25-4992-8f48-c523b37ce667&login=true&ui_locales=ko-KR&scope=openid",
        "color": "#E31837"
    },
    "11ë²ˆê°€": {
        "login_url": "https://login.11st.co.kr/auth/front/selleroffice/login.tmall",
        "color": "#FF5A00"
    },
    "ESMí†µí•©": {
        "login_url": "https://signin.esmplus.com/login",
        "color": "#6C5CE7",
        "tab_selector": "button[data-montelena-acode='700000273']"  # ESM PLUS íƒ­
    },
    "ì§€ë§ˆì¼“": {
        "login_url": "https://signin.esmplus.com/login",
        "color": "#00C73C",
        "tab_selector": "button[data-montelena-acode='700000274']"  # ì§€ë§ˆì¼“ íƒ­
    },
    "ì˜¥ì…˜": {
        "login_url": "https://signin.esmplus.com/login",
        "color": "#FF0000",
        "tab_selector": "button[data-montelena-acode='700000275']"  # ì˜¥ì…˜ íƒ­
    }
}

# ========== ë°ì´í„° ëª¨ë¸ ==========
class LoginRequest(BaseModel):
    username: str
    password: str

class AccountModel(BaseModel):
    platform: str
    login_id: str
    password: str = ""
    shop_alias: str = ""  # í•˜ìœ„ í˜¸í™˜ì„± ìœ ì§€
    store_name: str = ""  # í†µì¼ëœ í•„ë“œëª…
    business_number: str = ""
    # ìŠ¤ë§ˆíŠ¸ìŠ¤í† ì–´
    ss_seller_id: str = ""
    ss_app_id: str = ""
    ss_app_secret: str = ""
    # ì¿ íŒ¡
    cp_vendor_code: str = ""
    cp_access_key: str = ""
    cp_secret_key: str = ""
    # 11ë²ˆê°€
    st_api_key: str = ""
    api_key: str = ""  # 11ë²ˆê°€ API KEY (ë³„ì¹­)
    # ESM ì—°ê²°
    esm_master: str = ""
    esm_master_pw: str = ""
    esm_id: str = ""  # ESM ID
    esm_pw: str = ""  # ESM PW
    # ê¸°íƒ€
    owner: str = ""
    usage: str = ""

class SMSRequest(BaseModel):
    phone_profile: str
    to_number: str
    message: str

@dataclass
class SMSMessage:
    phone_profile: str
    sender: str
    content: str
    timestamp: str
    auth_code: Optional[str] = None
    unread: bool = False  # ì•ˆì½ìŒ ìƒíƒœ

# ========== ê¶Œí•œ ë ˆë²¨ ==========
# êµ¬ê¸€ ì‹œíŠ¸ "ë‹´ë‹¹ì" íƒ­ì˜ "ê¶Œí•œ" ì—´ì— ì…ë ¥
ROLE_ADMIN = "admin"      # ëª¨ë“  ê¶Œí•œ
ROLE_OPERATOR = "oper"    # ìš´ì˜ì (ì‚­ì œ ë¶ˆê°€)
ROLE_VIEWER = "viewer"    # ë·°ì–´

# ê¶Œí•œë³„ í—ˆìš© ê¸°ëŠ¥
ROLE_PERMISSIONS = {
    ROLE_ADMIN: ["view", "edit", "delete", "sms_send", "sms_control", "bulsaja"],
    ROLE_OPERATOR: ["view", "edit", "sms_send", "sms_control", "bulsaja"],
    ROLE_VIEWER: ["view", "sms_view"],
}

def get_role_permissions(role: str) -> list:
    """ê¶Œí•œ ë ˆë²¨ì— ë”°ë¥¸ í—ˆìš© ê¸°ëŠ¥ ëª©ë¡ ë°˜í™˜"""
    return ROLE_PERMISSIONS.get(role, ROLE_PERMISSIONS[ROLE_VIEWER])

def has_permission(role: str, permission: str) -> bool:
    """íŠ¹ì • ê¶Œí•œì´ ìˆëŠ”ì§€ í™•ì¸"""
    return permission in get_role_permissions(role)

# ========== ì„¸ì…˜ ê´€ë¦¬ ==========
sessions: Dict[str, Dict] = {}  # token -> {username, name, role, expires}

def create_session(username: str, name: str = None, role: str = ROLE_VIEWER) -> str:
    token = secrets.token_urlsafe(32)
    sessions[token] = {
        "username": username,
        "name": name or username,
        "role": role,
        "expires": datetime.now() + timedelta(hours=SESSION_EXPIRE_HOURS)
    }
    return token

def verify_session(token: str) -> Optional[Dict]:
    if not token or token not in sessions:
        return None
    session = sessions[token]
    if datetime.now() > session["expires"]:
        del sessions[token]
        return None
    return {
        "username": session["username"],
        "name": session["name"],
        "role": session.get("role", ROLE_VIEWER)
    }

# API Key (í¬ë¡¬ í™•ì¥ìš©)
API_KEY = os.environ.get("API_KEY", "pkonomiautokey2024")

def get_current_user(request: Request) -> Dict:
    # ë‚´ë¶€ ìš”ì²­(127.0.0.1)ì€ ì‹œìŠ¤í…œ ê³„ì •ìœ¼ë¡œ bypass
    client_host = request.client.host if request.client else ""
    if client_host in ("127.0.0.1", "localhost", "::1"):
        return {"username": "system", "name": "ì‹œìŠ¤í…œ", "role": ROLE_ADMIN}
    
    # API Key ì¸ì¦ (í¬ë¡¬ í™•ì¥ìš©)
    api_key = request.headers.get("X-API-Key")
    if api_key and api_key == API_KEY:
        return {"username": "extension", "name": "í¬ë¡¬í™•ì¥", "role": ROLE_ADMIN}
    
    token = request.cookies.get("session_token")
    user = verify_session(token)
    if not user:
        raise HTTPException(status_code=401, detail="ë¡œê·¸ì¸ì´ í•„ìš”í•©ë‹ˆë‹¤")
    return user

def require_permission(request: Request, permission: str) -> Dict:
    """íŠ¹ì • ê¶Œí•œ í•„ìš”í•œ APIì—ì„œ ì‚¬ìš©"""
    user = get_current_user(request)
    if not has_permission(user.get("role", ROLE_VIEWER), permission):
        raise HTTPException(status_code=403, detail="ê¶Œí•œì´ ì—†ìŠµë‹ˆë‹¤")
    return user

# ========== êµ¬ê¸€ ì‹œíŠ¸ ê´€ë¦¬ ==========
class GoogleSheetManager:
    def __init__(self):
        self.client = None
        self.sheet = None
        self.connected = False
        self._cache = {}  # (sheet_key, worksheet_name) -> records
        self._cache_time = {}  # (sheet_key, worksheet_name) -> timestamp
        self._worksheet_names_cache = {} # sheet_key -> list of names
        self._worksheet_names_time = {} # sheet_key -> timestamp
        self._cache_lock = threading.Lock()
        self.CACHE_TTL = 12 * 3600  # 12ì‹œê°„ ìºì‹œ
        self.LIST_CACHE_TTL = 3600   # ì›Œí¬ì‹œíŠ¸ ëª©ë¡ì€ 1ì‹œê°„ ìºì‹œ
        self.MARKETING_CACHE_FILE = os.path.join(APP_DIR, "marketing_cache.json")
    
    def connect(self):
        try:
            print(f"ğŸ“‚ ì¸ì¦ íŒŒì¼ ê²½ë¡œ: {CREDENTIALS_FILE}")
            print(f"ğŸ“‚ ì¸ì¦ íŒŒì¼ ì¡´ì¬: {os.path.exists(CREDENTIALS_FILE)}")
            scopes = [
                'https://www.googleapis.com/auth/spreadsheets',
                'https://www.googleapis.com/auth/drive'
            ]
            creds = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=scopes)
            self.client = gspread.authorize(creds)
            
            # ìŠ¤í”„ë ˆë“œì‹œíŠ¸ í‚¤ë¡œ ì—´ê¸° (ìš°ì„ ) ë˜ëŠ” ì´ë¦„ìœ¼ë¡œ ì—´ê¸°
            if SPREADSHEET_KEY:
                self.sheet = self.client.open_by_key(SPREADSHEET_KEY)
            else:
                self.sheet = self.client.open("ê³„ì •ê´€ë¦¬")
            
            self.connected = True
            print(f"âœ… êµ¬ê¸€ ì‹œíŠ¸ ì—°ê²°ë¨: {self.sheet.title}")
            return True
        except Exception as e:
            print(f"âŒ êµ¬ê¸€ ì‹œíŠ¸ ì—°ê²° ì‹¤íŒ¨: {e}")
            self.connected = False
            return False

    def get_market_status(self, force_refresh=False) -> Dict[str, Dict]:
        """ë§ˆì¼“ìƒíƒœí˜„í™© íƒ­ì—ì„œ ìŠ¤í† ì–´ë³„ íŒë§¤ì•¡ ë° ìƒíƒœ ë“± ì¡°íšŒ"""
        if not self.connected:
            if not self.connect():
                return {}
        
        # ìºì‹œ í™•ì¸ (5ë¶„)
        cache_key = (SPREADSHEET_KEY, "ë§ˆì¼“ìƒíƒœí˜„í™©")
        now = time.time()
        if not force_refresh and cache_key in self._cache:
            if now - self._cache_time.get(cache_key, 0) < 300: # 5ë¶„ ìºì‹œ
                return self._cache[cache_key]
        
        try:
            ws = self.sheet.worksheet("ë§ˆì¼“ìƒíƒœí˜„í™©")
            all_values = ws.get_all_values()
            
            if not all_values: return {}
            
            headers = all_values[0]
            # Headers Check: ['ìŠ¤í† ì–´ëª…', ..., 'íŒë§¤ì•¡']
            
            # ì¸ë±ìŠ¤ ì°¾ê¸°
            try:
                idx_name = -1
                idx_revenue = -1
                for i, h in enumerate(headers):
                    h_clean = h.strip()
                    if h_clean in STORE_NAME_ALIASES: idx_name = i
                    elif h_clean == "íŒë§¤ì•¡": idx_revenue = i
                
                if idx_name == -1 or idx_revenue == -1:
                    print(f"âŒ ë§ˆì¼“ìƒíƒœí˜„í™© í—¤ë” ë§¤ì¹­ ì‹¤íŒ¨: {headers} (ìŠ¤í† ì–´ëª… ë˜ëŠ” íŒë§¤ì•¡ ì—†ìŒ)")
                    return {}
            except Exception as e:
                print(f"âŒ ë§ˆì¼“ìƒíƒœí˜„í™© í—¤ë” íŒŒì‹± ì˜¤ë¥˜: {e}")
                return {}

            status_map = {}
            for row in all_values[1:]:
                if len(row) <= max(idx_name, idx_revenue): continue
                
                store_name = row[idx_name].strip()
                if not store_name: continue
                
                # íŒë§¤ì•¡ íŒŒì‹± (ì‰¼í‘œ, ì› ë“± ì œê±°)
                raw_revenue = row[idx_revenue].strip()
                try:
                    revenue = int(str(raw_revenue).replace(",", "").replace("ì›", ""))
                except:
                    revenue = 0
                
                # ìŠ¤í† ì–´ëª… ì •ê·œí™” (ê´„í˜¸ ë‚´ìš© ì œê±°, ê³µë°± ì œê±°)
                # ì˜ˆ: "ë£¨ì„¼ì½”(ìŠ¤ë§ˆíŠ¸ìŠ¤í† ì–´)" -> "ë£¨ì„¼ì½”"
                import re
                normalized_name = re.sub(r'\(.*?\)', '', store_name).strip()
                
                status_map[store_name] = {
                    "revenue": revenue,
                    "raw_row": row
                }
                # ì •ê·œí™”ëœ ì´ë¦„ìœ¼ë¡œë„ ë§¤í•‘ ì¶”ê°€ (ë‹¨, ì›ë³¸ ì´ë¦„ì´ ìš°ì„ )
                if normalized_name and normalized_name != store_name:
                    if normalized_name not in status_map:
                         status_map[normalized_name] = {
                            "revenue": revenue,
                            "raw_row": row
                        }
            
            # ìºì‹œ ì—…ë°ì´íŠ¸
            with self._cache_lock:
                self._cache[cache_key] = status_map
                self._cache_time[cache_key] = now
            
            print(f"âœ… ë§ˆì¼“ìƒíƒœí˜„í™© ë¡œë“œ ì™„ë£Œ: {len(status_map)}ê°œ ìŠ¤í† ì–´")
            return status_map
            
        except Exception as e:
            print(f"âŒ ë§ˆì¼“ìƒíƒœí˜„í™© ì¡°íšŒ ì‹¤íŒ¨: {e}")
            return {}

    async def sync_manual_marketing_data(self):
        """ê° ìŠ¤í† ì–´ ì‹œíŠ¸ì—ì„œ T2(ë°©ë¬¸ììˆ˜) ê°’ì„ ì§ì ‘ ì½ì–´ì™€ì„œ ì €ì¥ (ìˆ˜ë™ ì·¨í•©)"""
        if not self.connected:
            if not self.connect():
                return {"success": False, "message": "Google Sheets ì—°ê²° ì‹¤íŒ¨"}

        print("ğŸ”„ [Manual Sync] Starting T2 cell collection...")
        
        # 1. ê³„ì • ëª©ë¡ ê°€ì ¸ì˜¤ê¸° (ì‹œíŠ¸ URL í™•ì¸ìš©)
        accounts = self.get_accounts()
        if not accounts:
            return {"success": False, "message": "ë“±ë¡ëœ ê³„ì •ì´ ì—†ìŠµë‹ˆë‹¤."}
            
        today_str = datetime.now().strftime("%Y-%m-%d")
        updated_count = 0
        
        # ì´ë ¥ íŒŒì¼ ë¡œë“œ
        stats_file = os.path.join(APP_DIR, "marketing_stats.json")
        history_data = {}
        if os.path.exists(stats_file):
            try:
                with open(stats_file, 'r', encoding='utf-8') as f:
                    history_data = json.load(f)
            except: pass

        results = []

        for acc in accounts:
            store_name = acc.get('name')
            sheet_url = acc.get('sheet_url') # ê³„ì • ì‹œíŠ¸ URL
            if not store_name or not sheet_url:
                continue
                
            try:
                # ì‹œíŠ¸ ì—´ê¸°
                try:
                    doc = self.client.open_by_url(sheet_url)
                    # "ì¢…í•©" íƒ­ ë˜ëŠ” ì²« ë²ˆì§¸ íƒ­ ê°€ì •. 
                    # T2 ì…€ì´ ìˆëŠ” ì‹œíŠ¸ëª…ì„ ì •í™•íˆ ì•Œì•„ì•¼ í•¨. ë³´í†µ ì²«ë²ˆì§¸ ì‹œíŠ¸(dashboard)ì¼ ê°€ëŠ¥ì„± ë†’ìŒ.
                    # ì‚¬ìš©ì ìš”ì²­: "ë°ì´í„° ìˆ˜ë™ì·¨í•© ëˆ„ë¥´ë©´ ê° ê³„ì •ì— T2ê°’ì„ ê°€ì ¸ì™€ì„œ"
                    # ê¸°ë³¸ì ìœ¼ë¡œ ì²« ë²ˆì§¸ ì‹œíŠ¸ë¥¼ ëŒ€ìƒìœ¼ë¡œ í•¨.
                    ws = doc.get_worksheet(0) 
                except Exception as e:
                    print(f"âš ï¸ {store_name} ì‹œíŠ¸ ì—´ê¸° ì‹¤íŒ¨: {e}")
                    results.append(f"{store_name}: ì‹œíŠ¸ ì ‘ì† ì‹¤íŒ¨")
                    continue

                # T2 ì…€ ê°’ ì½ê¸° (ë°©ë¬¸ììˆ˜)
                t2_val = ws.acell('T2').value
                
                # ìˆ«ì íŒŒì‹±
                visitors = 0
                if t2_val:
                    try:
                        visitors = int(str(t2_val).replace(',', '').replace('ëª…', '').strip())
                    except:
                        visitors = 0
                
                print(f"âœ… {store_name}: T2 Visit Count = {visitors}")
                
                # ë°ì´í„° ì €ì¥
                if store_name not in history_data:
                    history_data[store_name] = {}
                
                history_data[store_name][today_str] = {
                    "visitors": visitors,
                    "updated_at": datetime.now().strftime("%H:%M:%S"),
                    "source": "manual_T2"
                }

                # T2 ê°’ ì—…ë°ì´íŠ¸ ê²°ê³¼ë¥¼ ë¦¬ìŠ¤íŠ¸ì— ì €ì¥
                results.append(f"{store_name}: {visitors}ëª…")
                updated_count += 1
                
            except Exception as e:
                print(f"âŒ {store_name} ì²˜ë¦¬ ì¤‘ ì˜¤ë¥˜: {e}")
                results.append(f"{store_name}: ì˜¤ë¥˜ ({str(e)[:20]})")

        # íŒŒì¼ ì €ì¥
        try:
            with open(stats_file, 'w', encoding='utf-8') as f:
                json.dump(history_data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"âŒ ì´ë ¥ íŒŒì¼ ì €ì¥ ì‹¤íŒ¨: {e}")
            
        return {
            "success": True, 
            "message": f"{updated_count}ê°œ ìŠ¤í† ì–´ ì—…ë°ì´íŠ¸ ì™„ë£Œ",
            "details": results
        }

    def get_marketing_data(self, force_refresh=False) -> Dict[str, Dict]:
        """ë§ˆì¼€íŒ… ì‹œíŠ¸ 'ì „ì²´ë°ì´í„°' íƒ­ì—ì„œ ìŠ¤í† ì–´ë³„ ë§¤ì¶œ/ìœ ì…/ì£¼ë¬¸ ì§‘ê³„ ë° ì´ë ¥ ê´€ë¦¬"""
        if not self.connected:
            if not self.connect():
                return {}

        # ìºì‹œ í™•ì¸ (5ë¶„)
        cache_key = (MARKETING_SPREADSHEET_KEY, "ì „ì²´ë°ì´í„°")
        now = time.time()
        
        if not force_refresh:
            # 1. ë©”ëª¨ë¦¬ ìºì‹œ í™•ì¸
            if cache_key in self._cache:
                if now - self._cache_time.get(cache_key, 0) < 300:
                    print("[CACHE] Using memory marketing data")
                    return self._cache[cache_key]
            
            # 2. íŒŒì¼ ìºì‹œ í™•ì¸ (ë©”ëª¨ë¦¬ì— ì—†ìœ¼ë©´)
            if os.path.exists(self.MARKETING_CACHE_FILE):
                try:
                    with open(self.MARKETING_CACHE_FILE, 'r', encoding='utf-8') as f:
                        print("[CACHE] Using persistent marketing data")
                        data = json.load(f)
                        
                        # [Smart Validation] ìºì‹œ ë°ì´í„°ê°€ ìœ íš¨í•œì§€ ê²€ì‚¬ (ë§¤ì¶œ ì •ë³´ê°€ ì „ë¶€ 0ì´ë©´ ë¬´íš¨)
                        has_valid_revenue = any(d.get("revenue", 0) > 0 for d in data.values()) if data else False
                        if not has_valid_revenue and data:
                            print("[CACHE] Cached data seems invalid (All revenue 0). Forcing refresh.")
                            # ìºì‹œ ë¬´ì‹œí•˜ê³  ì§„í–‰
                        else:
                            with self._cache_lock:
                                self._cache[cache_key] = data
                                self._cache_time[cache_key] = now
                            return data
                except: pass

        try:
            # 1. ì‹œíŠ¸ ë°ì´í„° ê°€ì ¸ì˜¤ê¸° (ì „ì²´ë°ì´í„°)
            print(f"[Marketing] Fetching data from {MARKETING_SPREADSHEET_KEY}...")
             # ì‹œíŠ¸ í‚¤ê°€ ë‹¤ë¥¼ ìˆ˜ ìˆìœ¼ë¯€ë¡œ ëª…ì‹œì ìœ¼ë¡œ ì—´ê¸°
            try:
                sheet = self.client.open_by_key(MARKETING_SPREADSHEET_KEY)
                ws = sheet.worksheet("ì „ì²´ë°ì´í„°")
            except Exception as e:
                print(f"âŒ ë§ˆì¼€íŒ… ì‹œíŠ¸ ì—´ê¸° ì‹¤íŒ¨: {e}")
                return {}

            all_values = ws.get_all_values()
            if not all_values: return {}
            
            headers = all_values[0]
            # Headers: ['ìˆ˜ì§‘ì¼ì‹œ', 'ìŠ¤í† ì–´ëª…', ..., 'ë°©ë¬¸íšŸìˆ˜', 'ìœ ì…ìˆ˜', 'í´ë¦­ìˆ˜', 'ì „í™˜ìˆ˜', 'íŒë§¤ì•¡']
            
            try:
                idx_store = -1
                
                # ìŠ¤í† ì–´ëª… ì»¬ëŸ¼ ì°¾ê¸° (STORE_NAME_ALIASES í™œìš©)
                for i, h in enumerate(headers):
                    if h.strip() in STORE_NAME_ALIASES:
                        idx_store = i
                        break
                        
                # íŒë§¤ì•¡, ìœ ì…ìˆ˜ ë“± ì»¬ëŸ¼ ì°¾ê¸° (ê³µë°± ë“± ì²˜ë¦¬ë¥¼ ìœ„í•´ ë£¨í”„ ì‚¬ìš© ê¶Œì¥ë˜ë‚˜ index ì‹œë„)
                idx_revenue = -1
                idx_visitors = -1
                idx_orders = -1

                for i, h in enumerate(headers):
                    h_clean = h.strip()
                    if h_clean in ["íŒë§¤ì•¡", "ë§¤ì¶œ", "ë§¤ì¶œì•¡", "ì´ê²°ì œê¸ˆì•¡", "ê²°ì œê¸ˆì•¡"]: idx_revenue = i
                    elif h_clean in ["ìœ ì…ìˆ˜", "ë°©ë¬¸íšŸìˆ˜", "ë°©ë¬¸ì", "ë°©ë¬¸ììˆ˜"]: idx_visitors = i
                    elif h_clean in ["ì „í™˜ìˆ˜", "ê²°ì œê±´ìˆ˜", "ì£¼ë¬¸ìˆ˜", "íŒë§¤ê±´ìˆ˜"]: idx_orders = i
                
                if idx_store == -1 or idx_revenue == -1:
                     print(f"âŒ ë§ˆì¼€íŒ… ë°ì´í„° í•„ìˆ˜ í—¤ë” ëˆ„ë½: {headers} (ìŠ¤í† ì–´ëª… í™•ì¸ í•„ìš”)")
                     return {}
                
            except ValueError:
                print("âŒ ë§ˆì¼€íŒ… ë°ì´í„° í—¤ë” ë§¤ì¹­ ì‹¤íŒ¨")
                return {}

            # 2. ë°ì´í„° ì§‘ê³„ (ìŠ¤í† ì–´ë³„)
            # {store_name: {revenue: 0, visitors: 0, orders: 0}}
            current_stats = {}
            
            for row in all_values[1:]:
                if len(row) <= idx_store: continue
                store_name = row[idx_store].strip()
                if not store_name: continue
                
                # ê°’ íŒŒì‹±
                def parse_int(val):
                    try: return int(str(val).replace(",", "").replace("ì›", "").strip())
                    except: return 0
                
                rev_raw = row[idx_revenue] if idx_revenue != -1 and idx_revenue < len(row) else "0"
                rev = parse_int(rev_raw)
                
                vis = parse_int(row[idx_visitors]) if idx_visitors != -1 and idx_visitors < len(row) else 0
                ord_ = parse_int(row[idx_orders]) if idx_orders != -1 and idx_orders < len(row) else 0

                if len(current_stats) < 5: # Debug first 5 rows
                    print(f"[DEBUG_REV] Store: {store_name}, Raw Revenue: '{rev_raw}', Parsed: {rev}")
                
                if store_name not in current_stats:
                    current_stats[store_name] = {"revenue": 0, "visitors": 0, "orders": 0}
                
                current_stats[store_name]["revenue"] += rev
                current_stats[store_name]["visitors"] += vis
                current_stats[store_name]["orders"] += ord_

            # 3. ì´ë ¥ ê´€ë¦¬ (marketing_stats.json) - ì¼ë³„ ë°ì´í„° ê¸°ë¡
            stats_file = os.path.join(APP_DIR, "marketing_stats.json")
            history_data = {}
            if os.path.exists(stats_file):
                try:
                    with open(stats_file, 'r', encoding='utf-8') as f:
                        history_data = json.load(f)
                except: pass

            today_str = datetime.now().strftime("%Y-%m-%d")
            
            for store, data in current_stats.items():
                if store not in history_data:
                    history_data[store] = {}
                
                # ì˜¤ëŠ˜ ë°ì´í„° ì—…ë°ì´íŠ¸ (ë®ì–´ì“°ê¸°)
                history_data[store][today_str] = {
                    "visitors": data["visitors"],
                    "orders": data["orders"],
                    "revenue": data["revenue"]
                }
                
                # 10ì¼ì¹˜ ìœ ì§€ (ì˜¤ë˜ëœ í‚¤ ì‚­ì œ)
                dates = sorted(history_data[store].keys())
                if len(dates) > 10:
                    for d in dates[:-10]:
                        del history_data[store][d]
            
            # íŒŒì¼ ì €ì¥
            try:
                with open(stats_file, 'w', encoding='utf-8') as f:
                    json.dump(history_data, f, ensure_ascii=False, indent=2)
            except Exception as e:
                print(f"âŒ ë§ˆì¼€íŒ… ì´ë ¥ ì €ì¥ ì‹¤íŒ¨: {e}")

            # 4. ê²°ê³¼ ë°˜í™˜ (í˜„ì¬ ì§‘ê³„ + ì´ë ¥ ë¶„ì„ ê²°ê³¼)
            final_data = {}
            for store, data in current_stats.items():
                # ë¦¬ë‰´ì–¼ ê¸°ì¤€ ë¶„ì„
                # 1) 5ì¼ ì—°ì† ë°©ë¬¸ì ê°ì†Œ
                # 2) 7ì¼ê°„ ì£¼ë¬¸ 0ê±´
                
                is_declining = False
                zero_orders_7days = False
                
                hist = history_data.get(store, {})
                sorted_dates = sorted(hist.keys(), reverse=True) # ìµœì‹ ìˆœ [ì˜¤ëŠ˜, ì–´ì œ, ...]
                
                # 5ì¼ ì—°ì† ê°ì†Œ ì²´í¬ (ë°ì´í„°ê°€ 5ì¼ ì´ìƒ ìˆì–´ì•¼ í•¨)
                if len(sorted_dates) >= 5:
                    v = [hist[d]["visitors"] for d in sorted_dates[:5]]
                    # v[0] (ì˜¤ëŠ˜) < v[1] (ì–´ì œ) < v[2] < v[3] < v[4]
                    if v[4] > v[3] > v[2] > v[1] > v[0]:
                        is_declining = True
                
                # 7ì¼ê°„ ì£¼ë¬¸ 0ê±´ ì²´í¬
                check_days = sorted_dates[:7]
                if len(check_days) > 0:
                     recent_orders = sum(hist[d]["orders"] for d in check_days)
                     if recent_orders == 0 and len(check_days) >= 1: # ë°ì´í„°ê°€ ìˆëŠ” í•œ 0ê±´ì¸ì§€ ì²´í¬
                        zero_orders_7days = True

                final_data[store] = {
                    "revenue": data["revenue"],
                    "visitors": data["visitors"],
                    "orders": data["orders"],
                    "is_declining": is_declining,
                    "zero_orders": zero_orders_7days,
                    "history": hist
                }
                
                # ì •ê·œí™”ëœ ì´ë¦„ ë§¤í•‘ ì¶”ê°€
                import re
                norm_name = re.sub(r'\(.*?\)', '', store).strip()
                if norm_name and norm_name != store:
                    if norm_name not in final_data:
                        final_data[norm_name] = final_data[store]

            # ìºì‹œ ì—…ë°ì´íŠ¸
            with self._cache_lock:
                self._cache[cache_key] = final_data
                self._cache_time[cache_key] = now
                # íŒŒì¼Persistent ìºì‹œ ì €ì¥
                try:
                    with open(self.MARKETING_CACHE_FILE, 'w', encoding='utf-8') as f:
                        json.dump(final_data, f, ensure_ascii=False, indent=2)
                except: pass
            
            print(f"âœ… ë§ˆì¼€íŒ… ë°ì´í„° ë¡œë“œ ì™„ë£Œ: {len(final_data)}ê°œ ìŠ¤í† ì–´")
            return final_data

        except Exception as e:
            print(f"âŒ ë§ˆì¼€íŒ… ë°ì´í„° ì¡°íšŒ ì‹¤íŒ¨: {e}")
            import traceback
            traceback.print_exc()
            return {}
    
    def get_worksheet(self, name: str):
        if not self.connected:
            return None
        try:
            return self.sheet.worksheet(name)
        except:
            return None

    def get_worksheet_names_with_cache(self, sheet_key: str = None, force_refresh: bool = False) -> List[str]:
        """ì›Œí¬ì‹œíŠ¸ ëª©ë¡ ê°€ì ¸ì˜¤ê¸° (ìºì‹œ ì ìš©)"""
        s_key = sheet_key or SPREADSHEET_KEY
        with self._cache_lock:
            now = time.time()
            if not force_refresh and s_key in self._worksheet_names_cache:
                if (now - self._worksheet_names_time.get(s_key, 0)) < self.LIST_CACHE_TTL:
                    return self._worksheet_names_cache[s_key]

        try:
            if not self.connected: self.connect()
            target_sheet = self.client.open_by_key(s_key) if sheet_key else self.sheet
            worksheets = target_sheet.worksheets()
            names = [ws.title for ws in worksheets]
            
            with self._cache_lock:
                self._worksheet_names_cache[s_key] = names
                self._worksheet_names_time[s_key] = now
            return names
        except Exception as e:
            print(f"âŒ ì›Œí¬ì‹œíŠ¸ ëª©ë¡ ë¡œë“œ ì‹¤íŒ¨ ({s_key}): {e}")
            return self._worksheet_names_cache.get(s_key, [])

    def get_external_worksheet(self, key: str, name: str):
        """íŠ¹ì • í‚¤ë¥¼ ê°€ì§„ ì™¸ë¶€ ìŠ¤í”„ë ˆë“œì‹œíŠ¸ì˜ ì›Œí¬ì‹œíŠ¸ ê°€ì ¸ì˜¤ê¸°"""
        if not self.connected:
            return None
        try:
            external_sheet = self.client.open_by_key(key)
            return external_sheet.worksheet(name)
        except Exception as e:
            print(f"âŒ ì™¸ë¶€ ì‹œíŠ¸ ì›Œí¬ì‹œíŠ¸ ë¡œë“œ ì‹¤íŒ¨ ({key}, {name}): {e}")
            return None

    def open_worksheet_with_creds(self, creds_path: str, sheet_key: str, ws_name: str):
        """íŠ¹ì • ì¸ì¦ íŒŒì¼ë¡œ íŠ¹ì • ì‹œíŠ¸ì˜ ì›Œí¬ì‹œíŠ¸ ì—´ê¸°"""
        try:
            if not os.path.exists(creds_path):
                print(f"âŒ ì¸ì¦ íŒŒì¼ ì—†ìŒ: {creds_path}")
                return None
            scopes = ['https://www.googleapis.com/auth/spreadsheets', 'https://www.googleapis.com/auth/drive']
            creds = Credentials.from_service_account_file(creds_path, scopes=scopes)
            client = gspread.authorize(creds)
            sheet = client.open_by_key(sheet_key)
            return sheet.worksheet(ws_name)
        except Exception as e:
            print(f"âŒ ì™¸ë¶€ ì¸ì¦ ì‹œíŠ¸ ë¡œë“œ ì‹¤íŒ¨ ({ws_name}): {e}")
            return None
    
    def get_records_with_cache(self, ws_name: str, sheet_key: str = None, force_refresh: bool = False) -> List[Dict]:
        """ìºì‹œë¥¼ ì‚¬ìš©í•˜ì—¬ ì›Œí¬ì‹œíŠ¸ ë ˆì½”ë“œ ê°€ì ¸ì˜¤ê¸°"""
        s_key = sheet_key or SPREADSHEET_KEY
        cache_key = (s_key, ws_name)
        
        with self._cache_lock:
            now = time.time()
            if not force_refresh and cache_key in self._cache and (now - self._cache_time.get(cache_key, 0)) < self.CACHE_TTL:
                return self._cache[cache_key]
        
        # ìºì‹œ ì—†ê±°ë‚˜ ë§Œë£Œë¨ ë˜ëŠ” ê°•ì œ ìƒˆë¡œê³ ì¹¨ -> API í˜¸ì¶œ
        try:
            if sheet_key and sheet_key != SPREADSHEET_KEY:
                ws = self.get_external_worksheet(sheet_key, ws_name)
            else:
                ws = self.get_worksheet(ws_name)
            
            if not ws:
                return []
                
            print(f"[CACHE] Fetching fresh records: {ws_name} (force={force_refresh})")
            # get_all_records ëŒ€ì‹  get_all_valuesë¥¼ ì‚¬ìš©í•˜ì—¬ ìˆ˜ë™ íŒŒì‹± (í—¤ë” ì¤‘ë³µì´ë‚˜ ê³µë°± ë¬¸ì œ ë°©ì§€)
            all_values = ws.get_all_values()
            if not all_values:
                return []
                
            headers = [str(h).strip() for h in all_values[0]]
            records = []
            for row in all_values[1:]:
                # í–‰ì˜ ê¸¸ì´ë¥¼ í—¤ë”ì™€ ë§ì¶¤
                row_data = row + [""] * (len(headers) - len(row))
                record = {headers[i]: row_data[i] for i in range(len(headers)) if headers[i]}
                records.append(record)
                
            print(f"[CACHE] Loaded {len(records)} records from {ws_name}")
            
            with self._cache_lock:
                self._cache[cache_key] = records
                self._cache_time[cache_key] = time.time()
            return records
        except Exception as e:
            print(f"âŒ ì‹œíŠ¸ ë°ì´í„° ì½ê¸° ì‹¤íŒ¨ ({ws_name}): {e}")
            with self._cache_lock:
                return self._cache.get(cache_key, [])

    def get_values_with_cache(self, ws_name: str, sheet_key: str = None, force_refresh: bool = False) -> List[List]:
        """ìºì‹œë¥¼ ì‚¬ìš©í•˜ì—¬ ì›Œí¬ì‹œíŠ¸ ì „ì²´ ê°’(get_all_values) ê°€ì ¸ì˜¤ê¸°"""
        s_key = sheet_key or SPREADSHEET_KEY
        cache_key = (s_key, ws_name, "values")
        
        with self._cache_lock:
            now = time.time()
            if not force_refresh and cache_key in self._cache and (now - self._cache_time.get(cache_key, 0)) < self.CACHE_TTL:
                return self._cache[cache_key]
        
        try:
            if sheet_key and sheet_key != SPREADSHEET_KEY:
                ws = self.get_external_worksheet(sheet_key, ws_name)
            else:
                ws = self.get_worksheet(ws_name)
                
            if not ws:
                return []
                
            print(f"[CACHE] Fetching fresh values: {ws_name} (force={force_refresh})")
            values = ws.get_all_values()
            with self._cache_lock:
                self._cache[cache_key] = values
                self._cache_time[cache_key] = time.time()
            return values
        except Exception as e:
            print(f"âŒ ì‹œíŠ¸ ê°’ ì½ê¸° ì‹¤íŒ¨ ({ws_name}): {e}")
            with self._cache_lock:
                return self._cache.get(cache_key, [])
    
    # ë‹´ë‹¹ì ì¸ì¦
    def verify_user(self, username: str, password: str, force_refresh: bool = False) -> Dict:
        try:
            records = self.get_records_with_cache(USERS_TAB, force_refresh=force_refresh)
            if not records:
                # ì‹œíŠ¸ ì—†ê±°ë‚˜ ë¹„ì–´ìˆìœ¼ë©´ ê¸°ë³¸ ê³„ì • í—ˆìš© (admin/admin)
                if username == "admin" and password == "admin":
                    return {"success": True, "name": "ê´€ë¦¬ì", "role": ROLE_ADMIN}
                return {"success": False, "name": "", "role": ""}
            
            pw_hash = hashlib.sha256(password.encode()).hexdigest()
            
            for row in records:
                # ì•„ì´ë”” ì»¬ëŸ¼ ì°¾ê¸° (ê³µë°± í—ˆìš© ë° ìœ ì—°í•œ ë§¤ì¹­)
                row_user = None
                for k, v in row.items():
                    if k.strip() == "ì•„ì´ë””":
                        row_user = v
                        break
                
                if row_user and str(row_user).strip() == str(username).strip():
                    # íŒ¨ìŠ¤ì›Œë“œ ì»¬ëŸ¼ ì°¾ê¸°
                    stored_pw = ""
                    for k, v in row.items():
                        if k.strip() in ["íŒ¨ìŠ¤ì›Œë“œ", "ë¹„ë°€ë²ˆí˜¸"]:
                            stored_pw = v
                            break
                    
                    staff_name = row.get("ë‹´ë‹¹ìëª…", username)
                    
                    # í‰ë¬¸ ë˜ëŠ” í•´ì‹œ ë¹„êµ (ê³µë°± ì œê±°)
                    s_pw = str(stored_pw).strip()
                    if s_pw == str(password).strip() or s_pw == pw_hash:
                        # ê¶Œí•œ ë§¤í•‘ (í•œê¸€ -> ì˜ì–´)
                        role_val = str(row.get("ê¶Œí•œ", "")).strip()
                        if role_val == "ê´€ë¦¬ì":
                            role = ROLE_ADMIN
                        elif role_val == "ìš´ì˜ì":
                            role = ROLE_OPERATOR
                        elif role_val == "ë·°ì–´":
                            role = ROLE_VIEWER
                        else:
                            role = role_val if role_val in [ROLE_ADMIN, ROLE_OPERATOR, ROLE_VIEWER] else ROLE_VIEWER
                        return {"success": True, "name": staff_name, "role": role}
            
            return {"success": False, "name": "", "role": ""}
        except Exception as e:
            print(f"ì‚¬ìš©ì ì¸ì¦ ì˜¤ë¥˜: {e}")
            return {"success": False, "name": "", "role": ""}
    
    # ê³„ì • ëª©ë¡ ì¡°íšŒ
    def get_accounts(self, platform: str = None, force_refresh: bool = False) -> List[Dict]:
        try:
            records = self.get_records_with_cache(ACCOUNTS_TAB, force_refresh=force_refresh)
            if not records:
                return []
            
            accounts = []
            
            # ì²« ë²ˆì§¸ ë ˆì½”ë“œì˜ ì»¬ëŸ¼ëª… ì¶œë ¥ (ë””ë²„ê·¸)
            if records:
                first_row_keys = list(records[0].keys())
                print(f"[ê³„ì •ëª©ë¡] ì»¬ëŸ¼ëª…: {first_row_keys}")
                # ESM ê´€ë ¨ ì»¬ëŸ¼ í™•ì¸
                esm_keys = [k for k in first_row_keys if 'ESM' in k or 'esm' in k]
                print(f"[ê³„ì •ëª©ë¡] ESM ê´€ë ¨ ì»¬ëŸ¼: {esm_keys}")
            
            for idx, row in enumerate(records):
                # ìŠ¤í† ì–´ëª…: STORE_NAME_ALIASES í™œìš©í•˜ì—¬ ìë™ ë§¤ì¹­
                store_name = ""
                for alias in STORE_NAME_ALIASES:
                    if row.get(alias):
                        store_name = row.get(alias)
                        break
                
                # ê·¸ë˜ë„ ì—†ìœ¼ë©´ ì²« ë²ˆì§¸ í‚¤ê°’ì„ ìŠ¤í† ì–´ëª…ìœ¼ë¡œ ê°„ì£¼í•˜ê±°ë‚˜ ë¹ˆê°’
                if not store_name:
                    # fallback (ê±°ì˜ ë°œìƒ ì•ˆí•¨)
                    pass
                
                # ESM ID/PW: ê³µë°± í¬í•¨ëœ í‚¤ë„ ì‹œë„
                esm_id = ""
                esm_pw = ""
                for key in row.keys():
                    key_stripped = key.strip()
                    if key_stripped == "ESM ID":
                        esm_id = row[key]
                    elif key_stripped == "ESM PW":
                        esm_pw = row[key]
                
                # ì²« 3ê°œ ê³„ì •ë§Œ ESM ì •ë³´ ë¡œê·¸
                if idx < 3 and (row.get("í”Œë«í¼") in ["ì§€ë§ˆì¼“", "ì˜¥ì…˜"]):
                    print(f"[ê³„ì •ëª©ë¡] {row.get('í”Œë«í¼')} {row.get('ì•„ì´ë””')}: esm_id='{esm_id}', esm_pw='{esm_pw}'")
                
                # í•œê¸€ í‚¤ ìš°ì„  + ì˜ì–´ í‚¤ í•˜ìœ„ í˜¸í™˜
                acc = {
                    # ê¸°ë³¸ ì •ë³´ (í•œê¸€ + ì˜ì–´)
                    "í”Œë«í¼": row.get("í”Œë«í¼", ""),
                    "platform": row.get("í”Œë«í¼", ""),  # í•˜ìœ„ í˜¸í™˜
                    
                    "ì•„ì´ë””": row.get("ì•„ì´ë””", ""),
                    "login_id": row.get("ì•„ì´ë””", ""),  # í•˜ìœ„ í˜¸í™˜
                    
                    "íŒ¨ìŠ¤ì›Œë“œ": row.get("íŒ¨ìŠ¤ì›Œë“œ") or row.get("ë¹„ë°€ë²ˆí˜¸", ""),
                    "password": row.get("íŒ¨ìŠ¤ì›Œë“œ") or row.get("ë¹„ë°€ë²ˆí˜¸", ""),  # í•˜ìœ„ í˜¸í™˜
                    
                    "ìŠ¤í† ì–´ëª…": store_name,
                    
                    "ì‚¬ì—…ìë²ˆí˜¸": row.get("ì‚¬ì—…ìë²ˆí˜¸", ""),
                    "business_number": row.get("ì‚¬ì—…ìë²ˆí˜¸", ""),  # í•˜ìœ„ í˜¸í™˜
                    
                    "ìš©ë„": row.get("ìš©ë„", ""),
                    "usage": row.get("ìš©ë„", ""),  # í•˜ìœ„ í˜¸í™˜
                    
                    "ì†Œìœ ì": row.get("ì†Œìœ ì", ""),
                    "owner": row.get("ì†Œìœ ì", ""),  # í•˜ìœ„ í˜¸í™˜
                    
                    # ìŠ¤ë§ˆíŠ¸ìŠ¤í† ì–´
                    "ss_seller_id": row.get("ìŠ¤ë§ˆíŠ¸ìŠ¤í† ì–´ API ì—°ë™ìš© íŒë§¤ìID", ""),
                    "ss_app_id": row.get("ìŠ¤ë§ˆíŠ¸ìŠ¤í† ì–´ ì• í”Œë¦¬ì¼€ì´ì…˜ ID", ""),
                    "ss_app_secret": row.get("ìŠ¤ë§ˆíŠ¸ìŠ¤í† ì–´ ì• í”Œë¦¬ì¼€ì´ì…˜ ì‹œí¬ë¦¿", ""),
                    
                    # ì¿ íŒ¡
                    "cp_vendor_code": row.get("ì¿ íŒ¡ ì—…ì²´ì½”ë“œ", ""),
                    "cp_access_key": row.get("ì¿ íŒ¡ Access Key", ""),
                    "cp_secret_key": row.get("ì¿ íŒ¡ Secret Key", ""),
                    
                    # 11ë²ˆê°€
                    "st_api_key": row.get("11ë²ˆê°€ API KEY", ""),
                    "api_key": row.get("11ë²ˆê°€ API KEY", ""),  # ë³„ì¹­
                    
                    # ESM ì—°ê²°
                    "esm_master": row.get("ESMí†µí•©ê³„ì •", ""),
                    "esm_master_pw": row.get("ESMí†µí•©ë¹„ë°€ë²ˆí˜¸", ""),
                    "esm_id": esm_id,
                    "esm_pw": esm_pw
                }
                if acc["ì•„ì´ë””"]:
                    if platform is None or acc["í”Œë«í¼"] == platform:
                        accounts.append(acc)
            return accounts
        except Exception as e:
            print(f"ê³„ì • ì¡°íšŒ ì˜¤ë¥˜: {e}")
            import traceback
            traceback.print_exc()
            return []
    
    # ê³„ì • ì¶”ê°€
    def add_account(self, account: Dict) -> bool:
        ws = self.get_worksheet(ACCOUNTS_TAB)
        if not ws:
            return False
        
        try:
            row = [
                account.get("platform", ""),
                account.get("login_id", ""),
                account.get("password", ""),
                account.get("ìŠ¤í† ì–´ëª…", ""),
                account.get("business_number", ""),
                account.get("ss_seller_id", ""),
                account.get("ss_app_id", ""),
                account.get("ss_app_secret", ""),
                account.get("cp_vendor_code", ""),
                account.get("cp_access_key", ""),
                account.get("cp_secret_key", ""),
                account.get("st_api_key", ""),
                account.get("esm_master", ""),
                account.get("esm_master_pw", "")
            ]
            ws.append_row(row)
            return True
        except Exception as e:
            print(f"ê³„ì • ì¶”ê°€ ì˜¤ë¥˜: {e}")
            return False
    
    # ê³„ì • ìˆ˜ì •
    def update_account(self, old_id: str, platform: str, account: Dict) -> bool:
        ws = self.get_worksheet(ACCOUNTS_TAB)
        if not ws:
            return False
        
        try:
            # í—¤ë” ê°€ì ¸ì˜¤ê¸°
            headers = ws.row_values(1)
            header_map = {h.strip(): i for i, h in enumerate(headers)}
            
            records = ws.get_all_records()
            
            # í•„ë“œ-í—¤ë” ë§¤í•‘
            field_to_header = {
                "platform": "í”Œë«í¼",
                "login_id": "ì•„ì´ë””",
                "password": "íŒ¨ìŠ¤ì›Œë“œ",
                "ìŠ¤í† ì–´ëª…": "ì‡¼í•‘ëª° ë³„ì¹­",
                "business_number": "ì‚¬ì—…ìë²ˆí˜¸",
                "ss_seller_id": "ìŠ¤ë§ˆíŠ¸ìŠ¤í† ì–´ API ì—°ë™ìš© íŒë§¤ìID",
                "ss_app_id": "ìŠ¤ë§ˆíŠ¸ìŠ¤í† ì–´ ì• í”Œë¦¬ì¼€ì´ì…˜ ID",
                "ss_app_secret": "ìŠ¤ë§ˆíŠ¸ìŠ¤í† ì–´ ì• í”Œë¦¬ì¼€ì´ì…˜ ì‹œí¬ë¦¿",
                "cp_vendor_code": "ì¿ íŒ¡ ì—…ì²´ì½”ë“œ",
                "cp_access_key": "ì¿ íŒ¡ Access Key",
                "cp_secret_key": "ì¿ íŒ¡ Secret Key",
                "st_api_key": "11ë²ˆê°€ API KEY",
                "api_key": "11ë²ˆê°€ API KEY",
                "esm_master": "ESMí†µí•©ê³„ì •",
                "esm_master_pw": "ESMí†µí•©ë¹„ë°€ë²ˆí˜¸",
                "esm_id": "ESM ID",
                "esm_pw": "ESM PW",
                "owner": "ì†Œìœ ì",
                "usage": "ìš©ë„"
            }
            
            # ì§€ë§ˆì¼“/ì˜¥ì…˜ ESM ë™ê¸°í™” í•„ìš” ì—¬ë¶€
            esm_sync = platform in ["ì§€ë§ˆì¼“", "ì˜¥ì…˜"] and ("esm_id" in account or "esm_pw" in account)
            other_platform = "ì˜¥ì…˜" if platform == "ì§€ë§ˆì¼“" else "ì§€ë§ˆì¼“"
            
            all_updates = []
            found = False
            
            for i, row in enumerate(records):
                row_id = row.get("ì•„ì´ë””")
                row_platform = row.get("í”Œë«í¼")
                row_num = i + 2
                
                # í˜„ì¬ ê³„ì • ì—…ë°ì´íŠ¸
                if row_id == old_id and row_platform == platform:
                    found = True
                    for field, value in account.items():
                        header = field_to_header.get(field)
                        if header and header in header_map:
                            col_idx = header_map[header]
                            col_letter = chr(65 + col_idx) if col_idx < 26 else f"{chr(64 + col_idx // 26)}{chr(65 + col_idx % 26)}"
                            all_updates.append({"range": f"{col_letter}{row_num}", "values": [[value if value else ""]]})
                
                # ì§€ë§ˆì¼“â†”ì˜¥ì…˜ ESM ë™ê¸°í™”
                elif esm_sync and row_id == old_id and row_platform == other_platform:
                    for field in ["esm_id", "esm_pw"]:
                        if field in account:
                            header = field_to_header.get(field)
                            if header and header in header_map:
                                col_idx = header_map[header]
                                col_letter = chr(65 + col_idx) if col_idx < 26 else f"{chr(64 + col_idx // 26)}{chr(65 + col_idx % 26)}"
                                all_updates.append({"range": f"{col_letter}{row_num}", "values": [[account[field] if account[field] else ""]]})
            
            if found and all_updates:
                ws.batch_update(all_updates)
            return found
        except Exception as e:
            print(f"ê³„ì • ìˆ˜ì • ì˜¤ë¥˜: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    # ê³„ì • ì‚­ì œ
    def delete_account(self, account_id: str, platform: str) -> bool:
        ws = self.get_worksheet(ACCOUNTS_TAB)
        if not ws:
            return False
        
        try:
            records = ws.get_all_records()
            for i, row in enumerate(records):
                if row.get("ì•„ì´ë””") == account_id and row.get("í”Œë«í¼") == platform:
                    ws.delete_rows(i + 2)
                    return True
            return False
        except Exception as e:
            print(f"ê³„ì • ì‚­ì œ ì˜¤ë¥˜: {e}")
            return False

# ========== SMS ë¸Œë¼ìš°ì € ê´€ë¦¬ (sms_gui.py ê¸°ë°˜) ==========
# ì°½ ì„¤ì • (3ê°œ ì°½ì´ 1920px ëª¨ë‹ˆí„°ì— ë“¤ì–´ì˜¤ë„ë¡)
WIN_LEFT = 10
WIN_TOP = 50
WIN_WIDTH = 400  # 3ê°œ ì°½ì´ ì—¬ìœ ìˆê²Œ ë“¤ì–´ì˜¤ë„ë¡ ë” ì¶•ì†Œ (500 -> 400)
WIN_HEIGHT = 800
WIN_GAP = 10  # ì°½ ê°„ê²©

class SMSBrowserManager:
    def __init__(self):
        self.playwright = None
        self.browsers: Dict[str, BrowserContext] = {}
        self.pages: Dict[str, Page] = {}
        self.ready: Dict[str, bool] = {p: False for p in PHONE_PROFILES}
        self.messages: List[SMSMessage] = []
        self.auth_codes: Dict[str, dict] = {}  # {phone: {"code": "123456", "time": "12:34:56"}}
        self.image_cache: Dict[str, Dict[str, str]] = {}  # {sender: {element_idx: filepath}} - ë²ˆí˜¸ë³„ ì´ë¯¸ì§€ ìºì‹œ
        self.image_cache_limit = 10  # ë²ˆí˜¸ë‹¹ ìµœëŒ€ ìºì‹œ ìˆ˜
        self.cache_file = APP_DIR / "sms_cache.json"
        self.api_log_file = APP_DIR / "sms_api_log.json"  # API ë¶„ì„ìš© ë¡œê·¸
        self.captured_apis: List[dict] = []  # ìº¡ì²˜ëœ API ìš”ì²­ë“¤
        self.api_capture_enabled = False  # API ìº¡ì²˜ í™œì„±í™” í”Œë˜ê·¸
        self.lock = asyncio.Lock()  # ë™ì‹œ ì‹¤í–‰ ë°©ì§€ë¥¼ ìœ„í•œ ë½ ì¶”ê°€
        self.load_cache()

    def load_cache(self):
        """íŒŒì¼ì—ì„œ SMS ìºì‹œ ë¡œë“œ"""
        if self.cache_file.exists():
            try:
                with open(self.cache_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    self.messages = [SMSMessage(**m) for m in data.get("messages", [])]
                    self.auth_codes = data.get("auth_codes", {})
                print(f"ğŸ“¦ [SMS] ìºì‹œ ë¡œë“œ ì™„ë£Œ: {len(self.messages)}ê±´")
            except Exception as e:
                print(f"âš ï¸ [SMS] ìºì‹œ ë¡œë“œ ì˜¤ë¥˜: {e}")
                self.messages = []
        else:
            self.messages = []

    def save_cache(self):
        """SMS ìºì‹œë¥¼ íŒŒì¼ì— ì €ì¥"""
        try:
            # ìµœê·¼ 1000ê°œ ì •ë„ë§Œ ìœ ì§€ (ì„±ëŠ¥ ê³ ë ¤)
            save_msgs = [asdict(m) for m in self.messages[-1000:]]
            data = {
                "messages": save_msgs,
                "auth_codes": self.auth_codes,
                "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }
            with open(self.cache_file, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"âš ï¸ [SMS] ìºì‹œ ì €ì¥ ì˜¤ë¥˜: {e}")
    
    def _add_to_image_cache(self, sender: str, element_idx: str, filepath: str):
        """ì´ë¯¸ì§€ ìºì‹œì— ì¶”ê°€ (ë²ˆí˜¸ë³„ 10ê°œ ì œí•œ)"""
        if sender not in self.image_cache:
            self.image_cache[sender] = {}
        
        cache = self.image_cache[sender]
        cache[element_idx] = filepath
        
        # ìºì‹œ ì œí•œ ì´ˆê³¼ ì‹œ ì˜¤ë˜ëœ ê²ƒ ì‚­ì œ
        if len(cache) > self.image_cache_limit:
            oldest_key = list(cache.keys())[0]
            del cache[oldest_key]
    
    def _get_from_image_cache(self, sender: str, element_idx: str) -> Optional[str]:
        """ì´ë¯¸ì§€ ìºì‹œì—ì„œ ê°€ì ¸ì˜¤ê¸°"""
        if sender in self.image_cache:
            return self.image_cache[sender].get(element_idx)
        return None
    
    async def init_playwright(self):
        """Playwright ì´ˆê¸°í™” (async)"""
        if self.playwright is None:
            from playwright.async_api import async_playwright
            self.playwright = await async_playwright().start()
    
    async def _clear_emulation_overrides(self, page: Page, context: BrowserContext):
        """Emulation(ë·°í¬íŠ¸) ì˜¤ë²„ë¼ì´ë“œ ì™„ì „ í•´ì œ"""
        try:
            cdp = await context.new_cdp_session(page)
            try:
                await cdp.send("Emulation.clearDeviceMetricsOverride", {})
            except:
                pass
            try:
                await cdp.send("Emulation.setVisibleSize", {"width": 0, "height": 0})
            except:
                pass
        except Exception as e:
            print(f"[warn] emulation clear failed: {e}")
    
    async def _wait_ui_ready(self, page: Page, timeout_ms: int = 15000) -> bool:
        """UI ì¤€ë¹„ ëŒ€ê¸°"""
        sel_any = (
            'a[data-e2e-start-button], '
            'textarea[data-e2e-message-input-box], '
            '[data-e2e-conversation-list]'
        )
        try:
            await page.wait_for_selector(sel_any, timeout=timeout_ms)
            return True
        except:
            return False
    
    async def _stabilize_messages_ui(self, page: Page, context: BrowserContext):
        """ì´ˆê¸° ì§„ì… ì‹œ ë¹ˆ í™”ë©´ì´ë©´ ë‹¨ê³„ì ìœ¼ë¡œ ë³µêµ¬"""
        if await self._wait_ui_ready(page, 8000):
            return
        
        # 1) soft reload
        try:
            await page.evaluate("""
                (()=>{
                  const href = location.href;
                  if (href.includes('/web/conversations/new')) {
                    location.href = 'https://messages.google.com/web';
                  } else {
                    location.reload();
                  }
                })();
            """)
        except:
            pass
        
        if await self._wait_ui_ready(page, 12000):
            return
        
        # 2) hard reload
        try:
            cdp = await context.new_cdp_session(page)
            await cdp.send("Page.reload", {"ignoreCache": True})
        except:
            pass
        
        if await self._wait_ui_ready(page, 12000):
            return
        
        # 3) ë‹¤ì‹œ ì´ë™
        try:
            await page.goto("https://messages.google.com/web", wait_until="domcontentloaded", timeout=60000)
        except:
            pass
        
        await self._wait_ui_ready(page, 15000)
    
    async def _ensure_conversation_visible(self, page: Page):
        """ëŒ€í™” ë¯¸ì„ íƒ ìƒíƒœë©´ ì²« ëŒ€í™” í´ë¦­ í˜¹ì€ 'ì±„íŒ… ì‹œì‘' ëˆ„ë¥´ê¸°"""
        try:
            if await page.locator('textarea[data-e2e-message-input-box]').count():
                return
            item = page.locator('[role="listitem"], mws-conversation-list-item').first
            if await item.count() and await item.is_visible():
                await item.click()
                await page.wait_for_timeout(300)
                return
            start = page.locator('a[data-e2e-start-button], a[aria-label="ì±„íŒ… ì‹œì‘"], a[aria-label="Start chat"]')
            if await start.count() and await start.first.is_visible():
                await start.first.click()
                await page.wait_for_timeout(300)
        except:
            pass
    
    async def launch_browser(self, profile_id: str):
        """ë¸Œë¼ìš°ì € ì‹¤í–‰ (async - sms_gui.py ë°©ì‹)"""
        if self.playwright is None:
            from playwright.async_api import async_playwright
            self.playwright = await async_playwright().start()
        
        # ê¸°ì¡´ ì¸ì¦ëœ í”„ë¡œí•„ ì‚¬ìš© (pw_sessions ë°”ë¡œ ì•„ë˜)
        actual_dir_name = PROFILE_DIR_MAPPING.get(profile_id, profile_id)
        profile_dir = APP_DIR / "pw_sessions" / actual_dir_name
        
        # ìƒˆ í”„ë¡œí•„ì´ë©´ sms_gui í”„ë¡œí•„ì˜ ì‹ë³„ìë¥¼ ë³µì‚¬ (ë‹¤ë¥¸ PCì—ì„œë„ ê°™ì€ ë¸Œë¼ìš°ì €ë¡œ ì¸ì‹ë˜ë„ë¡)
        source_profile = APP_DIR / "auto_sms" / "chrome_profile"
        if not profile_dir.exists() and source_profile.exists():
            profile_dir.mkdir(parents=True, exist_ok=True)
            # Local State íŒŒì¼ ë³µì‚¬ (client_id2, machine_id ë“± ë¸Œë¼ìš°ì € ì‹ë³„ì í¬í•¨)
            source_local_state = source_profile / "Local State"
            if source_local_state.exists():
                import shutil
                shutil.copy2(source_local_state, profile_dir / "Local State")
                print(f"[BROWSER] [{profile_id}] ë¸Œë¼ìš°ì € ì‹ë³„ì ë³µì‚¬ë¨ (from sms_gui)")
        else:
            profile_dir.mkdir(parents=True, exist_ok=True)
        
        try:
            APP_URL = "https://messages.google.com/web"
            idx = PHONE_PROFILES.index(profile_id)
            
            # sms_gui.pyì™€ ì™„ì „íˆ ë™ì¼í•œ ì„¤ì • (ì´ë™ ê°€ëŠ¥í•œ í”„ë¡œí•„)
            args = [
                "--disable-infobars",
                "--no-sandbox", "--disable-setuid-sandbox",
                "--disable-dev-shm-usage",
                "--no-first-run", "--no-default-browser-check",
                "--disable-features=TranslateUI,CalculateNativeWinOcclusion",
                "--wm-window-animations-disabled",
            ]
            args.insert(0, f"--app={APP_URL}")
            args.append(f"--window-position={WIN_LEFT + idx * (WIN_WIDTH + WIN_GAP)},{WIN_TOP}")
            args.append(f"--window-size={WIN_WIDTH},{WIN_HEIGHT}")
            
            context = await self.playwright.chromium.launch_persistent_context(
                user_data_dir=str(profile_dir),
                headless=False,
                viewport=None,
                args=args,
            )
            
            pages = context.pages
            page = pages[0] if pages else await context.new_page()
            
            # íƒ€ì„ì•„ì›ƒ ì„¤ì •
            context.set_default_timeout(45000)
            page.set_default_timeout(45000)
            page.set_default_navigation_timeout(60000)
            
            # sms_gui.pyì™€ ë™ì¼: í˜ì´ì§€ê°€ ë¹„ì–´ìˆê±°ë‚˜ ë‹¤ë¥¸ URLì´ë©´ ì´ë™
            APP_URL = "https://messages.google.com/web"
            try:
                current_url = page.url or ""
                if current_url == "about:blank":
                    await page.goto(APP_URL, wait_until="domcontentloaded", timeout=60000)
                elif "messages.google.com/web" not in current_url:
                    await page.goto(APP_URL, wait_until="domcontentloaded", timeout=60000)
            except Exception:
                pass
            
            # sms_gui.pyì™€ ë™ì¼: UI ë¡œë“œ ëŒ€ê¸°
            try:
                await page.wait_for_selector(
                    'a[data-e2e-start-button], textarea[data-e2e-message-input-box]', 
                    timeout=60000
                )
            except Exception:
                pass
            
            # CDPë¥¼ ì‚¬ìš©í•˜ì—¬ ì°½ í¬ê¸° ê°•ì œ ì„¤ì • (sms_gui.pyì™€ ë™ì¼ - í”„ë¡œí•„ ì €ì¥ê°’ ë¬´ì‹œ)
            try:
                cdp = await context.new_cdp_session(page)
                target = await cdp.send("Browser.getWindowForTarget")
                wid = target.get("windowId")
                if wid:
                    left_pos = WIN_LEFT + idx * (WIN_WIDTH + WIN_GAP)
                    await cdp.send("Browser.setWindowBounds", {
                        "windowId": wid,
                        "bounds": {
                            "left": left_pos,
                            "top": WIN_TOP,
                            "width": WIN_WIDTH,
                            "height": WIN_HEIGHT,
                            "windowState": "normal"
                        }
                    })
            except Exception as e:
                print(f"[WARN] [{profile_id}] ì°½ í¬ê¸° ê°•ì œ ì„¤ì • ì‹¤íŒ¨: {e}")
            
            self.browsers[profile_id] = context
            self.pages[profile_id] = page
            self.ready[profile_id] = True

            # API ìº¡ì²˜ ë¦¬ìŠ¤ë„ˆ ë“±ë¡ (ë¶„ì„ìš©)
            if self.api_capture_enabled:
                await self._setup_api_capture(page, profile_id)

            # sms_gui.pyì™€ ë™ì¼í•˜ê²Œ ì´ˆê¸°í™” (ë©€í‹° ì„¸ì…˜ ì§€ì›ì„ ìœ„í•´ í•„ìˆ˜)
            await self._clear_emulation_overrides(page, context)
            await self._stabilize_messages_ui(page, context)
            await self._ensure_conversation_visible(page)

            print(f"[SUCCESS] [{profile_id}] ë¸Œë¼ìš°ì € ì‹œì‘ë¨")
            
        except Exception as e:
            print(f"[ERROR] [{profile_id}] ë¸Œë¼ìš°ì € ì‹œì‘ ì‹¤íŒ¨: {e}")
            self.ready[profile_id] = False
    
    async def launch_all(self):
        """ëª¨ë“  ë¸Œë¼ìš°ì € ì‹¤í–‰ (async)"""
        for profile_id in PHONE_PROFILES:
            await self.launch_browser(profile_id)

    async def _setup_api_capture(self, page: Page, profile_id: str):
        """ë„¤íŠ¸ì›Œí¬ ìš”ì²­ ìº¡ì²˜ ì„¤ì • (API ë¶„ì„ìš©)"""
        async def on_request(request):
            url = request.url
            # êµ¬ê¸€ ë©”ì‹œì§€ ê´€ë ¨ APIë§Œ ìº¡ì²˜
            if any(x in url for x in ['messages.google.com', 'instantmessaging', 'tachyon', 'mobilemessaging']):
                try:
                    api_entry = {
                        "profile": profile_id,
                        "timestamp": datetime.now().isoformat(),
                        "method": request.method,
                        "url": url,
                        "headers": dict(request.headers),
                        "post_data": request.post_data[:2000] if request.post_data else None
                    }
                    self.captured_apis.append(api_entry)
                    # ìµœëŒ€ 500ê°œ ìœ ì§€
                    if len(self.captured_apis) > 500:
                        self.captured_apis = self.captured_apis[-500:]
                    print(f"  [API] {request.method} {url[:100]}")
                except:
                    pass

        async def on_response(response):
            url = response.url
            if any(x in url for x in ['messages.google.com', 'instantmessaging', 'tachyon', 'mobilemessaging']):
                try:
                    # ê¸°ì¡´ ìº¡ì²˜ì— ì‘ë‹µ ì •ë³´ ì¶”ê°€
                    for entry in reversed(self.captured_apis):
                        if entry.get("url") == url and "response_status" not in entry:
                            entry["response_status"] = response.status
                            try:
                                body = await response.text()
                                entry["response_body"] = body[:5000] if body else None
                            except:
                                pass
                            break
                except:
                    pass

        page.on("request", on_request)
        page.on("response", on_response)
        print(f"[APIë¶„ì„] [{profile_id}] ë„¤íŠ¸ì›Œí¬ ìº¡ì²˜ í™œì„±í™”ë¨")

    def enable_api_capture(self, enable: bool = True):
        """API ìº¡ì²˜ í™œì„±í™”/ë¹„í™œì„±í™”"""
        self.api_capture_enabled = enable
        if enable:
            print("[APIë¶„ì„] API ìº¡ì²˜ í™œì„±í™”ë¨ - ë¸Œë¼ìš°ì € ì¬ì‹œì‘ í•„ìš”")
        else:
            print("[APIë¶„ì„] API ìº¡ì²˜ ë¹„í™œì„±í™”ë¨")

    def get_captured_apis(self) -> List[dict]:
        """ìº¡ì²˜ëœ API ëª©ë¡ ë°˜í™˜"""
        return self.captured_apis

    def save_api_log(self):
        """ìº¡ì²˜ëœ APIë¥¼ íŒŒì¼ë¡œ ì €ì¥"""
        try:
            with open(self.api_log_file, 'w', encoding='utf-8') as f:
                json.dump(self.captured_apis, f, ensure_ascii=False, indent=2)
            print(f"[APIë¶„ì„] {len(self.captured_apis)}ê°œ API ë¡œê·¸ ì €ì¥ë¨: {self.api_log_file}")
        except Exception as e:
            print(f"[APIë¶„ì„] ì €ì¥ ì˜¤ë¥˜: {e}")

    def clear_api_log(self):
        """ìº¡ì²˜ëœ API ë¡œê·¸ ì´ˆê¸°í™”"""
        self.captured_apis = []
        print("[APIë¶„ì„] API ë¡œê·¸ ì´ˆê¸°í™”ë¨")
    
    async def refresh_messages(self) -> List[SMSMessage]:
        """ë©”ì‹œì§€ ìƒˆë¡œê³ ì¹¨ (ë¶€ë¶„ ì—…ë°ì´íŠ¸ ë° ìºì‹œ ê°•í™”)"""
        async with self.lock:  # ë½ ì ìš©í•˜ì—¬ ì¤‘ë³µ ì‹¤í–‰ ë° ê²½í•© ë°©ì§€
            updated_any = False

            # ë”•ì…”ë„ˆë¦¬ ë³µì‚¬ë³¸ìœ¼ë¡œ ìˆœíšŒ (ë™ì‹œ ìˆ˜ì • ë°©ì§€)
            for profile_id, page in list(self.pages.items()):
                if not self.ready.get(profile_id):
                    continue

                try:
                    # í•´ë‹¹ í”„ë¡œí•„ì˜ ìƒˆë¡œìš´ ë©”ì‹œì§€ ëª©ë¡ ìˆ˜ì§‘ (ì„ì‹œ ë¦¬ìŠ¤íŠ¸)
                    new_profile_messages = []
                    # ì•ˆì½ìŒ ìƒíƒœì˜€ë˜ í•­ëª©ë“¤ ì¶”ì  (ìºì‹œ í›„ ë‹¤ì‹œ ì•ˆì½ìŒìœ¼ë¡œ í‘œì‹œìš©)
                    unread_items_to_restore = []

                    # ëŒ€í™” ëª©ë¡ í•­ëª©ë“¤ ê°€ì ¸ì˜¤ê¸°
                    items = await page.locator('mws-conversation-list-item').all()

                    # ìµœëŒ€ 50ê°œê¹Œì§€ ê°€ì ¸ì˜¤ê¸°
                    for idx, item in enumerate(items[:50]):
                        try:
                            # ì•ˆì½ìŒ ìƒíƒœ í™•ì¸ (ì—¬ëŸ¬ ë°©ë²• ì‹œë„)
                            is_unread = await item.evaluate('''el => {
                                // 1. is-unread í´ë˜ìŠ¤ ì²´í¬
                                if (el.classList.contains("is-unread")) return true;
                                // 2. unread í´ë˜ìŠ¤ ì²´í¬
                                if (el.classList.contains("unread")) return true;
                                // 3. aria-labelì— "unread" ë˜ëŠ” "ì½ì§€ ì•ŠìŒ" í¬í•¨ ì²´í¬
                                const ariaLabel = el.getAttribute("aria-label") || "";
                                if (ariaLabel.toLowerCase().includes("unread") || ariaLabel.includes("ì½ì§€ ì•ŠìŒ")) return true;
                                // 4. ë‚´ë¶€ì— unread indicator ìš”ì†Œê°€ ìˆëŠ”ì§€ ì²´í¬
                                if (el.querySelector(".unread-indicator, .unread-count, [data-unread]")) return true;
                                // 5. ë³¼ë“œì²´ ì´ë¦„ (ì•ˆì½ìŒ ë©”ì‹œì§€ëŠ” ë³´í†µ ë³¼ë“œì²´)
                                const nameEl = el.querySelector(".name");
                                if (nameEl) {
                                    const style = window.getComputedStyle(nameEl);
                                    if (parseInt(style.fontWeight) >= 600 || style.fontWeight === "bold") return true;
                                }
                                return false;
                            }''')

                            sender = await item.locator('.name').first.inner_text() if await item.locator('.name').count() else ""
                            content = await item.locator('.snippet, .text-content').first.inner_text() if await item.locator('.snippet, .text-content').count() else ""
                            
                            # íƒ€ì„ìŠ¤íƒ¬í”„ ì¶”ì¶œ ë° ì •ë°€í™”
                            timestamp = ""
                            ts_loc = item.locator('mws-relative-timestamp.snippet-timestamp')
                            if await ts_loc.count():
                                raw_ts = await ts_loc.first.inner_text()
                                full_ts = await ts_loc.first.get_attribute("title") or ""
                                
                                # 1. 'ìˆ˜ìš”ì¼ ìˆ˜' ê°™ì€ ì¤‘ë³µ ì œê±°
                                import re
                                # 'ìˆ˜ìš”ì¼ ìˆ˜', 'ê¸ˆìš”ì¼ ê¸ˆ' ë“± íŒ¨í„´ ì œê±°
                                raw_val = raw_ts or ""
                                clean_ts = re.sub(r'([ê°€-í£]ìš”ì¼)\s+\1', r'\1', raw_val)
                                # 'ê¸ˆìš”ì¼ ê¸ˆ' ë“± ì²« ê¸€ì ì¤‘ë³µ íŒ¨í„´ ì œê±° (ì˜ˆ: ê¸ˆìš”ì¼ ê¸ˆ -> ê¸ˆìš”ì¼)
                                clean_ts = re.sub(r'(([ê°€-í£])ìš”ì¼)\s+\2(?!\w)', r'\1', clean_ts)
                                
                                # 2. ì˜¤ëŠ˜ ë¬¸ìì¸ ê²½ìš° ì‹œê°„ í‘œì‹œ (title í™œìš©)
                                import datetime
                                now = datetime.datetime.now()
                                # Windows í™˜ê²½ì—ì„œëŠ” %-m ëŒ€ì‹  f-string ì‚¬ìš© (íŒŒë”© ì œê±° í˜¸í™˜ì„±)
                                today_str = f"{now.month}. {now.day}." # '1. 16.'
                                alternative_today = f"{now.year}. {now.month}. {now.day}." # '2026. 1. 16.'
                                
                                # titleì— ì˜¤ëŠ˜ ë‚ ì§œë‚˜ 'ì˜¤ëŠ˜', 'Today'ê°€ ìˆìœ¼ë©´ ì‹œê°„ ì¶”ì¶œ
                                has_today = False
                                if full_ts:
                                    if today_str in full_ts or alternative_today in full_ts:
                                        has_today = True
                                    elif any(x in full_ts.lower() for x in ['ì˜¤ëŠ˜', 'today']):
                                        has_today = True
                                
                                if has_today:
                                    # 'ì˜¤í›„ 8:15' ë˜ëŠ” '20:15' í˜•íƒœ ì¶”ì¶œ
                                    time_match = re.search(r'((ì˜¤ì „|ì˜¤í›„)\s+\d{1,2}:\d{2}|\d{1,2}:\d{2})', full_ts)
                                    if time_match:
                                        timestamp = time_match.group(1)
                                
                                if not timestamp:
                                    timestamp = clean_ts.strip()
                            
                            auth_code = self._extract_auth_code(content)
                            
                            if content:
                                msg = SMSMessage(
                                    phone_profile=profile_id,
                                    sender=sender.strip(),
                                    content=content.strip()[:100],
                                    timestamp=timestamp,
                                    auth_code=auth_code,
                                    unread=is_unread  # ì•ˆì½ìŒ ìƒíƒœ ì €ì¥
                                )
                                new_profile_messages.append(msg)

                                # ì•ˆì½ìŒ ìƒíƒœì˜€ìœ¼ë©´ ë³µì› ëŒ€ìƒì— ì¶”ê°€
                                if is_unread:
                                    unread_items_to_restore.append(item)

                                # ìµœì‹  ì¸ì¦ì½”ë“œ ì—…ë°ì´íŠ¸ (í™”ë©´ì˜ ê°€ì¥ ì²« ë²ˆì§¸ í•­ëª©ì¼ ë•Œ)
                                if auth_code and idx == 0:
                                    from datetime import datetime
                                    msg_ts = self._parse_relative_time(timestamp)
                                    self.auth_codes[profile_id] = {
                                        "code": auth_code,
                                        "time": timestamp or datetime.now().strftime("%H:%M:%S"),
                                        "timestamp": msg_ts
                                    }
                        except:
                            continue

                    # í•´ë‹¹ í”„ë¡œí•„ì—ì„œ ë©”ì‹œì§€ ìˆ˜ì§‘ì— ì„±ê³µí–ˆì„ ê²½ìš°ì—ë§Œ ëª©ë¡ ê°±ì‹ 
                    if new_profile_messages:
                        # ê¸°ì¡´ ë¦¬ìŠ¤íŠ¸ì—ì„œ í•´ë‹¹ í”„ë¡œí•„ ë©”ì‹œì§€ë§Œ ì œê±° í›„ í•©ì¹˜ê¸°
                        self.messages = [m for m in self.messages if m.phone_profile != profile_id]
                        self.messages.extend(new_profile_messages)
                        updated_any = True

                    # ì•ˆì½ìŒ ìƒíƒœì˜€ë˜ ë©”ì‹œì§€ë“¤ ë‹¤ì‹œ ì•ˆì½ìŒìœ¼ë¡œ í‘œì‹œ (ìš°í´ë¦­ ë©”ë‰´ ì‚¬ìš©)
                    for unread_item in unread_items_to_restore:
                        try:
                            await self._mark_as_unread(page, unread_item)
                        except Exception as e:
                            print(f"[{profile_id}] ì•ˆì½ìŒ í‘œì‹œ ì‹¤íŒ¨: {e}")

                except Exception as e:
                    print(f"[{profile_id}] ë©”ì‹œì§€ ì½ê¸° ì˜¤ë¥˜: {e}")
            
            # í•˜ë‚˜ë¼ë„ ê°±ì‹ ë˜ì—ˆë‹¤ë©´ ìºì‹œ ì €ì¥
            if updated_any:
                self.save_cache()
            
            return self.messages

    def _parse_relative_time(self, time_str: str) -> float:
        """ìƒëŒ€ ì‹œê°„ ë¬¸ìì—´ì„ Unix timestampë¡œ ë³€í™˜
        ì˜ˆ: "14ë¶„", "2ì‹œê°„", "AM 10:42", "ì˜¤í›„ 3:23", "ì–´ì œ", "ì›”ìš”ì¼", "ì¡°ê¸ˆ ì „", "ë°©ê¸ˆ"
        """
        import time
        from datetime import datetime, timedelta
        
        now = datetime.now()
        time_str = time_str.strip()
        
        try:
            # "ì¡°ê¸ˆ ì „", "ë°©ê¸ˆ", "ì§€ê¸ˆ" ë“± - ê°€ì¥ ìµœì‹ 
            if any(x in time_str for x in ['ì¡°ê¸ˆ', 'ë°©ê¸ˆ', 'ì§€ê¸ˆ', 'now', 'just']):
                return now.timestamp()
            
            # "Në¶„" ë˜ëŠ” "Në¶„ ì „"
            if 'ë¶„' in time_str:
                match = re.search(r'(\d+)', time_str)
                if match:
                    minutes = int(match.group(1))
                    return (now - timedelta(minutes=minutes)).timestamp()
            
            # "Nì‹œê°„" ë˜ëŠ” "Nì‹œê°„ ì „"
            if 'ì‹œê°„' in time_str:
                match = re.search(r'(\d+)', time_str)
                if match:
                    hours = int(match.group(1))
                    return (now - timedelta(hours=hours)).timestamp()
            
            # "AM/PM HH:MM" ë˜ëŠ” "ì˜¤ì „/ì˜¤í›„ HH:MM"
            am_pm_match = re.search(r'(AM|PM|ì˜¤ì „|ì˜¤í›„)\s*(\d{1,2}):(\d{2})', time_str, re.IGNORECASE)
            if am_pm_match:
                period, hour, minute = am_pm_match.groups()
                hour = int(hour)
                minute = int(minute)
                if period.upper() in ['PM', 'ì˜¤í›„'] and hour != 12:
                    hour += 12
                elif period.upper() in ['AM', 'ì˜¤ì „'] and hour == 12:
                    hour = 0
                msg_time = now.replace(hour=hour, minute=minute, second=0, microsecond=0)
                if msg_time > now:  # ë¯¸ë˜ë©´ ì–´ì œ
                    msg_time -= timedelta(days=1)
                return msg_time.timestamp()
            
            # "ì–´ì œ"
            if 'ì–´ì œ' in time_str:
                return (now - timedelta(days=1)).timestamp()
            
            # ìš”ì¼ ("ì›”ìš”ì¼", "í™”ìš”ì¼" ë“±)
            weekdays = {'ì›”': 0, 'í™”': 1, 'ìˆ˜': 2, 'ëª©': 3, 'ê¸ˆ': 4, 'í† ': 5, 'ì¼': 6}
            for day_name, day_num in weekdays.items():
                if day_name in time_str:
                    days_ago = (now.weekday() - day_num) % 7
                    if days_ago == 0:
                        days_ago = 7  # ê°™ì€ ìš”ì¼ì´ë©´ ì§€ë‚œì£¼
                    return (now - timedelta(days=days_ago)).timestamp()
            
        except:
            pass

        # íŒŒì‹± ì‹¤íŒ¨ ì‹œ í˜„ì¬ ì‹œê°„
        return time.time()

    async def _mark_as_unread(self, page, item):
        """ëŒ€í™” í•­ëª©ì„ ì½ì§€ ì•ŠìŒìœ¼ë¡œ í‘œì‹œ (êµ¬ê¸€ ë©”ì‹œì§€ ìš°í´ë¦­ ë©”ë‰´ ì‚¬ìš©)"""
        try:
            # 1. í•´ë‹¹ í•­ëª© ìš°í´ë¦­
            await item.click(button="right")
            await page.wait_for_timeout(300)

            # 2. ì»¨í…ìŠ¤íŠ¸ ë©”ë‰´ì—ì„œ "ì½ì§€ ì•ŠìŒìœ¼ë¡œ í‘œì‹œ" ë²„íŠ¼ ì°¾ê¸° (ì—¬ëŸ¬ ì„ íƒì ì‹œë„)
            selectors = [
                'button.mat-menu-item:has-text("ì½ì§€ ì•ŠìŒ")',
                'button.mat-menu-item:has-text("Mark as unread")',
                '[role="menuitem"]:has-text("ì½ì§€ ì•ŠìŒ")',
                '[role="menuitem"]:has-text("Mark as unread")',
                '.mat-menu-content button:has-text("ì½ì§€ ì•ŠìŒ")',
                '.mat-menu-content button:has-text("Mark as unread")'
            ]

            clicked = False
            for selector in selectors:
                unread_btn = page.locator(selector)
                if await unread_btn.count() > 0:
                    await unread_btn.first.click()
                    await page.wait_for_timeout(100)
                    print(f"  [SMS] ì•ˆì½ìŒìœ¼ë¡œ í‘œì‹œ ì™„ë£Œ")
                    clicked = True
                    break

            if not clicked:
                # ë©”ë‰´ê°€ ì—´ë ¸ì§€ë§Œ ì˜µì…˜ì´ ì—†ìœ¼ë©´ ESCë¡œ ë‹«ê¸°
                await page.keyboard.press("Escape")
        except Exception as e:
            # ì‹¤íŒ¨ ì‹œ ESCë¡œ ë©”ë‰´ ë‹«ê¸° ì‹œë„
            try:
                await page.keyboard.press("Escape")
            except:
                pass
            raise e

    def _extract_auth_code(self, text: str) -> Optional[str]:
        if not text:
            return None
        
        # ì¸ì¦ ê´€ë ¨ í‚¤ì›Œë“œê°€ ìˆì–´ì•¼ ì¸ì¦ì½”ë“œë¡œ ì¸ì‹
        keywords = ['ì¸ì¦', 'í™•ì¸', 'code', 'verify', 'ë³¸ì¸', 'OTP', 'ì½”ë“œ', 'ìŠ¹ì¸', 'ë²ˆí˜¸ì…ë‹ˆë‹¤', 'ë²ˆí˜¸ëŠ”']
        has_keyword = any(kw.lower() in text.lower() for kw in keywords)
        
        # í‚¤ì›Œë“œ ì—†ìœ¼ë©´ ì¸ì¦ì½”ë“œ ì•„ë‹˜
        if not has_keyword:
            return None
        
        # ì „í™”ë²ˆí˜¸ íŒ¨í„´ ì œì™¸ (010-XXXX-XXXX, 02-XXXX-XXXX ë“±)
        # ë©”ì‹œì§€ì—ì„œ ì „í™”ë²ˆí˜¸ ë¶€ë¶„ ì œê±°
        text_cleaned = re.sub(r'01[0-9]-?\d{3,4}-?\d{4}', '', text)
        text_cleaned = re.sub(r'0\d{1,2}-?\d{3,4}-?\d{4}', '', text_cleaned)
        
        patterns = [
            r'\[(\d{6})\]', r'\((\d{6})\)',  # ëŒ€ê´„í˜¸/ê´„í˜¸ ì•ˆì˜ 6ìë¦¬ ìš°ì„  (ê°€ì¥ í™•ì‹¤í•¨)
            r'(?<!\d)(\d{6})(?!\d)',        # ì—°ì†ëœ 6ìë¦¬
            r'\[(\d{5})\]', r'\((\d{5})\)',
            r'\[(\d{4})\]', r'\((\d{4})\)',
            r'(?<!\d)(\d{5})(?!\d)',
            r'(?<!\d)(\d{4})(?!\d)',
        ]
        
        for pattern in patterns:
            match = re.search(pattern, text_cleaned)
            if match:
                code = match.group(1)
                
                # 4ìë¦¬ ì˜¤íƒ í•„í„° ê°•í™”
                if len(code) == 4:
                    # 1. ì—°ë„ íŒ¨í„´ (20XX, 19XX)
                    if code.startswith('20') or code.startswith('19'):
                        continue
                    # 2. ì‹œê°„ íŒ¨í„´ (HHMM) - ì—„ê²©í•œ í‚¤ì›Œë“œ ì²´í¬
                    if int(code[:2]) <= 24 and int(code[2:]) < 60:
                        strict_keywords = ['ì¸ì¦', 'code', 'OTP', 'ì½”ë“œ', 'í™•ì¸', 'ìŠ¹ì¸']
                        if not any(kw.lower() in text.lower() for kw in strict_keywords):
                            continue
                
                return code
        
        return None
    
    async def send_message(self, profile_id: str, to_number: str, message: str, file_path: str = None) -> bool:
        """ë©”ì‹œì§€ ì „ì†¡ (sms_gui.pyì˜ send_via_messages ê¸°ë°˜) - íŒŒì¼ ì²¨ë¶€ ì§€ì›"""
        if profile_id not in self.pages or not self.ready.get(profile_id):
            print(f"[{profile_id}] ë¸Œë¼ìš°ì € ë¯¸ì‹¤í–‰ ë˜ëŠ” ì¤€ë¹„ ì•ˆë¨")
            return False

        page = self.pages[profile_id]
        context = self.browsers[profile_id]

        try:
            # UI ì•ˆì •í™”
            await self._stabilize_messages_ui(page, context)
            await self._ensure_conversation_visible(page)

            # í™ˆìœ¼ë¡œ(ë’¤ë¡œ)
            try:
                back = page.locator('button[aria-label="ë’¤ë¡œ ê°€ê¸°"], button[aria-label="Back"]')
                if await back.count() and await back.first.is_visible():
                    await back.first.click()
                    await page.wait_for_timeout(300)
            except:
                pass

            # ìƒˆ ëŒ€í™” ì‹œì‘
            start = page.locator('a[data-e2e-start-button], a[aria-label="ì±„íŒ… ì‹œì‘"], a[aria-label="Start chat"]')
            await start.wait_for(state="visible", timeout=15000)
            await start.first.click()
            await page.wait_for_timeout(200)

            # ë²ˆí˜¸ ì…ë ¥
            contact = page.locator('input[data-e2e-contact-input]')
            if not await contact.count():
                contact = page.locator('input[type="text"]')
            await contact.wait_for(state="visible", timeout=15000)
            await contact.fill("")
            await contact.type(to_number, delay=40)
            await page.keyboard.press("Enter")
            await page.wait_for_timeout(600)

            # ì…ë ¥ì°½ í¬ì»¤ìŠ¤/ìŠ¤í¬ë¡¤ ë³´ì • + ì˜¤ë²„ë¼ì´ë“œ í•´ì œ
            try:
                await page.evaluate("""
                  (()=>{
                    const ta=document.querySelector('textarea[data-e2e-message-input-box]')||
                             document.querySelector('textarea[aria-label="ë©”ì‹œì§€"]')||
                             document.querySelector('textarea[aria-label="Message"]')||
                             document.querySelector('textarea');
                    if(ta){ ta.scrollIntoView({block:'center'}); ta.focus(); }
                  })();
                """)
                await self._clear_emulation_overrides(page, context)
            except:
                pass

            # íŒŒì¼ ì²¨ë¶€ê°€ ìˆìœ¼ë©´ ë¨¼ì € ì²˜ë¦¬
            if file_path and os.path.exists(file_path):
                print(f"[{profile_id}] íŒŒì¼ ì²¨ë¶€ ì‹œì‘: {file_path}")
                try:
                    # ì²¨ë¶€ ë²„íŠ¼ í´ë¦­ (+ ë²„íŠ¼ ë˜ëŠ” í´ë¦½ ì•„ì´ì½˜)
                    attach_btn = page.locator('button[data-e2e-attach-button], button[aria-label="ì²¨ë¶€"], button[aria-label="Attach"], mws-attachment-button button')
                    if await attach_btn.count():
                        await attach_btn.first.click()
                        await page.wait_for_timeout(300)

                    # íŒŒì¼ input ìš”ì†Œ ì°¾ê¸° (hidden input)
                    file_input = page.locator('input[type="file"]')
                    if await file_input.count():
                        await file_input.first.set_input_files(file_path)
                        print(f"[{profile_id}] íŒŒì¼ ì²¨ë¶€ë¨: {file_path}")
                        await page.wait_for_timeout(1000)  # íŒŒì¼ ì—…ë¡œë“œ ëŒ€ê¸°
                    else:
                        print(f"[{profile_id}] íŒŒì¼ input ìš”ì†Œë¥¼ ì°¾ì„ ìˆ˜ ì—†ìŒ")
                except Exception as e:
                    print(f"[{profile_id}] íŒŒì¼ ì²¨ë¶€ ì˜¤ë¥˜: {e}")

            # ë©”ì‹œì§€ ì…ë ¥ (í…ìŠ¤íŠ¸ê°€ ìˆìœ¼ë©´)
            if message:
                textarea = page.locator('textarea[data-e2e-message-input-box]')
                if not await textarea.count():
                    textarea = page.locator("textarea").last
                await textarea.wait_for(state="visible", timeout=15000)
                await textarea.fill(message)

            # ì „ì†¡ (Enter ë˜ëŠ” ì „ì†¡ ë²„íŠ¼)
            try:
                # ì „ì†¡ ë²„íŠ¼ì´ ìˆìœ¼ë©´ í´ë¦­
                send_btn = page.locator('button[data-e2e-send-button], button[aria-label="ì „ì†¡"], button[aria-label="Send"]')
                if await send_btn.count() and await send_btn.first.is_visible():
                    await send_btn.first.click()
                else:
                    await page.keyboard.press("Enter")
            except:
                await page.keyboard.press("Enter")
            await page.wait_for_timeout(400)

            print(f"[{profile_id}] ì „ì†¡ ì™„ë£Œ: {to_number}" + (f" (íŒŒì¼: {file_path})" if file_path else ""))
            return True

        except Exception as e:
            print(f"[{profile_id}] ì „ì†¡ ì˜¤ë¥˜: {e}")
            return False
    
    async def close_all(self):
        for context in self.browsers.values():
            try:
                await context.close()
            except:
                pass
        self.browsers.clear()
        self.pages.clear()
        
        if self.playwright:
            await self.playwright.stop()
            self.playwright = None
    
    async def get_conversation_detail(self, profile_id: str, sender: str, offset: int = 0, limit: int = 20) -> Dict[str, Any]:
        """ëŒ€í™” ìƒì„¸ ë‚´ìš© ê°€ì ¸ì˜¤ê¸° (í´ë¦­í•´ì„œ ì—´ê¸°)
        offset: 0ì´ë©´ ìµœê·¼ 20ê°œ, 20ì´ë©´ ì´ì „ 20ê°œ...
        limit: ê°€ì ¸ì˜¬ ë©”ì‹œì§€ ìˆ˜
        """
        if profile_id not in self.pages or not self.ready.get(profile_id):
            return {"error": "ë¸Œë¼ìš°ì € ë¯¸ì‹¤í–‰"}
        
        page = self.pages[profile_id]
        context = self.browsers[profile_id]
        
        # ë‹¤ìš´ë¡œë“œ í´ë” ìƒì„±
        download_dir = APP_DIR / "downloads" / profile_id
        download_dir.mkdir(parents=True, exist_ok=True)
        
        try:
            # ë’¤ë¡œê°€ê¸° (ëŒ€í™” ëª©ë¡ìœ¼ë¡œ)
            try:
                back = page.locator('button[aria-label="ë’¤ë¡œ ê°€ê¸°"], button[aria-label="Back"]')
                if await back.count() and await back.first.is_visible():
                    await back.first.click()
                    await page.wait_for_timeout(300)
            except:
                pass
            
            # 1. ëŒ€í™” ëª©ë¡ì—ì„œ í•´ë‹¹ ë°œì‹ ì ì°¾ì•„ í´ë¦­
            items = await page.locator('mws-conversation-list-item').all()
            found = False
            
            # senderì—ì„œ ìˆ«ìë§Œ ì¶”ì¶œ
            import re
            sender_digits = re.sub(r'[^0-9]', '', sender)
            
            for item in items:
                try:
                    name = await item.locator('.name').first.inner_text() if await item.locator('.name').count() else ""
                    name_digits = re.sub(r'[^0-9]', '', name)
                    
                    # ë§¤ì¹­ ì •ë°€ë„ ê°•í™”
                    is_match = False
                    if sender_digits and name_digits:
                        if sender_digits == name_digits:
                            is_match = True
                        elif len(sender_digits) >= 10 and len(name_digits) >= 10:
                            # í•œêµ­ í° ë²ˆí˜¸ ë’·ìë¦¬ ë§¤ì¹­ (ë 8ìë¦¬ ì´ìƒ ì¼ì¹˜)
                            if sender_digits[-8:] == name_digits[-8:]:
                                is_match = True
                    
                    if is_match:
                        await item.click()
                        await page.wait_for_timeout(800)
                        found = True
                        break
                except:
                    continue
            
            # 2. ëª©ë¡ì—ì„œ ëª» ì°¾ì•˜ì„ ê²½ìš° ê²€ìƒ‰(Fallback) ì‹œë„
            if not found and sender_digits:
                try:
                    # 'ì±„íŒ… ì‹œì‘' ë²„íŠ¼ í´ë¦­
                    start_chat = page.locator('div.floating-button-container button, a[href*="compose"]')
                    if await start_chat.count():
                        await start_chat.first.click()
                        await page.wait_for_timeout(1000)
                        
                        # ë²ˆí˜¸ ì…ë ¥
                        input_box = page.locator('mws-contact-search-input input')
                        if await input_box.count():
                            await input_box.first.fill(sender_digits)
                            await page.keyboard.press("Enter")
                            await page.wait_for_timeout(1500)
                            
                            # ëŒ€í™”ë°© ì§„ì… í™•ì¸ (ë©”ì‹œì§€ ì…ë ¥ì°½ ì¡´ì¬ ì—¬ë¶€)
                            if await page.locator('mws-message-compose textarea').count():
                                found = True
                except Exception as e:
                    print(f"[ëŒ€í™”ìƒì„¸] ê²€ìƒ‰ í´ë°± ì‹œë„ ì¤‘ ì˜¤ë¥˜: {e}")
            
            if not found:
                return {"error": f"ëŒ€í™”ë¥¼ ì°¾ì„ ìˆ˜ ì—†ìŒ: {sender}"}
            
            # offset > 0ì´ë©´ ìŠ¤í¬ë¡¤í•´ì„œ ì´ì „ ë©”ì‹œì§€ ë¡œë“œ
            if offset > 0:
                scroll_count = (offset // 20) + 1
                for _ in range(scroll_count):
                    # ëŒ€í™” ì˜ì—­ ë§¨ ìœ„ë¡œ ìŠ¤í¬ë¡¤
                    conv_container = page.locator('mws-messages-list, .message-list, [role="list"]')
                    if await conv_container.count():
                        await conv_container.first.evaluate('el => el.scrollTop = 0')
                        await page.wait_for_timeout(500)
            
            # ë©”ì‹œì§€ ì½ê¸° (ì¼ë°˜ ë©”ì‹œì§€ + íƒ€ì„ìŠ¤íƒ¬í”„ tombstone í¬í•¨)
            messages = []
            
            # ë©”ì‹œì§€ ì½ê¸° (ì¼ë°˜ ë©”ì‹œì§€ + íƒ€ì„ìŠ¤íƒ¬í”„ tombstone í¬í•¨)
            messages = []
            
            # ëª¨ë“  ë©”ì‹œì§€ ê´€ë ¨ ìš”ì†Œ ê°€ì ¸ì˜¤ê¸° (ë©”ì‹œì§€ + íƒ€ì„ìŠ¤íƒ¬í”„)
            # mws-message-partë¥¼ ì œì™¸í•˜ì—¬ ì¤‘ë³µ ìˆ˜ì§‘ ë°©ì§€ (wrapperë§Œ ìˆ˜ì§‘)
            all_elements = await page.locator('mws-message-wrapper, mws-tombstone-message-wrapper').all()
            
            total_count = len(all_elements)
            
            # offsetê³¼ limit ì ìš©
            if offset == 0:
                # ìµœê·¼ ë©”ì‹œì§€ (ëì—ì„œ limitê°œ)
                start_idx = max(0, total_count - limit)
                end_idx = total_count
            else:
                # ì´ì „ ë©”ì‹œì§€ (offsetë§Œí¼ ì•ì—ì„œ)
                end_idx = max(0, total_count - offset)
                start_idx = max(0, end_idx - limit)
            
            target_elements = all_elements[start_idx:end_idx]
            
            for idx, el in enumerate(target_elements):
                try:
                    # ì›ë˜ ì¸ë±ìŠ¤ ê³„ì‚°
                    original_idx = start_idx + idx
                    
                    # íƒœê·¸ ì´ë¦„ í™•ì¸
                    try:
                        tag_name = await el.evaluate('el => el.tagName.toLowerCase()')
                    except:
                        # ìš”ì†Œê°€ ì‚¬ë¼ì¡Œê±°ë‚˜ ì ‘ê·¼ ë¶ˆê°€ ì‹œ ìŠ¤í‚µ
                        continue
                    
                    if tag_name == 'mws-tombstone-message-wrapper':
                        # íƒ€ì„ìŠ¤íƒ¬í”„ êµ¬ë¶„ì„ 
                        ts_loc = el.locator('mws-relative-timestamp.tombstone-timestamp')
                        if await ts_loc.count():
                            ts_text = await ts_loc.first.inner_text()
                            if ts_text and ts_text.strip():
                                messages.append({
                                    "type": "timestamp_divider",
                                    "timestamp": ts_text.strip(),
                                    "direction": "system"
                                })
                    else:
                        # ì¼ë°˜ ë©”ì‹œì§€
                        msg_data = await self._parse_message_element(el, page, download_dir, original_idx)
                        if msg_data:
                            msg_data["type"] = "message"
                            messages.append(msg_data)
                            
                            # ì¸ì¦ì½”ë“œ ì¶”ì¶œ í™•ì¸
                            if msg_data.get("text"):
                                auth_code = self._extract_auth_code(msg_data["text"])
                                if auth_code:
                                    msg_data["auth_code"] = auth_code
                                    
                except Exception as e:
                    print(f"ìš”ì†Œ íŒŒì‹± ì˜¤ë¥˜(idx={original_idx}): {e}")
                    continue
            
            return {
                "success": True,
                "profile_id": profile_id,
                "sender": sender,
                "messages": messages,
                "total_count": total_count,
                "offset": offset,
                "has_more": start_idx > 0  # ë” ì´ì „ ë©”ì‹œì§€ê°€ ìˆëŠ”ì§€
            }
            
        except Exception as e:
            print(f"[{profile_id}] ëŒ€í™” ìƒì„¸ ì˜¤ë¥˜: {e}")
            return {"error": str(e)}
    
    async def _parse_message_element(self, msg_el, page: Page, download_dir: Path, idx: int) -> Optional[Dict]:
        """ê°œë³„ ë©”ì‹œì§€ ìš”ì†Œ íŒŒì‹±"""
        try:
            # ë°œì‹ /ìˆ˜ì‹  êµ¬ë¶„ - ì—¬ëŸ¬ íŒ¨í„´ ì²´í¬ (wrapper ë ˆë²¨ì—ì„œ ì²´í¬)
            class_attr = await msg_el.get_attribute("class") or ""
            
            # wrapperì˜ í´ë˜ìŠ¤ë¡œ 1ì°¨ íŒë‹¨
            is_outgoing = any(kw in class_attr.lower() for kw in ["outgoing", "from-me", "sent", "self"])
            is_incoming = any(kw in class_attr.lower() for kw in ["incoming", "received"])
            
            # ìì‹ ìš”ì†Œ(mws-message-part)ì˜ ì†ì„±ìœ¼ë¡œ 2ì°¨ íŒë‹¨ (ë” ì •í™•í•¨)
            if not is_outgoing and not is_incoming:
                part_el = msg_el.locator('mws-message-part').first
                if await part_el.count():
                    part_class = await part_el.get_attribute("class") or ""
                    is_outgoing = "is-me" in part_class.lower()
                    is_incoming = "is-not-me" in part_class.lower()

            if is_outgoing:
                direction = "outgoing"
            else:
                # ê¸°ë³¸ì ìœ¼ë¡œ ìˆ˜ì‹ ìœ¼ë¡œ ê°„ì£¼
                direction = "incoming"
            
            # ë””ë²„ê·¸: ì²« 5ê°œ ë©”ì‹œì§€ë§Œ í´ë˜ìŠ¤ ì¶œë ¥ (ë¹„í™œì„±í™”)
            # if idx < 5:
            #     print(f"[ë©”ì‹œì§€ {idx}] class='{class_attr[:100]}' â†’ {direction}")
            
            # í…ìŠ¤íŠ¸ ë‚´ìš©
            text_content = ""
            text_el = msg_el.locator('.text-msg, .message-text, [data-e2e-message-text]')
            if await text_el.count():
                text_content = await text_el.first.inner_text()
                # Debug: í…ìŠ¤íŠ¸ ë‚´ìš© í™•ì¸ (ë¹„í™œì„±í™”)
                # if len(text_content) > 0:
                #      safe_text = text_content[:50].replace('\n', ' ')
                #      print(f"  [ë©”ì‹œì§€ {idx}] í…ìŠ¤íŠ¸ ì¶”ì¶œ: {safe_text}...")
            
            # URL ì¶”ì¶œ (í…ìŠ¤íŠ¸ì—ì„œ)
            urls = re.findall(r'https?://[^\s]+', text_content)
            
            # ì´ë¯¸ì§€ í™•ì¸ (ìŠ¤í¬ë¦°ìƒ· ì—†ì´ ì¡´ì¬ ì—¬ë¶€ë§Œ)
            images = []
            img_elements = await msg_el.locator('img.image-msg, img[data-e2e-message-image], .mms-image img, img[src*="blob:"], img[src*="data:"]').all()
            for img_idx, img_el in enumerate(img_elements):
                try:
                    # ì´ë¯¸ì§€ ì¡´ì¬ë§Œ í™•ì¸, ìŠ¤í¬ë¦°ìƒ·ì€ ë‚˜ì¤‘ì—
                    images.append({
                        "type": "image",
                        "thumbnail": None,  # ë‚˜ì¤‘ì— ë¡œë“œ
                        "element_idx": f"{idx}_{img_idx}"
                    })
                except Exception as e:
                    print(f"ì´ë¯¸ì§€ í™•ì¸ ì˜¤ë¥˜: {e}")
            
            # ë™ì˜ìƒ í™•ì¸
            videos = []
            video_elements = await msg_el.locator('video, .video-msg, [data-e2e-message-video]').all()
            for vid_idx, vid_el in enumerate(video_elements):
                videos.append({
                    "type": "video",
                    "element_idx": f"{idx}_{vid_idx}"
                })
            
            # íŒŒì¼ í™•ì¸
            files = []
            file_elements = await msg_el.locator('.file-msg, .attachment, [data-e2e-message-attachment]').all()
            for file_idx, file_el in enumerate(file_elements):
                try:
                    filename = await file_el.inner_text() if await file_el.count() else f"íŒŒì¼_{file_idx}"
                    files.append({
                        "type": "file",
                        "filename": filename.strip()[:50],
                        "element_idx": f"{idx}_{file_idx}"
                    })
                except:
                    pass
            
            # íƒ€ì„ìŠ¤íƒ¬í”„ (ë©”ì‹œì§€ ë‚´: mws-absolute-timestamp, ë‚ ì§œ: mws-relative-timestamp.tombstone-timestamp)
            timestamp = ""
            # ë¨¼ì € ë©”ì‹œì§€ ë‚´ ì‹œê°„ (mws-absolute-timestamp)
            ts_loc = msg_el.locator('mws-absolute-timestamp')
            if await ts_loc.count():
                ts_text = await ts_loc.first.inner_text()
                if ts_text and ts_text.strip():
                    timestamp = ts_text.strip()
            # ì—†ìœ¼ë©´ ë‚ ì§œ+ì‹œê°„ (tombstone-timestamp)
            if not timestamp:
                ts_loc = msg_el.locator('mws-relative-timestamp.tombstone-timestamp')
                if await ts_loc.count():
                    ts_text = await ts_loc.first.inner_text()
                    if ts_text and ts_text.strip():
                        timestamp = ts_text.strip()
            
            return {
                "direction": direction,
                "text": text_content.strip(),
                "urls": urls,
                "images": images,
                "videos": videos,
                "files": files,
                "timestamp": timestamp
            }
            
        except Exception as e:
            print(f"ë©”ì‹œì§€ íŒŒì‹± ì˜¤ë¥˜: {e}")
            return None
    
    async def download_media(self, profile_id: str, sender: str, media_type: str, element_idx: str, get_thumbnail: bool = False) -> Optional[str]:
        """ì´ë¯¸ì§€/ë™ì˜ìƒ/íŒŒì¼ ë‹¤ìš´ë¡œë“œ
        get_thumbnail=True: ì¸ë„¤ì¼(ì‘ì€ ì´ë¯¸ì§€) ê°€ì ¸ì˜¤ê¸°
        get_thumbnail=False: ì›ë³¸ ì´ë¯¸ì§€ ê°€ì ¸ì˜¤ê¸°
        """
        # ìºì‹œ í™•ì¸ (ì¸ë„¤ì¼ë§Œ)
        if get_thumbnail and sender:
            cache_key = f"thumb_{element_idx}"
            cached = self._get_from_image_cache(sender, cache_key)
            if cached:
                # íŒŒì¼ì´ ì¡´ì¬í•˜ëŠ”ì§€ í™•ì¸
                full_path = APP_DIR / cached.lstrip('/')
                if full_path.exists():
                    print(f"[{profile_id}] ìºì‹œì—ì„œ ì¸ë„¤ì¼ ë°˜í™˜: {cached}")
                    return cached
        
        if profile_id not in self.pages or not self.ready.get(profile_id):
            return None
        
        page = self.pages[profile_id]
        context = self.browsers[profile_id]
        
        download_dir = APP_DIR / "downloads" / profile_id
        download_dir.mkdir(parents=True, exist_ok=True)
        
        try:
            idx, sub_idx = element_idx.split("_")
            idx, sub_idx = int(idx), int(sub_idx)
            
            # ë©”ì‹œì§€ ìš”ì†Œ ì°¾ê¸° (get_conversation_detailê³¼ ë™ì¼í•œ ì…€ë ‰í„°)
            msg_elements = await page.locator('mws-message-wrapper, mws-tombstone-message-wrapper').all()
            if not msg_elements:
                print(f"[{profile_id}] ë©”ì‹œì§€ ìš”ì†Œ ì—†ìŒ - ëŒ€í™”ì°½ì´ ì—´ë ¤ìˆì§€ ì•ŠìŒ")
                return None
            
            # idxê°€ ì „ì²´ ë©”ì‹œì§€ ì¤‘ ì–´ë””ì¸ì§€ í™•ì¸
            if idx >= len(msg_elements):
                print(f"[{profile_id}] idx={idx} ê°€ ë²”ìœ„ ì´ˆê³¼ (total={len(msg_elements)})")
                return None
            
            msg_el = msg_elements[idx]
            
            if media_type == "image":
                img_elements = await msg_el.locator('img.image-msg, img[data-e2e-message-image], .mms-image img, img[src*="blob:"], img[src*="data:"]').all()
                if sub_idx < len(img_elements):
                    img_el = img_elements[sub_idx]
                    
                    if get_thumbnail:
                        # ì¸ë„¤ì¼: ì´ë¯¸ì§€ src ì§ì ‘ ë‹¤ìš´ë¡œë“œ ì‹œë„
                        try:
                            src = await img_el.get_attribute('src')
                            print(f"[{profile_id}] ì¸ë„¤ì¼ src: {src[:80] if src else 'None'}...")
                            
                            if src and src.startswith('http'):
                                # HTTP URLì´ë©´ ì§ì ‘ ë‹¤ìš´ë¡œë“œ
                                import aiohttp
                                async with aiohttp.ClientSession() as session:
                                    async with session.get(src, timeout=aiohttp.ClientTimeout(total=10)) as resp:
                                        if resp.status == 200:
                                            ext = 'jpg' if 'jpeg' in resp.content_type or 'jpg' in resp.content_type else 'png'
                                            filename = f"thumb_{element_idx}_{int(datetime.now().timestamp())}.{ext}"
                                            filepath = download_dir / filename
                                            with open(filepath, 'wb') as f:
                                                f.write(await resp.read())
                                            print(f"[{profile_id}] ì¸ë„¤ì¼ HTTP ë‹¤ìš´ë¡œë“œ ì„±ê³µ")
                                            result_path = f"/downloads/{profile_id}/{filename}"
                                            if sender:
                                                self._add_to_image_cache(sender, f"thumb_{element_idx}", result_path)
                                            return result_path
                            
                            elif src and src.startswith('blob:'):
                                # blob URLì€ JavaScriptë¡œ fetch
                                print(f"[{profile_id}] blob URL ë‹¤ìš´ë¡œë“œ ì‹œë„")
                                blob_data = await page.evaluate('''async (src) => {
                                    try {
                                        const response = await fetch(src);
                                        const blob = await response.blob();
                                        const reader = new FileReader();
                                        return new Promise((resolve) => {
                                            reader.onloadend = () => resolve(reader.result);
                                            reader.readAsDataURL(blob);
                                        });
                                    } catch (e) {
                                        return null;
                                    }
                                }''', src)
                                
                                if blob_data and blob_data.startswith('data:'):
                                    import base64
                                    header, b64_data = blob_data.split(',', 1)
                                    ext = 'jpg' if 'jpeg' in header else 'png'
                                    filename = f"thumb_{element_idx}_{int(datetime.now().timestamp())}.{ext}"
                                    filepath = download_dir / filename
                                    with open(filepath, 'wb') as f:
                                        f.write(base64.b64decode(b64_data))
                                    print(f"[{profile_id}] ì¸ë„¤ì¼ blob ë‹¤ìš´ë¡œë“œ ì„±ê³µ")
                                    result_path = f"/downloads/{profile_id}/{filename}"
                                    if sender:
                                        self._add_to_image_cache(sender, f"thumb_{element_idx}", result_path)
                                    return result_path
                            
                            elif src and src.startswith('data:'):
                                # data URL ì§ì ‘ ë””ì½”ë”©
                                import base64
                                header, b64_data = src.split(',', 1)
                                ext = 'jpg' if 'jpeg' in header else 'png'
                                filename = f"thumb_{element_idx}_{int(datetime.now().timestamp())}.{ext}"
                                filepath = download_dir / filename
                                with open(filepath, 'wb') as f:
                                    f.write(base64.b64decode(b64_data))
                                print(f"[{profile_id}] ì¸ë„¤ì¼ data URL ë””ì½”ë”© ì„±ê³µ")
                                result_path = f"/downloads/{profile_id}/{filename}"
                                if sender:
                                    self._add_to_image_cache(sender, f"thumb_{element_idx}", result_path)
                                return result_path
                                
                        except Exception as e:
                            print(f"[{profile_id}] ì¸ë„¤ì¼ ë‹¤ìš´ë¡œë“œ ì‹¤íŒ¨: {e}")

                        # Fallback: ìŠ¤í¬ë¦°ìƒ·
                        try:
                            # ìš”ì†Œê°€ visibleí•˜ê³  DOMì— ìˆëŠ”ì§€ í™•ì¸
                            is_visible = await img_el.is_visible()
                            if not is_visible:
                                print(f"[{profile_id}] ì¸ë„¤ì¼ ìš”ì†Œê°€ ë³´ì´ì§€ ì•ŠìŒ, ìŠ¤í¬ë¡¤ ì‹œë„")
                                await img_el.scroll_into_view_if_needed(timeout=3000)
                                await page.wait_for_timeout(500)

                            # bounding box í™•ì¸ (DOMì— ìˆëŠ”ì§€)
                            box = await img_el.bounding_box()
                            if not box:
                                print(f"[{profile_id}] ì¸ë„¤ì¼ ìš”ì†Œ bounding box ì—†ìŒ")
                                return None

                            filename = f"thumb_{element_idx}_{int(datetime.now().timestamp())}.png"
                            filepath = download_dir / filename
                            await img_el.screenshot(path=str(filepath), timeout=5000)
                            result_path = f"/downloads/{profile_id}/{filename}"
                            if sender:
                                self._add_to_image_cache(sender, f"thumb_{element_idx}", result_path)
                            return result_path
                        except Exception as e:
                            print(f"[{profile_id}] ì¸ë„¤ì¼ ìŠ¤í¬ë¦°ìƒ· ì‹¤íŒ¨: {str(e)[:100]}")
                            return None
                    else:
                        # ì›ë³¸: ì´ë¯¸ì§€ í´ë¦­í•´ì„œ íŒì—… ì—´ê³  ì›ë³¸ src ê°€ì ¸ì˜¤ê¸°
                        await img_el.click()
                        await page.wait_for_timeout(1500)
                        
                        # ì›ë³¸ ì´ë¯¸ì§€ ì°¾ê¸°
                        full_img_selectors = [
                            '.cdk-overlay-container img[alt="ì „ì²´ í¬ê¸° ì´ë¯¸ì§€"]',
                            '.cdk-overlay-container img[alt="Full size image"]',
                            '.cdk-overlay-container img.ng-star-inserted',
                            '.cdk-overlay-container img',
                            'mat-dialog-container img',
                            '.mdc-dialog__surface img',
                            'div[role="dialog"] img',
                        ]
                        
                        full_img = None
                        for selector in full_img_selectors:
                            locator = page.locator(selector)
                            if await locator.count() > 0:
                                all_imgs = await locator.all()
                                for img in all_imgs:
                                    try:
                                        box = await img.bounding_box()
                                        if box and box['width'] > 100 and box['height'] > 100:
                                            full_img = img
                                            print(f"[{profile_id}] ì›ë³¸ ì´ë¯¸ì§€ ì°¾ìŒ: {selector}")
                                            break
                                    except:
                                        continue
                                if full_img:
                                    break
                        
                        if full_img:
                            # src ì§ì ‘ ë‹¤ìš´ë¡œë“œ ì‹œë„
                            try:
                                src = await full_img.get_attribute('src')
                                print(f"[{profile_id}] ì›ë³¸ ì´ë¯¸ì§€ src: {src[:100] if src else 'None'}...")
                                
                                if src and src.startswith('http'):
                                    import aiohttp
                                    async with aiohttp.ClientSession() as session:
                                        async with session.get(src) as resp:
                                            if resp.status == 200:
                                                content_type = resp.content_type or ''
                                                ext = 'jpg' if 'jpeg' in content_type or 'jpg' in content_type else 'png'
                                                filename = f"full_{element_idx}_{int(datetime.now().timestamp())}.{ext}"
                                                filepath = download_dir / filename
                                                with open(filepath, 'wb') as f:
                                                    f.write(await resp.read())
                                                await page.keyboard.press("Escape")
                                                await page.wait_for_timeout(300)
                                                print(f"[{profile_id}] ì›ë³¸ ë‹¤ìš´ë¡œë“œ ì„±ê³µ: {filename}")
                                                return f"/downloads/{profile_id}/{filename}"
                                
                                elif src and src.startswith('blob:'):
                                    # blob URLì€ JavaScriptë¡œ fetchí•´ì„œ ë‹¤ìš´ë¡œë“œ
                                    print(f"[{profile_id}] blob URL ê°ì§€, JSë¡œ ë‹¤ìš´ë¡œë“œ ì‹œë„")
                                    blob_data = await page.evaluate('''async (src) => {
                                        try {
                                            const response = await fetch(src);
                                            const blob = await response.blob();
                                            const reader = new FileReader();
                                            return new Promise((resolve) => {
                                                reader.onloadend = () => resolve(reader.result);
                                                reader.readAsDataURL(blob);
                                            });
                                        } catch (e) {
                                            return null;
                                        }
                                    }''', src)
                                    
                                    if blob_data and blob_data.startswith('data:'):
                                        import base64
                                        # data:image/jpeg;base64,.... í˜•ì‹
                                        header, b64_data = blob_data.split(',', 1)
                                        ext = 'jpg' if 'jpeg' in header else 'png'
                                        filename = f"full_{element_idx}_{int(datetime.now().timestamp())}.{ext}"
                                        filepath = download_dir / filename
                                        with open(filepath, 'wb') as f:
                                            f.write(base64.b64decode(b64_data))
                                        await page.keyboard.press("Escape")
                                        await page.wait_for_timeout(300)
                                        print(f"[{profile_id}] blob ë‹¤ìš´ë¡œë“œ ì„±ê³µ: {filename}")
                                        return f"/downloads/{profile_id}/{filename}"
                                
                            except Exception as e:
                                print(f"[{profile_id}] src ë‹¤ìš´ë¡œë“œ ì‹¤íŒ¨: {e}")
                            
                            # Fallback: ìŠ¤í¬ë¦°ìƒ· (íƒ€ì„ì•„ì›ƒ 5ì´ˆ)
                            try:
                                filename = f"full_{element_idx}_{int(datetime.now().timestamp())}.png"
                                filepath = download_dir / filename
                                await full_img.screenshot(path=str(filepath), timeout=5000)
                                await page.keyboard.press("Escape")
                                await page.wait_for_timeout(300)
                                return f"/downloads/{profile_id}/{filename}"
                            except Exception as e:
                                print(f"[{profile_id}] ì›ë³¸ ìŠ¤í¬ë¦°ìƒ· ì‹¤íŒ¨: {e}")
                                await page.keyboard.press("Escape")
                                return None
                        else:
                            # ì›ë³¸ ëª» ì°¾ìŒ
                            print(f"[{profile_id}] ì›ë³¸ ì´ë¯¸ì§€ ëª» ì°¾ìŒ")
                            await page.keyboard.press("Escape")
                            await page.wait_for_timeout(300)
                            return None
            
            elif media_type == "video":
                # ë™ì˜ìƒ ë‹¤ìš´ë¡œë“œ ë²„íŠ¼ í´ë¦­
                video_elements = await msg_el.locator('video, .video-msg').all()
                if sub_idx < len(video_elements):
                    vid_el = video_elements[sub_idx]
                    
                    # ë‹¤ìš´ë¡œë“œ ë²„íŠ¼ ì°¾ê¸°
                    download_btn = msg_el.locator('[aria-label="ë‹¤ìš´ë¡œë“œ"], [aria-label="Download"], .download-btn')
                    if await download_btn.count():
                        async with page.expect_download() as download_info:
                            await download_btn.first.click()
                        download = await download_info.value
                        filename = download.suggested_filename or f"video_{element_idx}.mp4"
                        filepath = download_dir / filename
                        await download.save_as(str(filepath))
                        return f"/downloads/{profile_id}/{filename}"
            
            elif media_type == "file":
                # íŒŒì¼ ë‹¤ìš´ë¡œë“œ
                file_elements = await msg_el.locator('.file-msg, .attachment').all()
                if sub_idx < len(file_elements):
                    file_el = file_elements[sub_idx]
                    
                    async with page.expect_download() as download_info:
                        await file_el.click()
                    download = await download_info.value
                    filename = download.suggested_filename or f"file_{element_idx}"
                    filepath = download_dir / filename
                    await download.save_as(str(filepath))
                    return f"/downloads/{profile_id}/{filename}"
            
            return None
            
        except Exception as e:
            print(f"[{profile_id}] ë¯¸ë””ì–´ ë‹¤ìš´ë¡œë“œ ì˜¤ë¥˜: {e}")
            return None
    
    async def search_by_phone(self, profile_id: str, phone_number: str) -> Dict[str, Any]:
        """ì „í™”ë²ˆí˜¸ë¡œ ëŒ€í™” ê²€ìƒ‰ (ìƒˆ ëŒ€í™” ì‹œì‘ ë²„íŠ¼ â†’ ë²ˆí˜¸ ì…ë ¥)"""
        if profile_id not in self.pages or not self.ready.get(profile_id):
            return {"error": "ë¸Œë¼ìš°ì € ë¯¸ì‹¤í–‰"}
        
        page = self.pages[profile_id]
        
        # ì „í™”ë²ˆí˜¸ ì •ê·œí™” (ìˆ«ìë§Œ)
        clean_number = re.sub(r'[^0-9]', '', phone_number)
        if len(clean_number) < 10:
            return {"error": "ì˜¬ë°”ë¥¸ ì „í™”ë²ˆí˜¸ë¥¼ ì…ë ¥í•˜ì„¸ìš”"}
        
        try:
            # ë’¤ë¡œê°€ê¸° (ëŒ€í™” ëª©ë¡ìœ¼ë¡œ)
            try:
                back = page.locator('button[aria-label="ë’¤ë¡œ ê°€ê¸°"], button[aria-label="Back"]')
                if await back.count() and await back.first.is_visible():
                    await back.first.click()
                    await page.wait_for_timeout(300)
            except:
                pass
            
            # "ì±„íŒ… ì‹œì‘" ë²„íŠ¼ í´ë¦­
            start_chat = page.locator('button[aria-label="ì±„íŒ… ì‹œì‘"], button[aria-label="Start chat"], a[href*="new"], [data-e2e-start-chat]')
            if not await start_chat.count():
                return {"error": "ì±„íŒ… ì‹œì‘ ë²„íŠ¼ì„ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤"}
            
            await start_chat.first.click()
            await page.wait_for_timeout(500)
            
            # ì „í™”ë²ˆí˜¸ ì…ë ¥ í•„ë“œ ì°¾ê¸°
            phone_input = page.locator('input[type="text"], input[aria-label*="ë²ˆí˜¸"], input[aria-label*="number"], input[placeholder*="ë²ˆí˜¸"], input[placeholder*="number"]')
            if not await phone_input.count():
                # ESCë¡œ ë‹«ê³  ì—ëŸ¬ ë°˜í™˜
                await page.keyboard.press("Escape")
                return {"error": "ì „í™”ë²ˆí˜¸ ì…ë ¥ í•„ë“œë¥¼ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤"}
            
            # ì…ë ¥ í•„ë“œ í´ë¦­í•˜ì—¬ í¬ì»¤ìŠ¤
            await phone_input.first.click()
            await page.wait_for_timeout(200)
            
            # ë²ˆí˜¸ ì…ë ¥ (íƒ€ì´í•‘ ë°©ì‹)
            await phone_input.first.fill("")
            await phone_input.first.type(clean_number, delay=50)
            await page.wait_for_timeout(500)
            
            # ì—”í„° í‚¤ ì…ë ¥ (ê²€ìƒ‰ ì‹¤í–‰)
            await phone_input.first.press("Enter")
            await page.wait_for_timeout(1500)
            
            # ê²€ìƒ‰ ê²°ê³¼ í™•ì¸
            found = False
            message_count = 0
            formatted_number = clean_number
            
            # ë©”ì‹œì§€ ì…€ë ‰í„°
            message_selectors = [
                'mws-message-wrapper',
                '.message-wrapper',
                '[data-e2e-message]',
                '.message-row',
                'mws-message-part',
                '.text-msg',
                'mws-bottom-nav ~ div [role="listitem"]'
            ]
            
            # ë°©ë²• 1: ëŒ€í™” ë‚´ìš©ì´ ë°”ë¡œ í‘œì‹œë˜ëŠ” ê²½ìš°
            for selector in message_selectors:
                msg_elements = await page.locator(selector).all()
                if len(msg_elements) > 0:
                    found = True
                    message_count = len(msg_elements)
                    print(f"[ê²€ìƒ‰] ëŒ€í™” ë‚´ìš© ì§ì ‘ í‘œì‹œë¨ ({selector}): {message_count}ê°œ ë©”ì‹œì§€")
                    break
            
            # ë°©ë²• 2: ì—°ë½ì²˜ ëª©ë¡ì´ ë‚˜íƒ€ë‚œ ê²½ìš° - í´ë¦­
            if not found:
                contact_item = page.locator('mws-contact-row, .contact-row, [data-e2e-contact], .contact-item')
                if await contact_item.count() > 0:
                    await contact_item.first.click()
                    await page.wait_for_timeout(1000)
                    
                    for selector in message_selectors:
                        msg_elements = await page.locator(selector).all()
                        if len(msg_elements) > 0:
                            found = True
                            message_count = len(msg_elements)
                            print(f"[ê²€ìƒ‰] ì—°ë½ì²˜ í´ë¦­ í›„ í‘œì‹œë¨ ({selector}): {message_count}ê°œ ë©”ì‹œì§€")
                            break
            
            # ë°©ë²• 3: URLì´ ëŒ€í™” í˜ì´ì§€ë¡œ ë³€ê²½ë˜ì—ˆëŠ”ì§€ í™•ì¸
            if not found:
                current_url = page.url
                if '/conversations/' in current_url:
                    # URL ë³€ê²½ë¨ = ëŒ€í™”ë°© ì¡´ì¬, ë©”ì‹œì§€ ë¡œë”© ëŒ€ê¸° (ìµœëŒ€ 3ì´ˆ ì¶”ê°€)
                    for retry in range(3):
                        await page.wait_for_timeout(1000)
                        for selector in message_selectors:
                            msg_elements = await page.locator(selector).all()
                            if len(msg_elements) > 0:
                                found = True
                                message_count = len(msg_elements)
                                print(f"[ê²€ìƒ‰] URL ë³€ê²½ í›„ ë©”ì‹œì§€ ë¡œë”©ë¨: {message_count}ê°œ (ëŒ€ê¸° {retry+1}ì´ˆ)")
                                break
                        if found:
                            break
            
            if found and message_count > 0:
                # ê²€ìƒ‰í•œ ë²ˆí˜¸ë¥¼ ê·¸ëŒ€ë¡œ ë°˜í™˜ (í™”ë©´ ì œëª©ì´ ì•„ë‹Œ)
                # í¬ë§·íŒ…ë§Œ ì¶”ê°€
                display_number = clean_number
                if len(clean_number) == 11:
                    display_number = f"{clean_number[:3]}-{clean_number[3:7]}-{clean_number[7:]}"
                elif len(clean_number) == 10:
                    display_number = f"{clean_number[:3]}-{clean_number[3:6]}-{clean_number[6:]}"
                
                return {
                    "success": True,
                    "found": True,
                    "phone_number": display_number,
                    "message_count": message_count,
                    "profile_id": profile_id
                }
            else:
                # ëŒ€í™”ë°©ì€ ìˆì§€ë§Œ ë©”ì‹œì§€ê°€ ì—†ê±°ë‚˜, ëŒ€í™”ë°© ìì²´ê°€ ì—†ìŒ
                await page.keyboard.press("Escape")
                await page.wait_for_timeout(300)
                
                return {
                    "success": True,
                    "found": False,
                    "phone_number": clean_number,
                    "message": "í•´ë‹¹ ë²ˆí˜¸ì™€ì˜ ëŒ€í™” ì´ë ¥ì´ ì—†ìŠµë‹ˆë‹¤"
                }
            
        except Exception as e:
            print(f"[{profile_id}] ì „í™”ë²ˆí˜¸ ê²€ìƒ‰ ì˜¤ë¥˜: {e}")
            try:
                await page.keyboard.press("Escape")
            except:
                pass
            return {"error": str(e)}

# ========== WebSocket ê´€ë¦¬ ==========
class ConnectionManager:
    def __init__(self):
        self.active_connections: List[WebSocket] = []
    
    async def connect(self, websocket: WebSocket):
        await websocket.accept()
        self.active_connections.append(websocket)
    
    def disconnect(self, websocket: WebSocket):
        if websocket in self.active_connections:
            self.active_connections.remove(websocket)
    
    async def broadcast(self, message: dict):
        for connection in self.active_connections:
            try:
                await connection.send_json(message)
            except:
                pass

# ========== ë¶ˆì‚¬ì ë§¤ë‹ˆì € ==========
import subprocess
import threading

class BulsajaManager:
    def __init__(self):
        self.is_running = False
        self.groups: List[Dict] = []  # [{num, status, message}]
        self.process: Optional[subprocess.Popen] = None
        self.stop_flag = False
        self.run_thread = None
        self.base_folder = "C:\\ìë™í™”ì‹œìŠ¤í…œ"
        self.logs = []
    
    def set_folder(self, folder: str):
        """í´ë” ê²½ë¡œ ì„¤ì •"""
        self.base_folder = folder
        return {"success": True, "folder": folder}
    
    def get_folder(self) -> str:
        """í˜„ì¬ í´ë” ê²½ë¡œ"""
        return self.base_folder
    
    def get_active_folder(self) -> str:
        """í˜„ì¬ í™œì„±í™”ëœ í´ë”"""
        return self.base_folder
    
    def find_exe(self) -> Optional[Path]:
        """C:\\ìë™í™”ì‹œìŠ¤í…œ\\*.exe íŒ¨í„´ìœ¼ë¡œ exe ì°¾ê¸°"""
        import glob
        exe_pattern = str(Path(self.base_folder) / "*.exe")
        exe_files = glob.glob(exe_pattern)
        if exe_files:
            # ê°€ì¥ ìµœê·¼ ìˆ˜ì •ëœ exe ë°˜í™˜
            return Path(max(exe_files, key=os.path.getmtime))
        return None
    
    def parse_groups(self, text: str) -> List[int]:
        """ê·¸ë£¹ ë¬¸ìì—´ íŒŒì‹±"""
        groups, seen = [], set()
        for part in text.replace(" ", "").split(","):
            if not part:
                continue
            if "-" in part:
                a, b = map(int, part.split("-", 1))
                step = 1 if a <= b else -1
                for n in range(a, b + step, step):
                    if 1 <= n <= 99 and n not in seen:
                        groups.append(n)
                        seen.add(n)
            else:
                n = int(part)
                if 1 <= n <= 99 and n not in seen:
                    groups.append(n)
                    seen.add(n)
        return groups
    
    def group_to_market_name(self, num: int) -> str:
        """ê·¸ë£¹ ë²ˆí˜¸ë¥¼ ë§ˆì¼“ê·¸ë£¹ëª…ìœ¼ë¡œ ë³€í™˜ (1 â†’ '1ë²ˆ ë§ˆì¼“ê·¸ë£¹')"""
        return f"{num}ë²ˆ ë§ˆì¼“ê·¸ë£¹"
    
    def start(self, groups_text: str, max_concurrent: int, group_gap: int, settings: dict = None):
        """ê·¸ë£¹ë³„ ì‹œíŠ¸ ë³€ê²½ + exe ì‹¤í–‰"""
        if self.is_running:
            return {"success": False, "message": "ì´ë¯¸ ì‹¤í–‰ ì¤‘ì…ë‹ˆë‹¤"}
        
        groups = self.parse_groups(groups_text)
        if not groups:
            return {"success": False, "message": "ìœ íš¨í•œ ê·¸ë£¹ì´ ì—†ìŠµë‹ˆë‹¤"}
        
        # exe ì°¾ê¸°
        exe_path = self.find_exe()
        if not exe_path:
            return {"success": False, "message": f"exe íŒŒì¼ì„ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤: {self.base_folder}\\*.exe"}
        
        # ì´ˆê¸°í™”
        self.groups = [{"num": g, "status": "pending", "message": ""} for g in groups]
        self.stop_flag = False
        self.is_running = True
        self.logs = []
        
        # ë°±ê·¸ë¼ìš´ë“œ ìŠ¤ë ˆë“œì—ì„œ ì‹¤í–‰
        self.run_thread = threading.Thread(
            target=self._run_groups,
            args=(exe_path, groups, max_concurrent, group_gap, settings or {}),
            daemon=True
        )
        self.run_thread.start()
        
        return {"success": True, "groups": groups}
    
    def _run_groups(self, exe_path: Path, groups: List[int], max_concurrent: int, group_gap: int, settings: dict):
        """ê·¸ë£¹ë³„ ì‹œíŠ¸ ë³€ê²½ + exe ì‹¤í–‰ (ìŠ¤ë ˆë“œ)"""
        import asyncio
        from dotenv import dotenv_values
        
        try:
            # ìë™í™”ì‹œìŠ¤í…œ í´ë”ì˜ .env ì½ê¸°
            env_path = Path(self.base_folder) / ".env"
            if env_path.exists():
                env_config = dotenv_values(env_path)
                creds_file = env_config.get("SERVICE_ACCOUNT_JSON", CREDENTIALS_FILE)
                sheet_key = env_config.get("SPREADSHEET_KEY", BULSAJA_SHEET_KEY)
                self._add_log(f".env ë¡œë“œ: {env_path}")
            else:
                creds_file = CREDENTIALS_FILE
                sheet_key = BULSAJA_SHEET_KEY
                self._add_log(f".env ì—†ìŒ, ê¸°ë³¸ê°’ ì‚¬ìš©")
            
            # ìƒëŒ€ê²½ë¡œë©´ base_folder ê¸°ì¤€ìœ¼ë¡œ ë³€í™˜
            creds_path = Path(creds_file)
            if not creds_path.is_absolute():
                creds_path = Path(self.base_folder) / creds_file
            
            # êµ¬ê¸€ì‹œíŠ¸ ì—°ê²°
            creds = Credentials.from_service_account_file(
                str(creds_path),
                scopes=["https://www.googleapis.com/auth/spreadsheets"]
            )
            gc = gspread.authorize(creds)
            ws = gc.open_by_key(sheet_key).worksheet(BULSAJA_TAB_NAME)
            
            # ========== ì„¤ì •ê°’ ë¨¼ì € ì €ì¥ ==========
            program = settings.get("program", "")
            if program:
                ws.update_acell("C10", program)
                self._add_log(f"[LOG] C10 â†’ {program}")
                
                if "ìƒí’ˆì—…ë¡œë“œ" in program:
                    target_cell = "C15"
                    if settings.get("uploadMarket"):
                        ws.update_acell("C17", settings["uploadMarket"])
                        self._add_log(f"[LOG] C17 â†’ {settings['uploadMarket']}")
                    if settings.get("uploadCount"):
                        ws.update_acell("C18", settings["uploadCount"])
                        self._add_log(f"[LOG] C18 â†’ {settings['uploadCount']}")
                        
                elif "ìƒí’ˆì‚­ì œ" in program:
                    target_cell = "C29"
                    if settings.get("deleteCount"):
                        ws.update_acell("C32", settings["deleteCount"])
                        self._add_log(f"[LOG] C32 â†’ {settings['deleteCount']}")
                        
                elif "ìƒí’ˆë³µì‚¬" in program:
                    target_cell = "C35"
                    if settings.get("copySourceMarket"):
                        ws.update_acell("C29", settings["copySourceMarket"])
                        self._add_log(f"[LOG] C29 â†’ {settings['copySourceMarket']}")
                    if settings.get("copyCount"):
                        ws.update_acell("C37", settings["copyCount"])
                        self._add_log(f"[LOG] C37 â†’ {settings['copyCount']}")
                else:
                    target_cell = "C15"
            else:
                # ì„¤ì •ì´ ì—†ìœ¼ë©´ í˜„ì¬ ì‹œíŠ¸ ê°’ ì‚¬ìš©
                program = ws.acell("C10").value or ""
                if "ìƒí’ˆì—…ë¡œë“œ" in program:
                    target_cell = "C15"
                elif "ìƒí’ˆì‚­ì œ" in program:
                    target_cell = "C29"
                elif "ìƒí’ˆë³µì‚¬" in program:
                    target_cell = "C35"
                else:
                    target_cell = "C15"
            
            self._add_log(f"í”„ë¡œê·¸ë¨: {program}")
            self._add_log(f"ì‹¤í–‰íŒŒì¼: {exe_path.name}")
            self._add_log(f"ëŒ€ìƒ ì…€: {target_cell}, ê·¸ë£¹: {groups}")
            
            # ë™ì‹œ ì‹¤í–‰ ê´€ë¦¬
            running_processes = []
            group_idx = 0
            
            while group_idx < len(groups) or running_processes:
                if self.stop_flag:
                    self._add_log("[STOP] ì¤‘ì§€ ìš”ì²­ë¨")
                    break
                
                # ì™„ë£Œëœ í”„ë¡œì„¸ìŠ¤ ì •ë¦¬
                for proc_info in running_processes[:]:
                    proc, gnum = proc_info
                    if proc.poll() is not None:  # í”„ë¡œì„¸ìŠ¤ ì¢…ë£Œë¨
                        exit_code = proc.returncode
                        if exit_code == 0:
                            self._update_group_status(gnum, "completed", "ì™„ë£Œ")
                            self._add_log(f"[SUCCESS] ê·¸ë£¹ {gnum} ì™„ë£Œ")
                        else:
                            self._update_group_status(gnum, "failed", f"exit: {exit_code}")
                            self._add_log(f"[ERROR] ê·¸ë£¹ {gnum} ì‹¤íŒ¨ (exit={exit_code})")
                        running_processes.remove(proc_info)
                
                # ìƒˆ í”„ë¡œì„¸ìŠ¤ ì‹œì‘ (ë™ì‹œ ì‹¤í–‰ ìˆ˜ ì´ë‚´)
                while len(running_processes) < max_concurrent and group_idx < len(groups):
                    if self.stop_flag:
                        break
                    
                    gnum = groups[group_idx]
                    group_idx += 1
                    
                    # ì‹œíŠ¸ ì…€ ë³€ê²½
                    market_name = self.group_to_market_name(gnum)
                    try:
                        ws.update_acell(target_cell, market_name)
                        self._add_log(f"[LOG] {target_cell} â†’ {market_name}")
                    except Exception as e:
                        self._update_group_status(gnum, "failed", f"ì‹œíŠ¸ ì˜¤ë¥˜: {e}")
                        self._add_log(f"[ERROR] ê·¸ë£¹ {gnum} ì‹œíŠ¸ ë³€ê²½ ì‹¤íŒ¨: {e}")
                        continue
                    
                    # exe ì‹¤í–‰ (ë³„ë„ ì°½ì—ì„œ)
                    try:
                        self._update_group_status(gnum, "running", "ì‹¤í–‰ ì¤‘")
                        self._add_log(f"[ENV] exe ê²½ë¡œ: {exe_path}")
                        
                        # Windowsì—ì„œ ë³„ë„ ì°½ìœ¼ë¡œ exe ì‹¤í–‰
                        if os.name == "nt":
                            proc = subprocess.Popen(
                                [str(exe_path)],
                                cwd=str(exe_path.parent),
                                creationflags=subprocess.CREATE_NEW_CONSOLE
                            )
                        else:
                            proc = subprocess.Popen(
                                [str(exe_path)],
                                cwd=str(exe_path.parent)
                            )
                        running_processes.append((proc, gnum))
                        self._add_log(f"[RUN] ê·¸ë£¹ {gnum} exe ì‹¤í–‰ (PID: {proc.pid})")
                        
                        # ë‹¤ìŒ ê·¸ë£¹ ì „ ê°„ê²© ëŒ€ê¸°
                        if group_idx < len(groups):
                            self._add_log(f"[WAIT] {group_gap}ì´ˆ ëŒ€ê¸°...")
                            for _ in range(group_gap):
                                if self.stop_flag:
                                    break
                                time.sleep(1)
                    except Exception as e:
                        self._update_group_status(gnum, "failed", f"ì‹¤í–‰ ì˜¤ë¥˜: {e}")
                        self._add_log(f"[ERROR] ê·¸ë£¹ {gnum} exe ì‹¤í–‰ ì‹¤íŒ¨: {e}")
                
                time.sleep(1)  # í´ë§ ê°„ê²©
            
            # ë‚¨ì€ í”„ë¡œì„¸ìŠ¤ ëŒ€ê¸°
            for proc, gnum in running_processes:
                try:
                    proc.wait(timeout=300)  # 5ë¶„ íƒ€ì„ì•„ì›ƒ
                    if proc.returncode == 0:
                        self._update_group_status(gnum, "completed", "ì™„ë£Œ")
                    else:
                        self._update_group_status(gnum, "failed", f"exit: {proc.returncode}")
                except:
                    self._update_group_status(gnum, "failed", "íƒ€ì„ì•„ì›ƒ")
            
            self._add_log("[SUCCESS] ëª¨ë“  ì‘ì—… ì™„ë£Œ!")
            
        except Exception as e:
            self._add_log(f"[ERROR] ì˜¤ë¥˜: {e}")
            for g in self.groups:
                if g["status"] in ["pending", "running"]:
                    g["status"] = "failed"
                    g["message"] = str(e)
        finally:
            self.is_running = False
    
    def _add_log(self, msg: str):
        """ë¡œê·¸ ì¶”ê°€ + WebSocket ë¸Œë¡œë“œìºìŠ¤íŠ¸"""
        import asyncio
        
        timestamp = datetime.now().strftime("%H:%M:%S")
        print(f"[ë¶ˆì‚¬ì {timestamp}] {msg}")
        self.logs.append({"time": timestamp, "msg": msg})
        
        # ìµœê·¼ 100ì¤„ë§Œ ìœ ì§€
        if len(self.logs) > 100:
            self.logs = self.logs[-100:]
        
        # WebSocketìœ¼ë¡œ ì „ì†¡
        try:
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            loop.run_until_complete(ws_manager.broadcast({
                "type": "bulsaja_log",
                "message": msg,
                "timestamp": timestamp
            }))
            loop.close()
        except:
            pass
    
    def _update_group_status(self, group_num: int, status: str, message: str):
        """ê·¸ë£¹ ìƒíƒœ ì—…ë°ì´íŠ¸"""
        for g in self.groups:
            if g["num"] == group_num:
                g["status"] = status
                g["message"] = message
                break
    
    def stop(self):
        """ì‹¤í–‰ ì¤‘ì§€ - njbul exeë“¤ë„ ì¢…ë£Œ"""
        self.stop_flag = True
        
        # njbul exeë“¤ ê°•ì œ ì¢…ë£Œ (Windows)
        if os.name == "nt":
            try:
                result = subprocess.run(
                    'wmic process where "name like \'njbul%\'" call terminate',
                    shell=True,
                    capture_output=True,
                    text=True
                )
                if "ì„±ê³µ" in result.stdout or "successfully" in result.stdout.lower():
                    print(f"[ë¶ˆì‚¬ì] njbul exe ì¢…ë£Œë¨")
                elif "ì¸ìŠ¤í„´ìŠ¤ê°€ ì—†ìŠµë‹ˆë‹¤" in result.stdout or "No Instance" in result.stdout:
                    print(f"[ë¶ˆì‚¬ì] ì‹¤í–‰ ì¤‘ì¸ njbul exe ì—†ìŒ")
            except Exception as e:
                print(f"[ë¶ˆì‚¬ì] wmic ì˜¤ë¥˜: {e}")
        
        self.is_running = False
        for g in self.groups:
            if g["status"] == "running" or g["status"] == "pending":
                g["status"] = "failed"
                g["message"] = "ì¤‘ì§€ë¨"
        return {"success": True, "message": "njbul exe ì¢…ë£Œë¨"}
    
    def get_status(self):
        """í˜„ì¬ ìƒíƒœ ë°˜í™˜"""
        counts = {"pending": 0, "running": 0, "completed": 0, "failed": 0}
        for g in self.groups:
            counts[g["status"]] = counts.get(g["status"], 0) + 1
        return {
            "is_running": self.is_running,
            "groups": self.groups,
            "counts": counts
        }

# ========== ì „ì—­ ì¸ìŠ¤í„´ìŠ¤ ==========
gsheet = GoogleSheetManager()
sms_manager = SMSBrowserManager()
ws_manager = ConnectionManager()
bulsaja_manager = BulsajaManager()

# ========== FastAPI ì•± ==========
@asynccontextmanager
async def lifespan(app: FastAPI):
    # ì‹œì‘ì‹œ
    print(f"[DEBUG] Starting lifespan. Active loop: {type(asyncio.get_running_loop()).__name__}")
    gsheet.connect()
    
    # ìŠ¤ì¼€ì¤„ëŸ¬ ì‹œì‘
    scheduler.start()
    print("[ì„œë²„ì‹œì‘] ìŠ¤ì¼€ì¤„ëŸ¬ ì‹œì‘ë¨")
    
    # ì €ì¥ëœ ìŠ¤ì¼€ì¤„ ë³µì›
    schedules = load_schedules()
    for s in schedules:
        if s.get('enabled', True):
            try:
                add_schedule_job(s)
                print(f"[ìŠ¤ì¼€ì¤„ëŸ¬] ë³µì›: {s.get('name')} ({s.get('id')})")
            except Exception as e:
                print(f"[ìŠ¤ì¼€ì¤„ëŸ¬] ë³µì› ì‹¤íŒ¨: {s.get('name')} - {e}")
    
    # SMS ë¸Œë¼ìš°ì € ìë™ ì‹¤í–‰
    async def delayed_sms_launch():
        await asyncio.sleep(5)  # ì„œë²„ ì™„ì „íˆ ëœ¬ í›„ 5ì´ˆ ëŒ€ê¸°
        print("[ì„œë²„ì‹œì‘] SMS ë¸Œë¼ìš°ì € ì „ì²´ ì‹¤í–‰ ì¤‘...")
        try:
            await sms_manager.launch_all()
            print("[ì„œë²„ì‹œì‘] SMS ë¸Œë¼ìš°ì € ì „ì²´ ì‹¤í–‰ ì™„ë£Œ")
        except Exception as e:
            print(f"[ì„œë²„ì‹œì‘] SMS ë¸Œë¼ìš°ì € ì‹¤í–‰ ì˜¤ë¥˜: {e}")
    
    asyncio.create_task(delayed_sms_launch())
    
    yield
    # ì¢…ë£Œì‹œ
    scheduler.shutdown(wait=False)
    print("[ì„œë²„ì¢…ë£Œ] ìŠ¤ì¼€ì¤„ëŸ¬ ì¢…ë£Œë¨")
    
    # SMS ë¸Œë¼ìš°ì € ì¢…ë£Œ
    await sms_manager.close_all()
    
    # ë¼ì´ë¸Œ ì•Œë¦¬ ë¸Œë¼ìš°ì € ì¢…ë£Œ (ìˆëŠ” ê²½ìš°)
    try:
        global ali_browser
        if ali_browser:
            # ë™ê¸°/ë¹„ë™ê¸° ë¸Œë¼ìš°ì € íƒ€ì…ì— ë”°ë¼ ëŒ€ì‘
            if hasattr(ali_browser, 'close'):
                if asyncio.iscoroutinefunction(ali_browser.close):
                    await ali_browser.close()
                else:
                    ali_browser.close()
            print("[ì„œë²„ì¢…ë£Œ] ì•Œë¦¬ ë¸Œë¼ìš°ì € ì¢…ë£Œ ì™„ë£Œ")
    except Exception as e:
        print(f"[ì„œë²„ì¢…ë£Œ] ì•Œë¦¬ ë¸Œë¼ìš°ì € ì¢…ë£Œ ì¤‘ ì˜¤ë¥˜: {e}")

def add_schedule_job(schedule: Dict):
    """ìŠ¤ì¼€ì¤„ ì‘ì—… ë“±ë¡"""
    job_id = schedule['id']
    schedule_type = schedule.get('schedule_type', 'cron')
    
    # ê¸°ì¡´ ì‘ì—… ì œê±°
    if scheduler.get_job(job_id):
        scheduler.remove_job(job_id)
    
    if schedule_type == 'cron':
        # cron í‘œí˜„ì‹: ë¶„ ì‹œ ì¼ ì›” ìš”ì¼
        cron_expr = schedule.get('cron', '0 9 * * *')  # ê¸°ë³¸ ë§¤ì¼ 09:00
        parts = cron_expr.split()
        trigger = CronTrigger(
            minute=parts[0] if len(parts) > 0 else '0',
            hour=parts[1] if len(parts) > 1 else '9',
            day=parts[2] if len(parts) > 2 else '*',
            month=parts[3] if len(parts) > 3 else '*',
            day_of_week=parts[4] if len(parts) > 4 else '*'
        )
    elif schedule_type == 'interval':
        # ê°„ê²© (ë¶„ ë‹¨ìœ„)
        interval_minutes = schedule.get('interval_minutes', 60)
        trigger = IntervalTrigger(minutes=interval_minutes)
    else:
        return
    
    scheduler.add_job(
        execute_scheduled_task,
        trigger=trigger,
        id=job_id,
        args=[
            job_id,
            schedule.get('platform', 'ìŠ¤ë§ˆíŠ¸ìŠ¤í† ì–´'),
            schedule.get('task', 'ë“±ë¡ê°¯ìˆ˜'),
            schedule.get('stores', []),
            schedule.get('options', {})
        ],
        replace_existing=True
    )

app = FastAPI(title="êµ¬ë§¤ëŒ€í–‰ í†µí•©ê´€ë¦¬", lifespan=lifespan)

# CORS - í¬ë¡¬ í™•ì¥í”„ë¡œê·¸ë¨ì—ì„œ credentials í¬í•¨ ìš”ì²­ í—ˆìš©
# allow_origins=["*"]ì™€ allow_credentials=TrueëŠ” í•¨ê»˜ ì‚¬ìš© ë¶ˆê°€
# ëŒ€ì‹  allow_origin_regexë¡œ ëª¨ë“  origin í—ˆìš©
app.add_middleware(
    CORSMiddleware,
    allow_origin_regex=r".*",  # ëª¨ë“  origin í—ˆìš© (í¬ë¡¬ í™•ì¥í”„ë¡œê·¸ë¨ í¬í•¨)
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ì •ì  íŒŒì¼ & í…œí”Œë¦¿
static_dir = APP_DIR / "static"
static_dir.mkdir(exist_ok=True)
templates_dir = APP_DIR / "templates"
templates_dir.mkdir(exist_ok=True)

app.mount("/static", StaticFiles(directory=static_dir), name="static")
templates = Jinja2Templates(directory=templates_dir)

# ========== API ë¼ìš°íŠ¸ ==========

# ë¡œê·¸ì¸ í˜ì´ì§€
@app.get("/login", response_class=HTMLResponse)
async def login_page(request: Request):
    return templates.TemplateResponse("login.html", {"request": request})

# ë¡œê·¸ì¸ ì²˜ë¦¬
@app.post("/api/login")
async def login(req: LoginRequest, refresh: bool = False):
    result = gsheet.verify_user(req.username, req.password, force_refresh=refresh)
    if result["success"]:
        staff_name = result["name"] or req.username
        role = result.get("role", ROLE_VIEWER)
        token = create_session(req.username, staff_name, role)
        response = JSONResponse({
            "success": True,
            "username": req.username,
            "name": staff_name,
            "role": role
        })
        # samesite=laxë¡œ ì„¤ì • (ì™¸ë¶€ IP ì ‘ì† ì‹œì—ë„ ì¿ í‚¤ ì „ì†¡)
        response.set_cookie(
            "session_token", 
            token, 
            httponly=True, 
            max_age=SESSION_EXPIRE_HOURS*3600,
            samesite="lax"
        )
        return response
    raise HTTPException(status_code=401, detail="ì•„ì´ë”” ë˜ëŠ” ë¹„ë°€ë²ˆí˜¸ê°€ í‹€ë ¸ìŠµë‹ˆë‹¤")

# ë¡œê·¸ì•„ì›ƒ
@app.post("/api/logout")
async def logout(request: Request):
    token = request.cookies.get("session_token")
    if token in sessions:
        del sessions[token]
    response = JSONResponse({"success": True})
    response.delete_cookie("session_token")
    return response

# ë©”ì¸ í˜ì´ì§€
@app.get("/", response_class=HTMLResponse)
async def main_page(request: Request):
    token = request.cookies.get("session_token")
    user = verify_session(token)
    if not user:
        return RedirectResponse("/login")
    return templates.TemplateResponse("index.html", {
        "request": request,
        "username": user["name"],
        "role": user.get("role", ROLE_VIEWER)
    })

# í˜„ì¬ ì‚¬ìš©ì
@app.get("/api/me")
async def get_me(request: Request):
    user = get_current_user(request)
    role = user.get("role", ROLE_VIEWER)

    response = {
        "username": user["username"],
        "name": user["name"],
        "role": role,
        "permissions": get_role_permissions(role)
    }

    # ìš´ì˜ìì¸ ê²½ìš° íƒ­ ê¶Œí•œ ì •ë³´ë„ ë°˜í™˜
    if role == ROLE_OPERATOR:
        try:
            if TAB_PERMISSIONS_FILE.exists():
                with open(TAB_PERMISSIONS_FILE, 'r', encoding='utf-8') as f:
                    response["tab_permissions"] = json.load(f)
            else:
                # ê¸°ë³¸ê°’: ëª¨ë“  íƒ­ í—ˆìš©
                response["tab_permissions"] = {
                    "sms": True, "monitor": True, "market": True, "sales": True,
                    "accounts": True, "marketing": True, "aio": True, "scheduler": True,
                    "bulsaja": True, "tools": True, "calendar": True
                }
        except Exception as e:
            print(f"[íƒ­ ê¶Œí•œ ë¡œë“œ ì˜¤ë¥˜] {e}")
            response["tab_permissions"] = None

    return response

# ê³„ì • ëª©ë¡
@app.get("/api/accounts")
async def get_accounts(request: Request, platform: str = None, refresh: bool = False):
    get_current_user(request)
    accounts = gsheet.get_accounts(platform, force_refresh=refresh)
    
    # í”Œë«í¼ë³„ ìˆ˜ëŸ‰ ê³„ì‚°
    all_accounts = gsheet.get_accounts(None, force_refresh=refresh)
    platform_counts = {}
    for acc in all_accounts:
        p = acc.get("platform", "")
        platform_counts[p] = platform_counts.get(p, 0) + 1
    
    # ë¹„ë°€ë²ˆí˜¸/ì‹œí¬ë¦¿ ë§ˆìŠ¤í‚¹
    for acc in accounts:
        if acc.get("password"):
            acc["password_masked"] = "â—" * min(len(acc["password"]), 8)
        if acc.get("ss_app_secret"):
            acc["ss_app_secret_masked"] = acc["ss_app_secret"][:4] + "â—â—â—â—" if len(acc["ss_app_secret"]) > 4 else "â—â—â—â—"
        if acc.get("cp_secret_key"):
            acc["cp_secret_key_masked"] = acc["cp_secret_key"][:4] + "â—â—â—â—" if len(acc["cp_secret_key"]) > 4 else "â—â—â—â—"
    
    return {
        "accounts": accounts, 
        "platforms": list(PLATFORM_CONFIG.keys()),
        "platform_counts": platform_counts,
        "total_count": len(all_accounts)
    }

# ê³„ì • ìƒì„¸ (ë¹„ë°€ë²ˆí˜¸ í¬í•¨)
@app.get("/api/accounts/{platform}/{account_id}")
async def get_account_detail(request: Request, platform: str, account_id: str):
    get_current_user(request)
    accounts = gsheet.get_accounts(platform)
    for acc in accounts:
        login_id = acc.get("ì•„ì´ë””") or acc.get("login_id") or ""
        if login_id == account_id:
            return acc
    raise HTTPException(status_code=404, detail="ê³„ì •ì„ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤")

@app.get("/api/accounts/search")
async def search_account(request: Request, shop_alias: str = None, store_name: str = None, platform: str = None, login_id: str = None):
    """ê³„ì • ê²€ìƒ‰ - store_name ë˜ëŠ” login_idë¡œ ê²€ìƒ‰"""
    get_current_user(request)
    
    # store_nameì´ ì—†ìœ¼ë©´ shop_alias ì‚¬ìš© (í•˜ìœ„ í˜¸í™˜)
    search_name = store_name or shop_alias
    
    # ëª¨ë“  í”Œë«í¼ì—ì„œ ê²€ìƒ‰
    platforms_to_search = [platform] if platform else ["ìŠ¤ë§ˆíŠ¸ìŠ¤í† ì–´", "ì¿ íŒ¡", "11ë²ˆê°€", "ì§€ë§ˆì¼“", "ì˜¥ì…˜", "ESMí†µí•©"]
    
    for p in platforms_to_search:
        accounts = gsheet.get_accounts(p)
        for acc in accounts:
            # store_nameìœ¼ë¡œ ê²€ìƒ‰
            acc_name = acc.get("ìŠ¤í† ì–´ëª…") or acc.get("ìŠ¤í† ì–´ëª…", "")
            if search_name and acc_name == search_name:
                return acc
            # login_idë¡œ ê²€ìƒ‰
            if login_id and acc.get("login_id") == login_id:
                return acc
    
    return {"login_id": None, "message": "ê³„ì •ì„ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤"}

# ê³„ì • ì¶”ê°€
@app.post("/api/accounts")
async def add_account(request: Request, account: AccountModel):
    require_permission(request, "edit")  # ìš´ì˜ì ì´ìƒ
    if gsheet.add_account(account.dict()):
        await ws_manager.broadcast({"type": "account_update"})
        return {"success": True}
    raise HTTPException(status_code=500, detail="ì¶”ê°€ ì‹¤íŒ¨")

# ê³„ì • ìˆ˜ì •
@app.put("/api/accounts/{platform}/{account_id}")
async def update_account(request: Request, platform: str, account_id: str):
    require_permission(request, "edit")  # ìš´ì˜ì ì´ìƒ
    
    # ìš”ì²­ ë³¸ë¬¸ì—ì„œ ì‹¤ì œ ì „ë‹¬ëœ í•„ë“œë§Œ ê°€ì ¸ì˜¤ê¸°
    body = await request.json()
    
    if gsheet.update_account(account_id, platform, body):
        await ws_manager.broadcast({"type": "account_update"})
        return {"success": True}
    raise HTTPException(status_code=500, detail="ìˆ˜ì • ì‹¤íŒ¨")

# ê³„ì • ì‚­ì œ
@app.delete("/api/accounts/{platform}/{account_id}")
async def delete_account(request: Request, platform: str, account_id: str):
    require_permission(request, "delete")  # ê´€ë¦¬ìë§Œ
    if gsheet.delete_account(account_id, platform):
        await ws_manager.broadcast({"type": "account_update"})
        return {"success": True}
    raise HTTPException(status_code=500, detail="ì‚­ì œ ì‹¤íŒ¨")

# ========== ê´€ì œì„¼í„° API ==========

@app.get("/api/monitor/daily-status")
async def get_daily_status(request: Request):
    """ë§ˆì¼“ë³„ ìƒíƒœ ì¡°íšŒ - ìƒˆë¡œìš´ ì „ìš© ë°ì´í„° ì‹œíŠ¸(1r-ROJ...) ì°¸ì¡°"""
    get_current_user(request)
    
    try:
        # 1. ì™¸ë¶€ ì „ìš© ì‹œíŠ¸ì—ì„œ ë“±ë¡ê°¯ìˆ˜/11ë²ˆê°€ ë°ì´í„° ê°€ì ¸ì˜¤ê¸° (ì „ìš© ì¸ì¦ íŒŒì¼ ì‚¬ìš©)
        ss_counts = {}
        st_counts = {}
        ss_reg_map = {}
        st_reg_map = {}
        today = datetime.now().date()
        
        # ìŠ¤ë§ˆíŠ¸ìŠ¤í† ì–´ (ë“±ë¡ê°¯ìˆ˜ íƒ­)
        ws_counts = gsheet.open_worksheet_with_creds(COUNT_CREDENTIALS_FILE, SPREADSHEET_KEY, "ë“±ë¡ê°¯ìˆ˜")
        if ws_counts:
            data = ws_counts.get_all_values()
            if len(data) > 1:
                h = data[0]
                n_idx = next((i for i, v in enumerate(h) if v in ["store_name", "ìŠ¤í† ì–´ëª…"]), None)
                c_idx = next((i for i, v in enumerate(h) if v == "íŒë§¤ì¤‘"), None)
                r_idx = next((i for i, v in enumerate(h) if "ë§ˆì§€ë§‰" in v and "ë“±ë¡" in v), None)
                
                if n_idx is not None:
                    for row in data[1:]:
                        if len(row) > n_idx:
                            store = row[n_idx].strip()
                            if not store: continue
                            # ìƒí’ˆìˆ˜
                            if c_idx is not None and len(row) > c_idx:
                                try: ss_counts[store] = int(row[c_idx]) if row[c_idx] else 0
                                except: ss_counts[store] = 0
                            # ë§ˆì§€ë§‰ ë“±ë¡ì¼
                            if r_idx is not None and len(row) > r_idx and row[r_idx]:
                                try:
                                    d_str = row[r_idx][:10]
                                    d_val = datetime.strptime(d_str, "%Y-%m-%d").date()
                                    ss_reg_map[store] = {"date": d_str, "days": (today - d_val).days}
                                except: pass

        # 11ë²ˆê°€ (11ë²ˆê°€ íƒ­)
        ws_11st = gsheet.open_worksheet_with_creds(COUNT_CREDENTIALS_FILE, SPREADSHEET_KEY, "11ë²ˆê°€")
        if ws_11st:
            data = ws_11st.get_all_values()
            if len(data) > 1:
                h = data[0]
                n_idx = next((i for i, v in enumerate(h) if v in ["store_name", "ì‡¼í•‘ëª° ë³„ì¹­", "ìŠ¤í† ì–´ëª…"]), None)
                c_idx = next((i for i, v in enumerate(h) if v == "íŒë§¤ì¤‘"), None)
                r_idx = next((i for i, v in enumerate(h) if "ë§ˆì§€ë§‰" in v and "ë“±ë¡" in v), None)
                
                if n_idx is not None:
                    for row in data[1:]:
                        if len(row) > n_idx:
                            store = row[n_idx].strip()
                            if not store: continue
                            # ìƒí’ˆìˆ˜
                            if c_idx is not None and len(row) > c_idx:
                                try: st_counts[store] = int(row[c_idx]) if row[c_idx] else 0
                                except: st_counts[store] = 0
                            # ë§ˆì§€ë§‰ ë“±ë¡ì¼
                            if r_idx is not None and len(row) > r_idx and row[r_idx]:
                                try:
                                    d_str = row[r_idx][:10]
                                    d_val = datetime.strptime(d_str, "%Y-%m-%d").date()
                                    st_reg_map[store] = {"date": d_str, "days": (today - d_val).days}
                                except: pass

        # 2. ê³„ì • ì •ë³´ ê°€ì ¸ì˜¤ê¸°
        all_accounts = gsheet.get_accounts()
        
        # 3. ì‘ì—…ë¡œê·¸ (ì‚­ì œ ì‘ì—… ë“±) - ê¸°ì¡´ ì‹œíŠ¸ ì°¸ì¡°
        last_work_map = {}
        try:
            ws_worklog = gsheet.sheet.worksheet("ì‘ì—…ë¡œê·¸")
            worklog_data = ws_worklog.get_all_values()
            if len(worklog_data) > 1:
                h = worklog_data[0]
                d_idx = next((i for i, v in enumerate(h) if 'ì¼ì‹œ' in v or 'ë‚ ì§œ' in v), 0)
                a_idx = next((i for i, v in enumerate(h) if 'ê³„ì •' in v or 'ìŠ¤í† ì–´' in v), 2)
                for row in worklog_data[1:]:
                    if len(row) > max(d_idx, a_idx):
                        date_str = row[d_idx].strip()
                        acc_name = row[a_idx].strip()
                        if not date_str or not acc_name: continue
                        try:
                            w_date = datetime.strptime(date_str.split()[0].replace('/', '-'), "%Y-%m-%d").date()
                            if acc_name not in last_work_map or w_date > datetime.strptime(last_work_map[acc_name]["date"], "%Y-%m-%d").date():
                                last_work_map[acc_name] = {"date": w_date.strftime("%Y-%m-%d"), "days": (today - w_date).days}
                        except: pass
        except: pass

        result_data = []
        markets_set = set()
        usages_set = set()
        
        for idx, acc in enumerate(all_accounts):
            platform = acc.get("í”Œë«í¼") or acc.get("platform") or ""
            store_name = (acc.get("ìŠ¤í† ì–´ëª…") or "").strip()
            login_id = (acc.get("ì•„ì´ë””") or acc.get("login_id") or "").strip()
            usage = (acc.get("ìš©ë„") or acc.get("usage") or "").strip()
            owner = (acc.get("ì†Œìœ ì") or acc.get("owner") or "").strip()
            if not platform: continue
            
            market = platform
            if "ìŠ¤ë§ˆíŠ¸" in platform or "ë„¤ì´ë²„" in platform: market = "ìŠ¤ë§ˆíŠ¸ìŠ¤í† ì–´"
            elif "11" in platform: market = "11ë²ˆê°€"
            elif "ì¿ íŒ¡" in platform: market = "ì¿ íŒ¡"
            elif "ì§€ë§ˆì¼“" in platform: market = "ì§€ë§ˆì¼“"
            elif "ì˜¥ì…˜" in platform: market = "ì˜¥ì…˜"
            elif "ESM" in platform: market = "ESM"
            markets_set.add(market)
            if usage: usages_set.add(usage)
            
            # ë§¤ì¹­ìš© ì´ë¦„ ì •ê·œí™”
            account_name = store_name if store_name else login_id
            
            # ìƒí’ˆìˆ˜ ë§¤ì¹­
            product_count = 0
            if market == "ìŠ¤ë§ˆíŠ¸ìŠ¤í† ì–´":
                product_count = ss_counts.get(account_name, 0)
                if product_count == 0 and "_" in account_name:
                    product_count = ss_counts.get(account_name.split("_", 1)[1], 0)
            elif market == "11ë²ˆê°€":
                product_count = st_counts.get(account_name, 0)
                if product_count == 0: product_count = st_counts.get(login_id, 0)
            
            # ë§ˆì§€ë§‰ ë“±ë¡ì¼ ë§¤ì¹­
            last_cleanup_date = datetime.now().strftime("%Y-%m-%d")
            days_since_cleanup = 0
            
            # í”Œë«í¼ë³„ ì „ìš© ì‹œíŠ¸ ë§¤ì¹­ ì‹œë„
            match_name = account_name
            found = False
            target_reg_map = ss_reg_map if market == "ìŠ¤ë§ˆíŠ¸ìŠ¤í† ì–´" else (st_reg_map if market == "11ë²ˆê°€" else {})
            
            if match_name in target_reg_map: 
                found = True
            elif "_" in match_name and match_name.split("_", 1)[1] in target_reg_map:
                match_name = match_name.split("_", 1)[1]
                found = True
            
            if found:
                last_cleanup_date = target_reg_map[match_name]['date']
                days_since_cleanup = target_reg_map[match_name]['days']
            elif account_name in last_work_map:
                last_cleanup_date = last_work_map[account_name]['date']
                days_since_cleanup = last_work_map[account_name]['days']

            # ìƒíƒœ (20ì¼/30ì¼ ê¸°ì¤€ ìœ ì§€í•˜ë˜ ì£¼ ë‹¨ìœ„ í•„í„°ëŠ” í”„ë¡ íŠ¸ì—”ë“œì—ì„œ ì²˜ë¦¬)
            cleanup_status = 'normal'
            if days_since_cleanup > 30: cleanup_status = 'urgent'
            elif days_since_cleanup > 20: cleanup_status = 'warning'

            result_data.append({
                "row": idx + 2,
                "account": account_name,
                "login_id": login_id,
                "market": market,
                "platform": platform,
                "usage": usage,
                "owner": owner,
                "count": product_count,
                "status": "normal",
                "last_cleanup_date": last_cleanup_date,
                "days_since_cleanup": days_since_cleanup,
                "cleanup_status": cleanup_status
            })
        
        # 5. ë§ˆì¼“ìƒíƒœí˜„í™© ì‹œíŠ¸ì—ì„œ ìƒíƒœ ì •ë³´ ê°€ì ¸ì˜¤ê¸°
        try:
            ws_status = gsheet.sheet.worksheet(MARKET_STATUS_TAB)
            status_records = ws_status.get_all_records()
            status_map = {}
            for row in status_records:
                store = row.get("ìŠ¤í† ì–´ëª…", "")
                plat = row.get("í”Œë«í¼", "")
                if store and plat:
                    status_map[f"{store}_{plat}"] = {
                        "status": row.get("ìƒíƒœ", "ì •ìƒ"),
                        "note": row.get("ë¹„ê³ ", "")
                    }
            
            # ìƒíƒœ ì ìš©
            for item in result_data:
                key = f"{item['account']}_{item['market']}"
                if key in status_map:
                    status_info = status_map[key]

                    status = status_info["status"]
                    item["note"] = status_info["note"]
                    if status == "ì •ì§€": item["status"] = "stopped"
                    elif status == "ì¼ì‹œì •ì§€": item["status"] = "suspended"
                    elif status == "ê²½ê³ ": item["status"] = "warning"
                    elif status == "ì£¼ì˜": item["status"] = "caution"
        except Exception as e:
            print(f"[ê´€ì œì„¼í„°] ë§ˆì¼“ìƒíƒœí˜„í™© ì¡°íšŒ ì˜¤ë¥˜: {e}")
        
        return {
            "success": True,
            "data": result_data,
            "markets": sorted(list(markets_set)),
            "usages": sorted(list(usages_set))
        }
        
    except Exception as e:
        print(f"[ê´€ì œì„¼í„°] ì¡°íšŒ ì˜¤ë¥˜: {e}")
        import traceback
        traceback.print_exc()
        return {"success": False, "message": str(e), "data": [], "markets": [], "usages": []}


# ========== ë§ˆì¼“ìƒíƒœí˜„í™© API ==========
MARKET_STATUS_TAB = "ë§ˆì¼“ìƒíƒœí˜„í™©"
MARKET_STATUS_HEADERS = ["ìŠ¤í† ì–´ëª…", "í”Œë«í¼", "ìƒíƒœ", "ë³€ê²½ì¼ì‹œ", "ë¹„ê³ "]

def get_or_create_market_status_sheet():
    """ë§ˆì¼“ìƒíƒœí˜„í™© ì‹œíŠ¸ ê°€ì ¸ì˜¤ê¸° (ì—†ìœ¼ë©´ ìƒì„±)"""
    try:
        ws = gsheet.sheet.worksheet(MARKET_STATUS_TAB)
        return ws
    except:
        # ì‹œíŠ¸ ìƒì„±
        ws = gsheet.sheet.add_worksheet(title=MARKET_STATUS_TAB, rows=500, cols=len(MARKET_STATUS_HEADERS))
        ws.update('A1:E1', [MARKET_STATUS_HEADERS])
        # í—¤ë” ì„œì‹ (êµµê²Œ)
        ws.format('A1:E1', {'textFormat': {'bold': True}})
        print(f"âœ… '{MARKET_STATUS_TAB}' ì‹œíŠ¸ ìƒì„±ë¨")
        return ws

@app.get("/api/market-status")
async def get_market_status(request: Request):
    """ë§ˆì¼“ìƒíƒœí˜„í™© ì¡°íšŒ"""
    get_current_user(request)
    
    try:
        ws = get_or_create_market_status_sheet()
        records = ws.get_all_records()
        
        # {ìŠ¤í† ì–´ëª…_í”Œë«í¼: ìƒíƒœ} ë§µ ìƒì„±
        status_map = {}
        for row in records:
            store = row.get("ìŠ¤í† ì–´ëª…", "")
            platform = row.get("í”Œë«í¼", "")
            status = row.get("ìƒíƒœ", "ì •ìƒ")
            if store and platform:
                key = f"{store}_{platform}"
                status_map[key] = {
                    "status": status,
                    "updated_at": row.get("ë³€ê²½ì¼ì‹œ", ""),
                    "note": row.get("ë¹„ê³ ", "")
                }
        
        return {"success": True, "data": status_map}
    except Exception as e:
        print(f"[ë§ˆì¼“ìƒíƒœ] ì¡°íšŒ ì˜¤ë¥˜: {e}")
        return {"success": False, "message": str(e), "data": {}}


class MarketStatusUpdateRequest(BaseModel):
    store_name: str
    platform: str
    status: str
    note: Optional[str] = ""

@app.post("/api/market-status/update")
async def update_market_status(request: Request, req: MarketStatusUpdateRequest):
    """ë§ˆì¼“ìƒíƒœ ì—…ë°ì´íŠ¸ (ì¶”ê°€ ë˜ëŠ” ìˆ˜ì •)"""
    require_permission(request, "edit")
    
    try:
        ws = get_or_create_market_status_sheet()
        records = ws.get_all_records()
        
        now = datetime.now().strftime("%Y-%m-%d %H:%M")
        
        # ê¸°ì¡´ í–‰ ì°¾ê¸°
        row_idx = None
        for idx, row in enumerate(records):
            if row.get("ìŠ¤í† ì–´ëª…") == req.store_name and row.get("í”Œë«í¼") == req.platform:
                row_idx = idx + 2  # í—¤ë” + 0-based index
                break
        
        if req.status == "ì •ìƒ":
            # ì •ìƒì´ë©´ í–‰ ì‚­ì œ (ì €ì¥í•  í•„ìš” ì—†ìŒ)
            if row_idx:
                ws.delete_rows(row_idx)
                print(f"[ë§ˆì¼“ìƒíƒœ] ì‚­ì œ: {req.store_name} ({req.platform})")
            return {"success": True, "action": "deleted"}
        else:
            if row_idx:
                # ê¸°ì¡´ í–‰ ì—…ë°ì´íŠ¸
                ws.update(f'C{row_idx}:E{row_idx}', [[req.status, now, req.note or ""]])
                print(f"[ë§ˆì¼“ìƒíƒœ] ì—…ë°ì´íŠ¸: {req.store_name} ({req.platform}) â†’ {req.status}")
            else:
                # ìƒˆ í–‰ ì¶”ê°€
                ws.append_row([req.store_name, req.platform, req.status, now, req.note or ""])
                print(f"[ë§ˆì¼“ìƒíƒœ] ì¶”ê°€: {req.store_name} ({req.platform}) â†’ {req.status}")
            return {"success": True, "action": "updated"}
            
    except Exception as e:
        print(f"[ë§ˆì¼“ìƒíƒœ] ì—…ë°ì´íŠ¸ ì˜¤ë¥˜: {e}")
        return {"success": False, "message": str(e)}


@app.post("/api/market-status/bulk-update")
async def bulk_update_market_status(request: Request, items: List[MarketStatusUpdateRequest]):
    """ë§ˆì¼“ìƒíƒœ ì¼ê´„ ì—…ë°ì´íŠ¸"""
    require_permission(request, "edit")
    
    try:
        ws = get_or_create_market_status_sheet()
        records = ws.get_all_records()
        now = datetime.now().strftime("%Y-%m-%d %H:%M")
        
        # ê¸°ì¡´ ë°ì´í„° ë§µ
        existing = {}
        for idx, row in enumerate(records):
            key = f"{row.get('ìŠ¤í† ì–´ëª…')}_{row.get('í”Œë«í¼')}"
            existing[key] = idx + 2
        
        updates = 0
        for item in items:
            key = f"{item.store_name}_{item.platform}"
            row_idx = existing.get(key)
            
            if item.status == "ì •ìƒ":
                if row_idx:
                    ws.delete_rows(row_idx)
                    # ì¸ë±ìŠ¤ ì¬ì¡°ì •
                    existing = {k: v-1 if v > row_idx else v for k, v in existing.items()}
                    updates += 1
            else:
                if row_idx:
                    ws.update(f'C{row_idx}:E{row_idx}', [[item.status, now, item.note or ""]])
                else:
                    ws.append_row([item.store_name, item.platform, item.status, now, item.note or ""])
                updates += 1
        
        return {"success": True, "updated": updates}
    except Exception as e:
        print(f"[ë§ˆì¼“ìƒíƒœ] ì¼ê´„ ì—…ë°ì´íŠ¸ ì˜¤ë¥˜: {e}")
        return {"success": False, "message": str(e)}


# ì¼ì¼ì¥ë¶€ ìƒíƒœ ì—…ë°ì´íŠ¸
class DailyStatusUpdateRequest(BaseModel):
    row: int
    column: str
    value: str

@app.post("/api/monitor/daily-status/update")
async def update_daily_status(request: Request, req: DailyStatusUpdateRequest):
    """ì¼ì¼ì¥ë¶€ ì…€ ê°’ ì—…ë°ì´íŠ¸"""
    require_permission(request, "edit")
    
    try:
        # ë°˜ëŒ€ëŸ‰ ì—…ë¡œë“œ í˜„í™© ì‹œíŠ¸ì˜ 12ì›” íƒ­
        upload_sheet = gsheet.client.open_by_key("1MHhu1GdvV1OGS8Wy3NxWOKuqFvgZpqgwn08kG70EDsY")
        ws = upload_sheet.worksheet("12ì›”")
        headers = ws.row_values(1)
        
        # ì»¬ëŸ¼ ì¸ë±ìŠ¤ ì°¾ê¸°
        col_idx = None
        for idx, h in enumerate(headers):
            if h.strip() == req.column:
                col_idx = idx + 1
                break
        
        if col_idx is None:
            return {"success": False, "message": f"ì»¬ëŸ¼ '{req.column}' ì—†ìŒ"}
        
        ws.update_cell(req.row, col_idx, req.value)
        return {"success": True}
        
    except Exception as e:
        print(f"[ê´€ì œì„¼í„°] ìƒíƒœ ì—…ë°ì´íŠ¸ ì˜¤ë¥˜: {e}")
        return {"success": False, "message": str(e)}


class MarketExportRequest(BaseModel):
    headers: List[str]
    data: List[List[Any]]

@app.post("/api/market-table/export")
async def export_market_table(request: Request, req: MarketExportRequest):
    """ë§ˆì¼“í˜„í™© ë°ì´í„°ë¥¼ ì—‘ì…€ë¡œ ë‚´ë³´ë‚´ê¸°"""
    get_current_user(request) # ë¡œê·¸ì¸ ì²´í¬
    
    try:
        # ë°ì´í„°í”„ë ˆì„ ìƒì„±
        df = pd.DataFrame(req.data, columns=req.headers)
        
        # ë©”ëª¨ë¦¬ ë²„í¼ì— ì—‘ì…€ ì“°ê¸°
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='ë§ˆì¼“í˜„í™©')
        
        output.seek(0)
        
        headers = {
            'Content-Disposition': 'attachment; filename="market_status.xlsx"'
        }
        
        return StreamingResponse(
            output, 
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers=headers
        )
        
    except Exception as e:
        print(f"[ë§ˆì¼“ìƒíƒœ] ì—‘ì…€ ë‚´ë³´ë‚´ê¸° ì˜¤ë¥˜: {e}")
        return JSONResponse(status_code=500, content={"success": False, "message": str(e)})


@app.get("/api/monitor/accounts")
async def get_monitor_accounts(request: Request):
    """ê´€ì œì„¼í„°ìš© ê³„ì • ëª©ë¡ (ìƒíƒœ ì •ë³´ í¬í•¨)"""
    get_current_user(request)

    try:
        accounts = gsheet.get_accounts()

        # monitor ì‹œíŠ¸ì—ì„œ ìƒíƒœ ì •ë³´ ê°€ì ¸ì˜¤ê¸°
        try:
            ws_monitor = gsheet.sheet.worksheet("monitor")
            monitor_data = ws_monitor.get_all_records()
            # í•œê¸€/ì˜ì–´ í‚¤ ëª¨ë‘ ì§€ì›
            monitor_map = {}
            for m in monitor_data:
                platform = m.get('í”Œë«í¼') or m.get('platform') or ''
                login_id = m.get('ì•„ì´ë””') or m.get('login_id') or ''
                key = f"{platform}_{login_id}"
                monitor_map[key] = m
        except:
            monitor_map = {}

        # ì‘ì—…ë¡œê·¸ì—ì„œ ê³„ì •ë³„ ë§ˆì§€ë§‰ ì‘ì—…ì¼ ì§‘ê³„
        last_work_map = {}  # "ìŠ¤í† ì–´ëª…": {"date": "2026-01-13", "days": 5}
        try:
            ws_worklog = gsheet.sheet.worksheet("ì‘ì—…ë¡œê·¸")
            worklog_data = ws_worklog.get_all_values()
            if worklog_data and len(worklog_data) > 1:
                headers = worklog_data[0]
                date_idx = next((i for i, h in enumerate(headers) if 'ì¼ì‹œ' in h or 'ë‚ ì§œ' in h), 0)
                type_idx = next((i for i, h in enumerate(headers) if 'ì‘ì—…' in h and 'ìœ í˜•' in h), 1)
                account_idx = next((i for i, h in enumerate(headers) if 'ê³„ì •' in h or 'ìŠ¤í† ì–´' in h), 2)

                today = datetime.now().date()
                for row in worklog_data[1:]:
                    if len(row) > max(date_idx, account_idx):
                        work_type = row[type_idx] if len(row) > type_idx else ""
                        # ì‚­ì œ ì‘ì—…ë§Œ ì¶”ì  (í•„ìš”ì‹œ ë‹¤ë¥¸ ì‘ì—…ë„ ì¶”ê°€ ê°€ëŠ¥)
                        if "ì‚­ì œ" not in work_type:
                            continue

                        date_str = row[date_idx].strip()
                        account_name = row[account_idx].strip()
                        if not date_str or not account_name:
                            continue

                        # ë‚ ì§œ íŒŒì‹± (ì—¬ëŸ¬ í˜•ì‹ ì§€ì›)
                        work_date = None
                        for fmt in ["%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d"]:
                            try:
                                work_date = datetime.strptime(date_str.split()[0], fmt.split()[0]).date()
                                break
                            except:
                                continue

                        if work_date:
                            # ê¸°ì¡´ ê¸°ë¡ë³´ë‹¤ ìµœì‹ ì´ë©´ ì—…ë°ì´íŠ¸
                            if account_name not in last_work_map or work_date > datetime.strptime(last_work_map[account_name]["date"], "%Y-%m-%d").date():
                                days_ago = (today - work_date).days
                                last_work_map[account_name] = {
                                    "date": work_date.strftime("%Y-%m-%d"),
                                    "days": days_ago
                                }
        except Exception as e:
            print(f"[ê´€ì œì„¼í„°] ì‘ì—…ë¡œê·¸ ì¡°íšŒ ì˜¤ë¥˜: {e}")

        # ë§ˆì§€ë§‰ë“±ë¡ì¼ ë°ì´í„° ì¡°íšŒ (ë“±ë¡ê°¯ìˆ˜ ì‹œíŠ¸ - ìŠ¤ë§ˆíŠ¸ìŠ¤í† ì–´)
        last_reg_map = {}  # "ìŠ¤í† ì–´ëª…": {"date": "2026-01-13", "days": 5}
        today = datetime.now().date()
        try:
            ws_counts = gsheet.sheet.worksheet("ë“±ë¡ê°¯ìˆ˜")
            counts_data = ws_counts.get_all_values()
            if counts_data and len(counts_data) > 1:
                headers = counts_data[0]
                name_idx = None
                reg_idx = None
                print(f"[ê´€ì œì„¼í„°] ë“±ë¡ê°¯ìˆ˜ ì‹œíŠ¸ í—¤ë”: {headers}")
                for i, h in enumerate(headers):
                    hl = h.strip().replace(" ", "")
                    if h in ["store_name", "ìŠ¤í† ì–´ëª…"]:
                        name_idx = i
                    if "ë§ˆì§€ë§‰" in hl and "ë“±ë¡" in hl:
                        reg_idx = i
                        print(f"[ê´€ì œì„¼í„°] ë“±ë¡ê°¯ìˆ˜ ë§ˆì§€ë§‰ë“±ë¡ì¼ ì»¬ëŸ¼: idx={i}, name='{h}'")

                if name_idx is not None and reg_idx is not None:
                    num_headers = len(headers)
                    for row in counts_data[1:]:
                        while len(row) < num_headers:
                            row.append('')
                        store_name = row[name_idx].strip()
                        reg_date_str = row[reg_idx].strip()
                        if store_name and reg_date_str:
                            try:
                                reg_date = datetime.strptime(reg_date_str[:10], "%Y-%m-%d").date()
                                days_ago = (today - reg_date).days
                                last_reg_map[f"ìŠ¤ë§ˆíŠ¸ìŠ¤í† ì–´_{store_name}"] = {"date": reg_date_str[:10], "days": days_ago}
                            except:
                                pass
                    print(f"[ê´€ì œì„¼í„°] ë“±ë¡ê°¯ìˆ˜ ì‹œíŠ¸ì—ì„œ ë§ˆì§€ë§‰ë“±ë¡ì¼ {len(last_reg_map)}ê°œ ë¡œë“œ")
                else:
                    print(f"[ê´€ì œì„¼í„°] ë“±ë¡ê°¯ìˆ˜ ì‹œíŠ¸ ì»¬ëŸ¼ ëª»ì°¾ìŒ: name_idx={name_idx}, reg_idx={reg_idx}")
        except Exception as e:
            print(f"[ê´€ì œì„¼í„°] ë“±ë¡ê°¯ìˆ˜ ì‹œíŠ¸ ë§ˆì§€ë§‰ë“±ë¡ì¼ ì¡°íšŒ ì˜¤ë¥˜: {e}")

        # ë§ˆì§€ë§‰ë“±ë¡ì¼ ë°ì´í„° ì¡°íšŒ (11ë²ˆê°€ ì‹œíŠ¸)
        try:
            ws_11st = gsheet.sheet.worksheet("11ë²ˆê°€")
            st_data = ws_11st.get_all_values()
            if st_data and len(st_data) > 1:
                headers = st_data[0]
                name_idx = None
                reg_idx = None
                for i, h in enumerate(headers):
                    hl = h.strip().replace(" ", "")
                    if h in ["store_name", "ì‡¼í•‘ëª° ë³„ì¹­", "ìŠ¤í† ì–´ëª…"]:
                        name_idx = i
                    if "ë§ˆì§€ë§‰" in hl and "ë“±ë¡" in hl:
                        reg_idx = i

                if name_idx is not None and reg_idx is not None:
                    num_headers = len(headers)
                    for row in st_data[1:]:
                        while len(row) < num_headers:
                            row.append('')
                        store_name = row[name_idx].strip()
                        reg_date_str = row[reg_idx].strip()
                        if store_name and reg_date_str:
                            try:
                                reg_date = datetime.strptime(reg_date_str[:10], "%Y-%m-%d").date()
                                days_ago = (today - reg_date).days
                                last_reg_map[f"11ë²ˆê°€_{store_name}"] = {"date": reg_date_str[:10], "days": days_ago}
                            except:
                                pass
                    print(f"[ê´€ì œì„¼í„°] 11ë²ˆê°€ ì‹œíŠ¸ í¬í•¨ ë§ˆì§€ë§‰ë“±ë¡ì¼ ì´ {len(last_reg_map)}ê°œ ë¡œë“œ")
        except Exception as e:
            print(f"[ê´€ì œì„¼í„°] 11ë²ˆê°€ ì‹œíŠ¸ ë§ˆì§€ë§‰ë“±ë¡ì¼ ì¡°íšŒ ì˜¤ë¥˜: {e}")

        # ê³„ì •ì— ìƒíƒœ ì •ë³´ ë³‘í•©
        for acc in accounts:
            platform = acc.get('í”Œë«í¼') or acc.get('platform') or ''
            login_id = acc.get('ì•„ì´ë””') or acc.get('login_id') or ''
            key = f"{platform}_{login_id}"
            if key in monitor_map:
                acc['monitor_status'] = monitor_map[key].get('status', 'green')
                acc['warning_count'] = monitor_map[key].get('warning_count', 0)
                acc['caution_count'] = monitor_map[key].get('caution_count', 0)
                acc['suspend_count'] = monitor_map[key].get('suspend_count', 0)
                acc['memo'] = monitor_map[key].get('memo', '')
            else:
                acc['monitor_status'] = 'green'
                acc['warning_count'] = 0
                acc['caution_count'] = 0
                acc['suspend_count'] = 0
                acc['memo'] = ''
            
            # owner, optype ê¸°ë³¸ê°’ (ì‹œíŠ¸ì— ì—†ìœ¼ë©´)
            if 'owner' not in acc or not acc['owner']:
                acc['owner'] = acc.get('owner', '')
            if 'optype' not in acc or not acc['optype']:
                acc['optype'] = acc.get('optype', 'ëŒ€ëŸ‰')
            
            # í†µê³„ í•„ë“œ (ì¶”í›„ ì—°ë™)
            acc['product_count'] = acc.get('product_count', 0)
            acc['total_sales'] = acc.get('total_sales', 0)
            acc['order_count'] = acc.get('order_count', 0)

            # ë§ˆì§€ë§‰ë“±ë¡ì¼ ì •ë³´ ë³‘í•© (ë“±ë¡ê°¯ìˆ˜/11ë²ˆê°€ ì‹œíŠ¸ ê¸°ì¤€) - í”Œë«í¼+ìŠ¤í† ì–´ëª… AND ì¡°ê±´
            store_name = acc.get('ìŠ¤í† ì–´ëª…') or acc.get('store_name') or ''
            reg_key = f"{platform}_{store_name}"
            if store_name == "ëª¨ìŒìƒì‚¬":
                print(f"[DEBUG] ëª¨ìŒìƒì‚¬: platform={platform}, reg_key={reg_key}, in_map={reg_key in last_reg_map}")
                print(f"[DEBUG] last_reg_map keys sample: {list(last_reg_map.keys())[:10]}")
            if reg_key in last_reg_map:
                # ë§ˆì§€ë§‰ë“±ë¡ì¼ ê¸°ì¤€ ê²½ê³¼ì¼
                acc['last_cleanup_date'] = last_reg_map[reg_key]['date']
                acc['days_since_cleanup'] = last_reg_map[reg_key]['days']
            elif store_name in last_work_map:
                # ë§ˆì§€ë§‰ë“±ë¡ì¼ì´ ì—†ìœ¼ë©´ ì‘ì—…ë¡œê·¸ ê¸°ì¤€ (fallback)
                acc['last_cleanup_date'] = last_work_map[store_name]['date']
                acc['days_since_cleanup'] = last_work_map[store_name]['days']
            else:
                # ë‘˜ ë‹¤ ì—†ìœ¼ë©´ ì˜¤ëŠ˜ ë‚ ì§œë¡œ (ì‹ ê·œ ê³„ì • ì·¨ê¸‰)
                acc['last_cleanup_date'] = datetime.now().strftime("%Y-%m-%d")
                acc['days_since_cleanup'] = 0

            # ê²½ê³¼ì¼ ìƒíƒœ (30ì¼ ì´ˆê³¼: urgent, 20~30ì¼: warning, 20ì¼ ì´ë‚´: normal)
            days = acc['days_since_cleanup']
            if days > 30:
                acc['cleanup_status'] = 'urgent'
            elif days > 20:
                acc['cleanup_status'] = 'warning'
            else:
                acc['cleanup_status'] = 'normal'

        return {"accounts": accounts}
    except Exception as e:
        print(f"[ê´€ì œì„¼í„°] ì˜¤ë¥˜: {e}")
        return {"accounts": []}

# ë§ˆì¼“í˜„í™©ìš© íŒë§¤ì¤‘ ìˆ˜ëŸ‰ API
@app.get("/api/monitor/product-counts")
async def get_product_counts(request: Request):
    """ë“±ë¡ê°¯ìˆ˜/11ë²ˆê°€/ESMíŒë§¤ì¤‘ ì‹œíŠ¸ì—ì„œ íŒë§¤ì¤‘ ìˆ˜ëŸ‰ + ë§ˆì§€ë§‰ë“±ë¡ì¼ ì¡°íšŒ"""
    get_current_user(request)

    try:
        result = {}  # "ìŠ¤í† ì–´ëª…_í”Œë«í¼": {"count": ìˆ˜ëŸ‰, "last_reg": "YYYY-MM-DD"}

        # 1. ë“±ë¡ê°¯ìˆ˜ ì‹œíŠ¸ (ìŠ¤ë§ˆíŠ¸ìŠ¤í† ì–´)
        try:
            ws_counts = gsheet.sheet.worksheet("ë“±ë¡ê°¯ìˆ˜")
            counts_data = ws_counts.get_all_values()
            if counts_data and len(counts_data) > 1:
                headers = counts_data[0]
                name_idx = None
                count_idx = None
                last_reg_idx = None
                print(f"[product-counts] ë“±ë¡ê°¯ìˆ˜ ì‹œíŠ¸ ì „ì²´ í—¤ë”: {headers}")
                for i, h in enumerate(headers):
                    hl = h.strip().lower().replace(" ", "")
                    if h == "ìŠ¤í† ì–´ëª…":
                        name_idx = i
                    elif h == "íŒë§¤ì¤‘":
                        count_idx = i
                    elif "ë§ˆì§€ë§‰" in hl and "ë“±ë¡" in hl:
                        last_reg_idx = i
                        print(f"[product-counts] ë§ˆì§€ë§‰ë“±ë¡ì¼ ì»¬ëŸ¼ ë°œê²¬: idx={i}, name='{h}'")

                print(f"[product-counts] ë“±ë¡ê°¯ìˆ˜ í—¤ë”: name_idx={name_idx}, count_idx={count_idx}, last_reg_idx={last_reg_idx}")

                # ì‹œíŠ¸ ë°ì´í„° ìƒ˜í”Œ ì¶œë ¥ (ì²« 3í–‰)
                if len(counts_data) > 1:
                    for i, row in enumerate(counts_data[1:4]):
                        last_val = row[last_reg_idx] if last_reg_idx is not None and last_reg_idx < len(row) else "N/A"
                        print(f"[product-counts] ë“±ë¡ê°¯ìˆ˜ ìƒ˜í”Œí–‰{i+1}: í–‰ê¸¸ì´={len(row)}, last_reg_idx={last_reg_idx}, last_regê°’='{last_val}'")

                if name_idx is not None and count_idx is not None:
                    sample_count = 0
                    num_headers = len(headers)
                    for row in counts_data[1:]:
                        # í–‰ ê¸¸ì´ë¥¼ í—¤ë” ê¸¸ì´ì— ë§ì¶¤ (ë¹ˆ ì…€ íŒ¨ë”©)
                        while len(row) < num_headers:
                            row.append('')

                        if len(row) > max(name_idx, count_idx):
                            store = row[name_idx].strip()
                            try:
                                cnt = int(row[count_idx]) if row[count_idx] else 0
                            except:
                                cnt = 0
                            last_reg = ""
                            if last_reg_idx is not None:
                                last_reg = row[last_reg_idx].strip()
                            if store:
                                result[f"{store}_ìŠ¤ë§ˆíŠ¸ìŠ¤í† ì–´"] = {"count": cnt, "last_reg": last_reg}
                                if sample_count < 3:
                                    print(f"[product-counts] ìƒ˜í”Œ: {store} -> count={cnt}, last_reg={last_reg}")
                                    sample_count += 1
        except Exception as e:
            print(f"[product-counts] ë“±ë¡ê°¯ìˆ˜ ì‹œíŠ¸ ì˜¤ë¥˜: {e}")

        # 2. 11ë²ˆê°€ ì‹œíŠ¸
        try:
            ws_11st = gsheet.sheet.worksheet("11ë²ˆê°€")
            st_data = ws_11st.get_all_values()
            if st_data and len(st_data) > 1:
                headers = st_data[0]
                print(f"[product-counts] 11ë²ˆê°€ ì‹œíŠ¸ ì „ì²´ í—¤ë”: {headers}")
                name_idx = None
                count_idx = None
                last_reg_idx = None
                for i, h in enumerate(headers):
                    hl = h.strip().lower().replace(" ", "")
                    if h in ["store_name", "ì‡¼í•‘ëª° ë³„ì¹­", "ìŠ¤í† ì–´ëª…"]:
                        name_idx = i
                    elif h == "íŒë§¤ì¤‘":
                        count_idx = i
                    elif "ë§ˆì§€ë§‰" in hl and "ë“±ë¡" in hl:
                        last_reg_idx = i
                        print(f"[product-counts] 11ë²ˆê°€ ë§ˆì§€ë§‰ë“±ë¡ì¼ ì»¬ëŸ¼ ë°œê²¬: idx={i}, name='{h}'")

                print(f"[product-counts] 11ë²ˆê°€ ì¸ë±ìŠ¤: name={name_idx}, count={count_idx}, last_reg={last_reg_idx}")
                if name_idx is not None and count_idx is not None:
                    num_headers = len(headers)
                    for row in st_data[1:]:
                        # í–‰ ê¸¸ì´ë¥¼ í—¤ë” ê¸¸ì´ì— ë§ì¶¤ (ë¹ˆ ì…€ íŒ¨ë”©)
                        while len(row) < num_headers:
                            row.append('')

                        if len(row) > max(name_idx, count_idx):
                            store = row[name_idx].strip()
                            try:
                                cnt = int(row[count_idx]) if row[count_idx] else 0
                            except:
                                cnt = 0
                            last_reg = ""
                            if last_reg_idx is not None:
                                last_reg = row[last_reg_idx].strip()
                            if store:
                                result[f"{store}_11ë²ˆê°€"] = {"count": cnt, "last_reg": last_reg}
        except Exception as e:
            print(f"[product-counts] 11ë²ˆê°€ ì‹œíŠ¸ ì˜¤ë¥˜: {e}")
        
        # 3. ESMíŒë§¤ì¤‘ ì‹œíŠ¸ (ì§€ë§ˆì¼“/ì˜¥ì…˜)
        try:
            ws_esm = gsheet.sheet.worksheet("ESMíŒë§¤ì¤‘")
            esm_data = ws_esm.get_all_values()
            if esm_data and len(esm_data) > 1:
                headers = esm_data[0]
                name_idx = None
                platform_idx = None
                count_idx = None
                for i, h in enumerate(headers):
                    if h == "ìŠ¤í† ì–´ëª…":
                        name_idx = i
                    elif h == "platform":
                        platform_idx = i
                    elif h == "product_count":
                        count_idx = i
                
                if name_idx is not None and count_idx is not None:
                    for row in esm_data[1:]:
                        if len(row) > max(name_idx, count_idx):
                            store = row[name_idx].strip()
                            platform = row[platform_idx].strip() if platform_idx is not None and platform_idx < len(row) else ""
                            try:
                                cnt = int(row[count_idx]) if row[count_idx] else 0
                            except:
                                cnt = 0
                            if store and platform:
                                result[f"{store}_{platform}"] = cnt
        except Exception as e:
            print(f"[product-counts] ESMíŒë§¤ì¤‘ ì‹œíŠ¸ ì˜¤ë¥˜ (ì‹œíŠ¸ ì—†ì„ ìˆ˜ ìˆìŒ): {e}")
        
        # ë””ë²„ê·¸: last_reg ìˆëŠ” ë°ì´í„° ê°œìˆ˜ ì¶œë ¥
        with_last_reg = [k for k, v in result.items() if isinstance(v, dict) and v.get('last_reg')]
        print(f"[product-counts] ì´ {len(result)}ê°œ ì¤‘ last_reg ìˆìŒ: {len(with_last_reg)}ê°œ")
        if with_last_reg[:5]:
            print(f"[product-counts] last_reg ìƒ˜í”Œ: {[(k, result[k]) for k in with_last_reg[:5]]}")

        return {
            "success": True,
            "data": result,
            "debug": {
                "total": len(result),
                "with_last_reg": len(with_last_reg),
                "samples": [(k, result[k]) for k in with_last_reg[:3]] if with_last_reg else []
            }
        }
    except Exception as e:
        print(f"[product-counts] ì˜¤ë¥˜: {e}")
        return {"success": False, "data": {}, "message": str(e)}


class ProductCountUpdateRequest(BaseModel):
    store_name: str
    platform: str
    count: int


@app.post("/api/market/update-product-count")
async def update_product_count(request: Request, req: ProductCountUpdateRequest):
    """ì§€ë§ˆì¼“/ì˜¥ì…˜ íŒë§¤ì¤‘ ìˆ˜ëŸ‰ ì—…ë°ì´íŠ¸"""
    require_permission(request, "edit")
    
    if req.platform not in ["ì§€ë§ˆì¼“", "ì˜¥ì…˜"]:
        return {"success": False, "message": "ì§€ë§ˆì¼“/ì˜¥ì…˜ë§Œ ìˆ˜ì • ê°€ëŠ¥í•©ë‹ˆë‹¤"}
    
    try:
        # ESMíŒë§¤ì¤‘ ì‹œíŠ¸ ê°€ì ¸ì˜¤ê¸° (ì—†ìœ¼ë©´ ìƒì„±)
        try:
            ws_esm = gsheet.sheet.worksheet("ESMíŒë§¤ì¤‘")
        except:
            # ì‹œíŠ¸ ìƒì„±
            ws_esm = gsheet.sheet.add_worksheet(title="ESMíŒë§¤ì¤‘", rows=500, cols=5)
            ws_esm.append_row(["store_name", "platform", "product_count", "updated_at"])
            print("[ESMíŒë§¤ì¤‘] ì‹œíŠ¸ ìƒì„±ë¨")
        
        # ê¸°ì¡´ ë°ì´í„° í™•ì¸
        all_data = ws_esm.get_all_values()
        headers = all_data[0] if all_data else ["store_name", "platform", "product_count", "updated_at"]
        
        # ì»¬ëŸ¼ ì¸ë±ìŠ¤
        name_idx = headers.index("store_name") if "store_name" in headers else 0
        platform_idx = headers.index("platform") if "platform" in headers else 1
        count_idx = headers.index("product_count") if "product_count" in headers else 2
        
        # ê¸°ì¡´ í–‰ ì°¾ê¸°
        found_row = None
        for i, row in enumerate(all_data[1:], start=2):
            if len(row) > max(name_idx, platform_idx):
                if row[name_idx].strip() == req.store_name and row[platform_idx].strip() == req.platform:
                    found_row = i
                    break
        
        from datetime import datetime
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        if found_row:
            # ê¸°ì¡´ í–‰ ì—…ë°ì´íŠ¸
            ws_esm.update_cell(found_row, count_idx + 1, req.count)
            if "updated_at" in headers:
                ws_esm.update_cell(found_row, headers.index("updated_at") + 1, now)
            print(f"[ESMíŒë§¤ì¤‘] {req.store_name}({req.platform}) ì—…ë°ì´íŠ¸: {req.count}")
        else:
            # ìƒˆ í–‰ ì¶”ê°€
            ws_esm.append_row([req.store_name, req.platform, req.count, now])
            print(f"[ESMíŒë§¤ì¤‘] {req.store_name}({req.platform}) ì¶”ê°€: {req.count}")
        
        return {"success": True, "message": "ì €ì¥ ì™„ë£Œ"}
        
    except Exception as e:
        print(f"[ESMíŒë§¤ì¤‘] ì €ì¥ ì˜¤ë¥˜: {e}")
        return {"success": False, "message": str(e)}

class MonitorUpdateRequest(BaseModel):
    platform: str
    login_id: str
    monitor_status: str = "green"
    warning_count: int = 0
    memo: str = ""

@app.post("/api/monitor/update")
async def update_monitor_status(request: Request, req: MonitorUpdateRequest):
    """ê³„ì • ìƒíƒœ ì—…ë°ì´íŠ¸"""
    require_permission(request, "edit")
    
    try:
        # monitor ì‹œíŠ¸ ê°€ì ¸ì˜¤ê¸° (ì—†ìœ¼ë©´ ìƒì„±)
        try:
            ws_monitor = gsheet.sheet.worksheet("monitor")
        except:
            ws_monitor = gsheet.sheet.add_worksheet(title="monitor", rows=1000, cols=10)
            ws_monitor.append_row(["platform", "login_id", "status", "warning_count", "memo", "updated_at"])
        
        # ê¸°ì¡´ ë°ì´í„° í™•ì¸
        all_data = ws_monitor.get_all_records()
        target_row = None
        for idx, row in enumerate(all_data):
            if row.get("platform") == req.platform and row.get("login_id") == req.login_id:
                target_row = idx + 2
                break
        
        from datetime import datetime
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        if target_row:
            # ì—…ë°ì´íŠ¸
            ws_monitor.update(f"C{target_row}:F{target_row}", 
                            [[req.monitor_status, req.warning_count, req.memo, now]])
        else:
            # ìƒˆë¡œ ì¶”ê°€
            ws_monitor.append_row([req.platform, req.login_id, req.monitor_status, 
                                  req.warning_count, req.memo, now])
        
        return {"success": True}
    except Exception as e:
        print(f"[ê´€ì œì„¼í„°] ì—…ë°ì´íŠ¸ ì˜¤ë¥˜: {e}")
        return {"success": False, "message": str(e)}

# ë¹„ë°€ë²ˆí˜¸ ì—…ë°ì´íŠ¸ API (í™•ì¥í”„ë¡œê·¸ë¨ì—ì„œ ìë™ ë³€ê²½ ì‹œ í˜¸ì¶œ)
class UpdatePasswordRequest(BaseModel):
    platform: str
    login_id: str
    new_password: str

@app.post("/api/update-password")
async def update_password(request: Request, req: UpdatePasswordRequest):
    """í¬ë¡¬ í™•ì¥ì—ì„œ ë¹„ë°€ë²ˆí˜¸ ë³€ê²½ ì‹œ í˜¸ì¶œ"""
    # API í‚¤ ì¸ì¦ ë˜ëŠ” ì„¸ì…˜ ì¸ì¦
    api_key = request.headers.get("X-API-Key")
    if api_key != "pkonomiautokey2024":
        try:
            get_current_user(request)
        except:
            raise HTTPException(status_code=401, detail="ì¸ì¦ í•„ìš”")
    
    try:
        # ê³„ì • ì‹œíŠ¸ì—ì„œ í•´ë‹¹ ê³„ì • ì°¾ê¸°
        ws = gsheet.sheet.worksheet(ACCOUNTS_TAB)  # "ê³„ì •ëª©ë¡" ì‚¬ìš©
        all_data = ws.get_all_records()
        
        target_row = None
        for idx, row in enumerate(all_data):
            # í”Œë«í¼ê³¼ ì•„ì´ë””ë¡œ ì°¾ê¸°
            row_platform = row.get("í”Œë«í¼") or row.get("platform", "")
            row_login_id = row.get("ì•„ì´ë””") or row.get("login_id", "")
            if row_platform == req.platform and row_login_id == req.login_id:
                target_row = idx + 2  # í—¤ë” + 0-indexed
                break
        
        if not target_row:
            return {"success": False, "message": "ê³„ì •ì„ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤"}
        
        # íŒ¨ìŠ¤ì›Œë“œ ì—´ ì°¾ê¸°
        headers = ws.row_values(1)
        pw_col = None
        for i, h in enumerate(headers):
            if h in ["íŒ¨ìŠ¤ì›Œë“œ", "password", "ë¹„ë°€ë²ˆí˜¸"]:
                pw_col = i + 1
                break
        
        if not pw_col:
            return {"success": False, "message": "íŒ¨ìŠ¤ì›Œë“œ ì—´ì„ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤"}
        
        # ë¹„ë°€ë²ˆí˜¸ ì—…ë°ì´íŠ¸
        ws.update_cell(target_row, pw_col, req.new_password)
        
        print(f"[ë¹„ë°€ë²ˆí˜¸ë³€ê²½] {req.platform}/{req.login_id} â†’ {req.new_password}")
        
        return {"success": True, "message": "ë¹„ë°€ë²ˆí˜¸ê°€ ì—…ë°ì´íŠ¸ë˜ì—ˆìŠµë‹ˆë‹¤"}
        
    except Exception as e:
        print(f"[ë¹„ë°€ë²ˆí˜¸ë³€ê²½] ì˜¤ë¥˜: {e}")
        return {"success": False, "message": str(e)}

@app.post("/api/accounts/update-password")
async def update_account_password(request: Request, req: UpdatePasswordRequest):
    """ë¹„ë°€ë²ˆí˜¸ ìë™ ë³€ê²½ í›„ êµ¬ê¸€ì‹œíŠ¸ ì—…ë°ì´íŠ¸"""
    get_current_user(request)
    
    try:
        # ê³„ì • ì‹œíŠ¸ì—ì„œ í•´ë‹¹ ê³„ì • ì°¾ê¸°
        ws = gsheet.sheet.worksheet("accounts")
        all_data = ws.get_all_records()
        
        target_row = None
        for idx, row in enumerate(all_data):
            if row.get("platform") == req.platform and row.get("login_id") == req.login_id:
                target_row = idx + 2  # í—¤ë” + 0-indexed
                break
        
        if not target_row:
            return {"success": False, "message": "ê³„ì •ì„ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤"}
        
        # password ì—´ ì°¾ê¸°
        headers = ws.row_values(1)
        pw_col = None
        for i, h in enumerate(headers):
            if h.lower() == "password":
                pw_col = i + 1
                break
        
        if not pw_col:
            return {"success": False, "message": "password ì—´ì„ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤"}
        
        # ë¹„ë°€ë²ˆí˜¸ ì—…ë°ì´íŠ¸
        ws.update_cell(target_row, pw_col, req.new_password)
        
        print(f"[ë¹„ë°€ë²ˆí˜¸ë³€ê²½] {req.platform}/{req.login_id} ë¹„ë°€ë²ˆí˜¸ ì—…ë°ì´íŠ¸ ì™„ë£Œ")
        
        await ws_manager.broadcast({"type": "account_update"})
        return {"success": True, "message": "ë¹„ë°€ë²ˆí˜¸ ì—…ë°ì´íŠ¸ ì™„ë£Œ"}
        
    except Exception as e:
        print(f"[ë¹„ë°€ë²ˆí˜¸ë³€ê²½] ì˜¤ë¥˜: {e}")
        return {"success": False, "message": str(e)}

# ìë™ ë¡œê·¸ì¸ API (í´ë¼ì´ì–¸íŠ¸ í”„ë¡œê·¸ë¨ì—ì„œ ì²˜ë¦¬)
class AutoLoginRequest(BaseModel):
    platform: str
    login_id: str

@app.post("/api/auto-login")
async def auto_login(request: Request, req: AutoLoginRequest):
    """ìë™ ë¡œê·¸ì¸ ìš”ì²­ - í´ë¼ì´ì–¸íŠ¸ í”„ë¡œê·¸ë¨ì´ ì²˜ë¦¬"""
    get_current_user(request)
    
    # ê³„ì • ì •ë³´ ê°€ì ¸ì˜¤ê¸°
    accounts = gsheet.get_accounts(req.platform)
    account = None
    for acc in accounts:
        login_id = acc.get("ì•„ì´ë””") or acc.get("login_id") or ""
        if login_id == req.login_id:
            account = acc
            break
    
    if not account:
        return {"success": False, "message": "ê³„ì •ì„ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤"}
    
    platform_config = PLATFORM_CONFIG.get(req.platform)
    if not platform_config:
        return {"success": False, "message": "í”Œë«í¼ ì„¤ì • ì—†ìŒ"}
    
    # í´ë¼ì´ì–¸íŠ¸ ì—°ê²° í™•ì¸
    if not client_status.get("connected"):
        return {"success": False, "message": "í´ë¼ì´ì–¸íŠ¸ í”„ë¡œê·¸ë¨ì´ ì—°ê²°ë˜ì§€ ì•Šì•˜ìŠµë‹ˆë‹¤. í´ë¼ì´ì–¸íŠ¸ë¥¼ ì‹¤í–‰í•´ì£¼ì„¸ìš”."}
    
    # í´ë¼ì´ì–¸íŠ¸ê°€ ì½ì–´ê°ˆ pending ì •ë³´ ì €ì¥
    global pending_login_info
    pending_login_info = {
        "platform": req.platform,
        "login_id": account["login_id"],
        "password": account["password"],
        "url": platform_config["login_url"],
        "timestamp": datetime.now().isoformat()
    }
    
    return {"pending": True, "message": "í´ë¼ì´ì–¸íŠ¸ì—ì„œ ë¡œê·¸ì¸ ì§„í–‰ ì¤‘..."}

# ìë™ ë¡œê·¸ì¸ ëŒ€ê¸° ì •ë³´ (í´ë¼ì´ì–¸íŠ¸ í”„ë¡œê·¸ë¨ìš©)
pending_login_info = {}

# í´ë¼ì´ì–¸íŠ¸ ìƒíƒœ (ì•ì—ì„œ ì •ì˜)
client_status = {"connected": False, "last_ping": None}

class PendingLoginRequest(BaseModel):
    platform: str
    login_id: str
    password: str
    url: str

@app.post("/api/auto-login/pending")
async def set_pending_login(request: Request, req: PendingLoginRequest):
    """í´ë¼ì´ì–¸íŠ¸ í”„ë¡œê·¸ë¨ì´ ì½ì–´ê°ˆ ë¡œê·¸ì¸ ì •ë³´ ì €ì¥"""
    get_current_user(request)
    
    global pending_login_info
    pending_login_info = {
        "platform": req.platform,
        "login_id": req.login_id,
        "password": req.password,
        "url": req.url,
        "timestamp": datetime.now().isoformat()
    }
    
    return {"success": True}

@app.get("/api/auto-login/pending")
async def get_pending_login(request: Request):
    """í´ë¼ì´ì–¸íŠ¸ í”„ë¡œê·¸ë¨ì´ ì½ì–´ê°ˆ ë¡œê·¸ì¸ ì •ë³´ ì¡°íšŒ"""
    # API í‚¤ ì¸ì¦ ë˜ëŠ” ì„¸ì…˜ ì¸ì¦
    api_key = request.headers.get("X-API-Key")
    if api_key != "pkonomiautokey2024":
        try:
            get_current_user(request)
        except:
            raise HTTPException(status_code=401, detail="ì¸ì¦ í•„ìš”")
    
    # í´ë¼ì´ì–¸íŠ¸ ìƒíƒœ ì—…ë°ì´íŠ¸
    client_status["connected"] = True
    client_status["last_ping"] = datetime.now()
    
    global pending_login_info
    if pending_login_info:
        result = pending_login_info.copy()
        result["pending"] = True
        # ì½ì€ í›„ ì‚­ì œ (1íšŒìš©)
        pending_login_info = {}
        return result
    
    return {"platform": None, "pending": False}

# í”Œë«í¼ ì„¤ì •
@app.get("/api/platforms")
async def get_platforms(request: Request):
    get_current_user(request)
    return {"platforms": PLATFORM_CONFIG}

# SMS ë¸Œë¼ìš°ì € ìƒíƒœ
@app.get("/api/sms/status")
async def get_sms_status(request: Request):
    get_current_user(request)
    return {
        "profiles": PHONE_PROFILES,
        "ready": sms_manager.ready,
        "auth_codes": sms_manager.auth_codes
    }

# ì¸ì¦ì½”ë“œ ì¡°íšŒ (í™•ì¥í”„ë¡œê·¸ë¨ìš© - API í‚¤ ì¸ì¦)
@app.get("/api/sms/auth-code")
async def get_auth_code(request: Request, refresh: bool = False):
    """ìµœì‹  ì¸ì¦ì½”ë“œ ë°˜í™˜ - í™•ì¥í”„ë¡œê·¸ë¨/ìˆ˜ì§‘ê¸°ì—ì„œ ì‚¬ìš©"""
    # API í‚¤ ì¸ì¦ ë˜ëŠ” ì„¸ì…˜ ì¸ì¦
    api_key = request.headers.get("X-API-Key")
    if api_key != "pkonomiautokey2024":
        try:
            get_current_user(request)
        except:
            raise HTTPException(status_code=401, detail="ì¸ì¦ í•„ìš”")
    
    # â˜… refresh=trueì´ë©´ ê°•ì œ ìƒˆë¡œê³ ì¹¨ (ì‹¤ì‹œê°„ ë©”ì‹œì§€ ìˆ˜ì§‘ ë³´ì¥)
    if refresh:
        await sms_manager.refresh_messages()

    # â˜… clear_time ê°€ì ¸ì˜¤ê¸° (ì—†ìœ¼ë©´ 0)
    clear_time = getattr(sms_manager, 'auth_code_clear_time', 0)
    
    # ì „ì²´ ë©”ì‹œì§€ì—ì„œ ê°€ì¥ ìµœê·¼ ìˆ˜ì‹  ì‹œê°„ì˜ ì¸ì¦ì½”ë“œ ì°¾ê¸°
    latest_code = None
    latest_time = None
    latest_timestamp = 0  # Unix timestamp
    
    for msg in sms_manager.messages:
        if msg.auth_code and msg.auth_code.isdigit() and len(msg.auth_code) >= 4:
            # timestampë¥¼ Unix timestampë¡œ ë³€í™˜
            msg_timestamp = sms_manager._parse_relative_time(msg.timestamp)
            
            # â˜… clear_time ì´í›„ ë©”ì‹œì§€ë§Œ ì‚¬ìš©
            if msg_timestamp <= clear_time:
                continue
            
            # ê°€ì¥ ìµœê·¼ ë©”ì‹œì§€ ì°¾ê¸°
            if msg_timestamp > latest_timestamp:
                latest_timestamp = msg_timestamp
                latest_code = msg.auth_code
                latest_time = msg.timestamp
    
    return {
        "code": latest_code,
        "time": latest_time,
        "auth_codes": sms_manager.auth_codes  # ì „ì²´ ì¸ì¦ì½”ë“œ ì •ë³´ (ì‹œê°„ í¬í•¨)
    }

# ì¸ì¦ì½”ë“œ ì´ˆê¸°í™” (í™•ì¥í”„ë¡œê·¸ë¨ìš© - API í‚¤ ì¸ì¦)
@app.post("/api/sms/auth-code/clear")
async def clear_auth_code(request: Request):
    """ì¸ì¦ì½”ë“œ ì´ˆê¸°í™” - ìƒˆ ì¸ì¦ì½”ë“œ ëŒ€ê¸° ì „ í˜¸ì¶œ"""
    # API í‚¤ ì¸ì¦
    api_key = request.headers.get("X-API-Key")
    if api_key != "pkonomiautokey2024":
        raise HTTPException(status_code=401, detail="API í‚¤ í•„ìš”")
    
    # â˜… clear_time ê¸°ë¡ (ì´ ì‹œì  ì´í›„ ë©”ì‹œì§€ë§Œ ì‚¬ìš©)
    sms_manager.auth_code_clear_time = time.time()
    
    # ëª¨ë“  ì¸ì¦ì½”ë“œ ì´ˆê¸°í™”
    sms_manager.auth_codes.clear()
    
    return {"success": True, "message": "ì¸ì¦ì½”ë“œ ì´ˆê¸°í™”ë¨"}

# SMS API ë¶„ì„ - ìº¡ì²˜ í™œì„±í™” (ê´€ë¦¬ì ì „ìš©)
@app.post("/api/sms/api-capture/enable")
async def enable_sms_api_capture(request: Request):
    """API ìº¡ì²˜ í™œì„±í™” - ë¸Œë¼ìš°ì € ì¬ì‹œì‘ í•„ìš”"""
    require_permission(request, "admin")
    sms_manager.enable_api_capture(True)
    return {"success": True, "message": "API ìº¡ì²˜ í™œì„±í™”ë¨. SMS ë¸Œë¼ìš°ì €ë¥¼ ì¬ì‹œì‘í•˜ì„¸ìš”."}

@app.post("/api/sms/api-capture/disable")
async def disable_sms_api_capture(request: Request):
    """API ìº¡ì²˜ ë¹„í™œì„±í™”"""
    require_permission(request, "admin")
    sms_manager.enable_api_capture(False)
    return {"success": True, "message": "API ìº¡ì²˜ ë¹„í™œì„±í™”ë¨"}

@app.get("/api/sms/api-capture/logs")
async def get_sms_api_logs(request: Request):
    """ìº¡ì²˜ëœ API ë¡œê·¸ ì¡°íšŒ"""
    require_permission(request, "admin")
    return {
        "enabled": sms_manager.api_capture_enabled,
        "count": len(sms_manager.captured_apis),
        "apis": sms_manager.captured_apis[-100:]  # ìµœê·¼ 100ê°œë§Œ
    }

@app.post("/api/sms/api-capture/save")
async def save_sms_api_logs(request: Request):
    """ìº¡ì²˜ëœ API ë¡œê·¸ë¥¼ íŒŒì¼ë¡œ ì €ì¥"""
    require_permission(request, "admin")
    sms_manager.save_api_log()
    return {"success": True, "file": str(sms_manager.api_log_file), "count": len(sms_manager.captured_apis)}

@app.post("/api/sms/api-capture/clear")
async def clear_sms_api_logs(request: Request):
    """ìº¡ì²˜ëœ API ë¡œê·¸ ì´ˆê¸°í™”"""
    require_permission(request, "admin")
    sms_manager.clear_api_log()
    return {"success": True, "message": "API ë¡œê·¸ ì´ˆê¸°í™”ë¨"}

# SMS ë¸Œë¼ìš°ì € ì‹œì‘ (ìš´ì˜ì ì´ìƒ)
@app.post("/api/sms/launch/{profile_id}")
async def launch_sms_browser(request: Request, profile_id: str):
    require_permission(request, "sms_control")  # ìš´ì˜ì ì´ìƒ
    if profile_id not in PHONE_PROFILES:
        raise HTTPException(status_code=400, detail="ì˜ëª»ëœ í”„ë¡œí•„")
    
    await sms_manager.launch_browser(profile_id)
    await ws_manager.broadcast({"type": "sms_status", "ready": sms_manager.ready})
    return {"success": True, "ready": sms_manager.ready.get(profile_id)}

# SMS ë¸Œë¼ìš°ì € ì „ì²´ ì‹œì‘ (ìš´ì˜ì ì´ìƒ)
@app.post("/api/sms/launch-all")
async def launch_all_sms(request: Request):
    require_permission(request, "sms_control")  # ìš´ì˜ì ì´ìƒ
    await sms_manager.launch_all()
    await ws_manager.broadcast({"type": "sms_status", "ready": sms_manager.ready})
    return {"success": True, "ready": sms_manager.ready}

# SMS ë©”ì‹œì§€ ìƒˆë¡œê³ ì¹¨ (ëª¨ë“  ì‚¬ìš©ì - ë·°ì–´ë„ ê°€ëŠ¥)
# SMS ìƒˆë¡œê³ ì¹¨ ì“°ë¡œí‹€ë§
_sms_last_refresh = None
_sms_refresh_interval = 3  # ìµœì†Œ 3ì´ˆ ê°„ê²©

@app.get("/api/sms/messages")
async def get_sms_messages(request: Request, refresh: bool = False):
    """SMS ë©”ì‹œì§€ ëª©ë¡ - refresh=trueì¼ ë•Œë§Œ ìƒˆë¡œê³ ì¹¨ (3ì´ˆ ì“°ë¡œí‹€ë§)"""
    global _sms_last_refresh
    get_current_user(request)  # ë¡œê·¸ì¸ë§Œ í™•ì¸
    
    # refresh íŒŒë¼ë¯¸í„°ê°€ trueì¼ ë•Œë§Œ ì‹¤ì œë¡œ ìƒˆë¡œê³ ì¹¨
    if refresh:
        now = datetime.now()
        # ë§ˆì§€ë§‰ ìƒˆë¡œê³ ì¹¨ í›„ 3ì´ˆ ì´ë‚´ë©´ ìºì‹œ ë°˜í™˜
        if _sms_last_refresh and (now - _sms_last_refresh).total_seconds() < _sms_refresh_interval:
            pass  # ìºì‹œ ì‚¬ìš©
        else:
            await sms_manager.refresh_messages()
            _sms_last_refresh = now
    
    messages = sms_manager.messages
    
    return {
        "messages": [asdict(m) for m in messages],
        "auth_codes": sms_manager.auth_codes
    }


@app.post("/api/sms/refresh")
async def refresh_sms_messages(request: Request):
    """SMS ë©”ì‹œì§€ ê°•ì œ ìƒˆë¡œê³ ì¹¨"""
    get_current_user(request)
    messages = await sms_manager.refresh_messages()
    return {
        "messages": [asdict(m) for m in messages],
        "auth_codes": sms_manager.auth_codes
    }

@app.post("/api/sms/reload-page")
async def reload_sms_pages(request: Request):
    """êµ¬ê¸€ë©”ì‹œì§€ í˜ì´ì§€ F5 ìƒˆë¡œê³ ì¹¨"""
    require_permission(request, "edit")  # ìš´ì˜ì ì´ìƒ
    reloaded = []
    errors = []
    for profile_id, page in list(sms_manager.pages.items()):
        try:
            if page and not page.is_closed():
                await page.reload(timeout=30000)
                reloaded.append(profile_id)
        except Exception as e:
            errors.append(f"{profile_id}: {str(e)}")
    if reloaded:
        return {"success": True, "message": f"ìƒˆë¡œê³ ì¹¨ ì™„ë£Œ: {', '.join(reloaded)}"}
    elif errors:
        return {"success": False, "message": f"ì˜¤ë¥˜: {', '.join(errors)}"}
    else:
        return {"success": False, "message": "í™œì„±í™”ëœ ë¸Œë¼ìš°ì €ê°€ ì—†ìŠµë‹ˆë‹¤"}

# SMS ì „ì†¡ (ìš´ì˜ì ì´ìƒ)
@app.post("/api/sms/send")
async def send_sms(request: Request, req: SMSRequest):
    require_permission(request, "sms_send")  # ìš´ì˜ì ì´ìƒ
    success = await sms_manager.send_message(req.phone_profile, req.to_number, req.message)
    if success:
        # ì‘ì—… ë¡œê·¸ ê¸°ë¡
        log_work("SMSì „ì†¡", req.phone_profile, 1, f"ìˆ˜ì‹ : {req.to_number}", "ì›¹")
        await ws_manager.broadcast({"type": "sms_sent", "profile": req.phone_profile})
        return {"success": True}
    raise HTTPException(status_code=500, detail="ì „ì†¡ ì‹¤íŒ¨")

# SMS ì „ì†¡ with íŒŒì¼ ì²¨ë¶€ (ìš´ì˜ì ì´ìƒ)
@app.post("/api/sms/send-with-file")
async def send_sms_with_file(
    request: Request,
    phone_profile: str = Form(...),
    to_number: str = Form(...),
    message: str = Form(""),
    file: UploadFile = File(None)
):
    require_permission(request, "sms_send")  # ìš´ì˜ì ì´ìƒ

    file_path = None
    try:
        # íŒŒì¼ì´ ìˆìœ¼ë©´ ì„ì‹œ ì €ì¥
        if file and file.filename:
            upload_dir = APP_DIR / "uploads"
            upload_dir.mkdir(exist_ok=True)

            # íŒŒì¼ëª… ì•ˆì „í•˜ê²Œ ì²˜ë¦¬
            import uuid
            ext = os.path.splitext(file.filename)[1]
            safe_filename = f"{uuid.uuid4().hex}{ext}"
            file_path = str(upload_dir / safe_filename)

            with open(file_path, "wb") as f:
                content = await file.read()
                f.write(content)
            print(f"[SMS] íŒŒì¼ ì—…ë¡œë“œë¨: {file_path} ({len(content)} bytes)")

        # ë©”ì‹œì§€ ì „ì†¡
        success = await sms_manager.send_message(phone_profile, to_number, message, file_path)

        if success:
            # ì‘ì—… ë¡œê·¸ ê¸°ë¡
            log_work("SMSì „ì†¡", phone_profile, 1, f"ìˆ˜ì‹ : {to_number}" + (" (íŒŒì¼ì²¨ë¶€)" if file_path else ""), "ì›¹")
            await ws_manager.broadcast({"type": "sms_sent", "profile": phone_profile})
            return {"success": True, "message": "ì „ì†¡ ì™„ë£Œ"}
        raise HTTPException(status_code=500, detail="ì „ì†¡ ì‹¤íŒ¨")

    finally:
        # ì„ì‹œ íŒŒì¼ ì‚­ì œ
        if file_path and os.path.exists(file_path):
            try:
                os.remove(file_path)
            except:
                pass

# SMS ëŒ€í™” ìƒì„¸ ê°€ì ¸ì˜¤ê¸°
class ConversationDetailRequest(BaseModel):
    profile_id: str
    sender: str
    offset: int = 0  # 0: ìµœê·¼ 20ê°œ, 20: ì´ì „ 20ê°œ, 40: ë” ì´ì „...
    limit: int = 20

@app.post("/api/sms/conversation")
async def get_conversation_detail(request: Request, req: ConversationDetailRequest):
    get_current_user(request)
    # print(f"[ëŒ€í™”ìƒì„¸] ìš”ì²­: profile={req.profile_id}, sender={req.sender}, offset={req.offset}")
    result = await sms_manager.get_conversation_detail(req.profile_id, req.sender, req.offset, req.limit)
    # print(f"[ëŒ€í™”ìƒì„¸] ê²°ê³¼: messages={len(result.get('messages', []))} ê°œ, error={result.get('error', 'None')}")
    return result

# ì „í™”ë²ˆí˜¸ë¡œ ëŒ€í™” ê²€ìƒ‰
class SearchByPhoneRequest(BaseModel):
    profile_id: str
    phone_number: str

@app.post("/api/sms/search")
async def search_by_phone(request: Request, req: SearchByPhoneRequest):
    get_current_user(request)
    result = await sms_manager.search_by_phone(req.profile_id, req.phone_number)
    return result

# ë¯¸ë””ì–´ ë‹¤ìš´ë¡œë“œ
class MediaDownloadRequest(BaseModel):
    profile_id: str
    sender: str
    media_type: str  # image, video, file
    element_idx: str
    get_thumbnail: bool = False  # Trueë©´ ì¸ë„¤ì¼, Falseë©´ ì›ë³¸

@app.post("/api/sms/download")
async def download_media(request: Request, req: MediaDownloadRequest):
    get_current_user(request)
    try:
        filepath = await sms_manager.download_media(req.profile_id, req.sender, req.media_type, req.element_idx, req.get_thumbnail)
        if filepath:
            return {"success": True, "filepath": filepath}
        return {"success": False, "message": "ë‹¤ìš´ë¡œë“œ ì‹¤íŒ¨ ë˜ëŠ” ì´ë¯¸ì§€ ì—†ìŒ"}
    except Exception as e:
        print(f"[ë‹¤ìš´ë¡œë“œ ì˜¤ë¥˜] {e}")
        return {"success": False, "message": str(e)}

# ë‹¤ìš´ë¡œë“œ íŒŒì¼ ì„œë¹™
from fastapi.responses import FileResponse

@app.get("/downloads/{profile_id}/{filename}")
async def serve_download(request: Request, profile_id: str, filename: str):
    get_current_user(request)
    filepath = APP_DIR / "downloads" / profile_id / filename
    if filepath.exists():
        return FileResponse(str(filepath))
    raise HTTPException(status_code=404, detail="íŒŒì¼ ì—†ìŒ")

# WebSocket
@app.websocket("/ws")
async def websocket_endpoint(websocket: WebSocket):
    await ws_manager.connect(websocket)
    try:
        while True:
            data = await websocket.receive_text()
            # í´ë¼ì´ì–¸íŠ¸ë¡œë¶€í„° ë©”ì‹œì§€ ì²˜ë¦¬ (í•„ìš”ì‹œ)
    except WebSocketDisconnect:
        ws_manager.disconnect(websocket)

# ========== ë¶ˆì‚¬ì API ==========
class BulsajaRunRequest(BaseModel):
    groups: str
    max_concurrent: int = 3
    group_gap: int = 60
    # ì„¤ì •ê°’
    program: str = ""
    uploadMarket: str = ""
    uploadCount: str = ""
    deleteCount: str = ""
    copySourceMarket: str = ""
    copyCount: str = ""

@app.post("/api/bulsaja/run")
async def run_bulsaja(request: Request, req: BulsajaRunRequest):
    get_current_user(request)
    print(f"[API] /api/bulsaja/run í˜¸ì¶œ: groups={req.groups}, program={req.program}")

    # ì‘ì—… ë¡œê·¸ ê¸°ë¡
    group_names = ", ".join([str(g) for g in req.groups[:5]]) + ("..." if len(req.groups) > 5 else "")
    log_work(f"ë¶ˆì‚¬ì-{req.program}", "ë¶ˆì‚¬ì", len(req.groups), f"ê·¸ë£¹: {group_names}", "ì›¹")

    # ì„¤ì •ê°’ì„ ë”•ì…”ë„ˆë¦¬ë¡œ ì „ë‹¬
    settings = {
        "program": req.program,
        "uploadMarket": req.uploadMarket,
        "uploadCount": req.uploadCount,
        "deleteCount": req.deleteCount,
        "copySourceMarket": req.copySourceMarket,
        "copyCount": req.copyCount
    }

    result = bulsaja_manager.start(req.groups, req.max_concurrent, req.group_gap, settings)
    print(f"[API] /api/bulsaja/run ê²°ê³¼: {result}")
    return result

@app.post("/api/bulsaja/stop")
async def stop_bulsaja(request: Request):
    get_current_user(request)
    result = bulsaja_manager.stop()
    return result

@app.get("/api/bulsaja/status")
async def get_bulsaja_status(request: Request):
    get_current_user(request)
    return bulsaja_manager.get_status()

@app.post("/api/bulsaja/update_account_info")
async def update_bulsaja_account_info(request: Request):
    """ê³„ì •ë³„ ìš´ì˜ ì •ë³´(ìš´ì˜ì¼ìˆ˜ ë“±) ì—…ë°ì´íŠ¸"""
    get_current_user(request)
    data = await request.json()
    store_name = data.get("name")
    
    if not store_name:
        return {"success": False, "message": "ìŠ¤í† ì–´ëª…ì´ í•„ìš”í•©ë‹ˆë‹¤."}

    # 3. [New] ë§ˆì¼“ìƒíƒœí˜„í™©(ë§¤ì¶œ) ê°€ì ¸ì˜¤ê¸°
    # account_info APIì—ì„œëŠ” ì‚¬ìš©í•˜ì§€ ì•Šì§€ë§Œ ë¡œì§ ì¼ê´€ì„±ì„ ìœ„í•´ ì¶”ê°€í•˜ê±°ë‚˜ ìƒëµ ê°€ëŠ¥
    # ì—¬ê¸°ì„œëŠ” dashboard_data APIë¥¼ ìˆ˜ì •í•´ì•¼ í•¨.
    
    settings_path = os.path.join(os.path.dirname(__file__), "bulsaja_settings.json")
    settings = {}
    if os.path.exists(settings_path):
        try:
            with open(settings_path, 'r', encoding='utf-8') as f:
                settings = json.load(f)
        except: pass
        
    if store_name not in settings:
        settings[store_name] = {}
        
    # ì§€ì›í•˜ëŠ” í•„ë“œ ì—…ë°ì´íŠ¸
    if "operationDays" in data:
        settings[store_name]["operationDays"] = int(data["operationDays"])
    if "inflowDecreaseDays" in data:
        settings[store_name]["inflowDecreaseDays"] = int(data["inflowDecreaseDays"])
    if "revenue30d" in data:
        settings[store_name]["revenue30d"] = int(data["revenue30d"])
        
    try:
        with open(settings_path, 'w', encoding='utf-8') as f:
            json.dump(settings, f, ensure_ascii=False, indent=4)
        return {"success": True}
    except Exception as e:
        return {"success": False, "message": str(e)}

@app.get("/api/bulsaja/dashboard_data_deprecated")
async def get_bulsaja_dashboard_data_deprecated(request: Request, refresh: bool = False):
    """ìë™í™” ëŒ€ì‹œë³´ë“œìš© ì‹¤ì œ ê³„ì • ë°ì´í„° ì¡°íšŒ"""
    get_current_user(request)
    
    try:
        # [2026-01-19 Updated] ë§ˆì¼€íŒ… ë°ì´í„° ë¡œë“œ (ë§¤ì¶œ, ìœ ì…ìˆ˜, ì£¼ë¬¸ìˆ˜, ì´ë ¥ë¶„ì„)
        market_stats = gsheet.get_marketing_data(force_refresh=refresh)
        
        # 1. ë§ˆì¼“ë³„ ìƒí’ˆìˆ˜ ì •ë³´ ê°€ì ¸ì˜¤ê¸° (get_market_summary ë¡œì§ ì¬ì‚¬ìš©)
        ss_counts = {} # ìŠ¤ë§ˆíŠ¸ìŠ¤í† ì–´
        st_counts = {} # 11ë²ˆê°€
        
        # ìŠ¤ë§ˆíŠ¸ìŠ¤í† ì–´ ìƒí’ˆìˆ˜ (ë“±ë¡ê°¯ìˆ˜ ì‹œíŠ¸)
        try:
            ws_counts = gsheet.sheet.worksheet("ë“±ë¡ê°¯ìˆ˜")
            counts_data = ws_counts.get_all_values()
            if counts_data and len(counts_data) > 1:
                headers = counts_data[0]
                name_idx = next((i for i, h in enumerate(headers) if h == "ìŠ¤í† ì–´ëª…"), None)
                count_idx = next((i for i, h in enumerate(headers) if h == "íŒë§¤ì¤‘"), None)
                if name_idx is not None and count_idx is not None:
                    for row in counts_data[1:]:
                        if len(row) > max(name_idx, count_idx):
                            store = row[name_idx].strip()
                            try: cnt = int(row[count_idx]) if row[count_idx] else 0
                            except: cnt = 0
                            if store: ss_counts[store] = cnt
        except: pass

        # 11ë²ˆê°€ ìƒí’ˆìˆ˜ (11ë²ˆê°€ ì‹œíŠ¸)
        try:
            ws_11st = gsheet.sheet.worksheet("11ë²ˆê°€")
            st_data = ws_11st.get_all_values()
            if st_data and len(st_data) > 1:
                headers = st_data[0]
                name_idx = next((i for i, h in enumerate(headers) if h in ["store_name", "ì‡¼í•‘ëª° ë³„ì¹­", "ìŠ¤í† ì–´ëª…"]), None)
                count_idx = next((i for i, h in enumerate(headers) if h in ["on_sale", "íŒë§¤ì¤‘"]), None)
                if name_idx is not None and count_idx is not None:
                    for row in st_data[1:]:
                        if len(row) > max(name_idx, count_idx):
                            store = row[name_idx].strip()
                            try: cnt = int(row[count_idx]) if row[count_idx] else 0
                            except: cnt = 0
                            if store: st_counts[store] = cnt
        except: pass

        # ESMíŒë§¤ì¤‘ ì‹œíŠ¸ì—ì„œ ì§€ë§ˆì¼“/ì˜¥ì…˜ ìƒí’ˆìˆ˜ ê°€ì ¸ì˜¤ê¸°
        esm_counts = {}
        try:
            ws_esm = gsheet.sheet.worksheet("ESMíŒë§¤ì¤‘")
            esm_data = ws_esm.get_all_values()
            if esm_data and len(esm_data) > 1:
                headers = esm_data[0]
                name_idx = next((i for i, h in enumerate(headers) if h == "ìŠ¤í† ì–´ëª…"), None)
                platform_idx = next((i for i, h in enumerate(headers) if h == "platform"), None)
                count_idx = next((i for i, h in enumerate(headers) if h == "product_count"), None)
                if name_idx is not None and count_idx is not None:
                    for row in esm_data[1:]:
                        if len(row) > max(name_idx, count_idx):
                            store = row[name_idx].strip()
                            plat = row[platform_idx].strip() if platform_idx is not None and platform_idx < len(row) else ""
                            try: cnt = int(row[count_idx]) if row[count_idx] else 0
                            except: cnt = 0
                            if store:
                                # ìŠ¤í† ì–´ëª…_í”Œë«í¼ í˜•ì‹ê³¼ ìŠ¤í† ì–´ëª… í˜•ì‹ ë‘˜ ë‹¤ ì €ì¥
                                esm_counts[f"{store}_{plat}"] = cnt
                                # ìŠ¤í† ì–´ëª…ë§Œìœ¼ë¡œë„ ë§¤ì¹­ ê°€ëŠ¥í•˜ê²Œ (ì²« ë²ˆì§¸ ê°’ ìš°ì„ )
                                if store not in esm_counts:
                                    esm_counts[store] = cnt
            print(f"[ESMíŒë§¤ì¤‘] ë¡œë“œ ì™„ë£Œ: {len(esm_counts)}ê°œ ìŠ¤í† ì–´")
        except Exception as e:
            print(f"[ESMíŒë§¤ì¤‘] ë¡œë“œ ì‹¤íŒ¨: {e}")

        # 3. [New] ë§ˆì¼“ìƒíƒœí˜„í™©(ë§¤ì¶œ) ê°€ì ¸ì˜¤ê¸°
        market_status = gsheet.get_market_status(force_refresh=refresh)

        # 2. ê³„ì • ì •ë³´ ê°€ì ¸ì˜¤ê¸°
        accounts = gsheet.get_accounts(None, force_refresh=refresh)
        
        dashboard_accounts = []
        for acc in accounts:
            platform = acc.get("platform", "ê¸°íƒ€")
            store_name = acc.get("ìŠ¤í† ì–´ëª…") or acc.get("login_id", "ê³„ì •")
            
            # ë§ˆì¼“ëª… ì •ê·œí™” ë° ìƒí’ˆìˆ˜ ë§¤ì¹­
            product_count = 0
            mapped_platform = "etc"
            
            if "ìŠ¤ë§ˆíŠ¸" in platform or "ë„¤ì´ë²„" in platform:
                mapped_platform = "naver"
                product_count = ss_counts.get(store_name, 0)
                if product_count == 0 and "_" in store_name:
                    product_count = ss_counts.get(store_name.split("_", 1)[1], 0)
            elif "ì¿ íŒ¡" in platform:
                mapped_platform = "coupang"
                # ì¿ íŒ¡ì€ í˜„ì¬ ë³„ë„ ìƒí’ˆìˆ˜ ì‹œíŠ¸ê°€ ì—†ìœ¼ë¯€ë¡œ ë§ˆì¼€íŒ… ë°ì´í„° ë“±ì—ì„œ ê°€ì ¸ì™€ì•¼ í•˜ë‚˜ ì¼ë‹¨ 0ìœ¼ë¡œ ì‹œì‘
                # (ì¶”í›„ ì¿ íŒ¡ ìƒí’ˆìˆ˜ ìˆ˜ì§‘ ë¡œì§ ì¶”ê°€ í•„ìš”)
            elif "11" in platform:
                mapped_platform = "11st"
                product_count = st_counts.get(store_name, 0) or st_counts.get(acc.get("login_id"), 0)
            elif "ì§€ë§ˆì¼“" in platform or "Gmarket" in platform:
                mapped_platform = "gmarket"
                # ESMíŒë§¤ì¤‘ ì‹œíŠ¸ì—ì„œ ìŠ¤í† ì–´ëª…ìœ¼ë¡œ ë§¤ì¹­
                product_count = esm_counts.get(f"{store_name}_gmarket", 0) or esm_counts.get(store_name, 0)
            elif "ì˜¥ì…˜" in platform or "Auction" in platform:
                mapped_platform = "auction"
                # ESMíŒë§¤ì¤‘ ì‹œíŠ¸ì—ì„œ ìŠ¤í† ì–´ëª…ìœ¼ë¡œ ë§¤ì¹­
                product_count = esm_counts.get(f"{store_name}_auction", 0) or esm_counts.get(store_name, 0)
            elif "ESM" in platform:
                mapped_platform = "esm"

            # 3. ìš´ì˜ ë‹¨ê³„ íŒë³„ ë° ë¦¬ë‰´ì–¼ ë¡œì§ (ì‚¬ìš©ì ìš”ì²­ ê¸°ì¤€ ê°œí¸)
            # ì„¤ì • íŒŒì¼ì—ì„œ ìˆ˜ë™ ì„¤ì •ê°’ ë¡œë“œ
            settings_path = os.path.join(os.path.dirname(__file__), "bulsaja_settings.json")
            acc_settings = {}
            if os.path.exists(settings_path):
                try:
                    with open(settings_path, 'r', encoding='utf-8') as f:
                        all_settings = json.load(f)
                        acc_settings = all_settings.get(store_name, {})
                except: pass

            # 11ë²ˆê°€: 4000ê°œ ì´ìƒì´ë©´ ìš´ì˜ì¤‘, ë‚˜ë¨¸ì§€ëŠ” ë“±ë¡ìœ¨ 90% ê¸°ì¤€
            # ìŠ¤ë§ˆíŠ¸ìŠ¤í† ì–´/ì¿ íŒ¡: 10,000ê°œ * 0.9 = 9,000ê°œ
            # ì§€ë§ˆì¼“/ì˜¥ì…˜: 2,000ê°œ * 0.9 = 1,800ê°œ
            if mapped_platform == "11st":
                # 11ë²ˆê°€: 4000ê°œ ì´ìƒì´ë©´ ìš´ì˜ì¤‘
                is_operating = product_count >= 4000
            else:
                target_products = 10000
                if mapped_platform in ["gmarket", "auction"]:
                    target_products = 2000
                # 90% ì´ìƒ ìƒí’ˆë“±ë¡ ì‹œ ìš´ì˜ ë‹¨ê³„ë¡œ ì „í™˜
                is_operating = product_count >= target_products * 0.9
            stage = "ìš´ì˜" if is_operating else "ì—…ë¡œë“œ"
            
            # ìˆ˜ë™ ì„¤ì •ê°’ì´ ìˆìœ¼ë©´ ìš°ì„  ì ìš©, ì•„ë‹ˆë©´ ìë™ ê³„ì‚°
            if "operationDays" in acc_settings:
                operation_days = int(acc_settings["operationDays"])
            elif is_operating:
                operation_days = 30 # ê¸°ë³¸ê°’ (ë¡œì§ì— ë”°ë¼ ìˆ˜ì • ê°€ëŠ¥)
            else:
                operation_days = 0 
            
            # [Cleaned up duplicated logic]

            # ë¦¬ë‰´ì–¼ íŒë³„ ì •ì±…: 60ì¼ ê²½ê³¼ OR ë§¤ì¶œ 50ë§Œ ì´í•˜ OR ìœ ì…ìˆ˜ 5ì¼ ì—°ì† í•˜ë½
            
            # [2026-01-19 Updated] ë§ˆì¼€íŒ… ë°ì´í„°(ë§¤ì¶œ/ìœ ì…/ì£¼ë¬¸) ë§¤ì¹­
            # market_statsëŠ” get_marketing_data()ë¡œ ì´ë¯¸ ë¡œë“œë¨
            
            m_data = {"revenue": 0, "visitors": 0, "orders": 0, "is_declining": False, "zero_orders": False}
            
            # 1. ìŠ¤í† ì–´ëª… ë§¤ì¹­
            # [Debug] ë§¤ì¶œ ë°ì´í„° ë§¤ì¹­ ë””ë²„ê¹… (ìƒì„¸ ë¡œê·¸)
            if len(dashboard_accounts) < 5: # ì²˜ìŒ 5ê°œë§Œ ë¡œê·¸
                 print(f"[Dashboard Debug] Checking revenue for '{store_name}' (Platform: {mapped_platform})")
                 # print(f"  - Available keys in market stats (sample): {list(market_stats.keys())[:5]} ... total {len(market_stats)}")
            
            if store_name in market_stats:
                m_data = market_stats[store_name]
                if len(dashboard_accounts) < 5: print(f"  - Matched exactly: {store_name} -> Revenue: {m_data.get('revenue', 0):,}")
            # 2. '_' ë’¤ ì´ë¦„ ë§¤ì¹­
            elif "_" in store_name:
                 short_name = store_name.split("_", 1)[1]
                 if short_name in market_stats:
                     m_data = market_stats[short_name]
                     if len(dashboard_accounts) < 5: print(f"  - Matched by short name '{short_name}' -> Revenue: {m_data.get('revenue', 0):,}")
            # 3. ì •ê·œí™” ë§¤ì¹­ (ì´ë¯¸ get_marketing_dataì—ì„œ ë§¤í•‘ ì²˜ë¦¬í–ˆì§€ë§Œ ì•ˆì „ì¥ì¹˜)
            if m_data["revenue"] == 0:
                 import re
                 norm_name = re.sub(r'\(.*?\)', '', store_name).strip()
                 if norm_name in market_stats:
                     m_data = market_stats[norm_name]
                     if len(dashboard_accounts) < 5: print(f"  - Matched by norm name '{norm_name}' -> Revenue: {m_data.get('revenue', 0):,}")
            
            real_revenue = m_data.get("revenue", 0)
            visitors = m_data.get("visitors", 0)
            orders = m_data.get("orders", 0)

            # ë§Œì•½ ì‹œíŠ¸ì— ì—†ìœ¼ë©´ ê¸°ì¡´ ì„¤ì •ê°’ ì‚¬ìš© (í•˜ìœ„ í˜¸í™˜)
            if real_revenue == 0:
                settings_rev = acc_settings.get("revenue30d", 0)
                if settings_rev > 0:
                    real_revenue = settings_rev
                    if len(dashboard_accounts) < 5: print(f"  - Using manual settings revenue: {real_revenue:,}")
                else:
                    if len(dashboard_accounts) < 5: print(f"  - No revenue found (0)")

            revenue_30d = real_revenue
            
            # ë¦¬ë‰´ì–¼ íŒë³„ ì •ì±… ì—…ë°ì´íŠ¸
            renewal_reasons = []
            if is_operating:
                # 1. ìš´ì˜ 60ì¼ ê²½ê³¼
                if operation_days >= 60:
                    renewal_reasons.append(f"ìš´ì˜ê¸°ê°„ {operation_days}ì¼ ê²½ê³¼")
                
                # 2. ìš´ì˜ 30ì¼ ì´ìƒ & ë§¤ì¶œ 50ë§Œ ì´í•˜
                if operation_days >= 30 and revenue_30d < 500000:
                    renewal_reasons.append(f"ìš´ì˜ {operation_days}ì¼ì°¨ ë§¤ì¶œ {revenue_30d//10000}ë§Œì› (50ë§Œ ë¯¸ë§Œ)")
                
                # 3. [New] ìœ ì…ìˆ˜ 5ì¼ ì—°ì† í•˜ë½
                if m_data.get("is_declining"):
                    renewal_reasons.append(f"ë°©ë¬¸ì 5ì¼ ì—°ì† ê°ì†Œ ({visitors}ëª…)")
                
                # 4. [New] 7ì¼ê°„ ì£¼ë¬¸ 0ê±´
                if m_data.get("zero_orders"):
                    renewal_reasons.append(f"ìµœê·¼ 7ì¼ ì£¼ë¬¸ 0ê±´")
                
                # ê¸°ì¡´ ì„¤ì •ê°’ í•˜ë½ì¼ìˆ˜ ì²´í¬ (í•˜ìœ„ í˜¸í™˜ ìœ ì§€)
                inflow_decrease_days = acc_settings.get("inflowDecreaseDays", 0)
                if inflow_decrease_days >= 5 and not m_data.get("is_declining"):
                     renewal_reasons.append(f"ìœ ì…ìˆ˜ {inflow_decrease_days}ì¼ ì—°ì† í•˜ë½(ìˆ˜ë™ì„¤ì •)")
                
            renewal_reason = ""
            if renewal_reasons:
                stage = "ë¦¬ë‰´ì–¼ëŒ€ìƒ"
                renewal_reason = " | ".join(renewal_reasons)
            
            # 4. ëª©í‘œ ë§¤ì¶œ ì •ì±… ë°˜ì˜
            target_revenue = 2000000  # ëª¨ë“  í”Œë«í¼ ëª©í‘œ ë§¤ì¶œ 200ë§Œì›ìœ¼ë¡œ í†µì¼
            if len(dashboard_accounts) < 1: # Log only once
                print(f"[Dashboard] Force setting target revenue to {target_revenue} for {store_name}")
            
            # ì—…ë¡œë“œ ì§„í–‰ë¥  ê³„ì‚°
            progress = min(int((product_count / target_products) * 100), 100) if target_products > 0 else 0

            dashboard_accounts.append({
                "name": store_name,
                "platform": mapped_platform,
                "stage": stage,
                "products": product_count,
                "targetProducts": target_products,
                "progress": progress,
                "revenue": revenue_30d, 
                "targetRevenue": target_revenue,
                "visitors": visitors, # New
                "orders": orders,    # New
                "status": "success" if stage == "ìš´ì˜" else ("danger" if stage == "ë¦¬ë‰´ì–¼ëŒ€ìƒ" else "warning"),
                "operationDays": operation_days,
                "renewalReason": renewal_reason
            })
            
        return {"success": True, "accounts": dashboard_accounts}
    except Exception as e:
        print(f"[ëŒ€ì‹œë³´ë“œ] ë°ì´í„° ì¡°íšŒ ì˜¤ë¥˜: {e}")
        return {"success": False, "message": str(e)}

@app.get("/api/bulsaja/logs")
async def get_bulsaja_logs(request: Request):
    get_current_user(request)
    return {"logs": getattr(bulsaja_manager, 'logs', [])}

# ì‘ì—… ë¡œê·¸ ëª¨ë¸
class WorkLogRequest(BaseModel):
    work_type: str  # ìƒí’ˆì‚­ì œ, ìƒí’ˆë“±ë¡, ìƒí’ˆìˆ˜ì •, ë§ˆì¼€íŒ…ìˆ˜ì§‘, ì˜ˆì•½ì‘ì—…
    account: str    # ê³„ì •ëª… ë˜ëŠ” ê·¸ë£¹ëª…
    count: int = 0  # ì²˜ë¦¬ ìƒí’ˆ/ê³„ì • ìˆ˜
    detail: str = ""  # ìƒì„¸ ë‚´ìš©
    method: str = ""  # ì‹¤í–‰ ë°©ë²•
    datetime: str = ""  # ë‚ ì§œ/ì‹œê°„ (YYYY-MM-DD HH:MM:SS) - ë¹„ì–´ìˆìœ¼ë©´ í˜„ì¬ ì‹œê°„

@app.post("/api/bulsaja/update_settings")
async def update_bulsaja_settings(request: Request):
    """ë¶ˆì‚¬ì ëŒ€ì‹œë³´ë“œ ìˆ˜ë™ ì„¤ì • ì €ì¥ (ìš´ì˜ì¼ìˆ˜ ë“±)"""
    get_current_user(request)

    # JSON body ì§ì ‘ íŒŒì‹±
    data = await request.json()
    store_name = data.get("store_name")
    operation_days = data.get("operation_days")

    print(f"[DEBUG] update_bulsaja_settings called: store_name={store_name}, operation_days={operation_days}")

    if not store_name:
        return {"success": False, "message": "ìŠ¤í† ì–´ëª…ì´ í•„ìš”í•©ë‹ˆë‹¤."}

    settings_path = os.path.join(os.path.dirname(__file__), "bulsaja_settings.json")
    settings = {}
    if os.path.exists(settings_path):
        try:
            with open(settings_path, 'r', encoding='utf-8') as f:
                settings = json.load(f)
        except: pass

    if store_name not in settings:
        settings[store_name] = {}

    if operation_days is not None:
        settings[store_name]["operationDays"] = int(operation_days)
        print(f"[DEBUG] Setting operationDays for {store_name} = {operation_days}")

    try:
        with open(settings_path, 'w', encoding='utf-8') as f:
            json.dump(settings, f, ensure_ascii=False, indent=4)
        print(f"[DEBUG] Settings saved successfully")
        return {"success": True}
    except Exception as e:
        print(f"[DEBUG] Error saving settings: {e}")
        return {"success": False, "message": str(e)}

class BulsajaFolderRequest(BaseModel):
    folder: str

@app.post("/api/bulsaja/folder")
async def set_bulsaja_folder(request: Request, req: BulsajaFolderRequest):
    get_current_user(request)
    return bulsaja_manager.set_folder(req.folder)

# ========== ì‘ì—… ë¡œê·¸ API ==========
WORK_LOG_SHEET = "ì‘ì—…ë¡œê·¸"

def log_work(work_type: str, account: str, count: int = 0, detail: str = "", method: str = "", datetime_str: str = ""):
    """ì‘ì—… ë¡œê·¸ ê¸°ë¡"""
    try:
        from datetime import datetime
        
        # Google Sheetsì— ê¸°ë¡
        creds = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=["https://www.googleapis.com/auth/spreadsheets"])
        gc = gspread.authorize(creds)
        wb = gc.open_by_key(SPREADSHEET_KEY)
        
        # ì‘ì—…ë¡œê·¸ ì‹œíŠ¸ ê°€ì ¸ì˜¤ê¸° (ì—†ìœ¼ë©´ ìƒì„±)
        try:
            ws = wb.worksheet(WORK_LOG_SHEET)
        except:
            ws = wb.add_worksheet(title=WORK_LOG_SHEET, rows=1000, cols=10)
            ws.update('A1', [['ì¼ì‹œ', 'ì‘ì—…ìœ í˜•', 'ê³„ì •ëª…', 'ìƒí’ˆìˆ˜', 'ìƒì„¸ë‚´ìš©', 'ì‹¤í–‰ë°©ë²•']])
        
        # ë‚ ì§œ/ì‹œê°„ ê²°ì • (ì œê³µë˜ë©´ ì‚¬ìš©, ì•„ë‹ˆë©´ í˜„ì¬ ì‹œê°„)
        if datetime_str:
            timestamp = datetime_str
        else:
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # ìƒˆ í–‰ ì¶”ê°€
        ws.append_row([timestamp, work_type, account, count, detail, method])
        
        print(f"[ì‘ì—…ë¡œê·¸] {timestamp} | {work_type} | {account} | {count}ê°œ | {detail}")
        
    except Exception as e:
        print(f"[ì‘ì—…ë¡œê·¸ ì˜¤ë¥˜] {e}")

@app.post("/api/work-log/add")
async def add_work_log(request: Request, req: WorkLogRequest):
    """ì‘ì—… ë¡œê·¸ ì¶”ê°€ (ìˆ˜ë™)"""
    get_current_user(request)
    log_work(req.work_type, req.account, req.count, req.detail, req.method, req.datetime)
    return {"success": True}

@app.get("/api/work-log/calendar")
async def get_work_calendar(request: Request, year: int, month: int):
    """ì›”ë³„ ì‘ì—… ë¡œê·¸ ì¡°íšŒ (ì‘ì—… ìœ í˜•ë³„ ê·¸ë£¹í™”)"""
    get_current_user(request)
    
    try:
        from datetime import datetime
        from collections import defaultdict
        
        creds = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=["https://www.googleapis.com/auth/spreadsheets"])
        gc = gspread.authorize(creds)
        wb = gc.open_by_key(SPREADSHEET_KEY)
        ws = wb.worksheet(WORK_LOG_SHEET)
        
        # ëª¨ë“  ë°ì´í„° ê°€ì ¸ì˜¤ê¸°
        data = ws.get_all_values()[1:]  # í—¤ë” ì œì™¸
        
        # í•´ë‹¹ ì›” í•„í„°ë§ ë° ê·¸ë£¹í™” ì¤€ë¹„
        # êµ¬ì¡°: {ë‚ ì§œ: {ì‘ì—…ìœ í˜•: [ë¡œê·¸ë“¤]}}
        grouped_data = defaultdict(lambda: defaultdict(list))
        
        for row in data:
            if len(row) >= 6 and row[0]:
                datetime_str = row[0].strip()
                row_year = None
                row_month = None
                row_day = None
                
                # ë‹¤ì–‘í•œ ë‚ ì§œ í˜•ì‹ íŒŒì‹±
                try:
                    # YYYY-MM-DD HH:MM:SS
                    dt = datetime.strptime(datetime_str.split()[0], "%Y-%m-%d")
                    row_year = dt.year
                    row_month = dt.month
                    row_day = dt.day
                except:
                    try:
                        # M/D/YYYY ë˜ëŠ” MM/DD/YYYY
                        parts = datetime_str.split()[0].split('/')
                        if len(parts) == 3:
                            row_month = int(parts[0])
                            row_day = int(parts[1])
                            row_year = int(parts[2])
                    except:
                        pass
                
                if row_year == year and row_month == month:
                    # ë‚ ì§œ í‚¤ ìƒì„± (YYYY-MM-DD)
                    date_key = f"{row_year:04d}-{row_month:02d}-{row_day:02d}"
                    work_type = row[1].strip()
                    account = row[2].strip()
                    count = int(row[3]) if row[3].isdigit() else 0
                    detail = row[4].strip()
                    method = row[5].strip()
                    
                    # ê·¸ë£¹ì— ì¶”ê°€
                    grouped_data[date_key][work_type].append({
                        "datetime": datetime_str,
                        "account": account,
                        "count": count,
                        "detail": detail,
                        "method": method
                    })
        
        # ê·¸ë£¹í™”ëœ ë°ì´í„°ë¥¼ ìµœì¢… í˜•ì‹ìœ¼ë¡œ ë³€í™˜
        month_data = []
        for date_key, work_types in grouped_data.items():
            for work_type, logs in work_types.items():
                # ê³„ì • ëª©ë¡ ë° ì´ ê°œìˆ˜ ê³„ì‚°
                accounts = [log["account"] for log in logs]
                total_count = sum(log["count"] for log in logs)
                
                month_data.append({
                    "datetime": f"{date_key} 00:00:00",  # ë‚ ì§œë§Œ ì‚¬ìš©
                    "work_type": work_type,
                    "account": f"{len(accounts)}ê°œ ìŠ¤í† ì–´",  # "2ê°œ ìŠ¤í† ì–´" í˜•ì‹
                    "count": total_count,
                    "detail": f"{', '.join(accounts[:3])}{'...' if len(accounts) > 3 else ''}",  # ì²˜ìŒ 3ê°œë§Œ í‘œì‹œ
                    "method": logs[0]["method"] if logs else "",
                    "store_count": len(accounts)  # í”„ë¡ íŠ¸ì—”ë“œì—ì„œ ì‚¬ìš©í•  ìˆ˜ ìˆë„ë¡
                })
        
        print(f"[ì›”ë³„ ì¡°íšŒ] {year}ë…„ {month}ì›”: {len(month_data)}ê°œ ê·¸ë£¹")
        return {"logs": month_data}
        
    except Exception as e:
        print(f"[ì‘ì—…ë¡œê·¸ ì¡°íšŒ ì˜¤ë¥˜] {e}")
        import traceback
        traceback.print_exc()
        return {"logs": []}

@app.get("/api/work-log/day")
async def get_work_day(request: Request, date: str):
    """íŠ¹ì • ë‚ ì§œ ì‘ì—… ë¡œê·¸ ì¡°íšŒ"""
    get_current_user(request)
    
    print(f"[ë””ë²„ê·¸] ì¡°íšŒ ìš”ì²­ ë‚ ì§œ: {date}")
    
    try:
        creds = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=["https://www.googleapis.com/auth/spreadsheets"])
        gc = gspread.authorize(creds)
        wb = gc.open_by_key(SPREADSHEET_KEY)
        ws = wb.worksheet(WORK_LOG_SHEET)
        
        data = ws.get_all_values()[1:]
        print(f"[ë””ë²„ê·¸] ì „ì²´ ë¡œê·¸ ìˆ˜: {len(data)}ê°œ")
        
        # ë‚ ì§œ íŒŒì‹± (YYYY-MM-DD)
        from datetime import datetime
        try:
            target_date = datetime.strptime(date, "%Y-%m-%d").date()
            print(f"[ë””ë²„ê·¸] íŒŒì‹±ëœ ë‚ ì§œ: {target_date}")
        except:
            print(f"[ë””ë²„ê·¸] ë‚ ì§œ íŒŒì‹± ì‹¤íŒ¨: {date}")
            return {"logs": []}
        
        day_data = []
        for row in data:
            if len(row) >= 6 and row[0]:
                # ë‹¤ì–‘í•œ ë‚ ì§œ í˜•ì‹ ì‹œë„
                row_date = None
                datetime_str = row[0].strip()
                
                # ì‹œë„1: YYYY-MM-DD HH:MM:SS
                try:
                    row_date = datetime.strptime(datetime_str.split()[0], "%Y-%m-%d").date()
                except:
                    pass
                
                # ì‹œë„2: MM/DD/YYYY ë˜ëŠ” M/D/YYYY
                if not row_date:
                    try:
                        # "1/2/2026" í˜•ì‹
                        parts = datetime_str.split()[0].split('/')
                        if len(parts) == 3:
                            row_date = datetime(int(parts[2]), int(parts[0]), int(parts[1])).date()
                    except:
                        pass
                
                # ì‹œë„3: DD/MM/YYYY
                if not row_date:
                    try:
                        parts = datetime_str.split()[0].split('/')
                        if len(parts) == 3:
                            row_date = datetime(int(parts[2]), int(parts[1]), int(parts[0])).date()
                    except:
                        pass
                
                print(f"[ë””ë²„ê·¸] '{datetime_str}' â†’ {row_date} vs {target_date} â†’ ì¼ì¹˜={row_date == target_date if row_date else False}")
                
                if row_date == target_date:
                    # datetimeì„ YYYY-MM-DD HH:MM:SS í˜•ì‹ìœ¼ë¡œ í‘œì¤€í™”
                    normalized_datetime = datetime_str
                    try:
                        if '-' in datetime_str.split()[0]:
                            normalized_datetime = datetime_str  # ì´ë¯¸ í‘œì¤€ í˜•ì‹
                        elif '/' in datetime_str.split()[0]:
                            parts = datetime_str.split()
                            date_parts = parts[0].split('/')
                            if len(date_parts) == 3:
                                m, d, y = int(date_parts[0]), int(date_parts[1]), int(date_parts[2])
                                time_part = parts[1] if len(parts) > 1 else "00:00:00"
                                normalized_datetime = f"{y:04d}-{m:02d}-{d:02d} {time_part}"
                    except:
                        pass
                    
                    day_data.append({
                        "datetime": normalized_datetime,
                        "work_type": row[1],
                        "account": row[2],
                        "count": int(row[3]) if row[3].isdigit() else 0,
                        "detail": row[4],
                        "method": row[5]
                    })
        
        print(f"[ë””ë²„ê·¸] ì°¾ì€ ì‘ì—…: {len(day_data)}ê°œ")
        return {"logs": day_data}
        
    except Exception as e:
        print(f"[ì¼ë³„ ë¡œê·¸ ì¡°íšŒ ì˜¤ë¥˜] {e}")
        import traceback
        traceback.print_exc()
        return {"logs": []}

@app.get("/api/work-log/stats")
async def get_work_stats(request: Request, year: int, month: int):
    """ì›”ê°„ í†µê³„"""
    get_current_user(request)
    
    try:
        from datetime import datetime
        
        creds = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=["https://www.googleapis.com/auth/spreadsheets"])
        gc = gspread.authorize(creds)
        wb = gc.open_by_key(SPREADSHEET_KEY)
        ws = wb.worksheet(WORK_LOG_SHEET)
        
        data = ws.get_all_values()[1:]
        
        stats = {
            "total_works": 0,
            "deleted_products": 0,
            "uploaded_products": 0,
            "processed_accounts": set()
        }
        
        for row in data:
            if len(row) >= 6 and row[0]:
                datetime_str = row[0].strip()
                row_year = None
                row_month = None
                
                # ë‚ ì§œ íŒŒì‹±
                try:
                    # YYYY-MM-DD HH:MM:SS
                    dt = datetime.strptime(datetime_str.split()[0], "%Y-%m-%d")
                    row_year = dt.year
                    row_month = dt.month
                except:
                    try:
                        # M/D/YYYY
                        parts = datetime_str.split()[0].split('/')
                        if len(parts) == 3:
                            row_month = int(parts[0])
                            row_year = int(parts[2])
                    except:
                        pass
                
                if row_year == year and row_month == month:
                    stats["total_works"] += 1
                    count = int(row[3]) if row[3].isdigit() else 0
                    
                    if row[1] == "ìƒí’ˆì‚­ì œ":
                        stats["deleted_products"] += count
                    elif row[1] == "ìƒí’ˆë“±ë¡":
                        stats["uploaded_products"] += count
                    
                    if row[2]:
                        stats["processed_accounts"].add(row[2])
        
        stats["processed_accounts"] = len(stats["processed_accounts"])
        
        print(f"[í†µê³„] {year}ë…„ {month}ì›”: ì‘ì—… {stats['total_works']}ê°œ")
        return stats
        
    except Exception as e:
        print(f"[í†µê³„ ì¡°íšŒ ì˜¤ë¥˜] {e}")
        return {
            "total_works": 0,
            "deleted_products": 0,
            "uploaded_products": 0,
            "processed_accounts": 0
        }

class WorkLogUpdateRequest(BaseModel):
    datetime: str  # ì›ë³¸ ì¼ì‹œ (ê³ ìœ  ì‹ë³„ì)
    work_type: str = None
    account: str = None
    count: int = None
    detail: str = None
    method: str = None
    new_datetime: str = None  # ë‚ ì§œ/ì‹œê°„ ë³€ê²½ ì‹œ

@app.put("/api/work-log/update")
async def update_work_log(request: Request, req: WorkLogUpdateRequest):
    """ì‘ì—… ë¡œê·¸ ìˆ˜ì •"""
    get_current_user(request)
    
    try:
        creds = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=["https://www.googleapis.com/auth/spreadsheets"])
        gc = gspread.authorize(creds)
        wb = gc.open_by_key(SPREADSHEET_KEY)
        ws = wb.worksheet(WORK_LOG_SHEET)
        
        # ëª¨ë“  ë°ì´í„° ê°€ì ¸ì˜¤ê¸°
        all_data = ws.get_all_values()
        
        # ì›ë³¸ ì¼ì‹œë¡œ í–‰ ì°¾ê¸°
        row_index = None
        for i, row in enumerate(all_data):
            if i == 0:  # í—¤ë” ìŠ¤í‚µ
                continue
            if row[0] == req.datetime:
                row_index = i + 1  # gspreadëŠ” 1-indexed
                break
        
        if row_index is None:
            return {"success": False, "message": "ì‘ì—…ì„ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤"}
        
        # í˜„ì¬ ê°’ ê°€ì ¸ì˜¤ê¸°
        current = all_data[row_index - 1]
        
        # ì—…ë°ì´íŠ¸í•  ê°’ ì¤€ë¹„
        new_datetime = req.new_datetime if req.new_datetime else current[0]
        new_work_type = req.work_type if req.work_type is not None else current[1]
        new_account = req.account if req.account is not None else current[2]
        new_count = req.count if req.count is not None else current[3]
        new_detail = req.detail if req.detail is not None else current[4]
        new_method = req.method if req.method is not None else current[5]
        
        # í–‰ ì—…ë°ì´íŠ¸
        ws.update(values=[[new_datetime, new_work_type, new_account, str(new_count), new_detail, new_method]], range_name=f'A{row_index}:F{row_index}')
        
        print(f"[ì‘ì—…ë¡œê·¸ ìˆ˜ì •] {req.datetime} â†’ {new_datetime}")
        return {"success": True}
        
    except Exception as e:
        print(f"[ì‘ì—…ë¡œê·¸ ìˆ˜ì • ì˜¤ë¥˜] {e}")
        return {"success": False, "message": str(e)}

class WorkLogDeleteRequest(BaseModel):
    datetime: str  # ì¼ì‹œ (ê³ ìœ  ì‹ë³„ì)

@app.delete("/api/work-log/delete")
async def delete_work_log(request: Request, req: WorkLogDeleteRequest):
    """ì‘ì—… ë¡œê·¸ ì‚­ì œ"""
    get_current_user(request)
    
    try:
        creds = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=["https://www.googleapis.com/auth/spreadsheets"])
        gc = gspread.authorize(creds)
        wb = gc.open_by_key(SPREADSHEET_KEY)
        ws = wb.worksheet(WORK_LOG_SHEET)
        
        # ëª¨ë“  ë°ì´í„° ê°€ì ¸ì˜¤ê¸°
        all_data = ws.get_all_values()
        
        # ì›ë³¸ ì¼ì‹œë¡œ í–‰ ì°¾ê¸°
        row_index = None
        for i, row in enumerate(all_data):
            if i == 0:  # í—¤ë” ìŠ¤í‚µ
                continue
            if row[0] == req.datetime:
                row_index = i + 1  # gspreadëŠ” 1-indexed
                break
        
        if row_index is None:
            return {"success": False, "message": "ì‘ì—…ì„ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤"}
        
        # í–‰ ì‚­ì œ
        ws.delete_rows(row_index)
        
        print(f"[ì‘ì—…ë¡œê·¸ ì‚­ì œ] {req.datetime}")
        return {"success": True}
        
    except Exception as e:
        print(f"[ì‘ì—…ë¡œê·¸ ì‚­ì œ ì˜¤ë¥˜] {e}")
        return {"success": False, "message": str(e)}

@app.get("/api/bulsaja/folder")
async def get_bulsaja_folder(request: Request):
    get_current_user(request)
    return {"folder": bulsaja_manager.get_folder()}

# ë¶ˆì‚¬ì ì‹œíŠ¸ ì„¤ì • API
class BulsajaSettingsRequest(BaseModel):
    program: str
    # ìƒí’ˆì—…ë¡œë“œìš©
    uploadMarket: str = ""
    uploadCount: str = ""
    # ìƒí’ˆì‚­ì œìš© (C29ëŠ” ì‹¤í–‰ ì‹œ ê·¸ë£¹ë³„ë¡œ ë³€ê²½)
    deleteCount: str = ""
    # ìƒí’ˆë³µì‚¬ìš© (C35ëŠ” ì‹¤í–‰ ì‹œ ê·¸ë£¹ë³„ë¡œ ë³€ê²½)
    copySourceMarket: str = ""
    copyCount: str = ""

@app.post("/api/bulsaja/settings")
async def save_bulsaja_settings(request: Request, req: BulsajaSettingsRequest):
    """ë¶ˆì‚¬ì êµ¬ê¸€ì‹œíŠ¸ ì„¤ì • ì €ì¥"""
    get_current_user(request)
    
    try:
        from dotenv import dotenv_values
        
        # ìë™í™”ì‹œìŠ¤í…œ í´ë”ì˜ .env ì½ê¸°
        env_path = Path(bulsaja_manager.base_folder) / ".env"
        if env_path.exists():
            env_config = dotenv_values(env_path)
            creds_file = env_config.get("SERVICE_ACCOUNT_JSON", CREDENTIALS_FILE)
            sheet_key = env_config.get("SPREADSHEET_KEY", BULSAJA_SHEET_KEY)
        else:
            creds_file = CREDENTIALS_FILE
            sheet_key = BULSAJA_SHEET_KEY
        
        # ìƒëŒ€ê²½ë¡œë©´ base_folder ê¸°ì¤€ìœ¼ë¡œ ë³€í™˜
        creds_path = Path(creds_file)
        if not creds_path.is_absolute():
            creds_path = Path(bulsaja_manager.base_folder) / creds_file
        
        # ë¶ˆì‚¬ì ì‹œíŠ¸ ì—´ê¸°
        creds = Credentials.from_service_account_file(
            str(creds_path),
            scopes=["https://www.googleapis.com/auth/spreadsheets"]
        )
        gc = gspread.authorize(creds)
        ws = gc.open_by_key(sheet_key).worksheet(BULSAJA_TAB_NAME)
        
        # C10: ì‚¬ìš©í• í”„ë¡œê·¸ë¨
        ws.update_acell("C10", req.program)
        
        # í”„ë¡œê·¸ë¨ë³„ ì„¤ì •
        if req.program == "2. ìƒí’ˆì—…ë¡œë“œ":
            # C17: ì—…ë¡œë“œë§ˆì¼“ì„¤ì •, C18: ì—…ë¡œë“œìˆ˜
            if req.uploadMarket:
                ws.update_acell("C17", req.uploadMarket)
            if req.uploadCount:
                ws.update_acell("C18", req.uploadCount)
                
        elif req.program == "4. ìƒí’ˆì‚­ì œ":
            # C32: ì‚­ì œìˆ˜ëŸ‰ë§Œ (C29ëŠ” ì‹¤í–‰ ì‹œ ê·¸ë£¹ë³„ë¡œ ë³€ê²½)
            if req.deleteCount:
                ws.update_acell("C32", req.deleteCount)
                
        elif req.program == "4-3. ë¶ˆì‚¬ììƒí’ˆë³µì‚¬":
            # C29: ì†ŒìŠ¤ë§ˆì¼“, C37: ë³µì‚¬ìˆ˜ëŸ‰ (C35ëŠ” ì‹¤í–‰ ì‹œ ê·¸ë£¹ë³„ë¡œ ë³€ê²½)
            if req.copySourceMarket:
                ws.update_acell("C29", req.copySourceMarket)
            if req.copyCount:
                ws.update_acell("C37", req.copyCount)
        
        return {"success": True}
    except Exception as e:
        print(f"[ë¶ˆì‚¬ì] ì‹œíŠ¸ ì„¤ì • ì €ì¥ ì˜¤ë¥˜: {e}")
        return {"success": False, "message": str(e)}

@app.get("/api/bulsaja/settings")
async def get_bulsaja_settings(request: Request):
    """ë¶ˆì‚¬ì êµ¬ê¸€ì‹œíŠ¸ ì„¤ì • ì¡°íšŒ"""
    get_current_user(request)
    
    try:
        from dotenv import dotenv_values
        
        # ìë™í™”ì‹œìŠ¤í…œ í´ë”ì˜ .env ì½ê¸°
        env_path = Path(bulsaja_manager.base_folder) / ".env"
        if env_path.exists():
            env_config = dotenv_values(env_path)
            creds_file = env_config.get("SERVICE_ACCOUNT_JSON", CREDENTIALS_FILE)
            sheet_key = env_config.get("SPREADSHEET_KEY", BULSAJA_SHEET_KEY)
        else:
            creds_file = CREDENTIALS_FILE
            sheet_key = BULSAJA_SHEET_KEY
        
        # ìƒëŒ€ê²½ë¡œë©´ base_folder ê¸°ì¤€ìœ¼ë¡œ ë³€í™˜
        creds_path = Path(creds_file)
        if not creds_path.is_absolute():
            creds_path = Path(bulsaja_manager.base_folder) / creds_file
        
        # ë¶ˆì‚¬ì ì‹œíŠ¸ ì—´ê¸°
        creds = Credentials.from_service_account_file(
            str(creds_path),
            scopes=["https://www.googleapis.com/auth/spreadsheets"]
        )
        gc = gspread.authorize(creds)
        ws = gc.open_by_key(sheet_key).worksheet(BULSAJA_TAB_NAME)
        
        # ê¸°ë³¸ ì„¤ì •ê°’ ì½ê¸°
        program = ws.acell("C10").value or ""
        # ìƒí’ˆì—…ë¡œë“œìš©
        uploadMarket = ws.acell("C17").value or ""
        uploadCount = ws.acell("C18").value or ""
        # ìƒí’ˆì‚­ì œìš© (C32: ìˆ˜ëŸ‰)
        deleteCount = ws.acell("C32").value or ""
        # ìƒí’ˆë³µì‚¬ìš© (C29: ì†ŒìŠ¤ë§ˆì¼“, C37: ìˆ˜ëŸ‰)
        copySourceMarket = ws.acell("C29").value or ""
        copyCount = ws.acell("C37").value or ""
        
        # ========== ì¶”ê°€ ì„¤ì •ê°’ (í‘œì‹œìš©) ==========
        # ë§ˆì§„ì„¤ì • (í–‰ 9~11)
        margin = {
            "exchangeRate": ws.acell("E9").value or "",       # ê¸°ì¤€ í™˜ìœ¨(ìœ„ì•ˆ)
            "cardFee": ws.acell("E10").value or "",           # ì¹´ë“œìˆ˜ìˆ˜ë£Œ(%)
            "marketDiscount": ws.acell("E11").value or "",    # ë§ˆì¼“ í• ì¸ìœ¨(%)
            "priceRounding": ws.acell("G9").value or "",      # ê°€ê²©ë‹¨ìœ„ì˜¬ë¦¼(ì›)
            "percentMargin": ws.acell("G10").value or "",     # í¼ì„¼íŠ¸ë§ˆì§„(%)
            "addMargin": ws.acell("G11").value or ""          # ë”í•˜ê¸° ë§ˆì§„(ì›)
        }
        
        # ìƒí’ˆì—…ë¡œë“œ ì„¤ì • (í–‰ 15~25)
        upload = {
            "productName": ws.acell("C19").value or "",       # ìƒí’ˆëª…
            "uploadCount": uploadCount,                        # ì—…ë¡œë“œìˆ˜ (C18)
            "optionSort": ws.acell("C21").value or "",        # ì˜µì…˜ì„¤ì •
            "uploadCondition": ws.acell("C23").value or "",   # ì—…ë¡œë“œì¡°ê±´
            "minPrice": ws.acell("C24").value or "",          # ì˜µì…˜ ìµœì € ê°€ê²©
            "maxPrice": ws.acell("C25").value or ""           # ì˜µì…˜ ìµœëŒ€ ê°€ê²©
        }
        
        # ìƒí’ˆì‚­ì œ/ë³µì‚¬ ì„¤ì • (í–‰ 29~38)
        deleteCopy = {
            "deleteScope": ws.acell("C31").value or "",       # ìƒí’ˆì‚­ì œì„¤ì • (ì—…ë¡œë“œí•œ ë§ˆì¼“ì—ì„œë§Œ)
            "deleteOrder": ws.acell("C33").value or "",       # ì‚­ì œë°©ì‹ (ê³¼ê±°ìˆœ)
            "baseMarket": ws.acell("C34").value or "",        # ê¸°ì¤€ë§ˆì¼“ì„¤ì • (ìŠ¤ë§ˆíŠ¸ìŠ¤í† ì–´)
            "copyCondition": ws.acell("C38").value or ""      # ë³µì‚¬ì¡°ê±´ (ì „ì²´ìƒí’ˆë³µì‚¬)
        }
        
        return {
            "success": True,
            "program": program,
            "uploadMarket": uploadMarket,
            "uploadCount": uploadCount,
            "deleteCount": deleteCount,
            "copySourceMarket": copySourceMarket,
            "copyCount": copyCount,
            # ì¶”ê°€ ì„¤ì • (í‘œì‹œìš©)
            "margin": margin,
            "upload": upload,
            "deleteCopy": deleteCopy
        }
        
        # [Alias Logic] ë¡œì»¬ ì„¤ì • íŒŒì¼ ë³‘í•©
        try:
            settings_path = os.path.join(os.path.dirname(__file__), "bulsaja_settings.json")
            if os.path.exists(settings_path):
                with open(settings_path, 'r', encoding='utf-8') as f:
                    local_settings = json.load(f)
                    # sales_key_alias ë³‘í•©
                    if "sales_key_alias" in local_settings:
                        result["sales_key_alias"] = local_settings["sales_key_alias"]
        except Exception as e:
            print(f"[ë¶ˆì‚¬ì] ë¡œì»¬ ì„¤ì • ë³‘í•© ì˜¤ë¥˜: {e}")
            
        return result
    except Exception as e:
        print(f"[ë¶ˆì‚¬ì] ì‹œíŠ¸ ì„¤ì • ì¡°íšŒ ì˜¤ë¥˜: {e}")
        return {"success": False, "message": str(e)}

# ========== íƒ­ ê¶Œí•œ ì„¤ì • API ==========

TAB_PERMISSIONS_FILE = APP_DIR / "tab_permissions.json"

@app.get("/api/settings/tab-permissions")
async def get_tab_permissions(request: Request):
    """ìš´ì˜ì íƒ­ ê¶Œí•œ ì„¤ì • ì¡°íšŒ (ê´€ë¦¬ì ì „ìš©)"""
    user = get_current_user(request)
    if user.get("role") not in ("admin", "ê´€ë¦¬ì"):
        raise HTTPException(status_code=403, detail="ê´€ë¦¬ìë§Œ ì ‘ê·¼ ê°€ëŠ¥í•©ë‹ˆë‹¤")
    
    try:
        if TAB_PERMISSIONS_FILE.exists():
            with open(TAB_PERMISSIONS_FILE, 'r', encoding='utf-8') as f:
                permissions = json.load(f)
        else:
            # ê¸°ë³¸ê°’: ëª¨ë“  íƒ­ í—ˆìš©
            permissions = {
                "accounts": True,
                "sms": True,
                "monitor": True,
                "aio": True,
                "bulsaja": True,
                "tools": True,
                "marketing": True,
                "calendar": True
            }

        return {"success": True, "permissions": permissions}
    except Exception as e:
        print(f"[íƒ­ ê¶Œí•œ ì¡°íšŒ ì˜¤ë¥˜] {e}")
        return {"success": False, "message": str(e)}

@app.post("/api/settings/tab-permissions")
async def save_tab_permissions(request: Request):
    """ìš´ì˜ì íƒ­ ê¶Œí•œ ì„¤ì • ì €ì¥ (ê´€ë¦¬ì ì „ìš©)"""
    user = get_current_user(request)
    print(f"[íƒ­ ê¶Œí•œ ì €ì¥ ì‹œë„] user={user}, role={user.get('role')}")
    
    if user.get("role") not in ("admin", "ê´€ë¦¬ì"):
        print(f"[íƒ­ ê¶Œí•œ ì €ì¥ ê±°ë¶€] role '{user.get('role')}'ëŠ” admin/ê´€ë¦¬ìê°€ ì•„ë‹˜")
        raise HTTPException(status_code=403, detail="ê´€ë¦¬ìë§Œ ì ‘ê·¼ ê°€ëŠ¥í•©ë‹ˆë‹¤")
    
    try:
        data = await request.json()
        
        # JSON íŒŒì¼ë¡œ ì €ì¥
        with open(TAB_PERMISSIONS_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        
        print(f"[íƒ­ ê¶Œí•œ ì„¤ì • ì €ì¥] {data}")
        return {"success": True}
    except Exception as e:
        print(f"[íƒ­ ê¶Œí•œ ì €ì¥ ì˜¤ë¥˜] {e}")
        import traceback
        traceback.print_exc()
        return {"success": False, "message": str(e)}

# ========== All-in-One API ==========
import bcrypt
import base64

class AioRunRequest(BaseModel):
    platform: str
    task: str
    options: dict = {}  # ì‘ì—… ì˜µì…˜ (mode, count, date, delete_count ë“±)
    stores: List[str] = []  # ì„ íƒëœ ìŠ¤í† ì–´ ëª©ë¡

class AioUpdateActiveRequest(BaseModel):
    platform: str
    task: str
    active_stores: List[str]
    all_stores: List[str]

# ì˜¬ì¸ì› ìƒíƒœ ì €ì¥
def create_aio_status():
    return {
        "running": False,
        "progress": 0,
        "status": "idle",  # idle, running, completed, stopped
        "results": [],
        "process": None,
        "options": {},  # í˜„ì¬ ì‹¤í–‰ ì˜µì…˜
        "current_store": "",  # í˜„ì¬ ì²˜ë¦¬ ì¤‘ì¸ ìŠ¤í† ì–´
        "current_action": "",  # í˜„ì¬ ì‘ì—… ë‚´ìš©
        "total": 0,  # ì „ì²´ ìŠ¤í† ì–´ ìˆ˜
        "completed": 0,  # ì™„ë£Œëœ ìŠ¤í† ì–´ ìˆ˜
        "logs": [],  # ì‹¤ì‹œê°„ ë¡œê·¸ (ìµœê·¼ 50ê°œ)
        "log_file": None,
        "log_pos": 0
    }

# í”Œë«í¼ë³„ ìƒíƒœ ê´€ë¦¬
# í”Œë«í¼ë³„ ìƒíƒœ ê´€ë¦¬
aio_status_by_platform = {
    "ìŠ¤ë§ˆíŠ¸ìŠ¤í† ì–´": {},
    "11ë²ˆê°€": {},
    "ì¿ íŒ¡": {},
    "ESM": {},
}

# ê¸°ì¡´ í˜¸í™˜ì„±ì„ ìœ„í•œ ê¸°ë³¸ ìƒíƒœ (í”Œë«í¼ ë¯¸ì§€ì •ì‹œ ì‚¬ìš©)
aio_status = create_aio_status()

def get_aio_status(platform: str = None, task: str = None):
    """í”Œë«í¼/ì‘ì—…ë³„ ìƒíƒœ ë°˜í™˜"""
    if platform and platform in aio_status_by_platform:
        platform_dict = aio_status_by_platform[platform]
        
        # taskê°€ ì£¼ì–´ì§€ë©´ í•´ë‹¹ task ìƒíƒœ ë°˜í™˜ (ì—†ìœ¼ë©´ ìƒì„±)
        if task:
            if task not in platform_dict:
                platform_dict[task] = create_aio_status()
            return platform_dict[task]
            
        # taskê°€ ì—†ìœ¼ë©´(legacy), ì²« ë²ˆì§¸ ì‹¤í–‰ ì¤‘ì¸ ì‘ì—… ë°˜í™˜ or ì²« ë²ˆì§¸ ì‘ì—… ë°˜í™˜
        if platform_dict:
            # 1. ì‹¤í–‰ ì¤‘ì¸ ì‘ì—… ê²€ìƒ‰
            for t, status in platform_dict.items():
                if status.get("running"):
                    return status
            # 2. ì—†ìœ¼ë©´ ì•„ë¬´ê±°ë‚˜ ë°˜í™˜ (ì²«ë²ˆì§¸)
            return list(platform_dict.values())[0]
            
        # 3. ì•„ì˜ˆ ì—†ìœ¼ë©´ ê¸°ë³¸ ìƒì„± (default í‚¤ ì‚¬ìš©)
        if "default" not in platform_dict:
            platform_dict["default"] = create_aio_status()
        return platform_dict["default"]
        
    return aio_status

# ìŠ¤ë§ˆíŠ¸ìŠ¤í† ì–´ API ì¸ì¦
def ss_sign_client_secret(client_id: str, client_secret: str, ts_ms: int) -> str:
    pwd = f"{client_id}_{ts_ms}".encode("utf-8")
    hashed = bcrypt.hashpw(pwd, client_secret.strip().encode("utf-8"))
    return base64.b64encode(hashed).decode("utf-8")

def ss_get_access_token(client_id: str, client_secret: str) -> str:
    ts = int(time.time() * 1000)
    data = {
        "grant_type": "client_credentials",
        "client_id": client_id,
        "timestamp": ts,
        "client_secret_sign": ss_sign_client_secret(client_id, client_secret, ts),
        "type": "SELF",
    }
    headers = {"Content-Type": "application/x-www-form-urlencoded", "Accept": "application/json"}
    r = requests.post("https://api.commerce.naver.com/external/v1/oauth2/token", data=data, headers=headers, timeout=30)
    r.raise_for_status()
    return r.json()["access_token"]

# ìŠ¤ë§ˆíŠ¸ìŠ¤í† ì–´ ìƒí’ˆ ìˆ˜ëŸ‰ ì¡°íšŒ (ìƒì„¸)
def ss_get_product_count(access_token: str) -> Dict[str, int]:
    headers = {
        "Authorization": f"Bearer {access_token}", 
        "Accept": "application/json",
        "Content-Type": "application/json"
    }
    
    def total_elements(body: dict) -> int:
        base = {"page": 1, "size": 1}
        base.update(body)
        r = requests.post("https://api.commerce.naver.com/external/v1/products/search", 
                          headers=headers, json=base, timeout=30)
        r.raise_for_status()
        data = r.json()
        return int(data.get("totalElements") or data.get("total") or 0)
    
    def count_by_status(status_code: str) -> int:
        return total_elements({"productStatusTypes": [status_code]})
    
    total_all = total_elements({})
    on_sale = count_by_status("SALE")
    stop_selling = count_by_status("SUSPENSION")
    approval_wait = count_by_status("WAIT")
    
    return {
        "ì „ì²´": total_all,
        "íŒë§¤ì¤‘": on_sale,
        "íŒë§¤ì¤‘ì§€": stop_selling,
        "ìŠ¹ì¸ëŒ€ê¸°": approval_wait
    }

# ìŠ¤ë§ˆíŠ¸ìŠ¤í† ì–´ ë§ˆì§€ë§‰ ë“±ë¡ì¼ ì¡°íšŒ
def ss_get_last_registration_date(access_token: str) -> str:
    """íŒë§¤ì¤‘ ìƒí’ˆ ì¤‘ ê°€ì¥ ìµœê·¼ ë“±ë¡ì¼ ì¡°íšŒ"""
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Accept": "application/json",
        "Content-Type": "application/json"
    }

    try:
        # íŒë§¤ì¤‘ ìƒí’ˆì„ ë“±ë¡ì¼ ìµœì‹ ìˆœìœ¼ë¡œ 1ê°œë§Œ ì¡°íšŒ
        body = {
            "page": 1,
            "size": 1,
            "productStatusTypes": ["SALE"],
            "sortType": "RECENTLY_REGISTERED"  # ìµœê·¼ ë“±ë¡ìˆœ
        }

        r = requests.post(
            "https://api.commerce.naver.com/external/v1/products/search",
            headers=headers,
            json=body,
            timeout=30
        )
        r.raise_for_status()
        data = r.json()

        contents = data.get("contents", [])
        if contents and len(contents) > 0:
            product = contents[0]
            # ë“±ë¡ì¼ í•„ë“œ: regDate ë˜ëŠ” createdDate
            reg_date = product.get("regDate") or product.get("createdDate") or product.get("registrationDate")
            if reg_date:
                # ISO formatì„ YYYY-MM-DDë¡œ ë³€í™˜
                if "T" in str(reg_date):
                    return str(reg_date).split("T")[0]
                return str(reg_date)[:10]

        return ""
    except Exception as e:
        print(f"[ìŠ¤ë§ˆíŠ¸ìŠ¤í† ì–´] ë§ˆì§€ë§‰ ë“±ë¡ì¼ ì¡°íšŒ ì˜¤ë¥˜: {e}")
        return ""

# ìƒí’ˆìˆ˜ëŸ‰ êµ¬ê¸€ì‹œíŠ¸ ê¸°ë¡
def ss_save_product_count_to_sheet(store_name: str, counts: Dict[str, int], last_reg_date: str = ""):
    """ìƒí’ˆìˆ˜ëŸ‰ì„ êµ¬ê¸€ì‹œíŠ¸ 'ë“±ë¡ê°¯ìˆ˜' íƒ­ì— ê¸°ë¡ (ë§ˆì§€ë§‰ë“±ë¡ì¼ í¬í•¨)"""
    SHEET_NAME = "ë“±ë¡ê°¯ìˆ˜"
    HEADERS = ["ìŠ¤í† ì–´ëª…", "ì „ì²´", "íŒë§¤ì¤‘", "íŒë§¤ì¤‘ì§€", "ìŠ¹ì¸ëŒ€ê¸°", "ë§ˆì§€ë§‰ë“±ë¡ì¼", "updated_at"]

    try:
        # ì‹œíŠ¸ ê°€ì ¸ì˜¤ê¸° ë˜ëŠ” ìƒì„±
        try:
            ws = gsheet.sheet.worksheet(SHEET_NAME)
        except:
            ws = gsheet.sheet.add_worksheet(title=SHEET_NAME, rows=200, cols=len(HEADERS))
            ws.update(range_name="1:1", values=[HEADERS], value_input_option="RAW")

        # í—¤ë” í™•ì¸ ë° ë§ˆì§€ë§‰ë“±ë¡ì¼ ì»¬ëŸ¼ ì¶”ê°€
        all_vals = ws.get_all_values() or []
        if not all_vals:
            ws.update(range_name="1:1", values=[HEADERS], value_input_option="RAW")
            all_vals = [HEADERS]

        headers = [h.strip() for h in all_vals[0]] if all_vals else HEADERS

        # ë§ˆì§€ë§‰ë“±ë¡ì¼ ì»¬ëŸ¼ì´ ì—†ìœ¼ë©´ ì¶”ê°€
        if "ë§ˆì§€ë§‰ë“±ë¡ì¼" not in headers:
            # updated_at ì•ì— ì¶”ê°€
            try:
                updated_idx = headers.index("updated_at")
                headers.insert(updated_idx, "ë§ˆì§€ë§‰ë“±ë¡ì¼")
            except ValueError:
                headers.append("ë§ˆì§€ë§‰ë“±ë¡ì¼")
            ws.update(range_name="1:1", values=[headers], value_input_option="RAW")
            print(f"[ë“±ë¡ê°¯ìˆ˜] ë§ˆì§€ë§‰ë“±ë¡ì¼ ì»¬ëŸ¼ ì¶”ê°€ë¨")

        # store_name/ìŠ¤í† ì–´ëª… ì—´ ì¸ë±ìŠ¤
        store_col_idx = 0
        for i, h in enumerate(headers):
            if h in ["store_name", "ìŠ¤í† ì–´ëª…"]:
                store_col_idx = i
                break

        # ê¸°ì¡´ í–‰ ì°¾ê¸°
        row_map = {}
        for i, row in enumerate(all_vals[1:], start=2):
            if len(row) > store_col_idx and row[store_col_idx]:
                row_map[row[store_col_idx]] = i

        # ë°ì´í„° ì¤€ë¹„
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        row_values = {
            "ìŠ¤í† ì–´ëª…": store_name,
            "store_name": store_name,
            "ì „ì²´": counts.get("ì „ì²´", 0),
            "íŒë§¤ì¤‘": counts.get("íŒë§¤ì¤‘", 0),
            "íŒë§¤ì¤‘ì§€": counts.get("íŒë§¤ì¤‘ì§€", 0),
            "ìŠ¹ì¸ëŒ€ê¸°": counts.get("ìŠ¹ì¸ëŒ€ê¸°", 0),
            "ë§ˆì§€ë§‰ë“±ë¡ì¼": last_reg_date,
            "updated_at": now_str
        }
        row_to_write = [row_values.get(h, "") for h in headers]

        # ê¸°ì¡´ í–‰ ì—…ë°ì´íŠ¸ ë˜ëŠ” ìƒˆ í–‰ ì¶”ê°€
        target_row = row_map.get(store_name)
        if target_row is None:
            ws.append_row(row_to_write, value_input_option="RAW")
        else:
            def col_letter(n):
                result = ""
                while n > 0:
                    n, remainder = divmod(n - 1, 26)
                    result = chr(65 + remainder) + result
                return result

            end_col = col_letter(len(headers))
            rng = f"A{target_row}:{end_col}{target_row}"
            ws.update(range_name=rng, values=[row_to_write], value_input_option="RAW")

        print(f"[ë“±ë¡ê°¯ìˆ˜] {store_name} ì‹œíŠ¸ ê¸°ë¡ ì™„ë£Œ (ë§ˆì§€ë§‰ë“±ë¡ì¼: {last_reg_date})")
        return True
    except Exception as e:
        print(f"[ë“±ë¡ê°¯ìˆ˜] {store_name} ì‹œíŠ¸ ê¸°ë¡ ì˜¤ë¥˜: {e}")
        return False

@app.get("/api/allinone/stores")
async def get_allinone_stores(request: Request, platform: str, task: str):
    """ì‘ì—…ë³„ ìŠ¤í† ì–´ ëª©ë¡ ì¡°íšŒ - ìƒí’ˆìˆ˜ ì •ë³´ í¬í•¨"""
    require_permission(request, "edit")
    
    # ì‘ì—…-íƒ­ ë§¤í•‘ (ê° ì‘ì—…ë³„ ì‹œíŠ¸ ì‚¬ìš©)
    task_to_sheet = {
        # ìŠ¤ë§ˆíŠ¸ìŠ¤í† ì–´ - ê° ì‘ì—…ë³„ ì‹œíŠ¸
        "ë“±ë¡ê°¯ìˆ˜": "stores",
        "ë°°ì†¡ì½”ë“œ": "stores",
        "ë°°ì†¡ë³€ê²½": "ë°°ì†¡ë³€ê²½",
        "ìƒí’ˆì‚­ì œ": "ìƒí’ˆì‚­ì œ",
        "í˜œíƒì„¤ì •": "í˜œíƒì„¤ì •",
        "ì¤‘ë³µì‚­ì œ": "ì¤‘ë³µì‚­ì œ",
        "KCì¸ì¦": "KCì¸ì¦",
        "ê¸°íƒ€ê¸°ëŠ¥": "ê¸°íƒ€ê¸°ëŠ¥",
        "ë§¤ì¶œì¡°íšŒ": "stores",
        # 11ë²ˆê°€
        "íŒë§¤ì¤‘ì§€": "11ë²ˆê°€",
        "íŒë§¤ì¬ê°œ": "11ë²ˆê°€",
        "11stë§¤ì¶œ": "11ë²ˆê°€"
    }
    
    # 11ë²ˆê°€ í”Œë«í¼ì€ 11ë²ˆê°€ ì‹œíŠ¸ ì‚¬ìš©
    if platform == "11ë²ˆê°€":
        sheet_name = "11ë²ˆê°€"
    else:
        sheet_name = task_to_sheet.get(task, "stores")
    
    try:
        # ê³„ì •ëª©ë¡ íƒ­ì—ì„œ ì†Œìœ ì/ìš©ë„ ì •ë³´ ë¡œë“œ (í”Œë«í¼ë³„ë¡œ êµ¬ë¶„)
        # ê³„ì •ëª©ë¡ì˜ ì‡¼í•‘ëª° ë³„ì¹­ì—ì„œ ì–¸ë”ë°” ì•ë¶€ë¶„ ì œê±° í›„ ë§¤ì¹­
        account_info = {}
        try:
            ws_accounts = gsheet.sheet.worksheet("ê³„ì •ëª©ë¡")
            acc_data = ws_accounts.get_all_values()
            if acc_data and len(acc_data) > 1:
                acc_headers = acc_data[0]
                acc_name_idx = None
                owner_idx = None
                usage_idx = None
                platform_idx = None
                
                for i, h in enumerate(acc_headers):
                    if h in ["ì‡¼í•‘ëª° ë³„ì¹­", "ê³„ì •ëª…", "account_name", "ìŠ¤í† ì–´ëª…", "shop_alias"]:
                        acc_name_idx = i
                    elif h in ["ì†Œìœ ì", "owner"]:
                        owner_idx = i
                    elif h in ["ìš©ë„", "usage"]:
                        usage_idx = i
                    elif h in ["í”Œë«í¼", "platform"]:
                        platform_idx = i
                
                print(f"[ì˜¬ì¸ì›] ê³„ì •ëª©ë¡ í—¤ë”: acc_name_idx={acc_name_idx}, owner_idx={owner_idx}, usage_idx={usage_idx}, platform_idx={platform_idx}")
                
                if acc_name_idx is not None and platform_idx is not None:
                    for row in acc_data[1:]:
                        if len(row) > acc_name_idx and len(row) > platform_idx:
                            acc_name_raw = row[acc_name_idx].strip()
                            acc_platform = row[platform_idx].strip()
                            
                            # í”Œë«í¼ ë§¤ì¹­ (ìŠ¤ë§ˆíŠ¸ìŠ¤í† ì–´/11ë²ˆê°€ ë“±)
                            platform_match = False
                            if platform == "ìŠ¤ë§ˆíŠ¸ìŠ¤í† ì–´" and ("ìŠ¤ë§ˆíŠ¸" in acc_platform or "ë„¤ì´ë²„" in acc_platform):
                                platform_match = True
                            elif platform == "11ë²ˆê°€" and "11" in acc_platform:
                                platform_match = True
                            elif platform == acc_platform:
                                platform_match = True
                            
                            if acc_name_raw and platform_match:
                                owner = row[owner_idx].strip() if owner_idx is not None and len(row) > owner_idx else ""
                                usage = row[usage_idx].strip() if usage_idx is not None and len(row) > usage_idx else ""
                                
                                # ì‡¼í•‘ëª° ë³„ì¹­ì—ì„œ ì–¸ë”ë°” ìˆìœ¼ë©´ ì•ë¶€ë¶„ ì œê±° (01_í‘¸ë¡œí…Œì¹´ -> í‘¸ë¡œí…Œì¹´)
                                if "_" in acc_name_raw:
                                    match_name = acc_name_raw.split("_", 1)[1]
                                else:
                                    match_name = acc_name_raw
                                
                                account_info[match_name] = {"owner": owner, "usage": usage}
                                        
                    print(f"[ì˜¬ì¸ì›] ê³„ì •ëª©ë¡ ë¡œë“œ ì™„ë£Œ ({platform}): {len(account_info)}ê°œ, ì˜ˆì‹œ: {list(account_info.items())[:5]}")
        except Exception as e:
            print(f"[ì˜¬ì¸ì›] ê³„ì •ëª©ë¡ ì‹œíŠ¸ ë¡œë“œ ì˜¤ë¥˜: {e}")
        
        # ë°°ì†¡ì½”ë“œ ì‹œíŠ¸ ë°ì´í„° ë¡œë“œ (ì¶œê³ ì§€ ì½”ë“œ)
        shipping_codes = {}
        if task in ["ë°°ì†¡ì½”ë“œ", "ë°°ì†¡ë³€ê²½"]:
            try:
                ws_shipping = gsheet.sheet.worksheet("ë°°ì†¡ì½”ë“œ")
                ship_data = ws_shipping.get_all_values()
                if ship_data and len(ship_data) > 1:
                    ship_headers = ship_data[0]
                    ship_cols = {h: i for i, h in enumerate(ship_headers)}
                    
                    for row in ship_data[1:]:
                        store = row[ship_cols.get("ìŠ¤í† ì–´ëª…", 0)].strip() if len(row) > ship_cols.get("ìŠ¤í† ì–´ëª…", 0) else ""
                        if store:
                            shipping_codes[store] = {
                                "êµ­ë‚´ì¶œê³ ì§€": row[ship_cols["êµ­ë‚´ì¶œê³ ì§€"]] if "êµ­ë‚´ì¶œê³ ì§€" in ship_cols and len(row) > ship_cols["êµ­ë‚´ì¶œê³ ì§€"] else "",
                                "í•´ì™¸ì¶œê³ ì§€": row[ship_cols["í•´ì™¸ì¶œê³ ì§€"]] if "í•´ì™¸ì¶œê³ ì§€" in ship_cols and len(row) > ship_cols["í•´ì™¸ì¶œê³ ì§€"] else "",
                                "ë°˜í’ˆì§€": row[ship_cols["ë°˜í’ˆì§€"]] if "ë°˜í’ˆì§€" in ship_cols and len(row) > ship_cols["ë°˜í’ˆì§€"] else "",
                                "updated_at": row[ship_cols["updated_at"]] if "updated_at" in ship_cols and len(row) > ship_cols["updated_at"] else ""
                            }
                    print(f"[ì˜¬ì¸ì›] ë°°ì†¡ì½”ë“œ ì‹œíŠ¸ ë¡œë“œ: {len(shipping_codes)}ê°œ")
            except Exception as e:
                print(f"[ì˜¬ì¸ì›] ë°°ì†¡ì½”ë“œ ì‹œíŠ¸ ë¡œë“œ ì˜¤ë¥˜: {e}")
        
        # ë°°ì†¡ë³€ê²½ ì‹œíŠ¸ ë°ì´í„° ë¡œë“œ
        delivery_info = {}
        if task == "ë°°ì†¡ë³€ê²½":
            try:
                ws_delivery = gsheet.sheet.worksheet("ë°°ì†¡ë³€ê²½")
                del_data = ws_delivery.get_all_values()
                if del_data and len(del_data) > 1:
                    del_headers = del_data[0]
                    del_cols = {h: i for i, h in enumerate(del_headers)}
                    
                    # ìŠ¤í† ì–´ëª… ì»¬ëŸ¼ ì°¾ê¸° (í•œê¸€/ì˜ì–´ ë‘˜ ë‹¤ ì§€ì›)
                    store_name_idx = None
                    for i, h in enumerate(del_headers):
                        if h in ["ìŠ¤í† ì–´ëª…", "store_name", "ê³„ì •ëª…"]:
                            store_name_idx = i
                            break
                    if store_name_idx is None:
                        store_name_idx = 1  # fallback
                    
                    for row in del_data[1:]:
                        store = row[store_name_idx].strip() if len(row) > store_name_idx else ""
                        if store:
                            delivery_info[store] = {
                                "target_limit": row[del_cols["target_limit"]] if "target_limit" in del_cols and len(row) > del_cols["target_limit"] else "",
                                "shippingAddressId": row[del_cols["shippingAddressId"]] if "shippingAddressId" in del_cols and len(row) > del_cols["shippingAddressId"] else "",
                                "differentialFeeByArea": row[del_cols["differentialFeeByArea"]] if "differentialFeeByArea" in del_cols and len(row) > del_cols["differentialFeeByArea"] else "",
                                "cutofftime": row[del_cols["cutofftime"]] if "cutofftime" in del_cols and len(row) > del_cols["cutofftime"] else "",
                                "updated_at": row[del_cols["updated_at"]] if "updated_at" in del_cols and len(row) > del_cols["updated_at"] else ""
                            }
                    print(f"[ì˜¬ì¸ì›] ë°°ì†¡ë³€ê²½ ì‹œíŠ¸ ë¡œë“œ: {len(delivery_info)}ê°œ")
            except Exception as e:
                print(f"[ì˜¬ì¸ì›] ë°°ì†¡ë³€ê²½ ì‹œíŠ¸ ë¡œë“œ ì˜¤ë¥˜: {e}")
        
        # í˜œíƒì„¤ì • ì‹œíŠ¸ ë°ì´í„° ë¡œë“œ
        benefit_info = {}
        if task == "í˜œíƒì„¤ì •":
            try:
                ws_benefit = gsheet.sheet.worksheet("í˜œíƒì„¤ì •")
                ben_data = ws_benefit.get_all_values()
                if ben_data and len(ben_data) > 1:
                    ben_headers = ben_data[0]
                    ben_cols = {h: i for i, h in enumerate(ben_headers)}
                    
                    # ìŠ¤í† ì–´ëª… ì»¬ëŸ¼ ì°¾ê¸° (í•œê¸€/ì˜ì–´ ë‘˜ ë‹¤ ì§€ì›)
                    store_name_idx = None
                    for i, h in enumerate(ben_headers):
                        if h in ["ìŠ¤í† ì–´ëª…", "store_name", "ê³„ì •ëª…"]:
                            store_name_idx = i
                            break
                    if store_name_idx is None:
                        store_name_idx = 0  # fallback
                    
                    for row in ben_data[1:]:
                        store = row[store_name_idx].strip() if len(row) > store_name_idx else ""
                        if store:
                            benefit_info[store] = {
                                "í›„ê¸°í¬ì¸íŠ¸": row[ben_cols["í›„ê¸°í¬ì¸íŠ¸"]] if "í›„ê¸°í¬ì¸íŠ¸" in ben_cols and len(row) > ben_cols["í›„ê¸°í¬ì¸íŠ¸"] else "",
                                "í¬í† í›„ê¸°í¬ì¸íŠ¸": row[ben_cols["í¬í† í›„ê¸°í¬ì¸íŠ¸"]] if "í¬í† í›„ê¸°í¬ì¸íŠ¸" in ben_cols and len(row) > ben_cols["í¬í† í›„ê¸°í¬ì¸íŠ¸"] else "",
                                "í•œë‹¬í›„ê¸°í¬ì¸íŠ¸": row[ben_cols["í•œë‹¬í›„ê¸°í¬ì¸íŠ¸"]] if "í•œë‹¬í›„ê¸°í¬ì¸íŠ¸" in ben_cols and len(row) > ben_cols["í•œë‹¬í›„ê¸°í¬ì¸íŠ¸"] else "",
                                "í•œë‹¬í¬í† í›„ê¸°í¬ì¸íŠ¸": row[ben_cols["í•œë‹¬í¬í† í›„ê¸°í¬ì¸íŠ¸"]] if "í•œë‹¬í¬í† í›„ê¸°í¬ì¸íŠ¸" in ben_cols and len(row) > ben_cols["í•œë‹¬í¬í† í›„ê¸°í¬ì¸íŠ¸"] else "",
                                "ì´ë²¤íŠ¸ë¬¸êµ¬": row[ben_cols["ì´ë²¤íŠ¸ë¬¸êµ¬"]] if "ì´ë²¤íŠ¸ë¬¸êµ¬" in ben_cols and len(row) > ben_cols["ì´ë²¤íŠ¸ë¬¸êµ¬"] else "",
                                "ì‚¬ì€í’ˆ": row[ben_cols["ì‚¬ì€í’ˆ"]] if "ì‚¬ì€í’ˆ" in ben_cols and len(row) > ben_cols["ì‚¬ì€í’ˆ"] else "",
                                "ìµœì†ŒíŒë§¤ê°€": row[ben_cols["ìµœì†ŒíŒë§¤ê°€"]] if "ìµœì†ŒíŒë§¤ê°€" in ben_cols and len(row) > ben_cols["ìµœì†ŒíŒë§¤ê°€"] else "",
                                "ë³µìˆ˜êµ¬ë§¤": row[ben_cols["ë³µìˆ˜êµ¬ë§¤"]] if "ë³µìˆ˜êµ¬ë§¤" in ben_cols and len(row) > ben_cols["ë³µìˆ˜êµ¬ë§¤"] else "",
                                "ë³µìˆ˜êµ¬ë§¤í• ì¸": row[ben_cols["ë³µìˆ˜êµ¬ë§¤í• ì¸"]] if "ë³µìˆ˜êµ¬ë§¤í• ì¸" in ben_cols and len(row) > ben_cols["ë³µìˆ˜êµ¬ë§¤í• ì¸"] else "",
                                "ê²°ê³¼": row[ben_cols["ê²°ê³¼"]] if "ê²°ê³¼" in ben_cols and len(row) > ben_cols["ê²°ê³¼"] else "",
                                "updated_at": row[ben_cols["updated_at"]] if "updated_at" in ben_cols and len(row) > ben_cols["updated_at"] else ""
                            }
                    print(f"[ì˜¬ì¸ì›] í˜œíƒì„¤ì • ì‹œíŠ¸ ë¡œë“œ: {len(benefit_info)}ê°œ")
            except Exception as e:
                print(f"[ì˜¬ì¸ì›] í˜œíƒì„¤ì • ì‹œíŠ¸ ë¡œë“œ ì˜¤ë¥˜: {e}")
        
        # ë“±ë¡ê°¯ìˆ˜ ì‹œíŠ¸ì—ì„œ ìƒí’ˆìˆ˜ ì •ë³´ ë¡œë“œ
        product_counts = {}
        store_order = []  # êµ¬ê¸€ì‹œíŠ¸ ìˆœì„œ ìœ ì§€
        try:
            ws_counts = gsheet.sheet.worksheet("ë“±ë¡ê°¯ìˆ˜")
            counts_data = ws_counts.get_all_values()
            if counts_data and len(counts_data) > 1:
                headers = counts_data[0]
                name_idx = None
                total_idx = None
                on_sale_idx = None
                suspended_idx = None
                pending_idx = None  # ìŠ¹ì¸ëŒ€ê¸°
                updated_idx = None
                
                for i, h in enumerate(headers):
                    if h == "ìŠ¤í† ì–´ëª…":
                        name_idx = i
                    elif h == "ì „ì²´":
                        total_idx = i
                    elif h == "íŒë§¤ì¤‘":
                        on_sale_idx = i
                    elif h == "íŒë§¤ì¤‘ì§€":
                        suspended_idx = i
                    elif h == "ìŠ¹ì¸ëŒ€ê¸°":
                        pending_idx = i
                    elif h == "updated_at":
                        updated_idx = i
                
                if name_idx is not None:
                    for row_idx, row in enumerate(counts_data[1:], start=1):
                        if len(row) > name_idx:
                            store = row[name_idx].strip()
                            if store:
                                store_order.append(store)
                                product_counts[store] = {
                                    "row_num": row_idx,
                                    "total": int(row[total_idx]) if total_idx and len(row) > total_idx and row[total_idx].isdigit() else 0,
                                    "on_sale": int(row[on_sale_idx]) if on_sale_idx and len(row) > on_sale_idx and row[on_sale_idx].isdigit() else 0,
                                    "suspended": int(row[suspended_idx]) if suspended_idx and len(row) > suspended_idx and row[suspended_idx].isdigit() else 0,
                                    "pending": int(row[pending_idx]) if pending_idx and len(row) > pending_idx and row[pending_idx].isdigit() else 0,
                                    "updated_at": row[updated_idx] if updated_idx and len(row) > updated_idx else ""
                                }
        except Exception as e:
            print(f"[ì˜¬ì¸ì›] ë“±ë¡ê°¯ìˆ˜ ì‹œíŠ¸ ë¡œë“œ ì˜¤ë¥˜: {e}")
        
        # 11ë²ˆê°€ë„ ê°™ì€ ìŠ¤í”„ë ˆë“œì‹œíŠ¸, íƒ­ë§Œ ë‹¤ë¦„
        if sheet_name == "11ë²ˆê°€":
            ws = gsheet.sheet.worksheet("11ë²ˆê°€")
            
            all_values = ws.get_all_values()
            
            if not all_values or len(all_values) < 2:
                return {"stores": []}
            
            headers = all_values[0]
            print(f"[ì˜¬ì¸ì›] 11ë²ˆê°€ í—¤ë”: {headers[:15]}")
            
            # ì»¬ëŸ¼ ì¸ë±ìŠ¤ ì°¾ê¸°
            col_idx = {}
            for i, h in enumerate(headers):
                if h in ["store_name", "ì‡¼í•‘ëª° ë³„ì¹­", "ìŠ¤í† ì–´ëª…"]:
                    col_idx["ìŠ¤í† ì–´ëª…"] = i
                elif h in ["active", "í™œì„±", "ì‚¬ìš©"]:
                    col_idx["active"] = i
                elif h == "ì „ì²´":
                    col_idx["total"] = i
                elif h == "íŒë§¤ì¤‘":
                    col_idx["on_sale"] = i
                elif h == "íŒë§¤ì¤‘ì§€":
                    col_idx["suspended"] = i
                elif h == "ìŠ¹ì¸ëŒ€ê¸°":
                    col_idx["pending"] = i
                elif h == "updated_at":
                    col_idx["updated_at"] = i
            
            # active ì—´ì´ ì—†ìœ¼ë©´ ì²« ë²ˆì§¸ ì—´
            if "active" not in col_idx:
                col_idx["active"] = 0
            
            if "ìŠ¤í† ì–´ëª…" not in col_idx:
                print(f"[ì˜¬ì¸ì›] ìŠ¤í† ì–´ëª… ì—´ì„ ì°¾ì„ ìˆ˜ ì—†ìŒ")
                return {"stores": [], "error": "ìŠ¤í† ì–´ëª… ì—´ ì—†ìŒ"}
            
            stores = []
            for row_idx, row in enumerate(all_values[1:], start=1):
                if len(row) <= col_idx["ìŠ¤í† ì–´ëª…"]:
                    continue
                
                store_name = row[col_idx["ìŠ¤í† ì–´ëª…"]].strip() if len(row) > col_idx["ìŠ¤í† ì–´ëª…"] else ""
                active = row[col_idx["active"]].strip() if len(row) > col_idx["active"] else ""
                
                if not store_name:
                    continue
                
                is_active = str(active).upper() in ["TRUE", "ON", "Y", "1", "ì‚¬ìš©"]
                
                # 11ë²ˆê°€ ì‹œíŠ¸ì—ì„œ ì§ì ‘ ìˆ˜ëŸ‰ ì½ê¸°
                def get_int(key):
                    if key in col_idx and len(row) > col_idx[key]:
                        val = row[col_idx[key]]
                        return int(val) if val.isdigit() else 0
                    return 0
                
                # ê³„ì •ëª©ë¡ì—ì„œ ì†Œìœ ì/ìš©ë„ ë§¤ì¹­
                # store_nameì—ì„œ ì–¸ë”ë°” ì•ë¶€ë¶„ ì œê±° í›„ ë§¤ì¹­ (01_ë£¨ë¯¸ì¼“1 -> ë£¨ë¯¸ì¼“1)
                match_name = store_name
                if "_" in store_name:
                    match_name = store_name.split("_", 1)[1]
                acc_info = account_info.get(match_name, {})
                
                stores.append({
                    "row_num": row_idx,
                    "ìŠ¤í† ì–´ëª…": store_name,
                    "active": is_active,
                    "total": get_int("total"),
                    "on_sale": get_int("on_sale"),
                    "suspended": get_int("suspended"),
                    "pending": get_int("pending"),
                    "owner": acc_info.get("owner", ""),
                    "usage": acc_info.get("usage", ""),
                    "updated_at": row[col_idx["updated_at"]] if "updated_at" in col_idx and len(row) > col_idx["updated_at"] else ""
                })
            
            print(f"[ì˜¬ì¸ì›] 11ë²ˆê°€ ìŠ¤í† ì–´ {len(stores)}ê°œ ë¡œë“œ")
            return {"stores": stores}
        
        # ìŠ¤ë§ˆíŠ¸ìŠ¤í† ì–´ (ê¸°ì¡´ ë¡œì§)
        ws = gsheet.sheet.worksheet(sheet_name)
        
        # stores íƒ­ì€ A1ì´ ì‘ì—…ì„ íƒ ì¹¸ì´ë¯€ë¡œ get_all_values ì‚¬ìš©
        all_values = ws.get_all_values()
        
        if not all_values or len(all_values) < 2:
            return {"stores": []}
        
        # ì²« ë²ˆì§¸ í–‰ì€ í—¤ë” (A1ì€ ì‘ì—…ì„ íƒ ë“œë¡­ë‹¤ìš´)
        headers = all_values[0]
        print(f"[ì˜¬ì¸ì›] {sheet_name} í—¤ë”: {headers[:5]}")
        
        # ìŠ¤í† ì–´ëª… ì—´ ì¸ë±ìŠ¤ ì°¾ê¸°
        store_col = None
        for i, h in enumerate(headers):
            if h in ["store_name", "ì‡¼í•‘ëª° ë³„ì¹­", "ìŠ¤í† ì–´ëª…", "ê³„ì •ëª…"]:
                store_col = i
                break
        
        if store_col is None:
            print(f"[ì˜¬ì¸ì›] ìŠ¤í† ì–´ëª… ì—´ì„ ì°¾ì„ ìˆ˜ ì—†ìŒ")
            return {"stores": [], "error": "ìŠ¤í† ì–´ëª… ì—´ ì—†ìŒ"}
        
        # êµ¬ê¸€ì‹œíŠ¸ ìˆœì„œëŒ€ë¡œ ìŠ¤í† ì–´ ëª©ë¡ ìƒì„±
        stores = []
        for row_idx, row in enumerate(all_values[1:], start=1):  # í—¤ë” ì œì™¸
            if len(row) <= store_col:
                continue
            
            # Aì—´(ì¸ë±ìŠ¤ 0)ì´ active
            active = row[0] if len(row) > 0 else ""
            store_name = row[store_col] if len(row) > store_col else ""
            
            if not store_name:
                continue
            
            is_active = str(active).upper() == "TRUE"
            counts = product_counts.get(store_name, {})
            
            # ê³„ì •ëª©ë¡ì—ì„œ ì†Œìœ ì/ìš©ë„ ë§¤ì¹­
            # storesì˜ store_nameê³¼ ê³„ì •ëª©ë¡ì˜ (ì–¸ë”ë°” ì œê±°ëœ) ì‡¼í•‘ëª° ë³„ì¹­ ë§¤ì¹­
            acc_info = account_info.get(store_name, {})
            
            # ê¸°ë³¸ ë°ì´í„°
            store_data = {
                "row_num": counts.get("row_num", row_idx),
                "ìŠ¤í† ì–´ëª…": store_name,
                "active": is_active,
                "total": counts.get("total", 0),
                "on_sale": counts.get("on_sale", 0),
                "suspended": counts.get("suspended", 0),
                "pending": counts.get("pending", 0),
                "owner": acc_info.get("owner", ""),
                "usage": acc_info.get("usage", ""),
                "updated_at": counts.get("updated_at", "")
            }
            
            # ì–¸ë”ë°” ë’¤ ì´ë¦„ ì¶”ì¶œ (ë§¤ì¹­ìš©)
            match_name = store_name
            if "_" in store_name:
                match_name = store_name.split("_", 1)[1]
            
            # ë°°ì†¡ì½”ë“œ ë°ì´í„° ì¶”ê°€
            if task == "ë°°ì†¡ì½”ë“œ":
                ship_info = shipping_codes.get(store_name) or shipping_codes.get(match_name) or {}
                store_data["êµ­ë‚´ì¶œê³ ì§€"] = ship_info.get("êµ­ë‚´ì¶œê³ ì§€", "")
                store_data["í•´ì™¸ì¶œê³ ì§€"] = ship_info.get("í•´ì™¸ì¶œê³ ì§€", "")
                store_data["ë°˜í’ˆì§€"] = ship_info.get("ë°˜í’ˆì§€", "")
                store_data["shipping_updated_at"] = ship_info.get("updated_at", "")
            
            # ë°°ì†¡ë³€ê²½ ë°ì´í„° ì¶”ê°€ (store_name ë˜ëŠ” ì–¸ë”ë°” ë’¤ ì´ë¦„ìœ¼ë¡œ ë§¤ì¹­)
            if task == "ë°°ì†¡ë³€ê²½":
                del_info = delivery_info.get(store_name) or delivery_info.get(match_name) or {}
                ship_info = shipping_codes.get(store_name) or shipping_codes.get(match_name) or {}
                store_data["target_limit"] = del_info.get("target_limit", "")
                store_data["shippingAddressId"] = del_info.get("shippingAddressId", "")
                store_data["differentialFeeByArea"] = del_info.get("differentialFeeByArea", "")
                store_data["cutofftime"] = del_info.get("cutofftime", "")
                store_data["delivery_updated_at"] = del_info.get("updated_at", "")
                # ë°°ì†¡ì½”ë“œì—ì„œ ì¶œê³ ì§€ ì½”ë“œ (ë“œë¡­ë‹¤ìš´ ë³€í™˜ìš©)
                store_data["êµ­ë‚´ì¶œê³ ì§€ì½”ë“œ"] = ship_info.get("êµ­ë‚´ì¶œê³ ì§€", "")
                store_data["í•´ì™¸ì¶œê³ ì§€ì½”ë“œ"] = ship_info.get("í•´ì™¸ì¶œê³ ì§€", "")
            
            # í˜œíƒì„¤ì • ë°ì´í„° ì¶”ê°€ (store_name ë˜ëŠ” ì–¸ë”ë°” ë’¤ ì´ë¦„ìœ¼ë¡œ ë§¤ì¹­)
            if task == "í˜œíƒì„¤ì •":
                ben_info = benefit_info.get(store_name) or benefit_info.get(match_name) or {}
                store_data["í›„ê¸°í¬ì¸íŠ¸"] = ben_info.get("í›„ê¸°í¬ì¸íŠ¸", "")
                store_data["í¬í† í›„ê¸°í¬ì¸íŠ¸"] = ben_info.get("í¬í† í›„ê¸°í¬ì¸íŠ¸", "")
                store_data["í•œë‹¬í›„ê¸°í¬ì¸íŠ¸"] = ben_info.get("í•œë‹¬í›„ê¸°í¬ì¸íŠ¸", "")
                store_data["í•œë‹¬í¬í† í›„ê¸°í¬ì¸íŠ¸"] = ben_info.get("í•œë‹¬í¬í† í›„ê¸°í¬ì¸íŠ¸", "")
                store_data["ì´ë²¤íŠ¸ë¬¸êµ¬"] = ben_info.get("ì´ë²¤íŠ¸ë¬¸êµ¬", "")
                store_data["ì‚¬ì€í’ˆ"] = ben_info.get("ì‚¬ì€í’ˆ", "")
                store_data["ìµœì†ŒíŒë§¤ê°€"] = ben_info.get("ìµœì†ŒíŒë§¤ê°€", "")
                store_data["ë³µìˆ˜êµ¬ë§¤"] = ben_info.get("ë³µìˆ˜êµ¬ë§¤", "")
                store_data["ë³µìˆ˜êµ¬ë§¤í• ì¸"] = ben_info.get("ë³µìˆ˜êµ¬ë§¤í• ì¸", "")
                store_data["benefit_result"] = ben_info.get("ê²°ê³¼", "")
                store_data["benefit_updated_at"] = ben_info.get("updated_at", "")
            
            stores.append(store_data)
        
        print(f"[ì˜¬ì¸ì›] {platform}/{task} ìŠ¤í† ì–´ {len(stores)}ê°œ ë¡œë“œ")
        return {"stores": stores}
    except Exception as e:
        print(f"[ì˜¬ì¸ì›] ìŠ¤í† ì–´ ëª©ë¡ ì¡°íšŒ ì˜¤ë¥˜: {e}")
        import traceback
        traceback.print_exc()
        return {"stores": [], "error": str(e)}

@app.post("/api/allinone/update-active")
async def update_allinone_active(request: Request, req: AioUpdateActiveRequest):
    """êµ¬ê¸€ì‹œíŠ¸ active ì—´ ì—…ë°ì´íŠ¸"""
    require_permission(request, "edit")
    
    # ì‘ì—…ë³„ ì‹œíŠ¸ ë§¤í•‘ (ê° ì‘ì—…ë³„ ì‹œíŠ¸ ì‚¬ìš©)
    task_to_sheet = {
        "ë“±ë¡ê°¯ìˆ˜": "stores",
        "ë°°ì†¡ì½”ë“œ": "stores",
        "ë°°ì†¡ë³€ê²½": "ë°°ì†¡ë³€ê²½",
        "ìƒí’ˆì‚­ì œ": "ìƒí’ˆì‚­ì œ",
        "í˜œíƒì„¤ì •": "í˜œíƒì„¤ì •",
        "ì¤‘ë³µì‚­ì œ": "ì¤‘ë³µì‚­ì œ",
        "KCì¸ì¦": "KCì¸ì¦",
        "ê¸°íƒ€ê¸°ëŠ¥": "ê¸°íƒ€ê¸°ëŠ¥",
        "íŒë§¤ì¤‘ì§€": "11ë²ˆê°€",
        "íŒë§¤ì¬ê°œ": "11ë²ˆê°€"
    }
    
    # 11ë²ˆê°€ í”Œë«í¼ì˜ ë“±ë¡ê°¯ìˆ˜ëŠ” 11ë²ˆê°€ ì‹œíŠ¸ ì‚¬ìš©
    if req.platform == "11ë²ˆê°€" and req.task == "ë“±ë¡ê°¯ìˆ˜":
        sheet_name = "11ë²ˆê°€"
    else:
        sheet_name = task_to_sheet.get(req.task, "stores")
    
    try:
        # 11ë²ˆê°€ë„ ê°™ì€ ìŠ¤í”„ë ˆë“œì‹œíŠ¸
        if sheet_name == "11ë²ˆê°€":
            ws = gsheet.sheet.worksheet("11ë²ˆê°€")
        else:
            ws = gsheet.sheet.worksheet(sheet_name)
        
        all_data = ws.get_all_values()
        
        if not all_data:
            return {"success": False, "message": "ì‹œíŠ¸ê°€ ë¹„ì–´ìˆìŠµë‹ˆë‹¤"}
        
        headers = all_data[0]
        
        # active/store_name ì—´ ì¸ë±ìŠ¤ ì°¾ê¸°
        active_col = 0  # ê¸°ë³¸ê°’
        store_col = None
        
        for i, h in enumerate(headers):
            if h in ["active", "í™œì„±", "ì‚¬ìš©"]:
                active_col = i
            if h in ["store_name", "ì‡¼í•‘ëª° ë³„ì¹­", "ìŠ¤í† ì–´ëª…"]:
                store_col = i
        
        if store_col is None:
            return {"success": False, "message": "ìŠ¤í† ì–´ëª… ì—´ì„ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤"}
        
        # ì—…ë°ì´íŠ¸í•  ì…€ ëª©ë¡ ìƒì„±
        updates = []
        active_col_letter = chr(65 + active_col) if active_col < 26 else f"{chr(64 + active_col // 26)}{chr(65 + active_col % 26)}"
        
        for row_idx, row in enumerate(all_data[1:], start=2):
            if len(row) <= store_col:
                continue
            
            store_name = row[store_col].strip()
            if not store_name:
                continue
            
            new_value = "TRUE" if store_name in req.active_stores else "FALSE"
            cell = f"{active_col_letter}{row_idx}"
            updates.append({"range": cell, "values": [[new_value]]})
        
        # ì¼ê´„ ì—…ë°ì´íŠ¸
        if updates:
            ws.batch_update(updates)
        
        # A1ì— ì‘ì—…ëª… ì„¤ì • (stores íƒ­ì¸ ê²½ìš°)
        if sheet_name == "stores":
            ws.update_acell("A1", req.task)
            print(f"[ì˜¬ì¸ì›] A1 ì…€ì— ì‘ì—… ì„¤ì •: {req.task}")
        
        print(f"[ì˜¬ì¸ì›] {len(req.active_stores)}ê°œ ìŠ¤í† ì–´ í™œì„±í™” ì™„ë£Œ")
        return {"success": True, "message": f"{len(req.active_stores)}ê°œ ìŠ¤í† ì–´ í™œì„±í™”"}
    except Exception as e:
        print(f"[ì˜¬ì¸ì›] active ì—…ë°ì´íŠ¸ ì˜¤ë¥˜: {e}")
        import traceback
        traceback.print_exc()
        return {"success": False, "message": str(e)}

@app.post("/api/allinone/run")
async def run_allinone_task(request: Request, req: AioRunRequest):
    """ì˜¬ì¸ì› í”„ë¡œê·¸ë¨ ì‹¤í–‰"""
    require_permission(request, "edit")

    # í”Œë«í¼ë³„ ìƒíƒœ ê°€ì ¸ì˜¤ê¸°
    platform_status = get_aio_status(req.platform, req.task)

    if platform_status["running"]:
        return {"success": False, "message": f"{req.platform} ì‘ì—…ì´ ì´ë¯¸ ì‹¤í–‰ ì¤‘ì…ë‹ˆë‹¤"}

    print(f"[ì˜¬ì¸ì›] ì‹¤í–‰ ìš”ì²­: {req.platform} / {req.task}, ì˜µì…˜: {req.options}")

    # ì‘ì—… ë¡œê·¸ ê¸°ë¡
    store_count = len(req.stores) if req.stores else 0
    store_names = ", ".join(req.stores[:5]) + ("..." if len(req.stores) > 5 else "") if req.stores else "ì „ì²´"
    log_work(f"ì˜¬ì¸ì›-{req.task}", f"{req.platform}", store_count, f"ëŒ€ìƒ: {store_names}", "ì›¹")
    
    # í”„ë¡œê·¸ë¨ ê²½ë¡œ
    if req.platform == "ìŠ¤ë§ˆíŠ¸ìŠ¤í† ì–´":
        script_path = r"C:\autosystem\smartstore_all_in_one_v1_1.py"

        try:
            # ë°°ì†¡ë³€ê²½ ì˜µì…˜ ì„¤ì •
            if req.task == "ë°°ì†¡ë³€ê²½" and req.options:
                ws = gsheet.sheet.worksheet("ë°°ì†¡ë³€ê²½")
                
                # ì„ íƒëœ ìŠ¤í† ì–´ ëª©ë¡
                selected_stores = set(req.stores) if req.stores else set()
                
                # ë°°ì†¡ë³€ê²½ ì‹œíŠ¸ì—ì„œ ì§ì ‘ active í™•ì¸
                delivery_data = ws.get_all_values()
                delivery_headers = delivery_data[0] if delivery_data else []
                
                # active, store_name ì»¬ëŸ¼ ì°¾ê¸°
                active_col_idx = None
                store_name_col_idx = None
                for i, h in enumerate(delivery_headers):
                    if h in ["active", "í™œì„±"]:
                        active_col_idx = i
                    if h in ["store_name", "ìŠ¤í† ì–´ëª…"]:
                        store_name_col_idx = i
                
                if active_col_idx is not None and store_name_col_idx is not None:
                    if req.options.get('mode') == 'count':
                        count = req.options.get('count', 100)
                        processed = []
                        for row_idx, row in enumerate(delivery_data[1:], start=2):
                            if len(row) > max(active_col_idx, store_name_col_idx):
                                is_active = str(row[active_col_idx]).upper() == "TRUE"
                                store_name = row[store_name_col_idx]
                                
                                if is_active and (not selected_stores or store_name in selected_stores):
                                    ws.update_acell(f"F{row_idx}", str(count))
                                    ws.update_acell(f"Y{row_idx}", "")
                                    processed.append(store_name)
                        print(f"[ì˜¬ì¸ì›] ë°°ì†¡ë³€ê²½ ìˆ˜ëŸ‰ {count}ê°œ: {', '.join(processed) if processed else 'ì—†ìŒ'}")
                    
                    elif req.options.get('mode') == 'date':
                        date_val = req.options.get('date', '')
                        processed = []
                        for row_idx, row in enumerate(delivery_data[1:], start=2):
                            if len(row) > max(active_col_idx, store_name_col_idx):
                                is_active = str(row[active_col_idx]).upper() == "TRUE"
                                store_name = row[store_name_col_idx]
                                
                                if is_active and (not selected_stores or store_name in selected_stores):
                                    ws.update_acell(f"F{row_idx}", "")
                                    ws.update_acell(f"Y{row_idx}", date_val)
                                    processed.append(store_name)
                        print(f"[ì˜¬ì¸ì›] ë°°ì†¡ë³€ê²½ ë‚ ì§œ {date_val}: {', '.join(processed) if processed else 'ì—†ìŒ'}")
            
            # í˜œíƒì„¤ì • ì˜µì…˜ ì„¤ì •
            elif req.task == "í˜œíƒì„¤ì •" and req.options:
                ws = gsheet.sheet.worksheet("í˜œíƒì„¤ì •")
                
                # ì„ íƒëœ ìŠ¤í† ì–´ ëª©ë¡
                selected_stores = set(req.stores) if req.stores else set()
                
                # í˜œíƒì„¤ì • ì‹œíŠ¸ì—ì„œ ì§ì ‘ active í™•ì¸
                benefit_data = ws.get_all_values()
                benefit_headers = benefit_data[0] if benefit_data else []
                
                # active, store_name ì»¬ëŸ¼ ì°¾ê¸°
                active_col_idx = None
                store_name_col_idx = None
                for i, h in enumerate(benefit_headers):
                    if h in ["active", "í™œì„±"]:
                        active_col_idx = i
                    if h in ["store_name", "ìŠ¤í† ì–´ëª…"]:
                        store_name_col_idx = i
                
                if active_col_idx is not None and store_name_col_idx is not None:
                    if req.options.get('date'):
                        date_val = req.options.get('date', '')
                        processed = []
                        for row_idx, row in enumerate(benefit_data[1:], start=2):
                            if len(row) > max(active_col_idx, store_name_col_idx):
                                is_active = str(row[active_col_idx]).upper() == "TRUE"
                                store_name = row[store_name_col_idx]
                                
                                if is_active and (not selected_stores or store_name in selected_stores):
                                    ws.update_acell(f"M{row_idx}", date_val)
                                    processed.append(store_name)
                        print(f"[ì˜¬ì¸ì›] í˜œíƒì„¤ì • ë‚ ì§œ {date_val}: {', '.join(processed) if processed else 'ì—†ìŒ'}")
            
            elif req.task == "ìƒí’ˆì‚­ì œ" and req.options:
                ws_delete = gsheet.sheet.worksheet("ìƒí’ˆì‚­ì œ")
                
                # ì„ íƒëœ ìŠ¤í† ì–´ ëª©ë¡
                selected_stores = set(req.stores) if req.stores else set()
                
                # ìƒí’ˆì‚­ì œ ì‹œíŠ¸ì—ì„œ ì§ì ‘ active í™•ì¸
                delete_data = ws_delete.get_all_values()
                delete_headers = delete_data[0] if delete_data else []
                
                # active, store_name ì»¬ëŸ¼ ì°¾ê¸°
                active_col_idx = None
                store_name_col_idx = None
                for i, h in enumerate(delete_headers):
                    if h in ["active", "í™œì„±"]:
                        active_col_idx = i
                    if h in ["store_name", "ìŠ¤í† ì–´ëª…"]:
                        store_name_col_idx = i
                
                if active_col_idx is None or store_name_col_idx is None:
                    print(f"[ì˜¬ì¸ì›] ìƒí’ˆì‚­ì œ ì‹œíŠ¸ì—ì„œ active ë˜ëŠ” store_name ì—´ì„ ì°¾ì„ ìˆ˜ ì—†ìŒ")
                else:
                    if req.options.get('delete_excess_only'):
                        # ì´ˆê³¼ë¶„ë§Œ ì‚­ì œ
                        delete_limit = req.options.get('delete_limit', 9500)
                        print(f"[ì˜¬ì¸ì›] ì´ˆê³¼ë¶„ë§Œ ì‚­ì œ ëª¨ë“œ: ê¸°ì¤€={delete_limit}ê°œ")
                        
                        try:
                            ws_counts = gsheet.sheet.worksheet("ë“±ë¡ê°¯ìˆ˜")
                            counts_data = ws_counts.get_all_records()
                            store_sales = {}
                            for row in counts_data:
                                store_name = row.get("ìŠ¤í† ì–´ëª…", "")
                                sales_count = int(row.get("íŒë§¤ì¤‘", 0) or 0)
                                if store_name:
                                    store_sales[store_name] = sales_count
                            
                            processed_stores = []
                            for row_idx, row in enumerate(delete_data[1:], start=2):
                                if len(row) > max(active_col_idx, store_name_col_idx):
                                    is_active = str(row[active_col_idx]).upper() == "TRUE"
                                    store_name = row[store_name_col_idx]
                                    
                                    # active=TRUEì´ê³ , ì„ íƒëœ ìŠ¤í† ì–´ì¸ ê²½ìš°ë§Œ
                                    if is_active and (not selected_stores or store_name in selected_stores):
                                        sales = store_sales.get(store_name, 0)
                                        excess = max(0, sales - delete_limit)
                                        ws_delete.update_acell(f"C{row_idx}", str(excess))
                                        processed_stores.append(store_name)
                                        print(f"[ì˜¬ì¸ì›] {store_name}: íŒë§¤ì¤‘={sales}, ê¸°ì¤€={delete_limit}, ì‚­ì œ={excess}")
                            
                            if processed_stores:
                                print(f"[ì˜¬ì¸ì›] ì´ˆê³¼ë¶„ ì‚­ì œ ëŒ€ìƒ: {', '.join(processed_stores)}")
                        except Exception as e:
                            print(f"[ì˜¬ì¸ì›] ì´ˆê³¼ë¶„ ê³„ì‚° ì˜¤ë¥˜: {e}")
                    
                    elif req.options.get('delete_count'):
                        delete_count = req.options.get('delete_count', 50)
                        processed_stores = []
                        
                        for row_idx, row in enumerate(delete_data[1:], start=2):
                            if len(row) > max(active_col_idx, store_name_col_idx):
                                is_active = str(row[active_col_idx]).upper() == "TRUE"
                                store_name = row[store_name_col_idx]
                                
                                # active=TRUEì´ê³ , ì„ íƒëœ ìŠ¤í† ì–´ì¸ ê²½ìš°ë§Œ
                                if is_active and (not selected_stores or store_name in selected_stores):
                                    ws_delete.update_acell(f"C{row_idx}", str(delete_count))
                                    processed_stores.append(store_name)
                        
                        if processed_stores:
                            print(f"[ì˜¬ì¸ì›] ìƒí’ˆì‚­ì œ {delete_count}ê°œ: {', '.join(processed_stores)}")
                        else:
                            print(f"[ì˜¬ì¸ì›] ìƒí’ˆì‚­ì œ: ëŒ€ìƒ ìŠ¤í† ì–´ ì—†ìŒ (active=TRUEì¸ ì„ íƒ ìŠ¤í† ì–´ í™•ì¸)")
                    
        except Exception as e:
            print(f"[ì˜¬ì¸ì›] ì‘ì—… ì„¤ì • ì˜¤ë¥˜: {e}")
        
        # ë¡œê·¸ íŒŒì¼ ê²½ë¡œ (í”Œë«í¼ë³„)
        log_file = os.path.join(os.path.dirname(__file__), "logs", "allinone_smartstore.log")
        os.makedirs(os.path.dirname(log_file), exist_ok=True)
        
        # í”Œë«í¼ë³„ ìƒíƒœ ì´ˆê¸°í™”
        # get_aio_statusê°€ ì´ë¯¸ ìƒì„±/ë°˜í™˜í•˜ë¯€ë¡œ ì´ˆê¸°í™”ë§Œ ë‹¤ì‹œ í•´ì¤Œ
        platform_status["running"] = True
        platform_status["progress"] = 0
        platform_status["status"] = "running"
        platform_status["results"] = []
        platform_status["process"] = None
        platform_status["options"] = req.options
        platform_status["current_store"] = ""
        platform_status["current_action"] = "í”„ë¡œì„¸ìŠ¤ ì‹œì‘ ì¤‘..."
        platform_status["total"] = 0
        platform_status["completed"] = 0
        platform_status["logs"] = []
        platform_status["log_file"] = log_file
        platform_status["log_pos"] = 0
        
        # í™˜ê²½ë³€ìˆ˜ ì„¤ì •
        env = os.environ.copy()
        env["SERVICE_ACCOUNT_JSON"] = os.environ.get("SERVICE_ACCOUNT_JSON", "")
        env["SPREADSHEET_KEY"] = os.environ.get("SPREADSHEET_KEY", "")
        env["PARALLEL_STORES"] = "true"
        env["PARALLEL_WORKERS"] = "4"
        env["PYTHONIOENCODING"] = "utf-8"  # UTF-8 ì¶œë ¥ ê°•ì œ
        env["AIO_TASK"] = req.task  # ì‘ì—…ëª… ì „ë‹¬ (ëŒ€ìƒ ìŠ¤í† ì–´ëŠ” ì‹œíŠ¸ì˜ 'í™œì„±í™”' ì»¬ëŸ¼ ì°¸ì¡°)
        
        # subprocessë¡œ ë³„ë„ í”„ë¡œì„¸ìŠ¤ ì‹¤í–‰
        module_path = os.path.join(os.path.dirname(__file__), "modules", "smartstore_allinone.py")
        
        with open(log_file, "w", encoding="utf-8") as f:
            f.write(f"[{datetime.now().strftime('%H:%M:%S')}] ìŠ¤ë§ˆíŠ¸ìŠ¤í† ì–´ {req.task} ì‹œì‘\n")
        
        process = subprocess.Popen(
            [sys.executable, module_path],
            stdout=open(log_file, "a", encoding="utf-8"),
            stderr=subprocess.STDOUT,
            env=env,
            cwd=os.path.dirname(__file__)
        )
        platform_status["process"] = process
        
        return {"success": True, "message": f"{req.platform} {req.task} ì‹¤í–‰ ì‹œì‘ (PID: {process.pid})"}
    
    elif req.platform == "11ë²ˆê°€":
        # 11ë²ˆê°€ ì‘ì—… ì²˜ë¦¬
        
        # ë¡œê·¸ íŒŒì¼ ê²½ë¡œ (í”Œë«í¼ë³„)
        log_file = os.path.join(os.path.dirname(__file__), "logs", "allinone_11st.log")
        os.makedirs(os.path.dirname(log_file), exist_ok=True)
        
        # í”Œë«í¼ë³„ ìƒíƒœ ì´ˆê¸°í™”
        platform_status["running"] = True
        platform_status["progress"] = 0
        platform_status["status"] = "running"
        platform_status["results"] = []
        platform_status["process"] = None
        platform_status["options"] = req.options
        platform_status["current_store"] = ""
        platform_status["current_action"] = "í”„ë¡œì„¸ìŠ¤ ì‹œì‘ ì¤‘..."
        platform_status["total"] = 0
        platform_status["completed"] = 0
        platform_status["logs"] = []
        platform_status["log_file"] = log_file
        platform_status["log_pos"] = 0

        
        if req.task == "ë“±ë¡ê°¯ìˆ˜":
            # ë“±ë¡ê°¯ìˆ˜(íŒë§¤ì¤‘) ì¡°íšŒ - ì¸ë¼ì¸ ì²˜ë¦¬
            with open(log_file, "w", encoding="utf-8") as f:
                f.write(f"[{datetime.now().strftime('%H:%M:%S')}] 11ë²ˆê°€ íŒë§¤ì¤‘ ì¡°íšŒ ì‹œì‘\n")
            
            # ë°±ê·¸ë¼ìš´ë“œì—ì„œ ì‹¤í–‰
            import asyncio
            asyncio.create_task(run_11st_product_count_task(log_file, "11ë²ˆê°€"))
            
            return {"success": True, "message": f"11ë²ˆê°€ íŒë§¤ì¤‘ ì¡°íšŒ ì‹œì‘"}
        else:
            # íŒë§¤ì¤‘ì§€/íŒë§¤ì¬ê°œ ë“± - subprocessë¡œ ì‹¤í–‰
            env = os.environ.copy()
            env["GOOGLE_SERVICE_ACCOUNT_FILE"] = os.environ.get("GOOGLE_SERVICE_ACCOUNT_FILE", os.environ.get("SERVICE_ACCOUNT_JSON", ""))
            env["SPREADSHEET_KEY"] = os.environ.get("SPREADSHEET_KEY", "")
            env["ELEVENST_TASK"] = req.task  # ì‘ì—… ì¢…ë¥˜ ì „ë‹¬
            env["PYTHONIOENCODING"] = "utf-8"  # UTF-8 ì¶œë ¥ ê°•ì œ
            
            # subprocessë¡œ ë³„ë„ í”„ë¡œì„¸ìŠ¤ ì‹¤í–‰
            module_path = os.path.join(os.path.dirname(__file__), "modules", "elevenst.py")
            
            with open(log_file, "w", encoding="utf-8") as f:
                f.write(f"[{datetime.now().strftime('%H:%M:%S')}] 11ë²ˆê°€ {req.task} ì‹œì‘\n")
            
            process = subprocess.Popen(
                [sys.executable, module_path],
                stdout=open(log_file, "a", encoding="utf-8"),
                stderr=subprocess.STDOUT,
                env=env,
                cwd=os.path.dirname(__file__)
            )
            platform_status["process"] = process
            
            return {"success": True, "message": f"{req.platform} {req.task} ì‹¤í–‰ ì‹œì‘ (PID: {process.pid})"}
    
    else:
        return {"success": False, "message": f"ì§€ì›í•˜ì§€ ì•ŠëŠ” í”Œë«í¼: {req.platform}"}

@app.get("/api/allinone/progress")
async def get_allinone_progress(request: Request, platform: str = None, task: str = None):
    """ì˜¬ì¸ì› ì§„í–‰ìƒí™© ì¡°íšŒ (í”Œë«í¼ë³„)"""
    require_permission(request, "edit")
    
    # í”Œë«í¼ë³„ ìƒíƒœ ê°€ì ¸ì˜¤ê¸°
    status = get_aio_status(platform, task)
    
    # í”„ë¡œì„¸ìŠ¤ ìƒíƒœ í™•ì¸
    if status.get("process"):
        poll = status["process"].poll()
        if poll is not None:
            # í”„ë¡œì„¸ìŠ¤ ì¢…ë£Œë¨
            status["running"] = False
            status["status"] = "completed"
            status["progress"] = 100
            status["current_store"] = ""
            status["current_action"] = "ì™„ë£Œ"
    
    # ë¡œê·¸ íŒŒì¼ì—ì„œ ìƒˆ ë¡œê·¸ ì½ê¸°
    log_file = status.get("log_file")
    if log_file and os.path.exists(log_file):
        try:
            # Windows CP949 ë˜ëŠ” UTF-8 ì¸ì½”ë”© ì‹œë„
            for enc in ["utf-8", "cp949", "euc-kr"]:
                try:
                    with open(log_file, "r", encoding=enc, errors="replace") as f:
                        f.seek(status.get("log_pos", 0))
                        new_lines = f.readlines()
                        new_pos = f.tell()
                    break
                except UnicodeDecodeError:
                    continue
            else:
                new_lines = []
                new_pos = status.get("log_pos", 0)
            
            if new_lines:
                status["log_pos"] = new_pos
                for line in new_lines:
                    line = line.strip()
                    if line:
                        # ë¡œê·¸ì—ì„œ ì‹œê°„ ì¶”ì¶œ ì‹œë„ (ì˜ˆ: [23:36:04] ë˜ëŠ” [11ë²ˆê°€] í˜•íƒœ)
                        import re
                        time_match = re.match(r'\[(\d{2}:\d{2}:\d{2})\]', line)
                        if time_match:
                            log_time = time_match.group(1)
                        else:
                            log_time = datetime.now().strftime("%H:%M:%S")
                        
                        status["logs"].append({
                            "time": log_time,
                            "msg": line
                        })
                # ìµœê·¼ 100ê°œë§Œ ìœ ì§€
                if len(status["logs"]) > 100:
                    status["logs"] = status["logs"][-100:]
                
                # ë¡œê·¸ì—ì„œ ì§„í–‰ ìƒí™© íŒŒì‹±
                for line in new_lines:
                    line = line.strip()
                    import re
                    
                    # [RUN] ëŒ€ìƒ ê³„ì •: Nê°œ í˜•íƒœì—ì„œ total ì¶”ì¶œ
                    total_match = re.search(r'ëŒ€ìƒ ê³„ì •:\s*(\d+)ê°œ', line)
                    if total_match:
                        status["total"] = int(total_match.group(1))
                    
                    # [1/5] ìŠ¤í† ì–´ëª…: í˜•íƒœì—ì„œ í˜„ì¬ ì§„í–‰ìƒí™© ì¶”ì¶œ
                    progress_match = re.search(r'\[(\d+)/(\d+)\]\s*([^:]+):', line)
                    if progress_match:
                        completed = int(progress_match.group(1))
                        total = int(progress_match.group(2))
                        store_name = progress_match.group(3).strip()
                        status["completed"] = completed
                        status["total"] = total
                        status["current_store"] = store_name
                        status["current_action"] = line
                        if total > 0:
                            status["progress"] = int((completed / total) * 100)
        except Exception as e:
            print(f"ë¡œê·¸ ì½ê¸° ì˜¤ë¥˜: {e}")
    
    return {
        "running": status.get("running", False),
        "progress": status.get("progress", 0),
        "status": status.get("status", "idle"),
        "results": status.get("results", []),
        "current_store": status.get("current_store", ""),
        "current_action": status.get("current_action", ""),
        "total": status.get("total", 0),
        "completed": status.get("completed", 0),
        "logs": status.get("logs", [])
    }

@app.post("/api/allinone/stop")
async def stop_allinone_task(request: Request, platform: str = None, task: str = None):
    """ì˜¬ì¸ì› ì‘ì—… ì¤‘ì§€ (í”Œë«í¼ë³„)"""
    require_permission(request, "edit")
    
    # í”Œë«í¼ë³„ ìƒíƒœ ê°€ì ¸ì˜¤ê¸°
    status = get_aio_status(platform, task)
    
    if status.get("process"):
        try:
            status["process"].terminate()
        except:
            pass
    
    # 11ë²ˆê°€ ëª¨ë“ˆ ì¤‘ì§€
    if platform == "11ë²ˆê°€" or platform is None:
        try:
            from modules import elevenst
            elevenst.stop_all()
        except:
            pass
    
    status["running"] = False
    status["status"] = "stopped"
    
    return {"success": True, "message": f"{platform or 'ì „ì²´'} ì¤‘ì§€ ìš”ì²­ë¨"}


# ========== KC ì¸ì¦ ìˆ˜ì • API ==========
class KCModifyRequest(BaseModel):
    stores: List[str]  # store_name ëª©ë¡
    product_limit: int = 2000  # ë§ˆì¼“ë‹¹ ì²˜ë¦¬í•  ìƒí’ˆ ìˆ˜
    mode: str = "count"  # count ë˜ëŠ” date
    target_date: str = ""  # mode=dateì¼ ë•Œ ê¸°ì¤€ ë‚ ì§œ (YYYY-MM-DD)

# KC ìˆ˜ì • ìƒíƒœ ì €ì¥
kc_modify_status = {
    "running": False,
    "progress": {},  # store_name -> {progress, total, success, fail, status}
    "logs": [],
    "stop_requested": False
}

def get_naver_token(client_id: str, client_secret: str) -> str:
    """ë„¤ì´ë²„ ì»¤ë¨¸ìŠ¤ API í† í° ë°œê¸‰"""
    import bcrypt
    timestamp = int(time.time() * 1000)
    password = f"{client_id}_{timestamp}"
    hashed = bcrypt.hashpw(password.encode('utf-8'), client_secret.encode('utf-8'))
    signature = base64.b64encode(hashed).decode('utf-8')
    
    url = "https://api.commerce.naver.com/external/v1/oauth2/token"
    data = {
        "client_id": client_id,
        "timestamp": timestamp,
        "client_secret_sign": signature,
        "grant_type": "client_credentials",
        "type": "SELF"
    }
    
    response = requests.post(url, data=data)
    if response.status_code == 200:
        return response.json().get("access_token")
    raise Exception(f"í† í° ë°œê¸‰ ì‹¤íŒ¨: {response.text}")

def modify_kc_for_store(store_name: str, client_id: str, client_secret: str, product_limit: int, mode: str = "count", target_date: str = ""):
    """ë‹¨ì¼ ìŠ¤í† ì–´ KC ì¸ì¦ ìˆ˜ì •
    mode: count - ìµœì‹  Nê°œ ìƒí’ˆ
    mode: date - ì§€ì • ë‚ ì§œ ì´í›„ ë“±ë¡ ìƒí’ˆ
    """
    global kc_modify_status
    
    def add_log(msg, status="info"):
        log_entry = {
            "time": datetime.now().strftime("%H:%M:%S"),
            "store": store_name,
            "msg": msg,
            "status": status
        }
        kc_modify_status["logs"].append(log_entry)
        if len(kc_modify_status["logs"]) > 500:
            kc_modify_status["logs"] = kc_modify_status["logs"][-500:]
        print(f"[KC-{store_name}] {msg}")
    
    try:
        kc_modify_status["progress"][store_name] = {
            "progress": 0, "total": 0, "success": 0, "fail": 0, "status": "í† í° ë°œê¸‰ ì¤‘..."
        }
        
        # í† í° ë°œê¸‰
        token = get_naver_token(client_id, client_secret)
        headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json"
        }
        
        add_log("í† í° ë°œê¸‰ ì™„ë£Œ")
        kc_modify_status["progress"][store_name]["status"] = "ìƒí’ˆ ì¡°íšŒ ì¤‘..."
        
        # ìƒí’ˆ ëª©ë¡ ì¡°íšŒ (ìµœì‹  ë“±ë¡ìˆœ)
        products = []
        page = 1
        base_url = "https://api.commerce.naver.com/external"
        
        # ë‚ ì§œ ëª¨ë“œì¼ ë•Œ ê¸°ì¤€ ë‚ ì§œ íŒŒì‹±
        filter_date = None
        if mode == "date" and target_date:
            try:
                filter_date = datetime.strptime(target_date, "%Y-%m-%d")
                add_log(f"ë‚ ì§œ ê¸°ì¤€: {target_date} ì´í›„ ë“±ë¡ ìƒí’ˆ")
            except:
                add_log(f"ë‚ ì§œ íŒŒì‹± ì‹¤íŒ¨: {target_date}", "error")
        
        max_pages = 100 if mode == "date" else 20  # ë‚ ì§œ ëª¨ë“œëŠ” ë” ë§ì´ ì¡°íšŒ
        
        while (mode == "count" and len(products) < product_limit) or (mode == "date" and page <= max_pages):
            if kc_modify_status["stop_requested"]:
                add_log("ì¤‘ì§€ ìš”ì²­ë¨", "warning")
                break
            
            body = {"page": page, "size": 500, "sortType": "RECENTLY_REGISTERED"}
            resp = requests.post(f"{base_url}/v1/products/search", headers=headers, json=body)
            
            if resp.status_code != 200:
                add_log(f"ìƒí’ˆ ì¡°íšŒ ì‹¤íŒ¨: {resp.text}", "error")
                break
            
            data = resp.json()
            contents = data.get("contents", [])
            
            if not contents:
                break
            
            stop_fetching = False
            for item in contents:
                if mode == "count" and len(products) >= product_limit:
                    stop_fetching = True
                    break
                
                origin_no = item.get("originProductNo")
                reg_date_str = item.get("registrationDate", "")
                channel_products = item.get("channelProducts", [])
                
                # ë‚ ì§œ ëª¨ë“œì¼ ë•Œ í•„í„°ë§
                if mode == "date" and filter_date and reg_date_str:
                    try:
                        # ë“±ë¡ì¼ íŒŒì‹± (ISO í˜•ì‹)
                        reg_date = datetime.fromisoformat(reg_date_str.replace('Z', '+00:00').split('+')[0])
                        if reg_date < filter_date:
                            stop_fetching = True  # ë” ì´ì „ ìƒí’ˆì€ ì¡°íšŒ ì¤‘ì§€
                            break
                    except:
                        pass
                
                if channel_products:
                    products.append({
                        "originProductNo": origin_no,
                        "name": channel_products[0].get("name", "")[:30],
                        "registrationDate": reg_date_str
                    })
            
            if stop_fetching:
                break
            
            if page >= data.get("totalPages", 1):
                break
            page += 1
            time.sleep(0.3)
        
        total = len(products)
        add_log(f"ìƒí’ˆ {total}ê°œ ì¡°íšŒ ì™„ë£Œ")
        kc_modify_status["progress"][store_name]["total"] = total
        kc_modify_status["progress"][store_name]["status"] = "KC ìˆ˜ì • ì¤‘..."
        
        if total == 0:
            kc_modify_status["progress"][store_name]["status"] = "ì™„ë£Œ (ìƒí’ˆ ì—†ìŒ)"
            return {"success": 0, "fail": 0}
        
        # ìƒí’ˆë³„ KC ì¸ì¦ ìˆ˜ì •
        success = 0
        fail = 0
        
        for idx, product in enumerate(products):
            if kc_modify_status["stop_requested"]:
                add_log("ì¤‘ì§€ë¨", "warning")
                break
            
            product_no = product["originProductNo"]
            
            try:
                # ìƒì„¸ ì¡°íšŒ
                detail_resp = requests.get(
                    f"{base_url}/v2/products/origin-products/{product_no}",
                    headers=headers
                )
                
                if detail_resp.status_code != 200:
                    fail += 1
                    continue
                
                detail = detail_resp.json()
                origin_product = detail.get("originProduct", {})
                
                # KC ì¸ì¦ ì œì™¸ ì„¤ì •
                if "detailAttribute" not in origin_product:
                    origin_product["detailAttribute"] = {}
                
                origin_product["detailAttribute"]["certificationTargetExcludeContent"] = {
                    "kcCertifiedProductExclusionYn": "TRUE",
                    "childCertifiedProductExclusionYn": True,
                    "greenCertifiedProductExclusionYn": True
                }
                
                # ì—…ë°ì´íŠ¸
                update_data = {"originProduct": origin_product}
                if "smartstoreChannelProduct" in detail:
                    update_data["smartstoreChannelProduct"] = detail["smartstoreChannelProduct"]
                
                update_resp = requests.put(
                    f"{base_url}/v2/products/origin-products/{product_no}",
                    headers=headers,
                    json=update_data
                )
                
                if update_resp.status_code == 200:
                    success += 1
                else:
                    fail += 1
                
                # 10ê°œë§ˆë‹¤ ë¡œê·¸
                if success % 10 == 0 and success > 0:
                    add_log(f"{success}ê°œ ì™„ë£Œ...")
                
            except Exception as e:
                fail += 1
                if fail <= 3:  # ì²˜ìŒ 3ê°œë§Œ ë¡œê·¸
                    add_log(f"[{product_no}] ì˜¤ë¥˜: {str(e)[:50]}", "error")
            
            kc_modify_status["progress"][store_name]["progress"] = idx + 1
            kc_modify_status["progress"][store_name]["success"] = success
            kc_modify_status["progress"][store_name]["fail"] = fail
            
            time.sleep(1)  # API ì œí•œ
        
        status_text = f"ì™„ë£Œ (ì„±ê³µ:{success}, ì‹¤íŒ¨:{fail})"
        kc_modify_status["progress"][store_name]["status"] = status_text
        add_log(status_text, "success")
        
        return {"success": success, "fail": fail}
        
    except Exception as e:
        add_log(f"ì˜¤ë¥˜: {str(e)}", "error")
        kc_modify_status["progress"][store_name]["status"] = f"ì˜¤ë¥˜: {str(e)[:30]}"
        return {"success": 0, "fail": 0, "error": str(e)}

@app.post("/api/allinone/kc-modify")
async def run_kc_modify(request: Request, req: KCModifyRequest):
    """KC ì¸ì¦ ì¼ê´„ ìˆ˜ì • ì‹¤í–‰"""
    require_permission(request, "edit")

    global kc_modify_status

    if kc_modify_status["running"]:
        return {"success": False, "message": "ì´ë¯¸ ì‹¤í–‰ ì¤‘ì…ë‹ˆë‹¤"}

    if not req.stores:
        return {"success": False, "message": "ìŠ¤í† ì–´ë¥¼ ì„ íƒí•˜ì„¸ìš”"}

    # ì‘ì—… ë¡œê·¸ ê¸°ë¡
    store_names = ", ".join(req.stores[:5]) + ("..." if len(req.stores) > 5 else "")
    log_work("KCì¸ì¦ìˆ˜ì •", "ìŠ¤ë§ˆíŠ¸ìŠ¤í† ì–´", len(req.stores), f"ëŒ€ìƒ: {store_names}", "ì›¹")
    
    # ê³„ì •ëª©ë¡ ì‹œíŠ¸ì—ì„œ API ì •ë³´ ê°€ì ¸ì˜¤ê¸°
    try:
        ws = gsheet.sheet.worksheet(ACCOUNTS_TAB)  # ê³„ì •ëª©ë¡
        data = ws.get_all_values()
        headers = data[0]

        # ì»¬ëŸ¼ ì¸ë±ìŠ¤ ì°¾ê¸° (ê³„ì •ëª©ë¡ ì»¬ëŸ¼ëª…)
        name_idx = None
        id_idx = None
        secret_idx = None
        platform_idx = None
        for i, h in enumerate(headers):
            if h in ["ìŠ¤í† ì–´ëª…", "store_name"]:
                name_idx = i
            elif h in ["ìŠ¤ë§ˆíŠ¸ìŠ¤í† ì–´ ì• í”Œë¦¬ì¼€ì´ì…˜ ID", "client_id"]:
                id_idx = i
            elif h in ["ìŠ¤ë§ˆíŠ¸ìŠ¤í† ì–´ ì• í”Œë¦¬ì¼€ì´ì…˜ ì‹œí¬ë¦¿", "client_secret"]:
                secret_idx = i
            elif h in ["í”Œë«í¼", "platform"]:
                platform_idx = i

        if None in [name_idx, id_idx, secret_idx]:
            return {"success": False, "message": "ì‹œíŠ¸ì— í•„ìš”í•œ ì»¬ëŸ¼ì´ ì—†ìŠµë‹ˆë‹¤ (ìŠ¤í† ì–´ëª…, ìŠ¤ë§ˆíŠ¸ìŠ¤í† ì–´ ì• í”Œë¦¬ì¼€ì´ì…˜ ID/ì‹œí¬ë¦¿)"}

        # ì„ íƒëœ ìŠ¤í† ì–´ ì •ë³´ ì¶”ì¶œ (ìŠ¤ë§ˆíŠ¸ìŠ¤í† ì–´ë§Œ)
        store_info = {}
        for row in data[1:]:
            if len(row) > max(name_idx, id_idx, secret_idx):
                # í”Œë«í¼ ì²´í¬ (ìŠ¤ë§ˆíŠ¸ìŠ¤í† ì–´ë§Œ)
                platform = row[platform_idx].lower() if platform_idx and len(row) > platform_idx else ""
                if platform and platform not in ["ìŠ¤ë§ˆíŠ¸ìŠ¤í† ì–´", "smartstore", "ë„¤ì´ë²„", "naver"]:
                    continue
                name = row[name_idx]
                if name in req.stores:
                    store_info[name] = {
                        "client_id": row[id_idx],
                        "client_secret": row[secret_idx]
                    }

        if not store_info:
            return {"success": False, "message": "ì„ íƒëœ ìŠ¤í† ì–´ ì •ë³´ë¥¼ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤"}

    except Exception as e:
        return {"success": False, "message": f"ì‹œíŠ¸ ì¡°íšŒ ì˜¤ë¥˜: {str(e)}"}
    
    # ìƒíƒœ ì´ˆê¸°í™”
    kc_modify_status = {
        "running": True,
        "progress": {},
        "logs": [],
        "stop_requested": False
    }
    
    def run_parallel():
        global kc_modify_status
        
        try:
            from concurrent.futures import ThreadPoolExecutor, as_completed
            
            with ThreadPoolExecutor(max_workers=min(5, len(store_info))) as executor:
                futures = {}
                for store_name, info in store_info.items():
                    future = executor.submit(
                        modify_kc_for_store,
                        store_name,
                        info["client_id"],
                        info["client_secret"],
                        req.product_limit,
                        req.mode,
                        req.target_date
                    )
                    futures[future] = store_name
                
                for future in as_completed(futures):
                    store_name = futures[future]
                    try:
                        result = future.result()
                    except Exception as e:
                        print(f"[KC] {store_name} ìŠ¤ë ˆë“œ ì˜¤ë¥˜: {e}")
        
        finally:
            kc_modify_status["running"] = False
    
    # ë°±ê·¸ë¼ìš´ë“œ ì‹¤í–‰
    import threading
    threading.Thread(target=run_parallel, daemon=True).start()
    
    return {"success": True, "message": f"{len(store_info)}ê°œ ìŠ¤í† ì–´ KC ìˆ˜ì • ì‹œì‘"}

@app.get("/api/allinone/kc-progress")
async def get_kc_progress(request: Request):
    """KC ìˆ˜ì • ì§„í–‰ìƒí™© ì¡°íšŒ"""
    require_permission(request, "edit")
    return kc_modify_status

@app.post("/api/allinone/kc-stop")
async def stop_kc_modify(request: Request):
    """KC ìˆ˜ì • ì¤‘ì§€"""
    require_permission(request, "edit")
    global kc_modify_status
    kc_modify_status["stop_requested"] = True
    return {"success": True, "message": "ì¤‘ì§€ ìš”ì²­ë¨"}


# ========== êµ¬ê¸€ì‹œíŠ¸ ë§¤ì¶œ ì§‘ê³„ API ==========
# ë§¤ì¶œ ì‹œíŠ¸ ID
SALES_SHEET_ID = "1MHhu1GdvV1OGS8Wy3NxWOKuqFvgZpqgwn08kG70EDsY"

# ë§¤ì¶œ ìºì‹œ (ë©”ëª¨ë¦¬)
sales_cache = {
    "data": {},  # {ë§ˆì¼“ID: {today_sales, today_orders, month_sales, month_orders}}
    "updated_at": None,
    "daily": [],  # ì¼ìë³„ ì§‘ê³„
    "by_owner": {},  # ì†Œìœ ìë³„ ì§‘ê³„
    "raw_data": []  # ì›ë³¸ ë°ì´í„°
}

@app.get("/api/sales/from-sheet")
async def get_sales_from_sheet_v2(request: Request, force: bool = False):
    """êµ¬ê¸€ì‹œíŠ¸ì—ì„œ ë§ˆì¼“ë³„ ë§¤ì¶œ ì§‘ê³„"""
    import os
    log_path = os.path.abspath("debug_sales_log.txt")
    with open(log_path, "a", encoding="utf-8") as f:
        f.write(f"\n[DEBUG] get_sales_from_sheet CALLED. force={force}\n")
        f.flush()
    
    with open("debug_filtering_stats.txt", "w", encoding="utf-8") as f:
        f.write("[DEBUG] get_sales_from_sheet START\n")

    from datetime import datetime
    
    # ìºì‹œ í™•ì¸ (5ë¶„ ì´ë‚´ë©´ ìºì‹œ ì‚¬ìš©)
    if not force and sales_cache["updated_at"]:
        cache_age = (datetime.now() - sales_cache["updated_at"]).total_seconds()
        with open(log_path, "a", encoding="utf-8") as f:
            f.write(f"[DEBUG] Cache hit check. Age: {cache_age}\n")
        if cache_age < 300:  # 5ë¶„
            with open(log_path, "a", encoding="utf-8") as f:
                f.write(f"[DEBUG] RETURNING CACHED DATA\n")
            return {"success": True, "data": sales_cache["data"], "daily": sales_cache.get("daily", []), "total": sales_cache.get("total", {}), "cached": True}
    
    with open(log_path, "a", encoding="utf-8") as f:
        f.write(f"[DEBUG] Proceeding to fetch from sheet...\n")
    
    try:
        # í˜„ì¬ ì›” íƒ­ ì´ë¦„ (ì˜ˆ: "12ì›”")
        current_month = datetime.now().month
        current_tab = f"{current_month}ì›”"
        prev_month = current_month - 1 if current_month > 1 else 12
        prev_tab = f"{prev_month}ì›”"
        today = datetime.now().date()
        days_30_ago = today - timedelta(days=30)
        
        # ê³„ì •ëª©ë¡ ì‹œíŠ¸ì—ì„œ ê³„ì • ëª©ë¡ ë¡œë“œ
        all_store_keys = set()  # "store_name(í”Œë«í¼)" í˜•íƒœ
        account_usage = {}  # {"store_name(í”Œë«í¼)": "ëŒ€ëŸ‰" or "ë°˜ëŒ€ëŸ‰"}
        account_owner = {}  # {"store_name(í”Œë«í¼)": "ì†Œìœ ì(JSM, JJI ë“±)"}
        # ì‚¬ì—…ìë²ˆí˜¸ë¡œ ë§¤ì¹­í•˜ê¸° ìœ„í•œ ì¶”ê°€ ë”•ì…”ë„ˆë¦¬
        biz_to_owner = {}  # {"ì‚¬ì—…ìë²ˆí˜¸": "ì†Œìœ ì"}
        biz_to_usage = {}  # {"ì‚¬ì—…ìë²ˆí˜¸": "ìš©ë„"}
        biz_to_stores = {}  # {"ì‚¬ì—…ìë²ˆí˜¸": set(ìŠ¤í† ì–´ëª…ë“¤)} - ë§¤ì¶œ ìœ ë¬´ì™€ ê´€ê³„ì—†ì´ ëª¨ë“  ìŠ¤í† ì–´
        try:
            accounts = gsheet.get_accounts()
            for acc in accounts:
                # store_name = ì‡¼í•‘ëª° ë³„ì¹­ = ë§¤ì¶œ ì‹œíŠ¸ì˜ "ì‚¬ì—…ì"(Fì—´)
                store_name = acc.get("ìŠ¤í† ì–´ëª…", "") or acc.get("ìŠ¤í† ì–´ëª…", "")
                store_name = store_name.strip()
                platform = acc.get("platform", "").strip()
                usage = acc.get("usage", "").strip()
                owner = acc.get("owner", "").strip()  # ì‹¤ì œ ì†Œìœ ì (JSM, JJI ë“±)
                biz_number = acc.get("business_number", "").strip()  # ì‚¬ì—…ìë²ˆí˜¸
                
                # store_name(í”Œë«í¼) í‚¤ë¡œ ë§¤ì¹­
                if store_name and platform:
                    key = f"{store_name}({platform})"
                    all_store_keys.add(key)
                    account_usage[key] = usage
                    account_owner[key] = owner
                
                # ì‚¬ì—…ìë²ˆí˜¸ë¡œë„ ë§¤ì¹­ (ì‚¬ì—…ìë²ˆí˜¸ê°€ ìˆìœ¼ë©´)
                if biz_number and owner:
                    biz_to_owner[biz_number] = owner
                if biz_number and usage:
                    biz_to_usage[biz_number] = usage
                
                # ì‚¬ì—…ìë²ˆí˜¸ë³„ ëª¨ë“  ìŠ¤í† ì–´ ë§¤í•‘ (ë§¤ì¶œ ìœ ë¬´ì™€ ê´€ê³„ì—†ì´)
                if biz_number and store_name:
                    if biz_number not in biz_to_stores:
                        biz_to_stores[biz_number] = set()
                    biz_to_stores[biz_number].add(store_name)
                    
            print(f"[ë§¤ì¶œì§‘ê³„] ê³„ì •ëª©ë¡ì—ì„œ {len(all_store_keys)}ê°œ ê³„ì • ë¡œë“œ")
            print(f"[ë§¤ì¶œì§‘ê³„] ì‚¬ì—…ìë²ˆí˜¸ ë§¤í•‘: {len(biz_to_owner)}ê°œ owner, {len(biz_to_usage)}ê°œ usage")
            if all_store_keys:
                sample_keys = list(all_store_keys)[:5]
                print(f"[ë§¤ì¶œì§‘ê³„] ê³„ì •ëª©ë¡ í‚¤ ìƒ˜í”Œ: {sample_keys}")
                # owner, usage ìƒ˜í”Œ ì¶œë ¥
                for k in sample_keys:
                    print(f"[ë§¤ì¶œì§‘ê³„]   {k} â†’ owner='{account_owner.get(k, '')}', usage='{account_usage.get(k, '')}'")
                # ì¬ì´ë§ˆì¼“ ì°¾ê¸°
                for k in all_store_keys:
                    if "ì¬ì´ë§ˆì¼“" in k:
                        print(f"[ë§¤ì¶œì§‘ê³„] â˜… ê³„ì •ëª©ë¡ ì¬ì´ë§ˆì¼“: key='{k}', owner='{account_owner.get(k, '')}'")
        except Exception as e:
            print(f"[ë§¤ì¶œì§‘ê³„] ê³„ì • ëª©ë¡ ë¡œë“œ ì‹¤íŒ¨: {e}")
        
        # ì‹œíŠ¸ ì—´ê¸°
        sales_sheet = gsheet.client.open_by_key(SALES_SHEET_ID)
        
        # ì´ë²ˆë‹¬ + ì§€ë‚œë‹¬ ë°ì´í„° í•©ì¹˜ê¸°
        all_data = []
        headers = None
        
        for tab_name in [current_tab, prev_tab]:
            try:
                ws = sales_sheet.worksheet(tab_name)
                tab_data = ws.get_all_values()
                if len(tab_data) >= 3:
                    if headers is None:
                        headers = tab_data[1]  # 2í–‰ì´ í—¤ë”
                        all_data = tab_data[2:]  # 3í–‰ë¶€í„° ë°ì´í„°
                    else:
                        all_data.extend(tab_data[2:])
                    print(f"[ë§¤ì¶œì§‘ê³„] {tab_name} íƒ­ì—ì„œ {len(tab_data)-2}ê±´ ë¡œë“œ")
            except Exception as e:
                print(f"[ë§¤ì¶œì§‘ê³„] {tab_name} íƒ­ ë¡œë“œ ì‹¤íŒ¨: {e}")
        
        if not headers or len(all_data) < 1:
            with open("debug_filtering_stats.txt", "a", encoding="utf-8") as f:
                f.write(f"[DEBUG] Early Return: No Data (headers={bool(headers)}, len={len(all_data)})\n")
            return {"success": False, "message": "ë°ì´í„° ì—†ìŒ"}
        
        # ì»¬ëŸ¼ ì¸ë±ìŠ¤ ì°¾ê¸° (ì´ë¦„ìœ¼ë¡œ) - ì¤„ë°”ê¿ˆ ì œê±°
        def find_col(names):
            for name in names:
                for idx, h in enumerate(headers):
                    # í—¤ë”ì—ì„œ ì¤„ë°”ê¿ˆ, ê³µë°± ì œê±° í›„ ë¹„êµ
                    h_clean = h.replace('\n', '').replace('\r', '').replace(' ', '')
                    name_clean = name.replace(' ', '')
                    if name_clean in h_clean:
                        return idx
            return -1
        
        # ì—´ ë¬¸ìë¥¼ ì¸ë±ìŠ¤ë¡œ ë³€í™˜ (A=0, B=1, ..., Z=25, AA=26, ...)
        def col_letter_to_idx(letter):
            letter = letter.upper()
            result = 0
            for char in letter:
                result = result * 26 + (ord(char) - ord('A') + 1)
            return result - 1
        
        # ì´ë¦„ìœ¼ë¡œ ë¨¼ì € ì°¾ê³ , ëª» ì°¾ìœ¼ë©´ ì§ì ‘ ì—´ ì¸ë±ìŠ¤ ì‚¬ìš©
        col_market = find_col(["ë§ˆì¼“"])  # Eì—´: ìŠ¤ë§ˆíŠ¸ìŠ¤í† ì–´, 11ë²ˆê°€, ì¿ íŒ¡ ë“±
        col_owner = find_col(STORE_NAME_ALIASES) # ìŠ¤í† ì–´ëª…(ì‚¬ì—…ì) ì‹ë³„ ì»¬ëŸ¼ í†µí•©
        col_order_date = find_col(["ì£¼ë¬¸ì¼ì"])
        col_payment = find_col(["ì‹¤ê²°ì œê¸ˆì•¡(ë°°ì†¡ë¹„í¬í•¨)", "ì‹¤ê²°ì œê¸ˆì•¡"])
        col_settlement = find_col(["ì •ì‚°ê¸ˆì•¡(ë°°ì†¡ë¹„í¬í•¨)", "ì •ì‚°ê¸ˆì•¡"])
        col_profit = find_col(["ìˆ˜ìµê¸ˆ"])
        col_profit_rate = find_col(["ìˆ˜ìµë¥ "])
        col_order_status = find_col(["ì£¼ë¬¸í˜„í™©"])
        col_purchase = find_col(["êµ¬ë§¤ê¸ˆì•¡(ì›í™”)", "êµ¬ë§¤ê¸ˆì•¡"])
        col_int_shipping = find_col(["êµ­ì œë°°ì†¡ë¹„"])
        col_cargo_shipping = find_col(["í™”ë¬¼íƒë°°ë¹„"])
        col_biz_number = find_col(["ì‚¬ì—…ìë²ˆí˜¸", "ì‚¬ì—…ì ë²ˆí˜¸"])  # ANì—´: ì‚¬ì—…ìë²ˆí˜¸
        
        # ëª» ì°¾ì€ ì»¬ëŸ¼ì€ ì§ì ‘ ì—´ ì¸ë±ìŠ¤ ì§€ì • (ì´ë¯¸ì§€ ê¸°ì¤€)
        if col_order_status < 0: col_order_status = col_letter_to_idx('D')   # Dì—´: ì£¼ë¬¸í˜„í™©
        if col_market < 0: col_market = col_letter_to_idx('E')               # Eì—´: ë§ˆì¼“
        if col_owner < 0: col_owner = col_letter_to_idx('F')                 # Fì—´: ì‚¬ì—…ì
        if col_order_date < 0: col_order_date = col_letter_to_idx('G')       # Gì—´: ì£¼ë¬¸ì¼ì
        if col_payment < 0: col_payment = col_letter_to_idx('X')             # Xì—´: ì‹¤ê²°ì œê¸ˆì•¡
        if col_settlement < 0: col_settlement = col_letter_to_idx('AA')      # AAì—´: ì •ì‚°ê¸ˆì•¡
        if col_purchase < 0: col_purchase = col_letter_to_idx('AL')          # ALì—´: êµ¬ë§¤ê¸ˆì•¡(ì›í™”)
        if col_biz_number < 0: col_biz_number = col_letter_to_idx('AN')      # ANì—´: ì‚¬ì—…ìë²ˆí˜¸
        if col_int_shipping < 0: col_int_shipping = col_letter_to_idx('AR')  # ARì—´: êµ­ì œë°°ì†¡ë¹„
        if col_profit < 0: col_profit = col_letter_to_idx('AY')              # AYì—´: ìˆ˜ìµê¸ˆ
        if col_profit_rate < 0: col_profit_rate = col_letter_to_idx('AZ')    # AZì—´: ìˆ˜ìµë¥ 
        if col_cargo_shipping < 0: col_cargo_shipping = col_letter_to_idx('AU')  # AUì—´: í™”ë¬¼íƒë°°ë¹„
        
        debug_msg = f"[ë§¤ì¶œì§‘ê³„] ì»¬ëŸ¼ - ë§ˆì¼“:{col_market}, ì£¼ë¬¸ì¼ì:{col_order_date}, ì‹¤ê²°ì œ:{col_payment}, ì •ì‚°:{col_settlement}, ìˆ˜ìµê¸ˆ:{col_profit}, ì‚¬ì—…ìë²ˆí˜¸:{col_biz_number}"
        print(debug_msg)
        with open("debug_sales_log.txt", "a", encoding="utf-8") as f:
            f.write(debug_msg + "\n")
        
        if col_market < 0 or col_order_date < 0:
            return {"success": False, "message": "í•„ìˆ˜ ì»¬ëŸ¼ ì—†ìŒ (ë§ˆì¼“, ì£¼ë¬¸ì¼ì)"}
        
        # ë§ˆì¼“ë³„ ì§‘ê³„
        market_sales = {}
        daily_sales = {}  # ì¼ìë³„ ì§‘ê³„
        skip_cancel = 0
        skip_return = 0
        skip_old = 0  # 30ì¼ ì´ì „ ë°ì´í„°
        
        # 2ì£¼ ê¸°ì¤€ì¼
        days_14_ago = (datetime.now() - timedelta(days=14)).date()
        # 7ì¼ ê¸°ì¤€ì¼
        days_7_ago = (datetime.now() - timedelta(days=7)).date()
        # 7ì¼ ê¸°ì¤€ì¼
        days_7_ago = (datetime.now() - timedelta(days=7)).date()
        
        # stores ì‹œíŠ¸ì˜ ëª¨ë“  ê³„ì •ì„ 0ìœ¼ë¡œ ì´ˆê¸°í™”
        for store_key in all_store_keys:
            market_sales[store_key] = {
                "today_sales": 0,
                "today_orders": 0,
                "month_sales": 0,
                "month_orders": 0,
                "orders_2w": 0,  # 2ì£¼ ì£¼ë¬¸
                "orders_7d": 0,  # 7ì¼ ì£¼ë¬¸
                "month_profit": 0,
                "month_settlement": 0,
                "month_purchase": 0,
                "month_shipping": 0,
                "usage": account_usage.get(store_key, ""),
                "owner": account_owner.get(store_key, "")  # ì‹¤ì œ ì†Œìœ ì
            }
        
        # ê¸ˆì•¡ íŒŒì‹± í•¨ìˆ˜
        def parse_amount(row, idx):
            if idx < 0 or idx >= len(row):
                return 0
            val = row[idx].replace(",", "").replace("ì›", "").replace("â‚©", "").replace("%", "").strip()
            try:
                return int(float(val)) if val else 0
            except:
                return 0
        
        def parse_date(date_str):
            try:
                # 1. 2025-06-01 10:08 í˜•ì‹ (YYYY-MM-DD HH:MM)
                return datetime.strptime(date_str[:16], "%Y-%m-%d %H:%M").date()
            except:
                try:
                    # 2. 2025-06-01 í˜•ì‹ (YYYY-MM-DD)
                    return datetime.strptime(date_str[:10], "%Y-%m-%d").date()
                except:
                    return None
        
        # ë°ì´í„° ì§‘ê³„
        biz_sales = {}  # ì‚¬ì—…ìë²ˆí˜¸ë³„ ì§‘ê³„
        
        for row in all_data:
            if len(row) <= col_order_date:
                continue
            
            # ì‚¬ì—…ì + ë§ˆì¼“ ì¡°í•©
            owner_raw = row[col_owner].strip() if col_owner >= 0 and col_owner < len(row) else ""
            market_raw = row[col_market].strip() if col_market >= 0 and col_market < len(row) else ""
            order_date_str = row[col_order_date].strip() if col_order_date < len(row) else ""
            order_status = row[col_order_status].strip() if col_order_status >= 0 and col_order_status < len(row) else ""
            biz_number = row[col_biz_number].strip() if col_biz_number >= 0 and col_biz_number < len(row) else ""
            
            # ë§ˆì¼“ëª… ì •ê·œí™” (ë§¤ì¶œ ì‹œíŠ¸ â†’ ê³„ì •ëª©ë¡ í”Œë«í¼ëª…)
            market_normalized = market_raw
            if "ìŠ¤ë§ˆíŠ¸" in market_raw or "ë„¤ì´ë²„" in market_raw or market_raw.upper() == "SS":
                market_normalized = "ìŠ¤ë§ˆíŠ¸ìŠ¤í† ì–´"
            elif "11" in market_raw or market_raw.upper() == "ST":
                market_normalized = "11ë²ˆê°€"
            elif "ì¿ íŒ¡" in market_raw or market_raw.upper() == "CP":
                market_normalized = "ì¿ íŒ¡"
            elif "ì§€ë§ˆì¼“" in market_raw or market_raw.upper() == "GM":
                market_normalized = "ì§€ë§ˆì¼“"
            elif "ì˜¥ì…˜" in market_raw or market_raw.upper() == "AC":
                market_normalized = "ì˜¥ì…˜"
            
            if not owner_raw or not order_date_str:
                continue
            
            # 30ì¼ ì´ë‚´ ë°ì´í„°ë§Œ ì§‘ê³„
            order_date = parse_date(order_date_str)
            if not order_date or order_date < days_30_ago:
                skip_old += 1
                continue
            
            # ì·¨ì†Œì™„ë£Œ: ì§‘ê³„ì—ì„œ ì™„ì „ ì œì™¸
            if "ì·¨ì†Œì™„ë£Œ" in order_status:
                skip_cancel += 1
                continue
            
            # ì†Œìœ ì(ë§ˆì¼“) í‚¤ ìƒì„± - ì •ê·œí™”ëœ ë§ˆì¼“ëª… ì‚¬ìš©
            store_key = f"{owner_raw}({market_normalized})"
            
            # ë””ë²„ê·¸: ì²˜ìŒ 10ê°œ í–‰ì˜ í‚¤ì™€ ë§¤ì¹­ ê²°ê³¼
            if len(market_sales) < 10:
                matched_owner = account_owner.get(store_key, "")
                matched_usage = account_usage.get(store_key, "")
                in_account_list = store_key in all_store_keys
                print(f"[ë§¤ì¶œì§‘ê³„] ë§¤ì¶œì‹œíŠ¸ í‚¤: '{store_key}' â†’ ê³„ì •ëª©ë¡ ë§¤ì¹­={in_account_list}, owner='{matched_owner}', usage='{matched_usage}'")
            
            # ì¬ì´ë§ˆì¼“ ë””ë²„ê·¸
            if "ì¬ì´ë§ˆì¼“" in owner_raw:
                matched_owner = account_owner.get(store_key, "")
                in_account_list = store_key in all_store_keys
                print(f"[ë§¤ì¶œì§‘ê³„] â˜… ì¬ì´ë§ˆì¼“ ë°œê²¬: key='{store_key}', ë§¤ì¹­={in_account_list}, owner='{matched_owner}', ì›ë³¸ë§ˆì¼“='{market_raw}'")
            
            # ë°˜í’ˆì™„ë£Œ: ë§¤ì¶œì—ì„œ ì œì™¸
            is_return = "ë°˜í’ˆì™„ë£Œ" in order_status
            if is_return:
                skip_return += 1
                payment = 0
                settlement = 0
                profit = 0
                purchase = parse_amount(row, col_purchase)  # ë§¤ì…ì€ ìœ ì§€
                int_shipping = parse_amount(row, col_int_shipping)
                cargo_shipping = parse_amount(row, col_cargo_shipping)
            else:
                payment = parse_amount(row, col_payment)
                settlement = parse_amount(row, col_settlement)
                profit = parse_amount(row, col_profit)
                purchase = parse_amount(row, col_purchase)
                int_shipping = parse_amount(row, col_int_shipping)
                cargo_shipping = parse_amount(row, col_cargo_shipping)
            
            # [DEBUG RE-ADDED] ê°’ í™•ì¸ (ì²« 5ê°œë§Œ)
            if len(market_sales) < 5:
                print(f"[DEBUG_DATA_V2] Key: {store_key} | Payment Raw: '{row[col_payment]}' -> Parsed: {payment} | Return: {is_return}")

            # [DEBUG] íŠ¹ì • ìŠ¤í† ì–´ ìƒì„¸ í™•ì¸ (í•„ìš”ì‹œ í™œì„±í™”)
            # if "ì´ëª¨í‹°ë³´ì´" in store_key: 
            #     pass
            
            # [DEBUG] ê°’ í™•ì¸ (ì²« 5ê°œë§Œ)
            if len(market_sales) < 5:
                print(f"[DEBUG_DATA] Key: {store_key} | Date: {order_date} | Return: {is_return} | Payment Raw: '{row[col_payment]}' -> Parsed: {payment}")

            # [DEBUG] íŠ¹ì • ìŠ¤í† ì–´ ìƒì„¸ í™•ì¸
            if "ì´ëª¨í‹°ë³´ì´" in store_key and len(market_sales) < 20: 
                 print(f"[DEBUG_EMOTI] Key: {store_key} | Payment Raw: '{row[col_payment]}' | Status: '{order_status}'")
            
            # ìŠ¤í† ì–´ í‚¤ ì´ˆê¸°í™” (ëª©ë¡ì— ì—†ëŠ” ê²½ìš°)
            if store_key not in market_sales:
                # ë¨¼ì € store_keyë¡œ ë§¤ì¹­ ì‹œë„, ì—†ìœ¼ë©´ ì‚¬ì—…ìë²ˆí˜¸ë¡œ ë§¤ì¹­
                matched_owner = account_owner.get(store_key, "") or biz_to_owner.get(biz_number, "")
                matched_usage = account_usage.get(store_key, "") or biz_to_usage.get(biz_number, "")
                market_sales[store_key] = {
                    "today_sales": 0,
                    "today_orders": 0,
                    "month_sales": 0,
                    "month_orders": 0,
                    "orders_2w": 0,  # 2ì£¼ ì£¼ë¬¸
                    "orders_7d": 0,  # 7ì¼ ì£¼ë¬¸
                    "month_profit": 0,
                    "month_settlement": 0,
                    "month_purchase": 0,
                    "month_shipping": 0,
                    "usage": matched_usage,
                    "owner": matched_owner,
                    "biz_number": biz_number
                }
            # ê¸°ì¡´ í•­ëª©ë„ owner/usageê°€ ì—†ìœ¼ë©´ ì‚¬ì—…ìë²ˆí˜¸ë¡œ ë§¤ì¹­ ì‹œë„
            elif biz_number:
                if not market_sales[store_key].get("owner"):
                    market_sales[store_key]["owner"] = biz_to_owner.get(biz_number, "")
                if not market_sales[store_key].get("usage"):
                    market_sales[store_key]["usage"] = biz_to_usage.get(biz_number, "")
            
            # ì‚¬ì—…ìë²ˆí˜¸ë³„ ì§‘ê³„
            if biz_number and not is_return:
                if biz_number not in biz_sales:
                    biz_sales[biz_number] = {
                        "sales": 0, "settlement": 0, "profit": 0, 
                        "purchase": 0, "shipping": 0, "orders": 0,
                        "stores": set()
                    }
                biz_sales[biz_number]["sales"] += payment
                biz_sales[biz_number]["settlement"] += settlement
                biz_sales[biz_number]["profit"] += profit
                biz_sales[biz_number]["purchase"] += purchase
                biz_sales[biz_number]["shipping"] += int_shipping + cargo_shipping
                biz_sales[biz_number]["orders"] += 1
                biz_sales[biz_number]["stores"].add(owner_raw)
            
            # ì¼ì í‚¤
            date_key = order_date_str[:10]
            
            # ì¼ìë³„ ì´ˆê¸°í™”
            if date_key not in daily_sales:
                daily_sales[date_key] = {
                    "date": date_key,
                    "sales": 0,
                    "settlement": 0,
                    "purchase": 0,
                    "shipping": 0,
                    "profit": 0,
                    "orders": 0
                }
            
            # ì›” ë§¤ì¶œ/ìˆœìµ ì§‘ê³„
            if not is_return:
                market_sales[store_key]["month_sales"] += payment
                market_sales[store_key]["month_orders"] += 1
                market_sales[store_key]["month_profit"] += profit
                market_sales[store_key]["month_settlement"] += settlement
                market_sales[store_key]["month_purchase"] += purchase
                market_sales[store_key]["month_shipping"] += int_shipping + cargo_shipping
                
                # 2ì£¼ ì´ë‚´ ì£¼ë¬¸
                if order_date >= days_14_ago:
                    market_sales[store_key]["orders_2w"] = market_sales[store_key].get("orders_2w", 0) + 1
                
                # 7ì¼ ì´ë‚´ ì£¼ë¬¸
                if order_date >= days_7_ago:
                    market_sales[store_key]["orders_7d"] = market_sales[store_key].get("orders_7d", 0) + 1
                
                # 7ì¼ ì´ë‚´ ì£¼ë¬¸
                if order_date >= days_7_ago:
                    market_sales[store_key]["orders_7d"] = market_sales[store_key].get("orders_7d", 0) + 1
                
                # ì¼ìë³„ ì§‘ê³„
                daily_sales[date_key]["sales"] += payment
                daily_sales[date_key]["settlement"] += settlement
                daily_sales[date_key]["purchase"] += purchase
                daily_sales[date_key]["shipping"] += int_shipping + cargo_shipping
                daily_sales[date_key]["profit"] += profit
                daily_sales[date_key]["orders"] += 1
            
            # ì˜¤ëŠ˜ ë§¤ì¶œ ì§‘ê³„
            if order_date == today and not is_return:
                market_sales[store_key]["today_sales"] += payment
                market_sales[store_key]["today_orders"] += 1
        
        # ì¼ìë³„ ë°ì´í„° ì •ë ¬
        daily_list = sorted(daily_sales.values(), key=lambda x: x["date"])
        
        # ìºì‹œ ì—…ë°ì´íŠ¸
        sales_cache["data"] = market_sales
        sales_cache["daily"] = daily_list
        sales_cache["updated_at"] = datetime.now()
        
        
        # DEBUG: Write filtering stats
        with open("debug_filtering_stats.txt", "w", encoding="utf-8") as f:
            f.write(f"Total Rows: {len(all_data)}\n")
            f.write(f"Skipped Old: {skip_old}\n")
            f.write(f"Skipped Cancel: {skip_cancel}\n")
            f.write(f"Skipped Return: {skip_return}\n")
            f.write(f"Processed: {len(all_data) - skip_old - skip_cancel - skip_return}\n")
            f.write(f"Sample First 5 Keys: {list(market_sales.keys())[:5]}\n")
            
        print(f"[ë§¤ì¶œì§‘ê³„] {len(market_sales)}ê°œ ê³„ì • ì§‘ê³„ ì™„ë£Œ (ì·¨ì†Œ:{skip_cancel}, ë°˜í’ˆ:{skip_return}, 30ì¼ì´ì „:{skip_old} ì œì™¸)")
        
        # ì „ì²´ í•©ê³„ ê³„ì‚°
        total = {
            "sales": sum(m["month_sales"] for m in market_sales.values()),
            "settlement": sum(m["month_settlement"] for m in market_sales.values()),
            "purchase": sum(m["month_purchase"] for m in market_sales.values()),
            "shipping": sum(m["month_shipping"] for m in market_sales.values()),
            "profit": sum(m["month_profit"] for m in market_sales.values()),
            "orders": sum(m["month_orders"] for m in market_sales.values())
        }
        if total["sales"] > 0:
            total["profit_rate"] = round(total["profit"] / total["sales"] * 100, 2)
        else:
            total["profit_rate"] = 0
        
        # ì‚¬ì—…ìë²ˆí˜¸ë³„ ì§‘ê³„ ë°ì´í„° ë³€í™˜ (set -> list)
        biz_sales_result = {}
        for biz_num, data in biz_sales.items():
            biz_sales_result[biz_num] = {
                **{k: v for k, v in data.items() if k != 'stores'},
                "stores": sorted(list(biz_to_stores.get(biz_num, data["stores"])))  # ê³„ì •ëª©ë¡ì˜ ëª¨ë“  ìŠ¤í† ì–´ (ì—†ìœ¼ë©´ ë§¤ì¶œ ìˆëŠ” ìŠ¤í† ì–´)
            }
        
        # ë§¤ì¶œì€ ì—†ì§€ë§Œ ê³„ì •ëª©ë¡ì— ìˆëŠ” ì‚¬ì—…ìë²ˆí˜¸ë„ ì¶”ê°€
        for biz_num, stores in biz_to_stores.items():
            if biz_num not in biz_sales_result:
                biz_sales_result[biz_num] = {
                    "sales": 0, "settlement": 0, "profit": 0,
                    "purchase": 0, "shipping": 0, "orders": 0,
                    "stores": sorted(list(stores))
                }
        
        
        return {
            "success": True, 
            "data": market_sales,
            "daily": daily_list,
            "total": total,
            "biz_sales": biz_sales_result,  # ì‚¬ì—…ìë²ˆí˜¸ë³„ ì§‘ê³„
            "count": len(market_sales),
            "tab": current_tab
        }
        
    except Exception as e:
        print(f"[ë§¤ì¶œì§‘ê³„] ì˜¤ë¥˜: {e}")
        import traceback
        traceback.print_exc()
        return {"success": False, "message": str(e)}


@app.get("/api/sales/top-products")
async def get_top_products(request: Request, limit: int = 40):
    """ì›”ê°„ TOP íŒë§¤ ìƒí’ˆ ì¡°íšŒ"""
    get_current_user(request)
    
    from datetime import datetime
    from collections import defaultdict
    
    try:
        # í˜„ì¬ ì›” íƒ­ ì´ë¦„
        current_month = datetime.now().month
        current_tab = f"{current_month}ì›”"
        today = datetime.now().date()
        days_30_ago = today - timedelta(days=30)
        
        # ì‹œíŠ¸ ì—´ê¸°
        sales_sheet = gsheet.client.open_by_key(SALES_SHEET_ID)
        
        all_data = []
        headers = None
        
        # ì´ë²ˆë‹¬ ë°ì´í„° ë¡œë“œ
        try:
            ws = sales_sheet.worksheet(current_tab)
            tab_data = ws.get_all_values()
            if len(tab_data) >= 3:
                headers = tab_data[1]  # 2í–‰ì´ í—¤ë”
                all_data = tab_data[2:]  # 3í–‰ë¶€í„° ë°ì´í„°
        except Exception as e:
            print(f"[TOPìƒí’ˆ] {current_tab} íƒ­ ë¡œë“œ ì‹¤íŒ¨: {e}")
            return {"success": False, "message": f"íƒ­ ë¡œë“œ ì‹¤íŒ¨: {e}"}
        
        if not headers or len(all_data) < 1:
            return {"success": True, "data": [], "message": "ë°ì´í„° ì—†ìŒ"}
        
        # ì»¬ëŸ¼ ì¸ë±ìŠ¤ ì°¾ê¸°
        def find_col(names):
            for name in names:
                for idx, h in enumerate(headers):
                    h_clean = h.replace('\n', '').replace('\r', '').replace(' ', '')
                    name_clean = name.replace(' ', '')
                    if name_clean in h_clean:
                        return idx
            return -1
        
        def col_letter_to_idx(letter):
            letter = letter.upper()
            result = 0
            for char in letter:
                result = result * 26 + (ord(char) - ord('A') + 1)
            return result - 1
        
        col_market = find_col(["ë§ˆì¼“"])
        col_owner = find_col(["ì‚¬ì—…ì"])  # ìŠ¤í† ì–´ëª…ìœ¼ë¡œ ì‚¬ìš©
        col_order_date = find_col(["ì£¼ë¬¸ì¼ì"])
        col_order_status = find_col(["ì£¼ë¬¸í˜„í™©"])
        col_product_name = find_col(["ìƒí’ˆëª…", "í’ˆëª…", "ì œí’ˆëª…"])
        col_quantity = find_col(["ìˆ˜ëŸ‰", "ì£¼ë¬¸ìˆ˜ëŸ‰"])
        col_payment = find_col(["ì‹¤ê²°ì œê¸ˆì•¡(ë°°ì†¡ë¹„í¬í•¨)", "ì‹¤ê²°ì œê¸ˆì•¡"])
        col_seller_code = find_col(["íŒë§¤ììƒí’ˆì½”ë“œ", "ìƒí’ˆì½”ë“œ", "íŒë§¤ì ìƒí’ˆì½”ë“œ"])
        
        # ëª» ì°¾ì€ ì»¬ëŸ¼ ì§ì ‘ ì§€ì •
        if col_order_status < 0: col_order_status = col_letter_to_idx('D')
        if col_market < 0: col_market = col_letter_to_idx('E')
        if col_owner < 0: col_owner = col_letter_to_idx('F')
        if col_order_date < 0: col_order_date = col_letter_to_idx('G')
        if col_product_name < 0: col_product_name = col_letter_to_idx('K')  # Kì—´: ìƒí’ˆëª…
        if col_seller_code < 0: col_seller_code = col_letter_to_idx('J')  # Jì—´: íŒë§¤ììƒí’ˆì½”ë“œ
        if col_quantity < 0: col_quantity = col_letter_to_idx('T')  # Tì—´: ìˆ˜ëŸ‰
        if col_payment < 0: col_payment = col_letter_to_idx('X')
        
        # í”Œë«í¼ ìŠ¤í ë§ ë³€í™˜
        def get_platform_short(platform):
            platform = platform.strip()
            if 'ìŠ¤ë§ˆíŠ¸ìŠ¤í† ì–´' in platform or 'ë„¤ì´ë²„' in platform:
                return 'N'
            elif 'ì¿ íŒ¡' in platform:
                return 'C'
            elif '11ë²ˆê°€' in platform:
                return '11'
            elif 'ì§€ë§ˆì¼“' in platform:
                return 'G'
            elif 'ì˜¥ì…˜' in platform:
                return 'A'
            return platform[:2] if platform else '-'
        
        print(f"[TOPìƒí’ˆ] ì»¬ëŸ¼ - ë§ˆì¼“:{col_market}, ìŠ¤í† ì–´:{col_owner}, ìƒí’ˆëª…:{col_product_name}, íŒë§¤ìì½”ë“œ:{col_seller_code}")
        
        # ìƒí’ˆë³„ ì§‘ê³„ (íŒë§¤ììƒí’ˆì½”ë“œ + ìŠ¤í† ì–´ ê¸°ì¤€)
        product_sales = defaultdict(lambda: {"order_count": 0, "total_quantity": 0, "total_sales": 0, "ìŠ¤í† ì–´ëª…": "", "platform": "", "product_name": "", "seller_code": ""})
        
        for row in all_data:
            if len(row) <= max(col_market, col_order_date, col_payment, col_product_name):
                continue
            
            # ì£¼ë¬¸í˜„í™© ì²´í¬ (ì·¨ì†Œ/ë°˜í’ˆ ì œì™¸)
            order_status = row[col_order_status] if col_order_status < len(row) else ""
            if any(x in order_status for x in ["ì·¨ì†Œ", "ë°˜í’ˆ", "í™˜ë¶ˆ"]):
                continue
            
            # ì£¼ë¬¸ì¼ì ì²´í¬ (30ì¼ ì´ë‚´)
            try:
                date_str = row[col_order_date].strip()
                if len(date_str) >= 10:
                    order_date = datetime.strptime(date_str[:10], "%Y-%m-%d").date()
                    if order_date < days_30_ago:
                        continue
            except:
                continue
            
            # ë°ì´í„° ì¶”ì¶œ
            platform = row[col_market].strip() if col_market < len(row) else ""
            store_name = row[col_owner].strip() if col_owner < len(row) else ""
            product_name = row[col_product_name].strip() if col_product_name < len(row) else ""
            seller_code = row[col_seller_code].strip() if col_seller_code < len(row) else ""
            
            if not product_name:
                continue
            
            # ìˆ˜ëŸ‰ íŒŒì‹±
            quantity = 1
            if col_quantity < len(row):
                try:
                    qty_str = row[col_quantity].replace(",", "").strip()
                    if qty_str:
                        quantity = int(float(qty_str))
                except:
                    pass
            
            # ê¸ˆì•¡ íŒŒì‹±
            payment = 0
            if col_payment < len(row):
                try:
                    pay_str = row[col_payment].replace(",", "").replace("ì›", "").strip()
                    if pay_str:
                        payment = int(float(pay_str))
                except:
                    pass
            
            # ìƒí’ˆ í‚¤ ìƒì„± (ìŠ¤í† ì–´ëª… + íŒë§¤ììƒí’ˆì½”ë“œ ë˜ëŠ” ìƒí’ˆëª…)
            key = f"{store_name}||{seller_code or product_name}"
            product_sales[key]["order_count"] += 1  # ì£¼ë¬¸ ê±´ìˆ˜ (ìˆ˜ëŸ‰ê³¼ ìƒê´€ì—†ì´ 1ê±´)
            product_sales[key]["total_quantity"] += quantity  # ì´ ìˆ˜ëŸ‰
            product_sales[key]["total_sales"] += payment
            product_sales[key]["ìŠ¤í† ì–´ëª…"] = store_name
            product_sales[key]["platform"] = get_platform_short(platform)
            product_sales[key]["product_name"] = product_name
            product_sales[key]["seller_code"] = seller_code
        
        # TOP N ì •ë ¬ (ì£¼ë¬¸ ê±´ìˆ˜ ê¸°ì¤€)
        sorted_products = sorted(
            product_sales.values(),
            key=lambda x: x["order_count"],
            reverse=True
        )[:limit]
        
        print(f"[TOPìƒí’ˆ] {len(product_sales)}ê°œ ìƒí’ˆ ì¤‘ TOP {limit} ë°˜í™˜")
        
        return {
            "success": True,
            "data": sorted_products,
            "total_products": len(product_sales)
        }
        
    except Exception as e:
        print(f"[TOPìƒí’ˆ] ì˜¤ë¥˜: {e}")
        import traceback
        traceback.print_exc()
        return {"success": False, "message": str(e)}


# ========== ë§¤ì¶œ ì¡°íšŒ API ==========
class SalesQueryRequest(BaseModel):
    stores: List[str]  # store_name ëª©ë¡
    platform: str = "ìŠ¤ë§ˆíŠ¸ìŠ¤í† ì–´"

# ë§¤ì¶œ ì¡°íšŒ ìƒíƒœ
sales_query_status = {
    "running": False,
    "progress": {},
    "logs": [],
    "stop_requested": False
}

def query_smartstore_sales(store_name: str, client_id: str, client_secret: str):
    """ìŠ¤ë§ˆíŠ¸ìŠ¤í† ì–´ ë§¤ì¶œ ì¡°íšŒ"""
    global sales_query_status
    
    def add_log(msg, status="info"):
        log_entry = {
            "time": datetime.now().strftime("%H:%M:%S"),
            "store": store_name,
            "msg": msg,
            "status": status
        }
        sales_query_status["logs"].append(log_entry)
        print(f"[ë§¤ì¶œ-{store_name}] {msg}")
    
    try:
        sales_query_status["progress"][store_name] = {"status": "í† í° ë°œê¸‰ ì¤‘...", "today": 0, "month": 0}
        
        # í† í° ë°œê¸‰ (ê¸°ì¡´ í•¨ìˆ˜ ì‚¬ìš©)
        token = get_naver_token(client_id, client_secret)
        headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json"
        }
        
        add_log("í† í° ë°œê¸‰ ì™„ë£Œ")
        sales_query_status["progress"][store_name]["status"] = "ë§¤ì¶œ ì¡°íšŒ ì¤‘..."
        
        base_url = "https://api.commerce.naver.com/external"
        
        # ì˜¤ëŠ˜ ë‚ ì§œ
        today = datetime.now()
        today_str = today.strftime("%Y-%m-%d")
        month_start = today.replace(day=1).strftime("%Y-%m-%d")
        
        # ì£¼ë¬¸ ì¡°íšŒ (ì˜¤ëŠ˜)
        today_sales = 0
        today_orders = 0
        
        body = {
            "productOrderStatuses": ["PAYED", "DELIVERING", "DELIVERED", "PURCHASE_DECIDED"],
            "startPayedDate": f"{today_str}T00:00:00",
            "endPayedDate": f"{today_str}T23:59:59"
        }
        
        try:
            resp = requests.post(f"{base_url}/v1/pay-order/seller/product-orders/search", 
                               headers=headers, json=body)
            if resp.status_code == 200:
                orders = resp.json().get("data", [])
                today_orders = len(orders)
                for order in orders:
                    today_sales += int(order.get("totalPaymentAmount", 0))
        except Exception as e:
            add_log(f"ì˜¤ëŠ˜ ë§¤ì¶œ ì¡°íšŒ ì˜¤ë¥˜: {e}", "error")
        
        # ì´ë‹¬ ë§¤ì¶œ
        month_sales = 0
        month_orders = 0
        
        body["startPayedDate"] = f"{month_start}T00:00:00"
        body["endPayedDate"] = f"{today_str}T23:59:59"
        
        try:
            resp = requests.post(f"{base_url}/v1/pay-order/seller/product-orders/search", 
                               headers=headers, json=body)
            if resp.status_code == 200:
                orders = resp.json().get("data", [])
                month_orders = len(orders)
                for order in orders:
                    month_sales += int(order.get("totalPaymentAmount", 0))
        except Exception as e:
            add_log(f"ì´ë‹¬ ë§¤ì¶œ ì¡°íšŒ ì˜¤ë¥˜: {e}", "error")
        
        today_sales_str = format(today_sales, ',')
        month_sales_str = format(month_sales, ',')
        
        sales_query_status["progress"][store_name] = {
            "status": f"ì™„ë£Œ (ì˜¤ëŠ˜ {today_sales_str}ì›)",
            "today_sales": today_sales,
            "today_orders": today_orders,
            "month_sales": month_sales,
            "month_orders": month_orders
        }
        
        add_log(f"ì˜¤ëŠ˜ {today_sales_str}ì› ({today_orders}ê±´) / ì´ë‹¬ {month_sales_str}ì› ({month_orders}ê±´)", "success")
        
        return {
            "today_sales": today_sales,
            "today_orders": today_orders,
            "month_sales": month_sales,
            "month_orders": month_orders
        }
        
    except Exception as e:
        add_log(f"ì˜¤ë¥˜: {str(e)}", "error")
        sales_query_status["progress"][store_name]["status"] = f"ì˜¤ë¥˜: {str(e)[:30]}"
        return None

@app.post("/api/allinone/sales-query")
async def run_sales_query(request: Request, req: SalesQueryRequest):
    """ë§¤ì¶œ ì¡°íšŒ ì‹¤í–‰"""
    require_permission(request, "edit")

    global sales_query_status

    if sales_query_status["running"]:
        return {"success": False, "message": "ì´ë¯¸ ì‹¤í–‰ ì¤‘ì…ë‹ˆë‹¤"}

    if not req.stores:
        return {"success": False, "message": "ìŠ¤í† ì–´ë¥¼ ì„ íƒí•˜ì„¸ìš”"}

    # ì‘ì—… ë¡œê·¸ ê¸°ë¡
    store_names = ", ".join(req.stores[:5]) + ("..." if len(req.stores) > 5 else "")
    log_work("ë§¤ì¶œì¡°íšŒ", "ìŠ¤ë§ˆíŠ¸ìŠ¤í† ì–´", len(req.stores), f"ëŒ€ìƒ: {store_names}", "ì›¹")
    
    # ê³„ì •ëª©ë¡ ì‹œíŠ¸ì—ì„œ API ì •ë³´ ê°€ì ¸ì˜¤ê¸°
    try:
        ws = gsheet.sheet.worksheet(ACCOUNTS_TAB)  # ê³„ì •ëª©ë¡
        data = ws.get_all_values()
        headers = data[0]

        # ì»¬ëŸ¼ ì¸ë±ìŠ¤ ì°¾ê¸° (ê³„ì •ëª©ë¡ ì»¬ëŸ¼ëª…)
        name_idx = None
        id_idx = None
        secret_idx = None
        platform_idx = None
        for i, h in enumerate(headers):
            if h in ["ìŠ¤í† ì–´ëª…", "store_name"]:
                name_idx = i
            elif h in ["ìŠ¤ë§ˆíŠ¸ìŠ¤í† ì–´ ì• í”Œë¦¬ì¼€ì´ì…˜ ID", "client_id"]:
                id_idx = i
            elif h in ["ìŠ¤ë§ˆíŠ¸ìŠ¤í† ì–´ ì• í”Œë¦¬ì¼€ì´ì…˜ ì‹œí¬ë¦¿", "client_secret"]:
                secret_idx = i
            elif h in ["í”Œë«í¼", "platform"]:
                platform_idx = i

        if None in [name_idx, id_idx, secret_idx]:
            return {"success": False, "message": "ì‹œíŠ¸ì— í•„ìš”í•œ ì»¬ëŸ¼ì´ ì—†ìŠµë‹ˆë‹¤ (ìŠ¤í† ì–´ëª…, ìŠ¤ë§ˆíŠ¸ìŠ¤í† ì–´ ì• í”Œë¦¬ì¼€ì´ì…˜ ID/ì‹œí¬ë¦¿)"}

        store_info = {}
        for row in data[1:]:
            if len(row) > max(name_idx, id_idx, secret_idx):
                # í”Œë«í¼ ì²´í¬ (ìŠ¤ë§ˆíŠ¸ìŠ¤í† ì–´ë§Œ)
                platform = row[platform_idx].lower() if platform_idx and len(row) > platform_idx else ""
                if platform and platform not in ["ìŠ¤ë§ˆíŠ¸ìŠ¤í† ì–´", "smartstore", "ë„¤ì´ë²„", "naver"]:
                    continue
                name = row[name_idx]
                if name in req.stores:
                    store_info[name] = {
                        "client_id": row[id_idx],
                        "client_secret": row[secret_idx]
                    }

        if not store_info:
            return {"success": False, "message": "ì„ íƒëœ ìŠ¤í† ì–´ ì •ë³´ë¥¼ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤"}

    except Exception as e:
        return {"success": False, "message": f"ì‹œíŠ¸ ì¡°íšŒ ì˜¤ë¥˜: {str(e)}"}

    # ìƒíƒœ ì´ˆê¸°í™”
    sales_query_status = {
        "running": True,
        "progress": {},
        "logs": [],
        "stop_requested": False
    }
    
    def run_parallel():
        global sales_query_status
        
        try:
            from concurrent.futures import ThreadPoolExecutor, as_completed
            
            with ThreadPoolExecutor(max_workers=min(5, len(store_info))) as executor:
                futures = {}
                for store_name, info in store_info.items():
                    future = executor.submit(
                        query_smartstore_sales,
                        store_name,
                        info["client_id"],
                        info["client_secret"]
                    )
                    futures[future] = store_name
                
                for future in as_completed(futures):
                    store_name = futures[future]
                    try:
                        result = future.result()
                    except Exception as e:
                        print(f"[ë§¤ì¶œ] {store_name} ìŠ¤ë ˆë“œ ì˜¤ë¥˜: {e}")
        
        finally:
            sales_query_status["running"] = False
    
    import threading
    threading.Thread(target=run_parallel, daemon=True).start()
    
    return {"success": True, "message": f"{len(store_info)}ê°œ ìŠ¤í† ì–´ ë§¤ì¶œ ì¡°íšŒ ì‹œì‘"}

@app.get("/api/allinone/sales-progress")
async def get_sales_progress(request: Request):
    """ë§¤ì¶œ ì¡°íšŒ ì§„í–‰ìƒí™©"""
    require_permission(request, "edit")
    return sales_query_status

@app.post("/api/allinone/sales-stop")
async def stop_sales_query(request: Request):
    """ë§¤ì¶œ ì¡°íšŒ ì¤‘ì§€"""
    require_permission(request, "edit")
    global sales_query_status
    sales_query_status["stop_requested"] = True
    return {"success": True, "message": "ì¤‘ì§€ ìš”ì²­ë¨"}


class AioSingleRunRequest(BaseModel):
    platform: str
    login_id: str
    task: str

@app.post("/api/allinone/run-single")
async def run_single_allinone_task(request: Request, req: AioSingleRunRequest):
    """ê°œë³„ ê³„ì • ì˜¬ì¸ì› ì‘ì—… ì‹¤í–‰"""
    require_permission(request, "edit")

    # í”Œë«í¼ë³„ ìƒíƒœ ê°€ì ¸ì˜¤ê¸°
    platform_status = get_aio_status(req.platform)

    if platform_status["running"]:
        return {"success": False, "message": f"{req.platform} ì‘ì—…ì´ ì‹¤í–‰ ì¤‘ì…ë‹ˆë‹¤"}

    print(f"[ì˜¬ì¸ì›] ê°œë³„ ì‹¤í–‰: {req.platform} / {req.login_id} / {req.task}")

    # ì‘ì—… ë¡œê·¸ ê¸°ë¡
    log_work(f"ì˜¬ì¸ì›-{req.task}", req.platform, 1, f"ê³„ì •: {req.login_id}", "ì›¹")
    
    try:
        if req.platform == "ìŠ¤ë§ˆíŠ¸ìŠ¤í† ì–´":
            # stores ì‹œíŠ¸ì—ì„œ í•´ë‹¹ ê³„ì •ë§Œ active ì„¤ì •
            ws_stores = gsheet.sheet.worksheet("stores")
            data = ws_stores.get_all_values()
            headers = data[0]
            
            # activeì™€ shop_alias ì»¬ëŸ¼ ì°¾ê¸°
            active_col = headers.index("active") + 1 if "active" in headers else 1
            shop_col = headers.index("shop_alias") + 1 if "shop_alias" in headers else 2
            
            # ëª¨ë“  ê³„ì • ë¹„í™œì„±í™” í›„ ëŒ€ìƒ ê³„ì •ë§Œ í™œì„±í™”
            updates = []
            target_row = None
            for i, row in enumerate(data[1:], start=2):
                shop_alias = row[shop_col - 1] if len(row) >= shop_col else ""
                if shop_alias == req.login_id:
                    updates.append({"range": f"{chr(64+active_col)}{i}", "values": [["TRUE"]]})
                    target_row = i
                else:
                    updates.append({"range": f"{chr(64+active_col)}{i}", "values": [["FALSE"]]})
            
            if not target_row:
                return {"success": False, "message": "ê³„ì •ì„ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤"}
            
            # ì¼ê´„ ì—…ë°ì´íŠ¸
            ws_stores.batch_update(updates)
            
            # ì‘ì—… ì„¤ì •
            ws_stores.update_acell("A1", req.task)
            
            # ë¡œê·¸ íŒŒì¼ ê²½ë¡œ
            log_file = os.path.join(os.path.dirname(__file__), "logs", "allinone.log")
            os.makedirs(os.path.dirname(log_file), exist_ok=True)
            
            # ìƒíƒœ ì´ˆê¸°í™”
            aio_status = {
                "running": True,
                "process": None,
                "status": "running",
                "progress": 0,
                "results": [],
                "current_store": req.login_id,
                "current_action": f"{req.task} ì¤€ë¹„ ì¤‘...",
                "total": 1,
                "completed": 0,
                "logs": [],
                "log_file": log_file,
                "log_pos": 0
            }
            
            # ë¡œê·¸ íŒŒì¼ ì´ˆê¸°í™”
            with open(log_file, "w", encoding="utf-8") as f:
                f.write(f"[{datetime.now().strftime('%H:%M:%S')}] {req.login_id} - {req.task} ì‹œì‘\n")
            
            # subprocessë¡œ ì‹¤í–‰
            script_path = r"C:\autosystem\smartstore_all_in_one_v1_1.py"
            env = os.environ.copy()
            env["PYTHONIOENCODING"] = "utf-8"
            
            with open(log_file, "a", encoding="utf-8") as log_f:
                process = subprocess.Popen(
                    ["python", script_path],
                    stdout=log_f,
                    stderr=subprocess.STDOUT,
                    env=env,
                    cwd=os.path.dirname(script_path)
                )
            
            aio_status["process"] = process
            
            return {"success": True, "message": f"{req.login_id} - {req.task} ì‹œì‘ë¨"}
            
        elif req.platform == "11ë²ˆê°€":
            # 11ë²ˆê°€ ì‹œíŠ¸ì—ì„œ í•´ë‹¹ ê³„ì •ë§Œ active ì„¤ì •
            ws = gsheet.sheet.worksheet("11ë²ˆê°€")
            data = ws.get_all_values()
            headers = data[0]
            
            active_col = headers.index("active") + 1 if "active" in headers else 1
            store_col = headers.index("store_name") + 1 if "store_name" in headers else 2
            
            updates = []
            target_row = None
            for i, row in enumerate(data[1:], start=2):
                store_name = row[store_col - 1] if len(row) >= store_col else ""
                if store_name == req.login_id:
                    updates.append({"range": f"{chr(64+active_col)}{i}", "values": [["TRUE"]]})
                    target_row = i
                else:
                    updates.append({"range": f"{chr(64+active_col)}{i}", "values": [["FALSE"]]})
            
            if not target_row:
                return {"success": False, "message": "ê³„ì •ì„ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤"}
            
            ws.batch_update(updates)
            
            # ë¡œê·¸ íŒŒì¼ ê²½ë¡œ
            log_file = os.path.join(os.path.dirname(__file__), "logs", "allinone.log")
            os.makedirs(os.path.dirname(log_file), exist_ok=True)
            
            aio_status = {
                "running": True,
                "process": None,
                "status": "running",
                "progress": 0,
                "results": [],
                "current_store": req.login_id,
                "current_action": f"{req.task} ì¤€ë¹„ ì¤‘...",
                "total": 1,
                "completed": 0,
                "logs": [],
                "log_file": log_file,
                "log_pos": 0
            }
            
            with open(log_file, "w", encoding="utf-8") as f:
                f.write(f"[{datetime.now().strftime('%H:%M:%S')}] {req.login_id} - {req.task} ì‹œì‘\n")
            
            module_path = os.path.join(os.path.dirname(__file__), "modules", "elevenst.py")
            env = os.environ.copy()
            env["PYTHONIOENCODING"] = "utf-8"
            env["ELEVENST_TASK"] = req.task
            
            service_account_path = os.getenv("GOOGLE_SERVICE_ACCOUNT_FILE", "")
            if service_account_path:
                env["SERVICE_ACCOUNT_JSON"] = service_account_path
                env["GOOGLE_SERVICE_ACCOUNT_FILE"] = service_account_path
            spreadsheet_key = os.getenv("SPREADSHEET_KEY", "")
            if spreadsheet_key:
                env["SPREADSHEET_KEY"] = spreadsheet_key
            
            with open(log_file, "a", encoding="utf-8") as log_f:
                process = subprocess.Popen(
                    ["python", module_path],
                    stdout=log_f,
                    stderr=subprocess.STDOUT,
                    env=env,
                    cwd=os.path.dirname(module_path)
                )
            
            aio_status["process"] = process
            
            return {"success": True, "message": f"{req.login_id} - {req.task} ì‹œì‘ë¨"}
        
        else:
            return {"success": False, "message": f"{req.platform}ì€ ì•„ì§ ì§€ì›í•˜ì§€ ì•ŠìŠµë‹ˆë‹¤"}
            
    except Exception as e:
        print(f"[ì˜¬ì¸ì›] ê°œë³„ ì‹¤í–‰ ì˜¤ë¥˜: {e}")
        import traceback
        traceback.print_exc()
        return {"success": False, "message": str(e)}

# ========== ì¼ì¼ì¥ë¶€ ì£¼ë¬¸/ë°°ì†¡ ë™ê¸°í™” API ==========

sync_state = {
    "status": "ready",  # ready, running, completed, error
    "logs": [],
    "last_check_index": 0
}

def add_sync_log(message: str, type: str = "info"):
    sync_state["logs"].append({
        "timestamp": datetime.now().strftime("%H:%M:%S"),
        "message": message,
        "type": type
    })
    if len(sync_state["logs"]) > 200:
        sync_state["logs"] = sync_state["logs"][-200:]
    print(f"[ë™ê¸°í™”] {message}")

@app.post("/api/sync/daily-journal")
async def start_sync_daily_journal(
    request: Request,
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    sheet_url: str = Form(...),
    month: str = Form(...),
    sync_order_info: str = Form("false"),
    sync_logistics: str = Form("false")
):
    try:
        require_permission(request, "edit")

        if sync_state["status"] == "running":
            return {"success": False, "message": "ì´ë¯¸ ë™ê¸°í™”ê°€ ì§„í–‰ ì¤‘ì…ë‹ˆë‹¤."}

        # ì‘ì—… ë¡œê·¸ ê¸°ë¡
        log_work("ì¼ì¼ì¥ë¶€ë™ê¸°í™”", month, 0, f"íŒŒì¼: {file.filename}", "ì›¹")
        
        # ìƒíƒœ ì´ˆê¸°í™”
        sync_state["status"] = "running"
        sync_state["logs"] = []
        sync_state["last_check_index"] = 0
        
        # íŒŒì¼ ì½ê¸°
        try:
            content = await file.read()
            # pandasë¡œ ì½ê¸° ê°€ëŠ¥í•œ í˜•íƒœë¡œ ì „í™˜
            if file.filename.endswith('.csv'):
                df_source = pd.read_csv(io.BytesIO(content))
            else:
                # xlsx, xls ë“± ì²˜ë¦¬
                df_source = pd.read_excel(io.BytesIO(content))
                
            add_sync_log(f"íŒŒì¼ ì—…ë¡œë“œ ì„±ê³µ: {file.filename} ({len(df_source)}í–‰)")
        except Exception as e:
            sync_state["status"] = "error"
            print(f"[ë™ê¸°í™”] íŒŒì¼ ì½ê¸° ì˜¤ë¥˜: {e}")
            return {"success": False, "message": f"íŒŒì¼ ì½ê¸° ì‹¤íŒ¨: {e}"}
        
        # ì‘ì—… ë°ì´í„° êµ¬ì„±
        job_data = {
            "df_source": df_source,
            "sheet_url": sheet_url,
            "month": month,
            "sync_order_info": sync_order_info.lower() == "true",
            "sync_logistics": sync_logistics.lower() == "true"
        }
        
        background_tasks.add_task(run_daily_journal_sync, job_data)
        return {"success": True}
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        sync_state["status"] = "error"
        return {"success": False, "message": f"ì„œë²„ ì˜¤ë¥˜: {str(e)}"}

@app.get("/api/sync/status")
async def get_sync_status():
    # ë§ˆì§€ë§‰ í™•ì¸ ì´í›„ì˜ ë¡œê·¸ë§Œ ë°˜í™˜
    logs = sync_state["logs"][sync_state["last_check_index"]:]
    sync_state["last_check_index"] = len(sync_state["logs"])
    
    return {
        "status": sync_state["status"],
        "logs": logs
    }

async def run_daily_journal_sync(data: dict):
    """ì¼ì¼ì¥ë¶€ ë™ê¸°í™” ë°±ê·¸ë¼ìš´ë“œ ì‘ì—… (ì—‘ì…€ ê¸°ë°˜) - ë³„ë„ ëª¨ë“ˆ ë¶„ë¦¬ë¨"""
    global sync_state
    
    sync_state["status"] = "running"
    sync_state["logs"] = []

    def update_state(key, value):
        sync_state[key] = value

    try:
        month = data["month"]
        sheet_url = data["sheet_url"]
        df_source = data["df_source"]
        
        add_sync_log(f"ë™ê¸°í™” ì‘ì—… ì‹œì‘: {month} (ë°ì´í„° {len(df_source)}ê±´)")
        
        # ì¸ì¦ íŒŒì¼ ê²½ë¡œ
        sync_credentials_path = str(APP_DIR / "autosms-466614-951e91617c69.json")
        
        # ëª¨ë“ˆ ì‹¤í–‰
        syncer = DailyJournalSyncer(sync_credentials_path)
        syncer.run_sync(sheet_url, month, df_source, add_sync_log, update_state)

    except Exception as e:
        import traceback
        traceback.print_exc()
        add_sync_log(f"ì‹¤í–‰ ì¤‘ ì˜ˆì™¸ ë°œìƒ: {e}", "error")
        sync_state["status"] = "error"

# ========== ì•Œë¦¬ ì†¡ì¥ ìˆ˜ì§‘ API ==========

ali_collect_status = {
    "running": False,
    "connected": False,
    "progress": 0,
    "total": 0,
    "current": "",
    "logs": [],
    "collected": []  # ìˆ˜ì§‘ëœ ì†¡ì¥ ë°ì´í„°
}

ali_browser = None  # playwright browser instance (async)
ali_playwright = None  # playwright instance (async)
ali_sync_browser = None  # playwright sync browser for collection
ali_sync_playwright = None  # playwright sync instance
ali_debug_port = None  # ì—°ê²°ëœ Chrome ë””ë²„ê·¸ í¬íŠ¸

def find_available_port(start: int = 9300, end: int = 9320) -> int:
    """ì‚¬ìš© ê°€ëŠ¥í•œ í¬íŠ¸ ìë™ ì°¾ê¸°"""
    import socket
    
    for port in range(start, end):
        try:
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.settimeout(1)
            result = sock.connect_ex(('127.0.0.1', port))
            sock.close()
            
            if result != 0:  # í¬íŠ¸ ì‚¬ìš© ì•ˆ í•¨
                print(f"[ì•Œë¦¬] ì‚¬ìš© ê°€ëŠ¥í•œ í¬íŠ¸ ë°œê²¬: {port}")
                return port
            else:
                print(f"[ì•Œë¦¬] í¬íŠ¸ {port} ì‚¬ìš© ì¤‘, ë‹¤ìŒ í¬íŠ¸ í™•ì¸...")
        except:
            continue
    
    print(f"[ì•Œë¦¬] ì‚¬ìš© ê°€ëŠ¥í•œ í¬íŠ¸ ì—†ìŒ, ê¸°ë³¸ê°’ {start} ì‚¬ìš©")
    return start

def find_existing_ali_chrome() -> int:
    """ê¸°ì¡´ ì‹¤í–‰ ì¤‘ì¸ ì•Œë¦¬ Chrome í¬íŠ¸ ì°¾ê¸° (9300~9320 ë²”ìœ„)"""
    import socket
    
    for port in range(9300, 9320):
        try:
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.settimeout(1)
            result = sock.connect_ex(('127.0.0.1', port))
            sock.close()
            
            if result == 0:  # í¬íŠ¸ ì‚¬ìš© ì¤‘
                print(f"[ì•Œë¦¬] ê¸°ì¡´ Chrome ê°ì§€: í¬íŠ¸ {port}")
                return port
        except:
            continue
    
    return None

@app.post("/api/ali/connect")
async def connect_ali_browser(request: Request):
    """ì•Œë¦¬ ë¸Œë¼ìš°ì € ì—°ê²° (ê¸°ì¡´ Chrome ì—°ê²° ë˜ëŠ” ìƒˆë¡œ ì‹¤í–‰)"""
    get_current_user(request)
    
    data = await request.json()
    requested_port = int(data.get("debug_port", 9300))
    
    global ali_browser, ali_playwright, ali_collect_status, ali_debug_port
    
    try:
        from playwright.async_api import async_playwright
        import subprocess
        import shutil
        
        # 1. ê¸°ì¡´ ì•Œë¦¬ Chromeì´ ìˆëŠ”ì§€ í™•ì¸ (9300~9320 ë²”ìœ„)
        existing_port = find_existing_ali_chrome()
        
        if existing_port:
            # ê¸°ì¡´ Chromeì— ì—°ê²° ì‹œë„
            try:
                ali_playwright = await async_playwright().start()
                ali_browser = await ali_playwright.chromium.connect_over_cdp(f"http://localhost:{existing_port}")
                
                ali_debug_port = existing_port  # í¬íŠ¸ ì €ì¥
                ali_collect_status["connected"] = True
                ali_collect_status["logs"].append(f"[{datetime.now().strftime('%H:%M:%S')}] ê¸°ì¡´ Chrome ì—°ê²° ì„±ê³µ (í¬íŠ¸: {existing_port})")
                
                return {"success": True, "message": f"ê¸°ì¡´ Chrome ì—°ê²° ì„±ê³µ (í¬íŠ¸: {existing_port})"}
            except Exception as e:
                print(f"[ì•Œë¦¬] ê¸°ì¡´ Chrome ì—°ê²° ì‹¤íŒ¨: {e}")
        
        # 2. ìƒˆë¡œìš´ Chrome ì‹¤í–‰
        available_port = find_available_port(9300, 9320)
        
        # Chrome ê²½ë¡œ ì°¾ê¸°
        chrome_paths = [
            r"C:\Program Files\Google\Chrome\Application\chrome.exe",
            r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
            os.path.expandvars(r"%LOCALAPPDATA%\Google\Chrome\Application\chrome.exe"),
        ]
        
        chrome_path = None
        for path in chrome_paths:
            if os.path.exists(path):
                chrome_path = path
                break
        
        if not chrome_path:
            chrome_path = shutil.which("chrome") or shutil.which("google-chrome")
        
        if not chrome_path:
            return {"success": False, "message": "Chromeì„ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤"}
        
        # ì•Œë¦¬ ì „ìš© í”„ë¡œí•„ ë””ë ‰í† ë¦¬
        ali_profile_dir = os.path.join(APP_DIR, "chrome_ali_profile")
        
        # Chrome ë””ë²„ê·¸ ëª¨ë“œë¡œ ì‹¤í–‰
        cmd = [
            chrome_path,
            f"--remote-debugging-port={available_port}",
            f"--user-data-dir={ali_profile_dir}",
            "--no-first-run",
            "--no-default-browser-check",
            "https://www.aliexpress.com"
        ]
        
        print(f"[ì•Œë¦¬] Chrome ì‹¤í–‰ (í¬íŠ¸ {available_port}): {' '.join(cmd)}")
        subprocess.Popen(cmd)
        
        # Chrome ì‹œì‘ ëŒ€ê¸°
        await asyncio.sleep(4)
        
        # Playwrightë¡œ ì—°ê²° (ì¬ì‹œë„)
        ali_playwright = await async_playwright().start()
        
        for retry in range(5):
            try:
                ali_browser = await ali_playwright.chromium.connect_over_cdp(f"http://localhost:{available_port}")
                break
            except Exception as e:
                if retry < 4:
                    print(f"[ì•Œë¦¬] ì—°ê²° ëŒ€ê¸°ì¤‘... ({retry+1}/5)")
                    await asyncio.sleep(2)
                else:
                    raise e
        
        ali_collect_status["connected"] = True
        ali_collect_status["logs"].append(f"[{datetime.now().strftime('%H:%M:%S')}] ë¸Œë¼ìš°ì € ì—°ê²° ì„±ê³µ (í¬íŠ¸: {available_port})")
        ali_debug_port = available_port  # í¬íŠ¸ ì €ì¥
        
        return {"success": True, "message": f"ë¸Œë¼ìš°ì € ì‹¤í–‰ ë° ì—°ê²° ì„±ê³µ (í¬íŠ¸: {available_port}). ì•Œë¦¬ìµìŠ¤í”„ë ˆìŠ¤ ë¡œê·¸ì¸ í›„ 'ìˆ˜ì§‘ ì‹œì‘' í´ë¦­"}
        
    except Exception as e:
        ali_collect_status["logs"].append(f"[{datetime.now().strftime('%H:%M:%S')}] ì—°ê²° ì‹¤íŒ¨: {e}")
        return {"success": False, "message": f"ì—°ê²° ì‹¤íŒ¨: {e}"}

class AliCollectRequest(BaseModel):
    sheet_url: str
    month: str

@app.post("/api/tools/ali/collect")
async def ali_collect(request: Request, req: AliCollectRequest):
    require_permission(request, "edit")

    global ali_collect_status, ali_browser

    if not ali_browser or not ali_collect_status.get("connected"):
        return {"success": False, "message": "ë¸Œë¼ìš°ì €ê°€ ì—°ê²°ë˜ì§€ ì•Šì•˜ìŠµë‹ˆë‹¤. 'ë¸Œë¼ìš°ì € ì—°ê²°'ì„ ë¨¼ì € í´ë¦­í•˜ì„¸ìš”."}

    # ì´ë¯¸ ì‹¤í–‰ ì¤‘ì´ë©´ ê±°ë¶€
    if ali_collect_status["running"]:
        return {"success": False, "message": "ì´ë¯¸ ìˆ˜ì§‘ ì¤‘ì…ë‹ˆë‹¤."}

    # ì‘ì—… ë¡œê·¸ ê¸°ë¡
    log_work("ì•Œë¦¬ì†¡ì¥ìˆ˜ì§‘", "ì•Œë¦¬ìµìŠ¤í”„ë ˆìŠ¤", 0, f"ì›”: {req.month}", "ì›¹")
    
    # ì‹œíŠ¸ ID ì¶”ì¶œ
    import re
    sheet_id = req.sheet_url
    if 'docs.google.com' in sheet_id:
        match = re.search(r'/d/([a-zA-Z0-9_-]+)', sheet_id)
        if match:
            sheet_id = match.group(1)
    
    # ìƒíƒœ ì´ˆê¸°í™”
    ali_collect_status = {
        "running": True,
        "progress": 0,
        "total": 0,
        "completed": 0,
        "logs": [],
        "message": "ì‹œì‘ ì¤‘...",
        "collected": []  # ìˆ˜ì§‘ëœ ë°ì´í„°
    }
    
    def add_ali_log(msg, status="info"):
        ali_collect_status["logs"].append({
            "time": datetime.now().strftime("%H:%M:%S"),
            "msg": msg,
            "status": status
        })
        if len(ali_collect_status["logs"]) > 100:
            ali_collect_status["logs"] = ali_collect_status["logs"][-100:]
        print(f"[ì•Œë¦¬] {msg}")
    
    def run_collection():
        global ali_collect_status
        try:
            add_ali_log(f"ì‹œíŠ¸ ID: {sheet_id}, ì›”: {req.month}")
            
            # ì•Œë¦¬ ìˆ˜ì§‘ìš© ë³„ë„ JSON íŒŒì¼ (APP_DIR ê¸°ì¤€)
            ALI_CREDENTIALS = str(APP_DIR / "autosms-466614-951e91617c69.json")
            add_ali_log(f"ì¸ì¦ íŒŒì¼: {ALI_CREDENTIALS}")
            add_ali_log(f"ì¸ì¦ íŒŒì¼ ì¡´ì¬: {os.path.exists(ALI_CREDENTIALS)}")
            
            if not os.path.exists(ALI_CREDENTIALS):
                add_ali_log(f"ì¸ì¦ íŒŒì¼ ì—†ìŒ: {ALI_CREDENTIALS}", "error")
                ali_collect_status["running"] = False
                return
            
            # ì‹œíŠ¸ ì—°ê²°
            from google.oauth2.service_account import Credentials
            scopes = [
                'https://www.googleapis.com/auth/spreadsheets',
                'https://www.googleapis.com/auth/drive'
            ]
            creds = Credentials.from_service_account_file(ALI_CREDENTIALS, scopes=scopes)
            ali_client = gspread.authorize(creds)
            
            spreadsheet = ali_client.open_by_key(sheet_id)
            sheet = spreadsheet.worksheet(req.month)
            add_ali_log(f"ì‹œíŠ¸ ì—°ê²° ì„±ê³µ: {spreadsheet.title} / {req.month}")
            
            # ì—´ ìœ„ì¹˜
            COL_PLATFORM = 30  # AD
            COL_ORDER_ID = 31  # AE
            COL_CARRIER = 43   # AQ
            COL_TRACKING = 44  # AR
            START_ROW = 3
            
            all_data = sheet.get_all_values()
            add_ali_log(f"ì „ì²´ ë°ì´í„° í–‰ ìˆ˜: {len(all_data)}")
            
            # ëŒ€ìƒ ì°¾ê¸°
            targets = []
            for idx, row in enumerate(all_data):
                row_num = idx + 1
                if row_num < START_ROW:
                    continue
                
                platform = row[COL_PLATFORM - 1] if len(row) >= COL_PLATFORM else ""
                order_id = row[COL_ORDER_ID - 1] if len(row) >= COL_ORDER_ID else ""
                tracking = row[COL_TRACKING - 1] if len(row) >= COL_TRACKING else ""
                
                if platform.strip() == "ì•Œë¦¬" and order_id.strip() and not tracking.strip():
                    targets.append({'row': row_num, 'order_id': order_id.strip()})
            
            add_ali_log(f"ì¡°íšŒ ëŒ€ìƒ: {len(targets)}ê±´")
            ali_collect_status["total"] = len(targets)
            
            if not targets:
                add_ali_log("ì¡°íšŒí•  ê±´ì´ ì—†ìŠµë‹ˆë‹¤", "success")
                ali_collect_status["running"] = False
                return
            
            # ìˆ˜ì§‘ - sync_playwright ì‚¬ìš© (ìŠ¤ë ˆë“œì—ì„œ ë™ê¸° ë°©ì‹ìœ¼ë¡œ)
            updated = 0
            
            # íƒë°°ì‚¬ êµ¬ë¶„ í•¨ìˆ˜
            def get_carrier_name(tracking_no):
                tracking_no = str(tracking_no).strip()
                # CJëŒ€í•œí†µìš´: 30, 50, 52, 56
                if tracking_no.startswith('30') or tracking_no.startswith('50') or tracking_no.startswith('52') or tracking_no.startswith('56'):
                    return 'CJëŒ€í•œí†µìš´'
                # í•œì§„íƒë°°: 51, 55, 58
                elif tracking_no.startswith('51') or tracking_no.startswith('55') or tracking_no.startswith('58'):
                    return 'í•œì§„íƒë°°'
                # ê²½ë™íƒë°°: 68
                elif tracking_no.startswith('68'):
                    return 'ê²½ë™'
                # ë¡œì  íƒë°°: 54
                elif tracking_no.startswith('54'):
                    return 'ë¡œì  '
                # íˆ¬ë°ì´: 9
                elif tracking_no.startswith('9'):
                    return 'íˆ¬ë°ì´'
                return 'í™•ì¸í•„ìš”'
            
            # sync_playwrightë¡œ Chromeì— ì—°ê²°
            from playwright.sync_api import sync_playwright
            
            if not ali_debug_port:
                add_ali_log("Chrome ë””ë²„ê·¸ í¬íŠ¸ë¥¼ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤. ë¸Œë¼ìš°ì €ë¥¼ ë‹¤ì‹œ ì—°ê²°í•˜ì„¸ìš”.", "error")
                ali_collect_status["running"] = False
                return
            
            add_ali_log(f"Chrome ì—°ê²° ì¤‘ (í¬íŠ¸: {ali_debug_port})...")
            
            with sync_playwright() as p:
                try:
                    sync_browser = p.chromium.connect_over_cdp(f"http://localhost:{ali_debug_port}")
                    add_ali_log("sync_playwright ì—°ê²° ì„±ê³µ")
                except Exception as e:
                    add_ali_log(f"Chrome ì—°ê²° ì‹¤íŒ¨: {e}", "error")
                    ali_collect_status["running"] = False
                    return
                
                contexts = sync_browser.contexts
                if not contexts:
                    add_ali_log("ë¸Œë¼ìš°ì € ì»¨í…ìŠ¤íŠ¸ê°€ ì—†ìŠµë‹ˆë‹¤", "error")
                    ali_collect_status["running"] = False
                    return
                
                context = contexts[0]
                pages = context.pages
                if not pages:
                    add_ali_log("ì—´ë¦° í˜ì´ì§€ê°€ ì—†ìŠµë‹ˆë‹¤", "error")
                    ali_collect_status["running"] = False
                    return
                
                ali_page = pages[0]
                add_ali_log(f"í˜ì´ì§€ ì—°ê²° ì™„ë£Œ: {ali_page.url[:50]}...")
                
                for i, target in enumerate(targets):
                    if not ali_collect_status["running"]:
                        add_ali_log("ì‚¬ìš©ìì— ì˜í•´ ì¤‘ë‹¨ë¨", "info")
                        break
                    
                    ali_collect_status["completed"] = i
                    ali_collect_status["progress"] = int((i / len(targets)) * 100)
                    
                    add_ali_log(f"[{i+1}/{len(targets)}] ì£¼ë¬¸ë²ˆí˜¸ {target['order_id']} ì¡°íšŒ ì¤‘...")
                    
                    # ì†¡ì¥ë²ˆí˜¸ ì¡°íšŒ
                    tracking_no = None
                    try:
                        url = f"https://www.aliexpress.com/p/tracking/index.html?_addShare=no&_login=yes&tradeOrderId={target['order_id']}"
                        
                        ali_page.goto(url, timeout=20000, wait_until="domcontentloaded")
                        time.sleep(2)
                        
                        page_text = ali_page.inner_text("body", timeout=10000)
                        
                        # ì†¡ì¥ë²ˆí˜¸ ì¶”ì¶œ
                        import re
                        match = re.search(r'ìš´ì†¡ì¥\s*ë²ˆí˜¸[:\s]*(\d+)', page_text)
                        if match:
                            tracking_no = match.group(1)
                        else:
                            match = re.search(r'Tracking\s*(?:number|no)[:\s]*(\d+)', page_text, re.IGNORECASE)
                            if match:
                                tracking_no = match.group(1)
                            else:
                                # êµ­ë‚´ ë°°ì†¡ ì†¡ì¥ë²ˆí˜¸ íŒ¨í„´
                                match = re.search(r'(\d{10,14})', page_text)
                                if match:
                                    potential = match.group(1)
                                    if potential[:2] in ['50', '51', '52', '54', '56'] or potential.startswith('9'):
                                        tracking_no = potential
                    except Exception as e:
                        add_ali_log(f"ì¡°íšŒ ì˜¤ë¥˜: {e}", "error")
                    
                    if tracking_no:
                        carrier = get_carrier_name(tracking_no)
                        
                        # íˆ¬ë°ì´ íƒë°°ì¸ ê²½ìš° ë°°ì†¡ì™„ë£Œ ì²´í¬
                        delivery_completed = False
                        if carrier == 'íˆ¬ë°ì´':
                            try:
                                # ë°°ì†¡ì™„ë£Œ í…ìŠ¤íŠ¸ í™•ì¸
                                completed_el = ali_page.locator('.logistic-info-v2--nodeTitle--2rejjVx:has-text("ë°°ì†¡ ì™„ë£Œ"), .logistic-info-v2--nodeTitle--2rejjVx:has-text("ë°°ì†¡ì™„ë£Œ")')
                                if completed_el.count() > 0:
                                    delivery_completed = True
                                    add_ali_log(f"â†’ ì†¡ì¥ë²ˆí˜¸: {tracking_no} ({carrier}) [ë°°ì†¡ì™„ë£Œ]", "success")
                                else:
                                    add_ali_log(f"â†’ ì†¡ì¥ë²ˆí˜¸: {tracking_no} ({carrier})", "success")
                            except:
                                add_ali_log(f"â†’ ì†¡ì¥ë²ˆí˜¸: {tracking_no} ({carrier})", "success")
                        else:
                            add_ali_log(f"â†’ ì†¡ì¥ë²ˆí˜¸: {tracking_no} ({carrier})", "success")
                        
                        # ìˆ˜ì§‘ ë°ì´í„° ì €ì¥
                        ali_collect_status["collected"].append({
                            "customer_order": target.get('customer_order', ''),
                            "order_id": target['order_id'],
                            "carrier": carrier,
                            "tracking_no": tracking_no,
                            "delivery_completed": delivery_completed
                        })
                        
                        try:
                            sheet.update_cell(target['row'], COL_CARRIER, carrier)
                            sheet.update_cell(target['row'], COL_TRACKING, tracking_no)
                            updated += 1
                            add_ali_log(f"â†’ í–‰ {target['row']} ì—…ë°ì´íŠ¸ ì™„ë£Œ")
                            
                            # íˆ¬ë°ì´ ë°°ì†¡ì™„ë£Œë©´ ë…¸ë€ìƒ‰ ìƒ‰ì¹ 
                            if delivery_completed:
                                try:
                                    # AQ, AR ì»¬ëŸ¼ (43, 44)
                                    ws.format(f"AQ{target['row']}:AR{target['row']}", {
                                        "backgroundColor": {"red": 1, "green": 1, "blue": 0}
                                    })
                                    add_ali_log(f"â†’ í–‰ {target['row']} ë…¸ë€ìƒ‰ ìƒ‰ì¹  (ë°°ì†¡ì™„ë£Œ)")
                                except Exception as e:
                                    add_ali_log(f"â†’ ìƒ‰ì¹  ì˜¤ë¥˜: {e}", "error")
                            
                            time.sleep(1)
                        except Exception as e:
                            add_ali_log(f"â†’ ì‹œíŠ¸ ì—…ë°ì´íŠ¸ ì˜¤ë¥˜: {e}", "error")
                    else:
                        add_ali_log("â†’ ì†¡ì¥ë²ˆí˜¸ ì—†ìŒ")
                    
                    time.sleep(2)
                
                # for ë£¨í”„ ë (with ë¸”ë¡ ì•ˆ)
                add_ali_log(f"ì™„ë£Œ! {updated}ê±´ ì—…ë°ì´íŠ¸", "success")
                ali_collect_status["completed"] = len(targets)
                ali_collect_status["progress"] = 100
            
        except Exception as e:
            import traceback
            add_ali_log(f"ì˜¤ë¥˜: {e}", "error")
            traceback.print_exc()
        finally:
            ali_collect_status["running"] = False
    
    # ë°±ê·¸ë¼ìš´ë“œ ìŠ¤ë ˆë“œë¡œ ì‹¤í–‰
    import threading
    thread = threading.Thread(target=run_collection, daemon=True)
    thread.start()
    
    return {"success": True, "message": "ìˆ˜ì§‘ ì‹œì‘ë¨"}

@app.get("/api/tools/ali/progress")
async def ali_progress(request: Request):
    """ì•Œë¦¬ ìˆ˜ì§‘ ì§„í–‰ìƒí™© ì¡°íšŒ"""
    require_permission(request, "edit")
    return ali_collect_status

@app.get("/api/tools/ali/progress-stream")
async def ali_progress_stream(request: Request):
    """ì•Œë¦¬ ìˆ˜ì§‘ ì§„í–‰ìƒí™© SSE ìŠ¤íŠ¸ë¦¼"""
    from fastapi.responses import StreamingResponse

    async def event_generator():
        last_log_count = 0
        last_completed = 0
        while True:
            current_log_count = len(ali_collect_status.get("logs", []))
            current_completed = ali_collect_status.get("completed", 0)

            # ë³€ê²½ì‚¬í•­ì´ ìˆì„ ë•Œë§Œ ì „ì†¡
            if current_log_count != last_log_count or current_completed != last_completed or not ali_collect_status.get("running"):
                yield f"data: {json.dumps(ali_collect_status, ensure_ascii=False, default=str)}\n\n"
                last_log_count = current_log_count
                last_completed = current_completed

            # ìˆ˜ì§‘ ì™„ë£Œ ë˜ëŠ” ì¤‘ì§€ (running=Falseë©´ ì¢…ë£Œ)
            if ali_collect_status.get("running") == False:
                break

            await asyncio.sleep(0.5)

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "Connection": "keep-alive"}
    )

@app.post("/api/tools/ali/stop")
async def ali_stop(request: Request):
    require_permission(request, "edit")
    global ali_collect_status
    ali_collect_status["running"] = False
    return {"success": True}

@app.get("/api/tools/ali/download")
async def ali_download_excel(request: Request):
    """ìˆ˜ì§‘ëœ ì•Œë¦¬ ì†¡ì¥ ì—‘ì…€ ë‹¤ìš´ë¡œë“œ"""
    require_permission(request, "edit")
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from fastapi.responses import FileResponse
    import tempfile
    
    # ìˆ˜ì§‘ëœ ë°ì´í„° ê°€ì ¸ì˜¤ê¸°
    collected = ali_collect_status.get("collected", [])
    
    if not collected:
        raise HTTPException(status_code=400, detail="ìˆ˜ì§‘ëœ ë°ì´í„°ê°€ ì—†ìŠµë‹ˆë‹¤")
    
    # ì—‘ì…€ ìƒì„±
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # í—¤ë” (ì–‘ì‹ì— ë§ì¶¤)
    headers = ['ê³ ê° ì£¼ë¬¸ì¼', 'ê³ ê° ì£¼ë¬¸ ë²ˆí˜¸', 'í•´ì™¸ ì£¼ë¬¸ì¼', 'í•´ì™¸ ì£¼ë¬¸ ë²ˆí˜¸', 
               'í•´ì™¸ íƒë°°ì‚¬', 'í•´ì™¸ ìš´ì†¡ì¥ë²ˆí˜¸', 'êµ­ë‚´ íƒë°°ì‚¬', 'êµ­ë‚´ ìš´ì†¡ì¥ë²ˆí˜¸']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # ë°ì´í„° ì…ë ¥
    for row_idx, item in enumerate(collected, 2):
        ws.cell(row=row_idx, column=2, value=item.get("customer_order", ""))  # ê³ ê° ì£¼ë¬¸ ë²ˆí˜¸
        ws.cell(row=row_idx, column=4, value=item.get("order_id", ""))  # í•´ì™¸ ì£¼ë¬¸ ë²ˆí˜¸ (ì•Œë¦¬ ì£¼ë¬¸ë²ˆí˜¸)
        ws.cell(row=row_idx, column=7, value=item.get("carrier", ""))  # êµ­ë‚´ íƒë°°ì‚¬
        ws.cell(row=row_idx, column=8, value=item.get("tracking_no", ""))  # êµ­ë‚´ ìš´ì†¡ì¥ë²ˆí˜¸
    
    # ì»¬ëŸ¼ ë„ˆë¹„ ì¡°ì •
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 20
    
    # ì„ì‹œ íŒŒì¼ë¡œ ì €ì¥
    today = datetime.now().strftime('%y%m%d')
    filename = f"ì†¡ì¥_ë²ˆí˜¸_ë‹¤ìš´ë¡œë“œ_{today}.xlsx"
    filepath = Path(tempfile.gettempdir()) / filename
    wb.save(filepath)
    
    return FileResponse(
        path=filepath,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ========== í´ë¼ì´ì–¸íŠ¸ í”„ë¡œê·¸ë¨ ì—°ë™ API ==========
client_status = {"connected": False, "last_ping": None}

@app.get("/api/client/status")
async def get_client_status(request: Request):
    """í´ë¼ì´ì–¸íŠ¸ ì—°ê²° ìƒíƒœ í™•ì¸"""
    # ë§ˆì§€ë§‰ pingì´ 10ì´ˆ ì´ë‚´ë©´ ì—°ê²°ëœ ê²ƒìœ¼ë¡œ íŒë‹¨
    if client_status["last_ping"]:
        if datetime.now() - client_status["last_ping"] < timedelta(seconds=10):
            return {"connected": True}
    return {"connected": False}

@app.post("/api/client/ping")
async def client_ping(request: Request):
    """í´ë¼ì´ì–¸íŠ¸ì—ì„œ ì£¼ê¸°ì ìœ¼ë¡œ í˜¸ì¶œ (ì—°ê²° ìœ ì§€)"""
    client_status["connected"] = True
    client_status["last_ping"] = datetime.now()
    return {"success": True}

@app.post("/api/auto-login/complete")
async def complete_login(request: Request):
    """ë¡œê·¸ì¸ ì™„ë£Œ ì•Œë¦¼"""
    api_key = request.headers.get("X-API-Key")
    if api_key != "pkonomiautokey2024":
        raise HTTPException(status_code=401, detail="API í‚¤ í•„ìš”")
    data = await request.json()
    await ws_manager.broadcast({
        "type": "login_complete",
        "platform": data.get("platform"),
        "login_id": data.get("login_id"),
        "success": data.get("success", False)
    })
    return {"success": True}

@app.get("/download/PkonomyClient.exe")
async def download_client_legacy():
    """í´ë¼ì´ì–¸íŠ¸ í”„ë¡œê·¸ë¨ ë‹¤ìš´ë¡œë“œ (ë ˆê±°ì‹œ ê²½ë¡œ)"""
    from fastapi.responses import FileResponse
    # dist í´ë” ìš°ì„ 
    client_path = APP_DIR / "dist" / "PkonomyClient.exe"
    if not client_path.exists():
        client_path = APP_DIR / "PkonomyClient.exe"
    if not client_path.exists():
        client_path = Path(r"C:\autosystem\PkonomyClient.exe")
    if client_path.exists():
        return FileResponse(path=str(client_path), filename="PkonomyClient.exe", media_type="application/octet-stream")
    raise HTTPException(status_code=404, detail="í´ë¼ì´ì–¸íŠ¸ íŒŒì¼ ì—†ìŒ. ê´€ë¦¬ìì—ê²Œ ë¬¸ì˜í•˜ì„¸ìš”.")


# ========== 11ë²ˆê°€ API ìƒí’ˆìˆ˜ ì¡°íšŒ ==========
import xml.etree.ElementTree as ET

ST_API_BASE = "http://api.11st.co.kr"
ST_SEARCH_PATH = "/rest/prodmarketservice/prodmarket"


async def run_11st_product_count_task(log_file: str, platform: str = "11ë²ˆê°€"):
    """11ë²ˆê°€ íŒë§¤ì¤‘ ì¡°íšŒ ë°±ê·¸ë¼ìš´ë“œ ì‘ì—…"""
    status = get_aio_status(platform)
    
    def write_log(msg):
        timestamp = datetime.now().strftime('%H:%M:%S')
        with open(log_file, "a", encoding="utf-8") as f:
            f.write(f"[{timestamp}] {msg}\n")
    
    try:
        # 11ë²ˆê°€ ì‹œíŠ¸ì—ì„œ activeì¸ ìŠ¤í† ì–´ ì¡°íšŒ
        ws = gsheet.sheet.worksheet("11ë²ˆê°€")
        all_values = ws.get_all_values()
        
        if not all_values or len(all_values) < 2:
            write_log("11ë²ˆê°€ ì‹œíŠ¸ê°€ ë¹„ì–´ìˆìŠµë‹ˆë‹¤")
            status["running"] = False
            status["status"] = "completed"
            return
        
        headers = all_values[0]
        
        # ì»¬ëŸ¼ ì¸ë±ìŠ¤ ì°¾ê¸°
        store_col = None
        active_col = None
        api_key_col = None
        on_sale_col = None
        last_reg_col = None
        result_col = None
        updated_col = None

        for i, h in enumerate(headers):
            if h in ["store_name", "ì‡¼í•‘ëª° ë³„ì¹­", "ìŠ¤í† ì–´ëª…"]:
                store_col = i
            if h in ["active", "í™œì„±", "ì‚¬ìš©"]:
                active_col = i
            if h in ["api_key", "API KEY", "11ë²ˆê°€ API KEY"]:
                api_key_col = i
            if h in ["on_sale", "íŒë§¤ì¤‘"]:
                on_sale_col = i
            if h in ["ë§ˆì§€ë§‰ë“±ë¡ì¼"]:
                last_reg_col = i
            if h in ["ê²°ê³¼"]:
                result_col = i
            if h in ["updated_at"]:
                updated_col = i

        # í•„ìˆ˜ í—¤ë”ê°€ ì—†ìœ¼ë©´ ì¶”ê°€
        headers_modified = False
        if on_sale_col is None:
            headers.append("íŒë§¤ì¤‘")
            on_sale_col = len(headers) - 1
            headers_modified = True
        if last_reg_col is None:
            headers.append("ë§ˆì§€ë§‰ë“±ë¡ì¼")
            last_reg_col = len(headers) - 1
            headers_modified = True
        if updated_col is None:
            headers.append("updated_at")
            updated_col = len(headers) - 1
            headers_modified = True

        if headers_modified:
            ws.update(range_name="1:1", values=[headers], value_input_option="RAW")
            write_log(f"í—¤ë” ì¶”ê°€ë¨: íŒë§¤ì¤‘={on_sale_col}, ë§ˆì§€ë§‰ë“±ë¡ì¼={last_reg_col}, updated_at={updated_col}")

        if store_col is None or api_key_col is None:
            write_log(f"í•„ìˆ˜ ì—´ì„ ì°¾ì„ ìˆ˜ ì—†ìŒ: store_col={store_col}, api_key_col={api_key_col}")
            status["running"] = False
            status["status"] = "completed"
            return
        
        # activeì¸ ìŠ¤í† ì–´ í•„í„°ë§
        active_stores = []
        for row_idx, row in enumerate(all_values[1:], start=2):
            if len(row) <= max(store_col, active_col or 0, api_key_col):
                continue
            
            store_name = row[store_col].strip() if len(row) > store_col else ""
            active = row[active_col].strip() if active_col and len(row) > active_col else "TRUE"
            api_key = row[api_key_col].strip() if len(row) > api_key_col else ""
            
            if not store_name or not api_key:
                continue
            
            is_active = str(active).upper() in ["TRUE", "ON", "Y", "1", "ì‚¬ìš©", ""]
            if is_active:
                active_stores.append({
                    "row": row_idx,
                    "ìŠ¤í† ì–´ëª…": store_name,
                    "api_key": api_key
                })
        
        write_log(f"ì¡°íšŒ ëŒ€ìƒ: {len(active_stores)}ê°œ ìŠ¤í† ì–´")
        status["total"] = len(active_stores)
        
        # ë³‘ë ¬ë¡œ íŒë§¤ì¤‘ ìˆ˜ëŸ‰ ì¡°íšŒ (5ê°œì”© ë™ì‹œ ì²˜ë¦¬)
        total_done = 0
        batch_size = 5

        def get_col_letter(idx):
            if idx < 26:
                return chr(65 + idx)
            return f"{chr(64 + idx // 26)}{chr(65 + idx % 26)}"

        for batch_start in range(0, len(active_stores), batch_size):
            if not status["running"]:
                write_log("ì‘ì—… ì¤‘ì§€ë¨")
                break

            batch = active_stores[batch_start:batch_start + batch_size]
            batch_names = [s["ìŠ¤í† ì–´ëª…"] for s in batch]

            status["current_store"] = ", ".join(batch_names)
            status["current_action"] = "íŒë§¤ì¤‘ ì¡°íšŒ ì¤‘..."
            status["completed"] = batch_start
            status["progress"] = int((batch_start / len(active_stores)) * 100)

            write_log(f"ë°°ì¹˜ ì¡°íšŒ: {batch_names}")

            # ë°°ì¹˜ ë‚´ ë³‘ë ¬ ì¡°íšŒ
            async def fetch_store_count(store):
                try:
                    count, last_reg = await get_11st_product_count_and_last_reg(store["api_key"])
                    return {
                        "row": store["row"],
                        "ìŠ¤í† ì–´ëª…": store["ìŠ¤í† ì–´ëª…"],
                        "count": count,
                        "last_reg": last_reg
                    }
                except Exception as e:
                    return {
                        "row": store["row"],
                        "ìŠ¤í† ì–´ëª…": store["ìŠ¤í† ì–´ëª…"],
                        "count": 0,
                        "last_reg": ""
                    }

            batch_results = await asyncio.gather(*[fetch_store_count(s) for s in batch])

            # ë°°ì¹˜ë³„ë¡œ ë°”ë¡œ ì‹œíŠ¸ì— ê¸°ë¡
            updates = []
            now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            for r in batch_results:
                write_log(f"[{r['ìŠ¤í† ì–´ëª…']}] íŒë§¤ì¤‘: {r['count']}ê°œ, ë§ˆì§€ë§‰ë“±ë¡: {r['last_reg'] or '-'}")

                # íŒë§¤ì¤‘ ìˆ˜ëŸ‰
                if on_sale_col is not None:
                    cell = f"{get_col_letter(on_sale_col)}{r['row']}"
                    updates.append({"range": cell, "values": [[r["count"]]]})

                # ë§ˆì§€ë§‰ë“±ë¡ì¼
                if last_reg_col is not None and r.get("last_reg"):
                    cell = f"{get_col_letter(last_reg_col)}{r['row']}"
                    updates.append({"range": cell, "values": [[r["last_reg"]]]})

                # ê²°ê³¼
                if result_col is not None:
                    cell = f"{get_col_letter(result_col)}{r['row']}"
                    result_text = f"íŒë§¤ì¤‘ {r['count']}ê°œ"
                    updates.append({"range": cell, "values": [[result_text]]})

                # ë‚ ì§œ
                if updated_col is not None:
                    cell = f"{get_col_letter(updated_col)}{r['row']}"
                    updates.append({"range": cell, "values": [[now_str]]})

            if updates:
                ws.batch_update(updates)
                write_log(f"ë°°ì¹˜ {len(batch_results)}ê°œ ì‹œíŠ¸ ì €ì¥ ì™„ë£Œ")

            total_done += len(batch_results)
            await asyncio.sleep(0.3)  # ë°°ì¹˜ ê°„ ê°„ê²©

        write_log(f"ì™„ë£Œ: ì´ {total_done}ê°œ ìŠ¤í† ì–´ ì¡°íšŒ")
        
        status["running"] = False
        status["status"] = "completed"
        status["progress"] = 100
        status["current_store"] = ""
        status["current_action"] = "ì™„ë£Œ"
        
    except Exception as e:
        write_log(f"ì˜¤ë¥˜ ë°œìƒ: {e}")
        import traceback
        traceback.print_exc()
        status["running"] = False
        status["status"] = "error"


async def get_11st_product_count_and_last_reg(api_key: str) -> tuple:
    """11ë²ˆê°€ íŒë§¤ì¤‘ ìƒí’ˆìˆ˜ + ìµœì‹  ë“±ë¡ì¼ ì¡°íšŒ - ë³‘ë ¬ í˜ì´ì§•"""
    if not api_key:
        return 0, ""

    import re

    def fetch_page(page: int) -> tuple:
        """í•œ í˜ì´ì§€ ìƒí’ˆë²ˆí˜¸ + ë‚ ì§œ ì¡°íšŒ"""
        limit = 500
        parts = ["<SearchProduct>"]
        parts.append("    <selStatCd>103</selStatCd>")
        parts.append(f"    <limit>{limit}</limit>")
        if page > 0:
            start = page * limit + 1
            parts.append(f"    <start>{start}</start>")
        parts.append("</SearchProduct>")
        xml_body = "\n".join(parts)

        headers = {
            "openapikey": api_key,
            "Content-Type": "text/xml;charset=euc-kr",
            "Accept": "application/xml",
        }

        try:
            data = xml_body.encode("euc-kr", errors="ignore")
            res = requests.post(
                f"{ST_API_BASE}{ST_SEARCH_PATH}",
                headers=headers,
                data=data,
                timeout=30
            )

            if res.status_code != 200:
                return [], []

            raw = res.content.decode("euc-kr", errors="ignore")

            # prdNo + aplBgnDy íŒŒì‹±
            prd_list = []
            date_list = []
            try:
                root = ET.fromstring(raw)
                for prod in root.iter():
                    if prod.tag.endswith("product"):
                        prd_no = ""
                        apl_bgn = ""
                        for child in prod:
                            if child.tag.endswith("prdNo"):
                                prd_no = (child.text or "").strip()
                            elif "aplBgnDy" in child.tag:
                                apl_bgn = (child.text or "").strip()[:10]
                        if prd_no:
                            prd_list.append(prd_no)
                            if apl_bgn:
                                date_list.append(apl_bgn)
            except Exception as e:
                print(f"[11ë²ˆê°€] XML íŒŒì‹± ì˜¤ë¥˜: {e}")

            return prd_list, date_list
        except:
            return [], []

    def fetch_all():
        from concurrent.futures import ThreadPoolExecutor, as_completed

        all_prd = set()
        all_dates = []
        limit = 500

        # ì²« í˜ì´ì§€ ì¡°íšŒ
        first_page, first_dates = fetch_page(0)
        if not first_page:
            return 0, ""

        for p in first_page:
            all_prd.add(p)
        all_dates.extend(first_dates)

        print(f"[11ë²ˆê°€] ì²« í˜ì´ì§€: {len(first_page)}ê°œ")

        # ì²« í˜ì´ì§€ê°€ ê°€ë“ ì°¼ìœ¼ë©´ ì¶”ê°€ í˜ì´ì§€ ë³‘ë ¬ ì¡°íšŒ
        if len(first_page) >= limit:
            max_pages = 20
            with ThreadPoolExecutor(max_workers=5) as executor:
                futures = {executor.submit(fetch_page, p): p for p in range(1, max_pages)}

                for future in as_completed(futures):
                    page_result, page_dates = future.result()
                    if not page_result:
                        continue

                    for p in page_result:
                        all_prd.add(p)
                    all_dates.extend(page_dates)

        # ê°€ì¥ ìµœì‹  ë‚ ì§œ ì°¾ê¸°
        latest_date = ""
        print(f"[11ë²ˆê°€] ìˆ˜ì§‘ëœ ë‚ ì§œ ìˆ˜: {len(all_dates)}ê°œ")
        if all_dates:
            print(f"[11ë²ˆê°€] ë‚ ì§œ ìƒ˜í”Œ: {all_dates[:5]}")
            # ë‚ ì§œ í˜•ì‹ í†µì¼ í›„ ì •ë ¬
            normalized_dates = []
            for d in all_dates:
                if len(d) == 8 and d.isdigit():
                    # YYYYMMDD -> YYYY-MM-DD
                    normalized_dates.append(f"{d[:4]}-{d[4:6]}-{d[6:8]}")
                elif "-" in d:
                    normalized_dates.append(d[:10])
                elif "/" in d:
                    # YYYY/MM/DD í˜•ì‹
                    normalized_dates.append(d[:10].replace("/", "-"))
            if normalized_dates:
                normalized_dates.sort(reverse=True)
                latest_date = normalized_dates[0]
                print(f"[11ë²ˆê°€] ìµœì‹  ë‚ ì§œ: {latest_date}")

        print(f"[11ë²ˆê°€] ì´ ìƒí’ˆìˆ˜: {len(all_prd)}ê°œ, ìµœì‹ ë“±ë¡ì¼: {latest_date}")
        return len(all_prd), latest_date

    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(None, fetch_all)


async def get_11st_product_count(api_key: str) -> int:
    """11ë²ˆê°€ íŒë§¤ì¤‘ ìƒí’ˆìˆ˜ ì¡°íšŒ (í•˜ìœ„í˜¸í™˜ìš©)"""
    count, _ = await get_11st_product_count_and_last_reg(api_key)
    return count


def get_11st_last_registration_date(api_key: str) -> str:
    """11ë²ˆê°€ ìµœì‹  ë“±ë¡ì¼ ì¡°íšŒ (ë™ê¸° ë²„ì „ - í•˜ìœ„í˜¸í™˜ìš©)"""
    import asyncio
    try:
        loop = asyncio.get_event_loop()
        _, last_reg = loop.run_until_complete(get_11st_product_count_and_last_reg(api_key))
        return last_reg
    except:
        return ""


# 11ë²ˆê°€ ìƒí’ˆìˆ˜ ìºì‹œ
st_product_cache: Dict[str, dict] = {}  # {login_id: {"count": 1234, "time": datetime}}
ST_CACHE_TTL = 300  # 5ë¶„ ìºì‹œ

@app.get("/api/11st/product-count/{login_id}")
async def get_11st_count(request: Request, login_id: str):
    """11ë²ˆê°€ ê°œë³„ ê³„ì • ìƒí’ˆìˆ˜ ì¡°íšŒ"""
    get_current_user(request)
    
    # ìºì‹œ í™•ì¸
    cached = st_product_cache.get(login_id)
    if cached:
        elapsed = (datetime.now() - cached["time"]).total_seconds()
        if elapsed < ST_CACHE_TTL:
            return {"success": True, "count": cached["count"], "cached": True}
    
    # ê³„ì • ì°¾ê¸°
    accounts = gsheet.get_accounts("11ë²ˆê°€")
    acc = None
    for a in accounts:
        if a.get("login_id") == login_id:
            acc = a
            break
    
    if not acc:
        return {"success": False, "error": "ê³„ì •ì„ ì°¾ì„ ìˆ˜ ì—†ìŒ"}
    
    api_key = acc.get("st_api_key", "")
    if not api_key:
        return {"success": False, "error": "API KEY ì—†ìŒ"}
    
    count = await get_11st_product_count(api_key)
    
    # ìºì‹œ ì €ì¥
    st_product_cache[login_id] = {"count": count, "time": datetime.now()}
    
    return {"success": True, "count": count, "cached": False}


@app.get("/api/market-summary")
async def get_market_summary(request: Request):
    """ë§ˆì¼“ë³„ í˜„í™© ìš”ì•½ (í‘œ í˜•ì‹ìš©)"""
    get_current_user(request)
    
    try:
        # 1. ë“±ë¡ê°¯ìˆ˜ ì‹œíŠ¸ì—ì„œ ìŠ¤ë§ˆíŠ¸ìŠ¤í† ì–´ ìƒí’ˆìˆ˜ ê°€ì ¸ì˜¤ê¸°
        ss_counts = {}
        try:
            ws_counts = gsheet.sheet.worksheet("ë“±ë¡ê°¯ìˆ˜")
            counts_data = ws_counts.get_all_values()
            if counts_data and len(counts_data) > 1:
                headers = counts_data[0]
                name_idx = None
                count_idx = None
                for i, h in enumerate(headers):
                    if h == "ìŠ¤í† ì–´ëª…":
                        name_idx = i
                    elif h == "íŒë§¤ì¤‘":
                        count_idx = i
                
                if name_idx is not None and count_idx is not None:
                    for row in counts_data[1:]:
                        if len(row) > max(name_idx, count_idx):
                            store = row[name_idx].strip()
                            try:
                                cnt = int(row[count_idx]) if row[count_idx] else 0
                            except:
                                cnt = 0
                            if store:
                                ss_counts[store] = cnt
        except Exception as e:
            print(f"[ë§ˆì¼“í˜„í™©] ë“±ë¡ê°¯ìˆ˜ ì‹œíŠ¸ ì˜¤ë¥˜: {e}")
        
        # 1-2. 11ë²ˆê°€ ì‹œíŠ¸ì—ì„œ ìƒí’ˆìˆ˜ ê°€ì ¸ì˜¤ê¸° (store_name -> íŒë§¤ì¤‘ ë§¤í•‘)
        st_counts = {}
        try:
            ws_11st = gsheet.sheet.worksheet("11ë²ˆê°€")
            st_data = ws_11st.get_all_values()
            if st_data and len(st_data) > 1:
                headers = st_data[0]
                name_idx = None
                count_idx = None
                for i, h in enumerate(headers):
                    if h in ["store_name", "ì‡¼í•‘ëª° ë³„ì¹­", "ìŠ¤í† ì–´ëª…"]:
                        name_idx = i
                    elif h in ["on_sale", "íŒë§¤ì¤‘"]:
                        count_idx = i
                
                if name_idx is not None and count_idx is not None:
                    for row in st_data[1:]:
                        if len(row) > max(name_idx, count_idx):
                            store = row[name_idx].strip()
                            try:
                                cnt = int(row[count_idx]) if row[count_idx] else 0
                            except:
                                cnt = 0
                            if store:
                                st_counts[store] = cnt
        except Exception as e:
            print(f"[ë§ˆì¼“í˜„í™©] 11ë²ˆê°€ ì‹œíŠ¸ ì˜¤ë¥˜: {e}")
        
        # 2. ë§ˆì¼“ìƒíƒœí˜„í™© ì‹œíŠ¸ì—ì„œ ìƒíƒœ/í˜ë„í‹° ì •ë³´ ê°€ì ¸ì˜¤ê¸°
        status_map = {}
        try:
            ws_status = gsheet.sheet.worksheet(MARKET_STATUS_TAB)
            status_records = ws_status.get_all_records()
            for row in status_records:
                store = row.get("ìŠ¤í† ì–´ëª…", "")
                plat = row.get("í”Œë«í¼", "")
                if store and plat:
                    status_map[f"{store}_{plat}"] = {
                        "status": row.get("ìƒíƒœ", "ì •ìƒ"),
                        "caution_count": int(row.get("ì£¼ì˜", 0) or 0),
                        "warning_count": int(row.get("ê²½ê³ ", 0) or 0),
                        "suspend_count": int(row.get("ì •ì§€", 0) or 0),
                    }
        except Exception as e:
            print(f"[ë§ˆì¼“í˜„í™©] ë§ˆì¼“ìƒíƒœí˜„í™© ì‹œíŠ¸ ì˜¤ë¥˜: {e}")
        
        # 3. ê³„ì • ì •ë³´ ê°€ì ¸ì˜¤ê¸°
        accounts = gsheet.get_accounts(None)
        
        # í”Œë«í¼ë³„ ê·¸ë£¹í™”
        summary = {}
        for acc in accounts:
            platform = acc.get("platform", "ê¸°íƒ€")
            if platform not in summary:
                summary[platform] = []
            
            store_name = acc.get("ìŠ¤í† ì–´ëª…") or acc.get("ìŠ¤í† ì–´ëª…", "")
            login_id = acc.get("login_id", "")
            account_name = store_name if store_name else login_id
            
            # ë§ˆì¼“ëª… ì •ê·œí™”
            market = platform
            if "ìŠ¤ë§ˆíŠ¸" in platform or "ë„¤ì´ë²„" in platform:
                market = "ìŠ¤ë§ˆíŠ¸ìŠ¤í† ì–´"
            elif "11" in platform:
                market = "11ë²ˆê°€"
            elif "ì¿ íŒ¡" in platform:
                market = "ì¿ íŒ¡"
            elif "ì§€ë§ˆì¼“" in platform:
                market = "ì§€ë§ˆì¼“"
            elif "ì˜¥ì…˜" in platform:
                market = "ì˜¥ì…˜"
            
            # ìƒí’ˆìˆ˜
            product_count = 0
            if market == "ìŠ¤ë§ˆíŠ¸ìŠ¤í† ì–´":
                product_count = ss_counts.get(account_name, 0)
                if product_count == 0 and "_" in account_name:
                    name_after_underscore = account_name.split("_", 1)[1]
                    product_count = ss_counts.get(name_after_underscore, 0)
            elif market == "11ë²ˆê°€":
                # 11ë²ˆê°€ ì‹œíŠ¸ì—ì„œ ì½ê¸°
                product_count = st_counts.get(account_name, 0)
                if product_count == 0:
                    product_count = st_counts.get(login_id, 0)
            
            # ìƒíƒœ/í˜ë„í‹° ì •ë³´
            key = f"{account_name}_{market}"
            status_info = status_map.get(key, {})
            status = status_info.get("status", "ì •ìƒ")
            caution_count = status_info.get("caution_count", 0)
            warning_count = status_info.get("warning_count", 0)
            suspend_count = status_info.get("suspend_count", 0)
            
            summary[platform].append({
                "ìŠ¤í† ì–´ëª…": account_name,
                "ìŠ¤í† ì–´ëª…": account_name,  # í•˜ìœ„ í˜¸í™˜
                "login_id": login_id,
                "status": status,
                "product_count": product_count,
                "caution_count": caution_count,
                "warning_count": warning_count,
                "suspend_count": suspend_count,
                "owner": acc.get("owner", ""),
                "usage": acc.get("usage", ""),
            })
        
        # ì •ë ¬ (ìƒí’ˆìˆ˜ ë‚´ë¦¼ì°¨ìˆœ)
        for platform in summary:
            summary[platform].sort(key=lambda x: x["product_count"], reverse=True)
        
        return {"success": True, "data": summary}
        
    except Exception as e:
        print(f"[ë§ˆì¼“í˜„í™©] ì˜¤ë¥˜: {e}")
        import traceback
        traceback.print_exc()
        return {"success": False, "error": str(e)}


@app.post("/api/11st/refresh-all-counts")
async def refresh_all_11st_counts(request: Request):
    """ëª¨ë“  11ë²ˆê°€ ê³„ì • ìƒí’ˆìˆ˜ ì¼ê´„ ì¡°íšŒ"""
    get_current_user(request)
    
    accounts = gsheet.get_accounts("11ë²ˆê°€")  # 11ë²ˆê°€ë§Œ
    results = []
    
    for acc in accounts:
        api_key = acc.get("st_api_key", "")
        if not api_key:
            continue
        
        login_id = acc.get("login_id", "")
        store_name = acc.get("ìŠ¤í† ì–´ëª…") or acc.get("ìŠ¤í† ì–´ëª…", login_id)
        
        try:
            count = await get_11st_product_count(api_key)
            st_product_cache[login_id] = {"count": count, "time": datetime.now()}
            results.append({"shop": store_name, "count": count, "success": True})
        except Exception as e:
            results.append({"shop": store_name, "count": 0, "success": False, "error": str(e)})
    
    return {"success": True, "results": results}


# ========== ë°°ì†¡ì¡°íšŒ API (ëª¨ë“ˆ ì‚¬ìš©) ==========
from modules.delivery_check import DeliveryChecker

# ë°°ì†¡ì¡°íšŒ ì¸ìŠ¤í„´ìŠ¤ ì´ˆê¸°í™” (ì „ì—­ CREDENTIALS_FILE ì‚¬ìš©)
delivery_checker = DeliveryChecker(CREDENTIALS_FILE)

@app.post("/api/delivery/check")
async def start_delivery_check(request: Request):
    """ë°°ì†¡ì¡°íšŒ ì‹œì‘"""
    get_current_user(request)

    data = await request.json()
    sheet_id = data.get("sheet_id", "")
    sheet_name = data.get("sheet_name", "")
    carrier_col = int(data.get("carrier_col", 43))
    tracking_col = int(data.get("tracking_col", 44))
    start_row = int(data.get("start_row", 4))

    # ì‘ì—… ë¡œê·¸ ê¸°ë¡
    log_work("ë°°ì†¡ì¡°íšŒ", sheet_name, 0, f"ì‹œíŠ¸: {sheet_name}", "ì›¹")

    # ëª¨ë“ˆì˜ ë¹„ë™ê¸° ë©”ì„œë“œ í˜¸ì¶œ (create_taskëŠ” ëª¨ë“ˆ ë‚´ë¶€ì—ì„œ ì²˜ë¦¬í•¨)
    return await delivery_checker.start_check(sheet_id, sheet_name, carrier_col, tracking_col, start_row)

@app.post("/api/delivery/stop")
async def stop_delivery_check(request: Request):
    """ë°°ì†¡ì¡°íšŒ ì¤‘ì§€"""
    get_current_user(request)
    return delivery_checker.stop_check()

@app.get("/api/delivery/status")
async def get_delivery_status(request: Request):
    """ë°°ì†¡ì¡°íšŒ ìƒíƒœ ì¡°íšŒ"""
    get_current_user(request)
    return delivery_checker.get_status()


# ========== ìŠ¤ì¼€ì¤„ëŸ¬ API ==========
class ScheduleRequest(BaseModel):
    name: str
    platform: str = "ìŠ¤ë§ˆíŠ¸ìŠ¤í† ì–´"
    task: str = "ë“±ë¡ê°¯ìˆ˜"
    stores: List[str] = []
    schedule_type: str = "cron"  # cron ë˜ëŠ” interval
    cron: str = "0 9 * * *"  # ë¶„ ì‹œ ì¼ ì›” ìš”ì¼
    interval_minutes: int = 60
    options: Dict = {}
    enabled: bool = True

@app.get("/api/schedules")
async def get_schedules(request: Request):
    """ìŠ¤ì¼€ì¤„ ëª©ë¡ ì¡°íšŒ"""
    get_current_user(request)
    schedules = load_schedules()
    
    # ë‹¤ìŒ ì‹¤í–‰ ì‹œê°„ ì¶”ê°€
    for s in schedules:
        job = scheduler.get_job(s['id'])
        if job and job.next_run_time:
            s['next_run'] = job.next_run_time.strftime('%Y-%m-%d %H:%M:%S')
        else:
            s['next_run'] = None
    
    return {"schedules": schedules}

@app.post("/api/schedules")
async def create_schedule(request: Request, req: ScheduleRequest):
    """ìŠ¤ì¼€ì¤„ ìƒì„±"""
    require_permission(request, "edit")
    
    schedules = load_schedules()
    
    # ìƒˆ ID ìƒì„±
    schedule_id = f"schedule_{int(time.time()*1000)}"
    
    new_schedule = {
        "id": schedule_id,
        "name": req.name,
        "platform": req.platform,
        "task": req.task,
        "stores": req.stores,
        "schedule_type": req.schedule_type,
        "cron": req.cron,
        "interval_minutes": req.interval_minutes,
        "options": req.options,
        "enabled": req.enabled,
        "created_at": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        "last_run": None,
        "run_count": 0
    }
    
    schedules.append(new_schedule)
    save_schedules(schedules)
    
    # í™œì„±í™”ëœ ê²½ìš° ì‘ì—… ë“±ë¡
    if req.enabled:
        add_schedule_job(new_schedule)
    
    return {"success": True, "schedule": new_schedule}

@app.put("/api/schedules/{schedule_id}")
async def update_schedule(request: Request, schedule_id: str, req: ScheduleRequest):
    """ìŠ¤ì¼€ì¤„ ìˆ˜ì •"""
    require_permission(request, "edit")
    
    schedules = load_schedules()
    
    for i, s in enumerate(schedules):
        if s['id'] == schedule_id:
            schedules[i].update({
                "name": req.name,
                "platform": req.platform,
                "task": req.task,
                "stores": req.stores,
                "schedule_type": req.schedule_type,
                "cron": req.cron,
                "interval_minutes": req.interval_minutes,
                "options": req.options,
                "enabled": req.enabled,
                "updated_at": datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            })
            
            save_schedules(schedules)
            
            # ì‘ì—… ê°±ì‹ 
            if scheduler.get_job(schedule_id):
                scheduler.remove_job(schedule_id)
            if req.enabled:
                add_schedule_job(schedules[i])
            
            return {"success": True, "schedule": schedules[i]}
    
    raise HTTPException(status_code=404, detail="ìŠ¤ì¼€ì¤„ì„ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤")

@app.delete("/api/schedules/{schedule_id}")
async def delete_schedule(request: Request, schedule_id: str):
    """ìŠ¤ì¼€ì¤„ ì‚­ì œ"""
    require_permission(request, "edit")
    
    schedules = load_schedules()
    schedules = [s for s in schedules if s['id'] != schedule_id]
    save_schedules(schedules)
    
    # ì‘ì—… ì œê±°
    if scheduler.get_job(schedule_id):
        scheduler.remove_job(schedule_id)
    
    return {"success": True}

@app.post("/api/schedules/{schedule_id}/toggle")
async def toggle_schedule(request: Request, schedule_id: str):
    """ìŠ¤ì¼€ì¤„ í™œì„±í™”/ë¹„í™œì„±í™”"""
    require_permission(request, "edit")
    
    schedules = load_schedules()
    
    for s in schedules:
        if s['id'] == schedule_id:
            s['enabled'] = not s.get('enabled', True)
            save_schedules(schedules)
            
            if s['enabled']:
                add_schedule_job(s)
            else:
                if scheduler.get_job(schedule_id):
                    scheduler.remove_job(schedule_id)
            
            return {"success": True, "enabled": s['enabled']}
    
    raise HTTPException(status_code=404, detail="ìŠ¤ì¼€ì¤„ì„ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤")

@app.post("/api/schedules/{schedule_id}/run")
async def run_schedule_now(request: Request, schedule_id: str):
    """ìŠ¤ì¼€ì¤„ ì¦‰ì‹œ ì‹¤í–‰"""
    require_permission(request, "edit")
    
    schedules = load_schedules()
    
    for s in schedules:
        if s['id'] == schedule_id:
            # ë¹„ë™ê¸°ë¡œ ì¦‰ì‹œ ì‹¤í–‰
            asyncio.create_task(execute_scheduled_task(
                schedule_id,
                s.get('platform', 'ìŠ¤ë§ˆíŠ¸ìŠ¤í† ì–´'),
                s.get('task', 'ë“±ë¡ê°¯ìˆ˜'),
                s.get('stores', []),
                s.get('options', {})
            ))
            return {"success": True, "message": "ì‘ì—…ì´ ì‹œì‘ë˜ì—ˆìŠµë‹ˆë‹¤"}
    
    raise HTTPException(status_code=404, detail="ìŠ¤ì¼€ì¤„ì„ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤")

@app.get("/api/schedules/{schedule_id}/log")
async def get_schedule_log(request: Request, schedule_id: str, lines: int = 100):
    """ìŠ¤ì¼€ì¤„ ì‹¤í–‰ ë¡œê·¸ ì¡°íšŒ"""
    require_permission(request, "view")
    
    log_file = os.path.join(os.path.dirname(__file__), "logs", f"schedule_{schedule_id}.log")
    
    if not os.path.exists(log_file):
        return {"success": False, "log": "", "message": "ë¡œê·¸ íŒŒì¼ì´ ì—†ìŠµë‹ˆë‹¤ (ì•„ì§ ì‹¤í–‰ëœ ì  ì—†ìŒ)"}
    
    try:
        # íŒŒì¼ í¬ê¸° í™•ì¸
        file_size = os.path.getsize(log_file)
        
        # ë§ˆì§€ë§‰ Nì¤„ ì½ê¸°
        with open(log_file, "r", encoding="utf-8", errors="replace") as f:
            all_lines = f.readlines()
            last_lines = all_lines[-lines:] if len(all_lines) > lines else all_lines
            log_content = "".join(last_lines)
        
        # íŒŒì¼ ìˆ˜ì • ì‹œê°„
        mtime = os.path.getmtime(log_file)
        modified_at = datetime.fromtimestamp(mtime).strftime("%Y-%m-%d %H:%M:%S")
        
        return {
            "success": True, 
            "log": log_content,
            "total_lines": len(all_lines),
            "file_size": file_size,
            "modified_at": modified_at
        }
    except Exception as e:
        return {"success": False, "log": "", "message": str(e)}


# ========== ë§ˆì¼€íŒ… ë°ì´í„° ìˆ˜ì§‘ ==========
marketing_tasks = {}  # {task_id: {"status": "running"|"completed"|"error", "current": 0, "total": 0, "logs": []}}
marketing_processes = {}

class MarketingCollectRequest(BaseModel):
    account_ids: List[str]

@app.post("/api/marketing/collect")
async def start_marketing_collection(request: Request, req: MarketingCollectRequest):
    """ë§ˆì¼€íŒ… ë°ì´í„° ìˆ˜ì§‘ ì‹œì‘"""
    require_permission(request, "edit")

    if not req.account_ids or len(req.account_ids) == 0:
        return {"error": "ê³„ì •ì„ ì„ íƒí•˜ì„¸ìš”"}

    # ì‘ì—… ë¡œê·¸ ê¸°ë¡
    account_names = ", ".join(req.account_ids[:5]) + ("..." if len(req.account_ids) > 5 else "")
    log_work("ë§ˆì¼€íŒ…ìˆ˜ì§‘", "ìŠ¤ë§ˆíŠ¸ìŠ¤í† ì–´", len(req.account_ids), f"ëŒ€ìƒ: {account_names}", "ì›¹")

    task_id = f"marketing_{int(time.time())}"

    # íƒœìŠ¤í¬ ì´ˆê¸°í™”
    marketing_tasks[task_id] = {
        "status": "running",
        "current": 0,
        "total": len(req.account_ids),
        "logs": [f"[{datetime.now().strftime('%H:%M:%S')}] ë§ˆì¼€íŒ… ë°ì´í„° ìˆ˜ì§‘ ì‹œì‘ ({len(req.account_ids)}ê°œ ê³„ì •)"]
    }
    
    # ë°±ê·¸ë¼ìš´ë“œì—ì„œ ì‹¤í–‰
    async def run_collection():
        try:
            # marketing_collector ì‹¤í–‰
            env = os.environ.copy()
            env["MARKETING_ACCOUNT_IDS"] = ",".join(req.account_ids)
            env["MARKETING_SPREADSHEET_KEY"] = MARKETING_SPREADSHEET_KEY
            env["SPREADSHEET_KEY"] = SPREADSHEET_KEY
            env["SERVICE_ACCOUNT_JSON"] = str(CREDENTIALS_FILE)
            env["API_KEY"] = API_KEY
            env["SERVER_URL"] = f"http://localhost:{PORT}"
            
            collector_path = os.path.join(os.path.dirname(__file__), "modules", "marketing_collector.py")
            
            # ë¹„ë™ê¸° ì„œë¸Œí”„ë¡œì„¸ìŠ¤ ìƒì„± (ë¸”ë¡œí‚¹ ë°©ì§€)
            proc = await asyncio.create_subprocess_exec(
                sys.executable, "-u", collector_path,
                stdout=asyncio.subprocess.PIPE,
                stderr=asyncio.subprocess.STDOUT,
                env=env,
                cwd=os.path.dirname(__file__)
            )
            
            marketing_processes[task_id] = proc
            
            print(f"[ë§ˆì¼€íŒ…ìˆ˜ì§‘] í”„ë¡œì„¸ìŠ¤ ì‹œì‘ - PID: {proc.pid}")
            
            # ë¹„ë™ê¸°ë¡œ ë¡œê·¸ ì½ê¸°
            while True:
                line_bytes = await proc.stdout.readline()
                if not line_bytes:
                    break
                
                line = line_bytes.decode('utf-8', errors='replace').strip()
                if line:
                    # print(f"[ë§ˆì¼€íŒ…ìˆ˜ì§‘] {line}")
                    marketing_tasks[task_id]["logs"].append(f"[{datetime.now().strftime('%H:%M:%S')}] {line}")

                    # ì§„í–‰ë¥  ì¶”ì¶œ
                    if "/" in line:
                        try:
                            parts = line.split("/")
                            if len(parts) == 2 and parts[0].split()[-1].isdigit():
                                current = int(parts[0].split()[-1])
                                marketing_tasks[task_id]["current"] = current
                        except:
                            pass

            await proc.wait()
            print(f"[ë§ˆì¼€íŒ…ìˆ˜ì§‘] í”„ë¡œì„¸ìŠ¤ ì¢…ë£Œ - exit code: {proc.returncode}")
            
            if proc.returncode == 0:
                marketing_tasks[task_id]["status"] = "completed"
                marketing_tasks[task_id]["current"] = marketing_tasks[task_id]["total"]
                marketing_tasks[task_id]["logs"].append(f"[{datetime.now().strftime('%H:%M:%S')}] [SUCCESS] ìˆ˜ì§‘ ì™„ë£Œ!")
            elif proc.returncode is None or proc.returncode == -15 or proc.returncode == 1:
                if marketing_tasks[task_id]["status"] == "running":
                    marketing_tasks[task_id]["status"] = "stopped"
                    marketing_tasks[task_id]["logs"].append(f"[{datetime.now().strftime('%H:%M:%S')}] â¹ ì‚¬ìš©ìì— ì˜í•´ ì¤‘ì§€ë¨")
            else:
                marketing_tasks[task_id]["status"] = "error"
                marketing_tasks[task_id]["logs"].append(f"[{datetime.now().strftime('%H:%M:%S')}] âŒ ì˜¤ë¥˜ ë°œìƒ (exit code: {proc.returncode})")
            
            if task_id in marketing_processes:
                del marketing_processes[task_id]
                
        except Exception as e:
            marketing_tasks[task_id]["status"] = "error"
            marketing_tasks[task_id]["logs"].append(f"[{datetime.now().strftime('%H:%M:%S')}] âŒ ì˜¤ë¥˜: {str(e)}")
            if task_id in marketing_processes:
                del marketing_processes[task_id]
    
    # ë¹„ë™ê¸° ì‹¤í–‰
    asyncio.create_task(run_collection())

    return {"task_id": task_id, "total": len(req.account_ids)}

@app.get("/api/marketing/collected-today")
async def get_marketing_collected_today(request: Request):
    """ê° ìŠ¤í† ì–´ì˜ ë§ˆì§€ë§‰ ìˆ˜ì§‘ ì‹œê°„ ì¡°íšŒ"""
    require_permission(request, "view")

    try:
        from google.oauth2.service_account import Credentials
        scopes = ["https://www.googleapis.com/auth/spreadsheets"]
        creds = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=scopes)
        gc = gspread.authorize(creds)
        sheet = gc.open_by_key(MARKETING_SPREADSHEET_KEY)

        today = datetime.now().strftime("%Y-%m-%d")
        collected_stores = []  # ì˜¤ëŠ˜ ìˆ˜ì§‘ëœ ìŠ¤í† ì–´
        last_collected = {}    # {ìŠ¤í† ì–´ëª…: ë§ˆì§€ë§‰ìˆ˜ì§‘ì‹œê°„}

        # ì‹œìŠ¤í…œ ì‹œíŠ¸ ì œì™¸
        system_sheets = {"ì „ì²´ë°ì´í„°", "ì‡¼í•‘ëª°ì •ë³´", "í…œí”Œë¦¿", "ì„¤ì •", "ìŠ¤í† ì–´ìœ ì…ìˆ˜"}

        worksheets = sheet.worksheets()
        for ws in worksheets:
            if ws.title in system_sheets:
                continue

            try:
                # Aì—´(ìˆ˜ì§‘ì¼ì)ì—ì„œ ë§ˆì§€ë§‰ ë‚ ì§œ í™•ì¸ (ìƒˆ êµ¬ì¡°: ë‚ ì§œë³„ ëˆ„ì )
                a_col = ws.col_values(1)  # Aì—´ ì „ì²´
                if len(a_col) > 1:  # í—¤ë” ì œì™¸
                    # Aì—´ì—ì„œ ë‚ ì§œ í˜•ì‹ì¸ ë§ˆì§€ë§‰ ê°’ ì°¾ê¸°
                    last_date = None
                    for val in reversed(a_col):
                        if val and len(val) >= 10 and val[:4].isdigit():  # YYYY-MM-DD í˜•ì‹
                            last_date = val
                            break

                    if last_date:
                        last_collected[ws.title] = last_date
                        if today == last_date or today in last_date:
                            collected_stores.append(ws.title)
            except:
                pass

        return {
            "success": True,
            "date": today,
            "collected": collected_stores,
            "count": len(collected_stores),
            "last_collected": last_collected
        }

    except Exception as e:
        print(f"ìˆ˜ì§‘ í˜„í™© ì¡°íšŒ ì˜¤ë¥˜: {e}")
        return {"success": False, "error": str(e)}

@app.post("/api/marketing/stop")
async def stop_marketing_collection(request: Request):
    """ì‹¤í–‰ ì¤‘ì¸ ë§ˆì¼€íŒ… ìˆ˜ì§‘ ì¤‘ì§€"""
    require_permission(request, "edit")
    
    data = await request.json()
    task_id = data.get("task_id")
    
    if not task_id:
        # ê°€ì¥ ìµœê·¼ì˜ running íƒœìŠ¤í¬ ì¤‘ì§€
        for tid, t in marketing_tasks.items():
            if t["status"] == "running":
                task_id = tid
                break
    
    if not task_id or task_id not in marketing_processes:
        return {"success": False, "message": "ì‹¤í–‰ ì¤‘ì¸ ì‘ì—…ì„ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤."}
    
    try:
        proc = marketing_processes[task_id]
        import signal
        if sys.platform == "win32":
            proc.terminate() # Windowsì—ì„œëŠ” terminate()ê°€ ì˜ ì‘ë™í•¨
        else:
            proc.send_signal(signal.SIGTERM)
            
        marketing_tasks[task_id]["status"] = "stopped"
        marketing_tasks[task_id]["logs"].append(f"[{datetime.now().strftime('%H:%M:%S')}] â¹ ì¤‘ì§€ ìš”ì²­ë¨")
        
        return {"success": True, "message": "ì‘ì—…ì´ ì¤‘ì§€ë˜ì—ˆìŠµë‹ˆë‹¤."}
    except Exception as e:
        return {"success": False, "message": f"ì¤‘ì§€ ì¤‘ ì˜¤ë¥˜: {str(e)}"}

@app.get("/api/marketing/progress/{task_id}")
async def get_marketing_progress(request: Request, task_id: str):
    """ë§ˆì¼€íŒ… ë°ì´í„° ìˆ˜ì§‘ ì§„í–‰ ìƒí™© ì¡°íšŒ"""
    require_permission(request, "view")

    if task_id not in marketing_tasks:
        return {"error": "ì‘ì—…ì„ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤"}

    return marketing_tasks[task_id]

@app.get("/api/marketing/progress-stream/{task_id}")
async def get_marketing_progress_stream(request: Request, task_id: str):
    """ë§ˆì¼€íŒ… ë°ì´í„° ìˆ˜ì§‘ ì§„í–‰ ìƒí™© SSE ìŠ¤íŠ¸ë¦¼"""
    from fastapi.responses import StreamingResponse

    async def event_generator():
        last_log_count = 0
        while True:
            if task_id not in marketing_tasks:
                yield f"data: {json.dumps({'error': 'ì‘ì—… ì—†ìŒ'})}\n\n"
                break

            task = marketing_tasks[task_id]
            current_log_count = len(task.get("logs", []))

            # ë³€ê²½ì‚¬í•­ì´ ìˆì„ ë•Œë§Œ ì „ì†¡
            if current_log_count != last_log_count or task.get("status") in ["completed", "error"]:
                yield f"data: {json.dumps(task, ensure_ascii=False)}\n\n"
                last_log_count = current_log_count

            if task.get("status") in ["completed", "error"]:
                break

            await asyncio.sleep(0.5)  # 0.5ì´ˆë§ˆë‹¤ ì²´í¬

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "Connection": "keep-alive"}
    )

@app.post("/api/marketing/create-sheets")
async def create_marketing_sheets(request: Request):
    """ë§ˆì¼€íŒ… ìŠ¤í”„ë ˆë“œì‹œíŠ¸ ì´ˆê¸°í™”"""
    require_permission(request, "edit")

    try:
        from google.oauth2.service_account import Credentials

        scopes = ["https://www.googleapis.com/auth/spreadsheets"]
        creds = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=scopes)
        gc = gspread.authorize(creds)
        sheet = gc.open_by_key(MARKETING_SPREADSHEET_KEY)

        results = []
        system_sheets = {"ì „ì²´ë°ì´í„°", "ì‡¼í•‘ëª°ì •ë³´", "í…œí”Œë¦¿", "ì„¤ì •", "ìŠ¤í† ì–´ìœ ì…ìˆ˜"}

        # ê°œë³„ ë§ˆì¼“ ì‹œíŠ¸ ë°ì´í„° ì´ˆê¸°í™” (ì‚­ì œí•˜ì§€ ì•Šê³  ë‚´ìš©ë§Œ ë¹„ì›€)
        all_worksheets = sheet.worksheets()
        for ws in all_worksheets:
            if ws.title not in system_sheets:
                try:
                    ws.clear()
                    results.append({"sheet": ws.title, "status": "ë°ì´í„° ì´ˆê¸°í™”ë¨"})
                except Exception as e:
                    results.append({"sheet": ws.title, "status": f"ì´ˆê¸°í™” ì‹¤íŒ¨: {str(e)[:20]}"})

        # "ì „ì²´ë°ì´í„°" ì‹œíŠ¸ ìƒì„±/ì´ˆê¸°í™”
        try:
            ws_data = sheet.worksheet("ì „ì²´ë°ì´í„°")
            ws_data.clear()
            ws_data.update('A1', [[
                "ìˆ˜ì§‘ì¼ì‹œ", "ìŠ¤í† ì–´ëª…", "ìƒí’ˆë²ˆí˜¸", "ìƒí’ˆëª…",
                "ë°©ë¬¸íšŸìˆ˜", "ìœ ì…ìˆ˜", "í´ë¦­ìˆ˜", "ì „í™˜ìˆ˜", "íŒë§¤ì•¡"
            ]])
            results.append({"sheet": "ì „ì²´ë°ì´í„°", "status": "ì´ˆê¸°í™”ë¨"})
        except:
            ws_data = sheet.add_worksheet(title="ì „ì²´ë°ì´í„°", rows=50000, cols=10)
            ws_data.update('A1', [[
                "ìˆ˜ì§‘ì¼ì‹œ", "ìŠ¤í† ì–´ëª…", "ìƒí’ˆë²ˆí˜¸", "ìƒí’ˆëª…",
                "ë°©ë¬¸íšŸìˆ˜", "ìœ ì…ìˆ˜", "í´ë¦­ìˆ˜", "ì „í™˜ìˆ˜", "íŒë§¤ì•¡"
            ]])
            results.append({"sheet": "ì „ì²´ë°ì´í„°", "status": "ìƒˆë¡œ ìƒì„±ë¨"})

        # "ì‡¼í•‘ëª°ì •ë³´" ì‹œíŠ¸ ìƒì„±/ì´ˆê¸°í™”
        try:
            ws_mall = sheet.worksheet("ì‡¼í•‘ëª°ì •ë³´")
            ws_mall.clear()
            ws_mall.update('A1', [[
                "ìˆ˜ì§‘ì¼ì‹œ", "ìŠ¤í† ì–´ëª…", "ì´ë°©ë¬¸ììˆ˜", "ì¬ë°©ë¬¸ììˆ˜", "ì‹ ê·œë°©ë¬¸ììˆ˜",
                "ì´í˜ì´ì§€ë·°", "í‰ê· ì²´ë¥˜ì‹œê°„", "ì´íƒˆë¥ ", "êµ¬ë§¤ì „í™˜ìœ¨"
            ]])
            results.append({"sheet": "ì‡¼í•‘ëª°ì •ë³´", "status": "ì´ˆê¸°í™”ë¨"})
        except:
            ws_mall = sheet.add_worksheet(title="ì‡¼í•‘ëª°ì •ë³´", rows=1000, cols=10)
            ws_mall.update('A1', [[
                "ìˆ˜ì§‘ì¼ì‹œ", "ìŠ¤í† ì–´ëª…", "ì´ë°©ë¬¸ììˆ˜", "ì¬ë°©ë¬¸ììˆ˜", "ì‹ ê·œë°©ë¬¸ììˆ˜",
                "ì´í˜ì´ì§€ë·°", "í‰ê· ì²´ë¥˜ì‹œê°„", "ì´íƒˆë¥ ", "êµ¬ë§¤ì „í™˜ìœ¨"
            ]])
            results.append({"sheet": "ì‡¼í•‘ëª°ì •ë³´", "status": "ìƒˆë¡œ ìƒì„±ë¨"})

        spreadsheet_url = f"https://docs.google.com/spreadsheets/d/{MARKETING_SPREADSHEET_KEY}"
        return {"success": True, "results": results, "spreadsheet_url": spreadsheet_url}

    except Exception as e:
        return {"success": False, "error": str(e)}


@app.get("/api/marketing/data")
async def get_marketing_data(request: Request, store: str = None, refresh: bool = False):
    """ë§ˆì¼€íŒ… ìˆ˜ì§‘ ë°ì´í„° ì¡°íšŒ
    - store: íŠ¹ì • ìŠ¤í† ì–´ëª… (ì—†ìœ¼ë©´ ì „ì²´ ëª©ë¡ë§Œ ë°˜í™˜)
    """
    require_permission(request, "view")

    try:
        # ëª¨ë“  ì›Œí¬ì‹œíŠ¸ ì¡°íšŒ (ìºì‹œ í™œìš©)
        store_names = gsheet.get_worksheet_names_with_cache(sheet_key=MARKETING_SPREADSHEET_KEY, force_refresh=refresh)
        store_names = [name for name in store_names if name not in ["í…œí”Œë¦¿", "ì„¤ì •", "ì „ì²´ë°ì´í„°", "ì‡¼í•‘ëª°ì •ë³´", "ìŠ¤í† ì–´ìœ ì…ìˆ˜"]]
        
        result = {
            "success": True,
            "stores": store_names,
            "data": {}
        }

        if store:
            # íŠ¹ì • ìŠ¤í† ì–´ ë°ì´í„°ë§Œ (ìƒì„¸ ì •ë³´ ìš”ì²­ ì‹œì—ë§Œ ì‹œíŠ¸ ì ‘ê·¼)
            try:
                # gsheet ì¸ìŠ¤í„´ìŠ¤ì˜ ìºì‹œ ê¸°ëŠ¥ í™œìš©
                all_values = gsheet.get_values_with_cache(store, sheet_key=MARKETING_SPREADSHEET_KEY, force_refresh=refresh)

                # ë°ì´í„° íŒŒì‹± (ì—´ ìœ„ì¹˜ ê¸°ë°˜ - data111.py êµ¬ì¡°)
                biz_data = []       # ë§ˆì¼€íŒ…ë¶„ì„ (ìƒí’ˆë…¸ì¶œì„±ê³¼)
                partner_data = []   # ìƒí’ˆí´ë¦­ë¦¬í¬íŠ¸
                mall_info = {}      # ì‡¼í•‘ëª°ì •ë³´
                channel_data = []   # ì „ì²´ì±„ë„

                # í—¤ë” í‚¤ì›Œë“œ ëª©ë¡
                header_keywords = {"ìƒí’ˆëª…", "ìƒí’ˆID", "ìƒí’ˆ ID", "ì±„ë„ê·¸ë£¹", "ì±„ë„ëª…", "í‚¤ì›Œë“œ",
                                   "í‰ê· ë…¸ì¶œìˆœìœ„", "ìœ ì…ìˆ˜", "ë…¸ì¶œìˆ˜", "í´ë¦­ìˆ˜", "í´ë¦­ìœ¨", "í´ë¦­ë¥ ",
                                   "ì ìš©ìˆ˜ìˆ˜ë£Œ", "í´ë¦­ë‹¹ìˆ˜ìˆ˜ë£Œ", "í•­ëª©", "ë…¸ì¶œ", "ì„±ê³¼"}

                for i, row in enumerate(all_values):
                    if i == 0: continue

                    # ë§ˆì¼€íŒ…ë¶„ì„ (A~Gì—´)
                    if len(row) > 6 and row[0]:
                        product_name = row[0].strip()
                        if product_name and product_name not in header_keywords:
                            biz_data.append({
                                "ìƒí’ˆëª…": product_name,
                                "ìƒí’ˆID": row[1].strip() if len(row) > 1 else "",
                                "ì±„ë„ê·¸ë£¹": row[2].strip() if len(row) > 2 else "",
                                "ì±„ë„ëª…": row[3].strip() if len(row) > 3 else "",
                                "í‚¤ì›Œë“œ": row[4].strip() if len(row) > 4 else "",
                                "í‰ê· ë…¸ì¶œìˆœìœ„": row[5].strip() if len(row) > 5 else "",
                                "ìœ ì…ìˆ˜": row[6].strip() if len(row) > 6 else "0"
                            })

                    # ìƒí’ˆí´ë¦­ë¦¬í¬íŠ¸ (I~Oì—´)
                    if len(row) > 14 and row[8]:
                        product_id = row[8].strip()
                        if product_id and product_id not in header_keywords and product_id[0].isdigit():
                            partner_data.append({
                                "ìƒí’ˆID": product_id,
                                "ìƒí’ˆëª…": row[9].strip() if len(row) > 9 else "",
                                "ë…¸ì¶œìˆ˜": row[10].strip() if len(row) > 10 else "0",
                                "í´ë¦­ìˆ˜": row[11].strip() if len(row) > 11 else "0",
                                "í´ë¦­ìœ¨": row[12].strip() if len(row) > 12 else "",
                                "ì ìš©ìˆ˜ìˆ˜ë£Œ": row[13].strip() if len(row) > 13 else "",
                                "í´ë¦­ë‹¹ìˆ˜ìˆ˜ë£Œ": row[14].strip() if len(row) > 14 else ""
                            })

                    # ì‡¼í•‘ëª°ì •ë³´ (Q~Rì—´)
                    if len(row) > 17 and row[16]:
                        p_key = row[16].strip()
                        if p_key and p_key not in header_keywords:
                            mall_info[p_key] = row[17].strip() if len(row) > 17 else ""

                    # ì „ì²´ì±„ë„ (S~Tì—´)
                    if len(row) > 19 and row[18]:
                        channel_name = row[18].strip()
                        if channel_name and channel_name not in header_keywords:
                            channel_data.append({
                                "ì±„ë„ëª…": channel_name,
                                "ìœ ì…ìˆ˜": row[19].strip() if len(row) > 19 else "0"
                            })

                result["data"][store] = {
                    "biz_advisor": biz_data,
                    "shopping_partner": partner_data,
                    "mall_info": mall_info,
                    "channel_data": channel_data
                }
            except Exception as e:
                result["data"][store] = {"error": str(e)}
        
        return result

    except Exception as e:
        print(f"ë§ˆì¼€íŒ… ë°ì´í„° ì¡°íšŒ ì˜¤ë¥˜: {e}")
        return {"success": False, "error": str(e)}


@app.get("/api/marketing/store-traffic")
async def get_marketing_store_traffic(request: Request):
    """ìŠ¤í† ì–´ìœ ì…ìˆ˜ ì‹œíŠ¸ ë°ì´í„° ì¡°íšŒ (ë‚ ì§œë³„ ì¶”ì´) - ì „ì¼ëŒ€ë¹„ìš©"""
    require_permission(request, "view")
    try:
        from google.oauth2.service_account import Credentials
        scopes = ["https://www.googleapis.com/auth/spreadsheets"]
        creds = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=scopes)
        gc = gspread.authorize(creds)
        sheet = gc.open_by_key(MARKETING_SPREADSHEET_KEY)
        
        target_ws = None
        try:
            target_ws = sheet.worksheet("ìŠ¤í† ì–´ìœ ì…ìˆ˜")
        except:
            try:
                target_ws = sheet.worksheet("ì‡¼í•‘ëª°ì •ë³´")
            except:
                return {"success": False, "error": "'ìŠ¤í† ì–´ìœ ì…ìˆ˜' ë˜ëŠ” 'ì‡¼í•‘ëª°ì •ë³´' ì‹œíŠ¸ë¥¼ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤."}
        
        rows = target_ws.get_all_values()
        if len(rows) < 2:
            return {"success": True, "data": [], "dates": []}
            
        header = rows[0]
        # ë‚ ì§œ ì»¬ëŸ¼ ì°¾ê¸° (ìš©ë„, ìŠ¤í† ì–´ëª… ì œì™¸í•œ ë‚˜ë¨¸ì§€)
        date_cols = []
        usage_idx = -1
        store_idx = -1
        
        for i, col in enumerate(header):
            col = col.strip()
            if col == "ìš©ë„": usage_idx = i
            elif col == "ìŠ¤í† ì–´ëª…": store_idx = i
            elif col: # ë‚ ì§œë¡œ ì¶”ì • (ë‚˜ë¨¸ì§€ëŠ” ë‹¤ ë‚ ì§œ ì»¬ëŸ¼ìœ¼ë¡œ ê°„ì£¼)
                date_cols.append({"index": i, "label": col})
        
        if store_idx == -1:
             return {"success": False, "error": "'ìŠ¤í† ì–´ëª…' ì»¬ëŸ¼ì„ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤."}

        result_data = []
        for r in rows[1:]:
            if len(r) <= store_idx: continue
            store_name = r[store_idx].strip()
            if not store_name: continue
            
            usage = r[usage_idx].strip() if usage_idx != -1 and len(r) > usage_idx else ""
            
            date_values = {}
            for dc in date_cols:
                idx = dc["index"]
                val = r[idx].strip() if len(r) > idx else "0"
                # ìˆ«ì ë³€í™˜ (ì½¤ë§ˆ ì œê±°)
                try: 
                    val_int = int(val.replace(",", ""))
                except: 
                    val_int = 0
                date_values[dc["label"]] = val_int
            
            result_data.append({
                "store": store_name,
                "usage": usage,
                "values": date_values
            })
            
        return {
            "success": True,
            "dates": [d["label"] for d in date_cols],
            "data": result_data
        }

    except Exception as e:
        return {"success": False, "error": str(e)}


@app.post("/api/marketing/sync-manual")
async def sync_marketing_manual(request: Request):
    """ë§ˆì¼€íŒ… ë°ì´í„° ìˆ˜ë™ ì·¨í•© (T2 ì…€ ìˆ˜ì§‘)"""
    require_permission(request, "edit")
    
    # (ìˆ˜ì •) ë¶ˆí•„ìš”í•œ ìœ„ì„ ì½”ë“œ ì‚­ì œë¨ - ì•„ë˜ ì§ì ‘ êµ¬í˜„ëœ ë¡œì§ ì‚¬ìš©


    try:
        from google.oauth2.service_account import Credentials
        scopes = ["https://www.googleapis.com/auth/spreadsheets"]
        creds = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=scopes)
        gc = gspread.authorize(creds)
        sheet = gc.open_by_key(MARKETING_SPREADSHEET_KEY)

        # 1. ëª¨ë“  ì›Œí¬ì‹œíŠ¸ ëª©ë¡ ì¡°íšŒ
        all_worksheets = sheet.worksheets()
        system_sheets = {"ì „ì²´ë°ì´í„°", "ì‡¼í•‘ëª°ì •ë³´", "í…œí”Œë¦¿", "ì„¤ì •", "ìŠ¤í† ì–´ìœ ì…ìˆ˜"}
        store_sheets = [ws for ws in all_worksheets if ws.title not in system_sheets]

        if not store_sheets:
            return {"success": False, "error": "ì·¨í•©í•  ê°œë³„ ìŠ¤í† ì–´ ì‹œíŠ¸ê°€ ì—†ìŠµë‹ˆë‹¤."}

        performance_rows = []  # ì‡¼í•‘ëª°ì •ë³´ì— ë“¤ì–´ê°ˆ í–‰ë“¤
        summary_rows = []      # ì „ì²´ë°ì´í„°ì— ë“¤ì–´ê°ˆ í–‰ë“¤
        cleared_sheets = []    # ì´ˆê¸°í™”ëœ ì‹œíŠ¸ ëª©ë¡
        
        # í—¤ë” í‚¤ì›Œë“œ (ë°ì´í„° ì‹œì‘ì  íŒë³„ìš©)
        header_keywords = {"ìƒí’ˆëª…", "ìƒí’ˆID", "ìƒí’ˆ ID", "ì±„ë„ê·¸ë£¹", "ì±„ë„ëª…", "í‚¤ì›Œë“œ", "í‰ê· ë…¸ì¶œìˆœìœ„", "ìœ ì…ìˆ˜", "ë…¸ì¶œìˆ˜", "í´ë¦­ìˆ˜", "í´ë¦­ìœ¨", "í´ë¦­ë¥ ", "ì ìš©ìˆ˜ìˆ˜ë£Œ", "í´ë¦­ë‹¹ìˆ˜ìˆ˜ë£Œ", "í•­ëª©", "ë…¸ì¶œ", "ì„±ê³¼"}

        for ws in store_sheets:
            try:
                values = ws.get_all_values()
                if not values or len(values) < 2: continue
                
                store_name = ws.title
                collect_time = ""
                
                # Q-Rì—´ ë“±ì—ì„œ ê³µí†µ ì •ë³´(mall_info) ì¶”ì¶œ
                mall_info = {}
                for row in values:
                    # ìˆ˜ì§‘ì¼ì‹œ ì°¾ê¸° (ë³´í†µ íŠ¹ì • ì…€ì— ê¸°ë¡ë¨)
                    if len(row) > 1 and "ìˆ˜ì§‘ì¼ì‹œ" in str(row[0]):
                        collect_time = row[1]
                    
                    # mall_info (Q-Rì—´ ë“± - get_marketing_data ë¡œì§ ì°¸ì¡°)
                    if len(row) > 17 and row[16]:
                        key = row[16].strip()
                        if key and key not in header_keywords:
                            mall_info[key] = row[17].strip()
                
                # ìˆ˜ì§‘ì‹œê°„ì´ ì—†ìœ¼ë©´ í˜„ì¬ ì‹œê°„ìœ¼ë¡œ
                if not collect_time: collect_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

                # (1) ì‡¼í•‘ëª°ì •ë³´ ë°ì´í„° ìƒì„±
                if mall_info:
                    performance_rows.append([
                        collect_time,
                        store_name,
                        mall_info.get("ì´ ë°©ë¬¸ììˆ˜", "0"),
                        mall_info.get("ì¬ë°©ë¬¸ììˆ˜", "0"),
                        mall_info.get("ì‹ ê·œë°©ë¬¸ììˆ˜", "0"),
                        mall_info.get("í˜ì´ì§€ë·°", "0"),
                        mall_info.get("í‰ê·  ì²´ë¥˜ì‹œê°„", "-"),
                        mall_info.get("ì´íƒˆë¥ ", "-"),
                        mall_info.get("êµ¬ë§¤ì „í™˜ìœ¨", "-")
                    ])

                # (2) ì „ì²´ë°ì´í„° (ì œí’ˆë³„) ìƒì„± - biz_advisor ë° shopping_partner ë³‘í•© ì‹œë„
                # ì°¸ê³ : ê°œë³„ ì‹œíŠ¸ì—ì„œ ì •í™•í•œ ìƒí’ˆë²ˆí˜¸/ë°©ë¬¸íšŸìˆ˜ ë“±ì„ ê°€ì ¸ì˜¤ê¸° ì–´ë ¤ìš¸ ìˆ˜ ìˆìœ¼ë¯€ë¡œ
                # biz_advisor(A-G)ì™€ shopping_partner(I-O) ë°ì´í„°ë¥¼ ìµœëŒ€í•œ ì¡°í•©í•©ë‹ˆë‹¤.
                
                # ì‹¤ì œ 'ì „ì²´ë°ì´í„°' ì‹œíŠ¸ëŠ” ì™¸ë¶€ ìˆ˜ì§‘ í”„ë¡œê·¸ë¨(data111.py)ì—ì„œ ì§ì ‘ ê¸°ë¡í•˜ë„ë¡ ì„¤ê³„ë˜ì–´ ìˆìœ¼ë‚˜
                # ì—¬ê¸°ì„œë„ ê°œë³„ ì‹œíŠ¸ ê¸°ë°˜ìœ¼ë¡œ ë³´ì™„ ì§‘ê³„í•©ë‹ˆë‹¤.
                for row in values:
                    if not row or row[0] in header_keywords: continue
                    
                    # ìƒí’ˆëª…(A)ì´ ìˆëŠ” í–‰ í•„í„°ë§
                    product_name = row[0].strip() if len(row) > 0 else ""
                    if not product_name: continue
                    
                    # ë°ì´í„° ì¶”ì¶œ
                    summary_rows.append([
                        collect_time,
                        store_name,
                        row[1].strip() if len(row) > 1 else "",  # ìƒí’ˆë²ˆí˜¸
                        product_name,
                        "0", # ë°©ë¬¸íšŸìˆ˜ (ê°œë³„ ì‹œíŠ¸ì—” í•©ì‚°ë§Œ ìˆëŠ” ê²½ìš°ê°€ ë§ìŒ)
                        row[6].strip() if len(row) > 6 else "0", # ìœ ì…ìˆ˜
                        "0", # í´ë¦­ìˆ˜
                        "0", # ì „í™˜ìˆ˜
                        "0"  # íŒë§¤ì•¡
                    ])
            except Exception as e:
                print(f"ğŸ“¦ [ì·¨í•©] {ws.title} ì²˜ë¦¬ ì¤‘ ì˜¤ë¥˜: {e}")

        # 1.5. ê°œë³„ ìŠ¤í† ì–´ ì‹œíŠ¸ ì´ˆê¸°í™” (ìŠ¤í† ì–´ìœ ì…ìˆ˜ ì œì™¸)
        for ws in store_sheets:
            try:
                ws.clear()
                cleared_sheets.append(ws.title)
                print(f"ğŸ“¦ [ì·¨í•©] {ws.title} ì‹œíŠ¸ ì´ˆê¸°í™” ì™„ë£Œ")
            except Exception as e:
                print(f"ğŸ“¦ [ì·¨í•©] {ws.title} ì‹œíŠ¸ ì´ˆê¸°í™” ì‹¤íŒ¨: {e}")

        # 2. ë§ˆìŠ¤í„° ì‹œíŠ¸ì— ì“°ê¸° (Batch Update)
        # (1) ì‡¼í•‘ëª°ì •ë³´ ì—…ë°ì´íŠ¸
        try:
            # (1) ìŠ¤í† ì–´ìœ ì…ìˆ˜ (êµ¬ ì‡¼í•‘ëª°ì •ë³´) ì—…ë°ì´íŠ¸
            sheet_name = "ìŠ¤í† ì–´ìœ ì…ìˆ˜"
            try:
                ws_perf = sheet.worksheet(sheet_name)
            except:
                try:
                    ws_perf = sheet.worksheet("ì‡¼í•‘ëª°ì •ë³´") # êµ¬ ì´ë¦„ í˜¸í™˜
                    ws_perf.update_title(sheet_name)
                except:
                    ws_perf = sheet.add_worksheet(sheet_name, 1000, 20)
                    ws_perf.append_row(["ìš©ë„", "ìŠ¤í† ì–´ëª…"])

            # ê¸°ì¡´ ë°ì´í„° ì½ê¸° (ë°ì´í„° ë³´ì¡´)
            existing_data = ws_perf.get_all_values()
            
            # í—¤ë” ì´ˆê¸°í™”
            if not existing_data:
                header = ["ìš©ë„", "ìŠ¤í† ì–´ëª…"]
                existing_data = [header]
            else:
                header = existing_data[0]
                if len(header) < 2:
                    header = ["ìš©ë„", "ìŠ¤í† ì–´ëª…"]
                    existing_data[0] = header

            # ì˜¤ëŠ˜ ë‚ ì§œ (Mì›” Dì¼ í¬ë§·)
            now = datetime.now()
            today_str = f"{now.month}ì›” {now.day}ì¼"
            
            # ë‚ ì§œ ì»¬ëŸ¼ ì¶”ê°€
            if today_str not in header:
                header.append(today_str)
                # ê¸°ì¡´ í–‰ì—ë„ ë¹ˆ ì¹¸ ì¶”ê°€
                for row in existing_data[1:]:
                    row.append("")
            
            # ë‚ ì§œ ì»¬ëŸ¼ 30ê°œ ìœ ì§€ (ì¸ë±ìŠ¤ 2ë¶€í„°)
            # ë‚ ì§œ ì»¬ëŸ¼ë§Œ ì¶”ì¶œí•´ì„œ ì •ë ¬í•˜ê±°ë‚˜, ë‹¨ìˆœíˆ ê¸¸ì´ë¡œ íŒë‹¨
            # ì—¬ê¸°ì„œëŠ” ë‹¨ìˆœíˆ ê°€ì¥ ì™¼ìª½ì˜ ë‚ ì§œ(ì¸ë±ìŠ¤ 2)ë¥¼ ì‚­ì œí•˜ëŠ” ë°©ì‹ì„ ì‚¬ìš© (ì‚¬ìš©ì ìš”ì²­: ì™¼ìª½ìœ¼ë¡œ ë‹¹ê²¨ì§)
            # í—¤ë”ì˜ ê¸¸ì´ê°€ 2(ê¸°ë³¸) + 30(ë‚ ì§œ) = 32ê°œë¥¼ ë„˜ìœ¼ë©´
            while len(header) > 32:
                # ì¸ë±ìŠ¤ 2 ì œê±°
                header.pop(2)
                for row in existing_data[1:]:
                    if len(row) > 2:
                        row.pop(2)

            # ì˜¤ëŠ˜ ë‚ ì§œ ì»¬ëŸ¼ ì¸ë±ìŠ¤ ì¬ê³„ì‚°
            today_col_idx = header.index(today_str)

            # ë°ì´í„° ë³‘í•© (ìŠ¤í† ì–´ëª… ê¸°ì¤€)
            # performance_rows êµ¬ì¡°: [ìˆ˜ì§‘ì¼ì‹œ, ìŠ¤í† ì–´ëª…, ì´ë°©ë¬¸ììˆ˜, ...]
            
            # ìŠ¤í† ì–´ëª… ë§¤í•‘ (ìŠ¤í† ì–´ëª… -> í–‰ ì¸ë±ìŠ¤)
            store_row_map = {}
            for i, row in enumerate(existing_data):
                if i == 0: continue
                if len(row) > 1:
                    store_row_map[row[1].strip()] = i

            for p_row in performance_rows:
                s_name = p_row[1].strip()
                visits = p_row[2] # ì´ë°©ë¬¸ììˆ˜

                if s_name in store_row_map:
                    row_idx = store_row_map[s_name]
                    target_row = existing_data[row_idx]
                    # í–‰ ê¸¸ì´ ë³´ì •
                    while len(target_row) <= today_col_idx:
                        target_row.append("")
                    target_row[today_col_idx] = visits
                else:
                    # ì‹ ê·œ ìŠ¤í† ì–´
                    new_row = [""] * len(header)
                    new_row[0] = "" # ìš©ë„
                    new_row[1] = s_name
                    new_row[today_col_idx] = visits
                    existing_data.append(new_row)

            # ì‹œíŠ¸ ì—…ë°ì´íŠ¸
            ws_perf.clear()
            ws_perf.update('A1', existing_data)
        except Exception as e:
            print(f"ğŸ“¦ [ì·¨í•©] ì‡¼í•‘ëª°ì •ë³´ ì“°ê¸° ì‹¤íŒ¨: {e}")

        # (2) ì „ì²´ë°ì´í„° ì—…ë°ì´íŠ¸
        try:
            try:
                ws_total = sheet.worksheet("ì „ì²´ë°ì´í„°")
            except:
                ws_total = sheet.add_worksheet("ì „ì²´ë°ì´í„°", 10000, 26) # Zì—´ê¹Œì§€ ë„‰ë„‰í•˜ê²Œ

            ws_total.clear()
            header = ["ìˆ˜ì§‘ì¼ì‹œ", "ìŠ¤í† ì–´ëª…", "ìƒí’ˆë²ˆí˜¸", "ìƒí’ˆëª…", "ë°©ë¬¸íšŸìˆ˜", "ìœ ì…ìˆ˜", "í´ë¦­ìˆ˜", "ì „í™˜ìˆ˜", "íŒë§¤ì•¡"]
            ws_total.update('A1', [header] + summary_rows[:10000]) # ë„ˆë¬´ ë§ìœ¼ë©´ ìë¦„
        except Exception as e:
            print(f"ğŸ“¦ [ì·¨í•©] ì „ì²´ë°ì´í„° ì“°ê¸° ì‹¤íŒ¨: {e}")

        return {
            "success": True,
            "message": f"{len(store_sheets)}ê°œ ìŠ¤í† ì–´ ì·¨í•© ì™„ë£Œ (ì„±ê³¼ {len(performance_rows)}ê±´, ìƒì„¸ {len(summary_rows)}ê±´, ê°œë³„ì‹œíŠ¸ {len(cleared_sheets)}ê°œ ì´ˆê¸°í™”)",
            "perf_count": len(performance_rows),
            "summary_count": len(summary_rows),
            "cleared_sheets": len(cleared_sheets)
        }

    except Exception as e:
        print(f"ë§ˆì¼€íŒ… í†µí•© ì·¨í•© ì˜¤ë¥˜: {e}")
        return {"success": False, "error": str(e)}


@app.get("/api/marketing/summary")
async def get_marketing_summary(request: Request, refresh: bool = False):
    """ë§ˆì¼€íŒ… ë°ì´í„° ìš”ì•½ (ëŒ€ì‹œë³´ë“œìš©)"""
    require_permission(request, "view")

    try:
        # gsheet ì¸ìŠ¤í„´ìŠ¤ì˜ ìºì‹œ ê¸°ëŠ¥ í™œìš©
        all_values = gsheet.get_values_with_cache("ì „ì²´ë°ì´í„°", sheet_key=MARKETING_SPREADSHEET_KEY, force_refresh=refresh)

        if not all_values or len(all_values) < 2:
            # ì „ì²´ë°ì´í„° ì‹œíŠ¸ê°€ ì—†ìœ¼ë©´ ì›Œí¬ì‹œíŠ¸ ëª©ë¡ ë°˜í™˜ (ìºì‹œ í™œìš©)
            store_names = gsheet.get_worksheet_names_with_cache(sheet_key=MARKETING_SPREADSHEET_KEY, force_refresh=refresh)
            store_names = [name for name in store_names if name not in ["í…œí”Œë¦¿", "ì„¤ì •", "ì „ì²´ë°ì´í„°", "ì‡¼í•‘ëª°ì •ë³´"]]
            return {
                "success": True,
                "stores": store_names,
                "message": "ì „ì²´ë°ì´í„° ì‹œíŠ¸ ì—†ìŒ. ê°œë³„ ìŠ¤í† ì–´ ì¡°íšŒ í•„ìš”"
            }

        headers = all_values[0]
        data = []

        for row in all_values[1:]:
            if not any(row): continue
            row_dict = {header: row[i] if i < len(row) else "" for i, header in enumerate(headers)}
            data.append(row_dict)

        return {"success": True, "data": data, "total": len(data)}

    except Exception as e:
        print(f"ë§ˆì¼€íŒ… ìš”ì•½ ì¡°íšŒ ì˜¤ë¥˜: {e}")
        return {"success": False, "error": str(e)}


@app.get("/api/marketing/history")
async def get_marketing_history(request: Request):
    """ë§ˆì¼€íŒ… ì¼ë³„ ë°©ë¬¸ì ì´ë ¥ ì¡°íšŒ (marketing_stats.json)"""
    require_permission(request, "view")
    
    stats_file = os.path.join(APP_DIR, "marketing_stats.json")
    if not os.path.exists(stats_file):
        return {"success": True, "history": {}}
        
    try:
        with open(stats_file, 'r', encoding='utf-8') as f:
            history = json.load(f)
        return {"success": True, "history": history}
    except Exception as e:
        return {"success": False, "error": str(e)}

@app.get("/api/marketing/performance")
async def get_marketing_performance(request: Request, refresh: bool = False):
    """ì „ì²´ ë§ˆì¼€íŒ… ë°©ë¬¸/ì„±ê³¼ ë¶„ì„ (ì‡¼í•‘ëª°ì •ë³´ ì‹œíŠ¸ í™œìš©)"""
    require_permission(request, "view")

    try:
        # gsheet ì¸ìŠ¤í„´ìŠ¤ì˜ ìºì‹œ ê¸°ëŠ¥ í™œìš©í•˜ì—¬ 'ì‡¼í•‘ëª°ì •ë³´' ì‹œíŠ¸ ë°ì´í„° ë¡œë“œ
        all_values = gsheet.get_values_with_cache("ì‡¼í•‘ëª°ì •ë³´", sheet_key=MARKETING_SPREADSHEET_KEY, force_refresh=refresh)

        if not all_values or len(all_values) < 2:
            return {"success": True, "data": [], "total": 0, "message": "ë°ì´í„°ê°€ ì—†ìŠµë‹ˆë‹¤."}

        headers = all_values[0]
        data = []

        for row in all_values[1:]:
            if not any(row):
                continue

            row_dict = {}
            for i, header in enumerate(headers):
                if i < len(row):
                    row_dict[header] = row[i].strip() if isinstance(row[i], str) else row[i]
                else:
                    row_dict[header] = ""
            data.append(row_dict)

        return {"success": True, "data": data, "total": len(data)}

    except Exception as e:
        print(f"ë§ˆì¼€íŒ… ì„±ê³¼ ì¡°íšŒ ì˜¤ë¥˜: {e}")
        return {"success": False, "error": str(e)}


# ========== ë‹¤ìš´ë¡œë“œ API ==========
@app.get("/api/downloads/client")
async def download_client_api():
    """í´ë¼ì´ì–¸íŠ¸ í”„ë¡œê·¸ë¨ ë‹¤ìš´ë¡œë“œ (exe)"""

    # dist í´ë” ê²½ë¡œ (ìš°ì„ )
    exe_path = APP_DIR / "dist" / "PkonomyClient.exe"

    if not exe_path.exists():
        # ê¸°ì¡´ ê²½ë¡œ fallback
        exe_path = APP_DIR / "pkonomy_client" / "dist" / "PkonomyClient.exe"

    if not exe_path.exists():
        # exeê°€ ì—†ìœ¼ë©´ py íŒŒì¼ ì œê³µ
        py_path = APP_DIR.parent / "client_v1.5.py"
        if py_path.exists():
            return FileResponse(
                path=str(py_path),
                filename="client_v1.5.py",
                media_type="application/octet-stream"
            )
        return {"error": "í´ë¼ì´ì–¸íŠ¸ íŒŒì¼ì„ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤"}

    return FileResponse(
        path=str(exe_path),
        filename="PkonomyClient.exe",
        media_type="application/octet-stream"
    )


@app.get("/api/downloads/extension")
async def download_extension():
    """í¬ë¡¬ ìµìŠ¤í…ì…˜ ë‹¤ìš´ë¡œë“œ (zip)"""
    import zipfile
    from io import BytesIO

    # dist í´ë” ë‚´ chrome_extension ìš°ì„ 
    ext_dir = APP_DIR / "dist" / "chrome_extension"

    if not ext_dir.exists():
        # chrome_ext í´ë” ë‚´ chrome_extension
        ext_dir = APP_DIR / "chrome_ext" / "chrome_extension"

    if not ext_dir.exists():
        # ê¸°ì¡´ ê²½ë¡œ fallback
        ext_dir = APP_DIR.parent / "chrome_extension"

    if not ext_dir.exists():
        return {"error": "í¬ë¡¬ ìµìŠ¤í…ì…˜ í´ë”ë¥¼ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤"}

    # ë©”ëª¨ë¦¬ì— ZIP ìƒì„±
    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        for file_path in ext_dir.rglob('*'):
            if file_path.is_file():
                arcname = file_path.relative_to(ext_dir)
                zf.write(file_path, arcname)

    zip_buffer.seek(0)

    return StreamingResponse(
        zip_buffer,
        media_type="application/zip",
        headers={"Content-Disposition": "attachment; filename=pkonomy_chrome_extension.zip"}
    )


@app.get("/api/downloads/info")
async def get_download_info():
    """ë‹¤ìš´ë¡œë“œ ê°€ëŠ¥ íŒŒì¼ ì •ë³´"""
    # í´ë¼ì´ì–¸íŠ¸ ê²½ë¡œ (dist í´ë” ìš°ì„ )
    exe_path = APP_DIR / "dist" / "PkonomyClient.exe"
    if not exe_path.exists():
        exe_path = APP_DIR / "pkonomy_client" / "dist" / "PkonomyClient.exe"

    # ìµìŠ¤í…ì…˜ ê²½ë¡œ (dist í´ë” ìš°ì„ )
    ext_dir = APP_DIR / "dist" / "chrome_extension"
    if not ext_dir.exists():
        ext_dir = APP_DIR / "chrome_ext" / "chrome_extension"
    if not ext_dir.exists():
        ext_dir = APP_DIR.parent / "chrome_extension"

    info = {
        "client": {
            "available": exe_path.exists(),
            "filename": "PkonomyClient.exe",
            "path": str(exe_path) if exe_path.exists() else None
        },
        "extension": {
            "available": ext_dir.exists(),
            "filename": "pkonomy_chrome_extension.zip",
            "path": str(ext_dir) if ext_dir.exists() else None
        }
    }

    # íŒŒì¼ í¬ê¸° ë° ìˆ˜ì •ì¼
    if exe_path.exists():
        stat = exe_path.stat()
        info["client"]["size"] = stat.st_size
        info["client"]["modified"] = datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M")

    if ext_dir.exists():
        manifest_path = ext_dir / "manifest.json"
        if manifest_path.exists():
            import json
            with open(manifest_path, 'r', encoding='utf-8') as f:
                manifest = json.load(f)
                info["extension"]["version"] = manifest.get("version", "1.0")

    return info


# ========== ë¶ˆì‚¬ì ëŒ€ì‹œë³´ë“œ API ==========
@app.post("/api/bulsaja/settings")
async def update_bulsaja_settings(request: Request):
    """ë¶ˆì‚¬ì ëŒ€ì‹œë³´ë“œ ê°œë³„ ì„¤ì • ì €ì¥ (ìš´ì˜ì¼ ë“±)"""
    get_current_user(request)
    try:
        data = await request.json()
        store_name = data.get("store_name")
        operation_days = data.get("operationDays")
        
        if not store_name:
            return {"success": False, "message": "ìŠ¤í† ì–´ëª…ì´ ì—†ìŠµë‹ˆë‹¤."}
            
        settings_path = os.path.join(APP_DIR, "bulsaja_settings.json")
        settings = {}
        if os.path.exists(settings_path):
            with open(settings_path, "r", encoding="utf-8") as f:
                settings = json.load(f)
        
        if store_name not in settings:
            settings[store_name] = {}
        
        if operation_days is not None:
            settings[store_name]["operationDays"] = int(operation_days)
            
        with open(settings_path, "w", encoding="utf-8") as f:
            json.dump(settings, f, indent=4, ensure_ascii=False)
            
        return {"success": True, "message": "ì„¤ì •ì´ ì €ì¥ë˜ì—ˆìŠµë‹ˆë‹¤."}
    except Exception as e:
        return {"success": False, "message": str(e)}


@app.get("/api/bulsaja/dashboard_data")
async def get_bulsaja_dashboard_data(request: Request, refresh: bool = False):
    """ë¶ˆì‚¬ì ëŒ€ì‹œë³´ë“œ ë°ì´í„° ì¡°íšŒ (ë§¤ì¶œ + ë§ˆì¼€íŒ… í†µí•©)"""
    get_current_user(request)

    try:
        # ìƒí’ˆìˆ˜ ì •ë³´
        ss_counts = {}
        st_counts = {}
        try:
            ws_counts = gsheet.sheet.worksheet("ë“±ë¡ê°¯ìˆ˜")
            c_data = ws_counts.get_all_values()
            if c_data and len(c_data) > 1:
                h = c_data[0]
                ni = next((i for i,x in enumerate(h) if x=="ìŠ¤í† ì–´ëª…"), None)
                ci = next((i for i,x in enumerate(h) if x=="íŒë§¤ì¤‘"), None)
                if ni is not None and ci is not None:
                    for r in c_data[1:]:
                        if len(r) > max(ni, ci):
                            ss_counts[r[ni].strip()] = int(r[ci]) if r[ci].isdigit() else 0
        except: pass
        
        try:
            ws_11 = gsheet.sheet.worksheet("11ë²ˆê°€")
            s_data = ws_11.get_all_values()
            if s_data and len(s_data) > 1:
                h = s_data[0]
                ni = next((i for i,x in enumerate(h) if x in ["store_name", "ì‡¼í•‘ëª° ë³„ì¹­", "ìŠ¤í† ì–´ëª…"]), None)
                ci = next((i for i,x in enumerate(h) if x in ["on_sale", "íŒë§¤ì¤‘"]), None)
                if ni is not None and ci is not None:
                    for r in s_data[1:]:
                        if len(r) > max(ni, ci):
                            st_counts[r[ni].strip()] = int(r[ci]) if r[ci].isdigit() else 0
        except: pass

        accounts = gsheet.get_accounts(None, force_refresh=refresh)
        sales_res = await get_sales_from_sheet_v2(request, force=refresh)
        sales_map = sales_res.get("data", {}) if sales_res.get("success") else {}
        marketing_res = await get_marketing_data(request, refresh=refresh)
        marketing_map = marketing_res.get("data", {}) if marketing_res.get("success") else {}

        for acc in accounts:
            store_name = (acc.get("ìŠ¤í† ì–´ëª…") or "").strip()
            raw_platform = (acc.get("platform") or "").strip().lower()
            
            # í•„ìˆ˜ í•„ë“œ ì´ˆê¸°í™”
            acc["name"] = store_name or acc.get("login_id") or "Unknown"
            acc["revenue"] = 0
            acc["month_sales"] = 0
            acc["today_sales"] = 0
            acc["month_orders"] = 0
            acc["month_profit"] = 0
            acc["orders_7d"] = 0
            acc["operationDays"] = 30
            acc["products"] = 0
            acc["targetProducts"] = 10000
            acc["targetRevenue"] = 2000000
            acc["stage"] = "ì—…ë¡œë“œ"  # ê¸°ë³¸ê°’: ì—…ë¡œë“œ (ë¬¸ìì—´!)
            acc["month_visitors"] = 0
            acc["renewalReason"] = ""
            # ìš©ë„ í•„ë“œ (ì›ë³¸ì—ì„œ ê°€ì ¸ì˜¤ê¸°, ê¸°ë³¸ê°’ "ëŒ€ëŸ‰")
            acc["usage"] = acc.get("ìš©ë„") or acc.get("usage") or "ëŒ€ëŸ‰"

            # Platform Normalization
            if raw_platform in ["naver", "smartstore", "ìŠ¤ë§ˆíŠ¸ìŠ¤í† ì–´", "n"]:
                pk = "ìŠ¤ë§ˆíŠ¸ìŠ¤í† ì–´"
                platform_display = "naver"  # HTMLìš©
            elif raw_platform in ["11st", "11ë²ˆê°€", "11", "elevenst"]:
                pk = "11ë²ˆê°€"
                platform_display = "11st"  # HTMLìš©
            elif raw_platform in ["coupang", "ì¿ íŒ¡", "cp", "c"]:
                pk = "ì¿ íŒ¡"
                platform_display = "coupang"  # HTMLìš©
            elif raw_platform in ["gmarket", "ì§€ë§ˆì¼“", "gm", "g"]:
                pk = "ì§€ë§ˆì¼“"
                platform_display = "gmarket"  # HTMLìš© (Gë§ˆì¼“/ì˜¥ì…˜ í•„í„°ì— í¬í•¨)
            elif raw_platform in ["auction", "ì˜¥ì…˜", "ac", "a"]:
                pk = "ì˜¥ì…˜"
                platform_display = "auction"  # HTMLìš© (A ì•„ì´ì½˜ í‘œì‹œ, Gë§ˆì¼“/ì˜¥ì…˜ í•„í„°ì—ì„œ í•¨ê»˜ ì²˜ë¦¬)
            else:
                pk = raw_platform
                platform_display = raw_platform

            # Alias ë° ìˆ˜ë™ ì„¤ì •ê°’ ë¡œë“œ
            an = store_name
            manual_op_days = None  # ìˆ˜ë™ ì„¤ì •ëœ ìš´ì˜ì¼ìˆ˜
            s_p = os.path.join(os.path.dirname(os.path.abspath(__file__)), "bulsaja_settings.json")
            if os.path.exists(s_p):
                try:
                    with open(s_p, "r", encoding="utf-8") as sf:
                        all_settings = json.load(sf)
                        # sales_key_alias ì²˜ë¦¬
                        al = all_settings.get("sales_key_alias", {})
                        if store_name in al:
                            an = al[store_name]
                        # ìŠ¤í† ì–´ë³„ ìˆ˜ë™ ì„¤ì •ê°’ (operationDays)
                        store_settings = all_settings.get(store_name, {})
                        if "operationDays" in store_settings:
                            manual_op_days = int(store_settings["operationDays"])
                except: pass

            # HTMLìš© í”Œë«í¼ ê°’ ì„¤ì •
            acc["platform"] = platform_display

            sk = f"{an}({pk})"
            info = {}
            matched = False
            
            # ë§¤ì¶œ ë°ì´í„° ë§¤ì¹­
            if sk in sales_map:
                info = sales_map[sk]
                matched = True
            else:
                for k, v in sales_map.items():
                    if "(" in k:
                        kn = k.split("(")[0]
                        kp = k.split("(")[1].replace(")", "")
                        if kp == pk and (an in kn or kn in an):
                            info = v
                            matched = True
                            break
            
            if not matched:
                ss = an.replace(" ", "")
                for k, v in sales_map.items():
                    if ss in k.replace(" ", "") and pk in k:
                        info = v
                        matched = True
                        break
            
            if matched:
                acc["revenue"] = info.get("month_sales", 0)
                acc["month_sales"] = info.get("month_sales", 0)
                acc["today_sales"] = info.get("today_sales", 0)
                acc["month_orders"] = info.get("month_orders", 0)
                acc["month_profit"] = info.get("month_profit", 0)
                acc["operationDays"] = info.get("operation_days", 30)
                acc["orders_7d"] = info.get("orders_7d", 0)

            # ìˆ˜ë™ ì„¤ì •ëœ ìš´ì˜ì¼ìˆ˜ê°€ ìˆìœ¼ë©´ ë®ì–´ì“°ê¸°
            if manual_op_days is not None:
                acc["operationDays"] = manual_op_days

            # Marketing
            md = marketing_map.get(store_name, {})
            visitors = 0
            biz = md.get("biz_advisor", [])
            if isinstance(biz, list):
                for b in biz:
                    try:
                        v = str(b.get("ìœ ì…ìˆ˜", "0")).replace(",", "")
                        visitors += int(v) if v.isdigit() else 0
                    except: pass
            acc["month_visitors"] = visitors

            # Products & Stage
            pc = 0
            if pk == "ìŠ¤ë§ˆíŠ¸ìŠ¤í† ì–´":
                pc = ss_counts.get(store_name, 0)
                if pc == 0 and "_" in store_name:
                    pc = ss_counts.get(store_name.split("_", 1)[1], 0)
            elif pk == "11ë²ˆê°€":
                pc = st_counts.get(store_name, 0) or st_counts.get(acc.get("login_id"), 0)
            
            tp = 10000
            if pk == "11ë²ˆê°€":
                tp = 5000
            elif pk in ["ì§€ë§ˆì¼“", "ì˜¥ì…˜"]:
                tp = 2000
            
            acc["products"] = pc
            acc["targetProducts"] = tp
            
            # ìƒí’ˆ ë“±ë¡ë¥  ê³„ì‚°
            registration_rate = (pc / tp) if tp > 0 else 0
            
            # ìš´ì˜ì¼ ê°€ì ¸ì˜¤ê¸° (ë§¤ì¹­ëœ ê²½ìš°ì—ë§Œ, ê¸°ë³¸ê°’ 30ì¼)
            operation_days = acc.get("operationDays", 30)
            
            # ë§¤ì¶œ ë°ì´í„°
            month_revenue = acc.get("revenue", 0)
            orders_7d = acc.get("orders_7d", 0)
            
            # ìŠ¤í† ì–´ ìƒíƒœ íŒë³„
            # 1ë‹¨ê³„: ì—…ë¡œë“œ vs ìš´ì˜ êµ¬ë¶„
            if registration_rate >= 0.9:
                # 90% ì´ìƒ ë“±ë¡ -> ìš´ì˜ ìƒíƒœ
                stage = "ìš´ì˜"  # ìš´ì˜ (ë¬¸ìì—´!)
                
                # 2ë‹¨ê³„: ë¦¬ë‰´ì–¼ëŒ€ìƒ íŒë³„
                # ì¡°ê±´1: ìš´ì˜ì¼ 60ì¼ ê²½ê³¼
                if operation_days >= 60:
                    stage = "ë¦¬ë‰´ì–¼ëŒ€ìƒ"  # ë¦¬ë‰´ì–¼ëŒ€ìƒ (ë¬¸ìì—´!)
                    acc["renewalReason"] = f"ìš´ì˜ {operation_days}ì¼ ê²½ê³¼"
                
                # ì¡°ê±´2: ìš´ì˜ì¼ 30ì¼ ê²½ê³¼ + ìµœê·¼ 30ì¼ ë§¤ì¶œ 50ë§Œì› ì´í•˜
                elif operation_days >= 30 and month_revenue <= 500000:
                    stage = "ë¦¬ë‰´ì–¼ëŒ€ìƒ"  # ë¦¬ë‰´ì–¼ëŒ€ìƒ (ë¬¸ìì—´!)
                    acc["renewalReason"] = f"ë§¤ì¶œë¶€ì§„ ({month_revenue:,}ì›)"
                
                # ì¡°ê±´3: 7ì¼ ì£¼ë¬¸ 0ê±´ (ìœ ì…ìˆ˜ ê°ì†ŒëŠ” í–¥í›„ ì¶”ê°€)
                elif orders_7d == 0 and operation_days >= 7:
                    stage = "ë¦¬ë‰´ì–¼ëŒ€ìƒ"  # ë¦¬ë‰´ì–¼ëŒ€ìƒ (ë¬¸ìì—´!)
                    acc["renewalReason"] = "7ì¼ ì£¼ë¬¸ 0ê±´"
            else:
                # 90% ë¯¸ë§Œ ë“±ë¡ -> ì—…ë¡œë“œ ìƒíƒœ
                stage = "ì—…ë¡œë“œ"  # ì—…ë¡œë“œ (ë¬¸ìì—´!)
                # ì—…ë¡œë“œ ìƒíƒœì—ì„œëŠ” ìš´ì˜ì¼ ì´ˆê¸°í™” (í–¥í›„ DB ì—…ë°ì´íŠ¸ ì‹œ ë°˜ì˜)
            
            acc["stage"] = stage

        return {"success": True, "accounts": accounts}
    except Exception as e:
        print(f"[ë¶ˆì‚¬ì ì˜¤ë¥˜] {e}")
        import traceback
        traceback.print_exc()
        return {"success": False, "message": str(e), "accounts": []}


# ========== ì‹¤í–‰ ì œì–´ ë³€ìˆ˜ ==========
HOST = "0.0.0.0"
PORT = 8000

# ========== ì‹¤í–‰ë¶€ ì œê±° (run_server.py ì‚¬ìš©) ==========

# ì „ì—­ ì„œë²„ ì¸ìŠ¤í„´ìŠ¤ ìƒì„± (run_server.pyì—ì„œ server:appìœ¼ë¡œ ì‹¤í–‰ ì‹œ í•„ìš”)
# API í•¸ë“¤ëŸ¬(sync_manual_marketing_data ë“±)ì—ì„œ 'server' ì „ì—­ ë³€ìˆ˜ë¥¼ ì°¸ì¡°í•˜ê¸° ìœ„í•¨
try:
    server = Server()
    app.state.server = server
    print("[INIT] Server instance created globally in server.py")
except Exception as e:
    print(f"[ERROR] Failed to create global Server instance: {e}")
