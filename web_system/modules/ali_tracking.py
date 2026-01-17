#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
알리 송장번호 수집 모듈
- Chrome 디버그 모드 연결
- 알리익스프레스 송장 조회
- 구글 시트 업데이트
"""

import os
import re
import time
import socket
import subprocess
import threading
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional

import gspread
from google.oauth2.service_account import Credentials


class AliTrackingCollector:
    """알리 송장번호 수집기"""
    
    def __init__(self, credentials_path: str = None):
        # 기본 경로: 상위 폴더(web_system)의 JSON 파일
        default_path = str(Path(__file__).resolve().parent.parent / "autosms-466614-951e91617c69.json")
        self.credentials_path = credentials_path or default_path
        self.browser = None
        self.playwright = None
        self.debug_port = None
        
        self.status = {
            "running": False,
            "connected": False,
            "progress": 0,
            "total": 0,
            "completed": 0,
            "current": "",
            "logs": [],
            "collected": [],
            "message": ""
        }
    
    def add_log(self, msg: str, log_status: str = "info"):
        """로그 추가"""
        self.status["logs"].append({
            "time": datetime.now().strftime("%H:%M:%S"),
            "msg": msg,
            "status": log_status
        })
        if len(self.status["logs"]) > 100:
            self.status["logs"] = self.status["logs"][-100:]
        print(f"[알리] {msg}")
    
    def find_available_port(self, start: int = 9300, end: int = 9320) -> int:
        """사용 가능한 포트 찾기"""
        for port in range(start, end):
            try:
                sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                sock.settimeout(1)
                result = sock.connect_ex(('127.0.0.1', port))
                sock.close()
                if result != 0:
                    print(f"[알리] 사용 가능한 포트 발견: {port}")
                    return port
            except:
                continue
        return start
    
    def find_existing_chrome(self) -> Optional[int]:
        """기존 Chrome 디버그 포트 찾기"""
        for port in range(9300, 9320):
            try:
                sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                sock.settimeout(1)
                result = sock.connect_ex(('127.0.0.1', port))
                sock.close()
                if result == 0:
                    print(f"[알리] 기존 Chrome 감지: 포트 {port}")
                    return port
            except:
                continue
        return None
    
    async def connect_browser(self, app_dir: str, requested_port: int = 9300) -> Dict:
        """Chrome 브라우저 연결"""
        from playwright.async_api import async_playwright
        import shutil
        
        try:
            # 기존 Chrome 확인
            existing_port = self.find_existing_chrome()
            
            if existing_port:
                try:
                    self.playwright = await async_playwright().start()
                    self.browser = await self.playwright.chromium.connect_over_cdp(f"http://localhost:{existing_port}")
                    self.debug_port = existing_port
                    self.status["connected"] = True
                    self.add_log(f"기존 Chrome 연결 성공 (포트: {existing_port})")
                    return {"success": True, "message": f"기존 Chrome 연결 성공 (포트: {existing_port})"}
                except Exception as e:
                    print(f"[알리] 기존 Chrome 연결 실패: {e}")
            
            # 새 Chrome 실행
            available_port = self.find_available_port(9300, 9320)
            
            chrome_paths = [
                r"C:\Program Files\Google\Chrome\Application\chrome.exe",
                r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
                os.path.expandvars(r"%LOCALAPPDATA%\Google\Chrome\Application\chrome.exe"),
            ]
            
            chrome_path = None
            for path in chrome_paths:
                if os.path.exists(path):
                    chrome_path = path
                    break
            
            if not chrome_path:
                chrome_path = shutil.which("chrome") or shutil.which("google-chrome")
            
            if not chrome_path:
                return {"success": False, "message": "Chrome을 찾을 수 없습니다"}
            
            ali_profile_dir = os.path.join(app_dir, "chrome_ali_profile")
            
            cmd = [
                chrome_path,
                f"--remote-debugging-port={available_port}",
                f"--user-data-dir={ali_profile_dir}",
                "--no-first-run",
                "--no-default-browser-check",
                "https://www.aliexpress.com"
            ]
            
            print(f"[알리] Chrome 실행 (포트 {available_port}): {' '.join(cmd)}")
            subprocess.Popen(cmd)
            
            import asyncio
            await asyncio.sleep(4)
            
            self.playwright = await async_playwright().start()
            
            for retry in range(5):
                try:
                    self.browser = await self.playwright.chromium.connect_over_cdp(f"http://localhost:{available_port}")
                    break
                except Exception as e:
                    if retry < 4:
                        print(f"[알리] 연결 대기중... ({retry+1}/5)")
                        await asyncio.sleep(2)
                    else:
                        raise e
            
            self.status["connected"] = True
            self.debug_port = available_port
            self.add_log(f"브라우저 연결 성공 (포트: {available_port})")
            
            return {"success": True, "message": f"브라우저 실행 및 연결 성공 (포트: {available_port}). 알리익스프레스 로그인 후 '수집 시작' 클릭"}
            
        except Exception as e:
            self.add_log(f"연결 실패: {e}", "error")
            return {"success": False, "message": f"연결 실패: {e}"}
    
    def get_carrier_name(self, tracking_no: str) -> str:
        """송장번호로 택배사 구분"""
        tracking_no = str(tracking_no).strip()
        if tracking_no.startswith('50') or tracking_no.startswith('56'):
            return 'CJ대한통운'
        elif tracking_no.startswith('51'):
            return '한진택배'
        elif tracking_no.startswith('9'):
            return '투데이'
        elif tracking_no.startswith('52'):
            return '경동'
        elif tracking_no.startswith('54'):
            return '로젠'
        return '확인필요'
    
    def run_collection(self, sheet_id: str, month: str):
        """송장번호 수집 실행 (동기 - 스레드에서 실행)"""
        from playwright.sync_api import sync_playwright
        
        try:
            self.add_log(f"시트 ID: {sheet_id}, 월: {month}")
            
            if not os.path.exists(self.credentials_path):
                self.add_log(f"인증 파일 없음: {self.credentials_path}", "error")
                self.status["running"] = False
                return
            
            # 시트 연결
            scopes = [
                'https://www.googleapis.com/auth/spreadsheets',
                'https://www.googleapis.com/auth/drive'
            ]
            creds = Credentials.from_service_account_file(self.credentials_path, scopes=scopes)
            ali_client = gspread.authorize(creds)
            
            spreadsheet = ali_client.open_by_key(sheet_id)
            sheet = spreadsheet.worksheet(month)
            self.add_log(f"시트 연결 성공: {spreadsheet.title} / {month}")
            
            # 열 위치
            COL_PLATFORM = 30  # AD
            COL_ORDER_ID = 31  # AE
            COL_CARRIER = 43   # AQ
            COL_TRACKING = 44  # AR
            START_ROW = 3
            
            all_data = sheet.get_all_values()
            self.add_log(f"전체 데이터 행 수: {len(all_data)}")
            
            # 대상 찾기
            targets = []
            for idx, row in enumerate(all_data):
                row_num = idx + 1
                if row_num < START_ROW:
                    continue
                
                platform = row[COL_PLATFORM - 1] if len(row) >= COL_PLATFORM else ""
                order_id = row[COL_ORDER_ID - 1] if len(row) >= COL_ORDER_ID else ""
                tracking = row[COL_TRACKING - 1] if len(row) >= COL_TRACKING else ""
                
                if platform.strip() == "알리" and order_id.strip() and not tracking.strip():
                    targets.append({'row': row_num, 'order_id': order_id.strip()})
            
            self.add_log(f"조회 대상: {len(targets)}건")
            self.status["total"] = len(targets)
            
            if not targets:
                self.add_log("조회할 건이 없습니다", "success")
                self.status["running"] = False
                self.status["message"] = "완료 - 조회할 건 없음"
                return
            
            if not self.debug_port:
                self.add_log("Chrome 디버그 포트를 찾을 수 없습니다. 브라우저를 다시 연결하세요.", "error")
                self.status["running"] = False
                return
            
            self.add_log(f"Chrome 연결 중 (포트: {self.debug_port})...")
            
            updated = 0
            
            with sync_playwright() as p:
                try:
                    sync_browser = p.chromium.connect_over_cdp(f"http://localhost:{self.debug_port}")
                    self.add_log("sync_playwright 연결 성공")
                except Exception as e:
                    self.add_log(f"Chrome 연결 실패: {e}", "error")
                    self.status["running"] = False
                    return
                
                contexts = sync_browser.contexts
                if not contexts:
                    self.add_log("브라우저 컨텍스트가 없습니다", "error")
                    self.status["running"] = False
                    return
                
                context = contexts[0]
                pages = context.pages
                if not pages:
                    self.add_log("열린 페이지가 없습니다", "error")
                    self.status["running"] = False
                    return
                
                ali_page = pages[0]
                self.add_log(f"페이지 연결 완료: {ali_page.url[:50]}...")
                
                for i, target in enumerate(targets):
                    if not self.status["running"]:
                        self.add_log("사용자에 의해 중단됨", "info")
                        break
                    
                    self.status["completed"] = i
                    self.status["progress"] = int((i / len(targets)) * 100)
                    
                    self.add_log(f"[{i+1}/{len(targets)}] 주문번호 {target['order_id']} 조회 중...")
                    
                    tracking_no = None
                    try:
                        url = f"https://www.aliexpress.com/p/tracking/index.html?_addShare=no&_login=yes&tradeOrderId={target['order_id']}"
                        
                        ali_page.goto(url, timeout=20000, wait_until="domcontentloaded")
                        time.sleep(2)
                        
                        page_text = ali_page.inner_text("body", timeout=10000)
                        
                        # 송장번호 추출
                        match = re.search(r'운송장\s*번호[:\s]*(\d+)', page_text)
                        if match:
                            tracking_no = match.group(1)
                        else:
                            match = re.search(r'Tracking\s*(?:number|no)[:\s]*(\d+)', page_text, re.IGNORECASE)
                            if match:
                                tracking_no = match.group(1)
                            else:
                                match = re.search(r'(\d{10,14})', page_text)
                                if match:
                                    potential = match.group(1)
                                    if potential[:2] in ['50', '51', '52', '54', '56'] or potential.startswith('9'):
                                        tracking_no = potential
                    except Exception as e:
                        self.add_log(f"조회 오류: {e}", "error")
                    
                    if tracking_no:
                        carrier = self.get_carrier_name(tracking_no)
                        
                        delivery_completed = False
                        if carrier == '투데이':
                            try:
                                completed_el = ali_page.locator('.logistic-info-v2--nodeTitle--2rejjVx:has-text("배송 완료"), .logistic-info-v2--nodeTitle--2rejjVx:has-text("배송완료")')
                                if completed_el.count() > 0:
                                    delivery_completed = True
                                    self.add_log(f"→ 송장번호: {tracking_no} ({carrier}) [배송완료]", "success")
                                else:
                                    self.add_log(f"→ 송장번호: {tracking_no} ({carrier})", "success")
                            except:
                                self.add_log(f"→ 송장번호: {tracking_no} ({carrier})", "success")
                        else:
                            self.add_log(f"→ 송장번호: {tracking_no} ({carrier})", "success")
                        
                        self.status["collected"].append({
                            "customer_order": target.get('customer_order', ''),
                            "order_id": target['order_id'],
                            "carrier": carrier,
                            "tracking_no": tracking_no,
                            "delivery_completed": delivery_completed
                        })
                        
                        try:
                            sheet.update_cell(target['row'], COL_CARRIER, carrier)
                            sheet.update_cell(target['row'], COL_TRACKING, tracking_no)
                            updated += 1
                            self.add_log(f"→ 행 {target['row']} 업데이트 완료")
                            
                            if delivery_completed:
                                try:
                                    sheet.format(f"AQ{target['row']}:AR{target['row']}", {
                                        "backgroundColor": {"red": 1, "green": 1, "blue": 0}
                                    })
                                    self.add_log(f"→ 행 {target['row']} 노란색 색칠 (배송완료)")
                                except Exception as e:
                                    self.add_log(f"→ 색칠 오류: {e}", "error")
                            
                            time.sleep(1)
                        except Exception as e:
                            self.add_log(f"→ 시트 업데이트 오류: {e}", "error")
                    else:
                        self.add_log("→ 송장번호 없음")
                    
                    time.sleep(2)
                
                self.add_log(f"완료! {updated}건 업데이트", "success")
                self.status["completed"] = len(targets)
                self.status["progress"] = 100
                self.status["message"] = f"완료! {updated}건 업데이트"
            
        except Exception as e:
            import traceback
            self.add_log(f"오류: {e}", "error")
            traceback.print_exc()
        finally:
            self.status["running"] = False
            # 작업이 끝나면 브라우저 닫기 (리소스 해제)
            try:
                if 'sync_browser' in locals() and sync_browser:
                    sync_browser.close()
                    print("[알리] 브라우저 종료 완료")
            except:
                pass
    
    def start_collection(self, sheet_url: str, month: str) -> Dict:
        """수집 시작"""
        if not self.browser or not self.status.get("connected"):
            return {"success": False, "message": "브라우저가 연결되지 않았습니다. '브라우저 연결'을 먼저 클릭하세요."}
        
        if self.status["running"]:
            return {"success": False, "message": "이미 수집 중입니다."}
        
        # 시트 ID 추출
        sheet_id = sheet_url
        if 'docs.google.com' in sheet_id:
            match = re.search(r'/d/([a-zA-Z0-9_-]+)', sheet_id)
            if match:
                sheet_id = match.group(1)
        
        # 상태 초기화
        self.status = {
            "running": True,
            "connected": True,
            "progress": 0,
            "total": 0,
            "completed": 0,
            "logs": [],
            "message": "시작 중...",
            "collected": []
        }
        
        # 백그라운드 스레드로 실행
        thread = threading.Thread(target=self.run_collection, args=(sheet_id, month), daemon=True)
        thread.start()
        
        return {"success": True, "message": "수집 시작됨"}
    
    def stop_collection(self):
        """수집 중지"""
        self.status["running"] = False
        return {"success": True}
    
    def get_progress(self) -> Dict:
        """진행상황 조회"""
        return self.status
    
    async def close_all(self):
        """브라우저 및 Playwright 종료 (서버 종료 시)"""
        try:
            if self.browser:
                await self.browser.close()
            if self.playwright:
                await self.playwright.stop()
            self.status["connected"] = False
            self.browser = None
            self.playwright = None
            print("[알리] 브라우저 종료 및 리소스 해제 완료")
        except Exception as e:
            print(f"[알리] 브라우저 종료 중 오류: {e}")

    def generate_excel(self) -> Optional[str]:
        """수집된 데이터 Excel 생성"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        
        collected = self.status.get("collected", [])
        
        if not collected:
            return None
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        headers = ['고객 주문일', '고객 주문 번호', '해외 주문일', '해외 주문 번호', 
                   '해외 택배사', '해외 운송장번호', '국내 택배사', '국내 운송장번호']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        for row_idx, item in enumerate(collected, 2):
            ws.cell(row=row_idx, column=2, value=item.get("customer_order", ""))
            ws.cell(row=row_idx, column=4, value=item.get("order_id", ""))
            ws.cell(row=row_idx, column=7, value=item.get("carrier", ""))
            ws.cell(row=row_idx, column=8, value=item.get("tracking_no", ""))
        
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 20
        
        today = datetime.now().strftime('%y%m%d')
        filename = f"송장_번호_다운로드_{today}.xlsx"
        filepath = Path(tempfile.gettempdir()) / filename
        wb.save(filepath)
        
        return str(filepath)


# 전역 인스턴스 (server.py에서 사용)
ali_collector = AliTrackingCollector()
