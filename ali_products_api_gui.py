#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
스마트스토어 API 알리 상품 수집 및 분석 GUI
- 용도="대량" 계정 자동 로드
- 날짜 범위 선택하여 주문 데이터 수집
- 피벗 테이블 형식으로 집계
"""

import os
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime, timedelta
from collections import defaultdict
from dotenv import load_dotenv
import threading
import requests

# Google Sheets
import gspread
from google.oauth2.service_account import Credentials

# ========== 설정 ==========
load_dotenv()

CREDENTIALS_FILE = os.environ.get("SERVICE_ACCOUNT_JSON", r"C:\autosystem\web_system\autosms-466614-951e91617c69.json")
SPREADSHEET_KEY = os.environ.get("SPREADSHEET_KEY", "1r-ROJ7ksv6qOtOTXbkrprxu17EQmbO-n1J1pm_N5Hh8")
ACCOUNTS_TAB = "계정목록"

# ========== 네이버 토큰 발급 ==========
def get_naver_token(client_id: str, client_secret: str) -> str:
    """네이버 커머스 API 토큰 발급"""
    import time
    import hmac
    import hashlib
    import base64
    
    timestamp = str(int(time.time() * 1000))
    password = f"{client_id}_{timestamp}"
    signature = base64.b64encode(
        hmac.new(client_secret.encode('utf-8'), password.encode('utf-8'), hashlib.sha256).digest()
    ).decode('utf-8')
    
    url = "https://api.commerce.naver.com/external/v1/oauth2/token"
    data = {
        "client_id": client_id,
        "timestamp": timestamp,
        "client_secret_sign": signature,
        "grant_type": "client_credentials",
        "type": "SELF"
    }
    
    resp = requests.post(url, json=data)
    if resp.status_code != 200:
        raise Exception(f"토큰 발급 실패: {resp.status_code} {resp.text}")
    
    return resp.json()["access_token"]

# ========== 주문 데이터 수집 ==========
def collect_orders_from_api(accounts, start_date, end_date, progress_callback=None):
    """
    네이버 커머스 API로 주문 데이터 수집
    
    Args:
        accounts: [{"store_name": "xxx", "client_id": "xxx", "client_secret": "xxx"}, ...]
        start_date: datetime
        end_date: datetime
        progress_callback: 진행률 콜백
    
    Returns:
        (result_list, error_message)
    """
    
    def update_progress(msg):
        if progress_callback:
            progress_callback(msg)
    
    try:
        all_orders = []
        
        for idx, account in enumerate(accounts, 1):
            store_name = account["store_name"]
            client_id = account["client_id"]
            client_secret = account["client_secret"]
            
            update_progress(f"[{idx}/{len(accounts)}] {store_name} - 토큰 발급 중...")
            
            try:
                token = get_naver_token(client_id, client_secret)
            except Exception as e:
                update_progress(f"[{idx}/{len(accounts)}] {store_name} - 토큰 발급 실패: {e}")
                continue
            
            headers = {
                "Authorization": f"Bearer {token}",
                "Content-Type": "application/json"
            }
            
            update_progress(f"[{idx}/{len(accounts)}] {store_name} - 주문 조회 중...")
            
            # 주문 조회
            base_url = "https://api.commerce.naver.com/external"
            
            start_str = start_date.strftime("%Y-%m-%d")
            end_str = end_date.strftime("%Y-%m-%d")
            
            # 정상 주문만 조회 (server.py와 동일)
            body = {
                "productOrderStatuses": ["PAYED", "DELIVERING", "DELIVERED", "PURCHASE_DECIDED"],
                "startPayedDate": f"{start_str}T00:00:00",
                "endPayedDate": f"{end_str}T23:59:59"
            }
            
            try:
                resp = requests.post(
                    f"{base_url}/v1/pay-order/seller/product-orders/search",
                    headers=headers,
                    json=body,
                    timeout=30
                )
                
                print(f"[API 응답] {store_name}: status={resp.status_code}")
                
                if resp.status_code == 200:
                    response_json = resp.json()
                    orders = response_json.get("data", [])
                    
                    print(f"[API 응답] {store_name}: {len(orders)}건 수집")
                    if len(orders) > 0:
                        print(f"[API 응답] {store_name}: 첫 번째 주문 샘플 = {orders[0].get('productName', 'N/A')}")
                    
                    update_progress(f"[{idx}/{len(accounts)}] {store_name} - {len(orders)}건 수집")
                    
                    # 스토어명 추가
                    for order in orders:
                        order["_store_name"] = store_name
                    
                    all_orders.extend(orders)
                else:
                    error_msg = resp.text[:200] if resp.text else "응답 없음"
                    print(f"[API 오류] {store_name}: {resp.status_code} - {error_msg}")
                    update_progress(f"[{idx}/{len(accounts)}] {store_name} - API 오류: {resp.status_code}")
                    
            except Exception as e:
                update_progress(f"[{idx}/{len(accounts)}] {store_name} - 주문 조회 오류: {e}")
        
        if not all_orders:
            return None, "수집된 주문이 없습니다"
        
        print(f"\n[전체 수집 완료] 총 {len(all_orders)}건 주문 수집")
        print(f"[전체 수집 완료] 기간: {start_date.strftime('%Y-%m-%d')} ~ {end_date.strftime('%Y-%m-%d')}")
        
        update_progress(f"총 {len(all_orders)}건 수집 완료 - 피벗 집계 중...")
        
        # 피벗 집계 (상품명 + 옵션명 기준)
        pivot_data = defaultdict(lambda: {
            "판매자상품코드": "",
            "스토어명": set(),
            "판매건수": 0,
            "총수량": 0,
            "총금액": 0,
            "클레임건수": 0
        })
        
        for order in all_orders:
            # 상품 정보
            product_name = order.get("productName", "").strip()
            option_name = order.get("productOption", "").strip()
            seller_code = order.get("sellerProductCode", "").strip()
            store_name = order.get("_store_name", "")
            
            # 수량 및 금액
            quantity = int(order.get("quantity", 1))
            total_amount = int(order.get("totalPaymentAmount", 0))
            
            # 클레임 여부 체크
            claim_status = order.get("claimStatus", "")
            product_order_status = order.get("productOrderStatus", "")
            
            is_claim = (
                claim_status in ["CANCEL", "RETURN", "EXCHANGE"] or
                product_order_status in ["CANCELED", "RETURNED", "EXCHANGED", "CANCELED_BY_NOPAYMENT"]
            )
            
            # 키 생성
            key = f"{product_name}|||{option_name}"
            
            # 집계
            pivot_data[key]["판매자상품코드"] = seller_code
            pivot_data[key]["스토어명"].add(store_name)
            pivot_data[key]["판매건수"] += 1
            pivot_data[key]["총수량"] += quantity
            pivot_data[key]["총금액"] += total_amount
            
            if is_claim:
                pivot_data[key]["클레임건수"] += 1
        
        # 결과 리스트로 변환
        result = []
        for key, data in pivot_data.items():
            product_name, option_name = key.split("|||")
            result.append({
                "상품명": product_name,
                "옵션명": option_name,
                "판매자상품코드": data["판매자상품코드"],
                "스토어명": ", ".join(sorted(data["스토어명"])),
                "판매건수": data["판매건수"],
                "총수량": data["총수량"],
                "총금액": data["총금액"],
                "정상건수": data["판매건수"] - data["클레임건수"],
                "클레임건수": data["클레임건수"]
            })
        
        # 판매건수 기준 내림차순 정렬
        result.sort(key=lambda x: x["판매건수"], reverse=True)
        
        print(f"\n[피벗 집계 완료] {len(result)}개 상품")
        if result:
            print(f"[TOP 3 상품]")
            for i, item in enumerate(result[:3], 1):
                print(f"  {i}. {item['상품명'][:30]} - {item['판매건수']}건")
        
        update_progress(f"완료! {len(all_orders)}건 수집, {len(result)}개 상품 집계")
        
        return result, None
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return None, f"오류: {str(e)}"

# ========== 계정 로드 ==========
def load_accounts():
    """계정목록 시트에서 스마트스토어 계정 로드 (server.py 방식)"""
    try:
        scope = [
            'https://spreadsheets.google.com/feeds',
            'https://www.googleapis.com/auth/drive'
        ]
        
        creds = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=scope)
        client = gspread.authorize(creds)
        
        sheet = client.open_by_key(SPREADSHEET_KEY)
        ws = sheet.worksheet(ACCOUNTS_TAB)
        
        # server.py와 동일하게 get_all_records() 사용
        records = ws.get_all_records()
        
        if not records:
            return [], "계정 데이터가 없습니다"
        
        accounts = []
        debug_log = []
        
        for idx, row in enumerate(records, 2):  # 2행부터 (헤더 다음)
            platform = row.get("플랫폼", "").strip()
            store_name = row.get("쇼핑몰 별칭", "").strip()
            usage = row.get("용도", "").strip()
            client_id = row.get("스마트스토어 애플리케이션 ID", "").strip()
            client_secret = row.get("스마트스토어 애플리케이션 시크릿", "").strip()
            
            # 스마트스토어 + 용도=대량 + API 정보 있음
            if "스마트스토어" in platform and usage == "대량" and client_id and client_secret:
                accounts.append({
                    "store_name": store_name,
                    "client_id": client_id,
                    "client_secret": client_secret,
                    "usage": usage
                })
                debug_log.append(f"✅ 행{idx}: {store_name} (용도: {usage})")
            elif "스마트스토어" in platform and usage == "대량":
                debug_log.append(f"⚠️ 행{idx}: {store_name} - API 정보 없음 (ID: {bool(client_id)}, Secret: {bool(client_secret)})")
            elif "스마트스토어" in platform and client_id and client_secret:
                debug_log.append(f"ℹ️ 행{idx}: {store_name} - 용도가 '대량'이 아님 (용도: {usage})")
        
        # 디버그 출력
        print("\n" + "="*60)
        print("📋 계정 로딩 결과:")
        print("="*60)
        for log in debug_log:
            print(log)
        print(f"\n총 {len(accounts)}개 스마트스토어 계정 로드 완료")
        print("="*60 + "\n")
        
        return accounts, None
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return [], f"계정 로드 오류: {str(e)}"

# ========== GUI ==========
class AliProductsAPIGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("스마트스토어 API 알리 상품 수집 및 분석")
        self.root.geometry("1400x800")
        
        self.accounts = []
        self.selected_accounts = []
        self.data = []
        
        self.create_widgets()
        self.load_accounts_async()
    
    def create_widgets(self):
        # ========== 계정 정보 프레임 ==========
        account_frame = ttk.LabelFrame(self.root, text="📋 계정 정보 (용도=대량)", padding=10)
        account_frame.pack(fill=tk.X, padx=10, pady=5)
        
        # 로딩 상태
        self.account_status = ttk.Label(account_frame, text="계정 로드 중...", foreground="blue", font=("맑은 고딕", 10))
        self.account_status.pack(side=tk.LEFT, padx=5)
        
        # ========== 날짜 선택 프레임 ==========
        date_frame = ttk.LabelFrame(self.root, text="📅 날짜 범위", padding=10)
        date_frame.pack(fill=tk.X, padx=10, pady=5)
        
        # 시작일
        ttk.Label(date_frame, text="시작일:", font=("맑은 고딕", 10, "bold")).pack(side=tk.LEFT, padx=5)
        
        self.start_year = tk.IntVar(value=datetime.now().year)
        self.start_month = tk.IntVar(value=datetime.now().month)
        self.start_day = tk.IntVar(value=1)
        
        ttk.Spinbox(date_frame, from_=2020, to=2030, textvariable=self.start_year, width=8).pack(side=tk.LEFT, padx=2)
        ttk.Label(date_frame, text="년").pack(side=tk.LEFT)
        ttk.Spinbox(date_frame, from_=1, to=12, textvariable=self.start_month, width=5).pack(side=tk.LEFT, padx=2)
        ttk.Label(date_frame, text="월").pack(side=tk.LEFT)
        ttk.Spinbox(date_frame, from_=1, to=31, textvariable=self.start_day, width=5).pack(side=tk.LEFT, padx=2)
        ttk.Label(date_frame, text="일").pack(side=tk.LEFT, padx=10)
        
        # 종료일
        ttk.Label(date_frame, text="종료일:", font=("맑은 고딕", 10, "bold")).pack(side=tk.LEFT, padx=5)
        
        self.end_year = tk.IntVar(value=datetime.now().year)
        self.end_month = tk.IntVar(value=datetime.now().month)
        self.end_day = tk.IntVar(value=datetime.now().day)
        
        ttk.Spinbox(date_frame, from_=2020, to=2030, textvariable=self.end_year, width=8).pack(side=tk.LEFT, padx=2)
        ttk.Label(date_frame, text="년").pack(side=tk.LEFT)
        ttk.Spinbox(date_frame, from_=1, to=12, textvariable=self.end_month, width=5).pack(side=tk.LEFT, padx=2)
        ttk.Label(date_frame, text="월").pack(side=tk.LEFT)
        ttk.Spinbox(date_frame, from_=1, to=31, textvariable=self.end_day, width=5).pack(side=tk.LEFT, padx=2)
        ttk.Label(date_frame, text="일").pack(side=tk.LEFT)
        
        # 빠른 선택
        ttk.Label(date_frame, text="빠른 선택:", font=("맑은 고딕", 9)).pack(side=tk.LEFT, padx=20)
        ttk.Button(date_frame, text="이번달", command=self.set_this_month).pack(side=tk.LEFT, padx=2)
        ttk.Button(date_frame, text="지난달", command=self.set_last_month).pack(side=tk.LEFT, padx=2)
        ttk.Button(date_frame, text="최근 7일", command=self.set_last_7days).pack(side=tk.LEFT, padx=2)
        ttk.Button(date_frame, text="최근 30일", command=self.set_last_30days).pack(side=tk.LEFT, padx=2)
        
        # ========== 컨트롤 프레임 ==========
        control_frame = ttk.Frame(self.root, padding=10)
        control_frame.pack(fill=tk.X)
        
        # 수집 버튼
        self.collect_btn = ttk.Button(
            control_frame,
            text="🔍 데이터 수집",
            command=self.start_collection,
            state=tk.DISABLED
        )
        self.collect_btn.pack(side=tk.LEFT, padx=10)
        
        # 엑셀 저장 버튼
        self.save_btn = ttk.Button(
            control_frame,
            text="💾 엑셀 저장",
            command=self.save_to_excel,
            state=tk.DISABLED
        )
        self.save_btn.pack(side=tk.LEFT, padx=5)
        
        # 상태 표시
        self.status_label = ttk.Label(control_frame, text="날짜 범위를 지정한 후 '데이터 수집'을 클릭하세요", foreground="gray")
        self.status_label.pack(side=tk.LEFT, padx=20)
        
        # ========== 통계 프레임 ==========
        stats_frame = ttk.LabelFrame(self.root, text="📊 통계", padding=10)
        stats_frame.pack(fill=tk.X, padx=10, pady=5)
        
        self.stats_text = tk.Text(stats_frame, height=3, font=("맑은 고딕", 9), bg="#f0f0f0")
        self.stats_text.pack(fill=tk.X)
        self.stats_text.insert("1.0", "데이터를 수집하면 통계가 표시됩니다.")
        self.stats_text.config(state=tk.DISABLED)
        
        # ========== 테이블 프레임 ==========
        table_frame = ttk.LabelFrame(self.root, text="🛒 알리 상품 집계 (판매건수 순)", padding=10)
        table_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        # 스크롤바
        scroll_y = ttk.Scrollbar(table_frame, orient=tk.VERTICAL)
        scroll_y.pack(side=tk.RIGHT, fill=tk.Y)
        
        scroll_x = ttk.Scrollbar(table_frame, orient=tk.HORIZONTAL)
        scroll_x.pack(side=tk.BOTTOM, fill=tk.X)
        
        # Treeview
        columns = ("순위", "상품명", "옵션명", "판매자상품코드", "스토어명", "판매건수", "총수량", "총금액", "정상", "클레임")
        
        self.tree = ttk.Treeview(
            table_frame,
            columns=columns,
            show="headings",
            yscrollcommand=scroll_y.set,
            xscrollcommand=scroll_x.set,
            height=15
        )
        
        scroll_y.config(command=self.tree.yview)
        scroll_x.config(command=self.tree.xview)
        
        # 컬럼 설정
        self.tree.heading("순위", text="순위")
        self.tree.heading("상품명", text="상품명")
        self.tree.heading("옵션명", text="옵션명")
        self.tree.heading("판매자상품코드", text="판매자상품코드")
        self.tree.heading("스토어명", text="스토어명")
        self.tree.heading("판매건수", text="판매건수")
        self.tree.heading("총수량", text="총수량")
        self.tree.heading("총금액", text="총금액")
        self.tree.heading("정상", text="정상")
        self.tree.heading("클레임", text="클레임")
        
        self.tree.column("순위", width=50, anchor=tk.CENTER)
        self.tree.column("상품명", width=350, anchor=tk.W)
        self.tree.column("옵션명", width=150, anchor=tk.W)
        self.tree.column("판매자상품코드", width=120, anchor=tk.W)
        self.tree.column("스토어명", width=120, anchor=tk.W)
        self.tree.column("판매건수", width=80, anchor=tk.E)
        self.tree.column("총수량", width=80, anchor=tk.E)
        self.tree.column("총금액", width=120, anchor=tk.E)
        self.tree.column("정상", width=60, anchor=tk.E)
        self.tree.column("클레임", width=60, anchor=tk.E)
        
        self.tree.pack(fill=tk.BOTH, expand=True)
        
        # 테이블 스타일
        style = ttk.Style()
        style.configure("Treeview", rowheight=25, font=("맑은 고딕", 9))
        style.configure("Treeview.Heading", font=("맑은 고딕", 9, "bold"))
    
    def load_accounts_async(self):
        """계정 로드 (비동기)"""
        thread = threading.Thread(target=self.load_accounts_thread, daemon=True)
        thread.start()
    
    def load_accounts_thread(self):
        """계정 로드 스레드"""
        accounts, error = load_accounts()
        self.root.after(0, lambda: self.on_accounts_loaded(accounts, error))
    
    def on_accounts_loaded(self, accounts, error):
        """계정 로드 완료"""
        if error:
            self.account_status.config(text=f"❌ {error}", foreground="red")
            messagebox.showerror("오류", error)
            return
        
        if not accounts:
            self.account_status.config(text="⚠️ API 정보가 있는 용도=대량 계정이 없습니다", foreground="orange")
            messagebox.showwarning("알림", "API 정보(애플리케이션 ID/시크릿)가 설정된 용도=대량 스마트스토어 계정이 없습니다.\n\n계정목록 시트를 확인해주세요.")
            return
        
        self.accounts = accounts
        
        # 계정명 목록 표시
        account_names = [acc["store_name"] for acc in accounts]
        display_text = f"✅ {len(accounts)}개 계정 로드 완료: " + ", ".join(account_names[:5])
        if len(accounts) > 5:
            display_text += f" 외 {len(accounts)-5}개"
        
        self.account_status.config(text=display_text, foreground="green")
        self.collect_btn.config(state=tk.NORMAL)
    
    def start_collection(self):
        """데이터 수집 시작"""
        # 모든 계정 자동 사용
        if not self.accounts:
            messagebox.showwarning("경고", "로드된 계정이 없습니다")
            return
        
        # 날짜 범위 (IntVar 값을 안전하게 int로 변환)
        try:
            start_year = int(self.start_year.get())
            start_month = int(self.start_month.get())
            start_day = int(self.start_day.get())
            end_year = int(self.end_year.get())
            end_month = int(self.end_month.get())
            end_day = int(self.end_day.get())
            
            start_date = datetime(start_year, start_month, start_day)
            end_date = datetime(end_year, end_month, end_day)
        except (ValueError, TypeError) as e:
            messagebox.showerror("오류", f"유효하지 않은 날짜입니다: {e}")
            return
        
        if start_date > end_date:
            messagebox.showerror("오류", "시작일이 종료일보다 늦습니다")
            return
        
        self.collect_btn.config(state=tk.DISABLED)
        self.status_label.config(text=f"📥 {len(self.accounts)}개 계정에서 데이터 수집 중...", foreground="blue")
        
        # 백그라운드 스레드에서 실행
        thread = threading.Thread(target=self.collect_data, args=(self.accounts, start_date, end_date), daemon=True)
        thread.start()
    
    def set_this_month(self):
        """이번달 설정"""
        now = datetime.now()
        self.start_year.set(now.year)
        self.start_month.set(now.month)
        self.start_day.set(1)
        self.end_year.set(now.year)
        self.end_month.set(now.month)
        self.end_day.set(now.day)
    
    def set_last_month(self):
        """지난달 설정"""
        now = datetime.now()
        first_day = now.replace(day=1)
        last_month = first_day - timedelta(days=1)
        
        self.start_year.set(last_month.year)
        self.start_month.set(last_month.month)
        self.start_day.set(1)
        self.end_year.set(last_month.year)
        self.end_month.set(last_month.month)
        self.end_day.set(last_month.day)
    
    def set_last_7days(self):
        """최근 7일 설정"""
        now = datetime.now()
        start = now - timedelta(days=7)
        
        self.start_year.set(start.year)
        self.start_month.set(start.month)
        self.start_day.set(start.day)
        self.end_year.set(now.year)
        self.end_month.set(now.month)
        self.end_day.set(now.day)
    
    def set_last_30days(self):
        """최근 30일 설정"""
        now = datetime.now()
        start = now - timedelta(days=30)
        
        self.start_year.set(start.year)
        self.start_month.set(start.month)
        self.start_day.set(start.day)
        self.end_year.set(now.year)
        self.end_month.set(now.month)
        self.end_day.set(now.day)
    
    def collect_data(self, accounts, start_date, end_date):
        """데이터 수집 (백그라운드)"""
        
        def progress_callback(msg):
            self.root.after(0, lambda: self.status_label.config(text=msg))
        
        result, error = collect_orders_from_api(accounts, start_date, end_date, progress_callback)
        
        # UI 업데이트는 메인 스레드에서
        self.root.after(0, lambda: self.on_collection_complete(result, error, start_date, end_date))
    
    def on_collection_complete(self, result, error, start_date, end_date):
        """수집 완료 후 처리"""
        self.collect_btn.config(state=tk.NORMAL)
        
        if error:
            messagebox.showerror("오류", error)
            self.status_label.config(text=f"❌ 오류: {error}", foreground="red")
            return
        
        if not result:
            messagebox.showinfo("알림", "데이터가 없습니다")
            self.status_label.config(text="⚠️ 데이터 없음", foreground="orange")
            return
        
        self.data = result
        self.display_data()
        self.update_stats(start_date, end_date)
        self.save_btn.config(state=tk.NORMAL)
        
        date_range = f"{start_date.strftime('%Y-%m-%d')} ~ {end_date.strftime('%Y-%m-%d')}"
        self.status_label.config(text=f"✅ 수집 완료! ({len(result)}개 상품)", foreground="green")
    
    def display_data(self):
        """데이터 테이블에 표시"""
        # 기존 데이터 삭제
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        # 데이터 삽입
        for rank, item in enumerate(self.data, 1):
            values = (
                rank,
                item["상품명"][:50] + "..." if len(item["상품명"]) > 50 else item["상품명"],
                item["옵션명"][:25] + "..." if len(item["옵션명"]) > 25 else item["옵션명"],
                item["판매자상품코드"],
                item["스토어명"][:20] + "..." if len(item["스토어명"]) > 20 else item["스토어명"],
                f"{item['판매건수']:,}",
                f"{item['총수량']:,}",
                f"{item['총금액']:,}원",
                f"{item['정상건수']:,}",
                f"{item['클레임건수']:,}"
            )
            
            self.tree.insert("", tk.END, values=values)
    
    def update_stats(self, start_date, end_date):
        """통계 업데이트"""
        if not self.data:
            return
        
        total_products = len(self.data)
        total_orders = sum(item["판매건수"] for item in self.data)
        total_quantity = sum(item["총수량"] for item in self.data)
        total_amount = sum(item["총금액"] for item in self.data)
        total_normal = sum(item["정상건수"] for item in self.data)
        total_claim = sum(item["클레임건수"] for item in self.data)
        
        date_range = f"{start_date.strftime('%Y-%m-%d')} ~ {end_date.strftime('%Y-%m-%d')}"
        
        stats_text = f"""
📅 기간: {date_range} | 📦 총 {total_products:,}개 상품 | 📋 총 {total_orders:,}건 주문 | 🔢 총 {total_quantity:,}개 | 💰 총 {total_amount:,}원
✅ 정상: {total_normal:,}건 ({total_normal/total_orders*100:.1f}%) | ⚠️ 클레임: {total_claim:,}건 ({total_claim/total_orders*100:.1f}%)
        """
        
        self.stats_text.config(state=tk.NORMAL)
        self.stats_text.delete("1.0", tk.END)
        self.stats_text.insert("1.0", stats_text.strip())
        self.stats_text.config(state=tk.DISABLED)
    
    def save_to_excel(self):
        """엑셀 파일로 저장"""
        if not self.data:
            messagebox.showwarning("경고", "저장할 데이터가 없습니다")
            return
        
        # 파일명 생성
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_filename = f"알리상품집계_API_{timestamp}.xlsx"
        
        # 파일 저장 다이얼로그
        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile=default_filename,
            filetypes=[("Excel 파일", "*.xlsx"), ("모든 파일", "*.*")]
        )
        
        if not filepath:
            return
        
        try:
            # openpyxl로 저장
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "알리상품집계"
            
            # 헤더
            headers = ["순위", "상품명", "옵션명", "판매자상품코드", "스토어명", "판매건수", "총수량", "총금액", "정상건수", "클레임건수"]
            ws.append(headers)
            
            # 헤더 스타일
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 데이터
            for rank, item in enumerate(self.data, 1):
                ws.append([
                    rank,
                    item["상품명"],
                    item["옵션명"],
                    item["판매자상품코드"],
                    item["스토어명"],
                    item["판매건수"],
                    item["총수량"],
                    item["총금액"],
                    item["정상건수"],
                    item["클레임건수"]
                ])
            
            # 컬럼 너비 조정
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 50
            ws.column_dimensions['C'].width = 30
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 12
            ws.column_dimensions['H'].width = 15
            ws.column_dimensions['I'].width = 12
            ws.column_dimensions['J'].width = 12
            
            # 저장
            wb.save(filepath)
            
            messagebox.showinfo("저장 완료", f"파일이 저장되었습니다:\n{filepath}")
            
        except Exception as e:
            messagebox.showerror("저장 실패", f"엑셀 저장 중 오류:\n{str(e)}")

# ========== 메인 실행 ==========
if __name__ == "__main__":
    root = tk.Tk()
    app = AliProductsAPIGUI(root)
    root.mainloop()
