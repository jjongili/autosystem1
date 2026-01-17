# -*- coding: utf-8 -*-
"""
불사자 옵션 업데이터 v1.0

시뮬레이션 결과 엑셀을 읽어서:
1. 사용자가 선택한 대표옵션(A/B/C)을 불사자에 반영
2. 미끼옵션 SKU 삭제

워크플로우:
1. 시뮬레이터로 분석 → 엑셀 저장
2. 사용자가 엑셀에서 '선택' 컬럼 수정 (A/B/C)
3. 이 프로그램으로 선택 반영 + 미끼옵션 삭제
4. 상품명 변경 후 재등록 (재활용)

by 프코노미
"""

import os
import sys
import time
import threading
import json
from datetime import datetime
from typing import List, Dict, Tuple, Optional

import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox, filedialog

# 공통 모듈 임포트
try:
    from bulsaja_common import (
        BulsajaAPIClient, extract_tokens_from_browser,
        MARKET_IDS
    )
except ImportError:
    print("⚠️ bulsaja_common.py 모듈을 찾을 수 없습니다. 같은 폴더에 있는지 확인하세요.")
    sys.exit(1)

# 엑셀 라이브러리
try:
    from openpyxl import load_workbook
    EXCEL_AVAILABLE = True
except ImportError:
    print("⚠️ openpyxl이 설치되지 않았습니다. pip install openpyxl")
    EXCEL_AVAILABLE = False

# ==================== 설정 ====================
CONFIG_FILE = "bulsaja_option_updater_config.json"
DEBUG_PORT = 9222


def load_config():
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            pass
    return {}


def save_config(config):
    try:
        with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
        return True
    except:
        return False


# ==================== 옵션 업데이터 클래스 ====================
class OptionUpdater:
    """엑셀 기반 옵션 업데이터"""

    def __init__(self, gui):
        self.gui = gui
        self.api_client = None
        self.excel_data = []
        self.stop_flag = False

    def set_api_client(self, client: BulsajaAPIClient):
        self.api_client = client

    def log(self, message: str, level: str = 'INFO'):
        """GUI 로그 출력"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.gui.log_message(f"[{timestamp}] {message}")

    def load_excel(self, filepath: str) -> Tuple[bool, str]:
        """
        시뮬레이션 결과 엑셀 로드

        상세정보 시트에서 읽어올 컬럼:
        - B: 불사자ID
        - L: 선택 (A/B/C)
        - K: 미끼옵션목록
        - P: 최종옵션목록
        """
        if not os.path.exists(filepath):
            return False, f"파일 없음: {filepath}"

        try:
            wb = load_workbook(filepath, data_only=True)

            # 상세정보 시트 찾기
            if "상세정보" in wb.sheetnames:
                ws = wb["상세정보"]
            else:
                return False, "'상세정보' 시트를 찾을 수 없습니다"

            self.excel_data = []

            # 헤더 행 스킵 (row 1), 데이터는 row 2부터
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                product_id = row[1].value  # B: 불사자ID (인덱스 1)
                selection = row[11].value  # L: 선택 (인덱스 11)
                bait_options = row[10].value  # K: 미끼옵션목록 (인덱스 10)
                final_options = row[15].value  # P: 최종옵션목록 (인덱스 15)
                product_name = row[2].value  # C: 상품명 (인덱스 2)

                if not product_id:
                    continue

                # 선택값 정규화 (A/B/C 외 값은 A로 기본)
                if selection not in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
                    selection = 'A'

                # 미끼옵션 파싱 (A. xxx\nB. yyy 형태)
                bait_list = self._parse_option_list(bait_options)

                # 최종옵션 파싱
                final_list = self._parse_option_list(final_options)

                self.excel_data.append({
                    'row': row_idx,
                    'product_id': str(product_id),
                    'product_name': product_name or '',
                    'selection': selection,
                    'bait_options': bait_list,
                    'final_options': final_list
                })

            wb.close()

            # 통계
            total = len(self.excel_data)
            bait_count = sum(1 for d in self.excel_data if d['bait_options'])
            non_a_count = sum(1 for d in self.excel_data if d['selection'] != 'A')

            return True, f"로드 완료: {total}개 상품, 미끼옵션 {bait_count}개, A 외 선택 {non_a_count}개"

        except Exception as e:
            return False, f"엑셀 로드 에러: {e}"

    def _parse_option_list(self, text: str) -> List[str]:
        """
        옵션 목록 파싱 (A. xxx\nB. yyy 형태)
        """
        if not text:
            return []

        options = []
        for line in str(text).split('\n'):
            line = line.strip()
            if not line:
                continue
            # "A. 옵션명" 형태에서 옵션명만 추출
            if len(line) > 2 and line[1] == '.':
                options.append(line[3:].strip())
            else:
                options.append(line)

        return options

    def run_update(self, update_selection: bool, delete_bait: bool,
                   progress_callback=None) -> Tuple[int, int, List[str]]:
        """
        옵션 업데이트 실행

        Args:
            update_selection: 대표옵션 선택 반영 여부
            delete_bait: 미끼옵션 삭제 여부
            progress_callback: 진행률 콜백 (current, total)

        Returns:
            (성공 수, 실패 수, 에러 목록)
        """
        if not self.api_client:
            return 0, 0, ["API 클라이언트가 설정되지 않았습니다"]

        if not self.excel_data:
            return 0, 0, ["엑셀 데이터가 로드되지 않았습니다"]

        success_count = 0
        fail_count = 0
        errors = []
        total = len(self.excel_data)

        for idx, item in enumerate(self.excel_data):
            if self.stop_flag:
                self.log("사용자 중단 요청")
                break

            product_id = item['product_id']
            product_name = item['product_name'][:30]
            selection = item['selection']
            bait_options = item['bait_options']
            final_options = item['final_options']

            if progress_callback:
                progress_callback(idx + 1, total)

            self.log(f"[{idx+1}/{total}] {product_name} (ID: {product_id})")

            try:
                # 1. 상품 상세 정보 조회
                detail = self.api_client.get_product_detail(product_id)
                if not detail:
                    errors.append(f"{product_id}: 상품 조회 실패")
                    fail_count += 1
                    continue

                skus = detail.get('skuInfoList', [])
                if not skus:
                    self.log(f"  → SKU 없음, 스킵")
                    continue

                modified = False

                # 2. 대표옵션 선택 반영
                if update_selection and selection != 'A':
                    # 선택된 옵션 인덱스 계산 (A=0, B=1, C=2...)
                    selection_idx = ord(selection) - ord('A')

                    if selection_idx < len(final_options):
                        selected_option = final_options[selection_idx]
                        self.log(f"  → 대표옵션 변경: {selection}. {selected_option[:20]}")

                        # SKU 중 선택된 옵션을 대표로 설정
                        # mainImage 플래그 또는 순서 변경으로 처리
                        for sku in skus:
                            sku_name = sku.get('skuName', '') or sku.get('optionName', '')
                            if selected_option in sku_name or sku_name in selected_option:
                                sku['mainImage'] = True
                                sku['isRepresentative'] = True
                            else:
                                sku['mainImage'] = False
                                sku['isRepresentative'] = False
                        modified = True

                # 3. 미끼옵션 삭제
                if delete_bait and bait_options:
                    original_count = len(skus)

                    # 미끼옵션과 일치하는 SKU 제거
                    filtered_skus = []
                    deleted_names = []

                    for sku in skus:
                        sku_name = sku.get('skuName', '') or sku.get('optionName', '')
                        is_bait = False

                        for bait in bait_options:
                            if bait in sku_name or sku_name in bait:
                                is_bait = True
                                deleted_names.append(sku_name[:20])
                                break

                        if not is_bait:
                            filtered_skus.append(sku)

                    if len(filtered_skus) < original_count:
                        self.log(f"  → 미끼옵션 삭제: {deleted_names}")
                        detail['skuInfoList'] = filtered_skus
                        modified = True

                # 4. 변경사항 저장
                if modified:
                    success = self.api_client.update_product_fields(product_id, detail)
                    if success:
                        self.log(f"  → 저장 완료")
                        success_count += 1
                    else:
                        errors.append(f"{product_id}: 저장 실패")
                        fail_count += 1
                else:
                    self.log(f"  → 변경사항 없음")
                    success_count += 1

                # API 부하 방지
                time.sleep(0.3)

            except Exception as e:
                errors.append(f"{product_id}: {str(e)[:50]}")
                fail_count += 1
                self.log(f"  → 에러: {e}")

        return success_count, fail_count, errors

    def stop(self):
        """작업 중단"""
        self.stop_flag = True


# ==================== GUI 클래스 ====================
class App(tk.Tk):
    def __init__(self):
        super().__init__()

        self.title("불사자 옵션 업데이터 v1.0 - 엑셀 기반 옵션 반영")
        self.geometry("800x650")
        self.resizable(True, True)

        self.config_data = load_config()
        self.updater = OptionUpdater(self)
        self.worker_thread = None

        self.create_widgets()
        self.load_saved_settings()

        self.protocol("WM_DELETE_WINDOW", self.on_close)

    def create_widgets(self):
        main_frame = ttk.Frame(self, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # === 1. API 연결 ===
        conn_frame = ttk.LabelFrame(main_frame, text="API 연결", padding="5")
        conn_frame.pack(fill=tk.X, pady=(0, 5))

        row0 = ttk.Frame(conn_frame)
        row0.pack(fill=tk.X, pady=2)
        ttk.Button(row0, text="크롬", command=self.open_debug_chrome, width=8).pack(side=tk.LEFT)
        ttk.Button(row0, text="토큰", command=self.extract_tokens, width=8).pack(side=tk.LEFT, padx=2)
        ttk.Button(row0, text="연결", command=self.connect_api, width=8).pack(side=tk.LEFT, padx=2)
        self.api_status = ttk.Label(row0, text="연결 안 됨", foreground="gray")
        self.api_status.pack(side=tk.LEFT, padx=10)
        ttk.Label(row0, text="포트:").pack(side=tk.RIGHT)
        self.port_var = tk.StringVar(value="9222")
        ttk.Entry(row0, textvariable=self.port_var, width=6).pack(side=tk.RIGHT, padx=2)

        # === 2. 엑셀 파일 선택 ===
        excel_frame = ttk.LabelFrame(main_frame, text="엑셀 파일 (시뮬레이션 결과)", padding="5")
        excel_frame.pack(fill=tk.X, pady=(0, 5))

        row1 = ttk.Frame(excel_frame)
        row1.pack(fill=tk.X, pady=2)

        self.excel_path_var = tk.StringVar()
        ttk.Entry(row1, textvariable=self.excel_path_var, width=60).pack(side=tk.LEFT, fill=tk.X, expand=True)
        ttk.Button(row1, text="찾아보기", command=self.browse_excel, width=10).pack(side=tk.LEFT, padx=5)
        ttk.Button(row1, text="로드", command=self.load_excel, width=8).pack(side=tk.LEFT)

        self.excel_status = ttk.Label(excel_frame, text="파일 선택 필요", foreground="gray")
        self.excel_status.pack(anchor=tk.W, pady=2)

        # === 3. 업데이트 옵션 ===
        option_frame = ttk.LabelFrame(main_frame, text="업데이트 옵션", padding="5")
        option_frame.pack(fill=tk.X, pady=(0, 5))

        self.update_selection_var = tk.BooleanVar(value=True)
        self.delete_bait_var = tk.BooleanVar(value=True)

        ttk.Checkbutton(option_frame, text="대표옵션 선택 반영 (엑셀 '선택' 컬럼 A/B/C 기준)",
                        variable=self.update_selection_var).pack(anchor=tk.W, pady=2)
        ttk.Checkbutton(option_frame, text="미끼옵션 삭제 (엑셀 '미끼옵션목록' 기준)",
                        variable=self.delete_bait_var).pack(anchor=tk.W, pady=2)

        # 안내 메시지
        info_text = """
        워크플로우:
        1. 시뮬레이터로 분석 → 엑셀 저장 (기본 선택값: A)
        2. 엑셀에서 '선택' 컬럼 수정 (A/B/C... 중 선택)
        3. 이 프로그램으로 선택 반영 + 미끼옵션 삭제
        4. 상품명 변경 후 재등록 (재활용)
        """
        ttk.Label(option_frame, text=info_text.strip(), foreground="gray",
                  justify=tk.LEFT).pack(anchor=tk.W, pady=5)

        # === 4. 실행 버튼 ===
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=tk.X, pady=5)

        self.run_btn = ttk.Button(btn_frame, text="업데이트 실행", command=self.run_update, width=15)
        self.run_btn.pack(side=tk.LEFT)

        self.stop_btn = ttk.Button(btn_frame, text="중단", command=self.stop_update, width=10, state=tk.DISABLED)
        self.stop_btn.pack(side=tk.LEFT, padx=5)

        # 진행률
        self.progress_var = tk.DoubleVar(value=0)
        self.progress_bar = ttk.Progressbar(btn_frame, variable=self.progress_var, maximum=100)
        self.progress_bar.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=10)

        self.progress_label = ttk.Label(btn_frame, text="0/0")
        self.progress_label.pack(side=tk.RIGHT)

        # === 5. 로그 ===
        log_frame = ttk.LabelFrame(main_frame, text="로그", padding="5")
        log_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 5))

        self.log_text = scrolledtext.ScrolledText(log_frame, height=15, font=('Consolas', 9))
        self.log_text.pack(fill=tk.BOTH, expand=True)

    def log_message(self, message: str):
        """로그 메시지 출력"""
        self.log_text.insert(tk.END, message + "\n")
        self.log_text.see(tk.END)
        self.update_idletasks()

    def load_saved_settings(self):
        """저장된 설정 로드"""
        if 'port' in self.config_data:
            self.port_var.set(str(self.config_data['port']))
        if 'last_excel' in self.config_data:
            self.excel_path_var.set(self.config_data['last_excel'])

    def save_settings(self):
        """설정 저장"""
        self.config_data['port'] = int(self.port_var.get() or "9222")
        self.config_data['last_excel'] = self.excel_path_var.get()
        save_config(self.config_data)

    def open_debug_chrome(self):
        """디버그 모드 크롬 실행"""
        import subprocess
        port = self.port_var.get() or "9222"
        chrome_path = r"C:\Program Files\Google\Chrome\Application\chrome.exe"
        user_data = os.path.join(os.environ.get('LOCALAPPDATA', ''), 'Google', 'Chrome', 'User Data')

        cmd = f'"{chrome_path}" --remote-debugging-port={port} --user-data-dir="{user_data}"'

        try:
            subprocess.Popen(cmd, shell=True)
            self.log_message(f"크롬 실행 (포트: {port})")
            self.log_message("→ 불사자(bulsaja.com)에 로그인 후 '토큰' 버튼 클릭")
        except Exception as e:
            messagebox.showerror("오류", f"크롬 실행 실패: {e}")

    def extract_tokens(self):
        """토큰 추출"""
        port = int(self.port_var.get() or "9222")
        success, access, refresh, msg = extract_tokens_from_browser(port)

        if success:
            self.config_data['access_token'] = access
            self.config_data['refresh_token'] = refresh
            save_config(self.config_data)
            self.log_message("토큰 추출 성공")
            self.connect_api()
        else:
            messagebox.showerror("오류", f"토큰 추출 실패: {msg}")

    def connect_api(self):
        """API 연결"""
        access = self.config_data.get('access_token', '')
        refresh = self.config_data.get('refresh_token', '')

        if not access or not refresh:
            messagebox.showwarning("경고", "토큰이 없습니다. '토큰' 버튼을 먼저 클릭하세요.")
            return

        client = BulsajaAPIClient(access, refresh)
        success, msg, total = client.test_connection()

        if success:
            self.updater.set_api_client(client)
            self.api_status.config(text=f"연결됨 ({total}개 상품)", foreground="green")
            self.log_message(f"API 연결 성공: {total}개 상품")
        else:
            self.api_status.config(text="연결 실패", foreground="red")
            messagebox.showerror("오류", f"API 연결 실패: {msg}")

    def browse_excel(self):
        """엑셀 파일 선택"""
        filepath = filedialog.askopenfilename(
            title="시뮬레이션 결과 엑셀 선택",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialdir=os.getcwd()
        )
        if filepath:
            self.excel_path_var.set(filepath)

    def load_excel(self):
        """엑셀 로드"""
        filepath = self.excel_path_var.get()
        if not filepath:
            messagebox.showwarning("경고", "엑셀 파일을 선택하세요")
            return

        success, msg = self.updater.load_excel(filepath)

        if success:
            self.excel_status.config(text=msg, foreground="green")
            self.log_message(msg)
            self.save_settings()
        else:
            self.excel_status.config(text=msg, foreground="red")
            messagebox.showerror("오류", msg)

    def run_update(self):
        """업데이트 실행"""
        if not self.updater.api_client:
            messagebox.showwarning("경고", "API에 먼저 연결하세요")
            return

        if not self.updater.excel_data:
            messagebox.showwarning("경고", "엑셀 파일을 먼저 로드하세요")
            return

        update_selection = self.update_selection_var.get()
        delete_bait = self.delete_bait_var.get()

        if not update_selection and not delete_bait:
            messagebox.showwarning("경고", "최소 하나의 옵션을 선택하세요")
            return

        # 확인
        total = len(self.updater.excel_data)
        msg = f"{total}개 상품에 대해 업데이트를 실행합니다.\n\n"
        if update_selection:
            msg += "- 대표옵션 선택 반영\n"
        if delete_bait:
            msg += "- 미끼옵션 삭제\n"
        msg += "\n계속하시겠습니까?"

        if not messagebox.askyesno("확인", msg):
            return

        # 버튼 상태 변경
        self.run_btn.config(state=tk.DISABLED)
        self.stop_btn.config(state=tk.NORMAL)
        self.updater.stop_flag = False

        def progress_callback(current, total):
            self.progress_var.set((current / total) * 100)
            self.progress_label.config(text=f"{current}/{total}")
            self.update_idletasks()

        def worker():
            success, fail, errors = self.updater.run_update(
                update_selection, delete_bait, progress_callback
            )

            # 완료 후 GUI 업데이트
            self.after(0, lambda: self.on_update_complete(success, fail, errors))

        self.worker_thread = threading.Thread(target=worker, daemon=True)
        self.worker_thread.start()

    def on_update_complete(self, success: int, fail: int, errors: List[str]):
        """업데이트 완료 처리"""
        self.run_btn.config(state=tk.NORMAL)
        self.stop_btn.config(state=tk.DISABLED)

        msg = f"완료: 성공 {success}개, 실패 {fail}개"
        self.log_message(msg)

        if errors:
            self.log_message(f"에러 목록 ({len(errors)}개):")
            for err in errors[:10]:
                self.log_message(f"  - {err}")
            if len(errors) > 10:
                self.log_message(f"  ... 외 {len(errors) - 10}개")

        messagebox.showinfo("완료", msg)

    def stop_update(self):
        """업데이트 중단"""
        self.updater.stop()
        self.log_message("중단 요청...")

    def on_close(self):
        """종료 처리"""
        self.save_settings()
        self.updater.stop()
        self.destroy()


# ==================== 메인 ====================
if __name__ == "__main__":
    if not EXCEL_AVAILABLE:
        print("openpyxl이 필요합니다. pip install openpyxl")
        sys.exit(1)

    app = App()
    app.mainloop()
