# -*- coding: utf-8 -*-
"""
불사자 시뮬레이터 통합 v2.0 (13. bulsaja_simulator_v2.0.py)
기능:
1. [시뮬레이션] 탭: 불사자 API를 통해 상품 데이터를 수집 (모든 썸네일, 중국어 옵션 포함)
2. [검수] 탭: 수집된 데이터를 엑셀에서 불러와 시각적으로 검수
   - 썸네일 다중 비교 및 누끼(배경없음) 이미지 자동 추천
   - 옵션명 표준화 (중국어-한국어 사전 기반)
   - 미끼 옵션 필터링 및 대표 옵션 지정

사용법:
- 실행 후 [시뮬레이션] 탭에서 데이터 수집 -> 엑셀 저장
- [검수] 탭에서 엑셀 불러오기 -> 썸네일/옵션 선택 -> 저장
"""

import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox, filedialog
import os
import json
import threading
import time
import requests
from io import BytesIO
from datetime import datetime
from typing import List, Dict, Optional, Tuple, Any
import re
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    print("⚠️ pandas가 필요합니다.")
from PIL import Image, ImageTk

# ===== 외부 의존성 처리 =====
try:
    import cv2
    import numpy as np
    CV2_AVAILABLE = True
except ImportError:
    CV2_AVAILABLE = False
    print("⚠️ opencv-python이 설치되지 않았습니다. 누끼 정밀 분석이 제한됩니다.")

# OCR: EasyOCR (SOTA) - pytesseract 대체
try:
    import easyocr
    EASYOCR_AVAILABLE = True
    EASYOCR_READER = None # Lazy Loading
except ImportError:
    EASYOCR_AVAILABLE = False
    print("⚠️ easyocr이 설치되지 않았습니다. 텍스트 감지가 제한됩니다.")

# Nukki: rembg (SOTA)
try:
    import importlib.util
    if importlib.util.find_spec("onnxruntime") is None and importlib.util.find_spec("onnxruntime_gpu") is None:
        raise ImportError("onnxruntime not installed")
        
    from rembg import remove as rembg_remove
    REMBG_AVAILABLE = True
except Exception as e:
    REMBG_AVAILABLE = False
    # print(f"⚠️ rembg/onnx 로드 실패: {e}") # 너무 시끄러울 수 있으므로 주석 처리권장, 혹은 짧게 표시
    print("⚠️ rembg 기능을 사용할 수 없습니다. (onnxruntime 미설치)")

try:
    import pytesseract
    OCR_AVAILABLE = True
except ImportError:
    OCR_AVAILABLE = False # EasyOCR 있으면 굳이 필요 없으나 예비용

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("⚠️ openpyxl이 필요합니다.")

# 불사자 공통 모듈 (같은 폴더에 가정)
try:
    from bulsaja_common import (
        BulsajaAPIClient, load_banned_words, load_excluded_words, load_bait_keywords,
        check_product_safety, filter_bait_options, select_main_option, extract_tokens_from_browser
    )
    COMMON_AVAILABLE = True
except ImportError:
    COMMON_AVAILABLE = False
    print("⚠️ bulsaja_common.py가 없습니다. 일부 기능이 제한될 수 있습니다.")
    # 더미 클래스/함수 제공
    class BulsajaAPIClient:
        def __init__(self, *args): pass
        def test_connection(self): return False, "모듈 없음", 0

# ==================== 설정 및 상수 ====================
CONFIG_FILE = "bulsaja_simulator_v2_config.json"
TRANSLATION_DICT_FILE = "option_translation_dict.json"

# ==================== 유틸리티 클래스 ====================
class ThumbnailAnalyzer:
    """썸네일 분석기 (누끼/텍스트 감지 - SOTA 적용)"""
    def make_nukki(self, image_data: bytes) -> bytes:
        """rembg를 사용해 배경 제거 (투명 PNG 반환)"""
        if not REMBG_AVAILABLE:
            return image_data
        try:
            return rembg_remove(image_data)
        except Exception as e:
            print(f"누끼 생성 실패: {e}")
            return image_data

    def analyzed_score(self, image_url: str) -> Dict[str, Any]:
        global EASYOCR_READER
        result = {
            "score": 0, "is_nukki": False, "has_text": False, "recommendation": "normal"
        }
        if not CV2_AVAILABLE:
            return result
        
        try:
            # 이미지 다운로드 (메모리)
            resp = requests.get(image_url, timeout=5)
            img_bytes = bytearray(resp.content)
            arr = np.asarray(img_bytes, dtype=np.uint8)
            img = cv2.imdecode(arr, -1)

            if img is None: return result

            # 1. 배경 분석 (가장자리가 흰색/투명이면 누끼 가능성 높음)
            h, w = img.shape[:2]
            corners = [img[0,0], img[0, w-1], img[h-1, 0], img[h-1, w-1]]
            
            is_white_bg = True
            for c in corners:
                # BGR or BGRA
                if len(c) == 4 and c[3] == 0: # 투명
                    continue
                if np.mean(c[:3]) < 240: # 흰색 아님 (여유값)
                    is_white_bg = False
                    break
            
            result["is_nukki"] = is_white_bg
            
            # 2. 텍스트 감지 (SOTA: EasyOCR)
            has_text = False
            if EASYOCR_AVAILABLE:
                if EASYOCR_READER is None:
                    # 한국어, 영어 로드 (GPU 있으면 자동 사용)
                    print("🚀 EasyOCR 모델 로딩 중... (최초 1회)")
                    EASYOCR_READER = easyocr.Reader(['ko', 'en'], gpu=True, verbose=False)
                
                # EasyOCR은 이미지 경로, numpy array, bytes 모두 지원
                # detail=0: 텍스트만 리스트로 반환
                texts = EASYOCR_READER.readtext(img, detail=0)
                # 노이즈 필터링 (너무 짧은 텍스트 무시)
                valid_texts = [t for t in texts if len(t.strip()) > 1]
                if valid_texts:
                    has_text = True
                    # print(f"감지된 텍스트: {valid_texts}")

            elif OCR_AVAILABLE: # Fallback
                gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY) if len(img.shape)==3 else img
                text = pytesseract.image_to_string(gray, lang='eng+kor')
                if len(text.strip()) > 2:
                    has_text = True
            
            result["has_text"] = has_text

            # 점수 산정
            score = 50
            if result["is_nukki"]: score += 40
            if result["has_text"]: score -= 40 # 텍스트 있으면 감점 크게
            
            result["score"] = max(0, min(100, score))
            
            if score >= 80: result["recommendation"] = "best"
            elif score <= 20: result["recommendation"] = "trash"
            
        except Exception as e:
            print(f"이미지 분석 실패: {e}")
            
        return result

class TranslationManager:
    """옵션명 표준화 관리자"""
    def __init__(self):
        self.dictionary = self.load_dictionary()
        
    def load_dictionary(self):
        if os.path.exists(TRANSLATION_DICT_FILE):
            try:
                with open(TRANSLATION_DICT_FILE, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except:
                pass
        return {}

    def standardize(self, cn_text: str, kr_text: str) -> str:
        """중국어 기반으로 한국어 옵션명 표준화"""
        if not cn_text: return kr_text
        
        # 완전 일치 찾기
        for cat, maps in self.dictionary.items():
            if cn_text in maps:
                return maps[cn_text]
        
        # 부분 일치 교체
        new_kr = kr_text
        for cat, maps in self.dictionary.items():
            for cn_word, kr_standard in maps.items():
                if cn_word in cn_text:
                    # 기존 번역이 중구난방일 수 있으므로, 괄호 안이나 특수문자 제거 후 표준어 추가 등 전략 필요
                    # 여기서는 단순 치환 또는 덧붙이기 전략 사용
                    # 하지만 GPT 번역이 이미 되어있는 상태라면, GPT 번역을 무시하고 표준어로 덮어쓰는게 나을 수도 있음
                    # 위험성: 다른 의미가 사라질 수 있음.
                    # 안전한 방법: "[표준어]" 태그를 붙이거나, 완전히 매칭되는 단어가 있으면 그것만 남김
                    return kr_standard # 강력한 표준화 (발견되면 그걸로 대체)
                    
        return new_kr

# ==================== 메인 애플리케이션 ====================
class BulsajaSimulatorV2:
    def __init__(self, root):
        self.root = root
        self.root.title("불사자 시뮬레이터 통합 v2.1")
        self.root.geometry("1400x900")
        
        self.config = self.load_config()
        self.api_client = None
        self.stop_event = threading.Event()
        
        # 탭 구성
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # 탭 1: 시뮬레이션 (수집)
        self.tab_sim = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_sim, text=" 1. 데이터 수집 (시뮬레이션) ")
        self._init_simulation_tab()
        
        # 탭 2: 검수 (Inspector)
        self.tab_insp = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_insp, text=" 2. 데이터 검수 (Inspector) ")
        self._init_inspector_tab()
        
        # 초기화
        self._check_modules()

    def load_config(self):
        if os.path.exists(CONFIG_FILE):
            try:
                with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except:
                pass
        return {}

    def save_config(self):
        try:
            with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
                json.dump(self.config, f, ensure_ascii=False, indent=2)
        except:
            pass

    def _check_modules(self):
        msgs = []
        if not EXCEL_AVAILABLE: msgs.append("openpyxl 설치 필요")
        if not COMMON_AVAILABLE: msgs.append("bulsaja_common.py 없음")
        if msgs:
            messagebox.showwarning("경고", "\n".join(msgs))

    # ==================== 탭 1: 시뮬레이션 로직 ====================
    def _init_simulation_tab(self):
        # 상단: 설정
        f_cfg = ttk.LabelFrame(self.tab_sim, text=" API 설정 ", padding=10)
        f_cfg.pack(fill=tk.X, padx=10, pady=5)
        
        ttk.Label(f_cfg, text="Access Token:").pack(side=tk.LEFT)
        self.token_var = tk.StringVar(value=self.config.get("access_token", ""))
        ttk.Entry(f_cfg, textvariable=self.token_var, width=30).pack(side=tk.LEFT, padx=5)
        
        ttk.Label(f_cfg, text="Refresh Token:").pack(side=tk.LEFT)
        self.refresh_var = tk.StringVar(value=self.config.get("refresh_token", ""))
        ttk.Entry(f_cfg, textvariable=self.refresh_var, width=30).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(f_cfg, text="크롬 디버그 실행", command=self._open_debug_chrome).pack(side=tk.LEFT, padx=5)
        ttk.Button(f_cfg, text="포트9222 토큰추출", command=self._extract_tokens).pack(side=tk.LEFT, padx=5)
        ttk.Button(f_cfg, text="연결 확인", command=self._test_connection).pack(side=tk.LEFT, padx=10)
        
        # 중단: 실행 옵션
        f_opt = ttk.LabelFrame(self.tab_sim, text=" 시뮬레이션 옵션 ", padding=10)
        f_opt.pack(fill=tk.X, padx=10, pady=5)
        
        ttk.Label(f_opt, text="대상 그룹명 (쉼표 구분):").grid(row=0, column=0, sticky='w')
        self.groups_var = tk.StringVar(value=self.config.get("last_groups", ""))
        ttk.Entry(f_opt, textvariable=self.groups_var, width=50).grid(row=0, column=1, padx=5, sticky='w')
        
        ttk.Label(f_opt, text="그룹당 상품 수:").grid(row=0, column=2, sticky='w', padx=10)
        self.limit_var = tk.IntVar(value=self.config.get("limit_per_group", 50))
        ttk.Entry(f_opt, textvariable=self.limit_var, width=10).grid(row=0, column=3, sticky='w')
        
        ttk.Button(f_opt, text="▶ 시뮬레이션 시작", command=self._start_simulation, width=20).grid(row=0, column=4, padx=20)
        ttk.Button(f_opt, text="⏹ 중지", command=self._stop_simulation).grid(row=0, column=5)
        
        # 하단: 로그
        f_log = ttk.Frame(self.tab_sim, padding=10)
        f_log.pack(fill=tk.BOTH, expand=True)
        self.log_area = scrolledtext.ScrolledText(f_log, height=15)
        self.log_area.pack(fill=tk.BOTH, expand=True)
        
        # 진행바
        self.progress = ttk.Progressbar(f_log, mode='determinate')
        self.progress.pack(fill=tk.X, pady=5)

    def log_sim(self, msg):
        self.log_area.insert(tk.END, f"[{datetime.now().strftime('%H:%M:%S')}] {msg}\n")
        self.log_area.see(tk.END)

    def _open_debug_chrome(self):
        import subprocess
        port = 9222 # 고정 포트
        profile_dir = f"C:\\chrome_debug_profile_{port}"
        chrome_paths = [
            r"C:\Program Files\Google\Chrome\Application\chrome.exe",
            r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
            os.path.expandvars(r"%LOCALAPPDATA%\Google\Chrome\Application\chrome.exe"),
        ]
        chrome_path = None
        for path in chrome_paths:
            if os.path.exists(path):
                chrome_path = path
                break
        if not chrome_path:
            messagebox.showerror("오류", "Chrome을 찾을 수 없습니다")
            return
        
        # 불사자 URL로 바로 이동
        url = "https://www.bulsaja.com/products/manage/list/"
        cmd = f'"{chrome_path}" --remote-debugging-port={port} --user-data-dir="{profile_dir}" --remote-allow-origins=* "{url}"'
        try:
            subprocess.Popen(cmd, shell=True)
            self.log_sim(f"🌐 크롬 실행 (포트: {port})")
            messagebox.showinfo("안내", "크롬이 실행되었습니다.\n로그인 후 '토큰추출' 버튼을 누르세요.")
        except Exception as e:
            self.log_sim(f"❌ 크롬 실행 실패: {e}")

    def _extract_tokens(self):
        if not COMMON_AVAILABLE: return
        self.log_sim("🔍 크롬 디버그 포트(9222)에서 토큰 추출 시도...")
        ok, access, refresh, msg = extract_tokens_from_browser(9222)
        if ok:
            self.token_var.set(access)
            self.refresh_var.set(refresh)
            self.log_sim("✅ 토큰 추출 및 적용 완료")
            # 자동 연결 테스트
            self.root.after(500, self._test_connection)
        else:
            messagebox.showwarning("실패", f"토큰 추출 실패: {msg}\n크롬이 9222 포트로 실행 중인지 확인하세요.")

    def _test_connection(self):
        if not COMMON_AVAILABLE: return
        # [Fix] 토큰 공백 제거 (500 에러 방지)
        access = self.token_var.get().strip()
        refresh = self.refresh_var.get().strip()
        
        if not access:
            messagebox.showwarning("경고", "Access Token을 입력하세요")
            return
            
        client = BulsajaAPIClient(access, refresh)
        ok, msg, days = client.test_connection()
        if ok:
            messagebox.showinfo("성공", f"연결 성공! (남은기간: {days}일)")
            self.config["access_token"] = self.token_var.get()
            self.config["refresh_token"] = self.refresh_var.get()
            self.save_config()
            self.api_client = client
        else:
            messagebox.showerror("실패", f"연결 실패: {msg}")

    def _stop_simulation(self):
        self.stop_event.set()
        self.log_sim("🛑 중지 요청됨...")

    def _start_simulation(self):
        if not self.api_client:
            if not self._test_connection(): return # Try connect
            if not self.api_client: return

        groups = [g.strip() for g in self.groups_var.get().split(',') if g.strip()]
        if not groups:
            messagebox.showwarning("경고", "그룹명을 입력하세요")
            return
            
        self.config["last_groups"] = self.groups_var.get()
        self.config["limit_per_group"] = self.limit_var.get()
        self.save_config()
        
        self.stop_event.clear()
        threading.Thread(target=self._run_simulation_thread, args=(groups,)).start()

    def _run_simulation_thread(self, groups):
        self.log_sim("🚀 시뮬레이션 시작")
        limit = self.limit_var.get()
        results = []
        
        try:
            total_items = len(groups) * limit
            current_item = 0
            
            for group_name in groups:
                if self.stop_event.is_set(): break
                self.log_sim(f"📂 그룹 '{group_name}' 상품 조회 중...")
                
                products, _ = self.api_client.get_products_by_group(group_name, 0, limit)
                self.log_sim(f"   -> {len(products)}개 발견")
                
                for p in products:
                    if self.stop_event.is_set(): break
                    current_item += 1
                    res = self._analyze_single_product(p, group_name)
                    results.append(res)
                    
                    # UI Update
                    self.progress['value'] = (current_item / total_items) * 100
                    if current_item % 5 == 0:
                        self.log_sim(f"   ... {current_item}개 처리됨")
                        
            # Save Excel
            if results:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M")
                filename = f"simulation_{timestamp}.xlsx"
                self._save_to_excel(results, filename)
                self.log_sim(f"💾 엑셀 저장 완료: {filename}")
                self.log_sim("✅ 시뮬레이션 완료. [검수] 탭으로 이동하여 파일을 여세요.")
                
                # Auto-load in inspector if possible (unsafe from thread, better user manual load)
                
        except Exception as e:
            self.log_sim(f"❌ 오류 발생: {e}")
            import traceback
            traceback.print_exc()

    def _analyze_single_product(self, product, group_name):
        pid = product.get('ID', '')
        pname = product.get('uploadCommonProductName', '')
        
        # 1. 상세 정보 조회 (기존 로직)
        detail = self.api_client.get_product_detail(pid)
        
        # [추가] 업로드용 전체 필드 데이터 조회 (실제 업로드 시 필요한 모든 정보)
        upload_fields = self.api_client.get_upload_fields(pid)
        if not upload_fields:
            upload_fields = detail # 없을 경우 상세정보로 대체
            
        # 썸네일 수집 (전체)
        thumbs = detail.get('uploadThumbnails', [])
        
        # 옵션 수집
        skus = detail.get('uploadSkus', [])
        if not skus: skus = detail.get('original_skus', [])
        
        # [추가] 상품명 안전 검사 상세 수행
        safety = check_product_safety(pname, load_excluded_words(), check_level='strict') 
        unsafe_reason = ""
        if not safety['is_safe']:
            categories = []
            if safety['categories']['adult']: categories.append(f"성인:{','.join(safety['categories']['adult'][:2])}")
            if safety['categories']['medical']: categories.append(f"의료:{','.join(safety['categories']['medical'][:2])}")
            if safety['categories']['child']: categories.append(f"유아:{','.join(safety['categories']['child'][:2])}")
            if safety['categories']['prohibited']: categories.append(f"금지:{','.join(safety['categories']['prohibited'][:2])}")
            unsafe_reason = " / ".join(categories)

        # 미끼 옵션 정보 수집 (상세 내역)
        bait_details = []
        for sku in bait_skus:
            ko = sku.get('text_ko', '') or sku.get('text', '')
            cn = sku.get('text', '')
            price = sku.get('_origin_price', 0)
            keyword = sku.get('_bait_keyword', '')
            
            detail_str = f"{ko}"
            if cn and cn != ko: detail_str += f"({cn})"
            detail_str += f"[{price}]"
            if keyword: detail_str += f" <{keyword}>" # 탐지 키워드 표시
            bait_details.append(detail_str)
            
        bait_str = "\n".join(bait_details)

        # [로직 고도화] 이미지 있는 옵션 중 최저가 선택 & 썸네일 매칭 확인
        selected_label = "A" 
        selected_idx = 0     
        
        main_thumb = thumbs[0] if thumbs else ""
        thumb_match_idx = -1
        
        # 1. 썸네일 옵션 매칭 시도
        def get_img_id(u): return u.split('/')[-1].split('.')[0] if u else ""
        
        if main_thumb and valid_skus:
            mid = get_img_id(main_thumb)
            for i, sku in enumerate(valid_skus):
                u = sku.get('urlRef') or sku.get('image') or sku.get('img')
                if u and mid in u:
                    thumb_match_idx = i
                    break
        
        # 2. 최저가 선택 로직
        if valid_skus:
            candidates = []
            for idx, sku in enumerate(valid_skus):
                img_url = sku.get('urlRef') or sku.get('image') or sku.get('img')
                price = sku.get('_origin_price', float('inf'))
                if img_url:
                    candidates.append((idx, price))
            
            if not candidates:
                for idx, sku in enumerate(valid_skus):
                    price = sku.get('_origin_price', float('inf'))
                    candidates.append((idx, price))
            
            if candidates:
                candidates.sort(key=lambda x: x[1])
                best_idx = candidates[0][0]
                selected_label = chr(ord('A') + best_idx) if best_idx < 26 else str(best_idx + 1)
                selected_idx = best_idx

        # [추가] 업로드 필수 정보 추출 (Hijacker Schema 참조)
        # 1. 기본 정보
        sale_price = upload_fields.get('salePrice', detail.get('salePrice', 0))
        stock_qty = upload_fields.get('stockQuantity', detail.get('stockQuantity', 99))
        detail_html = upload_fields.get('detailContent', detail.get('detailContent', ''))
        
        # 2. AS 및 원산지
        detail_attr = upload_fields.get('detailAttribute', detail.get('detailAttribute', {}))
        as_info = detail_attr.get('afterServiceInfo', {})
        as_tel = as_info.get('afterServiceTelephoneNumber', '')
        as_guide = as_info.get('afterServiceGuideContent', '')
        origin_area = detail_attr.get('originAreaInfo', {})
        origin_code = origin_area.get('originAreaCode', '03')
        
        # 3. 배송 정보
        delivery = upload_fields.get('deliveryInfo', detail.get('deliveryInfo', {}))
        deliv_fee_obj = delivery.get('deliveryFee', {})
        deliv_type = deliv_fee_obj.get('deliveryFeeType', 'FREE')
        base_fee = deliv_fee_obj.get('baseFee', 0)
        
        claim_info = delivery.get('claimDeliveryInfo', {})
        ret_fee = claim_info.get('returnDeliveryFee', 3000)
        exch_fee = claim_info.get('exchangeDeliveryFee', 6000)
        
        # 4. 검색 정보 (태그, 브랜드 등)
        search_info = detail_attr.get('naverShoppingSearchInfo', {})
        brand = search_info.get('brandName', '')
        maker = search_info.get('manufacturerName', '')
        model = search_info.get('modelName', '')
        tags = ",".join(upload_fields.get('tags', detail.get('tags', [])))

        # 이미지 (추가 이미지)
        images_obj = upload_fields.get('images', detail.get('images', {}))
        add_imgs = []
        if isinstance(images_obj, dict):
            opts = images_obj.get('optionalImages', [])
            for img in opts:
                u = img.get('url') if isinstance(img, dict) else img
                if u: add_imgs.append(u)
        add_img_str = ",".join(add_imgs)

        # 데이터 구조핑 (업로드용 전체 데이터 포함)
        row = {
            "불사자ID": pid,
            "상품명": pname,
            "그룹명": group_name,
            "안전여부": "O" if safety['is_safe'] else "X",
            "위험사유": unsafe_reason,
            "카테고리ID": cat_id,
            "판매가": sale_price,
            "재고수량": stock_qty,
            "마진율": margin_rate,
            "배송비": delivery_fee, # 시뮬레이터 로직상의 배송비
            "기본배송비": base_fee, # API 구조상의 기본배송비
            "배송비유형": deliv_type,
            "반품배송비": ret_fee,
            "교환배송비": exch_fee,
            "AS전화번호": as_tel,
            "AS안내": as_guide,
            "원산지코드": origin_code,
            "제조사": maker,
            "브랜드": brand,
            "모델명": model,
            "태그": tags,
            "전체옵션": len(skus),
            "미끼옵션": bait_count,
            "미끼상세": bait_str,
            "썸네일매칭": "O (옵션{})".format(chr(ord('A')+thumb_match_idx)) if thumb_match_idx >= 0 else "X",
            "썸네일목록": "|".join(thumbs), # 시각적 확인용
            "추가이미지": add_img_str, # 업로드용
            "대표썸네일": main_thumb,
            "상세설명HTML": detail_html,
            "옵션목록_RAW": json.dumps(skus, ensure_ascii=False),
            "원본데이터_JSON": json.dumps(upload_fields, ensure_ascii=False)
        }
        
        # 옵션 목록 텍스트화
        opt_lines = []
        for i, sku in enumerate(valid_skus):
            ko = sku.get('text_ko', '') or sku.get('text', '')
            cn = sku.get('text', '')
            price = sku.get('_origin_price', 0)
            
            markers = []
            if i == selected_idx: markers.append("(선택V)")
            if i == thumb_match_idx: markers.append("(썸네일동일)")
            
            marker_str = " " + " ".join(markers) if markers else ""
            
            line = f"{ chr(ord('A')+i) }. {ko}"
            if cn and cn != ko: line += f" ({cn})"
            line += f" [{price}]{marker_str}"
            opt_lines.append(line)
            
        row["옵션명"] = "\n".join(opt_lines)
        row["선택"] = selected_label
        
        return row

    def _save_to_excel(self, data, filename):
        if not EXCEL_AVAILABLE: return
        wb = Workbook()
        ws = wb.active
        ws.title = "상세정보"
        
    def _save_to_excel(self, data, filename):
        if not EXCEL_AVAILABLE: return
        wb = Workbook()
        ws = wb.active
        ws.title = "상세정보"
        
        headers = [
            "불사자ID", "그룹명", "상품명", "안전여부", "위험사유", 
            "카테고리ID", "판매가", "재고수량", "마진율", "배송비", "기본배송비", "배송비유형", "반품배송비", "교환배송비",
            "AS전화번호", "AS안내", "원산지코드", "제조사", "브랜드", "모델명", "태그",
            "미끼옵션", "미끼상세", "전체옵션", "선택", "옵션명", 
            "썸네일매칭", "썸네일목록", "추가이미지", "대표썸네일", "상세설명HTML", 
            "옵션목록_RAW", "원본데이터_JSON"
        ]
        ws.append(headers)
        
        # 스타일 정의
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="44546A", fill_type="solid") # 더 진한 남색
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                             top=Side(style='thin'), bottom=Side(style='thin'))

        # 헤더 스타일 적용
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
            # 행 높이 조절
            ws.row_dimensions[1].height = 40 # 헤더 높이 증가
        
        # 데이터 쓰기 및 스타일 적용
        for i, d in enumerate(data):
            row_idx = i + 2
            row_values = [d.get(h, "") for h in headers]
            ws.append(row_values)
            
            # 행 높이 자동 조절 (옵션명 등 내용이 길 경우)
            max_lines = 1
            opt_lines = str(d.get("옵션명", "")).count('\n') + 1
            bait_lines = str(d.get("미끼상세", "")).count('\n') + 1
            max_lines = max(opt_lines, bait_lines, 1)
            ws.row_dimensions[row_idx].height = max(20, min(150, max_lines * 15))

            # 셀 스타일 적용
            for col_idx, cell in enumerate(ws[row_idx], 1):
                col_key = headers[col_idx-1]
                cell.border = thin_border
                
                # 왼쪽 정렬이 필요한 컬럼
                if col_key in ['상품명', '위험사유', '미끼상세', '옵션명', '썸네일목록', '추가이미지', '상세설명HTML', '옵션목록_RAW', '원본데이터_JSON', 'AS안내', '태그']: 
                     cell.alignment = left_align
                else:
                     cell.alignment = center_align

        # 열 너비 조정 (가독성 최적화)
        # 기본 10, 중요 15~20, 텍스트 많음 30~50
        ws.column_dimensions['C'].width = 30 # 상품명
        ws.column_dimensions['E'].width = 20 # 위험사유
        ws.column_dimensions['W'].width = 30 # 미끼상세 (인덱스 주의)
        ws.column_dimensions['Z'].width = 40 # 옵션명
        ws.column_dimensions['AE'].width = 20 # 상세설명HTML
        ws.column_dimensions['AC'].width = 20 # 추가이미지
        
        # 틀 고정
        ws.freeze_panes = "D2" # 상품명부터 스크롤
            
        wb.save(filename)

    # ==================== 탭 2: 검수 (Inspector) 로직 ====================
    # ==================== 탭 2: 검수 (Inspector) 로직 ====================
    def _init_inspector_tab(self):
        # 1. 상단 툴바 (파일/저장/ AI분석)
        toolbar = ttk.Frame(self.tab_insp, padding=5)
        toolbar.pack(fill=tk.X)
        
        ttk.Button(toolbar, text="📂 엑셀 열기", command=self._insp_load_excel).pack(side=tk.LEFT, padx=5)
        ttk.Button(toolbar, text="💾 저장", command=self._insp_save_excel).pack(side=tk.LEFT, padx=5)
        ttk.Button(toolbar, text="✨ 전체 AI 분석", command=self._run_full_analysis).pack(side=tk.LEFT, padx=5)
        
        self.lbl_insp_file = ttk.Label(toolbar, text="(파일 없음)", foreground="gray")
        self.lbl_insp_file.pack(side=tk.LEFT, padx=10)
        
        self.lbl_insp_count = ttk.Label(toolbar, text="상품: 0개")
        self.lbl_insp_count.pack(side=tk.RIGHT, padx=20)

        # 2. 필터 영역
        filter_frame = ttk.Frame(self.tab_insp, padding=5)
        filter_frame.pack(fill=tk.X)
        
        ttk.Label(filter_frame, text="필터:").pack(side=tk.LEFT)
        
        self.filter_safe_var = tk.BooleanVar(value=True)
        self.filter_unsafe_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(filter_frame, text="안전", variable=self.filter_safe_var, command=self._render_inspector_data).pack(side=tk.LEFT, padx=5)
        ttk.Checkbutton(filter_frame, text="위험", variable=self.filter_unsafe_var, command=self._render_inspector_data).pack(side=tk.LEFT, padx=5)
        
        ttk.Separator(filter_frame, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=10)
        
        ttk.Label(filter_frame, text="그룹:").pack(side=tk.LEFT, padx=5)
        self.insp_group_combo = ttk.Combobox(filter_frame, width=20, state="readonly")
        self.insp_group_combo.pack(side=tk.LEFT, padx=5)
        self.insp_group_combo.bind("<<ComboboxSelected>>", lambda e: self._render_inspector_data())

        # 3. 메인 영역 (스크롤 Canvas)
        main_frame = ttk.Frame(self.tab_insp)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        self.canvas = tk.Canvas(main_frame, bg="white")
        scrollbar_y = ttk.Scrollbar(main_frame, orient=tk.VERTICAL, command=self.canvas.yview)
        scrollbar_x = ttk.Scrollbar(main_frame, orient=tk.HORIZONTAL, command=self.canvas.xview)
        
        self.scrollable_frame = ttk.Frame(self.canvas)
        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        )
        
        self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)
        
        scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)
        scrollbar_x.pack(side=tk.BOTTOM, fill=tk.X)
        self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # 마우스 휠 스크롤
        self.canvas.bind_all("<MouseWheel>", lambda e: self.canvas.yview_scroll(int(-1*(e.delta/120)), "units"))
        
        # 4. 하단 상태바
        self.lbl_insp_status = ttk.Label(self.tab_insp, text="대기 중...", relief=tk.SUNKEN, anchor='w')
        self.lbl_insp_status.pack(fill=tk.X, side=tk.BOTTOM)

        # 데이터 저장소
        self.inspector_data = [] # List[Dict]
        self.current_excel_path = None
        self.option_frames = {}
        self.image_cache = {}
        self.thumb_images = {} 
        
        self.trans_manager = TranslationManager()
        self.thumb_analyzer = ThumbnailAnalyzer()

    # ==================== 탭 2: 검수 (Inspector) 이벤트 및 렌더링 ====================
    def _insp_load_excel(self):
        """엑셀 파일 로드 (검수용)"""
        if not PANDAS_AVAILABLE:
             messagebox.showerror("오류", "pandas 모듈이 필요합니다.")
             return

        filepath = filedialog.askopenfilename(title="시뮬레이션 결과 파일 선택", filetypes=[("Excel files", "*.xlsx")])
        if not filepath: return
        
        try:
            self.lbl_insp_status.config(text=f"로딩 중: {os.path.basename(filepath)}...", foreground="blue")
            self.root.update()
            
            # pandas 로드
            try:
                xls = pd.ExcelFile(filepath, engine='openpyxl')
                if "상세정보" in xls.sheet_names:
                    df = pd.read_excel(filepath, sheet_name="상세정보", engine='openpyxl')
                else:
                    df = pd.read_excel(filepath, engine='openpyxl')
            except Exception as e:
                df = pd.read_excel(filepath)
            
            self._parse_excel_data_for_inspector(df)
            self.current_excel_path = filepath
            self.lbl_insp_file.config(text=os.path.basename(filepath), foreground="black")
            
            # 그룹 필터 업데이트
            groups = sorted(set(item.get("group_name", "") for item in self.inspector_data if item.get("group_name")))
            self.insp_group_combo['values'] = ["(전체)"] + groups
            self.insp_group_combo.current(0)
            
            self._render_inspector_data()
            self.lbl_insp_status.config(text=f"로드 완료: {len(self.inspector_data)}개 파싱됨", foreground="green")
            
        except Exception as e:
            messagebox.showerror("오류", f"파일 로드 실패: {e}")
            self.lbl_insp_status.config(text="로드 실패", foreground="red")

    def _parse_excel_data_for_inspector(self, df):
        """엑셀 데이터를 검수 UI용 구조로 파싱"""
        self.inspector_data = []
        for idx, row in df.iterrows():
            try:
                # 안전 문자열 변환
                def safe_str(val): return str(val).strip() if pd.notna(val) else ""
                
                # 이미지 URL 추출
                thumb_raw = safe_str(row.get("썸네일\n이미지", "") or row.get("메인썸네일URL", ""))
                thumb_url = thumb_raw
                if thumb_raw.startswith('=IMAGE("') and thumb_raw.endswith('")'):
                    thumb_url = thumb_raw[8:-2]
                elif thumb_raw.startswith('http'):
                    thumb_url = thumb_raw
                else:
                    thumb_url = ""

                # 안전 여부 판단
                is_safe_val = safe_str(row.get("안전여부", "O")).upper()
                is_safe = is_safe_val in ["O", "안전", "TRUE", "1", "OK"]
                
                item = {
                    "row_idx": idx,
                    "product_name": safe_str(row.get("상품명", ""))[:40],
                    "product_id": safe_str(row.get("불사자ID", "") or row.get("상품ID", "")),
                    "is_safe": is_safe,
                    "unsafe_reason": safe_str(row.get("위험사유", ""))[:30],
                    "group_name": safe_str(row.get("그룹", "") or row.get("그룹명", "")),
                    "thumbnail_url": thumb_url,
                    "total_options": int(row.get("전체옵션", 0)) if pd.notna(row.get("전체옵션")) else 0,
                    "final_options": int(row.get("최종옵션", 0)) if pd.notna(row.get("최종옵션")) else 0,
                    "bait_options": int(row.get("미끼옵션", 0)) if pd.notna(row.get("미끼옵션")) else 0,
                    "main_option": safe_str(row.get("대표옵션", "")),
                    "selected": safe_str(row.get("선택", "A")).upper() or "A",
                    "option_raw": safe_str(row.get("옵션명", "") or row.get("최종옵션목록", "")),
                    
                    # [NEW] SOTA 결과 저장용
                    "sota_score": 0,
                    "sota_text": "",
                    "nukki_status": "none" # none, done
                }
                
                # 옵션 목록 파싱
                options = []
                if item["option_raw"]:
                    lines = item["option_raw"].split('\n')
                    for i, line in enumerate(lines):
                        if not line.strip(): continue
                        label = chr(ord('A') + i) if i < 26 else str(i+1)
                        if '. ' in line:
                            parts = line.split('. ', 1)
                            label = parts[0].strip()
                            name = parts[1].strip() if len(parts) > 1 else ""
                        else:
                            name = line.strip()
                        options.append({"label": label, "name": name})
                item["options"] = options
                item["option_count_str"] = f"{item['final_options']}/{item['total_options']}"
                
                self.inspector_data.append(item)
                
            except Exception as e:
                print(f"Row {idx} parsing error: {e}")

    def _render_inspector_data(self):
        """스크롤 프레임에 데이터 렌더링"""
        # 기존 위젯 제거
        for widget in self.scrollable_frame.winfo_children():
            widget.destroy()
        
        self.option_frames = {}
        
        if not self.inspector_data:
            ttk.Label(self.scrollable_frame, text="데이터가 없습니다.", font=("맑은 고딕", 14)).pack(pady=50)
            self.lbl_insp_count.config(text="상품: 0개")
            return

        # 필터링
        filtered = []
        target_group = self.insp_group_combo.get()
        show_safe = self.filter_safe_var.get()
        show_unsafe = self.filter_unsafe_var.get()
        
        for item in self.inspector_data:
            if item["is_safe"] and not show_safe: continue
            if not item["is_safe"] and not show_unsafe: continue
            if target_group and target_group != "(전체)" and item["group_name"] != target_group: continue
            filtered.append(item)
            
        self.lbl_insp_count.config(text=f"상품: {len(filtered)} / {len(self.inspector_data)}개")
        
        if not filtered:
            ttk.Label(self.scrollable_frame, text="표시할 상품이 없습니다 (필터 확인).", font=("맑은 고딕", 12)).pack(pady=50)
            return

        # 헤더 생성
        self._insp_create_header()
        
        # 행 생성 (최대 100개까지만 렌더링 권장? 아니면 페이지네이션? 일단 스크롤)
        # 성능 이슈 방지를 위해 처음 200개만 렌더링하고 '더보기' 버튼을 두는게 좋을 수도 있음.
        # 하지만 v3.1은 다 뿌림. 사용자가 수천개라고 했으니... Canvas window는 많아지면 느려짐.
        # 페이지네이션 없이 일단 300개 제한 혹은 전체 렌더링 시도.
        # "수천개 검수" -> tkinter widget 수천개는 매우 느림. 
        # 페이징 도입이 필수적이나, 일단 v3.1 로직(전부 렌더링) 따름.
        for item in filtered:
            self._insp_create_row(item)

    def _insp_create_header(self):
        h_frame = tk.Frame(self.scrollable_frame, bg="#4472C4")
        h_frame.pack(fill=tk.X, pady=(0, 2))
        
        headers = [
            ("이미지/누끼", 120), ("옵션 선택 (A,B,C...)", 450), 
            ("상품정보", 300), ("상태", 60), ("통계", 80)
        ]
        
        for text, w in headers:
            tk.Label(h_frame, text=text, width=w//8, bg="#4472C4", fg="white", font=("맑은 고딕", 9, "bold"), pady=5).pack(side=tk.LEFT, padx=1)
            
    def _on_select_product(self, event):
        sel = self.tree.selection()
        if not sel: return
        idx = int(sel[0])
        self.current_insp_idx = idx
        data = self.insp_data[idx]
        
        # 1. 썸네일 렌더링
        for w in self.frame_thumbs_inner.winfo_children(): w.destroy()
        self.thumb_images = {} # clear cache
        
        thumbs_str = str(data.get("썸네일목록", ""))
        thumbs = [t for t in thumbs_str.split("|") if t.strip()]
        if not thumbs: thumbs = []
        
        current_main = data.get("대표썸네일", "")
        
        for i, url in enumerate(thumbs):
            self._display_thumbnail_item(i, url, current_main == url)
        
        # 2. 옵션 표시
        self.txt_options.delete("1.0", tk.END)
        self.txt_options.insert("1.0", str(data.get("옵션명", "")))
        
        # 스크롤 초기화
        self.cv_thumb.xview_moveto(0)

    def _display_thumbnail_item(self, idx, url, is_main):
        """썸네일 개별 항목 표시 (이미지 로드는 비동기)"""
        f_item = tk.Frame(self.frame_thumbs_inner, bd=3, relief="solid" if is_main else "flat", bg="white")
        f_item.pack(side=tk.LEFT, padx=5, pady=5)
        
        # 플레이스홀더
        self._on_select_product(None)
        
        # 트리뷰 업데이트 (변경됨 표시)
        # self.tree.item(self.current_insp_idx, tags=("modified",))

    def _standardize_current_options(self):
        """현재 상품의 옵션명을 사전 기반으로 표준화"""
        current_text = self.txt_options.get("1.0", tk.END).strip()
        if not current_text: return
        
        # 원본 데이터(JSON)에서 중국어 추출이 가장 정확하지만, 화면 텍스트 파싱도 가능
        # 화면 포맷: A. 한국어명(중국어명) [가격]
        
        new_lines = []
        lines = current_text.split('\n')
        changed_count = 0
        
        for line in lines:
            # 정규식으로 중국어 추출: ([\u4e00-\u9fff]+)
            match = re.search(r'\(([\u4e00-\u9fff]+)\)', line)
            if match:
                cn_word = match.group(1)
                # 한국어 부분 추출 (A. 와 (중국어) 사이)
                # A. 블라블라 (중국어) -> "블라블라"
                pre_part = line.split('(')[0] # A. 블라블라 
                
                # 라벨(A.) 분리
                label_match = re.match(r'^([A-Z]\.\s*)', pre_part)
                label = label_match.group(1) if label_match else ""
                kr_part = pre_part[len(label):].strip()
                
                # 표준화
                std_kr = self.trans_manager.standardize(cn_word, kr_part)
                
                if std_kr != kr_part:
                    # 변경됨: A. 표준어(중국어) [가격]
                    # 뒷부분(가격 등) 유지
                    post_part = line[match.end():] # ) [가격]
                    new_line = f"{label}{std_kr}({cn_word}){post_part}"
                    new_lines.append(new_line)
                    changed_count += 1
                else:
                    new_lines.append(line)
            else:
                new_lines.append(line)
                
        if changed_count > 0:
            new_text = "\n".join(new_lines)
            self.txt_options.delete("1.0", tk.END)
            self.txt_options.insert("1.0", new_text)
            
            # 데이터 반영
            self.insp_data[self.current_insp_idx]["옵션명"] = new_text
            messagebox.showinfo("완료", f"{changed_count}개 옵션명이 표준화되었습니다.")
        else:
            messagebox.showinfo("알림", "적용할 표준어가 없거나 이미 표준화되어 있습니다.")

    def _insp_save_excel(self):
        if not self.current_file: return
        
        try:
            df = pd.DataFrame(self.insp_data)
            df.to_excel(self.current_file, index=False)
            messagebox.showinfo("저장 완료", "엑셀 파일이 저장되었습니다.")
        except Exception as e:
            messagebox.showerror("실패", f"저장 중 오류: {e}")

    def _make_selected_nukki(self):
        """현재 선택된 대표 이미지의 배경을 제거 (rembg)"""
        if self.current_insp_idx < 0: return
        if not REMBG_AVAILABLE:
            messagebox.showwarning("불가", "rembg 라이브러리가 설치되지 않았습니다.\npip install rembg")
            return
            
        data = self.insp_data[self.current_insp_idx]
        current_url = data.get("대표썸네일", "")
        if not current_url:
            messagebox.showwarning("알림", "선택된 대표 이미지가 없습니다.")
            return
            
        try:
            # 1. 이미지 로드
            if current_url.startswith("http"):
                resp = requests.get(current_url, timeout=10)
                img_bytes = resp.content
            else:
                # 로컬 파일
                with open(current_url, "rb") as f:
                    img_bytes = f.read()
            
            # 2. 누끼 생성 (알림)
            messagebox.showinfo("진행 중", "배경 제거 작업 중입니다... 잠시만 기다려주세요.")
            
            # 메인 스레드에서 실행 (GUI 멈춤 감수 - rembg는 무거움)
            # 개선: 별도 스레드에서 실행하고 완료 시 콜백 호출이 좋으나, 구조상 일단 동기 실행
            nukki_bytes = self.thumb_analyzer.make_nukki(img_bytes)
            
            if not nukki_bytes:
                messagebox.showerror("실패", "배경 제거에 실패했습니다.")
                return

            # 3. 저장
            cache_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "nukki_cache")
            os.makedirs(cache_dir, exist_ok=True)
            
            pid = data.get("불사자ID", "unknown")
            fname = f"{pid}_{int(time.time())}_nukki.png"
            save_path = os.path.join(cache_dir, fname)
            
            with open(save_path, "wb") as f:
                f.write(nukki_bytes)
                
            # 4. 데이터 업데이트 (대표 이미지를 로컬 경로로 변경)
            self._set_main_thumbnail(save_path)
            
            messagebox.showinfo("완료", "배경 제거 완료!\n대표 이미지가 누끼 이미지로 변경되었습니다.")
            
        except Exception as e:
            messagebox.showerror("오류", f"작업 중 오류 발생: {e}")

    def _run_full_analysis(self):
        """전체 항목에 대해 AI 분석(OCR/누끼) 실행"""
        if not self.inspector_data:
            messagebox.showwarning("경고", "데이터가 없습니다.")
            return
            
        if not messagebox.askyesno("확인", f"총 {len(self.inspector_data)}개 상품에 대해 AI 정밀 분석을 시작합니까?\n(시간이 다소 소요될 수 있습니다)"):
            return

        def _task():
            total = len(self.inspector_data)
            success = 0
            
            for i, item in enumerate(self.inspector_data):
                url = item["thumbnail_url"]
                if not url: continue
                
                try:
                    # 이미지 다운로드
                    res = requests.get(url, timeout=5)
                    img_bytes = res.content
                    
                    # CV2 변환
                    nparr = np.frombuffer(img_bytes, np.uint8)
                    img = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
                    
                    if img is None: continue
                    
                    # 분석 실행
                    # ThumbnailAnalyzer.analyzed_score 는 cv2 이미지를 받음
                    result = self.thumb_analyzer.analyzed_score(img)
                    
                    # 결과 저장
                    score = result["score"]
                    rec = result["recommendation"]
                    
                    sota_msg = []
                    if result["is_nukki"]: sota_msg.append("✨누끼됨")
                    if result["has_text"]: sota_msg.append("📝텍스트")
                    sota_msg.append(f"점수:{score}")
                    
                    item["sota_score"] = score
                    item["sota_text"] = " ".join(sota_msg)
                    item["nukki_status"] = "done" if result["is_nukki"] else "none"
                    
                    success += 1
                    
                    # UI 업데이트 (10개마다)
                    if i % 10 == 0:
                        self.lbl_insp_status.config(text=f"AI 분석 중... {i+1}/{total} (성공: {success})")
                        
                except Exception as e:
                    print(f"Item {i} analyze failed: {e}")
                    
            self.lbl_insp_status.config(text=f"AI 분석 완료. (총 {success}개 처리)", foreground="blue")
            
            # UI 리프레시 (메인스레드에서)
            self.root.after(0, self._render_inspector_data)

        threading.Thread(target=_task, daemon=True).start()

    def _insp_save_excel(self):
        """변경된 선택 옵션을 엑셀에 저장"""
        if not self.current_excel_path:
             messagebox.showwarning("경고", "열린 파일이 없습니다.")
             return
             
        try:
            from openpyxl import load_workbook
            wb = load_workbook(self.current_excel_path)
            if "상세정보" in wb.sheetnames:
                ws = wb["상세정보"]
            else:
                ws = wb.active
                
            # '선택' 컬럼 찾기
            header_row = 1
            sel_col = None
            for col in range(1, ws.max_column+2):
                h = ws.cell(row=header_row, column=col).value
                if str(h).strip() == "선택":
                    sel_col = col
                    break
            
            if not sel_col:
                messagebox.showerror("오류", "'선택' 컬럼을 엑셀에서 찾을 수 없습니다.")
                return
            
            count = 0
            for item in self.inspector_data:
                # item["row_idx"]는 DataFrame index (0부터 시작)
                # 엑셀 헤더가 1행이면 데이터는 2행부터. pd.read_excel이 0-index면 엑셀 행은 idx + 2
                r = item["row_idx"] + 2
                val = item["selected"]
                
                # 기존 값과 다르면 업데이트 (여기선 그냥 덮어쓰기)
                ws.cell(row=r, column=sel_col, value=val)
                count += 1
                
            wb.save(self.current_excel_path)
            messagebox.showinfo("저장 완료", f"총 {len(self.inspector_data)}개 상품의 선택값을 저장했습니다.")
            
        except Exception as e:
            messagebox.showerror("저장 실패", f"오류 발생: {e}")

if __name__ == "__main__":
    root = tk.Tk()
    app = BulsajaSimulatorV2(root)
    root.mainloop()
