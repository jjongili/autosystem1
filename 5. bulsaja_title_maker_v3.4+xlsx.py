# -*- coding: utf-8 -*-
"""
불사자 상품명 자동 변환기+검수 v3.0 (API 버전)
- v2.5 로직 기반 + Selenium → API 교체
- 백그라운드 동작 (브라우저 조작 없음)
- 속도 10배 이상 향상

기존 기능 유지:
- 가격추적기 이미지 검색 기반 상품명 생성
- 유사도 비교 + 이미지 검증 로직
- Claude 상품명 1개 + 키워드 20개 생성, Python이 2,3번 생성
- Google Sheets 금지단어/예외단어 동기화

by 프코노미
"""

import os
import re
import time
import threading
import json
import csv
import subprocess
import base64
import requests
import websocket
from datetime import datetime
from typing import List, Optional, Tuple, Dict, Set
from dataclasses import dataclass, field
from difflib import SequenceMatcher

import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox

# Selenium - 알리프라이스 검색용으로만 사용 (선택적)
try:
    from selenium import webdriver
    from selenium.webdriver.chrome.service import Service
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.common.by import By
    from selenium.webdriver.common.keys import Keys
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.common.action_chains import ActionChains
    from webdriver_manager.chrome import ChromeDriverManager
    SELENIUM_AVAILABLE = True
except ImportError:
    SELENIUM_AVAILABLE = False

import anthropic

# openpyxl for xlsx with colors
try:
    from openpyxl import Workbook
    from openpyxl.cell.text import InlineFont
    from openpyxl.cell.rich_text import TextBlock, CellRichText
    from openpyxl.styles import Font, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# ==================== v11 신규 설정 ====================
SIMILARITY_THRESHOLD = 0.3  # 유사도 30% 미만이면 이미지 검증
IMAGE_MATCH_THRESHOLD = 0.7  # Vision 신뢰도 70% 미만이면 불일치
SECOND_CHECK_TAG = "2차검수"  # 이미지 불일치시 태그

# ==================== 설정 ====================
BANNED_ADJECTIVES = ["강력한", "프리미엄", "고급", "최신형", "완벽한", "최고의"]
MIN_CHARS = 30
MAX_CHARS = 40
CONFIG_FILE = "bulsaja_config_claude.json"  # Claude 전용 config
BANNED_WORDS_FILE = "banned_words.json"
EXCLUDED_WORDS_FILE = "excluded_words.json"  # 탐지 제외 단어
REMOVE_WORDS_FILE = "remove_words.json"  # 상품명에서 무조건 제거할 단어
DEBUG_PORT = 9222
DEBUG_PORT_RANGE = (9222, 9240)  # 포트 탐색 범위
BULSAJA_PRODUCT_LIST_URL = "https://www.bulsaja.com/products/manage/list/"
CHROME_DEBUG_PROFILE = "C:\\chrome_debug_profile"
ALIPRICE_PROFILE = os.path.join(os.path.expanduser("~"), "aliprice_chrome_profile_titlemaker")

def cleanup_profile_locks(profile_path: str):
    """Chrome 프로필 잠금 파일 정리 - 충돌 방지"""
    if not os.path.exists(profile_path):
        return
    lock_files = ['SingletonLock', 'SingletonSocket', 'SingletonCookie']
    for lock_file in lock_files:
        lock_path = os.path.join(profile_path, lock_file)
        if os.path.exists(lock_path):
            try:
                os.remove(lock_path)
            except:
                pass

def find_available_port(start_port: int = 9222, end_port: int = 9240) -> int:
    """사용 가능한 포트 찾기"""
    import socket
    for port in range(start_port, end_port + 1):
        try:
            with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
                s.settimeout(0.5)
                result = s.connect_ex(('127.0.0.1', port))
                if result != 0:  # 포트가 사용 중이 아님
                    return port
        except:
            continue
    return start_port  # 기본값 반환

# ==================== 상품명 로직 선택 ====================
TITLE_LOGIC_OPTIONS = {
    "기본형": "basic",           # 메인키워드 앞 배치 (경쟁사 있으면 참고)
    "용도+메인": "usage_main",   # [용도/장소] + [메인키워드] + [세부] 구조
}

# ==================== 위험상품 카테고리 ====================
# 성인/성관련 단어
ADULT_KEYWORDS = {
    "성인용품", "섹시", "란제리", "가터벨트",
    "코르셋", "나이트웨어", "베이비돌", "테디", "캐미솔",
    "시스루", "노출", "야한", "에로", "19금",
    "콘돔", "러브젤", "바이브", "딜도", "오나홀", "리얼돌",
    "SM", "본디지", "채찍", "페티쉬", "코스프레", "메이드복",
    "끈팬티", "티팬티", "누드", "벗방",
}

# 의료기기/의료 관련 단어
MEDICAL_KEYWORDS = {
    "의료기기", "의료용", "의약품", "약품", "처방", "혈압계", "혈당계",
    "체온계", "산소포화도", "맥박", "심전도", "엑스레이", "MRI", "CT",
    "주사기", "주사바늘", "수액", "링거", "카테터", "스텐트", "임플란트",
    "보청기", "콘택트렌즈", "시력교정", "치과", "교정기", "틀니", "의치",
    "휠체어", "목발", "의족", "의수", "석고붕대", "깁스",
    "소독제", "살균제", "멸균", "수술", "마취", "진통제", "해열제", "항생제",
    "스테로이드", "호르몬제", "피임약", "발기부전", "탈모약", "다이어트약",
    "건강기능식품", "영양제", "비타민", "홍삼", "프로폴리스", "오메가3",
    "유산균", "프로바이오틱스", "콜라겐", "히알루론산", "글루코사민",
    "LED마스크", "피부관리기", "IPL",
}

# 유아/아동 관련 (구매대행 금지)
CHILD_KEYWORDS = {
    "유아용", "유아", "아기", "신생아", "영아", "베이비", "baby", "infant",
    "젖병", "분유", "이유식", "기저귀", "물티슈", "아기띠", "캐리어",
    "카시트", "바운서", "보행기", "점퍼루", "쏘서",
    "유아복", "아기옷", "배냇저고리", "턱받이", "손싸개", "발싸개",
    "장난감", "토이", "인형", "레고",
    "수유", "모유", "수유쿠션", "수유브라", "유축기", "젖꼭지",
    "치발기", "공갈젖꼭지", "노리개", "딸랑이", "모빌",
    "아동복", "키즈", "주니어", "어린이", "초등", "유치원",
}

# 판매금지/규제 상품
PROHIBITED_KEYWORDS = {
    # 자전거 관련 (KC인증 필요)
    "픽시", "픽시자전거", "fixie", "fixed gear", "고정기어",
    "전동킥보드", "전동휠", "전동스쿠터", "전기자전거", "전동자전거",
    # 안전인증 필요
    "가스렌지", "가스레인지", "전기장판", "전기매트", "전기담요",
    "온수매트", "전기히터", "석유난로", "가스난로",
    # 식품/건강 관련 (단독으로 쓰일 때만)
    "식품", "과자", "사탕", "초콜릿", "식용", "먹는",
    "화장품", "스킨케어", "로션", "세럼", "에센스", "마스크팩",
    "샴푸", "린스", "트리트먼트", "바디워시", "치약", "구강",
    # 위험물 (단독으로 쓰일 때만)
    "도검", "총", "비비탄", "석궁",
    "화약", "폭죽", "불꽃놀이", "가스통",
    # 기타 규제
    "담배", "전자담배", "액상", "니코틴", "베이프", "vape",
    "주류", "술", "소주", "보드카",
    "도박", "슬롯", "베팅",
    "복제", "짝퉁", "이미테이션", "레플리카", "replica", "fake",
}

# 문맥에 따라 안전한 복합어 패턴 (이 패턴이 있으면 위험단어 무시)
# 키: 위험단어, 값: 함께 있으면 안전한 단어들
SAFE_CONTEXT_PATTERNS = {
    "음식": ["음식물", "분리수거", "쓰레기", "처리기", "분쇄기"],  # 음식물 쓰레기통 OK
    "칼": ["꽂이", "거치대", "보관", "수납", "케이스", "블록", "세트홀더"],  # 칼꽂이 OK
    "나이프": ["꽂이", "거치대", "보관", "수납", "케이스", "블록", "홀더"],  # 나이프블록 OK
    "크림": ["보관", "케이스", "용기", "디스펜서"],  # 크림 용기 OK (화장품 아님)
    "라이터": ["케이스", "보관", "거치대"],  # 라이터 케이스 OK
    "성냥": ["케이스", "보관", "통"],  # 성냥통 OK
}

# 동의어 매핑 (완전히 같은 의미의 단어만 - 중복 제거용)
SYNONYM_HINTS = {
    "의자": ["체어", "chair"],
    "체어": ["의자"],
    "chair": ["의자", "체어"],
    "메쉬": ["메시", "mesh"],
    "메시": ["메쉬"],
    "mesh": ["메쉬", "메시"],
    "휴지통": ["쓰레기통"],
    "쓰레기통": ["휴지통"],
    "접이식": ["폴딩"],
    "폴딩": ["접이식"],
    "야외": ["실외", "아웃도어"],
    "실외": ["야외"],
    "아웃도어": ["야외"],
    "계란": ["달걀"],
    "달걀": ["계란"],
    "노트북": ["랩탑", "랩톱"],
    "랩탑": ["노트북"],
    "랩톱": ["노트북"],
    "핸드폰": ["휴대폰", "스마트폰"],
    "휴대폰": ["핸드폰"],
    "스마트폰": ["핸드폰", "휴대폰"],
    "선반": ["랙", "rack"],
    "랙": ["선반"],
    "rack": ["선반", "랙"],
    "조명": ["등", "라이트", "light"],
    "등": ["조명"],
    "라이트": ["조명"],
    "light": ["조명"],
    "텐트": ["천막"],
    "천막": ["텐트"],
    "캐노피": ["차양", "그늘막"],
    "차양": ["캐노피"],
    "그늘막": ["캐노피"],
    "워터파크": ["물놀이"],
    "물놀이": ["워터파크"],
    "돔쉘터": ["돔텐트"],
    "돔텐트": ["돔쉘터"],
    "대형": ["초대형", "빅사이즈"],
    "초대형": ["대형"],
}

# 이미지 확장자 목록
IMAGE_EXTENSIONS = ('.jpg', '.jpeg', '.png', '.gif', '.webp', '.bmp', '.tiff')

# 일반적인 한글 단어 (브랜드가 아닌 것으로 판단)
COMMON_KOREAN_WORDS = {
    # 일반 명사
    "가방", "거치대", "걸이", "공구", "그릇", "기계", "꽃병", "나무", "냄비",
    "다용도", "대형", "도구", "돌솥", "등받이", "라벨", "램프", "랙", "마감",
    "매트", "머그", "모형", "몰딩", "미니", "바구니", "바닥", "받침", "발판",
    "방수", "배수", "베개", "벽걸이", "보관", "보드", "봉투", "부엌", "분리",
    "블럭", "사무", "사이즈", "상자", "선반", "세트", "소형", "손잡이", "수납",
    "스탠드", "스틸", "슬라이드", "시트", "식기", "실리콘", "쌀통", "아기",
    "앞치마", "액자", "양면", "어린이", "업소", "여행", "연필", "오픈", "옷걸이",
    "우산", "원목", "원형", "유리", "의자", "이동식", "인테리어", "일회용",
    "자석", "잠금", "장난감", "장식", "전기", "접이식", "조립", "주방", "중형",
    "지갑", "직사각", "진열", "찬장", "책상", "책장", "철제", "청소", "체어",
    "추억", "출입", "측면", "카드", "카트", "캐비넷", "커버", "컵", "케이스",
    "코너", "코팅", "크기", "탁상", "태양", "테이블", "통풍", "트레이", "트롤리",
    "파티", "판매", "패드", "펜꽂이", "편지", "포장", "폴더", "플라스틱", "필통",
    "핸드", "행거", "향수", "홀더", "화분", "화이트", "확장", "휴대용", "휴지",
    # 색상
    "블랙", "화이트", "그레이", "브라운", "베이지", "네이비", "레드", "블루",
    "그린", "옐로우", "핑크", "퍼플", "오렌지", "골드", "실버", "아이보리",
    # 재질
    "우드", "메탈", "패브릭", "가죽", "스테인리스", "알루미늄", "아크릴",
    # 스타일
    "모던", "빈티지", "앤틱", "클래식", "심플", "내추럴", "미니멀", "북유럽",
    # 기능
    "방수", "접이식", "이동식", "휴대용", "다용도", "멀티", "자동", "수동",
    # 기타 흔한 단어
    "세트", "용품", "소품", "악세사리", "액세서리", "데코", "장식품",
}

# 상품명 생성 방식
TITLE_MODE_IMAGE_FIRST = "image_first"  # 이미지 + 기존상품명 (1순위 이미지) - AliPrice 실제 검색
TITLE_MODE_VISION = "vision"  # Claude Vision 이미지 분석 (타오바오 검색 없이)
TITLE_MODE_ORIGINAL_ONLY = "original_only"  # 기존상품명만

# ==================== 유명 브랜드 리스트 (2차검수 대상) ====================
FAMOUS_BRANDS = {
    # 글로벌 스포츠 브랜드
    "nike", "adidas", "puma", "reebok", "newbalance", "underarmour", "fila",
    "converse", "vans", "asics", "mizuno", "umbro", "kappa", "lotto",
    "나이키", "아디다스", "퓨마", "리복", "뉴발란스", "언더아머", "휠라",
    "컨버스", "반스", "아식스", "미즈노", "엄브로", "카파", "로또",
    # 명품 브랜드
    "gucci", "chanel", "louisvuitton", "prada", "hermes", "dior", "burberry",
    "versace", "armani", "balenciaga", "bottega", "celine", "fendi", "givenchy",
    "loewe", "moncler", "saintlaurent", "valentino", "ferragamo", "tiffany",
    "cartier", "bulgari", "rolex", "omega", "patek", "iwc", "breitling",
    "구찌", "샤넬", "루이비통", "프라다", "에르메스", "디올", "버버리",
    "베르사체", "아르마니", "발렌시아가", "보테가", "셀린느", "펜디", "지방시",
    "로에베", "몽클레어", "생로랑", "발렌티노", "페라가모", "티파니",
    "까르띠에", "불가리", "롤렉스", "오메가", "파텍", "브라이틀링",
    # 전자/IT 브랜드
    "apple", "samsung", "sony", "lg", "panasonic", "philips", "bose",
    "dyson", "xiaomi", "huawei", "dell", "hp", "lenovo", "asus", "acer",
    "애플", "삼성", "소니", "엘지", "파나소닉", "필립스", "보스",
    "다이슨", "샤오미", "화웨이", "델", "레노버", "아수스", "에이서",
    # 캐릭터/엔터테인먼트
    "disney", "marvel", "pokemon", "sanrio", "hellokitty", "kakao", "line",
    "nintendo", "playstation", "xbox", "bandai", "lego", "barbie", "transformer",
    "디즈니", "마블", "포켓몬", "산리오", "헬로키티", "카카오", "라인",
    "닌텐도", "플레이스테이션", "엑스박스", "반다이", "레고", "바비", "트랜스포머",
    "짱구", "뽀로로", "타요", "핑크퐁", "아기상어", "원피스", "나루토", "드래곤볼",
    # 아웃도어/스포츠
    "northface", "patagonia", "columbia", "arcteryx", "mammut", "salomon",
    "노스페이스", "파타고니아", "콜롬비아", "아크테릭스", "마무트", "살로몬",
    "blackyak", "kolon", "eider", "lafuma", "millet", "k2",
    "블랙야크", "코오롱", "아이더", "라푸마", "밀레", "케이투",
    # 국내 유명 브랜드
    "농심", "오뚜기", "cj", "풀무원", "동원", "삼양", "오리온", "롯데",
    "아모레퍼시픽", "lg생활건강", "이니스프리", "설화수", "라네즈", "헤라",
    # 자동차
    "bmw", "mercedes", "audi", "porsche", "ferrari", "lamborghini", "bentley",
    "현대", "기아", "제네시스", "벤츠", "아우디", "포르쉐", "페라리", "람보르기니",
}

# ==================== 일반 영어 단어 (제거하면 안 됨) ====================
COMMON_ENGLISH_WORDS = {
    # 크기/용량
    "mini", "micro", "small", "medium", "large", "big", "xl", "xxl",
    "slim", "compact", "portable", "lite", "light",
    # 기능/특성
    "pro", "plus", "max", "ultra", "super", "smart", "auto", "manual",
    "digital", "analog", "electric", "wireless", "bluetooth", "wifi",
    "usb", "led", "lcd", "hd", "fhd", "uhd", "oled", "qled",
    "waterproof", "dustproof", "shockproof", "fireproof",
    "foldable", "folding", "adjustable", "flexible", "portable",
    "rechargeable", "cordless", "battery", "solar", "magnetic",
    "silent", "quiet", "noise", "mute", "sound",
    # 용도/장소
    "home", "office", "outdoor", "indoor", "camping", "travel", "hiking",
    "kitchen", "bathroom", "bedroom", "living", "garden", "garage",
    "car", "bike", "desk", "table", "wall", "floor", "door", "window",
    # 재질
    "plastic", "metal", "wood", "wooden", "steel", "iron", "aluminum",
    "glass", "silicon", "silicone", "rubber", "leather", "fabric", "cotton",
    "stainless", "chrome", "brass", "copper", "zinc", "titanium",
    # 색상
    "black", "white", "gray", "grey", "red", "blue", "green", "yellow",
    "pink", "purple", "orange", "brown", "beige", "navy", "gold", "silver",
    # 형태
    "round", "square", "rectangle", "circle", "oval", "triangle",
    "flat", "curved", "straight", "long", "short", "wide", "narrow",
    # 기타 일반 단어
    "set", "kit", "pack", "box", "case", "cover", "holder", "stand",
    "rack", "shelf", "hook", "clip", "mount", "bracket", "hanger",
    "cup", "mug", "bottle", "pot", "pan", "bowl", "plate", "dish",
    "bag", "pouch", "basket", "bin", "container", "storage", "organizer",
    "tool", "device", "machine", "equipment", "accessory", "part",
    "new", "type", "style", "version", "model", "series", "edition",
    "multi", "dual", "double", "single", "triple", "pair",
    # 동사/형용사 파생
    "cleaning", "cooking", "charging", "cutting", "folding", "hanging",
    "rolling", "sliding", "spinning", "rotating", "tilting", "lifting",
}

# ==================== 영어/숫자 패턴 분류 함수 ====================
def classify_english_pattern(word: str) -> tuple:
    """
    영어/숫자 패턴을 분류하여 처리 방법 결정
    Returns: (분류, 처리방법, 설명)
        - 분류: FAMOUS_BRAND, COMMON_WORD, MODEL_NUMBER, SELLER_BRAND, NUMBER, UNKNOWN
        - 처리: review(2차검수), keep(유지), remove(제거), suspect(의심단어로)
    """
    if not word or len(word) < 2:
        return ("UNKNOWN", "keep", "")
    
    word_lower = word.lower().replace(" ", "").replace("-", "")
    word_clean = word.strip()
    
    # 1. 유명 브랜드 체크 → 2차검수
    if word_lower in FAMOUS_BRANDS:
        return ("FAMOUS_BRAND", "review", f"유명브랜드: {word}")
    
    # 2. 일반 영어 단어 체크 → 유지
    if word_lower in COMMON_ENGLISH_WORDS:
        return ("COMMON_WORD", "keep", "")
    
    # 3. 순수 숫자 → 제거
    if word_clean.isdigit():
        return ("NUMBER", "remove", f"숫자: {word}")
    
    # 4. 모델명/제품번호 패턴 → 제거
    # XK-2024, T500, A380, BT21 등
    if re.match(r'^[A-Za-z]{1,3}[-]?\d{2,5}$', word_clean, re.IGNORECASE):
        return ("MODEL_NUMBER", "remove", f"모델번호: {word}")
    # 500ML, 2024VER, 12V 등
    if re.match(r'^\d+[A-Za-z]{1,4}$', word_clean):
        return ("MODEL_NUMBER", "remove", f"스펙: {word}")
    # V2, V3, X1 등 (버전/모델)
    if re.match(r'^[VXS]\d{1,2}$', word_clean, re.IGNORECASE):
        return ("MODEL_NUMBER", "remove", f"버전: {word}")
    
    # 5. 영문만 4-10자 (사전에 없음) → 셀러 브랜드로 의심 → 의심단어로
    if re.match(r'^[A-Za-z]{4,10}$', word_clean):
        # 발음 가능한지 체크 (자음만 연속 3개 이상이면 의미없는 단어)
        consonants = re.findall(r'[bcdfghjklmnpqrstvwxz]{3,}', word_lower)
        if consonants:
            return ("SELLER_BRAND", "remove", f"셀러브랜드의심: {word}")
        else:
            # 발음 가능하지만 사전에 없음 → 의심단어로 보내서 사람이 판단
            return ("UNKNOWN_ENGLISH", "suspect", f"미확인영어: {word}")
    
    # 6. 영문+숫자 혼합 (모델명 패턴 아닌 것) → 제거
    if re.match(r'^[A-Za-z]+\d+[A-Za-z]*$', word_clean) or re.match(r'^\d+[A-Za-z]+\d*$', word_clean):
        return ("MODEL_NUMBER", "remove", f"모델명: {word}")
    
    # 7. 나머지 → 유지
    return ("UNKNOWN", "keep", "")


def process_suspect_words(words: list, remove_words_set: set = None, excluded_words_set: set = None) -> dict:
    """
    의심단어 리스트를 분류하여 처리
    Args:
        words: 의심단어 리스트
        remove_words_set: 구글시트 제거단어 셋 (있으면 해당 단어는 'remove'로 분류)
        excluded_words_set: 예외단어 셋 (있으면 해당 단어는 의심단어에서 제외)
    Returns: {
        'review': [(word, reason), ...],     # 2차검수 필요
        'remove': [(word, reason), ...],     # 제거할 단어
        'suspect': [(word, reason), ...],    # 의심단어 (사람 판단 필요)
        'keep': [word, ...]                  # 유지할 단어
    }
    """
    result = {
        'review': [],
        'remove': [],
        'suspect': [],
        'keep': []
    }
    
    if remove_words_set is None:
        remove_words_set = set()
    if excluded_words_set is None:
        excluded_words_set = set()
    
    for word in words:
        word = word.strip()
        if not word:
            continue
        
        # ★ v2.5: 예외단어에 있으면 바로 'keep' 처리 (의심단어에서 제외)
        if word in excluded_words_set or word.lower() in excluded_words_set:
            result['keep'].append(word)
            continue
        
        # ★ v2.5: 제거단어 시트에 있으면 바로 'remove' 처리 (의심단어에서 제외)
        if word in remove_words_set or word.lower() in remove_words_set:
            result['remove'].append((word, "제거단어 시트"))
            continue
            
        # 한글인 경우
        if re.search(r'[가-힣]', word):
            # 유명 한글 브랜드 체크
            if word.lower().replace(" ", "") in FAMOUS_BRANDS:
                result['review'].append((word, f"유명브랜드: {word}"))
            else:
                result['keep'].append(word)
            continue
        
        # 영어/숫자 패턴 분류
        category, action, reason = classify_english_pattern(word)
        
        if action == "review":
            result['review'].append((word, reason))
        elif action == "remove":
            result['remove'].append((word, reason))
        elif action == "suspect":
            result['suspect'].append((word, reason))
        else:
            result['keep'].append(word)
    
    return result


# ==================== 불사자 API 클라이언트 ====================
class BulsajaAPIClient:
    """불사자 API 클라이언트 - Selenium 대체"""
    
    BASE_URL = "https://api.bulsaja.com/api"
    
    def __init__(self, access_token: str = "", refresh_token: str = ""):
        self.access_token = access_token
        self.refresh_token = refresh_token
        self.session = requests.Session()
        self._created_tags = set()  # ★ 이미 생성/확인된 태그 캐시
        if access_token:
            self._setup_session()
    
    def _setup_session(self):
        """세션 헤더 설정"""
        self.session.headers.update({
            'accept': 'application/json, text/plain, */*',
            'accept-language': 'ko,en-US;q=0.9,en;q=0.8',
            'accesstoken': self.access_token,
            'refreshtoken': self.refresh_token,
            'content-type': 'application/json',
            'origin': 'https://www.bulsaja.com',
            'referer': 'https://www.bulsaja.com/',
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        })
    
    def update_tokens(self, access_token: str, refresh_token: str):
        """토큰 업데이트"""
        self.access_token = access_token
        self.refresh_token = refresh_token
        self._setup_session()
    
    def test_connection(self) -> Tuple[bool, str, int]:
        """연결 테스트 - (성공여부, 메시지, 총 상품수)"""
        try:
            products, total = self.get_products(0, 1)
            return True, f"연결 성공 (총 {total}개 상품)", total
        except Exception as e:
            return False, str(e), 0
    
    def get_products(
        self,
        start_row: int = 0,
        end_row: int = 100,
        filter_model: Dict = None
    ) -> Tuple[List[Dict], int]:
        """상품 목록 조회 - 원본 API 응답 반환"""
        url = f"{self.BASE_URL}/manage/list/serverside"
        
        payload = {
            "request": {
                "startRow": start_row,
                "endRow": end_row,
                "sortModel": [],
                "filterModel": filter_model or {}
            }
        }
        
        response = self.session.post(url, json=payload)
        response.raise_for_status()
        data = response.json()
        
        products = data.get('rowData', [])
        total_count = data.get('lastRow', len(products))
        return products, total_count
    
    def get_products_by_group(self, group_name: str, limit: int = 1000) -> Tuple[List[Dict], int]:
        """특정 그룹의 상품 조회"""
        filter_model = {
            "marketGroupName": {
                "filterType": "text",
                "type": "equals",
                "filter": group_name
            }
        }
        return self.get_products(0, limit, filter_model)
    
    def get_products_without_group(self, limit: int = 1000) -> Tuple[List[Dict], int]:
        """그룹 없는 상품 조회"""
        filter_model = {
            "marketGroupName": {
                "filterType": "text",
                "type": "blank"
            }
        }
        return self.get_products(0, limit, filter_model)
    
    def update_product_names(self, updates: List[Dict]) -> bool:
        """
        상품명 일괄 수정
        updates: [{"id": "...", "name": "새 상품명"}, ...]
        """
        url = f"{self.BASE_URL}/sourcing/bulk-update-names"
        
        update_items = []
        for item in updates:
            update_items.append({
                "id": item["id"],
                "uploadCommonProductName": item["name"],
                "uploadCoupangProductName": item.get("coupang_name", item["name"]),
                "uploadSmartStoreProductName": item.get("smartstore_name", item["name"])
            })
        
        payload = {"updateItems": update_items}
        response = self.session.post(url, json=payload)
        response.raise_for_status()
        return True
    
    def update_single_product(self, product_id: str, new_name: str) -> bool:
        """단일 상품명 수정"""
        return self.update_product_names([{"id": product_id, "name": new_name}])
    
    def get_existing_tags(self) -> List[str]:
        """서버에서 태그 목록 조회"""
        url = f"{self.BASE_URL}/manage/groups"
        try:
            response = self.session.get(url)
            response.raise_for_status()
            data = response.json()
            # 태그명 리스트 추출
            if isinstance(data, list):
                return [tag.get('name', '') for tag in data if tag.get('name')]
            return []
        except:
            return []
    
    def create_tag(self, tag_name: str) -> bool:
        """태그 생성 (중복 방지)"""
        # 이미 생성한 태그면 스킵
        if tag_name in self._created_tags:
            return True
        
        # 서버에 이미 있는지 확인
        existing_tags = self.get_existing_tags()
        if tag_name in existing_tags:
            self._created_tags.add(tag_name)
            return True
        
        url = f"{self.BASE_URL}/manage/groups"
        payload = {"name": tag_name}
        try:
            response = self.session.post(url, json=payload)
            response.raise_for_status()
            self._created_tags.add(tag_name)
            return True
        except:
            # 이미 존재하는 태그일 수 있음 - 캐시에 추가
            self._created_tags.add(tag_name)
            return True
    
    def apply_tag(self, product_ids: List[str], tag_name: str) -> bool:
        """상품에 태그 적용 (태그 없으면 자동 생성)"""
        url = f"{self.BASE_URL}/sourcing/bulk-update-groups"
        payload = {
            "productIds": product_ids if isinstance(product_ids, list) else [product_ids],
            "groupName": tag_name
        }
        
        try:
            response = self.session.post(url, json=payload)
            response.raise_for_status()
            return True
        except requests.exceptions.HTTPError as e:
            # 태그가 없으면 생성 후 재시도
            if response.status_code in [400, 404]:
                self.create_tag(tag_name)
                response = self.session.post(url, json=payload)
                response.raise_for_status()
                return True
            raise e
    
    def apply_tag_single(self, product_id: str, tag_name: str) -> bool:
        """단일 상품에 태그 적용"""
        return self.apply_tag([product_id], tag_name)


@dataclass
class ProductRow:
    index: int
    image_url: str
    original_title: str
    seller_code: str = ""  # 판매자 상품 코드
    row_element: any = None  # API 모드에서는 사용 안 함
    # v11 추가
    thumbnail_urls: List[str] = None  # 전체 썸네일 URL 리스트
    needs_image_check: bool = False  # 이미지 검증 필요 여부
    is_mismatch: bool = False  # 이미지 불일치 확정
    bulsaja_id: str = ""  # 불사자 코드 (uploadTrackcopyCode - 복사해도 동일)

# ==================== 설정 파일 관리 ====================
def load_config():
    """설정 파일 로드"""
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            pass
    return {}

def save_config(config):
    """설정 파일 저장"""
    try:
        with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
        return True
    except Exception as e:
        print(f"설정 저장 실패: {e}")
        return False

# ==================== 금지단어 관리 ====================
def load_banned_words():
    """금지단어 파일 로드 (여러 JSON 형태 지원)"""
    if os.path.exists(BANNED_WORDS_FILE):
        try:
            with open(BANNED_WORDS_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
                all_words = set()
                
                # Case 1: 단순 리스트 형태 ["단어1", "단어2", ...]
                if isinstance(data, list):
                    all_words.update(data)
                    return all_words, {"words": data}
                
                # Case 2: {"words": ["단어1", ...]} 형태
                if isinstance(data, dict) and 'words' in data:
                    words = data.get('words', [])
                    if isinstance(words, list):
                        all_words.update(words)
                
                # Case 3: {"categories": {...}} 형태
                if isinstance(data, dict) and 'categories' in data:
                    categories = data.get('categories', {})
                    if isinstance(categories, dict):
                        for cat_data in categories.values():
                            if isinstance(cat_data, dict):
                                words = cat_data.get('words', [])
                                if isinstance(words, list):
                                    all_words.update(words)
                            elif isinstance(cat_data, list):
                                all_words.update(cat_data)
                
                # Case 4: AI 감지 단어
                if isinstance(data, dict) and 'ai_detected' in data:
                    ai_detected = data.get('ai_detected', {})
                    if isinstance(ai_detected, dict):
                        approved = ai_detected.get('approved', [])
                        if isinstance(approved, list):
                            all_words.update(approved)
                
                return all_words, data
        except Exception as e:
            print(f"금지단어 로드 실패: {e}")
    return set(), {}

def sync_from_google_sheets(sheet_url: str, log_callback=None) -> dict:
    """
    Google Sheets에서 금지단어/예외단어 동기화 (gspread API 사용)
    
    시트 형식:
    - A열: 금지단어
    - B열: 예외단어 (탐지 제외)
    
    Returns: {'banned': [...], 'excluded': [...], 'success': bool, 'message': str}
    """
    result = {'banned': [], 'excluded': [], 'success': False, 'message': ''}
    
    try:
        import gspread
        from google.oauth2.service_account import Credentials
    except ImportError:
        result['message'] = "gspread 또는 google-auth 패키지가 필요합니다. pip install gspread google-auth"
        if log_callback:
            log_callback(f"❌ {result['message']}")
        return result
    
    try:
        # 서비스 계정 JSON 파일 찾기
        service_account_file = None
        for filename in os.listdir('.'):
            if filename.endswith('.json') and 'auto-smartstore' in filename.lower():
                service_account_file = filename
                break
        
        if not service_account_file:
            # 일반적인 서비스 계정 파일 이름 시도
            possible_files = ['auto-smartstore-update-61c3a948c45c.json', 'service_account.json', 'credentials.json']
            for pf in possible_files:
                if os.path.exists(pf):
                    service_account_file = pf
                    break
        
        if not service_account_file:
            result['message'] = "서비스 계정 JSON 파일을 찾을 수 없습니다"
            if log_callback:
                log_callback(f"❌ {result['message']}")
            return result
        
        if log_callback:
            log_callback(f"📥 시트 연결 중... ({service_account_file})")
        
        # Google Sheets API 인증
        scopes = [
            'https://www.googleapis.com/auth/spreadsheets',
            'https://www.googleapis.com/auth/drive'
        ]
        creds = Credentials.from_service_account_file(service_account_file, scopes=scopes)
        gc = gspread.authorize(creds)
        
        # Sheet ID 추출
        if '/d/' in sheet_url:
            sheet_id = sheet_url.split('/d/')[1].split('/')[0]
        else:
            sheet_id = sheet_url  # ID만 입력한 경우
        
        # 시트 열기
        spreadsheet = gc.open_by_key(sheet_id)
        
        # bulsaja_words 시트 찾기 (없으면 첫 번째 시트)
        try:
            worksheet = spreadsheet.worksheet('bulsaja_words')
        except:
            worksheet = spreadsheet.sheet1  # 폴백
        
        # 모든 데이터 가져오기
        all_values = worksheet.get_all_values()
        
        banned_words = []
        excluded_words = []
        remove_words = []
        
        for row_num, row in enumerate(all_values):
            if row_num == 0:
                # 헤더 스킵
                continue
            
            # A열: 금지단어
            if len(row) >= 1 and row[0].strip():
                banned_words.append(row[0].strip())
            
            # B열: 예외단어
            if len(row) >= 2 and row[1].strip():
                excluded_words.append(row[1].strip())
            
            # C열: 제거단어
            if len(row) >= 3 and row[2].strip():
                remove_words.append(row[2].strip())
        
        result['banned'] = banned_words
        result['excluded'] = excluded_words
        result['remove'] = remove_words
        result['success'] = True
        result['message'] = f"금지단어 {len(banned_words)}개, 예외단어 {len(excluded_words)}개, 제거단어 {len(remove_words)}개 로드"
        
        if log_callback:
            log_callback(f"✅ {result['message']}")
        
    except gspread.exceptions.SpreadsheetNotFound:
        result['message'] = "시트를 찾을 수 없습니다. 서비스 계정에 시트 공유 필요"
        if log_callback:
            log_callback(f"❌ {result['message']}")
    except Exception as e:
        result['message'] = f"동기화 실패: {e}"
        if log_callback:
            log_callback(f"❌ {result['message']}")
    
    return result

# ==================== 탐지 제외 단어 관리 ====================
def load_excluded_words() -> set:
    """탐지 제외 단어 파일 로드"""
    if os.path.exists(EXCLUDED_WORDS_FILE):
        try:
            with open(EXCLUDED_WORDS_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
                return set(data.get('words', []))
        except Exception as e:
            print(f"제외단어 로드 실패: {e}")
    return set()

def save_excluded_words(words: set) -> bool:
    """탐지 제외 단어 파일 저장"""
    try:
        data = {'words': sorted(list(words))}
        with open(EXCLUDED_WORDS_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        return True
    except Exception as e:
        print(f"제외단어 저장 실패: {e}")
        return False

# 전역 제외 단어 목록 (프로그램 시작 시 로드)
EXCLUDED_WORDS = load_excluded_words()

# ==================== 제거단어 관리 (상품명에서 무조건 삭제) ====================
def load_remove_words() -> set:
    """제거단어 파일 로드"""
    if os.path.exists(REMOVE_WORDS_FILE):
        try:
            with open(REMOVE_WORDS_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
                return set(data.get('words', []))
        except Exception as e:
            print(f"제거단어 로드 실패: {e}")
    return set()

def save_remove_words(words: set) -> bool:
    """제거단어 파일 저장"""
    try:
        data = {'words': sorted(list(words))}
        with open(REMOVE_WORDS_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        return True
    except Exception as e:
        print(f"제거단어 저장 실패: {e}")
        return False

def apply_remove_words(title: str, remove_words: set) -> str:
    """상품명에서 제거단어 삭제"""
    if not remove_words:
        return title
    
    result = title
    for word in remove_words:
        if word in result:
            result = result.replace(word, '')
    
    # 연속 공백 정리
    result = ' '.join(result.split())
    return result.strip()

# 전역 제거 단어 목록 (프로그램 시작 시 로드)
REMOVE_WORDS = load_remove_words()

def save_banned_words(data):
    """금지단어 파일 저장"""
    try:
        with open(BANNED_WORDS_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        return True
    except Exception as e:
        print(f"금지단어 저장 실패: {e}")
        return False

# ==================== v11: 유사도 및 이미지 검증 ====================
def calculate_similarity(str1: str, str2: str) -> float:
    """두 문자열의 유사도 계산 (0.0 ~ 1.0)"""
    if not str1 or not str2:
        return 0.0
    
    # 공백 제거하고 소문자로 비교
    s1 = str1.replace(" ", "").lower()
    s2 = str2.replace(" ", "").lower()
    
    return SequenceMatcher(None, s1, s2).ratio()

def extract_main_keywords(title: str) -> set:
    """상품명에서 핵심 키워드(2글자 이상 명사) 추출"""
    if not title:
        return set()
    
    # 특수문자 제거, 공백으로 분리
    cleaned = re.sub(r'[^\w\s가-힣]', ' ', title)
    words = cleaned.split()
    
    # 2글자 이상, 숫자 아닌 단어만
    keywords = set()
    for word in words:
        if len(word) >= 2 and not word.isdigit():
            keywords.add(word.lower())
    
    return keywords

def check_keyword_overlap(original_title: str, search_titles: List[str]) -> Tuple[float, set]:
    """원본과 검색 결과의 핵심 키워드 중복률 계산
    
    Returns:
        (중복률 0.0~1.0, 공통 키워드 set)
    """
    if not original_title or not search_titles:
        return 0.0, set()
    
    original_keywords = extract_main_keywords(original_title)
    if not original_keywords:
        return 0.0, set()
    
    # 검색 결과들에서 키워드 추출
    search_keywords = set()
    for title in search_titles[:5]:  # 상위 5개만
        search_keywords.update(extract_main_keywords(title))
    
    if not search_keywords:
        return 0.0, set()
    
    # 공통 키워드
    common = original_keywords & search_keywords
    
    # 원본 키워드 기준 중복률
    overlap_ratio = len(common) / len(original_keywords) if original_keywords else 0.0
    
    return overlap_ratio, common

def verify_images_with_claude(client, thumb_urls: List[str], log_callback=None) -> Tuple[bool, float, str]:
    """Claude Sonnet으로 이미지 일치 여부 검증
    
    Args:
        client: Anthropic 클라이언트
        thumb_urls: 썸네일 URL 리스트 (최소 3개: 1번스스, 2번타오, 6번마지막)
        log_callback: 로그 콜백
        
    Returns:
        (일치여부, 신뢰도, 사유)
    """
    try:
        if len(thumb_urls) < 3:
            return True, 1.0, "썸네일 부족으로 스킵"
        
        # 1번(스스), 2번(타오), 6번(마지막) 비교
        img1_url = thumb_urls[0]  # 스마트스토어 이미지
        img2_url = thumb_urls[1]  # 타오바오 첫번째
        img_last_url = thumb_urls[-1]  # 마지막 이미지
        
        if log_callback:
            log_callback(f"  📸 비교 이미지: 1번(스스) vs 2번(타오) vs {len(thumb_urls)}번(마지막)")
            log_callback(f"  🔍 Claude Vision API 호출 중...")
        
        # 이미지 다운로드
        images_data = []
        for i, url in enumerate([img1_url, img2_url, img_last_url], 1):
            img_b64, media_type = download_image_as_base64(url)
            if img_b64:
                images_data.append((img_b64, media_type))
            else:
                if log_callback:
                    log_callback(f"  ⚠️ 이미지 {i} 다운로드 실패")
        
        if len(images_data) < 3:
            if log_callback:
                log_callback(f"  📥 이미지 다운로드: {len(images_data)}/3개 성공")
            return True, 1.0, f"이미지 다운로드 실패 ({len(images_data)}/3)"
        
        if log_callback:
            log_callback(f"  📥 이미지 다운로드: 3/3개 성공")
        
        # 프롬프트
        prompt = """3개 이미지가 동일 상품인지 판단.
첫번째=스마트스토어, 두번째/세번째=타오바오.
JSON 한 줄로만 응답: {"match":true,"confidence":0.95,"reason":"3단어이내"}"""
        
        content = [
            {"type": "image", "source": {"type": "base64", "media_type": images_data[0][1], "data": images_data[0][0]}},
            {"type": "image", "source": {"type": "base64", "media_type": images_data[1][1], "data": images_data[1][0]}},
            {"type": "image", "source": {"type": "base64", "media_type": images_data[2][1], "data": images_data[2][0]}},
            {"type": "text", "text": prompt}
        ]
        
        message = client.messages.create(
            model="claude-3-5-haiku-20241022",  # Haiku 사용 (비용 절감)
            max_tokens=100,
            messages=[{"role": "user", "content": content}]
        )
        
        response_text = message.content[0].text.strip()
        
        # JSON 추출
        if "```" in response_text:
            match = re.search(r'```(?:json)?\s*(.*?)```', response_text, re.DOTALL)
            if match:
                response_text = match.group(1).strip()
        
        json_match = re.search(r'\{[^{}]*"match"[^{}]*\}', response_text, re.DOTALL)
        if json_match:
            response_text = json_match.group(0)
        
        response_text = response_text.replace('\n', ' ').replace('\r', '')
        result = json.loads(response_text)
        
        is_match = result.get("match", True)
        confidence = float(result.get("confidence", 0.5))
        reason = result.get("reason", "")
        
        # 신뢰도 기준 판정
        if confidence < IMAGE_MATCH_THRESHOLD:
            is_match = False
        
        if log_callback:
            status = "일치" if is_match else "불일치"
            log_callback(f"  ✅ Vision 분석: {status} ({confidence:.0%})")
        
        return is_match, confidence, reason
        
    except json.JSONDecodeError as e:
        if log_callback:
            log_callback(f"  ⚠️ JSON 파싱 오류: {e}")
        return True, 0.5, f"JSON 파싱 오류"
    except Exception as e:
        if log_callback:
            log_callback(f"  ⚠️ 이미지 검증 오류: {e}")
        return True, 0.5, f"검증 오류: {str(e)[:50]}"

def delete_thumbnail_at_position(driver, row_element, position: int, log_callback=None) -> bool:
    """특정 위치의 썸네일 삭제 (Ctrl+숫자)"""
    try:
        if log_callback:
            log_callback(f"🗑️ {position}번 썸네일 삭제 중... (Ctrl+{position})")
        
        # row 클릭하여 선택
        try:
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", row_element)
            time.sleep(0.3)
            row_element.click()
            time.sleep(0.3)
        except:
            pass
        
        # Ctrl+숫자 단축키
        actions = ActionChains(driver)
        actions.key_down(Keys.CONTROL)
        actions.send_keys(str(position))
        actions.key_up(Keys.CONTROL)
        actions.perform()
        
        time.sleep(1)
        
        if log_callback:
            log_callback(f"✅ {position}번 썸네일 삭제 완료")
        
        return True
        
    except Exception as e:
        if log_callback:
            log_callback(f"⚠️ 썸네일 삭제 실패: {e}")
        return False

def move_thumbnail_to_front(driver, row_element, position: int, log_callback=None) -> bool:
    """썸네일을 1번 위치로 이동 (Alt+숫자)"""
    try:
        if position == 1:
            return True  # 이미 1번
        
        if log_callback:
            log_callback(f"📦 {position}번 썸네일 → 1번으로 이동 중... (Alt+{position})")
        
        # row 클릭하여 선택
        try:
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", row_element)
            time.sleep(0.3)
            row_element.click()
            time.sleep(0.3)
        except:
            pass
        
        # Alt+숫자 단축키
        actions = ActionChains(driver)
        actions.key_down(Keys.ALT)
        actions.send_keys(str(position))
        actions.key_up(Keys.ALT)
        actions.perform()
        
        time.sleep(1)
        
        if log_callback:
            log_callback(f"✅ 이동 완료")
        
        return True
        
    except Exception as e:
        if log_callback:
            log_callback(f"⚠️ 썸네일 이동 실패: {e}")
        return False

def remove_background_at_position(driver, row_element, position: int, log_callback=None) -> bool:
    """썸네일 배경 제거 (누끼) - Ctrl+Alt+숫자"""
    try:
        if log_callback:
            log_callback(f"🎨 {position}번 썸네일 배경 제거 중... (Ctrl+Alt+{position})")
        
        # row 클릭하여 선택
        try:
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", row_element)
            time.sleep(0.3)
            row_element.click()
            time.sleep(0.3)
        except:
            pass
        
        # Ctrl+Alt+숫자 단축키
        actions = ActionChains(driver)
        actions.key_down(Keys.CONTROL)
        actions.key_down(Keys.ALT)
        actions.send_keys(str(position))
        actions.key_up(Keys.ALT)
        actions.key_up(Keys.CONTROL)
        actions.perform()
        
        time.sleep(2)  # 배경 제거는 시간 소요
        
        if log_callback:
            log_callback(f"✅ {position}번 썸네일 배경 제거 완료")
        
        return True
        
    except Exception as e:
        if log_callback:
            log_callback(f"⚠️ 배경 제거 실패: {e}")
        return False

def analyze_and_generate_title_sonnet(client, thumb_urls: List[str], original_title: str,
                                      banned_words: set, target_length: int, 
                                      log_callback=None) -> Tuple[str, int, List[str]]:
    """Sonnet으로 이미지 분석 + 최적 이미지 선택 + 상품명 생성
    
    Returns:
        (새상품명, 최적이미지위치(1-based), 감지브랜드리스트)
    """
    try:
        if len(thumb_urls) < 2:
            return "", 1, []
        
        # 2번(타오) vs 마지막 이미지 비교
        img2_url = thumb_urls[1] if len(thumb_urls) > 1 else thumb_urls[0]
        img_last_url = thumb_urls[-1]
        
        if log_callback:
            log_callback(f"  🔍 Sonnet 이미지 분석 중...")
        
        # 이미지 다운로드
        img2_b64, media2 = download_image_as_base64(img2_url)
        img_last_b64, media_last = download_image_as_base64(img_last_url)
        
        if not img2_b64 or not img_last_b64:
            if log_callback:
                log_callback("  ⚠️ 이미지 다운로드 실패")
            return "", 1, []
        
        min_length = target_length - 5
        max_length = target_length + 5
        
        prompt = f"""두 이미지를 분석하여 다음을 수행하세요:

1. **이미지 품질 비교**: 어느 이미지가 메인 썸네일로 더 적합한지 판단
   - 상품 전체가 명확히 보이는가
   - 배경이 깔끔한가 (흰색/단색 선호)
   - 텍스트/워터마크가 없는가

2. **상품명 생성**: 이미지를 보고 최적의 한국어 상품명 생성
   - 기존 상품명 참고 (제품과 불일치하면 무시): {original_title}
   - {min_length}-{max_length}자 길이
   - 핵심 키워드 앞쪽 배치

**절대 금지:** 숫자, 영문 모델명, 브랜드명, 특수기호, 과장 표현

**응답 형식 (JSON만 출력):**
{{"best_image": 1 또는 2, "best_reason": "선택 이유", "title": "생성된 상품명", "brands": ["발견된 브랜드명"]}}"""

        content = [
            {"type": "image", "source": {"type": "base64", "media_type": media2 or "image/jpeg", "data": img2_b64}},
            {"type": "image", "source": {"type": "base64", "media_type": media_last or "image/jpeg", "data": img_last_b64}},
            {"type": "text", "text": prompt}
        ]
        
        message = client.messages.create(
            model="claude-3-5-haiku-20241022",  # Haiku 사용 (비용 절감)
            max_tokens=400,
            messages=[{"role": "user", "content": content}]
        )
        
        response_text = message.content[0].text.strip()
        
        # JSON 파싱
        if "```" in response_text:
            response_text = re.sub(r'```json\s*', '', response_text)
            response_text = re.sub(r'```\s*', '', response_text)
            response_text = response_text.strip()
        
        result = json.loads(response_text)
        
        new_title = result.get("title", "")
        best_image = result.get("best_image", 1)
        best_reason = result.get("best_reason", "")
        detected_brands = result.get("brands", [])
        
        # best_image: 1=타오바오첫번째(2번), 2=마지막 → 실제 위치로 변환
        if best_image == 1:
            actual_position = 2  # 타오바오 첫번째 = 2번 위치
        else:
            actual_position = len(thumb_urls)  # 마지막
        
        if log_callback:
            log_callback(f"  → 최적 이미지: {actual_position}번 - {best_reason[:50]}")
            if new_title:
                log_callback(f"  ✨ 신규명: {new_title}")
        
        # 금지단어 필터링
        raw_length = len(new_title) if new_title else 0
        if banned_words and new_title:
            for word in banned_words:
                if word in new_title:
                    new_title = new_title.replace(word, "").strip()
                    new_title = re.sub(r'\s+', ' ', new_title)
        
        # 필터링 후 길이 체크
        final_length = len(new_title) if new_title else 0
        if final_length < min_length and log_callback:
            log_callback(f"  ⚠️ 상품명 짧음 ({final_length}자 < {min_length}자, 원본:{raw_length}자)")
        
        return new_title, actual_position, detected_brands
        
    except Exception as e:
        if log_callback:
            log_callback(f"  ⚠️ Sonnet 분석 오류: {e}")
        return "", 1, []

# ==================== Vision 기반 상품명 생성 ====================
def download_image_as_base64(image_url: str, log_callback=None) -> Tuple[Optional[str], str]:
    """이미지 URL을 base64로 변환
    
    Returns: (base64_data, media_type) 또는 (None, "")
    """
    try:
        if not image_url:
            return None, ""
        
        # URL 정리
        url = image_url.strip()
        
        # 프로토콜 없으면 추가
        if url.startswith('//'):
            url = 'https:' + url
        elif not url.startswith('http'):
            url = 'https://' + url
        
        if log_callback:
            # URL 앞부분만 로그
            log_callback(f"  📷 이미지: {url[:60]}...")
        
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Accept': 'image/webp,image/apng,image/*,*/*;q=0.8',
            'Referer': 'https://www.bulsaja.com/'
        }
        
        response = requests.get(url, headers=headers, timeout=15, allow_redirects=True)
        
        if response.status_code == 200:
            content = response.content
            
            # 이미지 타입 자동 감지 (매직 바이트)
            media_type = "image/jpeg"  # 기본값
            if content[:4] == b'\x89PNG':
                media_type = "image/png"
            elif content[:4] == b'RIFF' and content[8:12] == b'WEBP':
                media_type = "image/webp"
            elif content[:3] == b'GIF':
                media_type = "image/gif"
            elif content[:2] == b'\xff\xd8':
                media_type = "image/jpeg"
            else:
                # Content-Type 헤더에서 추출
                content_type = response.headers.get('Content-Type', '')
                if 'png' in content_type:
                    media_type = "image/png"
                elif 'webp' in content_type:
                    media_type = "image/webp"
                elif 'gif' in content_type:
                    media_type = "image/gif"
            
            if len(content) > 1000:
                return base64.b64encode(content).decode('utf-8'), media_type
            else:
                if log_callback:
                    log_callback(f"  ⚠️ 이미지 크기 너무 작음")
                return None, ""
        else:
            if log_callback:
                log_callback(f"  ⚠️ HTTP {response.status_code}")
            return None, ""
            
    except requests.exceptions.Timeout:
        if log_callback:
            log_callback("  ⚠️ 이미지 다운로드 타임아웃")
        return None, ""
    except Exception as e:
        if log_callback:
            log_callback(f"  ⚠️ 다운로드 오류: {str(e)[:50]}")
        return None, ""

def generate_title_with_vision_api(client, image_url: str, original_title: str = "", 
                                   model: str = "claude-3-5-haiku-20241022",
                                   banned_words: set = None,
                                   log_callback=None,
                                   target_length: int = 50) -> Tuple[str, List[str], bool]:
    """Claude Vision으로 이미지를 분석하여 바로 최종 상품명 생성 (1회 API 호출)
    
    Args:
        client: Anthropic 클라이언트
        image_url: 이미지 URL
        original_title: 기존 상품명 (참고용)
        model: 사용할 모델 (기본: haiku)
        banned_words: 금지단어 세트
        log_callback: 로그 콜백 함수
        target_length: 목표 상품명 길이 (기본: 50)
    
    Returns: (최종 상품명, 감지된 브랜드 리스트, 금지단어 발견 여부)
    """
    # 길이 범위 계산
    min_length = target_length - 5
    max_length = target_length + 5
    
    # 모델명 표시
    if 'haiku' in model.lower():
        model_name = "HAIKU"
    elif 'sonnet' in model.lower():
        model_name = "SONNET"
    else:
        model_name = model.split('-')[-1].upper()
    
    if log_callback:
        log_callback(f"🔍 Vision 분석 중... ({model_name})")
    
    # 이미지 다운로드 (media_type 자동 감지)
    img_b64, media_type = download_image_as_base64(image_url, log_callback)
    if not img_b64:
        return "", [], False
    
    prompt = f"""이 상품 이미지를 분석하여 한국 오픈마켓(쿠팡, 11번가)에 등록할 최적의 상품명 1개를 생성해주세요.

**기존 상품명 참고:** {original_title if original_title else "(없음)"}

**상품명 생성 규칙:**
1. 이미지에서 보이는 상품의 특징을 정확히 반영
2. {min_length}-{max_length}자 길이 (공백 포함) - 반드시 이 범위 내로 생성
3. 검색 최적화 키워드 포함 (색상, 재질, 용도, 사이즈 등)
4. 핵심 키워드를 앞쪽에 배치
5. 자연스러운 한국어 상품명

**절대 금지:**
- 숫자 (수량, 사이즈 숫자 등)
- 영문 모델명/코드
- 브랜드명
- 특수기호 (/, %, @ 등)
- 과장 광고 표현 (최고급, 명품, 프리미엄 등)

**응답 형식 (JSON):**
{{"title": "생성된 상품명", "brands": ["발견된 브랜드명 목록, 없으면 빈 배열"]}}

JSON만 출력하세요."""

    # 최대 3번 재시도 (529 Overloaded 오류 대응)
    max_retries = 3
    for attempt in range(max_retries):
        try:
            response = client.messages.create(
                model=model,
                max_tokens=300,
                messages=[{
                    "role": "user",
                    "content": [
                        {
                            "type": "image",
                            "source": {
                                "type": "base64",
                                "media_type": media_type,
                                "data": img_b64
                            }
                        },
                        {
                            "type": "text",
                            "text": prompt
                        }
                    ]
                }]
            )
            
            result_text = response.content[0].text.strip()
            
            # JSON 파싱
            import json
            if "```" in result_text:
                result_text = re.sub(r'```json\s*', '', result_text)
                result_text = re.sub(r'```\s*', '', result_text)
            
            try:
                result = json.loads(result_text)
                new_title = result.get('title', '')
                detected_brands = result.get('brands', [])
            except:
                # JSON 파싱 실패 시 텍스트 그대로 사용
                new_title = result_text.split('\n')[0].strip()
                new_title = re.sub(r'^\d+[\.\)\-\s]+', '', new_title).strip()
                detected_brands = []
            
            if not new_title:
                if log_callback:
                    log_callback("⚠️ Vision 상품명 생성 실패")
                return "", [], False
            
            # 금지단어 필터링
            forbidden_found = False
            raw_length = len(new_title)  # 필터링 전 길이
            
            if banned_words:
                filtered_title, found_words = filter_banned_words(new_title, banned_words)
                if found_words:
                    if log_callback:
                        log_callback(f"🚫 금지단어 발견: {', '.join(found_words[:3])}")
                    new_title = filtered_title
                    forbidden_found = True
            
            # 브랜드 의심단어 감지
            brand_suspects = detect_suspicious_words(new_title)
            if brand_suspects:
                detected_brands.extend(brand_suspects)
                detected_brands = list(set(detected_brands))
            
            # 필터링 후 길이 체크 - 너무 짧으면 재생성
            final_length = len(new_title)
            if final_length < min_length:
                if log_callback:
                    log_callback(f"⚠️ 상품명 너무 짧음 ({final_length}자 < {min_length}자, 원본:{raw_length}자) → 재생성")
                if attempt < max_retries - 1:
                    continue
                else:
                    if log_callback:
                        log_callback(f"⚠️ 재생성 횟수 초과 - 현재 상품명 사용")
            
            if log_callback:
                log_callback(f"✅ Vision 상품명: {new_title[:30]}...")
            
            return new_title, detected_brands, forbidden_found
                
        except Exception as e:
            error_str = str(e)
            # 529 Overloaded 오류 시 재시도
            if "529" in error_str or "overload" in error_str.lower():
                if attempt < max_retries - 1:
                    wait_time = (attempt + 1) * 3  # 3초, 6초, 9초
                    if log_callback:
                        log_callback(f"⚠️ API 과부하 - {wait_time}초 후 재시도 ({attempt + 1}/{max_retries})")
                    time.sleep(wait_time)
                    continue
                else:
                    if log_callback:
                        log_callback(f"❌ Vision 분석 실패 (재시도 초과): {e}")
                    return "", [], False
            else:
                if log_callback:
                    log_callback(f"⚠️ Vision 분석 오류: {e}")
                return "", [], False
    
    return "", [], False

def filter_banned_words(title: str, banned_words: set) -> tuple:
    """
    상품명에서 금지단어 필터링
    Returns: (필터링된 상품명, 발견된 금지단어 리스트)
    
    모든 금지단어는 공백으로 구분된 정확한 단어만 매칭
    예: "리스"는 "구리스"에서 매칭 안 됨, "리스 제거" 또는 "무료 리스"에서만 매칭
    """
    found_words = []
    filtered_title = title
    
    for word in banned_words:
        if len(word) <= 1:  # 1글자는 무시
            continue
        
        # 공백 경계 체크: 시작/끝/공백으로 구분된 단어만 매칭
        pattern = re.compile(r'(^|\s)' + re.escape(word) + r'(\s|$)', re.IGNORECASE)
        if pattern.search(filtered_title):
            found_words.append(word)
            # 단어만 제거하고 공백 하나 유지
            filtered_title = pattern.sub(r'\1', filtered_title)
    
    # 중복 공백 정리
    filtered_title = re.sub(r'\s+', ' ', filtered_title).strip()
    
    return filtered_title, found_words

def detect_dangerous_product(title: str) -> dict:
    """
    위험상품 감지 - 카테고리별로 분류 (문맥 고려)
    Returns: {
        'is_dangerous': bool,
        'categories': {
            'adult': [...],      # 성인/성관련
            'medical': [...],    # 의료기기
            'child': [...],      # 유아용품
            'prohibited': [...], # 판매금지
            'brand': [...]       # 브랜드 의심
        },
        'all_words': [...]       # 발견된 모든 위험단어
    }
    """
    title_lower = title.lower()
    title_words = set(re.findall(r'[가-힣a-zA-Z0-9]+', title_lower))
    
    # ✅ 규제 제외 품목 (이동수단 관련)
    REGULATION_EXEMPT = [
        "전동킥보드", "전동휠", "전동스쿠터", 
        "전기자전거", "전동자전거", "고정기어"
    ]
    
    # 제외 품목이 포함되어 있으면 안전 상품으로 처리
    for exempt_item in REGULATION_EXEMPT:
        if exempt_item in title_lower:
            return {
                'is_dangerous': False,
                'categories': {
                    'adult': [],
                    'medical': [],
                    'child': [],
                    'prohibited': [],
                    'brand': []
                },
                'all_words': []
            }
    
    result = {
        'is_dangerous': False,
        'categories': {
            'adult': [],
            'medical': [],
            'child': [],
            'prohibited': [],
            'brand': []
        },
        'all_words': []
    }
    
    def is_safe_context(keyword: str, title: str) -> bool:
        """문맥상 안전한지 체크"""
        if keyword not in SAFE_CONTEXT_PATTERNS:
            return False
        
        safe_words = SAFE_CONTEXT_PATTERNS[keyword]
        title_lower = title.lower()
        
        # 안전한 문맥 단어가 있으면 True
        for safe_word in safe_words:
            if safe_word in title_lower:
                return True
        return False
    
    # 성인/성관련 체크
    for keyword in ADULT_KEYWORDS:
        if keyword in EXCLUDED_WORDS or keyword.lower() in EXCLUDED_WORDS:
            continue
        if is_safe_context(keyword, title):
            continue
        if keyword.lower() in title_lower or keyword.lower() in title_words:
            result['categories']['adult'].append(keyword)
    
    # 의료기기 체크
    for keyword in MEDICAL_KEYWORDS:
        if keyword in EXCLUDED_WORDS or keyword.lower() in EXCLUDED_WORDS:
            continue
        if is_safe_context(keyword, title):
            continue
        if keyword.lower() in title_lower or keyword.lower() in title_words:
            result['categories']['medical'].append(keyword)
    
    # 유아용품 체크
    for keyword in CHILD_KEYWORDS:
        if keyword in EXCLUDED_WORDS or keyword.lower() in EXCLUDED_WORDS:
            continue
        if is_safe_context(keyword, title):
            continue
        if keyword.lower() in title_lower or keyword.lower() in title_words:
            result['categories']['child'].append(keyword)
    
    # 판매금지 체크
    for keyword in PROHIBITED_KEYWORDS:
        if keyword in EXCLUDED_WORDS or keyword.lower() in EXCLUDED_WORDS:
            continue
        if is_safe_context(keyword, title):
            continue
        if keyword.lower() in title_lower or keyword.lower() in title_words:
            result['categories']['prohibited'].append(keyword)
    
    # 브랜드 의심 체크 (기존 함수 활용 - 이미 EXCLUDED_WORDS 체크함)
    brand_suspects = detect_suspicious_words(title)
    result['categories']['brand'] = brand_suspects
    
    # 전체 위험단어 수집
    all_dangerous = []
    for cat, words in result['categories'].items():
        all_dangerous.extend(words)
    
    result['all_words'] = list(set(all_dangerous))
    result['is_dangerous'] = len(result['all_words']) > 0
    
    return result

def get_danger_category_name(category: str) -> str:
    """위험 카테고리 한글 이름 반환"""
    names = {
        'adult': '🔞 성인/성관련',
        'medical': '💊 의료기기/의약품',
        'child': '👶 유아/아동용품',
        'prohibited': '🚫 판매금지/규제',
        'brand': '™️ 브랜드 의심'
    }
    return names.get(category, category)

def remove_synonym_duplicates(words: List[str]) -> List[str]:
    """동의어 중복 제거 - 먼저 나온 단어만 유지"""
    result = []
    used_synonyms = set()  # 이미 사용된 단어와 그 동의어들을 추적
    
    for word in words:
        word_lower = word.lower()
        
        # 이미 사용된 동의어인지 확인
        if word_lower in used_synonyms:
            continue
        
        # 현재 단어를 결과에 추가
        result.append(word)
        used_synonyms.add(word_lower)
        
        # 현재 단어의 동의어들도 사용 불가 처리
        if word_lower in SYNONYM_HINTS:
            for syn in SYNONYM_HINTS[word_lower]:
                used_synonyms.add(syn.lower())
    
    return result

def remove_quantity_expressions(title: str) -> str:
    """
    수량/단수 표현 제거 (썸네일 불일치 방지)
    예: 2인용, 3단, 4칸, 5구, 10개, 3세트 등
    """
    # 수량/단수 표현 제거
    title = re.sub(r'\b\d+인용\b', '', title)
    title = re.sub(r'\b\d+단\b', '', title)
    title = re.sub(r'\b\d+칸\b', '', title)
    title = re.sub(r'\b\d+구\b', '', title)
    title = re.sub(r'\b\d+개\b', '', title)
    title = re.sub(r'\b\d+세트\b', '', title)
    title = re.sub(r'\b\d+층\b', '', title)
    title = re.sub(r'\b\d+P\b', '', title, flags=re.IGNORECASE)
    title = re.sub(r'\b\d+팩\b', '', title)
    title = re.sub(r'\b\d+매\b', '', title)
    # 연속 공백 제거
    title = re.sub(r'\s+', ' ', title).strip()
    return title

def remove_duplicate_words(title: str) -> str:
    """
    상품명에서 중복 단어 제거 (동의어 포함)
    1. 완전히 동일한 단어 중복 제거
    2. 동의어 중복 제거 (의자/체어, 야외/실외 등)
    """
    words = title.split()
    
    # 1. 동의어 중복 제거
    words = remove_synonym_duplicates(words)
    
    # 2. 완전히 동일한 단어 중복 제거 (대소문자 무시)
    unique_words = []
    seen_words = set()
    for w in words:
        w_lower = w.lower()
        if w_lower not in seen_words:
            unique_words.append(w)
            seen_words.add(w_lower)
    
    return ' '.join(unique_words)

def detect_suspicious_words(title: str) -> List[str]:
    words = title.split()
    suspicious = []  # 리스트 초기화
    
    # 브랜드 접미사 패턴
    brand_suffixes = ['텍', '코', '랩', '몰', '샵', '존', '플', '웍', '팜', '켓', '뷰', '온', '몬', '봇']
    
    # 일반적인 한글 단어 (자주 쓰이는 명사)
    common_words = {
        # 일반 명사
        '가방', '거울', '걸이', '고리', '그릇', '기름', '꽂이', '나무', '냄비', '다리',
        '덮개', '도구', '도마', '등받이', '램프', '마개', '매트', '머리', '면봉', '모자',
        '목걸이', '물건', '물병', '물통', '바구니', '바닥', '바퀴', '박스', '받침', '발판',
        '밥솥', '방석', '배낭', '베개', '벽걸이', '보관', '보드', '보온', '보조', '볼펜',
        '부품', '붓', '빗', '빨래', '사각', '사다리', '상자', '선반', '세면', '소형',
        '손잡이', '수건', '수납', '숟가락', '스탠드', '시계', '식기', '신발', '쓰레기', '안경',
        '양말', '어깨', '여행', '열쇠', '옷걸이', '용기', '우산', '운동', '원형', '의자',
        '이불', '인형', '자석', '잠금', '장갑', '장식', '전자', '접이식', '젓가락', '정리',
        '조각', '조립', '조명', '종이', '주걱', '주방', '주머니', '지갑', '진열', '찬장',
        '책상', '청소', '촛대', '충전', '침대', '칫솔', '카펫', '캐리어', '커튼', '케이스',
        '코팅', '쿠션', '크기', '클립', '탁자', '털실', '텀블러', '테이프', '통', '트레이',
        '파우치', '팔걸이', '패드', '포장', '필통', '핀', '하우스', '함', '행거', '헤드',
        '홀더', '화분', '화장', '후크',
        # 재질/소재
        '가죽', '고무', '금속', '나일론', '대나무', '면', '목재', '밤부', '스테인리스',
        '스틸', '아크릴', '알루미늄', '원목', '유리', '은박', '실리콘', '철제', '친환경',
        '플라스틱', '합성', '황동', '스텐',
        # 용도/기능
        '가정용', '다용도', '대용량', '미니', '방수', '방한', '보관용', '소형', '야외',
        '업소용', '여행용', '욕실', '이동식', '일회용', '접이식', '주방용', '차량용', '휴대용',
        # 형태/모양
        '각형', '구형', '긴', '네모', '넓은', '높은', '둥근', '사각', '삼각', '세로',
        '소', '얇은', '오각', '원형', '육각', '작은', '정사각', '직사각', '짧은', '큰', '타원',
        # 색상 (일반적)
        '검정', '노랑', '빨강', '초록', '파랑', '흰색', '회색', '투명', '블랙', '화이트',
        # 숫자 단위
        '개입', '묶음', '박스', '세트', '팩',
        # 동사형/부사형
        '걸이', '꽂이', '덮개', '받침', '보관', '수납', '정리', '청소', '충전',
        # 흔한 외래어
        '가드', '거치대', '그립', '도어', '라이트', '랙', '레일', '리모컨', '링', '마운트',
        '매직', '미러', '박스', '백', '버튼', '보틀', '볼', '브러시', '블록', '사이드',
        '샤워', '서랍', '세트', '소켓', '스위치', '스탠드', '스토리지', '슬라이드', '아이템',
        '인테리어', '체어', '커버', '컵', '케이블', '클리너', '키트', '타올', '태그', '테이블',
        '트레이', '파티션', '패널', '펜', '폴더', '푸시', '프레임', '플레이트', '핀', '필터',
        '행어', '헤드', '홀더', '휴지',
        # 안전 단어 (브랜드가 아닌 일반 명사)
        '황소', '홍콩', '호이스트', '포터', '챔피언', '마스터', '스타', '프리미엄', '로얄',
        '크라운', '킹', '퀸', '에이스', '골드', '실버', '플래티넘', '다이아', '루비',
        '사파이어', '에메랄드', '펄', '크리스탈', '파워', '터보', '하이퍼', '울트라',
        '메가', '기가', '슈퍼', '엑스트라', '플러스', '맥스', '프로', '스페셜',
    }
    
    for word in words:
        # 제외 단어 체크 (EXCLUDED_WORDS에 있으면 스킵)
        if word.lower() in EXCLUDED_WORDS or word in EXCLUDED_WORDS:
            continue
        
        # 1글자는 스킵
        if len(word) <= 1:
            continue
        
        # 숫자만 있는 것은 스킵
        if word.isdigit():
            continue
        
        # ========== 특이한 단어 감지 ==========
        
        # 1. 영문+숫자 조합 (모델명 가능성) - 예: GT500, RX7, A4
        if re.match(r'^[A-Za-z]+\d+[A-Za-z]*\d*$', word) or re.match(r'^\d+[A-Za-z]+\d*$', word):
            suspicious.append(word)
            continue
        
        # 2. 대문자만 3글자 이상 (약어/브랜드) - 예: BMW, USB는 예외처리 필요
        if re.match(r'^[A-Z]{3,}$', word) and word not in {'USB', 'LED', 'LCD', 'DIY', 'PVC', 'EVA', 'PET', 'ABS', 'TPU', 'TPE', 'PP', 'PE'}:
            suspicious.append(word)
            continue
        
        # 3. 영문+한글 조합 (브랜드 가능성 높음) - 예: LG전자, SK텔레콤
        if re.match(r'^[A-Za-z]+[가-힣]+$', word):
            suspicious.append(word)
            continue
        
        # 4. 한글+영문 조합 (브랜드 가능성 높음) - 예: 삼성PRO, 오스타MAX
        if re.match(r'^[가-힣]+[A-Za-z]+$', word):
            suspicious.append(word)
            continue
        
        # 5. 카멜케이스 또는 파스칼케이스 (브랜드/제품명) - 예: SHFiguarts, PlayStation
        if re.match(r'^[A-Z][a-z]+[A-Z]', word) or re.match(r'^[a-z]+[A-Z]', word):
            suspicious.append(word)
            continue
        
        # 6. 버전 표기 패턴 - 예: ver.2, v1.0, 2.0
        if re.match(r'^[Vv]er\.?\d', word) or re.match(r'^[Vv]\d', word):
            suspicious.append(word)
            continue
        
        # 7. 특수문자 포함 단어 (브랜드 가능성) - 예: E-mart, T&T
        if re.search(r'[&\-\'\.]', word) and len(word) > 2:
            suspicious.append(word)
            continue
        
        # 8. 순수 영문 단어 (흔하지 않은 것) - 일반적인 영문 단어 제외
        common_english = {
            'mini', 'set', 'box', 'bag', 'cup', 'pad', 'mat', 'key', 'led', 'usb', 'diy',
            'pro', 'max', 'plus', 'air', 'new', 'top', 'hot', 'big', 'one', 'two',
            'size', 'type', 'home', 'tool', 'rack', 'hook', 'clip', 'ring', 'case',
            'cover', 'holder', 'stand', 'light', 'lamp', 'cable', 'wire', 'cord',
            'storage', 'organizer', 'container', 'basket', 'tray', 'shelf',
            'black', 'white', 'blue', 'red', 'green', 'pink', 'gray', 'brown',
            'large', 'small', 'medium', 'long', 'short', 'wide', 'thin', 'thick',
        }
        if re.match(r'^[A-Za-z]{4,}$', word) and word.lower() not in common_english:
            # 첫글자 대문자인 경우 더 의심
            if word[0].isupper():
                suspicious.append(word)
                continue
        
        # 9. 2~4글자 한글 + 브랜드 접미사 패턴
        if re.match(r'^[가-힣]{2,4}$', word):
            if any(word.endswith(suffix) for suffix in brand_suffixes):
                # 일반 명사 제외
                if word not in COMMON_KOREAN_WORDS and word not in common_words:
                    suspicious.append(word)
                    continue
        
        # 10. 일반적이지 않은 한글 단어 → 제거 (오탐 너무 많음)
        # 순수 한글 일반명사는 브랜드 의심하지 않음
    
    return list(set(suspicious))  # 중복 제거

def find_chrome_path():
    """크롬 실행 파일 경로 찾기"""
    possible_paths = [
        r"C:\Program Files\Google\Chrome\Application\chrome.exe",
        r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
        os.path.expandvars(r"%LOCALAPPDATA%\Google\Chrome\Application\chrome.exe"),
    ]
    
    for path in possible_paths:
        if os.path.exists(path):
            return path
    
    return "chrome.exe"  # PATH에서 찾기

def is_valid_image_url(url: str) -> bool:
    """이미지 URL이 확장자를 가지고 있는지 확인"""
    if not url:
        return False
    
    # URL에서 쿼리 파라미터 제거 후 확장자 체크
    url_without_query = url.split('?')[0].lower()
    return url_without_query.endswith(IMAGE_EXTENSIONS)

# ==================== 메인 클래스 ====================
class BulsajaAutoFiller:
    def __init__(self, gui):
        self.gui = gui
        self.main_driver: Optional[webdriver.Chrome] = None  # Selenium (레거시)
        self.aliprice_driver: Optional[webdriver.Chrome] = None  # 알리프라이스 검색용
        self.claude_client: Optional[anthropic.Anthropic] = None
        self.is_running = False
        self.chrome_process = None
        self.main_window_handle = None
        
        # ★ v3.0: API 클라이언트 추가
        self.api_client: Optional[BulsajaAPIClient] = None
        self.use_api_mode = True  # True: API 모드, False: Selenium 모드
        
        # 금지단어 로드
        self.banned_words, self.banned_words_data = load_banned_words()
        if self.banned_words:
            print(f"금지단어 {len(self.banned_words)}개 로드됨")
        
        # 제거단어 로드
        self.remove_words = load_remove_words()
        if self.remove_words:
            print(f"제거단어 {len(self.remove_words)}개 로드됨")
    
    # ==================== API 연결 메서드 ====================
    def init_api_client(self, access_token: str, refresh_token: str) -> Tuple[bool, str, int]:
        """API 클라이언트 초기화"""
        self.api_client = BulsajaAPIClient(access_token, refresh_token)
        success, msg, total = self.api_client.test_connection()
        if success:
            self.use_api_mode = True
        return success, msg, total
    
    def extract_tokens_from_browser(self, port: int = 9222) -> Tuple[bool, str, str]:
        """크롬 디버깅 모드에서 토큰 자동 추출"""
        try:
            # 1. 열린 탭 목록 조회
            tabs_url = f"http://127.0.0.1:{port}/json"
            response = requests.get(tabs_url, timeout=3)
            tabs = response.json()
            
            # 2. 불사자 탭 찾기
            bulsaja_tab = None
            for tab in tabs:
                if 'bulsaja.com' in tab.get('url', ''):
                    bulsaja_tab = tab
                    break
            
            if not bulsaja_tab:
                return False, "", ""
            
            # 3. WebSocket으로 연결
            ws_url = bulsaja_tab.get('webSocketDebuggerUrl')
            if not ws_url:
                return False, "", ""
            
            ws = websocket.create_connection(ws_url)
            
            # 4. localStorage에서 토큰 추출
            cmd = {
                "id": 1,
                "method": "Runtime.evaluate",
                "params": {
                    "expression": """
                        (function() {
                            var tokenStr = localStorage.getItem('token');
                            if (tokenStr) {
                                try {
                                    var tokenObj = JSON.parse(tokenStr);
                                    if (tokenObj.state) {
                                        return JSON.stringify({
                                            accessToken: tokenObj.state.accessToken || '',
                                            refreshToken: tokenObj.state.refreshToken || ''
                                        });
                                    }
                                    return JSON.stringify({
                                        accessToken: tokenObj.accessToken || '',
                                        refreshToken: tokenObj.refreshToken || ''
                                    });
                                } catch(e) {
                                    return JSON.stringify({accessToken: '', refreshToken: ''});
                                }
                            }
                            return JSON.stringify({accessToken: '', refreshToken: ''});
                        })()
                    """,
                    "returnByValue": True
                }
            }
            
            ws.send(json.dumps(cmd))
            result = json.loads(ws.recv())
            ws.close()
            
            if 'result' in result and 'result' in result['result']:
                token_data = json.loads(result['result']['result'].get('value', '{}'))
                access_token = token_data.get('accessToken', '')
                refresh_token = token_data.get('refreshToken', '')
                
                if access_token and refresh_token:
                    return True, access_token, refresh_token
            
            return False, "", ""
            
        except Exception as e:
            return False, "", ""
        
    def reload_banned_words(self):
        """금지단어 다시 로드"""
        self.banned_words, self.banned_words_data = load_banned_words()
        return len(self.banned_words)
    
    def reset_filters_and_set_default(self):
        """필터 초기화 및 기본 필터 설정 - API 모드에서는 불필요"""
        
        # ★ v3.0: API 모드에서는 필터 초기화 불필요
        if self.use_api_mode and self.api_client:
            return True
        
        if not self.main_driver:
            return False
        
        try:
            self.gui.log("🔄 필터 초기화 중...")
            
            # 1. 필터 초기화 버튼 클릭
            try:
                reset_btn = self.main_driver.find_element(
                    By.XPATH, "//button[contains(text(), '필터 초기화')]"
                )
                self.main_driver.execute_script("arguments[0].click();", reset_btn)
                time.sleep(0.3)
                self.gui.log("  ✓ 필터 초기화 클릭")
            except:
                self.gui.log("  ⚠️ 필터 초기화 버튼 없음")
            
            # 2. 확인 버튼 클릭 (모달)
            try:
                confirm_btn = self.main_driver.find_element(
                    By.XPATH, "//button[contains(@class, 'bg-orange-50') and contains(text(), '확인')]"
                )
                self.main_driver.execute_script("arguments[0].click();", confirm_btn)
                time.sleep(0.5)
                self.gui.log("  ✓ 확인 클릭")
            except:
                pass  # 모달이 없을 수도 있음
            
            # 3. 기본 필터 탭 클릭 (수집완료 AND 번역완료)
            time.sleep(0.3)
            
            # 수집완료 탭 클릭
            try:
                tab = self.main_driver.find_element(
                    By.XPATH, "//button[contains(text(), '수집완료')]"
                )
                # ActionChains로 실제 클릭
                actions = ActionChains(self.main_driver)
                actions.move_to_element(tab).click().perform()
                time.sleep(0.3)
                self.gui.log("  ✓ '수집완료' 탭 선택")
            except Exception as e:
                self.gui.log(f"  ⚠️ 수집완료 탭 없음")
            
            # 번역완료 탭도 클릭 (이미지 번역 섹션)
            time.sleep(0.3)
            try:
                # role="tab"이고 텍스트가 "번역 완료"인 버튼 찾기
                tab = self.main_driver.find_element(
                    By.XPATH, "//button[@role='tab' and normalize-space(text())='번역 완료']"
                )
                
                # ActionChains로 실제 클릭 (Radix UI 탭은 JS 클릭이 안 먹힘)
                actions = ActionChains(self.main_driver)
                actions.move_to_element(tab).click().perform()
                time.sleep(0.3)
                
                # 클릭 후 상태 확인
                state = tab.get_attribute("data-state")
                if state == "active":
                    self.gui.log("  ✓ '번역 완료' 탭 선택")
                else:
                    # 다시 클릭 시도
                    tab.click()
                    time.sleep(0.3)
                    self.gui.log("  ✓ '번역 완료' 탭 선택 (재시도)")
            except Exception as e:
                self.gui.log(f"  ⚠️ 번역완료 탭 없음: {e}")
            
            time.sleep(0.5)
            self.gui.log("✅ 필터 초기화 완료")
            return True
            
        except Exception as e:
            self.gui.log(f"⚠️ 필터 초기화 오류: {e}")
            return False
    
    def ensure_main_window(self):
        """원래 탭(상품 리스트)으로 돌아가기"""
        if not self.main_driver or not self.main_window_handle:
            return
        
        try:
            current_handle = self.main_driver.current_window_handle
            
            # 현재 탭이 원래 탭이 아니면 전환
            if current_handle != self.main_window_handle:
                # 원래 탭이 아직 있는지 확인
                if self.main_window_handle in self.main_driver.window_handles:
                    self.main_driver.switch_to.window(self.main_window_handle)
                    self.gui.log("🔄 상품 리스트 탭으로 복귀")
                else:
                    # 원래 탭이 없으면 첫 번째 탭으로
                    self.main_driver.switch_to.window(self.main_driver.window_handles[0])
                    self.main_window_handle = self.main_driver.current_window_handle
                    self.gui.log("🔄 첫 번째 탭으로 전환")
            
            # 불필요한 탭 닫기 (원래 탭 제외)
            all_handles = self.main_driver.window_handles
            if len(all_handles) > 1:
                for handle in all_handles:
                    if handle != self.main_window_handle:
                        try:
                            self.main_driver.switch_to.window(handle)
                            self.main_driver.close()
                        except:
                            pass
                
                # 원래 탭으로 다시 전환
                self.main_driver.switch_to.window(self.main_window_handle)
                self.gui.log(f"🗑️ 불필요한 탭 {len(all_handles)-1}개 닫음")
        
        except Exception as e:
            self.gui.log(f"⚠️ 탭 전환 오류: {e}")
    
    def _add_keywords_to_banned_words(self, keywords: List[str]):
        """금지단어 JSON에 키워드 자동 추가"""
        if not keywords:
            return
        
        try:
            import json
            
            # 기존 JSON 파일 로드
            json_path = "banned_words.json"
            if os.path.exists(json_path):
                with open(json_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
            else:
                data = {"words": []}
            
            # 기존 단어 수집 (다양한 구조 지원)
            existing = set()
            
            # 구조 1: {"words": [...]}
            if "words" in data and isinstance(data["words"], list):
                existing.update(data["words"])
            
            # 구조 2: {"categories": {...}} 
            if "categories" in data and isinstance(data["categories"], dict):
                for cat_key, cat_val in data["categories"].items():
                    if isinstance(cat_val, list):
                        existing.update(cat_val)
                    elif isinstance(cat_val, dict) and "words" in cat_val:
                        existing.update(cat_val["words"])
            
            # 새 키워드 필터링 (중복 제외)
            new_keywords = [k for k in keywords if k not in existing and k.strip()]
            
            if new_keywords:
                # words 키가 없으면 생성
                if "words" not in data:
                    data["words"] = list(existing)  # 기존 단어들을 words로 이동
                
                data["words"].extend(new_keywords)
                
                # 파일 저장
                with open(json_path, 'w', encoding='utf-8') as f:
                    json.dump(data, f, ensure_ascii=False, indent=2)
                
                # 메모리에도 업데이트
                self.banned_words.update(new_keywords)
                
                self.gui.log(f"📝 금지단어 추가: {', '.join(new_keywords[:5])}{'...' if len(new_keywords) > 5 else ''}")
        
        except Exception as e:
            self.gui.log(f"⚠️ 금지단어 추가 실패: {e}")
    
    def close_aliprice(self):
        """AliPrice 브라우저 닫기 - main_driver 사용으로 불필요"""
        pass
    
    def create_tag(self, tag_name: str) -> bool:
        """태그 생성 - API 모드 또는 Selenium 모드"""
        
        # ★ v3.0: API 모드
        if self.use_api_mode and self.api_client:
            try:
                self.gui.log(f"🏷️ 태그 생성 중: {tag_name}")
                self.api_client.create_tag(tag_name)
                self.gui.log(f"✅ 태그 '{tag_name}' 생성 완료")
                return True
            except Exception as e:
                self.gui.log(f"⚠️ 태그 생성 실패 (이미 존재할 수 있음): {e}")
                return True  # 이미 존재해도 OK
        
        # 레거시 Selenium 모드
        try:
            # ★ 원래 탭으로 돌아가기 (새 탭 방지)
            self.ensure_main_window()
            
            self.gui.log(f"🏷️ 태그 생성 중: {tag_name}")
            
            # 1. 태그 관리 버튼 클릭 (명시적 대기)
            try:
                tag_btn = WebDriverWait(self.main_driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, "//button[contains(., '태그 관리')]"))
                )
            except:
                self.gui.log("❌ 태그 관리 버튼을 찾을 수 없습니다")
                self.gui.log("💡 상품 리스트 페이지에서 시도하세요")
                return False
            
            self.main_driver.execute_script("arguments[0].click();", tag_btn)
            time.sleep(1)  # 모달 열릴 때까지 대기
            
            # 2. 입력 필드 찾기 (대기 포함)
            try:
                input_field = WebDriverWait(self.main_driver, 5).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "input#new-group, input[placeholder='새 그룹 이름']"))
                )
            except:
                self.gui.log("⚠️ 태그 입력 필드를 찾을 수 없음")
                return False
            
            # 3. React 입력 필드에 값 설정 (nativeInputValueSetter 사용)
            self.main_driver.execute_script("""
                var input = arguments[0];
                var value = arguments[1];
                
                // React의 경우 native setter를 사용해야 함
                var nativeInputValueSetter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, 'value').set;
                nativeInputValueSetter.call(input, value);
                
                // React가 감지할 수 있도록 이벤트 발생
                var inputEvent = new Event('input', { bubbles: true });
                input.dispatchEvent(inputEvent);
            """, input_field, tag_name)
            time.sleep(0.5)
            
            # 4. 추가 버튼 찾아서 클릭
            try:
                add_btn = WebDriverWait(self.main_driver, 3).until(
                    EC.element_to_be_clickable((By.XPATH, "//button[text()='추가']"))
                )
                self.main_driver.execute_script("arguments[0].click();", add_btn)
                time.sleep(1)
            except:
                self.gui.log("⚠️ 추가 버튼 클릭 실패")
                return False
            
            # 5. 모달 닫기 - 닫기 버튼 클릭
            try:
                close_btn = self.main_driver.find_element(By.XPATH, "//button[text()='닫기']")
                self.main_driver.execute_script("arguments[0].click();", close_btn)
            except:
                pass
            
            time.sleep(0.3)
            self.gui.log(f"✅ 태그 '{tag_name}' 생성 완료")
            return True
            
        except Exception as e:
            self.gui.log(f"⚠️ 태그 생성 실패: {e}")
            # 모달 닫기 시도
            try:
                close_btn = self.main_driver.find_element(By.XPATH, "//button[text()='닫기']")
                self.main_driver.execute_script("arguments[0].click();", close_btn)
            except:
                pass
            return False
    
    def check_tag_exists(self, tag_name: str) -> bool:
        """태그 존재 여부 확인 - API 모드에서는 apply_tag에서 자동 생성하므로 항상 True"""
        
        # ★ v3.0: API 모드에서는 태그 적용 시 자동 생성되므로 확인만
        if self.use_api_mode and self.api_client:
            self.gui.log(f"✅ 태그 '{tag_name}' 확인됨 (적용 시 자동 생성)")
            return True
        
        # 레거시 Selenium 모드
        try:
            # ★ 원래 탭으로 돌아가기 (새 탭 방지)
            self.ensure_main_window()
            
            # 1. 태그 관리 버튼 클릭 (명시적 대기)
            try:
                tag_btn = WebDriverWait(self.main_driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, "//button[contains(., '태그 관리')]"))
                )
            except:
                self.gui.log("⚠️ 태그 확인 실패: 태그 관리 버튼을 찾을 수 없음")
                return False
            
            self.main_driver.execute_script("arguments[0].click();", tag_btn)
            time.sleep(1)
            
            # 2. 태그 목록에서 해당 태그 찾기
            try:
                # 수동 태그 목록 확인 (태그 관리 모달 내 목록)
                tag_items = self.main_driver.find_elements(By.XPATH, f"//*[contains(text(), '{tag_name}')]")
                tag_exists = len(tag_items) > 0
            except:
                tag_exists = False
            
            # 3. 모달 닫기
            try:
                close_btn = self.main_driver.find_element(By.XPATH, "//button[text()='닫기']")
                self.main_driver.execute_script("arguments[0].click();", close_btn)
            except:
                pass
            
            time.sleep(0.3)
            return tag_exists
            
        except Exception as e:
            self.gui.log(f"⚠️ 태그 확인 실패: {e}")
            try:
                close_btn = self.main_driver.find_element(By.XPATH, "//button[text()='닫기']")
                self.main_driver.execute_script("arguments[0].click();", close_btn)
            except:
                pass
            return False
    
    def apply_tag_to_product(self, product_index: int, tag_name: str, product_id: str = None) -> bool:
        """상품에 태그 적용 - API 모드 또는 Selenium 모드"""
        
        # ★ v3.0: API 모드
        if self.use_api_mode and self.api_client and product_id:
            try:
                self.api_client.apply_tag_single(product_id, tag_name)
                return True
            except Exception as e:
                self.gui.log(f"⚠️ API 태그 적용 실패: {e}")
                return False
        
        # 레거시 Selenium 모드
        try:
            # 1. 해당 상품 행으로 스크롤
            grid_body = self.main_driver.find_element(By.CSS_SELECTOR, ".ag-body-viewport")
            scroll_position = product_index * 126
            self.main_driver.execute_script(f"arguments[0].scrollTop = {scroll_position};", grid_body)
            time.sleep(0.3)
            
            # row-index로 해당 행 찾기
            row = None
            for attempt in range(5):
                rows = self.main_driver.find_elements(By.CSS_SELECTOR, f"div[role='row'][row-index='{product_index}']")
                if rows:
                    row = rows[0]
                    break
                self.main_driver.execute_script("arguments[0].scrollTop += 50;", grid_body)
                time.sleep(0.2)
            
            if not row:
                self.gui.log(f"⚠️ row-index {product_index} 찾을 수 없음")
                return False
            
            return self._apply_tag_to_row(row, product_index, tag_name)
            
        except Exception as e:
            self.gui.log(f"⚠️ 태그 적용 오류: {e}")
            return False
        
    def setup_drivers(self):
        """브라우저 드라이버 설정 - 아직 열지 않음"""
        try:
            self.gui.log("✅ 드라이버 준비 완료")
            return True
        except Exception as e:
            self.gui.log(f"❌ 드라이버 설정 실패: {e}")
            return False
    
    def setup_claude(self, api_key: str):
        """Claude API 설정"""
        try:
            self.claude_client = anthropic.Anthropic(api_key=api_key)
            self.gui.log("✅ Claude API 연결 완료")
            return True
        except Exception as e:
            self.gui.log(f"❌ Claude API 설정 실패: {e}")
            return False
    
    def launch_debug_chrome(self, port: int = DEBUG_PORT):
        """디버깅 모드 크롬 자동 실행"""
        import socket
        
        try:
            chrome_path = find_chrome_path()
            
            # 이미 실행 중인지 확인 (소켓으로 빠르게 체크)
            def is_port_open(p):
                try:
                    sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                    sock.settimeout(1)
                    result = sock.connect_ex(('127.0.0.1', p))
                    sock.close()
                    return result == 0
                except:
                    return False
            
            if is_port_open(port):
                self.gui.log("✅ 이미 실행 중인 디버깅 크롬 발견")
                return True
            
            self.gui.log("🚀 디버깅 모드 크롬 실행 중...")
            
            # 크롬 실행 (shell 명령어로)
            cmd = f'"{chrome_path}" --remote-debugging-port={port} --user-data-dir="{CHROME_DEBUG_PROFILE}" --remote-allow-origins=* "{BULSAJA_PRODUCT_LIST_URL}"'
            
            self.chrome_process = subprocess.Popen(
                cmd,
                shell=True,
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL
            )
            
            self.gui.log(f"📁 프로필: {CHROME_DEBUG_PROFILE}")
            self.gui.log(f"🌐 URL: {BULSAJA_PRODUCT_LIST_URL}")
            
            # 포트 열릴 때까지 대기 (최대 30초)
            self.gui.log("⏳ 크롬 시작 대기...")
            for i in range(30):
                if is_port_open(port):
                    self.gui.log("✅ 크롬 포트 열림")
                    return True
                time.sleep(1)
            
            self.gui.log("⚠️ 포트 열림 확인 실패, 연결 시도...")
            return True
            
        except Exception as e:
            self.gui.log(f"❌ 크롬 실행 실패: {e}")
            return False
    
    def connect_to_existing_chrome(self, port: int = DEBUG_PORT):
        """디버깅 모드로 실행 중인 크롬에 연결"""
        try:
            # 기존 드라이버가 있으면 정리
            if self.main_driver:
                try:
                    self.main_driver.quit()
                except:
                    pass
                self.main_driver = None
            
            opt = Options()
            opt.add_experimental_option("debuggerAddress", f"127.0.0.1:{port}")
            
            self.main_driver = webdriver.Chrome(options=opt)
            
            # ★ 창 최대화 (태그 컬럼 등 모든 컬럼이 보이도록)
            self.main_driver.maximize_window()
            
            # ★ 원래 탭 핸들 저장 (새 탭 방지용)
            self.main_window_handle = self.main_driver.current_window_handle
            
            # 연결 확인
            current_url = self.main_driver.current_url
            self.gui.log(f"✅ 크롬에 연결됨")
            self.gui.log(f"📍 현재 URL: {current_url}")
            
            return True
            
        except Exception as e:
            self.main_driver = None
            self.gui.log(f"❌ 크롬 연결 실패: {e}")
            return False
    
    def connect_to_bulsaja(self, url: str):
        """불사자 브라우저 연결 (새 창 열기)"""
        try:
            opt = Options()
            opt.add_argument("--window-size=1400,900")
            
            # 자동화 전용 프로필 (랜덤 디렉토리)
            import tempfile
            import random
            profile_dir = os.path.join(tempfile.gettempdir(), f'bulsaja_chrome_{random.randint(1000,9999)}')
            opt.add_argument(f"--user-data-dir={profile_dir}")
            
            # 자동화 감지 우회
            opt.add_experimental_option("excludeSwitches", ["enable-automation"])
            opt.add_experimental_option('useAutomationExtension', False)
            opt.add_argument("--disable-blink-features=AutomationControlled")
            opt.add_argument("--no-sandbox")
            opt.add_argument("--disable-dev-shm-usage")
            
            # 충돌 방지
            opt.add_argument("--disable-features=VizDisplayCompositor")
            opt.add_argument("--disable-gpu")
            
            try:
                service = Service(ChromeDriverManager().install())
                self.main_driver = webdriver.Chrome(service=service, options=opt)
            except Exception as e:
                self.gui.log(f"⚠️ 첫 시도 실패, 재시도 중...")
                time.sleep(1)
                service = Service(ChromeDriverManager().install())
                self.main_driver = webdriver.Chrome(service=service, options=opt)
            
            # User-Agent 변경
            try:
                self.main_driver.execute_cdp_cmd('Network.setUserAgentOverride', {
                    "userAgent": 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
                })
                self.main_driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
            except:
                pass
            
            self.main_driver.get(url)
            
            # ★ 원래 탭 핸들 저장 (새 탭 방지용)
            self.main_window_handle = self.main_driver.current_window_handle
            
            # ★ 창 최대화
            self.main_driver.maximize_window()
            
            self.gui.log("✅ 불사자 브라우저 열림")
            
            return True
            
        except Exception as e:
            self.gui.log(f"❌ 브라우저 연결 실패: {e}")
            return False
    
    def get_visible_products(self, start_index: int = 0, max_count: int = 100) -> List[ProductRow]:
        """현재 화면의 상품 리스트 추출 - API 모드 또는 Selenium 모드"""
        
        # ★ v3.0: API 모드
        if self.use_api_mode and self.api_client:
            return self._get_products_via_api(start_index, max_count)
        
        # 레거시 Selenium 모드
        return self._get_products_via_selenium(start_index, max_count)
    
    def _get_products_via_api(self, start_index: int = 0, max_count: int = 100) -> List[ProductRow]:
        """API로 상품 목록 조회 - 마켓그룹 + 태그 필터 동시 적용"""
        products = []
        
        try:
            # 필터 모델 구성
            filter_model = {}
            
            # 마켓 그룹 필터
            market_group = getattr(self.gui, 'current_market_group', None)
            if market_group and market_group not in ["(전체)", ""]:
                filter_model["marketGroupName"] = {
                    "filterType": "text",
                    "type": "equals",
                    "filter": market_group
                }
            
            # 태그 필터 (groupFile + contains)
            tag_filter = getattr(self.gui, 'current_tag_filter', None)
            if tag_filter:
                filter_model["groupFile"] = {
                    "filterType": "text",
                    "type": "contains",
                    "filter": tag_filter
                }
            
            # 상태 필터 비활성화 - 모든 상태의 상품 처리
            # filter_model["status"] = {
            #     "filterType": "text",
            #     "type": "equals",
            #     "filter": "0"
            # }
            
            # 번역 필터 비활성화 - 번역 상태 관계없이 처리
            # filter_model["uploadDetailContents.imageTranslated"] = {
            #     "filterType": "text",
            #     "type": "equals",
            #     "filter": "1"
            # }
            
            # API 호출
            api_products, total = self.api_client.get_products(start_index, start_index + max_count, filter_model)
            
            for idx, item in enumerate(api_products):
                # 썸네일 URL 목록
                thumbnails = item.get('uploadThumbnails', [])
                first_thumb = thumbnails[0] if thumbnails else ""
                
                product = ProductRow(
                    index=start_index + idx,
                    image_url=first_thumb,
                    original_title=item.get('uploadCommonProductName', ''),
                    seller_code=item.get('uploadBulsajaCode', ''),
                    row_element=None,  # API 모드에서는 사용 안 함
                    thumbnail_urls=thumbnails,
                    needs_image_check=False,
                    is_mismatch=False,
                    bulsaja_id=item.get('uploadTrackcopyCode', '')  # 불사자 코드 (복사해도 동일)
                )
                products.append(product)
            
            self.gui.log(f"  📦 API로 {len(products)}개 상품 로드 (총 {total}개)")
            
        except Exception as e:
            self.gui.log(f"❌ API 상품 조회 실패: {e}")
        
        return products
    
    def _get_products_via_selenium(self, start_index: int = 0, max_count: int = 100) -> List[ProductRow]:
        """Selenium으로 상품 리스트 추출 (레거시) - AG Grid"""
        products = []
        seen_indices = set()
        
        try:
            # AG Grid 컨테이너 찾기
            grid_body = self.main_driver.find_element(By.CSS_SELECTOR, ".ag-body-viewport")
            
            # ★ v1.6 방식: start_index가 0이면 현재 보이는 첫 번째 row의 인덱스 찾기
            if start_index == 0:
                first_row_index = self.main_driver.execute_script("""
                    var rows = document.querySelectorAll('div[role="row"][row-index]');
                    var minIndex = 999999;
                    for (var row of rows) {
                        var idx = parseInt(row.getAttribute('row-index'));
                        if (!isNaN(idx) && idx < minIndex) {
                            minIndex = idx;
                        }
                    }
                    return minIndex < 999999 ? minIndex : 0;
                """)
                start_index = first_row_index
                self.gui.log(f"  📍 첫 번째 row-index: {start_index}")
            
            # 시작 위치로 스크롤 (각 행 높이 약 126px)
            scroll_position = start_index * 126
            self.main_driver.execute_script(f"arguments[0].scrollTop = {scroll_position};", grid_body)
            time.sleep(0.2)
            
            # 스크롤하며 필요한 만큼만 수집
            no_new_count = 0
            
            while no_new_count < 3 and len(products) < max_count:
                # 현재 보이는 행 수집
                rows = self.main_driver.find_elements(By.CSS_SELECTOR, "div[role='row'][row-index]")
                new_found = 0
                
                for row in rows:
                    if len(products) >= max_count:
                        break
                    
                    try:
                        row_index = row.get_attribute("row-index")
                        if not row_index:
                            continue
                        
                        row_idx = int(row_index)
                        
                        # 시작 인덱스 이전은 스킵
                        if row_idx < start_index:
                            continue
                        
                        if row_index in seen_indices:
                            continue
                        
                        seen_indices.add(row_index)
                        new_found += 1
                        
                        # 상품명 찾기
                        title = ""
                        try:
                            title_elem = row.find_element(By.CSS_SELECTOR, ".whitespace-pre-wrap")
                            title = title_elem.text.strip()
                        except:
                            pass
                        
                        if not title or len(title) < 5:
                            continue
                        
                        # 이미지 찾기 - 썸네일 컬럼에서
                        image_url = ""
                        thumbnail_urls = []
                        try:
                            thumb_cell = row.find_element(By.CSS_SELECTOR, "div[col-id='uploadThumbnails']")
                            
                            # v11: 전체 썸네일 URL 리스트 추출
                            all_imgs = thumb_cell.find_elements(By.CSS_SELECTOR, "img")
                            for thumb_img in all_imgs:
                                thumb_src = thumb_img.get_attribute("src") or ""
                                if thumb_src and thumb_src not in thumbnail_urls:
                                    thumbnail_urls.append(thumb_src)
                            
                            # 확장자 있는 이미지 찾기 (최대 3개 확인)
                            for thumb_url in thumbnail_urls[:3]:
                                if is_valid_image_url(thumb_url):
                                    image_url = thumb_url
                                    break
                            
                            # 확장자 있는 이미지 없으면 첫 번째 이미지 사용
                            if not image_url and thumbnail_urls:
                                image_url = thumbnail_urls[0]
                        except:
                            pass
                        
                        # 판매자 상품 코드 찾기
                        seller_code = ""
                        try:
                            code_cell = row.find_element(By.CSS_SELECTOR, "div[col-id='uploadBulsajaCode']")
                            seller_code = code_cell.text.strip()
                        except:
                            pass
                        
                        # 불사자 상품 ID 찾기 (col-id="ID" 또는 uploadId)
                        bulsaja_id = ""
                        try:
                            id_cell = row.find_element(By.CSS_SELECTOR, "div[col-id='ID']")
                            bulsaja_id = id_cell.text.strip()
                        except:
                            try:
                                id_cell = row.find_element(By.CSS_SELECTOR, "div[col-id='uploadId']")
                                bulsaja_id = id_cell.text.strip()
                            except:
                                # 둘 다 없으면 row-index 사용
                                bulsaja_id = str(row_idx)
                        
                        products.append(ProductRow(
                            index=row_idx,
                            image_url=image_url,
                            original_title=title,
                            seller_code=seller_code,
                            row_element=None,
                            thumbnail_urls=thumbnail_urls if thumbnail_urls else None,
                            bulsaja_id=bulsaja_id
                        ))
                        
                    except Exception:
                        continue
                
                if len(products) >= max_count:
                    break
                
                # 새로운 행이 없으면 카운트 증가
                if new_found == 0:
                    no_new_count += 1
                else:
                    no_new_count = 0
                
                # 아래로 스크롤
                self.main_driver.execute_script(
                    "arguments[0].scrollTop += 400;", grid_body
                )
                time.sleep(0.1)
            
            # index 순으로 정렬
            products.sort(key=lambda p: p.index)
            
            self.gui.log(f"📋 {len(products)}개 상품 추출")
            return products
            
        except Exception as e:
            self.gui.log(f"❌ 상품 리스트 추출 실패: {e}")
            return []
    
    def search_with_aliprice(self, image_url: str) -> List[str]:
        """AliPrice로 유사 상품명 검색 - 디버그 크롬 사용"""
        if not image_url:
            return []
        
        driver = self.main_driver
        
        # main_driver 없으면 디버그 크롬에 연결
        if not driver:
            try:
                port = int(self.gui.port_var.get())
                self.gui.log(f"🔗 디버그 크롬 연결 중 (포트: {port})...")
                
                opt = Options()
                opt.add_experimental_option("debuggerAddress", f"127.0.0.1:{port}")
                
                service = Service(ChromeDriverManager().install())
                driver = webdriver.Chrome(service=service, options=opt)
                self.main_driver = driver  # 연결 유지
                self.gui.log("✅ 디버그 크롬 연결 완료")
            except Exception as e:
                self.gui.log(f"⚠️ 디버그 크롬 연결 실패: {e}")
                return []
        
        titles = []
        original_url = None
        
        try:
            # 현재 URL 저장
            original_url = driver.current_url
            
            # 이미지 검색
            self.gui.log(f"🔍 이미지 검색 중...")
            
            # 이미지 URL 직접 열기 (최대 3번 재시도)
            for page_attempt in range(3):
                driver.get(image_url)
                WebDriverWait(driver, 15).until(
                    lambda d: d.execute_script("return document.readyState") == "complete"
                )
                time.sleep(3)  # 확장 프로그램 로드 대기
                
                # AliPrice 버튼 선택자
                selectors = [
                    ".ap-abi-btn-search",
                    ".ap-abi-btn-search-wrapper",
                    "i.ap-icon-search",
                    ".ap-search-icon",
                    "*[class*='ap-abi']",
                    "*[class*='aliprice' i]",
                    "*[id*='aliprice' i]"
                ]
                
                # 이미지별로 호버 + 버튼 찾기 시도 (최대 3개 이미지)
                all_imgs = driver.find_elements(By.CSS_SELECTOR, "img")
                clicked = False
                
                for img_idx, img in enumerate(all_imgs[:3]):
                    if clicked:
                        break
                    
                    try:
                        # 1. 이미지 호버
                        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", img)
                        time.sleep(0.5)
                        
                        actions = ActionChains(driver)
                        actions.move_to_element(img).perform()
                        if img_idx == 0:
                            self.gui.log("✓ 이미지 호버 중...")
                        time.sleep(2)
                        
                        actions.move_to_element_with_offset(img, 50, 50).perform()
                        time.sleep(2)
                        
                        # 2. 버튼 찾기
                        for selector in selectors:
                            try:
                                buttons = driver.find_elements(By.CSS_SELECTOR, selector)
                                for btn in buttons:
                                    if btn.is_displayed():
                                        driver.execute_script("arguments[0].click();", btn)
                                        self.gui.log(f"✓ AliPrice 버튼 클릭")
                                        clicked = True
                                        break
                            except:
                                continue
                            if clicked:
                                break
                            
                    except Exception as e:
                        pass
                
                if clicked:
                    break
                else:
                    if page_attempt < 2:
                        self.gui.log(f"⚠️ 돋보기 안 나타남, 새로고침 재시도 ({page_attempt + 2}/3)...")
                        time.sleep(1)
            
            if clicked:
                # 결과 대기
                self.gui.log("⏳ 검색 결과 대기 중...")
                time.sleep(4)  # 결과 로딩 대기
                
                # 상품명 추출
                try:
                    cards = driver.find_elements(By.CSS_SELECTOR, ".ap-list-card")
                    for card in cards[:5]:
                        try:
                            title_elem = card.find_element(By.CSS_SELECTOR, ".ap-product-title")
                            title = title_elem.get_attribute("title") or title_elem.text
                            if title:
                                titles.append(title.strip())
                        except:
                            continue
                except Exception as e:
                    self.gui.log(f"⚠️ 결과 추출 실패: {e}")
                
                self.gui.log(f"✅ 유사 상품 {len(titles)}개 발견")
            else:
                self.gui.log("❌ AliPrice 버튼을 찾을 수 없음")
            
        except Exception as e:
            self.gui.log(f"⚠️ AliPrice 검색 실패: {e}")
        
        finally:
            # 원래 페이지로 복귀
            if original_url:
                try:
                    driver.get(original_url)
                    time.sleep(1)
                except:
                    pass
        
        return titles
    
    def generate_title_with_claude(self, original: str, similar_titles: List[str]) -> Tuple[str, str, str, List[str], bool]:
        """Claude로 상품명 1개 + 키워드 생성 → Python으로 2,3번 조합
        Returns: (상품명1, 상품명2, 상품명3, 의심 브랜드 리스트, 금지단어 발견 여부)
        """
        if not self.claude_client:
            return original, original, original, [], False
        
        # 경쟁사 상품명 텍스트 구성
        competitor_text = ""
        if similar_titles:
            for i, t in enumerate(similar_titles[:5], 1):
                competitor_text += f"경쟁사{i}: {t}\n"
        else:
            competitor_text = "(없음)"
        
        # 프롬프트 생성
        prompt = self._build_prompt(original, competitor_text, 35, 45, "basic")
        
        # 프롬프트 유효성 검사
        if not prompt or not isinstance(prompt, str):
            self.gui.log(f"⚠️ 프롬프트 생성 실패")
            return original, original, original, [], False

        selected_model = self.gui.model_var.get()
        temperature = float(self.gui.temp_var.get())
        
        # 최대 3번 재시도 (529 오류만)
        # ★ v2.5: 최대 3번 재시도 (529 오류 + 무효 상품명)
        max_retries = 3
        for attempt in range(max_retries):
            try:
                # temperature 살짝 올리며 재시도 (다양한 응답 유도)
                current_temp = min(temperature + (attempt * 0.1), 1.0)
                
                message = self.claude_client.messages.create(
                    model=selected_model,
                    max_tokens=350,
                    temperature=current_temp,
                    messages=[{"role": "user", "content": prompt}]
                )
                
                response_text = message.content[0].text.strip()
                
                # 응답 파싱 - 상품명 1개 + 키워드 리스트 + 의심단어
                base_title, keywords, claude_suspects = self._parse_title_response(response_text, log_invalid=False)
                
                # ★ v2.5: 무효 상품명이면 재시도
                if not base_title or len(base_title) < 10:
                    if attempt < max_retries - 1:
                        self.gui.log(f"⚠️ 무효 응답 - 재시도 ({attempt + 1}/{max_retries})")
                        time.sleep(1)
                        continue
                    else:
                        self.gui.log(f"⚠️ 재시도 초과 - 원본 사용")
                        base_title = original
                
                # Claude가 판별한 의심단어 사용
                brands = claude_suspects if claude_suspects else []
                
                # 모델 이름 표시
                if 'haiku' in selected_model.lower():
                    model_name = "HAIKU"
                elif 'sonnet' in selected_model.lower():
                    model_name = "SONNET"
                elif 'opus' in selected_model.lower():
                    model_name = "OPUS"
                else:
                    model_name = selected_model.split('-')[-1].upper()
                
                self.gui.log(f"📝 모델: {model_name} (temp: {current_temp})")
                
                # 키워드 로그
                if keywords:
                    self.gui.log(f"  🔑 키워드: {', '.join(keywords[:7])}{'...' if len(keywords) > 7 else ''}")
                
                # 1번 상품명 후처리
                base_title, forbidden_found = self._post_process_title(base_title, original)
                
                # 키워드로 3개 상품명 생성 (35자 미만이면 키워드 추가)
                new_title, new_title2, new_title3 = self._generate_titles_from_keywords(base_title, keywords)
                
                # 2, 3번도 후처리
                new_title2, _ = self._post_process_title(new_title2, original)
                new_title3, _ = self._post_process_title(new_title3, original)
                
                # 로그에 3개 상품명 표시
                self.gui.log(f"  ✨ 1: {new_title[:50]}{'...' if len(new_title) > 50 else ''}")
                if new_title2 != new_title:
                    self.gui.log(f"  ✨ 2: {new_title2[:50]}{'...' if len(new_title2) > 50 else ''}")
                if new_title3 != new_title:
                    self.gui.log(f"  ✨ 3: {new_title3[:50]}{'...' if len(new_title3) > 50 else ''}")
                
                return (new_title if new_title else original), new_title2, new_title3, brands, forbidden_found
                
            except Exception as e:
                error_str = str(e)
                # 529 Overloaded 오류 시 재시도
                if "529" in error_str or "overload" in error_str.lower():
                    if attempt < max_retries - 1:
                        wait_time = (attempt + 1) * 3
                        self.gui.log(f"⚠️ API 과부하 - {wait_time}초 후 재시도 ({attempt + 1}/{max_retries})")
                        time.sleep(wait_time)
                        continue
                    else:
                        self.gui.log(f"❌ Claude 생성 실패 (재시도 초과): {e}")
                        return original, original, original, [], False
                else:
                    self.gui.log(f"⚠️ Claude 생성 실패: {e}")
                    return original, original, original, [], False
        
        return original, original, original, [], False
    
    def _build_prompt(self, original: str, competitor_text: str, min_length: int, max_length: int, logic: str) -> str:
        """프롬프트 생성 - 상품명 1개 + 키워드 20개 (메인키워드 2~3번째 위치)"""
        
        # 경쟁사 데이터 유무 확인
        has_competitor = competitor_text and competitor_text.strip() != "(없음)"
        
        if has_competitor:
            return f"""네이버 스마트스토어 상품명 최적화 전문가입니다.

【유사 상품명 (네이버 상위 판매자)】
{competitor_text}

【원본 상품명】
{original}

【핵심 규칙 - 반드시 준수】
★ 메인키워드(핵심 상품명)는 반드시 2~3번째 단어에 위치!
★ 35-45자 (공백 포함), 7~10단어
★ 금지: 모델명, 숫자스펙, 수량, 광고문구, 특수기호, 색상, 브랜드명
★ 원본이 불완전해도 반드시 상품명 생성 (설명/질문/거절 금지)

【상품명 구조】
[수식어 1개] + [메인키워드] + [세부키워드들] + [대형키워드]

【올바른 예시】
원본: "차량 무선 전동 광택기 폴리싱"
→ "차량관리 광택기 전동 무선 폴리싱 자동차 세차용품 연마기"

원본: "알루미늄 계단 발판 작업대"
→ "현장용 계단발판 알루미늄 작업 보조대 공장 산업용 사다리"

【출력형식】
상품명: (35-45자, 메인키워드 2~3번째 위치)
키워드: (상품명에 미사용 관련 키워드 20개, 쉼표 구분)
의심단어: (브랜드/캐릭터/고유명사. 없으면 "없음")"""
        else:
            return f"""네이버 스마트스토어 상품명 최적화 전문가입니다.

【원본 상품명】
{original}

【핵심 규칙 - 반드시 준수】
★ 메인키워드(핵심 상품명)는 반드시 2~3번째 단어에 위치!
★ 35-45자 (공백 포함), 7~10단어
★ 금지: 모델명, 숫자스펙, 수량, 광고문구, 특수기호, 색상, 브랜드명
★ 원본이 불완전해도 반드시 상품명 생성 (설명/질문/거절 금지)

【상품명 구조】
[수식어 1개] + [메인키워드] + [세부키워드들] + [대형키워드]

【올바른 예시】
원본: "주방용 스텐 수납함 정리함 선반"
→ "가정용 스텐수납함 주방정리 다용도 선반 스테인리스 주방용품"

원본: "산업용 계단 사다리 발판"
→ "현장용 계단사다리 알루미늄 작업 발판 공장 산업용 보조대"

원본: "바코터 코팅 시험 잉크 측정"
→ "정밀 바코터 코팅기 시험장비 잉크측정 도포 실험용품 연구기자재"

【출력형식】
상품명: (35-45자, 메인키워드 2~3번째 위치)
키워드: (상품명에 미사용 관련 키워드 20개, 쉼표 구분)
의심단어: (브랜드/캐릭터/고유명사. 없으면 "없음")"""
    
    def _parse_title_response(self, response_text: str, log_invalid: bool = True) -> Tuple[str, List[str], List[str]]:
        """Claude 응답에서 상품명 1개 + 키워드 리스트 + 의심단어 추출
        Args:
            response_text: Claude 응답 텍스트
            log_invalid: 무효 상품명일 때 로그 출력 여부 (재시도 시 False)
        Returns: (상품명, 키워드 리스트, 의심단어 리스트)
        """
        lines = [l.strip() for l in response_text.splitlines() if l.strip()]
        
        title = ""
        keywords = []
        suspects = []
        
        # 무효 상품명 패턴 (이런 응답이 오면 상품명으로 사용하면 안됨)
        invalid_patterns = [
            "분석 결과", "최적화된 상품명", "제안해드리겠습니다", "제안드립니다",
            "다음과 같습니다", "생성하겠습니다", "알려드리겠습니다", "도와드리겠습니다",
            "확인해보겠습니다", "검토해보겠습니다", "살펴보겠습니다",
            "죄송합니다", "어렵습니다", "불가능합니다", "없습니다"
        ]
        
        def is_valid_title(t: str) -> bool:
            """상품명이 유효한지 검증"""
            if not t or len(t) < 10:
                return False
            for pattern in invalid_patterns:
                if pattern in t:
                    return False
            korean_chars = len(re.findall(r'[가-힣]', t))
            if korean_chars < 5:
                return False
            return True
        
        for line in lines:
            # "상품명:" 파싱
            if line.startswith("상품명:") or line.startswith("상품명："):
                title = re.sub(r"^상품명\s*[:：]\s*", "", line).strip()
                title = title.strip('"\'')
                continue
            
            # "키워드:" 파싱
            if line.startswith("키워드:") or line.startswith("키워드："):
                kw_text = re.sub(r"^키워드\s*[:：]\s*", "", line).strip()
                keywords = [k.strip() for k in kw_text.split(",") if k.strip()]
                continue
            
            # "의심단어:" 파싱 - v2.1: 괄호 포함 원본 유지 (GUI에서 분리)
            if line.startswith("의심단어:") or line.startswith("의심단어："):
                suspect_text = re.sub(r"^의심단어\s*[:：]\s*", "", line).strip()
                if suspect_text and suspect_text != "없음":
                    suspects = [s.strip() for s in suspect_text.split(",") if s.strip() and s.strip() != "없음"]
                continue
        
        # 상품명을 찾지 못한 경우 첫 번째 유효한 줄 사용
        if not title:
            for line in lines:
                if "【" in line or "】" in line:
                    continue
                if ":" in line or "：" in line:
                    continue
                if len(line) >= 10:
                    title = line.strip('"\'')
                    break
        
        # ★ 무효 상품명 검증
        if not is_valid_title(title):
            if log_invalid:
                self.gui.log(f"⚠️ 무효 상품명 감지 → 원본 사용")
            title = ""  # 빈 문자열 반환하면 원본 사용됨
        
        return title, keywords, suspects
    
    def _generate_titles_from_keywords(self, base_title: str, keywords: List[str]) -> Tuple[str, str, str]:
        """키워드를 활용하여 상품명 3개 생성
        - 1번: Claude 상품명에서 금지/제거단어 삭제 후 키워드로 보충
        - 2번: 용도+메인 키워드 형식
        - 3번: 1,2번과 겹치지 않는 키워드 조합
        Returns: (상품명1, 상품명2, 상품명3)
        """
        import random
        
        target_min = 30
        target_max = 45
        
        # ★ 1단계: 상품명에서 금지/제거단어 필터링
        base_words = base_title.split()
        filtered_words = []
        for word in base_words:
            # 금지단어 체크
            if self.banned_words and word in self.banned_words:
                continue
            # 제거단어 체크
            if REMOVE_WORDS and word in REMOVE_WORDS:
                continue
            filtered_words.append(word)
        
        # ★ 2단계: 키워드에서 금지/제거단어 필터링 (사용 가능한 키워드만)
        valid_keywords = []
        for kw in keywords:
            if len(kw) <= 1:
                continue
            if kw in filtered_words:  # 이미 상품명에 있음
                continue
            if self.banned_words and kw in self.banned_words:
                continue
            if REMOVE_WORDS and kw in REMOVE_WORDS:
                continue
            valid_keywords.append(kw)
        
        # ★ 3단계: 필터링된 상품명이 짧으면 키워드로 채우기
        title1_words = filtered_words.copy()
        title1 = " ".join(title1_words)
        
        if len(title1) < target_min and valid_keywords:
            for kw in valid_keywords:
                if len(title1) + len(kw) + 1 <= target_max:
                    title1_words.append(kw)
                    title1 = " ".join(title1_words)
                if len(title1) >= target_min:
                    break
        
        # 용도 키워드 목록
        usage_keywords = ['업소용', '가정용', '캠핑용', '휴대용', '농업용', '사무용', '주방용', 
                         '욕실용', '야외용', '실내용', '차량용', '산업용', '의료용', '학교용']
        
        # === 2번: 용도+메인 키워드 형식 ===
        usage_found = None
        for kw in valid_keywords:
            if kw in usage_keywords or kw.endswith('용'):
                usage_found = kw
                break
        
        if not usage_found:
            random.shuffle(usage_keywords)
            usage_found = usage_keywords[0]
        
        # 메인 키워드 (필터링된 상품명에서)
        main_keywords = filtered_words[:3] if len(filtered_words) >= 3 else filtered_words
        title2_words = [usage_found] + [w for w in main_keywords if w != usage_found]
        
        title2 = " ".join(title2_words)
        remaining_kw = [kw for kw in valid_keywords if kw not in title2_words and kw != usage_found]
        random.shuffle(remaining_kw)
        
        for kw in remaining_kw:
            if len(title2) + len(kw) + 1 <= target_max:
                title2_words.append(kw)
                title2 = " ".join(title2_words)
            if len(title2) >= target_min:
                break
        
        # === 3번: 1,2번과 겹치지 않는 키워드 조합 ===
        used_words = set(title1_words + title2_words)
        unique_keywords = [kw for kw in valid_keywords if kw not in used_words]
        
        title3_words = main_keywords.copy()
        random.shuffle(unique_keywords)
        
        for kw in unique_keywords:
            if len(" ".join(title3_words)) + len(kw) + 1 <= target_max:
                title3_words.append(kw)
            if len(" ".join(title3_words)) >= target_min:
                break
        
        # 부족하면 기존 키워드에서 추가
        if len(" ".join(title3_words)) < target_min:
            for kw in valid_keywords:
                if kw not in title3_words:
                    if len(" ".join(title3_words)) + len(kw) + 1 <= target_max:
                        title3_words.append(kw)
                if len(" ".join(title3_words)) >= target_min:
                    break
        
        title3 = " ".join(title3_words)
        
        return title1, title2, title3
    
    def _post_process_title(self, title: str, original: str) -> Tuple[str, bool]:
        """상품명 후처리 (필터링, 정리)
        Returns: (처리된 제목, 금지단어 발견 여부)
        """
        if not title:
            return original, False
        
        forbidden_found = False
        
        # 1) 특수기호 제거 (한글, 영문, 숫자, 공백만 허용)
        title = re.sub(r'[^\w\s가-힣]', '', title).strip()
        title = re.sub(r'\s+', ' ', title).strip()
        
        # 2) 금지단어 필터링
        if self.banned_words:
            filtered_title, found_words = filter_banned_words(title, self.banned_words)
            if found_words:
                self.gui.log(f"🚫 금지단어 발견: {', '.join(found_words[:3])}{'...' if len(found_words) > 3 else ''}")
                title = filtered_title
                forbidden_found = True
        
        # 3) 광고성 형용사 제거
        for bad in BANNED_ADJECTIVES:
            title = re.sub(re.escape(bad), " ", title, flags=re.IGNORECASE)
        
        # 4) 모델명 패턴 제거
        model_pattern = r'\b[A-Za-z]+[-]?[0-9]+[A-Za-z]*\b|\b[A-Z]{1,3}\b'
        title = re.sub(model_pattern, " ", title)
        
        # 5) 숫자+단위 패턴 제거
        size_pattern = r'\b\d+\.?\d*\s?(cm|mm|m|L|ml|kg|g|인치|리터)\b'
        title = re.sub(size_pattern, " ", title, flags=re.IGNORECASE)
        
        # 6) 수량/단수 표현 제거
        title = remove_quantity_expressions(title)
        
        # 7) 공백 정리
        title = re.sub(r'\s+', ' ', title).strip()
        
        # 8) 한 글자 단어 제거
        words = [w for w in title.split() if len(w) > 1]
        
        # 9) 동의어 중복 제거
        words = self._remove_synonym_duplicates(words)
        
        # 10) 완전 동일 단어 중복 제거
        unique_words = []
        seen = set()
        for w in words:
            w_lower = w.lower()
            if w_lower not in seen:
                unique_words.append(w)
                seen.add(w_lower)
        
        title = " ".join(unique_words)
        
        # 11) 글자 수 체크
        if len(title) > MAX_CHARS:
            words = title.split()
            result = []
            current_len = 0
            for w in words:
                space_len = 1 if result else 0
                if current_len + len(w) + space_len <= MAX_CHARS:
                    result.append(w)
                    current_len += len(w) + space_len
                else:
                    break
            title = " ".join(result)
        
        return (title if title else original), forbidden_found
    
    def verify_danger_with_context(self, title: str, detected_words: List[str]) -> dict:
        """맥락을 고려한 위험 검증 (Claude API)
        
        Args:
            title: 전체 상품명
            detected_words: 감지된 위험 단어들
            
        Returns:
            {
                'is_dangerous': bool,  # 실제 위험 여부
                'reason': str,         # 판단 이유
                'confidence': str      # 확신도 (high/medium/low)
            }
        """
        if not self.claude_client:
            # Claude API 없으면 기본 패턴 판단 유지
            return {'is_dangerous': True, 'reason': '패턴 기반 감지', 'confidence': 'low'}
        
        prompt = f"""당신은 전자상거래 상품 안전성 판단 전문가입니다.

【분석 대상 상품명】
{title}

【감지된 단어】
{', '.join(detected_words)}

【판단 기준】
다음 카테고리에 해당하면 "위험":
1. 성인용품 (란제리, 섹시, 성인 등)
2. 의료기기 (혈압계, 체온계, 레이저 치료기 등 - 단, 진열대/쇼케이스는 제외)
3. 유아용품 (젖병, 기저귀, 아기 등)
4. 판매금지 (도검, 총기, 담배, 가스 등)
5. 주류 (소주, 맥주, 와인 등 - 단, "미술", "예술" 같은 다른 의미는 제외)
6. 식품 (과자, 사탕 등 - 단, 진열대/쇼케이스/보관함은 제외)

【중요 규칙】
⚠️ 맥락을 반드시 고려할 것!
- "미술" → "술"이 포함되어도 미술용품이면 안전
- "보온 진열대 캔커피 온장고 호두과자 약국 쇼케이스" → "과자"가 있지만 "진열대", "쇼케이스"가 메인이므로 안전
- "강아지 보정틀" → 반려동물 용품이지만 위험 아님
- "엔진 체인 샤프너" → 공구이므로 안전

【출력 형식】
판정: 위험 또는 안전
이유: (한 줄로 간단히)
확신도: high 또는 medium 또는 low

예시:
판정: 안전
이유: 진열대와 쇼케이스가 핵심 키워드로 식품 판매 기구임
확신도: high"""

        # 최대 3번 재시도 (529 Overloaded 오류 대응)
        max_retries = 3
        for attempt in range(max_retries):
            try:
                message = self.claude_client.messages.create(
                    model="claude-3-5-haiku-20241022",  # 빠른 모델
                    max_tokens=150,
                    temperature=0,  # 일관성 있는 판단
                    messages=[{"role": "user", "content": prompt}]
                )
                
                response = message.content[0].text.strip()
                
                # 응답 파싱
                is_dangerous = "위험" in response.split('\n')[0]
                reason_line = [line for line in response.split('\n') if '이유:' in line]
                reason = reason_line[0].split('이유:')[1].strip() if reason_line else "맥락 분석 완료"
                
                conf_line = [line for line in response.split('\n') if '확신도:' in line]
                confidence = 'medium'
                if conf_line:
                    conf_text = conf_line[0].lower()
                    if 'high' in conf_text:
                        confidence = 'high'
                    elif 'low' in conf_text:
                        confidence = 'low'
                
                return {
                    'is_dangerous': is_dangerous,
                    'reason': reason,
                    'confidence': confidence
                }
                
            except Exception as e:
                error_str = str(e)
                # 529 Overloaded 오류 시 재시도
                if "529" in error_str or "overload" in error_str.lower():
                    if attempt < max_retries - 1:
                        wait_time = (attempt + 1) * 3  # 3초, 6초, 9초
                        self.gui.log(f"⚠️ API 과부하 - {wait_time}초 후 재시도 ({attempt + 1}/{max_retries})")
                        time.sleep(wait_time)
                        continue
                    else:
                        self.gui.log(f"⚠️ 맥락 분석 실패 (재시도 초과): {str(e)[:50]}")
                        return {'is_dangerous': False, 'reason': '분석 실패 - 안전으로 판단', 'confidence': 'low'}
                else:
                    self.gui.log(f"⚠️ 맥락 분석 실패: {str(e)[:50]}")
                    # 실패 시 안전으로 판단 (오탐 방지)
                    return {'is_dangerous': False, 'reason': '분석 실패 - 안전으로 판단', 'confidence': 'low'}
        
        return {'is_dangerous': False, 'reason': '분석 실패', 'confidence': 'low'}
    
    def _remove_synonym_duplicates(self, words: List[str]) -> List[str]:
        """동의어 중복 제거 - 먼저 나온 단어만 유지"""
        result = []
        used_synonyms = set()
        
        for word in words:
            word_lower = word.lower()
            
            # 이미 사용된 동의어인지 확인
            if word_lower in used_synonyms:
                continue
            
            result.append(word)
            used_synonyms.add(word_lower)
            
            # 현재 단어의 동의어들도 사용 불가 처리
            if word_lower in SYNONYM_HINTS:
                for syn in SYNONYM_HINTS[word_lower]:
                    used_synonyms.add(syn.lower())
        
        return result
    
    def generate_title_original_only(self, original: str) -> Tuple[str, str, str, List[str], bool]:
        """기존 상품명만으로 상품명 1개 + 키워드 생성 → Python으로 2,3번 조합
        Returns: (상품명1, 상품명2, 상품명3, 의심 브랜드 리스트, 금지단어 발견 여부)
        """
        if not self.claude_client:
            return original, original, original, [], False
        
        # 프롬프트 생성 (경쟁사 없음)
        prompt = self._build_prompt(original, "(없음)", 35, 45, "basic")
        
        # 프롬프트 유효성 검사
        if not prompt or not isinstance(prompt, str):
            self.gui.log(f"⚠️ 프롬프트 생성 실패")
            return original, original, original, [], False

        selected_model = self.gui.model_var.get()
        temperature = float(self.gui.temp_var.get())
        
        # ★ v2.5: 최대 3번 재시도 (529 오류 + 무효 상품명)
        max_retries = 3
        for attempt in range(max_retries):
            try:
                # temperature 살짝 올리며 재시도 (다양한 응답 유도)
                current_temp = min(temperature + (attempt * 0.1), 1.0)
                
                message = self.claude_client.messages.create(
                    model=selected_model,
                    max_tokens=350,
                    temperature=current_temp,
                    messages=[{"role": "user", "content": prompt}]
                )
                
                response_text = message.content[0].text.strip()
                
                # 응답 파싱 - 상품명 1개 + 키워드 리스트 + 의심단어
                base_title, keywords, claude_suspects = self._parse_title_response(response_text, log_invalid=False)
                
                # ★ v2.5: 무효 상품명이면 재시도
                if not base_title or len(base_title) < 10:
                    if attempt < max_retries - 1:
                        self.gui.log(f"⚠️ 무효 응답 - 재시도 ({attempt + 1}/{max_retries})")
                        time.sleep(1)
                        continue
                    else:
                        self.gui.log(f"⚠️ 재시도 초과 - 원본 사용")
                        base_title = original
                
                # Claude가 판별한 의심단어 사용
                brands = claude_suspects if claude_suspects else []
                
                # 모델 이름 표시
                if 'haiku' in selected_model.lower():
                    model_name = "HAIKU"
                elif 'sonnet' in selected_model.lower():
                    model_name = "SONNET"
                elif 'opus' in selected_model.lower():
                    model_name = "OPUS"
                else:
                    model_name = selected_model.split('-')[-1].upper()
                
                self.gui.log(f"📝 모델: {model_name} (temp: {current_temp})")
                
                # 키워드 로그
                if keywords:
                    self.gui.log(f"  🔑 키워드: {', '.join(keywords[:7])}{'...' if len(keywords) > 7 else ''}")
                
                # 1번 상품명 후처리
                base_title, forbidden_found = self._post_process_title(base_title, original)
                
                # 키워드로 3개 상품명 생성 (35자 미만이면 키워드 추가)
                new_title, new_title2, new_title3 = self._generate_titles_from_keywords(base_title, keywords)
                
                # 2, 3번도 후처리
                new_title2, _ = self._post_process_title(new_title2, original)
                new_title3, _ = self._post_process_title(new_title3, original)
                
                # 로그에 3개 상품명 표시
                self.gui.log(f"  ✨ 1: {new_title[:50]}{'...' if len(new_title) > 50 else ''}")
                if new_title2 != new_title:
                    self.gui.log(f"  ✨ 2: {new_title2[:50]}{'...' if len(new_title2) > 50 else ''}")
                if new_title3 != new_title:
                    self.gui.log(f"  ✨ 3: {new_title3[:50]}{'...' if len(new_title3) > 50 else ''}")
                
                return (new_title if new_title else original), new_title2, new_title3, brands, forbidden_found
                
            except Exception as e:
                error_str = str(e)
                # 529 Overloaded 오류 시 재시도
                if "529" in error_str or "overload" in error_str.lower():
                    if attempt < max_retries - 1:
                        wait_time = (attempt + 1) * 3
                        self.gui.log(f"⚠️ API 과부하 - {wait_time}초 후 재시도 ({attempt + 1}/{max_retries})")
                        time.sleep(wait_time)
                        continue
                    else:
                        self.gui.log(f"❌ Claude 생성 실패 (재시도 초과): {e}")
                        return original, original, original, [], False
                else:
                    self.gui.log(f"⚠️ Claude 생성 실패: {e}")
                    return original, original, original, [], False
        
        return original, original, original, [], False
    
    def analyze_brands_with_claude(self, titles: List[str]) -> List[str]:
        """Claude에게 상품명들에서 브랜드/상호 추출 요청"""
        if not self.claude_client or not titles:
            return []
        
        try:
            titles_text = "\n".join(f"- {t}" for t in titles[:20])  # 최대 20개 상품명
            
            prompt = f"""다음 상품명들에서 브랜드명, 상호명, 회사명으로 보이는 단어만 추출해주세요.

【상품명 목록】
{titles_text}

【제외 대상 - 이것들은 브랜드가 아닙니다】
- 일반 명사: 가방, 선반, 트레이, 거치대, 보관함, 정리함, 수납함, 케이스 등
- 재질: 스텐, 플라스틱, 우드, 원목, 철제, 알루미늄, 스테인리스 등
- 용도/장소: 주방, 욕실, 거실, 사무용, 가정용, 업소용, 산업용 등
- 상품 특성: 접이식, 휴대용, 자동, 수동, 방수, 대형, 소형, 미니 등
- 색상: 블랙, 화이트, 그레이, 브라운 등
- 일반 영단어: PC, CNC, LED, USB, DIY 등

【추출 대상 - 브랜드 가능성 높은 것】
- 영문+한글 조합: LG전자, 삼성테크, SK텔레콤
- 특이한 조어: 꾸꾸리빙, 오스타, 로보체인 등
- ~텍, ~코, ~몰, ~샵 등으로 끝나는 상호명

【출력 형식】
브랜드로 판단되는 단어만 쉼표로 구분하여 출력 (중복 제거)
없으면 "없음" 출력"""

            message = self.claude_client.messages.create(
                model="claude-3-5-haiku-20241022",
                max_tokens=300,
                temperature=0,
                messages=[{"role": "user", "content": prompt}]
            )
            
            result = message.content[0].text.strip()
            
            if result == "없음" or "없습니다" in result or not result:
                return []
            
            # 쉼표로 분리하고 정리
            confirmed_brands = [w.strip() for w in result.split(",") if w.strip() and w.strip() != "없음"]
            return confirmed_brands
            
        except Exception as e:
            self.gui.log(f"⚠️ 브랜드 분석 실패: {e}")
            return []
    
    def analyze_brand_descriptions(self, words: List[str]) -> dict:
        """Claude에게 의심단어들의 브랜드 여부 판단 요청
        Returns: {단어: {판정: "브랜드/일반명사/불확실", 설명: "..."}}
        """
        if not self.claude_client or not words:
            return {}
        
        try:
            words_text = ", ".join(words[:30])  # 최대 30개
            
            prompt = f"""다음 단어들이 브랜드/상표인지 일반명사인지 판단해주세요.

【단어 목록】
{words_text}

【판단 기준】
- 브랜드: 실제 등록된 상표, 회사명, 상호명
- 일반명사: 제품 특성, 재질, 용도를 나타내는 일반적인 단어
- 불확실: 판단하기 어려운 경우

【출력 형식 - 반드시 이 형식으로】
단어1|판정|간단한 이유
단어2|판정|간단한 이유
...

예시:
브라켓|일반명사|고정 부품을 뜻하는 일반 용어
삼성|브랜드|한국 대기업 상표
오스텍|불확실|브랜드일 수 있으나 확인 필요"""

            self.gui.log("🔍 브랜드 여부 분석 중...")
            
            message = self.claude_client.messages.create(
                model="claude-3-5-haiku-20241022",
                max_tokens=1000,
                temperature=0,
                messages=[{"role": "user", "content": prompt}]
            )
            
            result = message.content[0].text.strip()
            
            # 결과 파싱
            descriptions = {}
            for line in result.splitlines():
                line = line.strip()
                if not line or "|" not in line:
                    continue
                
                parts = line.split("|")
                if len(parts) >= 3:
                    word = parts[0].strip()
                    judgment = parts[1].strip()
                    reason = parts[2].strip()
                    
                    # 판정 정규화
                    if "브랜드" in judgment:
                        judgment = "🔴 브랜드"
                    elif "일반" in judgment:
                        judgment = "🟢 일반명사"
                    else:
                        judgment = "🟡 불확실"
                    
                    descriptions[word] = {
                        'judgment': judgment,
                        'reason': reason
                    }
            
            self.gui.log(f"✅ {len(descriptions)}개 단어 분석 완료")
            return descriptions
            
        except Exception as e:
            self.gui.log(f"⚠️ 브랜드 설명 분석 실패: {e}")
            return {}
    
    def analyze_products_risk(self, products_data: List[dict]) -> dict:
        """
        Claude로 상품명 일괄 위험 분석 (문맥 기반)
        
        Args:
            products_data: [{'index': 0, 'title': '상품명', 'original': '원본명'}, ...]
        
        Returns: {
            'danger': [{'index': 0, 'title': '...', 'reason': '...', 'keywords': [...]}, ...],
            'suspect': [{'index': 0, 'title': '...', 'reason': '...', 'keywords': [...]}, ...],
            'safe': [...]
        }
        """
        if not self.claude_client or not products_data:
            return {'danger': [], 'suspect': [], 'safe': products_data}
        
        try:
            # 상품명 리스트 생성
            products_text = "\n".join([
                f"{i+1}. {p['title']}" 
                for i, p in enumerate(products_data[:50])  # 최대 50개
            ])
            
            prompt = f"""다음 상품명들을 분석해주세요.

【상품명 목록】
{products_text}

【분석 기준】
1. 판매금지 품목:
   - 의료기기/의료용품 (레이저 치료기, 혈압계, 의료용 보조기구 등)
   - 가스류 (에어컨 냉매가스, 부탄가스 충전, LPG 관련)
   - 무기류 (도검, 총기, 석궁, 전기충격기)
   - 성인용품 (명확한 성인용품만)
   - 담배/니코틴 (전자담배, 액상 니코틴)

2. 브랜드/지재권 의심:
   - 유명 브랜드명이 포함된 경우만 (나이키, 아디다스, 샤넬 등)
   - 상표권 침해 가능성 있는 고유 브랜드명만

3. 안전:
   - 일반 생활용품, 문구, 공구, 수납용품 등
   - "액상 교반기"처럼 위험 키워드가 있어도 문맥상 안전한 제품

【중요】
- "액상"이 들어가도 "액상 교반기", "액상 디스펜서"는 안전 (기계/도구)
- "의료"가 들어가도 "의료 수납함", "의료 서류 보관함"은 안전 (수납용품)
- 문맥을 보고 실제 판매금지 품목인지 판단
- "대형", "업소용", "산업용", "휴대용", "이동식" 등 일반 형용사/명사는 의심단어가 아님!
- 브랜드가 아닌 일반 제품 설명어는 keywords에 포함하지 마세요

【출력 형식 - JSON으로】
{{
  "danger": [
    {{"num": 1, "reason": "판매금지 이유", "keywords": ["문제단어1"]}}
  ],
  "suspect": [
    {{"num": 3, "reason": "의심 이유", "keywords": ["브랜드명"]}}
  ],
  "safe": [2, 4, 5, 6]
}}

JSON만 출력하세요."""

            self.gui.log("🔍 상품 위험도 분석 중...")
            
            message = self.claude_client.messages.create(
                model="claude-3-5-haiku-20241022",
                max_tokens=2000,
                temperature=0,
                messages=[{"role": "user", "content": prompt}]
            )
            
            result_text = message.content[0].text.strip()
            
            # JSON 파싱
            import json
            # ```json ... ``` 제거
            if "```" in result_text:
                result_text = re.sub(r'```json\s*', '', result_text)
                result_text = re.sub(r'```\s*', '', result_text)
            
            try:
                analysis = json.loads(result_text)
            except:
                self.gui.log("⚠️ JSON 파싱 실패, 기본값 사용")
                return {'danger': [], 'suspect': [], 'safe': products_data}
            
            # 결과 변환
            result = {'danger': [], 'suspect': [], 'safe': []}
            
            # danger 처리
            for item in analysis.get('danger', []):
                num = item.get('num', 0) - 1  # 1-based to 0-based
                if 0 <= num < len(products_data):
                    result['danger'].append({
                        'index': products_data[num]['index'],
                        'title': products_data[num]['title'],
                        'reason': item.get('reason', ''),
                        'keywords': item.get('keywords', [])
                    })
            
            # suspect 처리
            for item in analysis.get('suspect', []):
                num = item.get('num', 0) - 1
                if 0 <= num < len(products_data):
                    result['suspect'].append({
                        'index': products_data[num]['index'],
                        'title': products_data[num]['title'],
                        'reason': item.get('reason', ''),
                        'keywords': item.get('keywords', [])
                    })
            
            # safe 처리
            safe_nums = analysis.get('safe', [])
            for num in safe_nums:
                idx = num - 1  # 1-based to 0-based
                if 0 <= idx < len(products_data):
                    result['safe'].append(products_data[idx])
            
            danger_count = len(result['danger'])
            suspect_count = len(result['suspect'])
            safe_count = len(result['safe'])
            
            self.gui.log(f"✅ 분석 완료: 위험 {danger_count}개, 의심 {suspect_count}개, 안전 {safe_count}개")
            
            return result
            
        except Exception as e:
            self.gui.log(f"⚠️ 위험 분석 실패: {e}")
            import traceback
            traceback.print_exc()
            return {'danger': [], 'suspect': [], 'safe': products_data}
    
    def update_product_title(self, product: ProductRow, new_title: str, tag_name: str = None) -> bool:
        """상품명 업데이트 (+ 태그 적용) - API 모드 또는 Selenium 모드"""
        
        # ★ v3.0: API 모드
        if self.use_api_mode and self.api_client:
            return self._update_product_via_api(product, new_title, tag_name)
        
        # 레거시 Selenium 모드
        return self._update_product_via_selenium(product, new_title, tag_name)
    
    def _update_product_via_api(self, product: ProductRow, new_title: str, tag_name: str = None) -> bool:
        """API로 상품명 업데이트"""
        try:
            if not product.bulsaja_id:
                self.gui.log(f"⚠️ 상품 ID 없음")
                return False
            
            # 상품명 수정
            self.api_client.update_single_product(product.bulsaja_id, new_title)
            self.gui.log(f"✅ 변경 완료: {new_title[:40]}...")
            
            # 태그 적용
            if tag_name:
                self.api_client.apply_tag_single(product.bulsaja_id, tag_name)
                self.gui.log(f"🏷️ 태그: {tag_name}")
            
            return True
            
        except Exception as e:
            self.gui.log(f"❌ API 업데이트 실패: {e}")
            return False
    
    def _update_product_via_selenium(self, product: ProductRow, new_title: str, tag_name: str = None) -> bool:
        """Selenium으로 상품명 업데이트 (레거시)"""
        try:
            # AG Grid 가상 스크롤: 해당 행이 보이도록 스크롤
            grid_body = self.main_driver.find_element(By.CSS_SELECTOR, ".ag-body-viewport")
            
            # row-index에 맞게 스크롤 (각 행 높이 약 126px)
            scroll_position = product.index * 126
            self.main_driver.execute_script(
                f"arguments[0].scrollTop = {scroll_position};", grid_body
            )
            time.sleep(0.2)
            
            # row-index로 행을 찾기
            row = None
            for attempt in range(3):
                try:
                    row = self.main_driver.find_element(By.CSS_SELECTOR, f"div[role='row'][row-index='{product.index}']")
                    break
                except:
                    # 스크롤 조정
                    self.main_driver.execute_script(
                        f"arguments[0].scrollTop = {scroll_position - 200};", grid_body
                    )
                    time.sleep(0.3)
            
            if not row:
                self.gui.log(f"⚠️ row-index {product.index} 찾을 수 없음")
                return False
            
            title_elem = row.find_element(By.CSS_SELECTOR, ".whitespace-pre-wrap")
            
            # 더블클릭 (JavaScript)
            self.main_driver.execute_script("""
                var evt = new MouseEvent('dblclick', {
                    bubbles: true,
                    cancelable: true,
                    view: window
                });
                arguments[0].dispatchEvent(evt);
            """, title_elem)
            time.sleep(0.3)
            
            # textarea 찾기 (여러 방법 시도)
            textarea = None
            selectors = [
                "textarea:focus",
                "textarea[class*='border']",
                "textarea",
                "input[type='text']:focus",
                "input[type='text']"
            ]
            
            for selector in selectors:
                try:
                    textarea = self.main_driver.find_element(By.CSS_SELECTOR, selector)
                    if textarea.is_displayed():
                        break
                except:
                    continue
            
            if not textarea or not textarea.is_displayed():
                self.gui.log("⚠️ 편집 필드를 찾을 수 없음 - 재시도")
                # 한 번 더 시도
                time.sleep(0.2)
                self.main_driver.execute_script("""
                    var evt = new MouseEvent('dblclick', {
                        bubbles: true,
                        cancelable: true,
                        view: window
                    });
                    arguments[0].dispatchEvent(evt);
                """, title_elem)
                time.sleep(0.3)
                
                for selector in selectors:
                    try:
                        textarea = self.main_driver.find_element(By.CSS_SELECTOR, selector)
                        if textarea.is_displayed():
                            break
                    except:
                        continue
            
            if not textarea or not textarea.is_displayed():
                self.gui.log("❌ 편집 필드를 찾을 수 없음")
                return False
            
            # execCommand insertText 방식으로 값 설정
            self.main_driver.execute_script("""
                var textarea = arguments[0];
                var newTitle = arguments[1];
                
                // 포커스 및 전체 선택
                textarea.focus();
                textarea.select();
                
                // 기존 내용 삭제 후 새 내용 삽입
                document.execCommand('selectAll', false, null);
                document.execCommand('insertText', false, newTitle);
                
                // React가 감지할 수 있도록 이벤트 발생
                textarea.dispatchEvent(new Event('input', { bubbles: true }));
                textarea.dispatchEvent(new Event('change', { bubbles: true }));
                
                // Enter 키 이벤트 발생
                var enterEvent = new KeyboardEvent('keydown', {
                    key: 'Enter',
                    code: 'Enter',
                    keyCode: 13,
                    which: 13,
                    bubbles: true
                });
                textarea.dispatchEvent(enterEvent);
                
                // blur 이벤트도 발생 (편집 종료)
                textarea.blur();
            """, textarea, new_title)
            time.sleep(0.5)  # DOM 업데이트 대기 증가 (0.2 → 0.5초)
            
            self.gui.log(f"✅ 변경 완료: {new_title[:40]}...")
            
            # 태그 적용 (tag_name이 있으면)
            if tag_name:
                # 상품명 변경 후 DOM이 안정화되도록 추가 대기
                time.sleep(0.3)
                self._apply_tag_to_row(row, product.index, tag_name)
            
            return True
            
        except Exception as e:
            self.gui.log(f"❌ 업데이트 실패: {e}")
            return False
    
    def _apply_tag_to_row(self, row, product_index: int, tag_name: str) -> bool:
        """row에서 직접 태그 적용 (JavaScript 방식 - 키보드/마우스 입력 없음)"""
        try:
            # ★ v2.5: smartstore v7 방식 - row 재탐색 (DOM 갱신 대응)
            try:
                row = self.main_driver.find_element(By.CSS_SELECTOR, f"div[role='row'][row-index='{product_index}']")
            except:
                self.gui.log(f"⚠️ row 재탐색 실패 - row-index {product_index}")
                return False
            
            # 행 내의 태그 버튼 찾기 (여러 방법 시도)
            tag_btn = None
            
            # 방법 1: col-id로 태그 셀 찾기
            try:
                tag_cell = row.find_element(By.CSS_SELECTOR, "div[col-id='uploadTag'], div[col-id='tag']")
                tag_btn = tag_cell.find_element(By.CSS_SELECTOR, "button")
            except:
                pass
            
            # 방법 2: aria-haspopup='dialog' 버튼
            if not tag_btn:
                try:
                    tag_btn = row.find_element(By.CSS_SELECTOR, "button[aria-haspopup='dialog']")
                except:
                    pass
            
            # 방법 3: 태그 아이콘이 있는 버튼
            if not tag_btn:
                try:
                    tag_btn = row.find_element(By.XPATH, ".//button[.//svg[contains(@class, 'lucide-tag')]]")
                except:
                    pass
            
            # 방법 4: "태그" 텍스트 포함 버튼
            if not tag_btn:
                try:
                    buttons = row.find_elements(By.CSS_SELECTOR, "button")
                    for btn in buttons:
                        if "태그" in btn.text or "없음" in btn.text:
                            tag_btn = btn
                            break
                except:
                    pass
            
            if not tag_btn:
                self.gui.log(f"⚠️ 태그 버튼 못 찾음")
                return False
            
            # 버튼 클릭 (JavaScript)
            self.main_driver.execute_script("arguments[0].click();", tag_btn)
            time.sleep(0.3)  # 대기 시간 증가 (0.2 → 0.3초)
            
            # 입력 필드 찾기 및 값 입력
            try:
                tag_input = WebDriverWait(self.main_driver, 5).until(  # 대기 시간 증가 (3 → 5초)
                    EC.presence_of_element_located((By.CSS_SELECTOR, "input[cmdk-input], input[placeholder='태그 변경...']"))
                )
                
                # React 입력 필드에 값 설정 (nativeInputValueSetter 사용)
                self.main_driver.execute_script("""
                    var input = arguments[0];
                    var tagName = arguments[1];
                    
                    // React의 경우 native setter를 사용해야 함
                    var nativeInputValueSetter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, 'value').set;
                    nativeInputValueSetter.call(input, tagName);
                    
                    // React가 감지할 수 있도록 이벤트 발생
                    input.dispatchEvent(new Event('input', { bubbles: true }));
                """, tag_input, tag_name)
                time.sleep(0.3)
                
                # 필터링된 옵션 중 첫 번째 매칭 항목 클릭
                try:
                    time.sleep(0.2)
                    options = self.main_driver.find_elements(By.CSS_SELECTOR, "[cmdk-item], [role='option']")
                    for opt in options:
                        if tag_name in (opt.get_attribute("data-value") or opt.text):
                            self.main_driver.execute_script("arguments[0].click();", opt)
                            self.gui.log(f"🏷️ 태그: {tag_name}")
                            return True
                            time.sleep(0.2)  # 태그 적용 후 대기
                    
                    # 옵션이 없으면 Enter 키 이벤트
                    self.main_driver.execute_script("""
                        var input = arguments[0];
                        var enterEvent = new KeyboardEvent('keydown', {
                            key: 'Enter',
                            code: 'Enter',
                            keyCode: 13,
                            which: 13,
                            bubbles: true
                        });
                        input.dispatchEvent(enterEvent);
                    """, tag_input)
                    time.sleep(0.1)
                    self.gui.log(f"🏷️ 태그: {tag_name}")
                    return True
                    
                except Exception as e:
                    # ESC 이벤트로 닫기 (JavaScript)
                    self.main_driver.execute_script("""
                        document.body.dispatchEvent(new KeyboardEvent('keydown', {
                            key: 'Escape', code: 'Escape', keyCode: 27, bubbles: true
                        }));
                    """)
                    return False
                
            except Exception as e:
                self.gui.log(f"⚠️ 태그 입력 실패: {e}")
                # ESC 이벤트로 닫기
                self.main_driver.execute_script("""
                    document.body.dispatchEvent(new KeyboardEvent('keydown', {
                        key: 'Escape', code: 'Escape', keyCode: 27, bubbles: true
                    }));
                """)
                return False
            
        except Exception as e:
            try:
                self.main_driver.execute_script("""
                    document.body.dispatchEvent(new KeyboardEvent('keydown', {
                        key: 'Escape', code: 'Escape', keyCode: 27, bubbles: true
                    }));
                """)
            except:
                pass
            return False
            
            # 방법 1: col-id로 태그 셀 찾기
            try:
                tag_cell = row.find_element(By.CSS_SELECTOR, "div[col-id='uploadTag'], div[col-id='tag']")
                tag_btn = tag_cell.find_element(By.CSS_SELECTOR, "button")
            except:
                pass
            
            # 방법 2: aria-haspopup='dialog' 버튼
            if not tag_btn:
                try:
                    tag_btn = row.find_element(By.CSS_SELECTOR, "button[aria-haspopup='dialog']")
                except:
                    pass
            
            # 방법 3: 태그 아이콘이 있는 버튼
            if not tag_btn:
                try:
                    tag_btn = row.find_element(By.XPATH, ".//button[.//svg[contains(@class, 'lucide-tag')]]")
                except:
                    pass
            
            # 방법 4: "태그" 텍스트 포함 버튼
            if not tag_btn:
                try:
                    buttons = row.find_elements(By.CSS_SELECTOR, "button")
                    for btn in buttons:
                        if "태그" in btn.text or "없음" in btn.text:
                            tag_btn = btn
                            break
                except:
                    pass
            
            if not tag_btn:
                self.gui.log(f"⚠️ 태그 버튼 못 찾음")
                return False
            
            # 버튼 클릭 (JavaScript)
            self.main_driver.execute_script("arguments[0].click();", tag_btn)
            time.sleep(0.2)
            
            # 입력 필드 찾기 및 값 입력
            try:
                tag_input = WebDriverWait(self.main_driver, 3).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "input[cmdk-input], input[placeholder='태그 변경...']"))
                )
                
                # React 입력 필드에 값 설정 (nativeInputValueSetter 사용)
                self.main_driver.execute_script("""
                    var input = arguments[0];
                    var tagName = arguments[1];
                    
                    // React의 경우 native setter를 사용해야 함
                    var nativeInputValueSetter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, 'value').set;
                    nativeInputValueSetter.call(input, tagName);
                    
                    // React가 감지할 수 있도록 이벤트 발생
                    input.dispatchEvent(new Event('input', { bubbles: true }));
                """, tag_input, tag_name)
                time.sleep(0.3)
                
                # 필터링된 옵션 중 첫 번째 매칭 항목 클릭
                try:
                    time.sleep(0.1)
                    options = self.main_driver.find_elements(By.CSS_SELECTOR, "[cmdk-item], [role='option']")
                    for opt in options:
                        if tag_name in (opt.get_attribute("data-value") or opt.text):
                            self.main_driver.execute_script("arguments[0].click();", opt)
                            self.gui.log(f"🏷️ 태그: {tag_name}")
                            return True
                    
                    # 옵션이 없으면 Enter 키 이벤트
                    self.main_driver.execute_script("""
                        var input = arguments[0];
                        var enterEvent = new KeyboardEvent('keydown', {
                            key: 'Enter',
                            code: 'Enter',
                            keyCode: 13,
                            which: 13,
                            bubbles: true
                        });
                        input.dispatchEvent(enterEvent);
                    """, tag_input)
                    time.sleep(0.1)
                    self.gui.log(f"🏷️ 태그: {tag_name}")
                    return True
                    
                except Exception as e:
                    # ESC 이벤트로 닫기 (JavaScript)
                    self.main_driver.execute_script("""
                        document.body.dispatchEvent(new KeyboardEvent('keydown', {
                            key: 'Escape', code: 'Escape', keyCode: 27, bubbles: true
                        }));
                    """)
                    return False
                
            except Exception as e:
                self.gui.log(f"⚠️ 태그 입력 실패: {e}")
                # ESC 이벤트로 닫기
                self.main_driver.execute_script("""
                    document.body.dispatchEvent(new KeyboardEvent('keydown', {
                        key: 'Escape', code: 'Escape', keyCode: 27, bubbles: true
                    }));
                """)
                return False
            
        except Exception as e:
            try:
                self.main_driver.execute_script("""
                    document.body.dispatchEvent(new KeyboardEvent('keydown', {
                        key: 'Escape', code: 'Escape', keyCode: 27, bubbles: true
                    }));
                """)
            except:
                pass
            return False
    
    def set_page_size(self, size: int = 1000):
        """페이지 크기 설정 - API 모드에서는 불필요"""
        
        # ★ v3.0: API 모드에서는 페이지 크기 설정 불필요 (API에서 직접 개수 지정)
        if self.use_api_mode and self.api_client:
            return True
        
        # 레거시 Selenium 모드
        try:
            # 페이지 크기 select 찾기
            select_elem = self.main_driver.find_element(By.CSS_SELECTOR, "select.rounded-md")
            
            # 현재 값 확인
            from selenium.webdriver.support.ui import Select
            select = Select(select_elem)
            current_value = select.first_selected_option.get_attribute("value")
            
            if current_value != str(size):
                self.gui.log(f"📊 페이지 크기 변경: {current_value} → {size}")
                select.select_by_value(str(size))
                self.gui.log("⏳ 데이터 로딩 대기 중...")
                time.sleep(2)  # 로딩 대기 늘림
                return True
            else:
                self.gui.log(f"📊 페이지 크기: {size} (이미 설정됨)")
                return True
                
        except Exception as e:
            self.gui.log(f"⚠️ 페이지 크기 설정 실패: {e}")
            return False
    
    def go_to_next_page(self) -> bool:
        """다음 페이지로 이동 - API 모드에서는 불필요"""
        
        # ★ v3.0: API 모드에서는 페이지 개념 없음 (한 번에 조회)
        if self.use_api_mode and self.api_client:
            return True
        
        # 레거시 Selenium 모드
        try:
            # "다음" 버튼 찾기
            next_btn = self.main_driver.find_element(By.XPATH, "//button[text()='다음']")
            
            # disabled 상태 확인
            if next_btn.get_attribute('disabled'):
                self.gui.log("⚠️ 마지막 페이지입니다 (버튼 비활성화)")
                return False
            
            # 클릭
            next_btn.click()
            time.sleep(0.5)
            
            # 상품 리스트 로딩 완료 대기
            self.gui.log("  ⏳ 페이지 로딩 대기...")
            if not self.wait_for_product_list_loaded():
                self.gui.log("  ⚠️ 페이지 로딩 타임아웃")
            
            time.sleep(0.5)  # 추가 안정화 대기
            return True
        except Exception as e:
            self.gui.log(f"⚠️ 다음 페이지 이동 실패: {e}")
            return False
    
    def select_market_group(self, group_name: str) -> bool:
        """마켓 그룹 선택 - API 모드 또는 Selenium 모드"""
        
        # 숫자인 경우 마켓 그룹 목록에서 실제 그룹 이름으로 변환
        actual_group_name = group_name
        if group_name.isdigit():
            market_groups_str = self.gui.market_groups_var.get().strip()
            market_groups_list = [g.strip() for g in market_groups_str.split(',') if g.strip()]
            idx = int(group_name) - 1  # 1-based -> 0-based
            if 0 <= idx < len(market_groups_list):
                actual_group_name = market_groups_list[idx]
                self.gui.log(f"  🔄 숫자 {group_name} → '{actual_group_name}'")
            else:
                self.gui.log(f"⚠️ 인덱스 {group_name} 범위 초과 (마켓 그룹 {len(market_groups_list)}개)")
                return False
        
        self.gui.log(f"📁 그룹 필터: {actual_group_name}")
        
        # ★ v3.0: API 모드
        if self.use_api_mode and self.api_client:
            self.gui.current_market_group = actual_group_name
            self.gui.log(f"✅ 그룹 '{actual_group_name}' 선택됨 (API)")
            return True
        
        # 레거시 Selenium 모드 - AG Grid 필터
        try:
            # AG Grid 마켓 그룹 필터 입력 필드 찾기
            filter_input = None
            
            # 방법 1: aria-label로 찾기
            try:
                filter_input = self.main_driver.find_element(
                    By.CSS_SELECTOR, "input[aria-label*='마켓 그룹 필터']"
                )
            except:
                pass
            
            # 방법 2: aria-label 마켓 그룹
            if not filter_input:
                try:
                    filter_input = self.main_driver.find_element(
                        By.CSS_SELECTOR, "input[aria-label*='마켓 그룹']"
                    )
                except:
                    pass
            
            # 방법 3: JavaScript로 찾기
            if not filter_input:
                try:
                    filter_input = self.main_driver.execute_script("""
                        var inputs = document.querySelectorAll('input.ag-input-field-input, input.ag-text-field-input');
                        for (var inp of inputs) {
                            var label = inp.getAttribute('aria-label') || '';
                            if (label.includes('마켓') && label.includes('그룹')) {
                                return inp;
                            }
                        }
                        return null;
                    """)
                except:
                    pass
            
            if not filter_input:
                self.gui.log("⚠️ 마켓 그룹 필터 입력 필드를 찾을 수 없음")
                return False
            
            # JavaScript로 필터 입력
            self.gui.log(f"  🔍 필터 입력: '{actual_group_name}'")
            
            self.main_driver.execute_script("""
                var input = arguments[0];
                var value = arguments[1];
                
                input.scrollIntoView({block: 'center'});
                input.focus();
                input.value = '';
                
                // React를 위한 native setter 사용
                var nativeSetter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, 'value').set;
                nativeSetter.call(input, value);
                
                // 이벤트 발생
                input.dispatchEvent(new Event('input', { bubbles: true }));
                input.dispatchEvent(new Event('change', { bubbles: true }));
                input.dispatchEvent(new KeyboardEvent('keyup', { bubbles: true }));
            """, filter_input, actual_group_name)
            
            time.sleep(1)  # 필터링 대기
            
            # ★ 필터 적용 후 그리드를 맨 위로 스크롤
            try:
                grid_body = self.main_driver.find_element(By.CSS_SELECTOR, ".ag-body-viewport")
                self.main_driver.execute_script("arguments[0].scrollTop = 0;", grid_body)
                time.sleep(0.3)
                self.gui.log(f"  ✓ 그리드 스크롤 초기화")
            except:
                pass
            
            self.gui.log(f"✅ 그룹 필터 '{actual_group_name}' 적용 완료")
            return True
            
        except Exception as e:
            self.gui.log(f"❌ 그룹 필터 실패: {e}")
            return False
    
    def get_current_page_info(self) -> str:
        """현재 페이지 정보 (예: 1-100 / 500)"""
        try:
            # 페이지 정보 텍스트 찾기
            info = self.main_driver.find_element(By.XPATH, "//*[contains(text(), '/')]")
            return info.text
        except:
            return ""
    
    def filter_by_tag(self, tag_name: str) -> bool:
        """태그 필터 - API 모드에서는 조회 시 필터 적용"""
        
        # ★ v3.0: API 모드에서는 current_tag_filter 설정
        if self.use_api_mode and self.api_client:
            # 태그명 그대로 저장 (API에서 groupFile + contains로 필터링)
            self.gui.current_tag_filter = tag_name
            self.gui.log(f"✅ 태그 필터 '{tag_name}' 설정됨 (API)")
            return True
        
        # 레거시 Selenium 모드 - AG Grid 필터
        try:
            self.gui.log(f"🔍 태그 필터: {tag_name}")
            
            # AG Grid 태그 필터 입력 필드 찾기
            search_input = None
            
            try:
                search_input = WebDriverWait(self.main_driver, 5).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "input.ag-input-field-input[aria-label*='태그']"))
                )
            except:
                pass
            
            if not search_input:
                try:
                    search_input = self.main_driver.find_element(By.CSS_SELECTOR, "input.ag-text-field-input[aria-label*='태그']")
                except:
                    pass
            
            if not search_input:
                self.gui.log("⚠️ 태그 필터 입력 필드를 찾을 수 없음")
                return False
            
            # React 입력 필드에 값 설정 (JavaScript)
            self.main_driver.execute_script("""
                var input = arguments[0];
                var value = arguments[1];
                
                var nativeInputValueSetter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, 'value').set;
                nativeInputValueSetter.call(input, value);
                
                var inputEvent = new Event('input', { bubbles: true });
                input.dispatchEvent(inputEvent);
                
                var changeEvent = new Event('change', { bubbles: true });
                input.dispatchEvent(changeEvent);
            """, search_input, tag_name)
            
            time.sleep(1)  # 필터 적용 대기
            self.gui.log(f"✅ 태그 필터 '{tag_name}' 적용 완료")
            return True
            
        except Exception as e:
            self.gui.log(f"❌ 태그 필터 실패: {e}")
            return False
    
    def clear_tag_filter(self) -> bool:
        """태그 필터 초기화"""
        try:
            search_input = None
            
            try:
                search_input = self.main_driver.find_element(By.CSS_SELECTOR, "input.ag-input-field-input[aria-label*='태그']")
            except:
                pass
            
            if not search_input:
                try:
                    search_input = self.main_driver.find_element(By.CSS_SELECTOR, "input.ag-text-field-input[aria-label*='태그']")
                except:
                    return True  # 못 찾아도 계속 진행
            
            # 입력 필드 비우기
            self.main_driver.execute_script("""
                var input = arguments[0];
                var nativeInputValueSetter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, 'value').set;
                nativeInputValueSetter.call(input, '');
                
                var inputEvent = new Event('input', { bubbles: true });
                input.dispatchEvent(inputEvent);
                
                var changeEvent = new Event('change', { bubbles: true });
                input.dispatchEvent(changeEvent);
            """, search_input)
            
            time.sleep(0.5)
            return True
            
        except Exception as e:
            self.gui.log(f"⚠️ 태그 필터 초기화 실패: {e}")
            return False
    
    def _is_last_page(self) -> bool:
        """마지막 페이지인지 확인 (다음 버튼 disabled 여부)"""
        try:
            # "다음" 버튼 찾기
            next_btn = self.main_driver.find_element(
                By.XPATH, "//button[text()='다음']"
            )
            # disabled 속성 확인
            is_disabled = next_btn.get_attribute("disabled")
            return is_disabled is not None
        except:
            # 버튼을 못찾으면 마지막 페이지로 간주
            return True
    
    def click_refresh_button(self) -> bool:
        """불사자 새로고침 버튼 클릭"""
        try:
            # lucide lucide-refresh-cw 클래스의 SVG가 있는 버튼 찾기
            refresh_btn = None
            
            # 방법 1: CSS 선택자
            try:
                refresh_btn = self.main_driver.find_element(
                    By.CSS_SELECTOR, "button svg.lucide-refresh-cw"
                )
                refresh_btn = refresh_btn.find_element(By.XPATH, "./ancestor::button")
            except:
                pass
            
            # 방법 2: XPath로 SVG 클래스 확인
            if not refresh_btn:
                try:
                    refresh_btn = self.main_driver.find_element(
                        By.XPATH, "//button[.//svg[contains(@class, 'lucide-refresh-cw')]]"
                    )
                except:
                    pass
            
            # 방법 3: JavaScript로 찾기
            if not refresh_btn:
                refresh_btn = self.main_driver.execute_script("""
                    var svgs = document.querySelectorAll('svg.lucide-refresh-cw');
                    for (var svg of svgs) {
                        var btn = svg.closest('button');
                        if (btn) return btn;
                    }
                    return null;
                """)
            
            if refresh_btn:
                self.main_driver.execute_script("arguments[0].click();", refresh_btn)
                time.sleep(1)
                return True
            else:
                self.gui.log("  ⚠️ 새로고침 버튼을 찾을 수 없음")
                return False
                
        except Exception as e:
            self.gui.log(f"⚠️ 새로고침 실패: {e}")
            return False
    
    def deselect_all_products(self) -> bool:
        """모든 상품 선택 해제"""
        try:
            self.gui.log("☐ 전체 선택 해제 중...")
            
            # "페이지 전체 해제" 버튼이 있으면 선택된 상태 → 클릭하여 해제
            deselect_btn = None
            try:
                deselect_btn = self.main_driver.find_element(By.XPATH, "//button[contains(text(), '페이지 전체 해제')]")
            except:
                pass
            
            if deselect_btn:
                self.main_driver.execute_script("arguments[0].click();", deselect_btn)
                time.sleep(0.3)
                self.gui.log("✅ 전체 선택 해제 완료")
                return True
            
            # "페이지 전체 선택" 버튼만 있으면 이미 해제된 상태
            select_btn = None
            try:
                select_btn = self.main_driver.find_element(By.XPATH, "//button[contains(text(), '페이지 전체 선택')]")
            except:
                pass
            
            if select_btn:
                self.gui.log("ℹ️ 이미 선택 해제 상태")
                return True
            
            self.gui.log("⚠️ 전체 선택/해제 버튼을 찾을 수 없음")
            return False
            
        except Exception as e:
            self.gui.log(f"⚠️ 선택 해제 실패: {e}")
            return False
    
    def select_all_in_page(self) -> bool:
        """페이지 전체 선택"""
        try:
            # "페이지 전체 선택" 버튼이 있으면 해제된 상태 → 클릭하여 선택
            select_btn = None
            try:
                select_btn = self.main_driver.find_element(By.XPATH, "//button[contains(text(), '페이지 전체 선택')]")
            except:
                pass
            
            if select_btn:
                self.main_driver.execute_script("arguments[0].click();", select_btn)
                time.sleep(0.3)
                return True
            
            # "페이지 전체 해제" 버튼만 있으면 이미 선택된 상태
            deselect_btn = None
            try:
                deselect_btn = self.main_driver.find_element(By.XPATH, "//button[contains(text(), '페이지 전체 해제')]")
            except:
                pass
            
            if deselect_btn:
                self.gui.log("ℹ️ 이미 전체 선택 상태")
                return True
            
            self.gui.log("⚠️ 페이지 전체 선택 버튼을 찾을 수 없음")
            return False
            
        except Exception as e:
            self.gui.log(f"⚠️ 전체 선택 실패: {e}")
            return False
    
    def wait_for_product_list_loaded(self, timeout: int = 15) -> bool:
        """상품 리스트가 로딩 완료될 때까지 대기 - API 모드에서는 불필요"""
        
        # ★ v3.0: API 모드에서는 대기 불필요
        if self.use_api_mode and self.api_client:
            return True
        
        # 레거시 Selenium 모드
        """상품 리스트가 로딩 완료될 때까지 대기
        
        조건:
        1. AG Grid 로딩 오버레이 사라짐
        2. 상품 행(row)이 나타나거나 "표시할 데이터 없음" 표시
        """
        try:
            start_time = time.time()
            
            while time.time() - start_time < timeout:
                # 1. 로딩 오버레이 확인
                loading_visible = self.main_driver.execute_script("""
                    var overlay = document.querySelector('.ag-overlay-loading-center');
                    return overlay && overlay.offsetParent !== null;
                """)
                
                if loading_visible:
                    time.sleep(0.2)
                    continue
                
                # 2. 상품 행 존재 확인
                row_count = self.main_driver.execute_script("""
                    return document.querySelectorAll('div[role="row"][row-index]').length;
                """)
                
                if row_count > 0:
                    # 추가로 첫 번째 행의 데이터가 실제로 렌더링되었는지 확인
                    has_content = self.main_driver.execute_script("""
                        var firstRow = document.querySelector('div[role="row"][row-index="0"]');
                        if (firstRow) {
                            var cells = firstRow.querySelectorAll('[role="gridcell"]');
                            return cells.length > 0;
                        }
                        return false;
                    """)
                    if has_content:
                        # ★ 리스트 렌더링 후 안정화 대기 (0.7초)
                        time.sleep(0.7)
                        return True
                
                # 3. "표시할 데이터 없음" 메시지 확인
                no_data = self.main_driver.execute_script("""
                    var noRowsOverlay = document.querySelector('.ag-overlay-no-rows-center');
                    return noRowsOverlay && noRowsOverlay.offsetParent !== null;
                """)
                
                if no_data:
                    self.gui.log("  ℹ️ 표시할 상품 없음")
                    time.sleep(0.5)
                    return True  # 데이터 없음도 로딩 완료로 간주
                
                time.sleep(0.2)
            
            self.gui.log(f"  ⚠️ 상품 리스트 로딩 타임아웃 ({timeout}초)")
            return False
            
        except Exception as e:
            self.gui.log(f"  ⚠️ 리스트 로딩 확인 오류: {e}")
            return False
    
    def _get_total_product_count(self) -> int:
        """총 상품 개수 확인 (1-100 / 1195 에서 1195 추출)"""
        try:
            # 페이지 정보 텍스트에서 총 개수 추출
            page_info = self.main_driver.execute_script("""
                var elements = document.querySelectorAll('*');
                for (var el of elements) {
                    var text = el.textContent.trim();
                    // "1-100 / 1195" 패턴 매칭
                    var match = text.match(/^\\d+-\\d+\\s*\\/\\s*(\\d+)$/);
                    if (match) {
                        return parseInt(match[1]);
                    }
                }
                return 0;
            """)
            return page_info if page_info else 0
        except:
            return 0
    
    def click_copy_button(self) -> bool:
        """상품복사 버튼 클릭"""
        try:
            self.gui.log("📋 상품복사 버튼 클릭 중...")
            
            # 상품복사 버튼 찾기 (에메랄드 색상)
            copy_btn = None
            
            try:
                copy_btn = self.main_driver.find_element(By.CSS_SELECTOR, "button.bg-emerald-500")
            except:
                pass
            
            if not copy_btn:
                try:
                    copy_btn = self.main_driver.find_element(By.XPATH, "//button[text()='상품복사']")
                except:
                    pass
            
            if not copy_btn:
                self.gui.log("⚠️ 상품복사 버튼을 찾을 수 없음")
                return False
            
            self.main_driver.execute_script("arguments[0].click();", copy_btn)
            
            # 다이얼로그가 열릴 때까지 대기 (최대 10초)
            try:
                WebDriverWait(self.main_driver, 10).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "[role='dialog']"))
                )
            except:
                self.gui.log("⚠️ 복사 다이얼로그가 열리지 않음")
                return False
            
            self.gui.log("✅ 상품복사 버튼 클릭 완료")
            return True
            
        except Exception as e:
            self.gui.log(f"❌ 상품복사 버튼 클릭 실패: {e}")
            return False
    
    def select_target_group_in_copy_dialog(self, group_name: str) -> bool:
        """복사 다이얼로그에서 대상 그룹 선택 (Radix UI)"""
        try:
            self.gui.log(f"📁 복사 대상 그룹 선택: {group_name}")
            
            # 다이얼로그가 열려있는지 확인
            dialog = None
            try:
                dialog = WebDriverWait(self.main_driver, 5).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "[role='dialog']"))
                )
            except:
                self.gui.log("⚠️ 복사 다이얼로그를 찾을 수 없음")
                return False
            
            # 다이얼로그 내부의 combobox 찾기
            dropdown = None
            try:
                dropdown = dialog.find_element(By.CSS_SELECTOR, "button[role='combobox']")
                before_text = dropdown.text.strip()
                self.gui.log(f"  📋 현재 선택: '{before_text}'")
            except Exception as e:
                self.gui.log(f"⚠️ 그룹 선택 드롭다운을 찾을 수 없음: {e}")
                return False
            
            # 드롭다운 클릭 시도 (여러 방법)
            # 1차: ActionChains 클릭
            try:
                actions = ActionChains(self.main_driver)
                actions.move_to_element(dropdown).click().perform()
                time.sleep(1)
            except:
                pass
            
            is_expanded = dropdown.get_attribute("aria-expanded")
            
            if is_expanded != "true":
                try:
                    dropdown.click()
                    time.sleep(0.5)
                    is_expanded = dropdown.get_attribute("aria-expanded")
                except:
                    pass
            
            if is_expanded != "true":
                try:
                    self.main_driver.execute_script("arguments[0].focus(); arguments[0].click();", dropdown)
                    time.sleep(0.5)
                    is_expanded = dropdown.get_attribute("aria-expanded")
                except:
                    pass
            
            if is_expanded != "true":
                try:
                    dropdown.send_keys(Keys.ENTER)
                    time.sleep(0.5)
                    is_expanded = dropdown.get_attribute("aria-expanded")
                except:
                    pass
            
            self.gui.log(f"  ✓ 드롭다운 열림: {is_expanded}")
            
            if is_expanded != "true":
                self.gui.log("⚠️ 드롭다운을 열 수 없음")
                return False
            
            time.sleep(0.5)  # 옵션 로딩 대기
            
            # 드롭다운이 열린 후 옵션 목록 확인
            options_after = self.main_driver.execute_script("""
                var items = document.querySelectorAll('[data-radix-collection-item]');
                var texts = [];
                for (var i = 0; i < items.length; i++) {
                    var t = items[i].textContent.trim();
                    if (t.match(/^\\d+_/)) {  // 숫자_이름 패턴만
                        texts.push(t);
                    }
                }
                return texts;
            """)
            self.gui.log(f"  📋 그룹 목록: {options_after[:10]}...")
            
            # 숫자로 입력된 경우 실제 그룹 이름으로 변환
            actual_group_name = group_name
            if group_name.isdigit():
                idx = int(group_name) - 1  # 1-based -> 0-based
                if 0 <= idx < len(options_after):
                    actual_group_name = options_after[idx]
                    self.gui.log(f"  🔄 숫자 {group_name} → '{actual_group_name}'")
                else:
                    self.gui.log(f"⚠️ 인덱스 {group_name} 범위 초과 (총 {len(options_after)}개)")
                    self.main_driver.find_element(By.TAG_NAME, "body").send_keys(Keys.ESCAPE)
                    return False
            
            # 그룹 찾기 및 Selenium으로 직접 클릭
            target_element = None
            
            # 방법 1: XPath로 정확히 텍스트 매칭
            try:
                target_element = self.main_driver.find_element(
                    By.XPATH, f"//*[@data-radix-collection-item][normalize-space(text())='{actual_group_name}']"
                )
            except:
                pass
            
            if not target_element:
                # 방법 2: contains로 부분 매칭
                try:
                    target_element = self.main_driver.find_element(
                        By.XPATH, f"//*[@data-radix-collection-item][contains(text(), '{actual_group_name}')]"
                    )
                except:
                    pass
            
            if not target_element:
                # 방법 3: 그룹명의 뒷부분으로 매칭 (11_썬이마켓 → 썬이마켓)
                group_parts = actual_group_name.split('_')
                if len(group_parts) > 1:
                    group_suffix = '_'.join(group_parts[1:])
                    try:
                        target_element = self.main_driver.find_element(
                            By.XPATH, f"//*[@data-radix-collection-item][contains(text(), '{group_suffix}')]"
                        )
                    except:
                        pass
            
            if not target_element:
                # 방법 4: JavaScript로 찾기
                result = self.main_driver.execute_script("""
                    var targetGroup = arguments[0];
                    var items = document.querySelectorAll('[data-radix-collection-item]');
                    for (var i = 0; i < items.length; i++) {
                        var text = items[i].textContent.trim();
                        if (text === targetGroup) {
                            return items[i];
                        }
                    }
                    // 부분 매칭
                    var groupParts = targetGroup.split('_');
                    var groupSuffix = groupParts.length > 1 ? groupParts.slice(1).join('_') : targetGroup;
                    for (var i = 0; i < items.length; i++) {
                        var text = items[i].textContent.trim();
                        if (text.includes(groupSuffix)) {
                            return items[i];
                        }
                    }
                    return null;
                """, group_name)
                target_element = result
            
            if not target_element:
                self.gui.log(f"⚠️ 그룹 '{group_name}'을 목록에서 찾을 수 없음")
                self.main_driver.find_element(By.TAG_NAME, "body").send_keys(Keys.ESCAPE)
                return False
            
            # 요소 클릭 (여러 방법 시도)
            clicked = False
            
            # 클릭 방법 1: ActionChains
            try:
                self.main_driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", target_element)
                time.sleep(0.3)
                actions = ActionChains(self.main_driver)
                actions.move_to_element(target_element).click().perform()
                clicked = True
                self.gui.log(f"  ✓ ActionChains 클릭")
            except Exception as e:
                self.gui.log(f"  ⚠️ ActionChains 실패: {e}")
            
            if not clicked:
                # 클릭 방법 2: 직접 클릭
                try:
                    target_element.click()
                    clicked = True
                    self.gui.log(f"  ✓ 직접 클릭")
                except:
                    pass
            
            if not clicked:
                # 클릭 방법 3: JavaScript click
                try:
                    self.main_driver.execute_script("arguments[0].click();", target_element)
                    clicked = True
                    self.gui.log(f"  ✓ JavaScript 클릭")
                except:
                    pass
            
            time.sleep(0.5)
            
            # 선택 검증: 선택된 그룹이 목표 그룹과 정확히 일치하는지 확인
            try:
                time.sleep(0.3)
                after_text = dropdown.text.strip()
                self.gui.log(f"  📋 선택 후: '{after_text}'")
                
                # 정확히 일치하는지 확인
                if after_text == actual_group_name:
                    self.gui.log(f"✅ 대상 그룹 '{actual_group_name}' 선택 완료")
                    return True
                
                # 부분 일치 확인 (11_썬이마켓 vs 썬이마켓)
                if actual_group_name in after_text or after_text in actual_group_name:
                    self.gui.log(f"✅ 대상 그룹 '{actual_group_name}' 선택 완료")
                    return True
                
                # 불일치 - 잘못된 그룹 선택됨
                self.gui.log(f"❌ 그룹 선택 불일치! 목표: '{actual_group_name}', 실제: '{after_text}'")
                return False
                
            except Exception as e:
                self.gui.log(f"⚠️ 선택 검증 실패: {e}")
                return False
            
            return clicked
            
        except Exception as e:
            self.gui.log(f"❌ 대상 그룹 선택 실패: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def confirm_copy(self) -> bool:
        """복사 확인 버튼 클릭"""
        try:
            self.gui.log("📋 복사 진행 중...")
            
            # 다이얼로그 내 복사 버튼 찾기 (bg-primary)
            confirm_btn = None
            
            try:
                confirm_btn = self.main_driver.find_element(By.CSS_SELECTOR, "div[role='dialog'] button.bg-primary")
            except:
                pass
            
            if not confirm_btn:
                try:
                    confirm_btn = self.main_driver.find_element(By.XPATH, "//div[@role='dialog']//button[text()='복사']")
                except:
                    pass
            
            if not confirm_btn:
                self.gui.log("⚠️ 복사 버튼을 찾을 수 없음")
                return False
            
            self.main_driver.execute_script("arguments[0].click();", confirm_btn)
            
            # 복사 처리 완료 대기 (다이얼로그 닫힘 또는 성공 토스트)
            try:
                # 다이얼로그가 닫힐 때까지 대기 (최대 15초)
                WebDriverWait(self.main_driver, 15).until(
                    EC.invisibility_of_element_located((By.CSS_SELECTOR, "[role='dialog']"))
                )
            except:
                # 타임아웃 시 다이얼로그 존재 여부 확인
                try:
                    dialog = self.main_driver.find_element(By.CSS_SELECTOR, "[role='dialog']")
                    if dialog.is_displayed():
                        self.gui.log("⚠️ 복사 다이얼로그가 아직 열려있음")
                        return False
                except:
                    pass  # 다이얼로그 없음 = 닫힘
            
            self.gui.log("✅ 복사 완료")
            return True
            
        except Exception as e:
            self.gui.log(f"❌ 복사 확인 실패: {e}")
            return False
    
    def close_copy_dialog(self) -> bool:
        """복사 다이얼로그 닫기"""
        try:
            close_btn = self.main_driver.find_element(By.XPATH, "//div[@role='dialog']//button[contains(., '취소')]")
            self.main_driver.execute_script("arguments[0].click();", close_btn)
            time.sleep(0.5)
            return True
        except:
            return False
    
    def copy_products_to_group(self, target_group: str, count: int, start_index: int = 0) -> Tuple[bool, int]:
        """상품을 지정된 그룹으로 복사 (특정 위치부터 시작)
        
        Returns: (성공여부, 다음 시작 인덱스)
        """
        try:
            self.gui.log(f"\n{'='*50}")
            self.gui.log(f"📋 복사 시작: {target_group} ({count}개, row {start_index}부터)")
            self.gui.log(f"{'='*50}")
            
            # 1. 전체 선택 해제
            self.deselect_all_products()
            time.sleep(0.3)
            
            # 2. 특정 위치로 스크롤 후 상품 선택
            selected = self.select_products_from_index(start_index, count)
            if selected == 0:
                self.gui.log("⚠️ 선택할 상품이 없습니다")
                return False, start_index
            
            # 3. 상품복사 버튼 클릭
            if not self.click_copy_button():
                return False, start_index
            
            # 4. 대상 그룹 선택
            if not self.select_target_group_in_copy_dialog(target_group):
                self.close_copy_dialog()
                return False, start_index
            
            # 5. 복사 확인
            if not self.confirm_copy():
                return False, start_index
            
            next_index = start_index + selected
            self.gui.log(f"✅ '{target_group}'으로 {selected}개 상품 복사 완료!")
            return True, next_index
            
        except Exception as e:
            self.gui.log(f"❌ 상품 복사 실패: {e}")
            return False, start_index
    
    def select_products_from_index(self, start_index: int, count: int) -> int:
        """특정 인덱스부터 상품 선택 (바로 점프)"""
        try:
            self.gui.log(f"☑️ row {start_index}~{start_index + count - 1} 선택 중...")
            
            # AG Grid 컨테이너 찾기
            grid_body = self.main_driver.find_element(By.CSS_SELECTOR, ".ag-body-viewport")
            
            # 행 높이 계산 (대략 40px, 실제 확인 필요)
            row_height = self.main_driver.execute_script("""
                var row = document.querySelector('div[role="row"][row-index]');
                return row ? row.offsetHeight : 40;
            """) or 40
            
            # 시작 위치로 바로 스크롤!
            target_scroll = start_index * row_height
            self.main_driver.execute_script(f"arguments[0].scrollTop = {target_scroll};", grid_body)
            time.sleep(0.5)
            self.gui.log(f"📍 row {start_index} 위치로 점프!")
            
            selected_count = 0
            target_indices = set(range(start_index, start_index + count))
            seen_indices = set()
            max_scroll_attempts = 50
            scroll_attempts = 0
            
            while selected_count < count and scroll_attempts < max_scroll_attempts:
                # 현재 보이는 행들 가져오기
                rows = self.main_driver.find_elements(By.CSS_SELECTOR, "div[role='row'][row-index]")
                
                for row in rows:
                    if selected_count >= count:
                        break
                    
                    try:
                        row_index = row.get_attribute("row-index")
                        if not row_index:
                            continue
                        
                        row_idx = int(row_index)
                        
                        if row_index in seen_indices:
                            continue
                        seen_indices.add(row_index)
                        
                        # 목표 범위에 있는 row만 선택
                        if row_idx not in target_indices:
                            continue
                        
                        # 체크박스 찾기
                        checkbox = row.find_element(By.CSS_SELECTOR, "input.ag-checkbox-input")
                        
                        # 체크박스 클릭
                        self.main_driver.execute_script("arguments[0].click();", checkbox)
                        time.sleep(0.03)
                        selected_count += 1
                        
                        if selected_count % 20 == 0:
                            self.gui.log(f"   ☑️ {selected_count}개 선택됨...")
                        
                    except:
                        continue
                
                if selected_count >= count:
                    break
                
                # 아래로 조금씩 스크롤
                self.main_driver.execute_script("arguments[0].scrollTop += 300;", grid_body)
                time.sleep(0.3)
                scroll_attempts += 1
            
            self.gui.log(f"✅ {selected_count}개 상품 선택 완료")
            return selected_count
            
        except Exception as e:
            self.gui.log(f"❌ 상품 선택 실패: {e}")
            return 0
    
    def get_total_rows_in_page(self) -> int:
        """현재 페이지의 총 상품 수 확인"""
        try:
            # AG Grid에서 총 row 수 확인
            total = self.main_driver.execute_script("""
                var grid = document.querySelector('.ag-body-viewport');
                if (!grid) return 0;
                
                // 방법 1: ag-row 개수 확인 (가상 스크롤이라 부정확할 수 있음)
                // 방법 2: 페이지네이션 정보에서 확인
                var pageInfo = document.querySelector('.ag-paging-panel');
                if (pageInfo) {
                    var text = pageInfo.textContent;
                    var match = text.match(/(\\d+)\\s*\\/\\s*(\\d+)/);
                    if (match) return parseInt(match[2]);
                }
                
                // 방법 3: row-model에서 확인
                var lastRow = document.querySelector('div[role="row"][row-index]:last-of-type');
                if (lastRow) {
                    // 마지막까지 스크롤해서 확인
                    grid.scrollTop = grid.scrollHeight;
                    return 1000;  // 기본값
                }
                return 1000;
            """)
            return total or 1000
        except:
            return 1000  # 기본값
    
    def process_copy_groups(self, copy_groups: List[Tuple[str, int]], search_tag: str = "", done_tag: str = "", work_group: str = ""):
        """복사 그룹 일괄 처리 - Mode2: 복사 후 태그 변경
        
        1. 작업 그룹 선택 (지정된 경우)
        2. 태그 검색 (예: 작업완료_251202)
        3. 페이지 크기 = 복사 수량 설정
        4. 전체 선택 → 복사
        5. 선택 상품에 완료 태그 추가 (예: 복사완료)
        6. 다시 태그 검색 (완료 태그 있는 상품 제외)
        7. 반복
        """
        try:
            total_groups = len(copy_groups)
            success_count = 0
            fail_count = 0
            total_copied = 0
            failed_groups = []  # 실패한 그룹 이름 추적
            
            copy_count = copy_groups[0][1] if copy_groups else 100
            
            self.gui.log(f"\n{'#'*60}")
            self.gui.log(f"📋 Mode2: 복사 후 태그 변경")
            if work_group:
                self.gui.log(f"📁 작업 그룹: {work_group}")
            self.gui.log(f"📁 복사 그룹: {total_groups}개")
            self.gui.log(f"📦 그룹당 복사 수량: {copy_count}개")
            if search_tag:
                self.gui.log(f"🏷️ 검색 태그: {search_tag}")
            if done_tag:
                self.gui.log(f"🏷️ 완료 태그: {done_tag}")
            self.gui.log(f"{'#'*60}")
            
            # 작업 그룹이 지정된 경우 해당 그룹으로 필터링
            if work_group:
                self.gui.log(f"\n📁 작업 그룹 '{work_group}' 선택 중...")
                if not self.select_market_group(work_group):
                    self.gui.log(f"❌ 작업 그룹 선택 실패")
                    return
                # 상품 리스트 로딩 완료 대기
                self.gui.log(f"  ⏳ 상품 리스트 로딩 대기...")
                self.wait_for_product_list_loaded()
            
            for idx, (group_name, count) in enumerate(copy_groups, 1):
                if not self.is_running:
                    self.gui.log("🛑 중지됨")
                    break
                
                self.gui.log(f"\n[{idx}/{total_groups}] {group_name}")
                
                # 0. 전체 선택 해제 (처음 시작시)
                if idx == 1:
                    self.deselect_all_products()
                    time.sleep(0.2)
                
                # 1. 태그 검색 (매번 새로 - 이전 복사 상품은 done_tag가 붙어서 제외됨)
                if search_tag:
                    self.gui.log(f"  🔍 태그 검색: {search_tag}")
                    if not self.search_by_tag(search_tag):
                        self.gui.log("  ❌ 태그 검색 실패")
                        fail_count += 1
                        failed_groups.append((group_name, "태그 검색 실패"))
                        continue
                    # 상품 리스트 로딩 완료 대기
                    self.gui.log(f"  ⏳ 검색 결과 로딩 대기...")
                    if not self.wait_for_product_list_loaded():
                        self.gui.log("  ❌ 검색 결과 로딩 실패")
                        fail_count += 1
                        failed_groups.append((group_name, "검색 결과 로딩 실패"))
                        continue
                
                # 2. 페이지 크기 설정
                self.gui.log(f"  📄 페이지 크기 {count} 설정")
                if not self.set_page_size(count):
                    self.gui.log("  ❌ 페이지 크기 설정 실패")
                    fail_count += 1
                    failed_groups.append((group_name, "페이지 크기 설정 실패"))
                    continue
                # 상품 리스트 로딩 완료 대기
                self.gui.log(f"  ⏳ 페이지 로딩 대기...")
                if not self.wait_for_product_list_loaded():
                    self.gui.log("  ❌ 페이지 로딩 실패")
                    fail_count += 1
                    failed_groups.append((group_name, "페이지 로딩 실패"))
                    continue
                
                # ★ 복사 전 총 상품 수 기록
                before_total = self._get_total_product_count()
                if before_total > 0:
                    self.gui.log(f"  📊 현재 상품 수: {before_total}개")
                
                # 3. 전체 선택 전 해제 상태 확인
                self.deselect_all_products()
                time.sleep(0.3)
                
                # 4. 전체 선택
                self.gui.log(f"  ☑️ 전체 선택 (목표: {count}개)")
                self.select_all_in_page()
                time.sleep(0.5)  # 선택 반영 대기
                
                # "선택 N개 상품" 레이블에서 실제 선택 개수 확인
                selected = self._get_selected_count_from_label()
                self.gui.log(f"  📍 실제 선택: {selected}개")
                
                # 선택 수량이 목표와 다르면 재시도
                if selected != count:
                    time.sleep(0.5)
                    selected = self._get_selected_count_from_label()
                    self.gui.log(f"  📍 재확인: {selected}개")
                
                if selected == 0:
                    self.gui.log(f"  ❌ 선택된 상품 없음 - 스킵")
                    fail_count += 1
                    failed_groups.append((group_name, "선택된 상품 없음"))
                    continue
                
                if selected < count:
                    # "다음" 버튼이 disabled면 마지막 페이지 → 진행
                    is_last_page = self._is_last_page()
                    if is_last_page:
                        self.gui.log(f"  ⚠️ 마지막 페이지: {selected}개로 진행 (목표: {count}개)")
                    else:
                        self.gui.log(f"  ❌ 선택 개수 불일치! (목표: {count}개, 실제: {selected}개) - 스킵")
                        fail_count += 1
                        failed_groups.append((group_name, f"선택 개수 불일치 (목표:{count}, 실제:{selected})"))
                        continue
                else:
                    self.gui.log(f"  ✅ {selected}개 선택 확인 완료")
                
                # 4. 복사
                if not self.click_copy_button():
                    fail_count += 1
                    failed_groups.append((group_name, "복사 버튼 클릭 실패"))
                    continue
                
                if not self.select_target_group_in_copy_dialog(group_name):
                    self.close_copy_dialog()
                    fail_count += 1
                    failed_groups.append((group_name, "대상 그룹 선택 실패"))
                    continue
                
                copy_success = self.confirm_copy()
                if not copy_success:
                    fail_count += 1
                    failed_groups.append((group_name, "복사 확인 실패"))
                    continue
                
                self.gui.log(f"  ✅ 복사 완료")
                
                # 복사 다이얼로그 닫히고 상품 리스트 나타날 때까지 대기
                self.gui.log(f"  ⏳ 다이얼로그 닫힘 대기...")
                try:
                    WebDriverWait(self.main_driver, 10).until(
                        EC.invisibility_of_element_located((By.CSS_SELECTOR, "[role='dialog']"))
                    )
                except:
                    pass
                
                # 상품 리스트 로딩 완료 대기
                self.gui.log(f"  ⏳ 상품 리스트 로딩 대기...")
                if not self.wait_for_product_list_loaded():
                    self.gui.log(f"  ⚠️ 리스트 로딩 타임아웃 - 계속 진행")
                
                # 5. 복사 후 선택 수량 재확인 (선택 유지되어야 함)
                selected_after_copy = self._get_selected_count_from_label()
                self.gui.log(f"  📍 복사 후 선택: {selected_after_copy}개")
                
                # ★ 복사 성공 & 선택 유지된 경우에만 태그 변경
                if selected_after_copy == 0:
                    self.gui.log(f"  ⚠️ 복사 후 선택 해제됨 - 태그 변경 스킵")
                    # 복사는 됐지만 선택이 해제되어 태그 변경 불가
                elif done_tag and copy_success:
                    # 6. 태그 변경
                    self.gui.log(f"  🏷️ 태그 '{done_tag}'로 변경 중...")
                    
                    if self.add_tag_to_selected(done_tag):
                        self.gui.log(f"  ✅ 태그 변경 완료")
                    else:
                        self.gui.log(f"  ⚠️ 태그 변경 실패")
                
                # 7. 새로고침 및 상품 수 감소 확인
                if done_tag:
                    expected_count = before_total - selected if before_total > 0 else 0
                    
                    for refresh_attempt in range(3):  # 최대 3회 시도
                        self.gui.log(f"  🔄 새로고침... (시도 {refresh_attempt + 1}/3)")
                        self.click_refresh_button()
                        
                        # 상품 리스트 로딩 완료 대기
                        self.gui.log(f"  ⏳ 새로고침 후 로딩 대기...")
                        if not self.wait_for_product_list_loaded():
                            self.gui.log(f"  ⚠️ 새로고침 후 로딩 타임아웃")
                            continue
                        
                        # 상품 수 감소 확인
                        after_total = self._get_total_product_count()
                        self.gui.log(f"  📊 새로고침 후 상품 수: {after_total}개")
                        
                        if before_total > 0 and after_total > 0:
                            if after_total <= expected_count:
                                self.gui.log(f"  ✅ 상품 감소 확인 ({before_total} → {after_total})")
                                break
                            else:
                                self.gui.log(f"  ⚠️ 상품 미감소 ({before_total} → {after_total}, 예상: {expected_count})")
                                time.sleep(1)  # 잠시 대기 후 재시도
                        else:
                            break  # 개수 확인 불가시 그냥 진행
                    
                    # 전체 선택 해제 (잔상 제거)
                    self.deselect_all_products()
                    time.sleep(0.3)
                
                success_count += 1
                total_copied += selected
                self.gui.log(f"📊 진행: {idx}/{total_groups}, 총 {total_copied}개 복사됨")
            
            self.gui.log(f"\n{'#'*60}")
            self.gui.log(f"📋 복사 완료!")
            self.gui.log(f"✅ 성공: {success_count}개 그룹")
            self.gui.log(f"❌ 실패: {fail_count}개 그룹")
            self.gui.log(f"📦 총 복사: {total_copied}개")
            
            # 실패한 그룹 목록 표시
            if failed_groups:
                self.gui.log(f"\n{'='*40}")
                self.gui.log(f"❌ 실패한 그룹 목록:")
                for group_name, reason in failed_groups:
                    self.gui.log(f"  • {group_name}: {reason}")
            self.gui.log(f"{'#'*60}")
            
        except Exception as e:
            self.gui.log(f"❌ 복사 오류: {e}")
        finally:
            self.is_running = False
            self.gui.after(0, self.gui.on_copy_finished)
    
    def search_by_tag(self, tag_value: str) -> bool:
        """태그 관리 컬럼 필터에 태그 입력"""
        try:
            self.gui.log(f"  🔍 태그 검색: {tag_value}")
            
            # aria-label로 정확히 찾기
            tag_input = self.main_driver.find_element(
                By.CSS_SELECTOR, "input[aria-label='태그 관리 필터 입력']"
            )
            
            tag_input.click()
            time.sleep(0.2)
            tag_input.clear()
            time.sleep(0.2)
            tag_input.send_keys(tag_value)
            time.sleep(0.3)
            tag_input.send_keys(Keys.ENTER)
            
            self.gui.log(f"  ✅ 태그 '{tag_value}' 검색 완료")
            time.sleep(1)
            return True
            
        except Exception as e:
            self.gui.log(f"  ❌ 태그 검색 실패: {e}")
            return False
    
    def _get_selected_count_from_label(self) -> int:
        """'선택 N개 상품' 레이블에서 선택 개수 파싱"""
        try:
            # "선택 N개 상품" 레이블 찾기
            label = self.main_driver.find_element(
                By.XPATH, "//label[contains(text(), '선택') and contains(text(), '상품')]"
            )
            text = label.text  # "선택 1개 상품" 형태
            
            # 숫자 추출
            import re
            match = re.search(r'선택\s*(\d+)\s*개', text)
            if match:
                return int(match.group(1))
            return 0
        except:
            return 0
    
    def add_tag_to_selected(self, tag_name: str) -> bool:
        """선택된 상품에 태그 변경 (상품복사와 동일한 드롭다운 방식)"""
        max_retries = 2
        
        for attempt in range(max_retries):
            try:
                # 이전 다이얼로그가 완전히 닫힐 때까지 대기
                time.sleep(0.5)
                
                # 열린 다이얼로그가 있으면 닫기
                try:
                    existing_dialogs = self.main_driver.find_elements(By.CSS_SELECTOR, "[role='dialog']")
                    for d in existing_dialogs:
                        try:
                            self.main_driver.find_element(By.TAG_NAME, "body").send_keys(Keys.ESCAPE)
                            time.sleep(0.3)
                        except:
                            pass
                except:
                    pass
                
                # 1. "태그 변경" 버튼 클릭
                tag_btn = WebDriverWait(self.main_driver, 5).until(
                    EC.element_to_be_clickable((By.XPATH, "//button[contains(@class, 'bg-violet') and contains(text(), '태그 변경')]"))
                )
                self.main_driver.execute_script("arguments[0].click();", tag_btn)
                time.sleep(0.8)
                self.gui.log("  📋 태그 변경 다이얼로그 열림")
                
                # 2. 다이얼로그 내부의 combobox 찾기 (새로 찾기)
                dialog = WebDriverWait(self.main_driver, 5).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "[role='dialog']"))
                )
                
                # dropdown을 다이얼로그 내부에서 새로 찾기
                dropdown = WebDriverWait(dialog, 3).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "button[role='combobox']"))
                )
                before_text = dropdown.text.strip()
                self.gui.log(f"  📋 현재 선택: '{before_text}'")
                
                # 마켓그룹 선택 드롭다운인지 확인 (태그 드롭다운이어야 함)
                if before_text and ('_' in before_text and before_text[0].isdigit()):
                    # 이건 복사 다이얼로그가 아직 열려있음 - 닫고 재시도
                    self.gui.log(f"  ⚠️ 잘못된 다이얼로그 감지 - 재시도")
                    self.main_driver.find_element(By.TAG_NAME, "body").send_keys(Keys.ESCAPE)
                    time.sleep(0.5)
                    continue
                
                # 3. 드롭다운 클릭 (여러 방법)
                # 1차: ActionChains
                try:
                    actions = ActionChains(self.main_driver)
                    actions.move_to_element(dropdown).click().perform()
                    time.sleep(0.5)
                except:
                    pass
                
                is_expanded = dropdown.get_attribute("aria-expanded")
                
                if is_expanded != "true":
                    try:
                        dropdown.click()
                        time.sleep(0.5)
                        is_expanded = dropdown.get_attribute("aria-expanded")
                    except:
                        pass
                
                if is_expanded != "true":
                    try:
                        self.main_driver.execute_script("arguments[0].focus(); arguments[0].click();", dropdown)
                        time.sleep(0.5)
                        is_expanded = dropdown.get_attribute("aria-expanded")
                    except:
                        pass
                
                self.gui.log(f"  ✓ 드롭다운 열림: {is_expanded}")
                
                if is_expanded != "true":
                    if attempt < max_retries - 1:
                        self.gui.log("  ⚠️ 드롭다운 열기 실패 - 재시도")
                        self.main_driver.find_element(By.TAG_NAME, "body").send_keys(Keys.ESCAPE)
                        time.sleep(0.5)
                        continue
                    self.gui.log("  ⚠️ 드롭다운을 열 수 없음")
                    return False
                
                time.sleep(0.5)
                
                # 4. 옵션 목록에서 태그 찾기 (data-radix-collection-item)
                target_element = None
                
                # 방법 1: XPath로 정확히 텍스트 매칭
                try:
                    target_element = self.main_driver.find_element(
                        By.XPATH, f"//*[@data-radix-collection-item][normalize-space(text())='{tag_name}']"
                    )
                except:
                    pass
                
                if not target_element:
                    # 방법 2: JavaScript로 정확히 일치하는 것만 찾기
                    result = self.main_driver.execute_script("""
                        var targetTag = arguments[0];
                        var items = document.querySelectorAll('[data-radix-collection-item]');
                        for (var i = 0; i < items.length; i++) {
                            var text = items[i].textContent.trim();
                            if (text === targetTag) {  // 정확히 일치
                                return items[i];
                            }
                        }
                        return null;
                    """, tag_name)
                    target_element = result
                
                if not target_element:
                    self.gui.log(f"  ⚠️ 태그 '{tag_name}'을 목록에서 찾을 수 없음")
                    self.main_driver.find_element(By.TAG_NAME, "body").send_keys(Keys.ESCAPE)
                    return False
                
                # 5. 옵션 클릭 (여러 방법)
                clicked = False
                
                try:
                    self.main_driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", target_element)
                    time.sleep(0.3)
                    actions = ActionChains(self.main_driver)
                    actions.move_to_element(target_element).click().perform()
                    clicked = True
                    self.gui.log(f"  ✓ ActionChains 클릭")
                except:
                    pass
                
                if not clicked:
                    try:
                        self.main_driver.execute_script("arguments[0].click();", target_element)
                        clicked = True
                        self.gui.log(f"  ✓ JavaScript 클릭")
                    except:
                        pass
                
                time.sleep(0.5)
                
                # 선택 검증 (dropdown을 다시 찾기 - stale 방지)
                try:
                    dropdown = dialog.find_element(By.CSS_SELECTOR, "button[role='combobox']")
                    after_text = dropdown.text.strip()
                    self.gui.log(f"  📋 선택 후: '{after_text}'")
                    
                    if after_text == tag_name:
                        self.gui.log(f"  ✅ 태그 '{tag_name}' 선택 완료")
                    else:
                        self.gui.log(f"  ⚠️ 선택 불일치! 기대: '{tag_name}', 실제: '{after_text}'")
                except:
                    pass
                
                # 6. 적용 버튼 클릭
                apply_btn = None
                try:
                    apply_btn = dialog.find_element(By.XPATH, ".//button[text()='적용']")
                except:
                    pass
                
                if not apply_btn:
                    try:
                        apply_btn = dialog.find_element(By.CSS_SELECTOR, "button.bg-primary")
                    except:
                        pass
                
                if not apply_btn:
                    self.gui.log(f"  ⚠️ 적용 버튼을 찾을 수 없음")
                    self.main_driver.find_element(By.TAG_NAME, "body").send_keys(Keys.ESCAPE)
                    return False
                
                self.gui.log(f"  📋 적용 버튼 클릭...")
                self.main_driver.execute_script("arguments[0].click();", apply_btn)
                
                # 7. 다이얼로그 닫힘 대기 (적용 완료 확인)
                try:
                    WebDriverWait(self.main_driver, 10).until(
                        EC.invisibility_of_element_located((By.CSS_SELECTOR, "[role='dialog']"))
                    )
                    self.gui.log(f"  ✅ 태그 적용 완료 (다이얼로그 닫힘)")
                except:
                    # 다이얼로그가 아직 열려있는지 확인
                    try:
                        still_open = self.main_driver.find_element(By.CSS_SELECTOR, "[role='dialog']")
                        if still_open.is_displayed():
                            self.gui.log(f"  ⚠️ 다이얼로그가 아직 열려있음 - 재시도")
                            self.main_driver.find_element(By.TAG_NAME, "body").send_keys(Keys.ESCAPE)
                            time.sleep(0.5)
                            continue
                    except:
                        pass  # 다이얼로그 없음 = 닫힘
                
                # 8. 성공 토스트 확인 (선택사항)
                time.sleep(0.5)
                
                return True
                
            except Exception as e:
                self.gui.log(f"  ❌ 태그 변경 오류: {e}")
                if attempt < max_retries - 1:
                    self.gui.log(f"  🔄 재시도 {attempt + 2}/{max_retries}")
                    try:
                        self.main_driver.find_element(By.TAG_NAME, "body").send_keys(Keys.ESCAPE)
                    except:
                        pass
                    time.sleep(0.5)
                    continue
                try:
                    self.main_driver.find_element(By.TAG_NAME, "body").send_keys(Keys.ESCAPE)
                except:
                    pass
                return False
        
        return False
    
    def remove_tag_from_selected(self, tag_name: str) -> bool:
        """선택된 상품에서 태그 제거"""
        try:
            time.sleep(0.3)
            
            # 1. "태그 변경" 버튼 클릭
            tag_btn = WebDriverWait(self.main_driver, 5).until(
                EC.element_to_be_clickable((By.XPATH, "//button[contains(@class, 'bg-violet') and contains(text(), '태그 변경')]"))
            )
            self.main_driver.execute_script("arguments[0].click();", tag_btn)
            time.sleep(0.8)
            
            # 2. 다이얼로그 내부의 combobox 찾기
            dialog = WebDriverWait(self.main_driver, 5).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "[role='dialog']"))
            )
            
            dropdown = WebDriverWait(dialog, 3).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "button[role='combobox']"))
            )
            current_tag = dropdown.text.strip()
            self.gui.log(f"  📋 현재 태그: '{current_tag}'")
            
            # 현재 선택된 태그가 제거할 태그와 같으면 "태그 없음"으로 변경
            if current_tag == tag_name:
                # 드롭다운 클릭
                try:
                    actions = ActionChains(self.main_driver)
                    actions.move_to_element(dropdown).click().perform()
                    time.sleep(0.5)
                except:
                    dropdown.click()
                    time.sleep(0.5)
                
                is_expanded = dropdown.get_attribute("aria-expanded")
                if is_expanded != "true":
                    self.main_driver.execute_script("arguments[0].focus(); arguments[0].click();", dropdown)
                    time.sleep(0.5)
                
                # "태그 없음" 옵션 선택
                try:
                    no_tag_option = self.main_driver.find_element(
                        By.XPATH, "//*[@data-radix-collection-item][contains(text(), '태그 없음')]"
                    )
                    actions = ActionChains(self.main_driver)
                    actions.move_to_element(no_tag_option).click().perform()
                    time.sleep(0.5)
                    self.gui.log(f"  ✓ '태그 없음' 선택")
                except:
                    # JavaScript로 찾기
                    result = self.main_driver.execute_script("""
                        var items = document.querySelectorAll('[data-radix-collection-item]');
                        for (var i = 0; i < items.length; i++) {
                            if (items[i].textContent.trim() === '태그 없음') {
                                items[i].click();
                                return true;
                            }
                        }
                        return false;
                    """)
                    if result:
                        self.gui.log(f"  ✓ '태그 없음' 선택 (JS)")
                    time.sleep(0.5)
                
                # 적용 버튼 클릭
                apply_btn = dialog.find_element(By.XPATH, ".//button[text()='적용']")
                self.main_driver.execute_script("arguments[0].click();", apply_btn)
                time.sleep(1)
                
                self.gui.log(f"  ✅ 태그 '{tag_name}' 제거 완료")
                return True
            else:
                # 제거할 태그가 선택되어 있지 않음 - 다이얼로그 닫기
                self.main_driver.find_element(By.TAG_NAME, "body").send_keys(Keys.ESCAPE)
                self.gui.log(f"  ℹ️ 태그 '{tag_name}'가 선택되어 있지 않음")
                return True
                
        except Exception as e:
            self.gui.log(f"  ⚠️ 태그 제거 오류: {e}")
            try:
                self.main_driver.find_element(By.TAG_NAME, "body").send_keys(Keys.ESCAPE)
            except:
                pass
            return False
    
    def process_products_from_excel(self, start_index: int, count: int, num_pages: int = 1, groups: List[str] = None):
        """엑셀 기반 상품명 적용 (그룹2,3용 - AI 사용 안 함)"""
        try:
            # 1. 엑셀 파일 확인
            excel_path = self.gui.excel_file_path.get().strip()
            if not excel_path or not os.path.exists(excel_path):
                self.gui.log("❌ 엑셀 파일을 선택하세요!")
                self.is_running = False
                return
            
            # 2. 엑셀 로드
            self.gui.log(f"📊 엑셀 로드 중: {os.path.basename(excel_path)}")
            try:
                import pandas as pd
                df = pd.read_excel(excel_path)
                self.gui.log(f"✅ 엑셀 로드 완료: {len(df)}개 행")
            except Exception as e:
                self.gui.log(f"❌ 엑셀 로드 실패: {e}")
                self.is_running = False
                return
            
            # 3. 적용할 상품명 번호
            title_num = self.gui.excel_title_number.get()
            col_name = f'상품명{title_num}'
            
            if col_name not in df.columns:
                self.gui.log(f"❌ 엑셀에 '{col_name}' 컬럼이 없습니다!")
                self.is_running = False
                return
            
            self.gui.log(f"📝 적용 대상: {col_name}")
            
            # 4. 태그 확인
            tag_name = self.gui.tag_var.get().strip()
            if tag_name:
                self.gui.log(f"🏷️ 태그 확인 중: {tag_name}")
                if not self.check_tag_exists(tag_name):
                    self.gui.log(f"❌ 오류: 태그 '{tag_name}'이(가) 존재하지 않습니다!")
                    self.gui.log("💡 먼저 '📌 태그 생성' 버튼으로 태그를 생성하세요.")
                    self.is_running = False
                    return
                self.gui.log(f"✅ 태그 '{tag_name}' 확인됨")
            else:
                self.gui.log("ℹ️ 태그 미설정 - 태그 변경 없이 진행")
                tag_name = None
            
            # 5. 그룹 처리
            if not groups:
                groups = [None]
            
            total_processed = 0
            total_failed = 0
            total_not_found = 0
            
            self.gui.reset_progress()
            
            for group_idx, group_name in enumerate(groups):
                if not self.is_running:
                    break
                
                # 그룹 선택
                if group_name:
                    self.gui.log(f"\n{'#'*60}")
                    self.gui.log(f"📁 그룹 {group_idx + 1}/{len(groups)}: {group_name}")
                    self.gui.log(f"{'#'*60}")
                    
                    if not self.select_market_group(group_name):
                        self.gui.log(f"⚠️ 그룹 '{group_name}' 선택 실패, 건너뜀")
                        continue
                    
                    time.sleep(1)
                
                # 페이지 크기 설정
                self.set_page_size(1000)
                
                # 6. 상품 로드 (API)
                if not self.api_client:
                    self.gui.log("❌ API 클라이언트 없음!")
                    continue
                
                self.gui.log("📦 상품 로드 중...")
                products = self._get_products_via_api(0, count)
                
                if not products:
                    self.gui.log("⚠️ 상품이 없습니다")
                    continue
                
                self.gui.log(f"📦 {len(products)}개 상품 로드 완료")
                
                # 7. 각 상품 처리
                for idx, product in enumerate(products, 1):
                    if not self.is_running:
                        break
                    
                    self.gui.log(f"\n[{idx}/{len(products)}] {product.original_title[:40]}...")
                    
                    # 불사자 코드로 엑셀 검색
                    bulsaja_code = product.bulsaja_id
                    
                    if not bulsaja_code:
                        self.gui.log(f"  ⚠️ 불사자 코드 없음 - 스킵")
                        total_failed += 1
                        continue
                    
                    # 엑셀에서 찾기
                    matched = df[df['불사자 코드'] == bulsaja_code]
                    
                    if matched.empty:
                        self.gui.log(f"  ⚠️ 엑셀에 없음: {bulsaja_code[:15]}... - 스킵")
                        total_not_found += 1
                        continue
                    
                    # 상품명 가져오기
                    new_title = matched.iloc[0][col_name]
                    
                    if pd.isna(new_title) or not str(new_title).strip():
                        self.gui.log(f"  ⚠️ 상품명 없음 - 스킵")
                        total_failed += 1
                        continue
                    
                    new_title = str(new_title).strip()
                    self.gui.log(f"  ✨ 적용: {new_title}")
                    
                    # 8. 상품명 적용
                    try:
                        self.api_client.update_single_product(bulsaja_code, new_title)
                        self.gui.log(f"  ✅ 적용 완료")
                        
                        # 태그 적용
                        if tag_name:
                            self.api_client.apply_tag_single(bulsaja_code, tag_name)
                            self.gui.log(f"  🏷️ 태그 적용: {tag_name}")
                        
                        total_processed += 1
                        
                    except Exception as e:
                        self.gui.log(f"  ❌ 적용 실패: {str(e)[:50]}")
                        total_failed += 1
                    
                    # 진행률 업데이트
                    progress = int((idx / len(products)) * 100)
                    self.gui.update_progress(progress, len(products), total_processed)
                    
                    time.sleep(0.3)  # API 과부하 방지
            
            # 9. 완료 메시지
            self.gui.log(f"\n{'='*60}")
            self.gui.log(f"✅ 전체 완료!")
            self.gui.log(f"   성공: {total_processed}")
            self.gui.log(f"   실패: {total_failed}")
            self.gui.log(f"   엑셀에 없음: {total_not_found}")
            self.gui.log(f"{'='*60}")
            
        except Exception as e:
            self.gui.log(f"❌ 엑셀 모드 오류: {e}")
            import traceback
            self.gui.log(traceback.format_exc())
        
        finally:
            self.is_running = False
    
    def process_products(self, start_index: int, count: int, num_pages: int = 1, groups: List[str] = None):
        """상품 일괄 처리 (그룹별, 배치 단위)"""
        
        # ★ v3.1: 엑셀 모드 체크
        if self.gui.excel_mode_enabled.get():
            self.process_products_from_excel(start_index, count, num_pages, groups)
            return
        
        # API 모드면 한 번에 가져오고, Selenium 모드면 100개씩
        BATCH_SIZE = count if (self.use_api_mode and self.api_client) else 100
        
        try:
            # ★ 원래 탭(상품 리스트)으로 돌아가기
            self.ensure_main_window()
            
            # 그룹 리스트가 없으면 현재 그룹만 처리
            if not groups:
                groups = [None]  # None = 현재 선택된 그룹 유지
            
            grand_total_processed = 0
            grand_total_failed = 0
            grand_total_skipped = 0
            confirmed_suspects = []  # 의심 단어 목록
            
            # 결과 데이터 저장용 리스트 (xlsx로 저장할 때 사용)
            result_data = []  # [(마켓그룹, 상품코드, 기존상품명, 변경상품명, 의심단어리스트), ...]
            
            # 태그 확인 (GUI에서 입력받은 태그) - 파일명에 사용하기 위해 먼저 가져옴
            tag_name = self.gui.tag_var.get().strip()
            
            # 결과 파일 이름 (result 폴더에 저장)
            # 형식: {그룹명}_{태그}.xlsx
            timestamp = datetime.now().strftime("%y%m%d_%H%M")
            result_dir = "result"
            if not os.path.exists(result_dir):
                os.makedirs(result_dir)
            
            # 그룹명 가져오기 (첫 번째 그룹, 없으면 "전체")
            group_name_for_file = groups[0] if (groups and groups[0]) else "전체"
            
            # 태그명 가져오기 (없으면 timestamp)
            tag_for_file = tag_name if tag_name else f"작업_{timestamp}"
            
            # 파일명 생성: 그룹명_태그.xlsx
            result_filename = os.path.join(result_dir, f"{group_name_for_file}_{tag_for_file}.xlsx")
            self.gui.log(f"📄 결과 파일: {result_filename}")
            
            # 태그 확인 (태그가 있을 경우에만 검증)
            if tag_name:
                self.gui.log(f"🏷️ 태그 확인 중: {tag_name}")
                if not self.check_tag_exists(tag_name):
                    self.gui.log(f"❌ 오류: 태그 '{tag_name}'이(가) 존재하지 않습니다!")
                    self.gui.log("💡 먼저 '📌 태그 생성' 버튼으로 태그를 생성하세요.")
                    self.is_running = False
                    return
                self.gui.log(f"✅ 태그 '{tag_name}' 확인됨")
            else:
                self.gui.log("ℹ️ 태그 미설정 - 태그 변경 없이 진행")
                tag_name = None  # 태그 적용 건너뛰기
            
            # 이미 태그가 달린 상품 제외를 위해 "태그 없음" 필터 적용
            self.gui.log("🏷️ 태그 필터 적용: 태그 없음 (작업 완료 상품 제외)")
            self.filter_by_tag("태그 없음")
            time.sleep(1)
            
            # 상품명 생성 방식 확인
            title_mode = self.gui.title_mode_var.get()
            if title_mode == TITLE_MODE_IMAGE_FIRST:
                mode_name = "이미지+기존상품명"
            elif title_mode == TITLE_MODE_VISION:
                mode_name = "Vision분석"
            else:
                mode_name = "기존상품명만"
            
            # 진행 상황 초기화
            self.gui.reset_progress()
            
            for group_idx, group_name in enumerate(groups):
                if not self.is_running:
                    break
                
                # 그룹 선택 (None이 아니면)
                if group_name:
                    self.gui.log(f"\n{'#'*60}")
                    self.gui.log(f"📁 그룹 {group_idx + 1}/{len(groups)}: {group_name}")
                    self.gui.log(f"{'#'*60}")
                    
                    if not self.select_market_group(group_name):
                        self.gui.log(f"⚠️ 그룹 '{group_name}' 선택 실패, 건너뜀")
                        continue
                    
                    time.sleep(1)
                
                # 페이지 크기 1000으로 설정
                self.set_page_size(1000)
                
                group_processed = 0
                group_failed = 0
                group_skipped = 0
                
                # ★ API 모드: 실제 가용 상품 개수 (첫 배치에서 확인)
                actual_available_count = None
                
                for page_num in range(num_pages):
                    if not self.is_running:
                        break
                    
                    self.gui.log(f"\n{'='*50}")
                    if group_name:
                        self.gui.log(f"📄 [{group_name}] 페이지 {page_num + 1}/{num_pages}")
                    else:
                        self.gui.log(f"📄 페이지 {page_num + 1}/{num_pages}")
                    self.gui.log(f"{'='*50}")
                    
                    # 배치 단위로 처리
                    # v1.7.1: 페이지 2 이후에서도 실제 row-index 범위로 시작
                    page_start = start_index if page_num == 0 else 0
                    batch_start = page_start
                    actual_page_min = None  # 페이지의 실제 row-index 최소값 저장
                    
                    while group_processed < count and self.is_running:
                        remaining = count - group_processed
                        batch_count = min(BATCH_SIZE, remaining)
                        
                        self.gui.log(f"\n📦 배치 수집: {batch_start}번부터 {batch_count}개")
                        
                        # 배치만큼 상품 추출
                        products = self.get_visible_products(
                            start_index=batch_start, 
                            max_count=batch_count
                        )
                        
                        if not products:
                            self.gui.log("ℹ️ 현재 페이지 상품 소진")
                            break  # while 루프 탈출 → 다음 페이지로
                        
                        # ★ 중복 처리 방지: 반환된 개수가 요청보다 적으면 더 이상 상품 없음
                        if len(products) < batch_count:
                            self.gui.log(f"ℹ️ 남은 상품: {len(products)}개 (요청: {batch_count}개)")
                        
                        # ★ 첫 배치: 실제 가용 개수 저장
                        if actual_available_count is None:
                            actual_available_count = len(products)
                            self.gui.log(f"📊 실제 가용 상품: {actual_available_count}개 (요청: {count}개)")
                        
                        # v1.7.1: 실제 row-index 범위 로그
                        first_idx = products[0].index
                        last_idx = products[-1].index
                        self.gui.log(f"  📍 row-index 범위: {first_idx}~{last_idx}")
                        
                        # ★ 실제 처리할 총 개수 계산 (중복 방지)
                        actual_total = min(len(products), count - group_processed)
                        self.gui.log(f"🚀 {len(products)}개 상품 처리 시작 (생성방식: {mode_name})")
                        
                        # 배치 처리
                        batch_start_count = group_processed  # 배치 시작 시점 처리 개수 저장
                        for i, product in enumerate(products, 1):
                            if not self.is_running:
                                self.gui.log("🛑 중지됨")
                                break
                            
                            current_num = batch_start_count + i  # 올바른 순번 계산
                            display_total = actual_available_count if actual_available_count else count
                            self.gui.update_progress(current_num, display_total, group_processed, group_failed)
                            self.gui.update_progress_detail(f"처리 중: {product.original_title[:30]}...")
                            self.gui.log(f"\n[{current_num}/{display_total}] {product.original_title[:40]}...")
                            
                            try:
                                is_mismatch = False  # v11: 이미지 불일치 여부
                                
                                if title_mode == TITLE_MODE_IMAGE_FIRST:
                                    similar_titles = []
                                    
                                    if product.image_url:
                                        try:
                                            similar_titles = self.search_with_aliprice(product.image_url)
                                        except Exception as e:
                                            self.gui.log(f"⚠️ 이미지 검색 오류: {str(e)[:100]}")
                                    
                                    # ===== v11: 유사도 비교 + 이미지 검증 =====
                                    if similar_titles:
                                        # 유사도 계산
                                        max_similarity = 0.0
                                        for s_title in similar_titles[:5]:
                                            sim = calculate_similarity(product.original_title, s_title)
                                            if sim > max_similarity:
                                                max_similarity = sim
                                        
                                        # 핵심 키워드 중복률 계산
                                        keyword_overlap, common_keywords = check_keyword_overlap(
                                            product.original_title, similar_titles
                                        )
                                        
                                        self.gui.log(f"📊 유사도: {max_similarity:.0%}, 키워드 중복: {keyword_overlap:.0%}")
                                        if common_keywords:
                                            self.gui.log(f"   공통 키워드: {', '.join(list(common_keywords)[:5])}")
                                        
                                        # 유사도 30% 미만 AND 키워드 중복 30% 미만 → 다른 상품 가능성
                                        if max_similarity < SIMILARITY_THRESHOLD and keyword_overlap < 0.3:
                                            self.gui.log(f"⚠️ 다른 상품 가능성 높음 → 이미지 검증 진행")
                                            
                                            # 썸네일 3개 이상일 때만 이미지 검증
                                            if product.thumbnail_urls and len(product.thumbnail_urls) >= 3:
                                                is_match, confidence, reason = verify_images_with_claude(
                                                    self.claude_client,
                                                    product.thumbnail_urls,
                                                    self.gui.log
                                                )
                                                
                                                if not is_match:
                                                    is_mismatch = True
                                                    self.gui.log(f"🚨 이미지 불일치! ({confidence:.0%})")
                                                else:
                                                    self.gui.log(f"✅ 이미지 일치 ({confidence:.0%})")
                                            else:
                                                self.gui.log(f"   ℹ️ 썸네일 부족 ({len(product.thumbnail_urls or [])}개) → 검증 스킵")
                                    else:
                                        self.gui.log("ℹ️ 유사 상품 없음 - 원본 기반으로 생성")
                                    
                                    # ===== 불일치 처리: 1번 삭제 + Sonnet 분석 =====
                                    if is_mismatch and product.thumbnail_urls and len(product.thumbnail_urls) >= 2:
                                        self.gui.log("\n" + "="*40)
                                        self.gui.log("🔄 이미지 불일치 처리")
                                        self.gui.log("="*40)
                                        
                                        # row_element 찾기
                                        try:
                                            grid_body = self.main_driver.find_element(By.CSS_SELECTOR, ".ag-body-viewport")
                                            scroll_position = product.index * 126
                                            self.main_driver.execute_script(f"arguments[0].scrollTop = {scroll_position};", grid_body)
                                            time.sleep(0.3)
                                            
                                            row_element = self.main_driver.find_element(
                                                By.CSS_SELECTOR, f"div[role='row'][row-index='{product.index}']"
                                            )
                                            
                                            # [1단계] 1번 썸네일 삭제
                                            self.gui.log("\n[1단계] 1번 썸네일(스스 이미지) 삭제")
                                            if delete_thumbnail_at_position(self.main_driver, row_element, 1, self.gui.log):
                                                time.sleep(0.5)
                                                
                                                # [2단계] Sonnet으로 이미지 분석 + 상품명 생성
                                                self.gui.log("\n[2단계] Sonnet 이미지 분석 + 상품명 생성")
                                                target_length = int(self.gui.title_length_var.get())
                                                new_title, best_position, detected_brands = analyze_and_generate_title_sonnet(
                                                    self.claude_client,
                                                    product.thumbnail_urls[1:],  # 1번 삭제 후 나머지
                                                    product.original_title,
                                                    self.banned_words,
                                                    target_length,
                                                    self.gui.log
                                                )
                                                # 이미지 불일치 시 상품명 2,3은 동일하게 설정
                                                new_title2 = new_title
                                                new_title3 = new_title
                                                
                                                # [3단계] 최적 이미지 이동 (1번이 아닌 경우)
                                                if best_position > 1 and new_title:
                                                    self.gui.log(f"\n[3단계] {best_position}번 이미지 → 1번으로 이동")
                                                    row_element = self.main_driver.find_element(
                                                        By.CSS_SELECTOR, f"div[role='row'][row-index='{product.index}']"
                                                    )
                                                    move_thumbnail_to_front(self.main_driver, row_element, best_position, self.gui.log)
                                                    time.sleep(0.5)
                                                
                                                # [4단계] 1번 썸네일 누끼 적용
                                                self.gui.log("\n[4단계] 1번 썸네일 배경 제거")
                                                try:
                                                    row_element = self.main_driver.find_element(
                                                        By.CSS_SELECTOR, f"div[role='row'][row-index='{product.index}']"
                                                    )
                                                    remove_background_at_position(self.main_driver, row_element, 1, self.gui.log)
                                                except Exception as e:
                                                    self.gui.log(f"⚠️ 누끼 적용 실패: {e}")
                                                
                                                forbidden_found = False
                                            else:
                                                self.gui.log("⚠️ 1번 썸네일 삭제 실패 → 일반 처리")
                                                is_mismatch = False
                                                
                                        except Exception as e:
                                            self.gui.log(f"⚠️ 불일치 처리 실패: {e}")
                                            is_mismatch = False
                                    
                                    # ===== 일치 처리 (기존 로직) =====
                                    if not is_mismatch:
                                        new_title, new_title2, new_title3, detected_brands, forbidden_found = self.generate_title_with_claude(
                                            product.original_title, 
                                            similar_titles
                                        )
                                elif title_mode == TITLE_MODE_VISION:
                                    # Vision 분석 모드 (1회 API 호출로 바로 최종 상품명 생성)
                                    new_title = ""
                                    new_title2 = ""
                                    new_title3 = ""
                                    detected_brands = []
                                    forbidden_found = False
                                    
                                    if product.image_url:
                                        # GUI에서 선택한 Vision 모델 사용
                                        vision_model = self.gui.vision_model_var.get()
                                        # GUI에서 설정된 상품명 길이 가져오기
                                        target_length = int(self.gui.title_length_var.get())
                                        new_title, detected_brands, forbidden_found = generate_title_with_vision_api(
                                            self.claude_client,
                                            product.image_url,
                                            product.original_title,
                                            model=vision_model,
                                            banned_words=self.banned_words,
                                            log_callback=self.gui.log,
                                            target_length=target_length
                                        )
                                        new_title2 = new_title  # Vision은 1개만 생성
                                        new_title3 = new_title
                                    else:
                                        self.gui.log("⚠️ 이미지 없음 → 기존상품명만 사용")
                                    
                                    # Vision 실패 시 기존상품명 기반으로 생성
                                    if not new_title:
                                        self.gui.log("ℹ️ Vision 실패 - 원본 기반으로 생성")
                                        new_title, new_title2, new_title3, detected_brands, forbidden_found = self.generate_title_original_only(product.original_title)
                                else:
                                    self.gui.log("ℹ️ 기존상품명 기반으로 생성")
                                    new_title, new_title2, new_title3, detected_brands, forbidden_found = self.generate_title_original_only(product.original_title)
                                
                                # 사용자가 선택한 번호에 따라 적용할 상품명 결정
                                apply_choice = self.gui.title_apply_var.get()
                                if apply_choice == "2":
                                    apply_title = new_title2
                                elif apply_choice == "3":
                                    apply_title = new_title3
                                else:
                                    apply_title = new_title
                                
                                # ✅ 제거단어 적용 (상품명에서 무조건 삭제)
                                if REMOVE_WORDS:
                                    original_apply = apply_title
                                    apply_title = apply_remove_words(apply_title, REMOVE_WORDS)
                                    # new_title 시리즈도 제거 적용 (결과 저장용)
                                    new_title = apply_remove_words(new_title, REMOVE_WORDS)
                                    new_title2 = apply_remove_words(new_title2, REMOVE_WORDS) if new_title2 else ""
                                    new_title3 = apply_remove_words(new_title3, REMOVE_WORDS) if new_title3 else ""
                                    if original_apply != apply_title:
                                        self.gui.log(f"🗑️ 제거단어 적용: {original_apply} → {apply_title}")
                                
                                # ✅ 1단계: 패턴 기반 위험 감지
                                danger_check = detect_dangerous_product(apply_title)
                                is_dangerous = danger_check['is_dangerous']
                                danger_categories = danger_check['categories']
                                danger_words = danger_check['all_words']
                                
                                # ✅ 2단계: 위험 단어 발견 시 맥락 분석 (Claude)
                                safe_words = set()  # 안전 판정 받은 단어들
                                if is_dangerous and danger_words:
                                    self.gui.log(f"🔍 위험 단어 감지: {', '.join(danger_words[:3])} - 맥락 분석 중...")
                                    
                                    # Claude로 맥락 기반 검증
                                    context_check = self.verify_danger_with_context(apply_title, danger_words)
                                    is_dangerous = context_check['is_dangerous']
                                    
                                    if is_dangerous:
                                        self.gui.log_warning(f"🚨 위험 확정: {context_check['reason']} (확신도: {context_check['confidence']})")
                                    else:
                                        self.gui.log(f"✅ 안전 판정: {context_check['reason']} (확신도: {context_check['confidence']})")
                                        # 안전 판정 받은 단어들 기록
                                        safe_words.update(danger_words)
                                
                                # ✅ 3단계: 의심단어 분류 처리 (API 호출 없이 패턴 기반)
                                # ★ v2.5: 실제 상품명에 포함된 의심단어만 처리
                                # ★ v3.4: 예외단어는 의심단어에서 미리 제외
                                # Claude가 반환한 의심단어 중 생성된 상품명에 없는 것은 제외
                                actual_suspects = []
                                for word in detected_brands:
                                    word_clean = word.strip()
                                    if not word_clean:
                                        continue
                                    # 예외단어면 스킵 (의심단어로 분류 안 함)
                                    if word_clean in EXCLUDED_WORDS or word_clean.lower() in EXCLUDED_WORDS:
                                        continue
                                    # apply_title (실제 적용될 상품명)에 포함된 것만
                                    if word_clean in apply_title or word_clean.lower() in apply_title.lower():
                                        actual_suspects.append(word_clean)
                                
                                # 의심단어 검출 로그 (있을 경우)
                                if actual_suspects:
                                    self.gui.log_warning(f"⚠️ 의심단어 검출: {', '.join(actual_suspects)}")
                                
                                # 실제 사용된 의심단어만 분류 (제거단어/예외단어 시트 참조)
                                suspect_result = process_suspect_words(actual_suspects, self.remove_words, EXCLUDED_WORDS)
                                
                                # 유명 브랜드 발견 → 무조건 2차검수
                                famous_brands_found = suspect_result['review']
                                # 제거할 단어 (모델명, 셀러브랜드 등)
                                words_to_remove = suspect_result['remove']
                                # 사람이 판단해야 할 의심단어
                                ambiguous_words = suspect_result['suspect']
                                
                                # 제거 대상 단어들 상품명에서 제거
                                if words_to_remove:
                                    removed_words = [w[0] for w in words_to_remove]
                                    original_apply = apply_title
                                    for word, reason in words_to_remove:
                                        apply_title = apply_title.replace(word, "").strip()
                                        new_title = new_title.replace(word, "").strip() if new_title else ""
                                        new_title2 = new_title2.replace(word, "").strip() if new_title2 else ""
                                        new_title3 = new_title3.replace(word, "").strip() if new_title3 else ""
                                    # 연속 공백 정리
                                    apply_title = re.sub(r'\s+', ' ', apply_title).strip()
                                    new_title = re.sub(r'\s+', ' ', new_title).strip() if new_title else ""
                                    new_title2 = re.sub(r'\s+', ' ', new_title2).strip() if new_title2 else ""
                                    new_title3 = re.sub(r'\s+', ' ', new_title3).strip() if new_title3 else ""
                                    if original_apply != apply_title:
                                        self.gui.log(f"🗑️ 자동 제거: {', '.join(removed_words)}")
                                
                                # 안전 판정 받은 단어 제외
                                filtered_brands = [b for b in detected_brands if b not in safe_words]
                                
                                # 최종 의심단어 판단
                                # - 유명 브랜드 → 2차검수
                                # - 애매한 단어 → 의심단어 리스트에 추가 (사람이 판단)
                                # - 제거된 단어 → 작업완료 (로그만)
                                has_famous_brand = len(famous_brands_found) > 0
                                has_ambiguous = len(ambiguous_words) > 0
                                
                                if has_famous_brand:
                                    brand_names = [w[0] for w in famous_brands_found]
                                    self.gui.log_warning(f"🚨 유명 브랜드 발견: {', '.join(brand_names)} → 2차검수")
                                    is_suspicious = True
                                elif has_ambiguous:
                                    ambig_names = [w[0] for w in ambiguous_words]
                                    self.gui.log_warning(f"⚠️ 미확인 단어: {', '.join(ambig_names)} → 의심단어로 분류")
                                    # 애매한 단어는 결과 데이터의 suspicious에 포함됨 (나중에 수집)
                                    is_suspicious = False  # 2차검수 아님, 작업완료 처리
                                else:
                                    # 의심단어는 있지만 모두 안전/제거 처리됨
                                    if detected_brands:
                                        self.gui.log(f"ℹ️ 의심단어 처리 완료: {', '.join(detected_brands[:3])} (안전/제거됨)")
                                    is_suspicious = False
                                
                                # 태그 결정 로직:
                                # - 이미지 불일치 확정 → 2차검수
                                # - 위험/금지단어 → 2차검수
                                # - 의심단어 발견 → 2차검수 (브랜드, 피규어 등)
                                # - 그 외 → 기본 태그
                                if title_mode == TITLE_MODE_IMAGE_FIRST and is_mismatch:
                                    actual_tag = SECOND_CHECK_TAG
                                elif is_dangerous or forbidden_found or is_suspicious:
                                    actual_tag = SECOND_CHECK_TAG
                                else:
                                    actual_tag = tag_name
                                
                                # 태그 적용하여 업데이트 (선택된 상품명 적용)
                                if self.update_product_title(product, apply_title, actual_tag):
                                    group_processed += 1
                                    
                                    # 결과 데이터에 저장 (나중에 Claude 문맥 분석용)
                                    result_data.append({
                                        'group': group_name or '(현재 마켓)',
                                        'bulsaja_id': product.bulsaja_id,
                                        'code': product.seller_code,
                                        'original': product.original_title,
                                        'new': new_title,  # 상품명1
                                        'new2': new_title2,  # 상품명2
                                        'new3': new_title3,  # 상품명3
                                        'applied': apply_choice,  # 적용된 상품명 번호 (1, 2, 3)
                                        'suspicious': detected_brands,
                                        'row_index': product.index,
                                        'is_dangerous': is_dangerous or forbidden_found,  # 진짜 위험 또는 금지단어 (빨간색)
                                        'danger_categories': danger_categories,
                                        'forbidden_found': forbidden_found,  # 금지단어 발견 여부
                                        'is_suspicious': is_suspicious,  # 유명브랜드 발견 여부 (노란색)
                                        'famous_brands': [w[0] for w in famous_brands_found] if famous_brands_found else [],
                                        'removed_words': [w[0] for w in words_to_remove] if words_to_remove else [],
                                        'ambiguous_words': [w[0] for w in ambiguous_words] if ambiguous_words else [],
                                    })
                                    
                                    # 진행 상황 업데이트
                                    self.gui.update_progress(current_num, count, group_processed, group_failed)
                                else:
                                    group_failed += 1
                                    self.gui.update_progress(current_num, count, group_processed, group_failed)
                                
                                time.sleep(1.5)  # 상품 간 대기 (API 과부하 방지)
                                
                            except Exception as e:
                                self.gui.log(f"❌ 처리 실패: {e}")
                                group_failed += 1
                        
                        # 다음 배치 시작 위치 - v1.6 방식 복원
                        batch_start += len(products)
                        
                        # ★ 중복 처리 방지: 실제 반환 개수가 요청보다 적으면 더 이상 상품 없음
                        if len(products) < batch_count:
                            self.gui.log(f"✅ 전체 상품 처리 완료 (가용: {group_processed}개)")
                            break
                        
                        # 목표 달성 확인
                        if group_processed >= count:
                            break
                    
                    # 목표 개수 달성하면 종료
                    if group_processed >= count:
                        break
                    
                    # 다음 페이지로 이동
                    if page_num < num_pages - 1:
                        self.gui.log("\n➡️ 다음 페이지로 이동...")
                        if not self.go_to_next_page():
                            self.gui.log("⚠️ 마지막 페이지입니다")
                            break
                
                # 그룹 결과 출력
                if group_name:
                    self.gui.log(f"\n📊 [{group_name}] 결과: 성공 {group_processed} / 실패 {group_failed}")
                
                grand_total_processed += group_processed
                grand_total_failed += group_failed
                grand_total_skipped += group_skipped
                
                # 중단 시 루프 탈출 (하지만 결과 분석은 진행)
                if not self.is_running:
                    break
            
            # ========== Claude 문맥 기반 위험 분석 (추가 검증) ==========
            # 중단되어도 그때까지 수집된 데이터로 분석 진행
            if result_data:
                self.gui.log(f"\n{'='*50}")
                self.gui.log("🔍 상품 위험도 추가 분석 중 (문맥 기반)...")
                self.gui.log(f"{'='*50}")
                
                # 분석용 데이터 준비
                products_for_analysis = [
                    {'index': i, 'title': r['new'], 'original': r['original']}
                    for i, r in enumerate(result_data)
                ]
                
                # Claude 분석 호출
                risk_result = self.analyze_products_risk(products_for_analysis)
                
                # 위험상품 처리 (패턴 기반에서 놓친 것만)
                danger_items = risk_result.get('danger', [])
                newly_found = []  # 새로 발견된 위험 상품
                
                if danger_items:
                    all_danger_keywords = []
                    
                    for item in danger_items:
                        idx = item.get('index', -1)
                        if 0 <= idx < len(result_data):
                            # 이미 위험으로 표시된 것은 스킵
                            if result_data[idx]['is_dangerous']:
                                continue
                            
                            # 새로 발견된 위험 상품
                            newly_found.append(item)
                            
                            # result_data 업데이트
                            result_data[idx]['is_dangerous'] = True
                            result_data[idx]['danger_categories'] = {'detected': item.get('keywords', [])}
                            
                            # 위험 키워드 수집
                            all_danger_keywords.extend(item.get('keywords', []))
                            
                            # 태그 변경 (패턴 기반에서 안전→위험으로 "위험상품" 태그 적용)
                            row_index = result_data[idx].get('row_index')
                            if row_index is not None:
                                try:
                                    row = self.main_driver.find_element(
                                        By.CSS_SELECTOR, f"div[role='row'][row-index='{row_index}']"
                                    )
                                    self._apply_tag_to_row(row, row_index, "위험상품")
                                except:
                                    pass
                    
                    # 새로 발견된 위험 상품 로그 (빨간색 + 아이콘)
                    if newly_found:
                        self.gui.log_warning(f"\n🚨 추가 위험상품 {len(newly_found)}개 감지! (문맥 분석)")
                        for item in newly_found:
                            self.gui.log_warning(f"  • {item['title'][:40]}...")
                            self.gui.log_warning(f"    이유: {item.get('reason', '')}")
                    
                    # ⚠️ 자동 금지단어 추가 제거 - 의심단어로만 표시
                    # 위험 키워드는 엑셀 파일에 의심단어로 기록됨
                    if all_danger_keywords:
                        unique_keywords = list(set(all_danger_keywords))
                        self.gui.log(f"📝 의심단어 {len(unique_keywords)}개 발견 (자동 추가 안 함)")
                
                # 의심 항목 수집
                suspect_items = risk_result.get('suspect', [])
                suspect_keywords = []
                for item in suspect_items:
                    suspect_keywords.extend(item.get('keywords', []))
                
                # 기존 브랜드 의심 + 새로운 의심 통합
                all_suspects = []
                for r in result_data:
                    all_suspects.extend(r['suspicious'])
                all_suspects.extend(suspect_keywords)
                confirmed_suspects = list(set(all_suspects))  # 중복 제거
                
                # 의심단어 패널 업데이트
                if confirmed_suspects:
                    self.gui.log_warning(f"\n🏷️ 의심 단어 {len(confirmed_suspects)}개 발견")
                    # v2.1: 괄호 안 내용은 함수 내부에서 설명으로 추출
                    # Claude 분석용으로는 괄호 제거한 단어만 전달
                    clean_for_analysis = []
                    for s in confirmed_suspects:
                        clean = re.sub(r'\([^)]*\)$', '', s).strip()
                        if clean:
                            clean_for_analysis.append(clean)
                    clean_for_analysis = list(set(clean_for_analysis))
                    
                    brand_descriptions = self.analyze_brand_descriptions(clean_for_analysis)
                    # GUI에는 원본(괄호 포함) 전달 → 함수에서 괄호 분리
                    self.gui.update_suspect_list_with_desc(confirmed_suspects, brand_descriptions)
                else:
                    self.gui.log("✅ 의심 단어 없음")
                    self.gui.update_suspect_list([])
            
            # 결과 파일 저장
            self._save_result_file(result_filename, result_data)
            
            # 위험 상품 통계
            danger_count = sum(1 for r in result_data if r.get('is_dangerous', False))
            safe_count = len(result_data) - danger_count
            
            # 최종 결과
            self.gui.log(f"\n{'#'*60}")
            self.gui.log(f"✅ 전체 완료: 성공 {grand_total_processed} / 실패 {grand_total_failed}")
            if danger_count > 0:
                self.gui.log_warning(f"🚨 위험 상품: {danger_count}개 (위험상품 태그 적용됨)")
                self.gui.log(f"✅ 안전 상품: {safe_count}개")
            if grand_total_skipped > 0:
                self.gui.log(f"ℹ️ 이미지 확장자 없음: {grand_total_skipped}개")
            self.gui.log(f"📄 결과 저장: {os.path.abspath(result_filename)}")
            self.gui.log(f"{'#'*60}")
            
            suspect_count = len(confirmed_suspects)
            messagebox.showinfo("완료", f"처리 완료!\n성공: {grand_total_processed} / 실패: {grand_total_failed}\n\n의심 단어: {suspect_count}개\n(오른쪽 패널에서 확인 후 금지단어 추가 가능)\n\n결과 파일: {result_filename}")
            
        except Exception as e:
            self.gui.log(f"❌ 오류 발생: {e}")
            messagebox.showerror("오류", str(e))
        finally:
            self.is_running = False
            self.gui.on_finished()
    
    def _save_result_file(self, filename: str, data: list):
        """결과 파일 저장 (xlsx 또는 csv)"""
        if OPENPYXL_AVAILABLE and filename.endswith('.xlsx'):
            self._save_xlsx(filename, data)
        else:
            self._save_csv(filename.replace('.xlsx', '.csv'), data)
    
    def _save_xlsx(self, filename: str, data: list):
        """xlsx 파일로 저장 (색상 포함)"""
        
        def highlight_suspicious_words(text: str, suspicious_words: list):
            """상품명에서 주의단어만 빨간색으로 표시"""
            if not suspicious_words or not text:
                return text
            
            # RichText 사용 가능한지 확인
            if not OPENPYXL_AVAILABLE:
                return text
            
            try:
                # 주의단어가 실제로 상품명에 포함되어 있는지 확인
                found_words = [w for w in suspicious_words if w in text]
                if not found_words:
                    return text
                
                # RichText 생성
                red_font = InlineFont(color='FF0000', bold=True)
                parts = []
                remaining_text = text
                
                # 각 주의단어를 빨간색으로 표시
                for word in found_words:
                    if word not in remaining_text:
                        continue
                    
                    # 단어 위치 찾기
                    idx = remaining_text.find(word)
                    if idx == -1:
                        continue
                    
                    # 앞부분 (일반 텍스트)
                    if idx > 0:
                        parts.append(TextBlock(InlineFont(), remaining_text[:idx]))
                    
                    # 주의단어 (빨간색)
                    parts.append(TextBlock(red_font, word))
                    
                    # 나머지 텍스트
                    remaining_text = remaining_text[idx + len(word):]
                
                # 마지막 남은 텍스트
                if remaining_text:
                    parts.append(TextBlock(InlineFont(), remaining_text))
                
                if parts:
                    return CellRichText(*parts)
                else:
                    return text
                    
            except Exception as e:
                # RichText 실패 시 원본 텍스트 반환
                return text
        
        wb = Workbook()
        ws = wb.active
        ws.title = "결과"
        
        # 스타일 정의
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # 적용된 헤더
        red_font = Font(color="FF0000", bold=True)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        applied_header_fill = PatternFill(start_color="228B22", end_color="228B22", fill_type="solid")  # 적용됨 헤더 (진한 초록)
        
        # 적용된 상품명 번호 확인 (첫 번째 데이터 기준)
        applied_num = data[0].get('applied', '1') if data else '1'
        
        # 헤더
        headers = ['마켓 그룹', '불사자 코드', '판매자 상품코드', '기존 상품명', '상품명1', '상품명2', '상품명3', '주의단어', '위험등급', '위험카테고리']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            
            # 적용된 상품명 헤더는 초록색
            if (col == 5 and applied_num == '1') or (col == 6 and applied_num == '2') or (col == 7 and applied_num == '3'):
                cell.fill = applied_header_fill
                cell.font = header_font
            else:
                cell.fill = header_fill
                cell.font = header_font
        
        # 데이터
        for row_idx, item in enumerate(data, 2):
            ws.cell(row=row_idx, column=1, value=item['group'])
            ws.cell(row=row_idx, column=2, value=item.get('bulsaja_id', ''))
            ws.cell(row=row_idx, column=3, value=item['code'])
            ws.cell(row=row_idx, column=4, value=item['original'])
            
            # 주의단어 리스트
            suspicious_words = item.get('suspicious', [])
            
            # 변경 상품명 1, 2, 3 (주의단어 하이라이트 적용)
            cell1 = ws.cell(row=row_idx, column=5)
            cell1.value = highlight_suspicious_words(item['new'], suspicious_words)
            
            cell2 = ws.cell(row=row_idx, column=6)
            cell2.value = highlight_suspicious_words(item.get('new2', ''), suspicious_words)
            
            cell3 = ws.cell(row=row_idx, column=7)
            cell3.value = highlight_suspicious_words(item.get('new3', ''), suspicious_words)
            
            # 위험 등급 판정
            is_dangerous = item.get('is_dangerous', False)
            is_suspicious = item.get('is_suspicious', False) or bool(item.get('suspicious'))
            categories = item.get('danger_categories', {})
            
            # 적용된 셀 찾기
            if applied_num == '1':
                applied_cell = cell1
            elif applied_num == '2':
                applied_cell = cell2
            else:
                applied_cell = cell3
            
            # 위험 등급 결정
            danger_level = ""
            if is_dangerous:
                if categories.get('weapon') or categories.get('drug') or categories.get('illegal'):
                    danger_level = "🚫 판매불가"
                else:
                    danger_level = "⚠️ 위험"
                applied_cell.fill = red_fill
            elif is_suspicious:
                danger_level = "🔶 브랜드주의"
                applied_cell.fill = yellow_fill
            
            # 주의단어 (빨간색 글씨 + 노란 배경)
            if item.get('suspicious'):
                suspicious_cell = ws.cell(row=row_idx, column=8, value=', '.join(item['suspicious']))
                suspicious_cell.font = red_font
                suspicious_cell.fill = yellow_fill
            
            # 위험등급 표시
            if danger_level:
                danger_cell = ws.cell(row=row_idx, column=9, value=danger_level)
                if "판매불가" in danger_level or "위험" in danger_level:
                    danger_cell.font = red_font
                    danger_cell.fill = red_fill
                elif "브랜드" in danger_level:
                    danger_cell.font = Font(color="FF6600", bold=True)
                    danger_cell.fill = yellow_fill
            
            # 위험 카테고리
            if categories:
                cat_texts = []
                for cat, words in categories.items():
                    if words:
                        cat_name = get_danger_category_name(cat)
                        cat_texts.append(f"{cat_name}: {', '.join(words[:3])}")
                if cat_texts:
                    cat_cell = ws.cell(row=row_idx, column=10, value='\n'.join(cat_texts))
                    cat_cell.font = red_font
        
        # 열 너비 조정
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 28
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 50
        ws.column_dimensions['E'].width = 50
        ws.column_dimensions['F'].width = 50
        ws.column_dimensions['G'].width = 50
        ws.column_dimensions['H'].width = 30  # 주의단어
        ws.column_dimensions['I'].width = 15  # 위험등급
        ws.column_dimensions['J'].width = 60  # 위험 카테고리/설명
        
        wb.save(filename)
        self.gui.log(f"✅ xlsx 파일 저장 완료")
    
    def _save_csv(self, filename: str, data: list):
        """csv 파일로 저장 (색상 없음)"""
        # 적용된 상품명 번호 확인 (첫 번째 데이터 기준)
        applied_num = data[0].get('applied', '1') if data else '1'
        
        with open(filename, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            # 헤더에 적용된 상품명 표시 (색상 없으므로 * 표시)
            headers = ['마켓 그룹', '불사자 코드', '판매자 상품코드', '기존 상품명', '상품명1', '상품명2', '상품명3', '주의단어', '위험등급']
            if applied_num == '1':
                headers[4] = '*상품명1'
            elif applied_num == '2':
                headers[5] = '*상품명2'
            else:
                headers[6] = '*상품명3'
            
            writer.writerow(headers)
            
            for item in data:
                # 위험등급 결정
                danger_level = ""
                if item.get('is_dangerous'):
                    categories = item.get('danger_categories', {})
                    if categories.get('weapon') or categories.get('drug') or categories.get('illegal'):
                        danger_level = "판매불가"
                    else:
                        danger_level = "위험"
                elif item.get('is_suspicious') or item.get('suspicious'):
                    danger_level = "브랜드주의"
                
                writer.writerow([
                    item['group'],
                    item.get('bulsaja_id', ''),
                    item['code'],
                    item['original'],
                    item['new'],
                    item.get('new2', ''),
                    item.get('new3', ''),
                    ', '.join(item['suspicious']) if item['suspicious'] else '',
                    danger_level
                ])
        self.gui.log(f"✅ csv 파일 저장 완료 (색상 미지원)")

# ==================== GUI ====================
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("불사자 상품명 자동 변환기+검수 v3.0 (API)")
        self.geometry("1350x950")
        
        self.filler = BulsajaAutoFiller(self)
        self.worker_thread = None
        self.config = load_config()
        
        # ★ v3.1: 엑셀 기반 모드 변수
        self.excel_mode_enabled = tk.BooleanVar(value=False)
        self.excel_file_path = tk.StringVar(value="")
        self.excel_title_number = tk.StringVar(value="2")  # 2번 or 3번
        
        # ★ v3.0: API 모드용 필터 변수
        self.current_market_group = None  # 마켓 그룹 필터
        self.current_tag_filter = None    # 태그 필터
        
        self._build_ui()
        self._load_saved_settings()
        
        # 프로그램 시작 시 시트 자동 동기화
        self.after(1000, self._auto_sync_on_start)
        
        # 프로그램 종료 시 자동 저장
        self.protocol("WM_DELETE_WINDOW", self.on_closing)
    
    def _auto_sync_on_start(self):
        """프로그램 시작 시 Google Sheets 자동 동기화"""
        # config에 없으면 기본 시트 ID 사용
        default_sheet_id = '1r-ROJ7ksv6qOtOTXbkrprxu17EQmbO-n1J1pm_N5Hh8'
        sheet_url = self.config.get('sheets_url', '') or default_sheet_id
        
        if not sheet_url:
            return
        
        def sync_task():
            self.log("🔄 시작 시 시트 자동 동기화...")
            result = sync_from_google_sheets(sheet_url, self.log)
            
            if result['success'] and (result['banned'] or result['excluded'] or result.get('remove')):
                # 금지단어 저장
                if result['banned']:
                    banned_data = {'words': result['banned']}
                    save_banned_words(banned_data)
                    self.filler.banned_words = set(result['banned'])
                    self.filler.banned_words_data = banned_data
                
                # 예외단어 저장
                if result['excluded']:
                    save_excluded_words(set(result['excluded']))
                    global EXCLUDED_WORDS
                    EXCLUDED_WORDS = set(result['excluded'])
                
                # 제거단어 저장
                if result.get('remove'):
                    save_remove_words(set(result['remove']))
                    global REMOVE_WORDS
                    REMOVE_WORDS = set(result['remove'])
                
                self.sheets_status.config(text="✅ 자동동기화", foreground="green")
            elif result['success']:
                self.log("⚠️ 시트가 비어있어 로컬 데이터 유지")
        
        threading.Thread(target=sync_task, daemon=True).start()
    
    def toggle_excel_mode(self):
        """엑셀 모드 토글"""
        enabled = self.excel_mode_enabled.get()
        state = 'normal' if enabled else 'disabled'
        
        self.excel_file_entry.config(state=state)
        self.excel_browse_btn.config(state=state)
        self.excel_radio2.config(state=state)
        self.excel_radio3.config(state=state)
        
        if enabled:
            self.log("📊 엑셀 적용 모드 활성화 - AI 사용 안 함")
        else:
            self.log("🤖 AI 생성 모드 활성화")
    
    def browse_excel_file(self):
        """엑셀 파일 선택"""
        from tkinter import filedialog
        
        filepath = filedialog.askopenfilename(
            title="엑셀 파일 선택",
            initialdir="result",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        
        if filepath:
            self.excel_file_path.set(filepath)
            self.log(f"📊 엑셀 파일 선택: {os.path.basename(filepath)}")
    
    def open_debug_chrome(self):
        """크롬 디버그 모드로 열기 (자동 포트 탐색)"""
        import subprocess
        
        # 사용 가능한 포트 자동 탐색
        port = find_available_port()
        self.port_var.set(str(port))  # GUI 포트칸 업데이트
        
        # 포트별 프로필 폴더 (충돌 방지)
        profile_dir = f"{CHROME_DEBUG_PROFILE}_{port}"
        
        # 크롬 경로 찾기
        chrome_paths = [
            r"C:\Program Files\Google\Chrome\Application\chrome.exe",
            r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
            os.path.expandvars(r"%LOCALAPPDATA%\Google\Chrome\Application\chrome.exe"),
        ]
        
        chrome_path = None
        for path in chrome_paths:
            if os.path.exists(path):
                chrome_path = path
                break
        
        if not chrome_path:
            messagebox.showerror("오류", "Chrome을 찾을 수 없습니다")
            return
        
        url = BULSAJA_PRODUCT_LIST_URL
        cmd = f'"{chrome_path}" --remote-debugging-port={port} --user-data-dir="{profile_dir}" --remote-allow-origins=* "{url}"'
        
        try:
            subprocess.Popen(cmd, shell=True)
            self.log(f"🌐 크롬 디버그 모드 실행 (포트: {port})")
            self.log(f"📁 프로필: {profile_dir}")
            self.log(f"   불사자에 로그인 후 '토큰 자동 추출' 버튼을 클릭하세요")
        except Exception as e:
            self.log(f"❌ 크롬 실행 실패: {e}")
    
    def extract_tokens(self):
        """크롬에서 토큰 자동 추출"""
        port = int(self.port_var.get().strip())
        self.log(f"🔍 크롬 포트 {port}에서 토큰 추출 중...")
        self.token_status.config(text="추출 중...", foreground="orange")
        
        def extract_task():
            success, access_token, refresh_token = self.filler.extract_tokens_from_browser(port)
            
            if success:
                self.access_token_var.set(access_token)
                self.refresh_token_var.set(refresh_token)
                self.log(f"✅ 토큰 추출 성공!")
                self.log(f"  Access Token: {access_token[:50]}...")
                self.log(f"  Refresh Token: {refresh_token[:50]}...")
                self.token_status.config(text="✅ 추출 완료", foreground="green")
                
                # 자동으로 API 연결
                self.after(500, self.connect_api)
            else:
                self.log("❌ 토큰 추출 실패")
                self.log("   1. 크롬을 디버깅 모드로 실행하세요:")
                self.log(f'      chrome.exe --remote-debugging-port={port}')
                self.log("   2. 불사자(bulsaja.com)에 로그인하세요")
                self.log("   3. 다시 '토큰 가져오기' 버튼을 클릭하세요")
                self.token_status.config(text="❌ 실패 (로그 확인)", foreground="red")
        
        threading.Thread(target=extract_task, daemon=True).start()
    
    def connect_api(self):
        """불사자 API 연결"""
        access_token = self.access_token_var.get().strip()
        refresh_token = self.refresh_token_var.get().strip()
        
        if not access_token or not refresh_token:
            messagebox.showwarning("경고", "토큰을 입력하세요")
            return
        
        self.log("🔗 불사자 API 연결 중...")
        
        success, msg, total = self.filler.init_api_client(access_token, refresh_token)
        
        if success:
            self.api_conn_status.config(text=f"✅ 연결됨 ({total}개)", foreground="green")
            self.log(f"✅ {msg}")
            
            # Claude API 초기화
            api_key = self.api_var.get().strip()
            if api_key:
                if self.filler.setup_claude(api_key):
                    self.api_status.config(text="✅", foreground="green")
        else:
            self.api_conn_status.config(text="❌ 실패", foreground="red")
            self.log(f"❌ 연결 실패: {msg}")
    
    def on_closing(self):
        """프로그램 종료 시 설정 자동 저장"""
        try:
            # 주요 설정 저장
            self.config['work_groups'] = self.groups_var.get().strip()
            self.config['market_groups'] = self.market_groups_var.get().strip()
            self.config['copy_groups'] = self.copy_groups_var.get().strip()
            self.config['copy_count'] = self.copy_count_var.get().strip()
            self.config['tag_name'] = self.tag_var.get().strip()
            self.config['danger_tag'] = self.danger_tag_var.get().strip()
            save_config(self.config)
        except:
            pass
        # AliPrice 창 닫기
        try:
            self.filler.close_aliprice()
        except:
            pass
        self.destroy()
    
    def _build_ui(self):
        main = ttk.Frame(self, padding="5")
        main.pack(fill="both", expand=True)
        
        
        # ========== 상단 설정 영역 (좌우 배치) ==========
        top_frame = ttk.Frame(main)
        top_frame.pack(fill="x", pady=(0,5))
        
        # 왼쪽: API 토큰 설정
        left_frame = ttk.LabelFrame(top_frame, text="🔑 불사자 API 연결 (v3.0)", padding="5")
        left_frame.pack(side="left", fill="both", expand=True, padx=(0,3))
        
        # 크롬 열기 + 토큰 추출
        token_row0 = ttk.Frame(left_frame)
        token_row0.pack(fill="x", pady=2)
        ttk.Button(token_row0, text="🌐 크롬 열기", command=self.open_debug_chrome, width=10).pack(side="left")
        ttk.Button(token_row0, text="🔑 토큰 가져오기", command=self.extract_tokens, width=12).pack(side="left", padx=5)
        self.token_status = ttk.Label(token_row0, text="", foreground="gray")
        self.token_status.pack(side="left", padx=5)
        ttk.Label(token_row0, text="포트:").pack(side="left", padx=(10,0))
        self.port_var = tk.StringVar(value=str(DEBUG_PORT))
        ttk.Entry(token_row0, textvariable=self.port_var, width=5).pack(side="left", padx=2)
        
        # Access Token
        token_row1 = ttk.Frame(left_frame)
        token_row1.pack(fill="x", pady=2)
        ttk.Label(token_row1, text="Access Token:").pack(side="left")
        self.access_token_var = tk.StringVar()
        ttk.Entry(token_row1, textvariable=self.access_token_var, width=55).pack(side="left", padx=5, fill="x", expand=True)
        
        # Refresh Token + 연결 버튼
        token_row2 = ttk.Frame(left_frame)
        token_row2.pack(fill="x", pady=2)
        ttk.Label(token_row2, text="Refresh Token:").pack(side="left")
        self.refresh_token_var = tk.StringVar()
        ttk.Entry(token_row2, textvariable=self.refresh_token_var, width=40).pack(side="left", padx=5)
        ttk.Button(token_row2, text="🔗 연결", command=self.connect_api, width=6).pack(side="left", padx=5)
        self.api_conn_status = ttk.Label(token_row2, text="미연결", foreground="gray")
        self.api_conn_status.pack(side="left", padx=5)
        
        # Claude API Key + 저장
        api_row = ttk.Frame(left_frame)
        api_row.pack(fill="x", pady=2)
        ttk.Label(api_row, text="Claude API:").pack(side="left")
        self.api_var = tk.StringVar()
        ttk.Entry(api_row, textvariable=self.api_var, width=45, show="•").pack(side="left", padx=5)
        ttk.Button(api_row, text="💾 저장", command=self.save_settings, width=6).pack(side="left")
        self.api_status = ttk.Label(api_row, text="", foreground="gray")
        self.api_status.pack(side="left", padx=5)
        
        # Google Sheets 동기화 (금지단어/예외단어)
        sheets_row = ttk.Frame(left_frame)
        sheets_row.pack(fill="x", pady=2)
        ttk.Label(sheets_row, text="시트URL:").pack(side="left")
        default_sheet = self.config.get('sheets_url', '1r-ROJ7ksv6qOtOTXbkrprxu17EQmbO-n1J1pm_N5Hh8')
        self.sheets_url_var = tk.StringVar(value=default_sheet)
        ttk.Entry(sheets_row, textvariable=self.sheets_url_var, width=38).pack(side="left", padx=5)
        ttk.Button(sheets_row, text="🔄 동기화", command=self.sync_words_from_sheets, width=8).pack(side="left")
        self.sheets_status = ttk.Label(sheets_row, text="", foreground="gray")
        self.sheets_status.pack(side="left", padx=5)
        
        # 오른쪽: 모델 + 다양성
        right_frame = ttk.LabelFrame(top_frame, text="⚙️ Claude 설정", padding="5")
        right_frame.pack(side="left", fill="both", expand=True, padx=(3,0))
        
        # 상품명 생성 모델 선택
        model_row = ttk.Frame(right_frame)
        model_row.pack(fill="x", pady=2)
        ttk.Label(model_row, text="상품명:").pack(side="left")
        self.model_var = tk.StringVar(value="claude-3-5-haiku-20241022")
        models = [("Haiku", "claude-3-5-haiku-20241022"), ("Sonnet", "claude-3-5-sonnet-20241022")]
        for label, value in models:
            ttk.Radiobutton(model_row, text=label, variable=self.model_var, value=value).pack(side="left", padx=5)
        
        # Vision 분석 모델 선택
        vision_model_row = ttk.Frame(right_frame)
        vision_model_row.pack(fill="x", pady=2)
        ttk.Label(vision_model_row, text="Vision:").pack(side="left")
        self.vision_model_var = tk.StringVar(value="claude-3-5-haiku-20241022")
        for label, value in models:
            ttk.Radiobutton(vision_model_row, text=label, variable=self.vision_model_var, value=value).pack(side="left", padx=5)
        
        # Temperature
        temp_row = ttk.Frame(right_frame)
        temp_row.pack(fill="x", pady=2)
        ttk.Label(temp_row, text="다양성:").pack(side="left")
        self.temp_var = tk.StringVar(value="0.7")
        temps = [("0.0", "0.0"), ("0.3", "0.3"), ("0.7✓", "0.7"), ("1.0", "1.0")]
        for label, value in temps:
            ttk.Radiobutton(temp_row, text=label, variable=self.temp_var, value=value).pack(side="left", padx=5)
        
        # 상품명 길이 고정 (30~40자) - GUI 옵션 제거, 변수만 유지
        self.title_length_var = tk.StringVar(value="35")  # 고정값
        self.title_logic_var = tk.StringVar(value="basic")  # 고정값
        
        # 상품명 생성 방식
        mode_row = ttk.Frame(right_frame)
        mode_row.pack(fill="x", pady=2)
        ttk.Label(mode_row, text="생성방식:").pack(side="left")
        self.title_mode_var = tk.StringVar(value=TITLE_MODE_IMAGE_FIRST)
        ttk.Radiobutton(mode_row, text="이미지+기존명", variable=self.title_mode_var, value=TITLE_MODE_IMAGE_FIRST).pack(side="left", padx=5)
        ttk.Radiobutton(mode_row, text="Vision분석", variable=self.title_mode_var, value=TITLE_MODE_VISION).pack(side="left", padx=5)
        ttk.Radiobutton(mode_row, text="기존명만", variable=self.title_mode_var, value=TITLE_MODE_ORIGINAL_ONLY).pack(side="left", padx=5)
        
        # 적용할 상품명 선택 (1번/2번/3번)
        apply_row = ttk.Frame(right_frame)
        apply_row.pack(fill="x", pady=2)
        ttk.Label(apply_row, text="적용상품명:").pack(side="left")
        self.title_apply_var = tk.StringVar(value="1")
        ttk.Radiobutton(apply_row, text="1번(기본)", variable=self.title_apply_var, value="1").pack(side="left", padx=5)
        ttk.Radiobutton(apply_row, text="2번(용도+메인)", variable=self.title_apply_var, value="2").pack(side="left", padx=5)
        ttk.Radiobutton(apply_row, text="3번(다른키워드)", variable=self.title_apply_var, value="3").pack(side="left", padx=5)
        
        # ========== 엑셀 적용 모드 (v3.1) ==========
        excel_frame = ttk.LabelFrame(main, text="📊 엑셀 적용 모드 (그룹2,3용)", padding="5")
        excel_frame.pack(fill="x", pady=2)
        
        excel_inner = ttk.Frame(excel_frame)
        excel_inner.pack(fill="x")
        
        # 체크박스
        self.excel_mode_check = ttk.Checkbutton(
            excel_inner, 
            text="엑셀 기반 모드 (AI 사용 안 함)",
            variable=self.excel_mode_enabled,
            command=self.toggle_excel_mode
        )
        self.excel_mode_check.pack(side="left", padx=5)
        
        # 파일 선택
        ttk.Label(excel_inner, text="엑셀:").pack(side="left", padx=(10,0))
        self.excel_file_entry = ttk.Entry(excel_inner, textvariable=self.excel_file_path, width=40, state='disabled')
        self.excel_file_entry.pack(side="left", padx=5)
        
        self.excel_browse_btn = ttk.Button(excel_inner, text="📁 찾기", command=self.browse_excel_file, state='disabled')
        self.excel_browse_btn.pack(side="left")
        
        # 상품명 번호 선택
        ttk.Label(excel_inner, text="적용:").pack(side="left", padx=(10,0))
        self.excel_radio2 = ttk.Radiobutton(excel_inner, text="2번", variable=self.excel_title_number, value="2", state='disabled')
        self.excel_radio2.pack(side="left", padx=2)
        
        self.excel_radio3 = ttk.Radiobutton(excel_inner, text="3번", variable=self.excel_title_number, value="3", state='disabled')
        self.excel_radio3.pack(side="left", padx=2)
        
        # ========== 태그 설정 ==========
        tag_frame = ttk.LabelFrame(main, text="🏷️ 태그 설정", padding="3")
        tag_frame.pack(fill="x", pady=2)
        
        tag_inner = ttk.Frame(tag_frame)
        tag_inner.pack(fill="x")
        
        ttk.Label(tag_inner, text="작업완료:").pack(side="left")
        self.tag_var = tk.StringVar(value=f"작업완료_{datetime.now().strftime('%y%m%d_%H%M')}")
        self.tag_entry = ttk.Entry(tag_inner, textvariable=self.tag_var, width=30)
        self.tag_entry.pack(side="left", padx=5)
        
        # 태그 새로고침 버튼
        self.btn_refresh_tag = ttk.Button(tag_inner, text="🔄", width=3, command=self.refresh_tag_time)
        self.btn_refresh_tag.pack(side="left")
        
        ttk.Label(tag_inner, text="위험상품:").pack(side="left", padx=(10,0))
        self.danger_tag_var = tk.StringVar(value="위험상품")
        self.danger_tag_entry = ttk.Entry(tag_inner, textvariable=self.danger_tag_var, width=20)
        self.danger_tag_entry.pack(side="left", padx=5)
        
        self.btn_create_tag = ttk.Button(tag_inner, text="📌 태그 생성", command=self.create_tag_manual)
        self.btn_create_tag.pack(side="left", padx=5)
        
        # ========== 작업 그룹 설정 ==========
        group_frame = ttk.LabelFrame(main, text="📁 작업 그룹 (순차 처리)", padding="3")
        group_frame.pack(fill="x", pady=2)
        
        group_inner = ttk.Frame(group_frame)
        group_inner.pack(fill="x")
        
        ttk.Label(group_inner, text="그룹:").pack(side="left")
        self.groups_var = tk.StringVar(value="")
        ttk.Entry(group_inner, textvariable=self.groups_var, width=60).pack(side="left", padx=5, fill="x", expand=True)
        ttk.Label(group_inner, text="(쉼표 구분, 비우면 현재 그룹만)", foreground="gray").pack(side="left")
        
        # ========== 마켓 그룹 목록 (숫자 맵핑용) ==========
        market_frame = ttk.LabelFrame(main, text="🏪 마켓 그룹 목록 (숫자 맵핑용)", padding="3")
        market_frame.pack(fill="x", pady=2)
        
        market_inner = ttk.Frame(market_frame)
        market_inner.pack(fill="x")
        
        self.market_groups_var = tk.StringVar(value="")
        ttk.Entry(market_inner, textvariable=self.market_groups_var, width=80).pack(side="left", padx=5, fill="x", expand=True)
        
        market_help = ttk.Frame(market_frame)
        market_help.pack(fill="x", pady=(1,0))
        ttk.Label(market_help, text="예: 01_마켓A,02_마켓B,03_마켓C,04_마켓D → 작업그룹/복사그룹에서 1, 1-3, 2,4 등으로 사용 가능", foreground="gray", font=("", 8)).pack(side="left")
        
        # ========== 복사 그룹 설정 (Mode2: 복사 후 태그 변경) ==========
        copy_frame = ttk.LabelFrame(main, text="📋 복사 (Mode2: 복사 후 태그변경)", padding="3")
        copy_frame.pack(fill="x", pady=2)
        
        # 첫째 줄: 검색 태그, 완료 태그
        copy_row1 = ttk.Frame(copy_frame)
        copy_row1.pack(fill="x", pady=(0,2))
        
        ttk.Label(copy_row1, text="검색 태그:").pack(side="left")
        self.copy_tag_var = tk.StringVar(value=f"작업완료_{datetime.now().strftime('%y%m%d_%H%M')}")
        ttk.Entry(copy_row1, textvariable=self.copy_tag_var, width=25).pack(side="left", padx=5)
        
        ttk.Label(copy_row1, text="완료 태그:").pack(side="left", padx=(10,0))
        self.copy_done_tag_var = tk.StringVar(value="작업완료")
        ttk.Entry(copy_row1, textvariable=self.copy_done_tag_var, width=15).pack(side="left", padx=5)
        
        ttk.Label(copy_row1, text="(복사 후 완료 태그 추가 → 다음 검색에서 제외)", foreground="gray").pack(side="left")
        
        # 둘째 줄: 복사 그룹, 수량
        copy_inner = ttk.Frame(copy_frame)
        copy_inner.pack(fill="x")
        
        ttk.Label(copy_inner, text="복사 그룹:").pack(side="left")
        self.copy_groups_var = tk.StringVar(value="")
        ttk.Entry(copy_inner, textvariable=self.copy_groups_var, width=40).pack(side="left", padx=5)
        
        ttk.Label(copy_inner, text="수량:").pack(side="left", padx=(10,0))
        self.copy_count_var = tk.StringVar(value="100")
        copy_count_combo = ttk.Combobox(copy_inner, textvariable=self.copy_count_var, width=6,
                                         values=["10", "20", "50", "100", "200", "300", "400", "500", "1000"],
                                         state="readonly")
        copy_count_combo.pack(side="left", padx=5)
        
        self.btn_copy = ttk.Button(copy_inner, text="📋 복사 시작", command=self.start_copy)
        self.btn_copy.pack(side="left", padx=10)
        
        # ========== 처리 범위 + 버튼 ==========
        action_frame = ttk.Frame(main)
        action_frame.pack(fill="x", pady=(0,10))
        
        # 처리 범위
        ttk.Label(action_frame, text="시작:").pack(side="left")
        self.start_var = tk.StringVar(value="0")
        ttk.Entry(action_frame, textvariable=self.start_var, width=6).pack(side="left", padx=2)
        
        ttk.Label(action_frame, text="개수(그룹당):").pack(side="left", padx=(10,0))
        self.count_var = tk.StringVar(value="1000")
        ttk.Entry(action_frame, textvariable=self.count_var, width=6).pack(side="left", padx=2)
        
        ttk.Label(action_frame, text="페이지:").pack(side="left", padx=(10,0))
        self.pages_var = tk.StringVar(value="1")
        ttk.Entry(action_frame, textvariable=self.pages_var, width=4).pack(side="left", padx=2)
        
        # 버튼 (API 모드: 크롬 연결 불필요)
        ttk.Separator(action_frame, orient="vertical").pack(side="left", padx=15, fill="y")
        
        self.btn_start = ttk.Button(action_frame, text="🚀 시작", command=self.start_automation)
        self.btn_start.pack(side="left", padx=5)
        
        self.btn_stop = ttk.Button(action_frame, text="⏹️ 중지", command=self.stop, state="disabled")
        self.btn_stop.pack(side="left", padx=5)
        
        # 금지단어 관리 버튼
        self.btn_banned = ttk.Button(action_frame, text="🚫 금지단어", command=self.show_banned_words)
        self.btn_banned.pack(side="left", padx=5)
        
        # ========== 진행 상황 표시 ==========
        progress_frame = ttk.LabelFrame(main, text="📊 진행 상황", padding="3")
        progress_frame.pack(fill="x", pady=2)
        
        # 진행률 바
        self.progress = ttk.Progressbar(progress_frame, mode='determinate', length=400)
        self.progress.pack(fill="x", pady=(0,5))
        
        # 상세 정보
        progress_info = ttk.Frame(progress_frame)
        progress_info.pack(fill="x")
        
        self.progress_label = ttk.Label(progress_info, text="진행 률:  0/0 (0.0%)", font=("", 9))
        self.progress_label.pack(side="left")
        
        self.progress_detail = ttk.Label(progress_info, text="", font=("", 9))
        self.progress_detail.pack(side="left", padx=(20,0))
        
        self.progress_stats = ttk.Label(progress_info, text="✅ 성공: 0  ❌ 실패: 0", font=("", 9))
        self.progress_stats.pack(side="right")
        
        # ========== 로그 + 의심단어 (2:1 분할) ==========
        bottom_frame = ttk.Frame(main)
        bottom_frame.pack(fill="both", expand=True)
        
        # 왼쪽: 실행 로그 (2/3)
        log_frame = ttk.LabelFrame(bottom_frame, text="📝 실행 로그", padding="5")
        log_frame.pack(side="left", fill="both", expand=True)
        
        self.log_text = scrolledtext.ScrolledText(log_frame, height=18, width=80)
        self.log_text.pack(fill="both", expand=True)
        
        # 컬러 태그 설정
        self.log_text.tag_configure("warning", foreground="red", font=("", 9, "bold"))        # ❌, 🚨
        self.log_text.tag_configure("success", foreground="green", font=("", 9, "bold"))      # ✅
        self.log_text.tag_configure("info", foreground="blue")          # ℹ️, 📋, 📝
        self.log_text.tag_configure("progress", foreground="purple")    # 🔍, ⏳
        self.log_text.tag_configure("tag", foreground="orange", font=("", 9, "bold"))         # 🏷️
        
        ttk.Button(log_frame, text="🗑️ 로그 지우기", command=self.clear_log).pack(anchor="w", pady=(3,0))
        
        # 오른쪽: 의심단어 패널 (1/3) - v1.7.1: 가로폭 확장
        suspect_frame = ttk.LabelFrame(bottom_frame, text="⚠️ 의심단어 (브랜드/상호)", padding="5")
        suspect_frame.pack(side="right", fill="both", padx=(5,0), expand=True)
        
        # 상단 버튼들 (전체선택/해제)
        top_btn_frame = ttk.Frame(suspect_frame)
        top_btn_frame.pack(fill="x", pady=(0,3))
        
        ttk.Button(top_btn_frame, text="☑️ 전체선택", command=self.select_all_suspects).pack(side="left", padx=2)
        ttk.Button(top_btn_frame, text="⬜ 전체해제", command=self.deselect_all_suspects).pack(side="left", padx=2)
        
        # 의심단어 리스트 (체크박스)
        self.suspect_listbox_frame = ttk.Frame(suspect_frame)
        self.suspect_listbox_frame.pack(fill="both", expand=True)
        
        # 스크롤바 + 캔버스 (체크박스 담기) - 가로폭 확장
        self.suspect_canvas = tk.Canvas(self.suspect_listbox_frame, width=400, height=200)
        suspect_scrollbar = ttk.Scrollbar(self.suspect_listbox_frame, orient="vertical", command=self.suspect_canvas.yview)
        self.suspect_inner_frame = ttk.Frame(self.suspect_canvas)
        
        self.suspect_canvas.configure(yscrollcommand=suspect_scrollbar.set)
        suspect_scrollbar.pack(side="right", fill="y")
        self.suspect_canvas.pack(side="left", fill="both", expand=True)
        
        self.suspect_canvas_window = self.suspect_canvas.create_window((0, 0), window=self.suspect_inner_frame, anchor="nw")
        self.suspect_inner_frame.bind("<Configure>", lambda e: self.suspect_canvas.configure(scrollregion=self.suspect_canvas.bbox("all")))
        
        # 체크박스 변수들 저장
        self.suspect_vars = {}  # {단어: BooleanVar}
        self.suspect_word_info = {}  # {단어: 괄호안설명} - v2.1: 리스트 갱신 시 설명 유지
        self.suspect_descriptions = {}  # {단어: {judgment, reason}} - v2.1: Claude 분석 결과
        
        # 하단 버튼들 (차단/제외/제거)
        ttk.Button(suspect_frame, text="🚫 선택 → 금지단어 추가", command=self.add_suspects_to_banned).pack(fill="x", pady=(3,0))
        ttk.Button(suspect_frame, text="🗑️ 선택 → 제거단어 추가", command=self.add_suspects_to_remove).pack(fill="x", pady=(3,0))
        ttk.Button(suspect_frame, text="✅ 선택 → 탐지제외 처리", command=self.exclude_suspects).pack(fill="x", pady=(3,0))
        
        # 상태 레이블
        self.suspect_status = ttk.Label(suspect_frame, text="작업 완료 후 표시됩니다", foreground="gray")
        self.suspect_status.pack(anchor="w", pady=(3,0))
        
        # ========== 하단 Footer ==========
        footer_frame = ttk.Frame(main)
        footer_frame.pack(fill="x", pady=(5,0))
        ttk.Label(footer_frame, text="by 프코노미", foreground="gray").pack(side="right")
    
    def _load_saved_settings(self):
        """저장된 설정 불러오기"""
        if self.config.get('api_key'):
            self.api_var.set(self.config['api_key'])
            self.api_status.config(text="✅ 저장됨", foreground="green")
        
        # ★ v3.0: 토큰 로드
        if self.config.get('access_token'):
            self.access_token_var.set(self.config['access_token'])
        if self.config.get('refresh_token'):
            self.refresh_token_var.set(self.config['refresh_token'])
        
        # 모델 로드 - gemini 모델이면 Claude로 대체
        if self.config.get('model'):
            saved_model = self.config['model']
            # gemini 모델이면 기본값(Haiku)으로 대체
            if 'gemini' in saved_model.lower():
                self.model_var.set("claude-3-5-haiku-20241022")
                self.gui.log("⚠️ Gemini 모델 감지 → Claude Haiku로 변경")
            else:
                self.model_var.set(saved_model)
        
        # Vision 모델 로드 - gemini 모델이면 Claude로 대체
        if self.config.get('vision_model'):
            saved_vision_model = self.config['vision_model']
            # gemini 모델이면 기본값(Haiku)으로 대체
            if 'gemini' in saved_vision_model.lower():
                self.vision_model_var.set("claude-3-5-haiku-20241022")
            else:
                self.vision_model_var.set(saved_vision_model)
        
        if self.config.get('title_mode'):
            self.title_mode_var.set(self.config['title_mode'])
        
        if self.config.get('port'):
            self.port_var.set(self.config['port'])
        
        if self.config.get('temperature'):
            self.temp_var.set(self.config['temperature'])
        
        # 그룹 정보 로드
        if self.config.get('work_groups'):
            self.groups_var.set(self.config['work_groups'])
        
        if self.config.get('market_groups'):
            self.market_groups_var.set(self.config['market_groups'])
        
        if self.config.get('copy_groups'):
            self.copy_groups_var.set(self.config['copy_groups'])
        
        if self.config.get('copy_count'):
            self.copy_count_var.set(self.config['copy_count'])
        
        # 태그 로드 시 날짜 체크 (프로그램 시작 시에만 자동 변경)
        today_date = datetime.now().strftime('%y%m%d')
        today_datetime = datetime.now().strftime('%y%m%d_%H%M')
        
        if self.config.get('tag_name'):
            saved_tag = self.config['tag_name']
            # 작업완료_YYMMDD_HHMM 또는 작업완료_YYMMDD 형식 처리
            if saved_tag.startswith("작업완료_"):
                # 날짜 부분 추출 (작업완료_ 다음 6글자)
                saved_date = saved_tag[5:11] if len(saved_tag) >= 11 else ""
                if saved_date != today_date:
                    # 날짜가 다르면 현재 날짜+시간으로 변경
                    self.tag_var.set(f"작업완료_{today_datetime}")
                elif len(saved_tag) <= 12:
                    # 시간 없는 형식(작업완료_YYMMDD)이면 시간 추가
                    self.tag_var.set(f"작업완료_{today_datetime}")
                else:
                    self.tag_var.set(saved_tag)
            else:
                self.tag_var.set(saved_tag)
        
        if self.config.get('danger_tag'):
            self.danger_tag_var.set(self.config['danger_tag'])
    
    def log(self, msg: str):
        """메시지를 로그에 출력 (이모티콘별 자동 컬러 적용)"""
        # 이모티콘별 컬러 매핑
        if any(emoji in msg for emoji in ['✅', '✓']):
            tag = "success"  # 초록색
        elif any(emoji in msg for emoji in ['❌', '🚨', '⚠️', '🔴']):
            tag = "warning"  # 빨간색
        elif any(emoji in msg for emoji in ['ℹ️', '📋', '📝', '📄', '💡']):
            tag = "info"  # 파란색
        elif any(emoji in msg for emoji in ['🔍', '⏳', '🔧']):
            tag = "progress"  # 보라색
        elif any(emoji in msg for emoji in ['🏷️']):
            tag = "tag"  # 주황색
        else:
            tag = None  # 기본 색상
        
        self.log_text.insert(tk.END, f"{msg}\n", tag)
        self.log_text.see(tk.END)
        self.update_idletasks()
    
    def log_warning(self, msg: str):
        """빨간색으로 경고 메시지 표시"""
        self.log_text.insert(tk.END, f"{msg}\n", "warning")
        self.log_text.see(tk.END)
        self.update_idletasks()
    
    def clear_log(self):
        self.log_text.delete("1.0", tk.END)
    
    def _update_logic_description(self):
        """선택된 로직에 따라 설명 업데이트"""
        logic = self.title_logic_var.get()
        descriptions = {
            "basic": "💡 메인키워드 앞 배치 (경쟁사 있으면 참고, 없으면 원본 기반)",
            "usage_main": "💡 [용도/장소] + [메인키워드] + [세부] 구조 (예: 업소용 반죽기 스테인리스)",
        }
        desc = descriptions.get(logic, "")
        self.logic_desc_label.config(text=desc)
    
    def select_all_suspects(self):
        """의심단어 전체 선택"""
        for var in self.suspect_vars.values():
            var.set(True)
    
    def deselect_all_suspects(self):
        """의심단어 전체 해제"""
        for var in self.suspect_vars.values():
            var.set(False)
    
    def update_suspect_list(self, words: list):
        """의심단어 리스트 업데이트"""
        # 기존 체크박스 제거
        for widget in self.suspect_inner_frame.winfo_children():
            widget.destroy()
        self.suspect_vars.clear()
        
        if not words:
            ttk.Label(self.suspect_inner_frame, text="의심단어 없음", foreground="gray").pack(anchor="w")
            self.suspect_status.config(text="의심단어 없음 ✅")
            return
        
        # 중복 제거하고 정렬
        unique_words = sorted(set(words))
        
        # ★ 이미 처리된 단어 제외 (금지단어, 예외단어, 제거단어)
        filtered_words = []
        for word in unique_words:
            if self.filler.banned_words and word in self.filler.banned_words:
                continue
            if EXCLUDED_WORDS and word in EXCLUDED_WORDS:
                continue
            if REMOVE_WORDS and word in REMOVE_WORDS:
                continue
            filtered_words.append(word)
        
        if not filtered_words:
            ttk.Label(self.suspect_inner_frame, text="의심단어 없음 (모두 처리됨)", foreground="gray").pack(anchor="w")
            self.suspect_status.config(text="의심단어 없음 ✅")
            return
        
        for word in filtered_words:
            var = tk.BooleanVar(value=True)  # 기본 체크됨
            self.suspect_vars[word] = var
            # tk.Checkbutton 사용 (ttk는 색상 변경 안 됨)
            cb = tk.Checkbutton(
                self.suspect_inner_frame, 
                text=word, 
                variable=var,
                fg="red",  # 빨간색
                selectcolor="white",  # 체크박스 배경
                anchor="w"
            )
            cb.pack(anchor="w", pady=1, fill="x")
        
        self.suspect_status.config(text=f"총 {len(filtered_words)}개 발견", foreground="red")
    
    def update_suspect_list_with_desc(self, words: list, descriptions: dict):
        """의심단어 리스트 업데이트 (설명 포함)"""
        # 기존 위젯 제거
        for widget in self.suspect_inner_frame.winfo_children():
            widget.destroy()
        self.suspect_vars.clear()
        self.suspect_word_info.clear()  # v2.1: 괄호안 설명 초기화
        self.suspect_descriptions = descriptions.copy()  # v2.1: Claude 분석 결과 저장
        
        if not words:
            ttk.Label(self.suspect_inner_frame, text="의심단어 없음", foreground="gray").pack(anchor="w")
            self.suspect_status.config(text="의심단어 없음 ✅")
            return
        
        # v2.1: 괄호 안 내용을 설명으로 추출
        word_info = {}  # {clean_word: paren_desc}
        for w in words:
            match = re.match(r'^(.+?)\(([^)]+)\)$', w.strip())
            if match:
                clean_word = match.group(1).strip()
                paren_desc = match.group(2).strip()
                word_info[clean_word] = paren_desc
            else:
                word_info[w.strip()] = ""
        
        # v2.1: 설명 정보 저장 (리스트 갱신 시 사용)
        self.suspect_word_info = word_info.copy()
        
        # 중복 제거하고 정렬
        unique_words = sorted(set(word_info.keys()))
        
        # ★ 이미 처리된 단어 제외 (금지단어, 예외단어, 제거단어)
        filtered_words = []
        for word in unique_words:
            # 금지단어에 있는지
            if self.filler.banned_words and word in self.filler.banned_words:
                continue
            # 예외단어에 있는지
            if EXCLUDED_WORDS and word in EXCLUDED_WORDS:
                continue
            # 제거단어에 있는지
            if REMOVE_WORDS and word in REMOVE_WORDS:
                continue
            filtered_words.append(word)
        
        if not filtered_words:
            ttk.Label(self.suspect_inner_frame, text="의심단어 없음 (모두 처리됨)", foreground="gray").pack(anchor="w")
            self.suspect_status.config(text="의심단어 없음 ✅")
            return
        
        for word in filtered_words:
            var = tk.BooleanVar(value=True)  # 기본 체크됨
            self.suspect_vars[word] = var  # 괄호 없는 단어로 저장
            
            # 행 프레임
            row_frame = ttk.Frame(self.suspect_inner_frame)
            row_frame.pack(anchor="w", fill="x", pady=2)
            
            # v2.1: 괄호 안 설명 또는 Claude 분석 결과 사용
            paren_desc = word_info.get(word, "")
            
            # Claude 분석 결과가 있으면 우선 사용
            if word in descriptions:
                desc = descriptions[word]
                judgment = desc.get('judgment', '')
                reason = desc.get('reason', '')
                
                if "브랜드" in judgment:
                    fg_color = "red"
                elif "일반" in judgment:
                    fg_color = "green"
                else:
                    fg_color = "orange"
            else:
                # Claude 분석 없으면 괄호 내용으로 색상 결정
                fg_color = "red"  # 기본값
                judgment = ""
                reason = paren_desc  # 괄호 안 내용을 이유로 사용
                
                # 괄호 내용에 따라 색상 결정
                if paren_desc:
                    if "브랜드" in paren_desc or "상표" in paren_desc or "고유명사" in paren_desc:
                        fg_color = "red"
                    elif "일반" in paren_desc or "용어" in paren_desc:
                        fg_color = "green"
                    else:
                        fg_color = "orange"
            
            # v2.1: 체크박스에는 단어만 표시 (괄호 제거됨)
            cb = tk.Checkbutton(
                row_frame, 
                text=word,  # 단어만 표시
                variable=var,
                fg=fg_color,
                selectcolor="white",
                anchor="w",
                width=15  # 고정 폭
            )
            cb.pack(side="left")
            
            # v2.1: 오른쪽에 설명 표시 (Claude 분석 또는 괄호 내용)
            display_desc = ""
            if judgment:
                display_desc = judgment
                if reason:
                    display_desc += f" - {reason[:20]}..." if len(reason) > 20 else f" - {reason}"
            elif paren_desc:
                display_desc = paren_desc
            
            if display_desc:
                ttk.Label(row_frame, text=display_desc, foreground="gray", font=("", 9)).pack(side="left", padx=(10,0))
        
        # 통계
        brand_count = sum(1 for w in unique_words if w in descriptions and "브랜드" in descriptions[w].get('judgment', ''))
        safe_count = sum(1 for w in unique_words if w in descriptions and "일반" in descriptions[w].get('judgment', ''))
        
        self.suspect_status.config(
            text=f"총 {len(unique_words)}개 (🔴브랜드:{brand_count} 🟢안전:{safe_count})", 
            foreground="red" if brand_count > 0 else "gray"
        )
    
    def _rebuild_suspect_list(self):
        """의심단어 리스트 재구성 (설명 유지) - v2.1"""
        # GUI에서 기존 위젯 제거
        for widget in self.suspect_inner_frame.winfo_children():
            widget.destroy()
        
        # 남은 단어들로 다시 표시
        remaining_words = list(self.suspect_vars.keys())
        
        if not remaining_words:
            ttk.Label(self.suspect_inner_frame, text="의심단어 없음 (모두 처리됨)", foreground="gray").pack(anchor="w")
            self.suspect_status.config(text="의심단어 없음 ✅", foreground="green")
            return
        
        for word in remaining_words:
            var = self.suspect_vars[word]
            
            # 행 프레임
            row_frame = ttk.Frame(self.suspect_inner_frame)
            row_frame.pack(anchor="w", fill="x", pady=2)
            
            # 설명 정보 가져오기
            paren_desc = self.suspect_word_info.get(word, "")
            descriptions = self.suspect_descriptions
            
            # Claude 분석 결과가 있으면 우선 사용
            if word in descriptions:
                desc = descriptions[word]
                judgment = desc.get('judgment', '')
                reason = desc.get('reason', '')
                
                if "브랜드" in judgment:
                    fg_color = "red"
                elif "일반" in judgment:
                    fg_color = "green"
                else:
                    fg_color = "orange"
            else:
                # Claude 분석 없으면 괄호 내용으로 색상 결정
                fg_color = "red"  # 기본값
                judgment = ""
                reason = paren_desc
                
                if paren_desc:
                    if "브랜드" in paren_desc or "상표" in paren_desc or "고유명사" in paren_desc:
                        fg_color = "red"
                    elif "일반" in paren_desc or "용어" in paren_desc:
                        fg_color = "green"
                    else:
                        fg_color = "orange"
            
            # 체크박스
            cb = tk.Checkbutton(
                row_frame, 
                text=word,
                variable=var,
                fg=fg_color,
                selectcolor="white",
                anchor="w",
                width=15
            )
            cb.pack(side="left")
            
            # 오른쪽에 설명 표시
            display_desc = ""
            if judgment:
                display_desc = judgment
                if reason:
                    display_desc += f" - {reason[:20]}..." if len(reason) > 20 else f" - {reason}"
            elif paren_desc:
                display_desc = paren_desc
            
            if display_desc:
                ttk.Label(row_frame, text=display_desc, foreground="gray", font=("", 9)).pack(side="left", padx=(10,0))
        
        self.suspect_status.config(text=f"총 {len(remaining_words)}개 남음", foreground="red")
    
    def add_suspects_to_banned(self):
        """선택된 의심단어를 금지단어에 추가"""
        selected_raw = [word for word, var in self.suspect_vars.items() if var.get()]
        
        if not selected_raw:
            messagebox.showwarning("경고", "추가할 단어를 선택하세요")
            return
        
        # ★ 괄호 안 설명 제거 (예: "렉산(브랜드의심)" → "렉산")
        selected = []
        for word in selected_raw:
            clean_word = re.sub(r'\([^)]*\)$', '', word).strip()
            if clean_word:
                selected.append(clean_word)
        
        if not selected:
            messagebox.showwarning("경고", "유효한 단어가 없습니다")
            return
        
        # 금지단어 파일 로드
        try:
            if os.path.exists(BANNED_WORDS_FILE):
                with open(BANNED_WORDS_FILE, 'r', encoding='utf-8') as f:
                    data = json.load(f)
            else:
                data = {"words": []}
            
            # 기존 단어 수집 (다양한 구조 지원)
            existing = set()
            
            # 구조 1: {"words": [...]}
            if "words" in data and isinstance(data["words"], list):
                existing.update(data["words"])
            
            # 구조 2: {"categories": {...}}
            if "categories" in data and isinstance(data["categories"], dict):
                for cat_val in data["categories"].values():
                    if isinstance(cat_val, list):
                        existing.update(cat_val)
                    elif isinstance(cat_val, dict) and "words" in cat_val:
                        existing.update(cat_val["words"])
            
            # words 키 확보
            if "words" not in data:
                data["words"] = list(existing)
            
            # 새 단어 추가
            added_count = 0
            for word in selected:
                if word not in existing:
                    data["words"].append(word)
                    existing.add(word)
                    added_count += 1
            
            # 저장
            with open(BANNED_WORDS_FILE, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            
            # filler의 금지단어 셋 업데이트
            self.filler.banned_words = existing
            self.filler.banned_words_data = data
            
            # v1.7.1: 선택된 단어 리스트에서 즉시 제거
            for word in selected:
                if word in self.suspect_vars:
                    del self.suspect_vars[word]
                # v2.1: 설명 정보도 함께 삭제
                if word in self.suspect_word_info:
                    del self.suspect_word_info[word]
            
            # v2.1: 공통 함수로 리스트 재구성 (설명 유지)
            self._rebuild_suspect_list()
            
            # 시트에도 자동 추가
            added_words = [w for w in selected if w not in existing or w in data["words"][-added_count:]]
            if added_words:
                self.add_words_to_sheet(added_words, 'banned')
            
            self.log(f"✅ {added_count}개 단어가 금지단어에 추가됨")
            messagebox.showinfo("완료", f"{added_count}개 단어가 금지단어에 추가되었습니다.\n\n추가된 단어:\n{', '.join(selected[:10])}{'...' if len(selected) > 10 else ''}")
            
        except Exception as e:
            self.log(f"❌ 금지단어 추가 실패: {e}")
            messagebox.showerror("오류", f"금지단어 추가 실패: {e}")
    
    def exclude_suspects(self):
        """선택된 의심단어를 탐지 제외 목록에 추가"""
        global EXCLUDED_WORDS
        
        selected_raw = [word for word, var in self.suspect_vars.items() if var.get()]
        
        if not selected_raw:
            messagebox.showwarning("경고", "제외할 단어를 선택하세요")
            return
        
        # ★ 괄호 안 설명 제거
        selected = [re.sub(r'\([^)]*\)$', '', w).strip() for w in selected_raw]
        selected = [w for w in selected if w]
        
        try:
            # 기존 제외 단어 로드
            excluded = load_excluded_words()
            
            added_count = 0
            for word in selected:
                if word not in excluded:
                    excluded.add(word)
                    added_count += 1
            
            # 저장
            if save_excluded_words(excluded):
                # 전역 변수 업데이트
                EXCLUDED_WORDS = excluded
                
                # v1.7.1: 선택된 단어 리스트에서 즉시 제거
                for word in selected:
                    if word in self.suspect_vars:
                        del self.suspect_vars[word]
                    # v2.1: 설명 정보도 함께 삭제
                    if word in self.suspect_word_info:
                        del self.suspect_word_info[word]
                
                # v2.1: 공통 함수로 리스트 재구성 (설명 유지)
                self._rebuild_suspect_list()
                
                # 시트에도 자동 추가
                self.add_words_to_sheet(selected, 'excluded')
                
                self.log(f"✅ {added_count}개 단어가 탐지 제외됨")
                messagebox.showinfo("완료", f"{added_count}개 단어가 탐지 제외되었습니다.\n다음부터 위험단어로 감지하지 않습니다.\n\n제외된 단어:\n{', '.join(selected[:10])}{'...' if len(selected) > 10 else ''}")
            else:
                raise Exception("파일 저장 실패")
            
        except Exception as e:
            self.log(f"❌ 탐지 제외 실패: {e}")
            messagebox.showerror("오류", f"탐지 제외 실패: {e}")
    
    def add_suspects_to_remove(self):
        """선택된 의심단어를 제거단어 목록에 추가 (상품명에서 삭제)"""
        global REMOVE_WORDS
        
        selected_raw = [word for word, var in self.suspect_vars.items() if var.get()]
        
        if not selected_raw:
            messagebox.showwarning("경고", "제거할 단어를 선택하세요")
            return
        
        # ★ 괄호 안 설명 제거
        selected = [re.sub(r'\([^)]*\)$', '', w).strip() for w in selected_raw]
        selected = [w for w in selected if w]
        
        try:
            added_count = 0
            for word in selected:
                if word not in REMOVE_WORDS:
                    REMOVE_WORDS.add(word)
                    added_count += 1
            
            # 저장
            if save_remove_words(REMOVE_WORDS):
                # v1.7.1: 선택된 단어 리스트에서 즉시 제거
                for word in selected:
                    if word in self.suspect_vars:
                        del self.suspect_vars[word]
                    # v2.1: 설명 정보도 함께 삭제
                    if word in self.suspect_word_info:
                        del self.suspect_word_info[word]
                
                # v2.1: 공통 함수로 리스트 재구성 (설명 유지)
                self._rebuild_suspect_list()
                
                # 시트에도 자동 추가
                self.add_words_to_sheet(selected, 'remove')
                
                self.log(f"🗑️ {added_count}개 단어가 제거단어에 추가됨")
                messagebox.showinfo("완료", f"{added_count}개 단어가 제거단어에 추가되었습니다.\n다음부터 상품명에서 자동 삭제됩니다.\n\n추가된 단어:\n{', '.join(selected[:10])}{'...' if len(selected) > 10 else ''}")
            else:
                raise Exception("파일 저장 실패")
            
        except Exception as e:
            self.log(f"❌ 제거단어 추가 실패: {e}")
            messagebox.showerror("오류", f"제거단어 추가 실패: {e}")
    
    def show_banned_words(self):
        """금지단어 관리 창 열기"""
        banned_win = tk.Toplevel(self)
        banned_win.title("🚫 금지단어 관리")
        banned_win.geometry("700x500")
        banned_win.transient(self)
        
        # 상단: 통계
        stat_frame = ttk.Frame(banned_win, padding="10")
        stat_frame.pack(fill="x")
        
        word_count = len(self.filler.banned_words)
        cat_count = len(self.filler.banned_words_data.get('categories', {}))
        ttk.Label(stat_frame, text=f"📊 총 {word_count}개 금지단어 ({cat_count}개 카테고리)", 
                  font=("", 11, "bold")).pack(side="left")
        
        ttk.Button(stat_frame, text="🔄 새로고침", command=lambda: self._refresh_banned_list(tree, stat_label)).pack(side="right")
        stat_label = ttk.Label(stat_frame, text="")
        stat_label.pack(side="right", padx=10)
        
        # 중앙: 카테고리별 단어 목록
        list_frame = ttk.Frame(banned_win, padding="10")
        list_frame.pack(fill="both", expand=True)
        
        # 트리뷰
        columns = ("category", "count", "examples")
        tree = ttk.Treeview(list_frame, columns=columns, show="headings", height=15)
        tree.heading("category", text="카테고리")
        tree.heading("count", text="개수")
        tree.heading("examples", text="예시")
        tree.column("category", width=150)
        tree.column("count", width=60)
        tree.column("examples", width=450)
        
        scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        
        tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # 데이터 채우기
        self._populate_banned_tree(tree)
        
        # 하단: 버튼
        btn_frame = ttk.Frame(banned_win, padding="10")
        btn_frame.pack(fill="x")
        
        ttk.Button(btn_frame, text="📂 파일 열기", 
                   command=lambda: self._open_banned_file()).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="🔍 단어 검색", 
                   command=lambda: self._search_banned_word(banned_win)).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="➕ 단어 추가", 
                   command=lambda: self._add_banned_word(banned_win, tree, stat_label)).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="닫기", command=banned_win.destroy).pack(side="right", padx=5)
    
    def _populate_banned_tree(self, tree):
        """금지단어 트리뷰 채우기"""
        # 기존 데이터 삭제
        for item in tree.get_children():
            tree.delete(item)
        
        categories = self.filler.banned_words_data.get('categories', {})
        for cat_name, cat_data in categories.items():
            words = cat_data.get('words', [])
            desc = cat_data.get('description', cat_name)
            examples = ', '.join(words[:5]) + ('...' if len(words) > 5 else '')
            tree.insert("", "end", values=(desc, len(words), examples))
        
        # AI 감지 단어
        ai_detected = self.filler.banned_words_data.get('ai_detected', {})
        approved = ai_detected.get('approved', [])
        pending = ai_detected.get('pending', [])
        if approved or pending:
            tree.insert("", "end", values=(f"AI 감지 (승인: {len(approved)}, 대기: {len(pending)})", 
                                           len(approved) + len(pending), 
                                           ', '.join(approved[:3] + pending[:3])))
    
    def _refresh_banned_list(self, tree, stat_label):
        """금지단어 새로고침"""
        count = self.filler.reload_banned_words()
        self._populate_banned_tree(tree)
        stat_label.config(text=f"✅ {count}개 로드됨")
        self.log(f"🔄 금지단어 새로고침: {count}개")
    
    def _open_banned_file(self):
        """금지단어 파일 열기"""
        import subprocess
        import platform
        
        if os.path.exists(BANNED_WORDS_FILE):
            if platform.system() == 'Windows':
                os.startfile(BANNED_WORDS_FILE)
            elif platform.system() == 'Darwin':
                subprocess.call(['open', BANNED_WORDS_FILE])
            else:
                subprocess.call(['xdg-open', BANNED_WORDS_FILE])
        else:
            messagebox.showwarning("경고", f"파일이 없습니다: {BANNED_WORDS_FILE}")
    
    def _search_banned_word(self, parent):
        """금지단어 검색"""
        search_win = tk.Toplevel(parent)
        search_win.title("🔍 금지단어 검색")
        search_win.geometry("400x300")
        search_win.transient(parent)
        
        ttk.Label(search_win, text="검색어:").pack(pady=5)
        search_var = tk.StringVar()
        search_entry = ttk.Entry(search_win, textvariable=search_var, width=30)
        search_entry.pack(pady=5)
        search_entry.focus()
        
        result_text = scrolledtext.ScrolledText(search_win, height=12, width=45)
        result_text.pack(pady=10, padx=10, fill="both", expand=True)
        
        def do_search(event=None):
            query = search_var.get().strip().lower()
            if not query:
                return
            result_text.delete("1.0", tk.END)
            found = [w for w in self.filler.banned_words if query in w.lower()]
            if found:
                result_text.insert("1.0", f"'{query}' 포함 단어 {len(found)}개:\n\n")
                result_text.insert(tk.END, '\n'.join(found[:50]))
                if len(found) > 50:
                    result_text.insert(tk.END, f"\n\n... 외 {len(found)-50}개")
            else:
                result_text.insert("1.0", f"'{query}'를 포함한 금지단어가 없습니다.")
        
        search_entry.bind('<Return>', do_search)
        ttk.Button(search_win, text="검색", command=do_search).pack(pady=5)
    
    def _add_banned_word(self, parent, tree, stat_label):
        """금지단어/제거단어 추가"""
        add_win = tk.Toplevel(parent)
        add_win.title("➕ 단어 추가")
        add_win.geometry("420x280")
        add_win.transient(parent)
        
        # 단어 타입 선택 (금지단어/제거단어)
        type_frame = ttk.Frame(add_win)
        type_frame.pack(pady=10)
        ttk.Label(type_frame, text="타입:").pack(side="left")
        word_type_var = tk.StringVar(value="banned")
        ttk.Radiobutton(type_frame, text="🚫 금지단어 (2차검수)", variable=word_type_var, value="banned").pack(side="left", padx=5)
        ttk.Radiobutton(type_frame, text="🗑️ 제거단어 (상품명에서 삭제)", variable=word_type_var, value="remove").pack(side="left", padx=5)
        
        ttk.Label(add_win, text="추가할 단어 (쉼표로 구분):").pack(pady=5)
        word_var = tk.StringVar()
        word_entry = ttk.Entry(add_win, textvariable=word_var, width=45)
        word_entry.pack(pady=5)
        word_entry.focus()
        
        # 카테고리 (금지단어용)
        cat_frame = ttk.Frame(add_win)
        cat_frame.pack(pady=5)
        ttk.Label(cat_frame, text="카테고리 (금지단어용):").pack(side="left")
        
        # categories가 없으면 기본 구조 생성
        if 'categories' not in self.filler.banned_words_data:
            self.filler.banned_words_data['categories'] = {
                'stores': {'name': '브랜드/스토어', 'words': []},
                'adult': {'name': '성인용품', 'words': []},
                'medical': {'name': '의료/약품', 'words': []},
                'child': {'name': '아동', 'words': []},
                'prohibited': {'name': '금지품목', 'words': []},
                'brand': {'name': '브랜드명', 'words': []},
                'etc': {'name': '기타', 'words': []}
            }
        
        cat_var = tk.StringVar(value="stores")
        categories = list(self.filler.banned_words_data.get('categories', {}).keys())
        if not categories:
            categories = ['stores', 'brand', 'etc']
        cat_combo = ttk.Combobox(cat_frame, textvariable=cat_var, values=categories, width=20)
        cat_combo.pack(side="left", padx=5)
        
        def do_add():
            words = [w.strip() for w in word_var.get().split(',') if w.strip()]
            word_type = word_type_var.get()
            cat = cat_var.get()
            
            if not words:
                messagebox.showwarning("경고", "단어를 입력하세요")
                return
            
            if word_type == "remove":
                # 제거단어 추가
                global REMOVE_WORDS
                added = []
                for w in words:
                    if w not in REMOVE_WORDS:
                        REMOVE_WORDS.add(w)
                        added.append(w)
                
                if added:
                    save_remove_words(REMOVE_WORDS)
                    stat_label.config(text=f"✅ 제거단어 {len(added)}개 추가됨")
                    self.log(f"🗑️ 제거단어 추가: {', '.join(added)}")
                    
                    # 시트에도 자동 추가
                    self.add_words_to_sheet(added, 'remove')
                    
                    add_win.destroy()
                else:
                    messagebox.showinfo("알림", "이미 존재하는 단어입니다")
            else:
                # 금지단어 추가
                # 카테고리가 없으면 생성
                if cat not in self.filler.banned_words_data.get('categories', {}):
                    self.filler.banned_words_data['categories'][cat] = {'name': cat, 'words': []}
                
                # 추가
                cat_data = self.filler.banned_words_data['categories'][cat]
                if isinstance(cat_data, dict):
                    existing = cat_data.get('words', [])
                    if not isinstance(existing, list):
                        existing = []
                        cat_data['words'] = existing
                else:
                    # 리스트 형태인 경우
                    existing = cat_data if isinstance(cat_data, list) else []
                    self.filler.banned_words_data['categories'][cat] = {'name': cat, 'words': existing}
                    existing = self.filler.banned_words_data['categories'][cat]['words']
                
                added = []
                for w in words:
                    if w not in existing:
                        existing.append(w)
                        self.filler.banned_words.add(w)
                        added.append(w)
                
                if added:
                    save_banned_words(self.filler.banned_words_data)
                    self._populate_banned_tree(tree)
                    stat_label.config(text=f"✅ 금지단어 {len(added)}개 추가됨")
                    self.log(f"➕ 금지단어 추가: {', '.join(added)}")
                    
                    # 시트에도 자동 추가
                    self.add_words_to_sheet(added, 'banned')
                    
                    add_win.destroy()
                else:
                    messagebox.showinfo("알림", "이미 존재하는 단어입니다")
        
        ttk.Button(add_win, text="추가", command=do_add).pack(pady=15)

    def update_progress(self, current: int, total: int, success: int = 0, failed: int = 0):
        if total > 0:
            pct = (current / total) * 100
            self.progress['value'] = pct
            self.progress_label.config(text=f"진행 률:  {current}/{total} ({pct:.1f}%)")
            self.progress_stats.config(text=f"✅ 성공: {success}  ❌ 실패: {failed}")
        self.update_idletasks()
    
    def update_progress_detail(self, text: str):
        """진행 상세 정보 업데이트"""
        self.progress_detail.config(text=text)
        self.update_idletasks()
    
    def reset_progress(self):
        """진행 상황 초기화"""
        self.progress['value'] = 0
        self.progress_label.config(text="진행 률:  0/0 (0.0%)")
        self.progress_detail.config(text="")
        self.progress_stats.config(text="✅ 성공: 0  ❌ 실패: 0")
    
    def save_settings(self):
        """설정 저장"""
        api_key = self.api_var.get().strip()
        
        # Claude API 키 검증 (있을 경우만)
        if api_key and not api_key.startswith('sk-ant-'):
            messagebox.showwarning("경고", "올바른 Claude API Key가 아닙니다")
            return
        
        self.config['api_key'] = api_key
        self.config['model'] = self.model_var.get()
        self.config['vision_model'] = self.vision_model_var.get()
        self.config['title_mode'] = self.title_mode_var.get()
        self.config['port'] = self.port_var.get()
        self.config['temperature'] = self.temp_var.get()
        self.config['sheets_url'] = self.sheets_url_var.get().strip()
        
        # ★ v3.0: 토큰 저장
        self.config['access_token'] = self.access_token_var.get().strip()
        self.config['refresh_token'] = self.refresh_token_var.get().strip()
        
        # 그룹 정보 저장
        self.config['work_groups'] = self.groups_var.get().strip()
        self.config['market_groups'] = self.market_groups_var.get().strip()
        self.config['copy_groups'] = self.copy_groups_var.get().strip()
        self.config['copy_count'] = self.copy_count_var.get().strip()
        self.config['tag_name'] = self.tag_var.get().strip()
        self.config['danger_tag'] = self.danger_tag_var.get().strip()
        
        if save_config(self.config):
            self.api_status.config(text="✅ 저장됨", foreground="green")
            self.log("✅ 설정이 저장되었습니다")
            messagebox.showinfo("완료", "설정이 저장되었습니다.")
        else:
            self.api_status.config(text="❌ 저장 실패", foreground="red")
    
    def refresh_tag_time(self):
        """태그 날짜시간 새로고침"""
        new_tag = f"작업완료_{datetime.now().strftime('%y%m%d_%H%M')}"
        self.tag_var.set(new_tag)
        self.log(f"🔄 태그 갱신: {new_tag}")
    
    def sync_words_from_sheets(self):
        """Google Sheets에서 금지단어/예외단어/제거단어 동기화"""
        sheet_url = self.sheets_url_var.get().strip()
        
        if not sheet_url:
            messagebox.showwarning("경고", "Google Sheets URL을 입력하세요")
            return
        
        self.sheets_status.config(text="동기화 중...", foreground="blue")
        self.update()
        
        def sync_task():
            result = sync_from_google_sheets(sheet_url, self.log)
            
            if result['success']:
                # ⚠️ 시트가 비어있으면 덮어쓰기 방지
                if len(result['banned']) == 0 and len(result['excluded']) == 0 and len(result.get('remove', [])) == 0:
                    self.sheets_status.config(text="⚠️ 시트 비어있음", foreground="orange")
                    self.log("⚠️ 시트에 데이터가 없습니다. 로컬 파일 유지됨.")
                    return
                
                # 금지단어 저장
                if result['banned']:
                    banned_data = {'words': result['banned']}
                    save_banned_words(banned_data)
                    self.filler.banned_words = set(result['banned'])
                    self.filler.banned_words_data = banned_data
                
                # 예외단어 저장
                if result['excluded']:
                    save_excluded_words(set(result['excluded']))
                    global EXCLUDED_WORDS
                    EXCLUDED_WORDS = set(result['excluded'])
                
                # 제거단어 저장
                if result.get('remove'):
                    save_remove_words(set(result['remove']))
                    global REMOVE_WORDS
                    REMOVE_WORDS = set(result['remove'])
                    self.filler.remove_words = set(result['remove'])
                
                # 설정에 URL 저장
                self.config['sheets_url'] = sheet_url
                save_config(self.config)
                
                self.sheets_status.config(text="✅ 완료", foreground="green")
                remove_count = len(result.get('remove', []))
                self.log(f"✅ 동기화 완료: 금지단어 {len(result['banned'])}개, 예외단어 {len(result['excluded'])}개, 제거단어 {remove_count}개")
            else:
                self.sheets_status.config(text="❌ 실패", foreground="red")
        
        threading.Thread(target=sync_task, daemon=True).start()
    
    def upload_words_to_sheets(self):
        """로컬 금지단어/예외단어/제거단어를 Google Sheets에 업로드"""
        sheet_url = self.sheets_url_var.get().strip()
        
        if not sheet_url:
            messagebox.showwarning("경고", "Google Sheets URL을 입력하세요")
            return
        
        # 현재 로컬 데이터 확인
        banned_words = list(self.filler.banned_words) if self.filler.banned_words else []
        excluded_words = list(EXCLUDED_WORDS) if EXCLUDED_WORDS else []
        remove_words = list(REMOVE_WORDS) if REMOVE_WORDS else []
        
        if not banned_words and not excluded_words and not remove_words:
            messagebox.showwarning("경고", "업로드할 로컬 데이터가 없습니다")
            return
        
        msg = f"로컬 데이터를 시트에 업로드합니다:\n\n"
        msg += f"• 금지단어: {len(banned_words)}개\n"
        msg += f"• 예외단어: {len(excluded_words)}개\n"
        msg += f"• 제거단어: {len(remove_words)}개\n\n"
        msg += "기존 시트 데이터는 덮어씌워집니다. 계속할까요?"
        
        if not messagebox.askyesno("확인", msg):
            return
        
        self.sheets_status.config(text="업로드 중...", foreground="blue")
        self.update()
        
        def upload_task():
            try:
                import gspread
                from google.oauth2.service_account import Credentials
                
                # 서비스 계정 파일 찾기
                service_account_file = None
                for filename in os.listdir('.'):
                    if filename.endswith('.json') and 'auto-smartstore' in filename.lower():
                        service_account_file = filename
                        break
                
                if not service_account_file:
                    self.log("❌ 서비스 계정 JSON 파일을 찾을 수 없습니다")
                    self.sheets_status.config(text="❌ 실패", foreground="red")
                    return
                
                self.log(f"📤 시트 업로드 중... ({service_account_file})")
                
                # 인증
                scopes = ['https://www.googleapis.com/auth/spreadsheets', 'https://www.googleapis.com/auth/drive']
                creds = Credentials.from_service_account_file(service_account_file, scopes=scopes)
                gc = gspread.authorize(creds)
                
                # Sheet ID 추출
                if '/d/' in sheet_url:
                    sheet_id = sheet_url.split('/d/')[1].split('/')[0]
                else:
                    sheet_id = sheet_url
                
                # 시트 열기
                spreadsheet = gc.open_by_key(sheet_id)
                try:
                    worksheet = spreadsheet.worksheet('bulsaja_words')
                except:
                    worksheet = spreadsheet.sheet1
                
                # 시트 초기화
                worksheet.clear()
                
                # 헤더 작성
                worksheet.update_cell(1, 1, '금지단어')
                worksheet.update_cell(1, 2, '예외단어')
                worksheet.update_cell(1, 3, '제거단어')
                
                # 데이터 작성 (최대 길이 맞추기)
                max_len = max(len(banned_words), len(excluded_words), len(remove_words))
                
                if max_len > 0:
                    # 데이터 준비
                    data = []
                    for i in range(max_len):
                        row = [
                            banned_words[i] if i < len(banned_words) else '',
                            excluded_words[i] if i < len(excluded_words) else '',
                            remove_words[i] if i < len(remove_words) else ''
                        ]
                        data.append(row)
                    
                    # 일괄 업데이트 (2행부터)
                    worksheet.update(f'A2:C{max_len + 1}', data)
                
                self.sheets_status.config(text="✅ 업로드 완료", foreground="green")
                self.log(f"✅ 업로드 완료: 금지단어 {len(banned_words)}개, 예외단어 {len(excluded_words)}개, 제거단어 {len(remove_words)}개")
                
            except Exception as e:
                self.log(f"❌ 업로드 실패: {e}")
                self.sheets_status.config(text="❌ 실패", foreground="red")
        
        threading.Thread(target=upload_task, daemon=True).start()
    
    def add_words_to_sheet(self, words: list, word_type: str = 'banned'):
        """시트에 단어 추가 (banned: A열, excluded: B열)"""
        sheet_url = self.sheets_url_var.get().strip()
        if not sheet_url or not words:
            return
        
        def add_task():
            try:
                import gspread
                from google.oauth2.service_account import Credentials
                
                # 서비스 계정 파일 찾기
                service_account_file = None
                for filename in os.listdir('.'):
                    if filename.endswith('.json') and 'auto-smartstore' in filename.lower():
                        service_account_file = filename
                        break
                
                if not service_account_file:
                    return
                
                # 인증
                scopes = ['https://www.googleapis.com/auth/spreadsheets', 'https://www.googleapis.com/auth/drive']
                creds = Credentials.from_service_account_file(service_account_file, scopes=scopes)
                gc = gspread.authorize(creds)
                
                # Sheet ID 추출
                if '/d/' in sheet_url:
                    sheet_id = sheet_url.split('/d/')[1].split('/')[0]
                else:
                    sheet_id = sheet_url
                
                # 시트 열기
                spreadsheet = gc.open_by_key(sheet_id)
                try:
                    worksheet = spreadsheet.worksheet('bulsaja_words')
                except:
                    worksheet = spreadsheet.sheet1
                
                # 현재 데이터 가져오기
                if word_type == 'banned':
                    col = 1  # A열
                elif word_type == 'excluded':
                    col = 2  # B열
                else:  # remove
                    col = 3  # C열
                
                current_values = worksheet.col_values(col)
                
                # 마지막 행 찾기 (헤더 제외)
                last_row = len(current_values) + 1 if current_values else 2
                
                # 필요한 행 수 확인 및 추가
                needed_rows = last_row + len(words)
                current_row_count = worksheet.row_count
                if needed_rows > current_row_count:
                    # 행 추가 (여유분 100개 추가)
                    worksheet.add_rows(needed_rows - current_row_count + 100)
                
                # 새 단어들 추가
                for i, word in enumerate(words):
                    worksheet.update_cell(last_row + i, col, word)
                
                type_name = {'banned': '금지단어', 'excluded': '예외단어', 'remove': '제거단어'}.get(word_type, word_type)
                self.log(f"☁️ 시트 동기화 ({type_name}): {', '.join(words)}")
                
            except Exception as e:
                self.log(f"⚠️ 시트 추가 실패: {e}")
        
        threading.Thread(target=add_task, daemon=True).start()
    
    def create_tag_manual(self):
        """태그 생성 버튼 클릭 시"""
        if not self.filler.main_driver:
            messagebox.showwarning("경고", "먼저 크롬에 연결하세요")
            return
        
        tag_name = self.tag_var.get().strip()
        if not tag_name:
            messagebox.showwarning("경고", "태그명을 입력하세요")
            return
        
        self.log(f"🏷️ 태그 생성 시도: {tag_name}")
        
        # 별도 스레드에서 실행
        def create_task():
            success = self.filler.create_tag(tag_name)
            if success:
                self.log(f"✅ 태그 '{tag_name}' 생성 완료!")
            else:
                self.log(f"❌ 태그 생성 실패")
        
        threading.Thread(target=create_task, daemon=True).start()
    
    def connect_chrome(self):
        """크롬 연결 - API 모드에서는 사용 안 함 (레거시)"""
        pass
    
    def _connect_chrome_thread(self):
        """크롬 연결 스레드"""
        api_key = self.api_var.get().strip()
        
        # 초기화
        self.log("🔧 초기화 중...")
        
        if not self.filler.setup_drivers():
            self._on_connect_failed()
            return
        
        if not self.filler.setup_claude(api_key):
            self._on_connect_failed()
            return
        
        self.log("✅ 초기화 완료")
        
        # 연결 방식에 따라 처리
        connect_mode = self.connect_mode_var.get()
        
        if connect_mode == "debug":
            try:
                port = int(self.port_var.get())
            except:
                port = DEBUG_PORT
            
            # 디버깅 크롬 자동 실행
            if not self.filler.launch_debug_chrome(port):
                self._on_connect_failed()
                return
            
            # 연결 시도
            self.log("🔗 크롬에 연결 시도...")
            
            # 연결 재시도 (최대 10번, 더 여유있게)
            connected = False
            for attempt in range(10):
                if self.filler.connect_to_existing_chrome(port):
                    connected = True
                    break
                self.log(f"⏳ 연결 대기 중... ({attempt + 1}/10)")
                time.sleep(1)
            
            if connected:
                self.after(0, self._on_connect_success)
            else:
                self.log("❌ 크롬 연결 실패")
                self.log("💡 크롬을 수동으로 닫고 다시 시도해보세요")
                self._on_connect_failed()
        else:
            # 새 브라우저 열기
            url = self.url_var.get().strip()
            if not url:
                self.after(0, lambda: messagebox.showwarning("경고", "불사자 URL을 입력하세요"))
                self._on_connect_failed()
                return
            
            self.log("🌐 새 브라우저 열기...")
            if self.filler.connect_to_bulsaja(url):
                self.after(0, self._on_connect_success_new)
            else:
                self._on_connect_failed()
    
    def _on_connect_success(self):
        """연결 성공 (레거시)"""
        self.btn_start.config(state="normal")
        self.log("✅ 연결 성공!")
    
    def _on_connect_success_new(self):
        """새 브라우저 연결 성공 (레거시)"""
        self.btn_start.config(state="normal")
        self.log("✅ 연결 성공!")
    
    def _on_connect_failed(self):
        """연결 실패 (레거시)"""
        self.log("❌ 연결 실패")
    
    def start_automation(self):
        """자동화 시작"""
        # ★ v3.0: API 연결 체크
        if not self.filler.api_client:
            messagebox.showwarning("경고", "먼저 불사자 API에 연결하세요\n(🔑 토큰 가져오기 → 🔗 연결)")
            return
        
        # Claude API 체크
        api_key = self.api_var.get().strip()
        if not api_key or not api_key.startswith('sk-ant-'):
            messagebox.showwarning("경고", "Claude API Key를 입력하세요")
            return
        
        # Claude 클라이언트 초기화
        if not self.filler.claude_client:
            self.filler.setup_claude(api_key)
        
        try:
            start_idx = int(self.start_var.get())
            count = int(self.count_var.get())
        except ValueError:
            messagebox.showerror("오류", "숫자를 입력하세요")
            return
        
        if count <= 0:
            messagebox.showwarning("경고", "처리 개수는 1 이상이어야 합니다")
            return
        
        # 페이지 수 자동 계산 (페이지당 1000개 기준) - 항상 자동 계산
        num_pages = max(1, (count + 999) // 1000)  # 올림 계산
        if num_pages > 1:
            self.log(f"📄 페이지 수 자동 계산: {num_pages} (1000개/페이지 기준)")
        
        # 그룹 리스트 파싱 (숫자 맵핑 지원)
        groups_str = self.groups_var.get().strip()
        if groups_str:
            # 마켓 그룹 목록 가져오기
            market_groups_str = self.market_groups_var.get().strip()
            market_groups_list = [g.strip() for g in market_groups_str.split(',') if g.strip()]
            
            # 숫자 맵핑 적용
            groups = self.resolve_group_input(groups_str, market_groups_list)
        else:
            groups = None  # 현재 그룹만
        
        # 생성 방식 확인
        title_mode = self.title_mode_var.get()
        if title_mode == TITLE_MODE_IMAGE_FIRST:
            mode_name = "이미지+기존상품명"
        elif title_mode == TITLE_MODE_VISION:
            mode_name = "Vision분석"
        else:
            mode_name = "기존상품명만"
        
        # 모델 이름 추출
        model_str = self.model_var.get()
        if 'haiku' in model_str:
            model_name = "Haiku"
        elif 'sonnet' in model_str:
            model_name = "Sonnet"
        elif 'opus' in model_str:
            model_name = "Opus"
        else:
            model_name = model_str
        
        msg = f"다음 설정으로 자동화를 시작합니다:\n\n"
        msg += f"• 시작 번호: {start_idx}\n"
        msg += f"• 처리 개수: {count} (그룹당)\n"
        msg += f"• 페이지 수: {num_pages}\n"
        if groups:
            msg += f"• 그룹: {', '.join(groups)} ({len(groups)}개)\n"
        else:
            msg += f"• 그룹: 현재 선택된 그룹\n"
        msg += f"• 모델: {model_name}\n"
        msg += f"• 생성 방식: {mode_name}\n"
        msg += f"• 다양성: {self.temp_var.get()}\n"
        
        if not messagebox.askyesno("확인", msg):
            return
        
        self.btn_start.config(state="disabled")
        self.btn_stop.config(state="normal")
        
        self.filler.is_running = True
        self.worker_thread = threading.Thread(
            target=self.filler.process_products,
            args=(start_idx, count, num_pages, groups),
            daemon=True
        )
        self.worker_thread.start()
    
    def stop(self):
        """처리 중지"""
        self.filler.is_running = False
        self.log("🛑 중지 요청...")
        # AliPrice 창 닫기
        self.filler.close_aliprice()
    
    def on_finished(self):
        """처리 완료 후"""
        self.filler.close_aliprice()  # AliPrice 브라우저 닫기
        self.btn_start.config(state="normal")
        self.btn_stop.config(state="disabled")
        self.btn_copy.config(state="normal")
        self.update_progress_detail("완료")
    
    def on_copy_finished(self):
        """복사 작업 완료 후"""
        self.btn_start.config(state="normal")
        self.btn_stop.config(state="disabled")
        self.btn_copy.config(state="normal")
        self.update_progress_detail("복사 완료")
    
    def resolve_group_input(self, input_str: str, market_groups_list: List[str]) -> List[str]:
        """그룹 입력 파싱 - 숫자, 범위, 풀네임 모두 지원
        
        입력 예시:
        - "1" or "01" -> 첫번째 마켓
        - "3-6" -> 3~6번째 마켓들
        - "3,6" -> 3번째, 6번째 마켓
        - "02_마켓B" -> 그대로 사용
        - "1-3,5,02_마켓B" -> 복합 사용
        """
        result = []
        
        # 쉼표로 분리
        parts = [p.strip() for p in input_str.split(',') if p.strip()]
        
        for part in parts:
            # 범위 표현인지 확인 (예: 3-6)
            if '-' in part and part.count('-') == 1:
                left, right = part.split('-')
                left = left.strip()
                right = right.strip()
                
                # 양쪽이 모두 숫자인 경우만 범위로 처리
                if left.isdigit() and right.isdigit():
                    start_idx = int(left) - 1  # 1-based -> 0-based
                    end_idx = int(right) - 1
                    
                    if start_idx < 0:
                        start_idx = 0
                    if end_idx >= len(market_groups_list):
                        end_idx = len(market_groups_list) - 1
                    
                    for i in range(start_idx, end_idx + 1):
                        if i < len(market_groups_list):
                            result.append(market_groups_list[i])
                else:
                    # 숫자가 아니면 풀네임으로 처리 (예: 02_스트롬-브린)
                    result.append(part)
            
            # 숫자인지 확인 (예: 1, 01, 3)
            elif part.isdigit():
                idx = int(part) - 1  # 1-based -> 0-based
                if 0 <= idx < len(market_groups_list):
                    result.append(market_groups_list[idx])
                else:
                    self.log(f"⚠️ 인덱스 {part} 범위 초과 (마켓 그룹 {len(market_groups_list)}개)")
            
            # 풀네임으로 처리
            else:
                result.append(part)
        
        return result
    
    def start_copy(self):
        """상품 복사 시작 - Mode2"""
        if not self.filler.main_driver:
            messagebox.showwarning("경고", "먼저 크롬에 연결하세요")
            return
        
        # 태그 설정 (사용자 입력 그대로 사용)
        search_tag = self.copy_tag_var.get().strip()
        done_tag = self.copy_done_tag_var.get().strip()
        
        copy_groups_str = self.copy_groups_var.get().strip()
        if not copy_groups_str:
            messagebox.showwarning("경고", "복사 그룹을 입력하세요\n예: 02_마켓B, 03_마켓C 또는 1-5, 3,6")
            return
        
        # 복사 수량
        try:
            copy_count = int(self.copy_count_var.get().strip() or "100")
        except ValueError:
            copy_count = 100
        
        # 마켓 그룹 목록 가져오기
        market_groups_str = self.market_groups_var.get().strip()
        market_groups_list = [g.strip() for g in market_groups_str.split(',') if g.strip()]
        
        # 복사 그룹 파싱 (숫자 맵핑 지원)
        group_names = self.resolve_group_input(copy_groups_str, market_groups_list)
        
        if not group_names:
            messagebox.showwarning("경고", "유효한 복사 그룹이 없습니다")
            return
        
        copy_groups = [(name, copy_count) for name in group_names]
        
        # 작업 그룹 가져오기
        work_group = self.groups_var.get().strip()
        
        # 확인 메시지
        msg = "📋 Mode2: 복사 후 태그 변경\n\n"
        if work_group:
            msg += f"📁 작업 그룹: {work_group}\n"
        msg += f"🏷️ 검색 태그: {search_tag or '(없음)'}\n"
        msg += f"🏷️ 완료 태그: {done_tag or '(없음)'}\n"
        msg += f"📦 그룹당 수량: {copy_count}개\n"
        msg += f"📁 복사 그룹: {len(copy_groups)}개\n"
        
        # 그룹 이름 표시 (최대 5개)
        if len(group_names) <= 5:
            msg += f"   {', '.join(group_names)}\n\n"
        else:
            msg += f"   {', '.join(group_names[:5])}... 외 {len(group_names)-5}개\n\n"
        
        msg += "진행 순서:\n"
        step = 1
        if work_group:
            msg += f"{step}. 작업 그룹 '{work_group}' 선택\n"
            step += 1
        msg += f"{step}. 태그 '{search_tag}' 검색\n"
        step += 1
        msg += f"{step}. 페이지 크기 {copy_count} 설정\n"
        step += 1
        msg += f"{step}. 전체 선택 → 복사\n"
        step += 1
        if done_tag:
            msg += f"{step}. 선택 상품에 '{done_tag}' 태그 추가\n"
            step += 1
            msg += f"{step}. 새로고침 → 다음 그룹\n"
        
        msg += "\n진행하시겠습니까?"
        
        if not messagebox.askyesno("상품 복사 확인", msg):
            return
        
        # 버튼 상태 변경
        self.btn_start.config(state="disabled")
        self.btn_copy.config(state="disabled")
        self.btn_stop.config(state="normal")
        
        self.filler.is_running = True
        
        # 스레드로 복사 실행
        threading.Thread(
            target=self.filler.process_copy_groups,
            args=(copy_groups, search_tag, done_tag, work_group),
            daemon=True
        ).start()

if __name__ == "__main__":
    app = App()
    app.mainloop()
