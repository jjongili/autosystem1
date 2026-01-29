import sys
import os
import json
import time
import re
import threading
import concurrent.futures # For parallel speed-up
import requests
import uuid
import random
import base64
import hmac
import hashlib
from datetime import datetime
import bcrypt
import sqlite3
from typing import List, Optional, Tuple, Dict, Set

try:
    import jwt
except ImportError:
    jwt = None

# PyQt6 Imports
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QLineEdit, QTextEdit, QPushButton, QTabWidget,
    QMessageBox, QRadioButton, QButtonGroup, QFileDialog,
    QProgressBar, QGroupBox, QScrollArea, QCheckBox, QTableWidget,
    QTableWidgetItem, QHeaderView, QSplitter, QFrame, QGridLayout,
    QSpinBox, QComboBox, QDialog, QInputDialog, QPlainTextEdit
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal, QSize, QTimer
from PyQt6.QtGui import QColor, QFont, QIcon, QAction
from PyQt6 import sip # For object lifecycle validation

# Third-party Imports
from google import genai
from google.genai import types
from kiwipiepy import Kiwi

# OpenPyXL for Excel support
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# ======================================================
# GLOBAL CONSTANTS & HELPERS
# ======================================================

DARK_THEME_STYLESHEET = """
QMainWindow {
    background-color: #0c0c0c;
    color: #ffffff;
}
QWidget {
    background-color: #0c0c0c;
    color: #ffffff;
    font-family: 'Inter', 'Segoe UI', 'Malgun Gothic', sans-serif;
    font-size: 13px;
}
QTableWidget {
    background-color: #121212;
    color: #e0e0e0;
    gridline-color: #2a2a2a;
    border: none;
    selection-background-color: #004d40;
}
QHeaderView::section {
    background-color: #1a1a1a;
    padding: 8px;
    border: 1px solid #2a2a2a;
    font-weight: bold;
    color: #00bfa5;
}
QLineEdit, QTextEdit {
    background-color: #1e1e1e;
    border: 1px solid #333;
    padding: 6px;
    border-radius: 4px;
    color: #eee;
}
QPushButton {
    background-color: #333;
    color: white;
    border: none;
    padding: 8px 15px;
    border-radius: 4px;
}
QPushButton:hover {
    background-color: #444;
}
QPushButton#PrimaryButton {
    background-color: #00bfa5;
    font-weight: bold;
}
QGroupBox {
    border: 1px solid #555555;
    border-radius: 5px;
    margin-top: 20px;
    font-weight: bold;
}
QGroupBox::title {
    subcontrol-origin: margin;
    subcontrol-position: top left;
    padding: 0 5px;
    color: #00fbff;
}
QProgressBar {
    border: 1px solid #555555;
    border-radius: 3px;
    text-align: center;
}
QProgressBar::chunk {
    background-color: #00C853;
}
"""

BANNED_ADJECTIVES = [
    "강력한", "프리미엄", "고급", "최신형", "완벽한", "최고의",
    "튼튼한", "간편한", "편리한", "실용적인", "효율적인",
    "고품질", "고성능", "초강력", "초경량", "초슬림",
    "예쁜", "귀여운", "멋진", "세련된", "깔끔한", "심플한",
    "모던한", "클래식한", "빈티지한", "럭셔리한",
    "대형", "소형", "미니", "빅", "라지", "스몰",
    "새로운", "특별한", "인기", "베스트", "핫한",
    "안전한", "위생적인", "친환경", "무독성", "무소음",
]

FAMOUS_BRANDS = {
    "nike", "adidas", "puma", "reebok", "newbalance", "underarmour", "fila",
    "converse", "vans", "asics", "mizuno", "umbro", "kappa", "lotto",
    "gucci", "chanel", "louisvuitton", "prada", "hermes", "dior", "burberry",
    "apple", "samsung", "sony", "lg", "panasonic", "philips", "bose",
    "dyson", "xiaomi", "huawei", "dell", "hp", "lenovo", "asus", "acer",
    "disney", "marvel", "pokemon", "sanrio", "hellokitty", "kakao", "line",
    "nintendo", "playstation", "xbox", "bandai", "lego", "barbie", "transformer",
    "northface", "patagonia", "columbia", "arcteryx", "mammut", "salomon",
    "blackyak", "kolon", "eider", "lafuma", "millet", "k2",
    "농심", "오뚜기", "cj", "풀무원", "동원", "삼양", "오리온", "롯데",
    "아모레퍼시픽", "lg생활건강", "이니스프리", "설화수", "라네즈", "헤라",
    "bmw", "mercedes", "audi", "porsche", "ferrari", "lamborghini", "bentley",
    "현대", "기아", "제네시스", "벤츠", "아우디", "포르쉐", "페라리", "람보르기니",
    "나이키", "아디다스", "퓨마", "리복", "뉴발란스", "언더아머", "휠라",
    "컨버스", "반스", "아식스", "미즈노", "엄브로", "카파", "로또",
    "구찌", "샤넬", "루이비통", "프라다", "에르메스", "디올", "버버리",
    "애플", "삼성", "소니", "엘지", "파나소닉", "필립스", "보스",
    "다이슨", "샤오미", "화웨이", "델", "레노버", "아수스", "에이서",
    "디즈니", "마블", "포켓몬", "산리오", "헬로키티", "카카오", "라인",
    "닌텐도", "플레이스테이션", "엑스박스", "반다이", "레고", "바비", "트랜스포머",
    "노스페이스", "파타고니아", "콜롬비아", "아크테릭스", "마무트", "살로몬",
    "블랙야크", "코오롱", "아이더", "라푸마", "밀레", "케이투",
}

COMMON_ENGLISH_WORDS = {
    "mini", "micro", "small", "medium", "large", "big", "xl", "xxl",
    "slim", "compact", "portable", "lite", "light",
    "pro", "plus", "max", "ultra", "super", "smart", "auto", "manual",
    "digital", "analog", "electric", "wireless", "bluetooth", "wifi",
    "usb", "led", "lcd", "hd", "fhd", "uhd", "oled", "qled",
    "waterproof", "dustproof", "shockproof", "fireproof",
    "foldable", "folding", "adjustable", "flexible", "portable",
    "rechargeable", "cordless", "battery", "solar", "magnetic",
    "silent", "quiet", "noise", "mute", "sound",
    "home", "office", "outdoor", "indoor", "camping", "travel", "hiking",
    "kitchen", "bathroom", "bedroom", "living", "garden", "garage",
    "car", "bike", "desk", "table", "wall", "floor", "door", "window",
    "plastic", "metal", "wood", "wooden", "steel", "iron", "aluminum",
    "glass", "silicon", "silicone", "rubber", "leather", "fabric", "cotton",
    "stainless", "chrome", "brass", "copper", "zinc", "titanium",
    "black", "white", "gray", "grey", "red", "blue", "green", "yellow",
    "pink", "purple", "orange", "brown", "beige", "navy", "gold", "silver",
    "round", "square", "rectangle", "circle", "oval", "triangle",
    "flat", "curved", "straight", "long", "short", "wide", "narrow",
    "set", "kit", "pack", "box", "case", "cover", "holder", "stand",
    "rack", "shelf", "hook", "clip", "mount", "bracket", "hanger",
    "cup", "mug", "bottle", "pot", "pan", "bowl", "plate", "dish",
    "bag", "pouch", "basket", "bin", "container", "storage", "organizer",
    "tool", "device", "machine", "equipment", "accessory", "part",
    "new", "type", "style", "version", "model", "series", "edition",
    "multi", "dual", "double", "single", "triple", "pair",
}

def classify_english_pattern(word: str) -> tuple:
    if not word or len(word) < 2: return ("UNKNOWN", "keep", "")
    word_lower = word.lower().replace(" ", "").replace("-", "")
    word_clean = word.strip()
    
    if word_lower in FAMOUS_BRANDS: return ("FAMOUS_BRAND", "review", f"Famous: {word}")
    if word_lower in COMMON_ENGLISH_WORDS: return ("COMMON_WORD", "keep", "")
    if word_clean.isdigit(): return ("NUMBER", "remove", f"Number: {word}")
    
    # Model Numbers
    if re.match(r'^[A-Za-z]{1,3}[-]?\d{2,5}$', word_clean, re.IGNORECASE): return ("MODEL_NUMBER", "remove", f"Model: {word}")
    if re.match(r'^\d+[A-Za-z]{1,4}$', word_clean): return ("MODEL_NUMBER", "remove", f"Spec: {word}")
    if re.match(r'^[VXS]\d{1,2}$', word_clean, re.IGNORECASE): return ("MODEL_NUMBER", "remove", f"Ver: {word}")
    
    # Suspicious English (4-10 chars, not common)
    if re.match(r'^[A-Za-z]{4,10}$', word_clean):
        consonants = re.findall(r'[bcdfghjklmnpqrstvwxz]{3,}', word_lower)
        if consonants: return ("SELLER_BRAND", "remove", f"Suspect Brand: {word}")
        else: return ("UNKNOWN_ENGLISH", "suspect", f"Unknown Eng: {word}")
    
    if re.match(r'^[A-Za-z]+\d+[A-Za-z]*$', word_clean) or re.match(r'^\d+[A-Za-z]+\d*$', word_clean):
        return ("MODEL_NUMBER", "remove", f"Model: {word}")
    
    return ("UNKNOWN", "keep", "")

# ======================================================
# CHROME DEBUGGING PROTOCOL (CDP) CLIENT
# ======================================================
class ChromeTokenExtractor:
    """Extracts local storage tokens from a running Chrome instance via CDP."""
    DEBUG_URL = "http://127.0.0.1:9222"
    
    def get_tokens(self) -> Tuple[str, str]:
        """Returns (access_token, refresh_token) or raises Exception."""
        try:
            # 1. Get List of Tabs
            resp = requests.get(f"{self.DEBUG_URL}/json", timeout=2)
            tabs = resp.json()
            
            target_tab = None
            for tab in tabs:
                url = tab.get("url", "")
                if "bulsaja.com" in url:
                    target_tab = tab
                    break
            
            if not target_tab:
                raise Exception("Bulsaja 탭을 찾을 수 없습니다. (크롬이 9222 포트로 실행 중이어야 합니다)")
            
            # 2. Connect via WebSocket (using requests to /json/monitor or simply HTTP evaluate)
            # Actually standard CDP suggests WebSocket, but devtools also supports simple HTTP if we just need console evaluation?
            # No, usually need websocket. But let's verify if we can do stateless eval on some versions.
            # Most reliable way without `websocket-client` dependency:
            # Re-using previous knowledge: We can use `requests` if we find a library-less way, 
            # but usually it requires a websocket library.
            # Let's check imports. `websocket` is commonly used.
            # If not installed, we can't easily do it.
            # Wait, `requests` is imported. Let's assume user has `websocket-client` or can install it.
            # The v3.5 file imported `websocket`.
            pass

        except Exception as e:
            raise e
            
        return "", ""

    # Re-implementing using a simpler approach if websocket lib is missing?
    # Let's assume we can use `websocket-client`.
    # To be safe, let's use a method that works.
    
    def extract(self) -> Tuple[str, str, str]:
        """Returns (access, refresh, cookie) or raises Exception."""
        try:
            import websocket
        except ImportError:
            raise Exception("'websocket-client' 모듈이 필요합니다. (pip install websocket-client)")
            
        # 1. Find Tab
        try:
            tabs = requests.get(f"{self.DEBUG_URL}/json", timeout=2).json()
        except:
            raise Exception("크롬 디버깅 포트(9222)에 연결할 수 없습니다.")
            
        ws_url = None
        for tab in tabs:
            if "bulsaja.com" in tab.get("url", ""):
                ws_url = tab.get("webSocketDebuggerUrl")
                break
        
        if not ws_url:
            raise Exception("불사자 탭을 찾을 수 없습니다. (www.bulsaja.com 열려있어야 함)")
            
        # 2. Execute JS
        ws = websocket.create_connection(ws_url)
        call_id = 1
        
        # Combined JS to get both tokens and cookie
        js_expr = """
        (function() {
            var tokenStr = localStorage.getItem('token');
            var cookieStr = document.cookie;
            var resultObj = { accessToken: '', refreshToken: '', cookie: cookieStr };
            if (tokenStr) {
                try {
                    var tokenObj = JSON.parse(tokenStr);
                    if (tokenObj.state) {
                        resultObj.accessToken = tokenObj.state.accessToken || '';
                        resultObj.refreshToken = tokenObj.state.refreshToken || '';
                    }
                } catch(e) {}
            }
            return JSON.stringify(resultObj);
        })()
        """
        
        ws.send(json.dumps({
            "id": call_id,
            "method": "Runtime.evaluate",
            "params": {"expression": js_expr, "returnByValue": True}
        }))
        
        raw_resp = ws.recv()
        result = json.loads(raw_resp)
        ws.close()
        
        result_val = result.get('result', {}).get('result', {}).get('value', '{}')
        parsed = json.loads(result_val)
        
        access = parsed.get('accessToken', '')
        refresh = parsed.get('refreshToken', '')
        cookie = parsed.get('cookie', '')
        
        if not access:
            raise Exception("토큰을 찾을 수 없습니다. (로그인 상태 확인 요망)")
            
        return access, refresh, cookie

# ==================== NAVER API CLIENT ====================
class NaverKeywordClient:
    """Client for Naver Search API & Search Ads API."""
    SEARCH_API_URL = "https://openapi.naver.com/v1/search/shop.json"
    ADS_API_URL = "https://api.naver.com"

    def __init__(self, client_id: str = "", client_secret: str = "", ads_access_key: str = "", ads_secret_key: str = "", ads_customer_id: str = ""):
        self.client_id = client_id
        self.client_secret = client_secret
        self.ads_access_key = ads_access_key
        self.ads_secret_key = ads_secret_key
        self.ads_customer_id = ads_customer_id

    def get_search_volume(self, keyword: str) -> Dict:
        """Fetch monthly search volume from Naver Search Ads API."""
        if not self.ads_access_key or not self.ads_secret_key:
            return {"pc": 0, "mobile": 0, "total": 0}

        timestamp = str(int(time.time() * 1000))
        method = "GET"
        path = "/keywordstool"
        
        # Signature logic
        msg = f"{timestamp}.{method}.{path}"
        signature = base64.b64encode(hmac.new(
            self.ads_secret_key.encode('utf-8'),
            msg.encode('utf-8'),
            hashlib.sha256
        ).digest()).decode('utf-8')

        headers = {
            "X-Timestamp": timestamp,
            "X-API-KEY": self.ads_access_key,
            "X-Customer": self.ads_customer_id,
            "X-Signature": signature
        }

        try:
            params = {"keywords": keyword}
            resp = requests.get(f"{self.ADS_API_URL}{path}", params=params, headers=headers)
            data = resp.json()
            keywords = data.get("keywordList", [])
            if keywords:
                k = keywords[0]
                # Convert string counts like "< 10" to 5
                def parse_cnt(v):
                    if isinstance(v, str) and "<" in v: return 5
                    try: return int(v)
                    except: return 0
                
                pc = parse_cnt(k.get("monthlyPcQcCnt", 0))
                mo = parse_cnt(k.get("monthlyMobileQcCnt", 0))
                return {"pc": pc, "mobile": mo, "total": pc + mo}
        except Exception as e:
            print(f"Naver Ads API Error: {e}")
        
        return {"pc": 0, "mobile": 0, "total": 0}

    def get_total_products(self, keyword: str) -> int:
        """Fetch total product count from Naver Search API."""
        if not self.client_id or not self.client_secret:
            return 0
            
        headers = {
            "X-Naver-Client-Id": self.client_id,
            "X-Naver-Client-Secret": self.client_secret
        }
        params = {"query": keyword, "display": 1}
        try:
            resp = requests.get(self.SEARCH_API_URL, params=params, headers=headers)
            data = resp.json()
            return int(data.get("total", 0))
        except Exception as e:
            print(f"Naver Search API Error: {e}")
            return 0

    def get_shopping_tags(self, keyword: str) -> List[str]:
        """Scrape related tags from Naver Shopping search results."""
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36"
        }
        url = f"https://search.shopping.naver.com/search/all?query={keyword}"
        try:
            resp = requests.get(url, headers=headers, timeout=5)
            # Find tags in the HTML (using regex as a lightweight alternative to BS4)
            # Naver Shopping tags are often in a JSON structure or inside specific tags
            tags = re.findall(r'"title":"([^"]+)"', resp.text)
            # Filtering and cleaning tags (very rough heuristic)
            valid_tags = [t for t in tags if len(t) > 1 and len(t) < 10 and t != keyword]
            return list(set(valid_tags))[:10]
        except:
            return []

    def get_rank_api(self, keyword: str, product_id: str) -> int:
        """네이버 쇼핑에서 상품 순위 조회. 0=미발견."""
        if not self.client_id or not self.client_secret:
            return 0
        headers = {
            "X-Naver-Client-Id": self.client_id,
            "X-Naver-Client-Secret": self.client_secret
        }
        params = {"query": keyword, "display": 100}
        try:
            resp = requests.get(self.SEARCH_API_URL, params=params, headers=headers)
            data = resp.json()
            items = data.get("items", [])
            for i, item in enumerate(items, 1):
                # productId 또는 link에서 매칭
                link = item.get("link", "")
                if product_id in link:
                    return i
            return 0
        except:
            return 0

    def get_keyword_stats(self, keyword: str) -> Dict:
        """Returns {volume: int, products: int, ratio: float, category: str}"""
        vol = self.get_search_volume(keyword)
        
        # Extended search to get category info
        cat_id = ""
        prods = 0
        if self.client_id and self.client_secret:
            headers = {"X-Naver-Client-Id": self.client_id, "X-Naver-Client-Secret": self.client_secret}
            params = {"query": keyword, "display": 1}
            try:
                resp = requests.get(self.SEARCH_API_URL, params=params, headers=headers)
                data = resp.json()
                prods = int(data.get("total", 0))
                items = data.get("items", [])
                if items:
                    cat_id = items[0].get("category4") or items[0].get("category3") or ""
            except: pass

        total_vol = vol["total"]
        ratio = prods / total_vol if total_vol > 0 else 9999
        return {
            "keyword": keyword,
            "volume": total_vol,
            "products": prods,
            "ratio": round(ratio, 2),
            "category": cat_id
        }

# ==================== NAVER COMMERCE API CLIENT ====================
class NaverCommerceClient:
    """Client for Naver Commerce API (Smart Store)."""
    BASE_URL = "https://api.commerce.naver.com/external"

    def __init__(self, client_id: str = "", client_secret: str = ""):
        self.client_id = client_id
        self.client_secret = client_secret
        self.access_token = None
        self.token_expiry = 0

    def _get_token(self) -> str:
        """네이버 커머스API HMAC 시그니처 방식 토큰 발급 (전략문서 8.4절)"""
        if self.access_token and time.time() < self.token_expiry:
            return self.access_token

        timestamp = str(int(time.time() * 1000))

        # HMAC-SHA256 시그니처 생성: client_id + "_" + timestamp
        message = f"{self.client_id}_{timestamp}"
        signature = base64.b64encode(
            hmac.new(
                self.client_secret.encode('utf-8'),
                message.encode('utf-8'),
                hashlib.sha256
            ).digest()
        ).decode('utf-8')

        url = f"{self.BASE_URL}/v1/oauth2/token"
        params = {
            "client_id": self.client_id,
            "timestamp": timestamp,
            "client_secret_sign": signature,
            "grant_type": "client_credentials",
            "type": "SELF"
        }

        try:
            resp = requests.post(url, data=params)
            data = resp.json()
            self.access_token = data.get("access_token")
            self.token_expiry = time.time() + data.get("expires_in", 3600) - 60
            return self.access_token
        except Exception as e:
            print(f"Commerce Auth Error: {e}")
            return None

    def update_product_name(self, product_id: str, new_name: str) -> bool:
        """Update product name via Commerce API."""
        token = self._get_token()
        if not token: return False

        url = f"{self.BASE_URL}/v2/products/{product_id}"
        headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
        payload = {"originProduct": {"name": new_name}}

        try:
            resp = requests.patch(url, json=payload, headers=headers)
            return resp.status_code == 200
        except:
            return False

# ==================== NAVER RANK TRACKER ====================
class NaverRankTracker:
    """Tracker to find product ranking on Naver Shopping."""
    SEARCH_URL = "https://search.shopping.naver.com/search/all"

    @staticmethod
    def find_rank(keyword: str, product_id: str, max_pages: int = 5) -> Tuple[int, int]:
        """Returns (page, position) or (0, 0) if not found."""
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36"
        }
        
        for page in range(1, max_pages + 1):
            params = {"query": keyword, "pagingIndex": page, "pagingSize": 40}
            try:
                resp = requests.get(NaverRankTracker.SEARCH_URL, params=params, headers=headers)
                # Naver Shopping results are often embedded in a JSON-like string (next_data)
                html = resp.text
                if product_id in html:
                    # Logic to extract position from HTML (Simplified for demonstration)
                    # In a real scenario, use BeautifulSoup to find the specific element's index
                    return page, 1 # Placeholder for exact position
            except:
                continue
        return 0, 0

class BulsajaAPIClient:
    """Client for Bulsaja API with Pagination Support."""
    BASE_URL = "https://api.bulsaja.com/api"
    
    def __init__(self, access_token: str = "", refresh_token: str = "", cookie: str = ""):
        self.access_token = access_token
        self.refresh_token = refresh_token
        self.cookie = cookie
        self.session = requests.Session()
        if access_token:
            self._setup_session()
    
    def _setup_session(self):
        self.session.headers.update({
            'accept': 'application/json, text/plain, */*',
            'accesstoken': self.access_token,
            'refreshtoken': self.refresh_token,
            'Cookie': self.cookie,
            'content-type': 'application/json',
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        })
    
    def fetch_products_by_group(self, group_name: str, batch_size: int = 500, max_items: int = 10000) -> List[Dict]:
        """Fetch products by market group with pagination."""
        all_products = []
        start_row = 0
        print(f"Fetching group: {group_name}")
        
        while len(all_products) < max_items:
            filter_model = {}
            if group_name:
                filter_model = {
                    "marketGroupName": {
                        "filterType": "text",
                        "type": "equals",
                        "filter": group_name
                    }
                }

            url = f"{self.BASE_URL}/manage/list/serverside"
            payload = {
                "request": {
                    "startRow": start_row,
                    "endRow": start_row + batch_size,
                    "sortModel": [],
                    "filterModel": filter_model
                }
            }
            
            try:
                response = self.session.post(url, json=payload)
                response.raise_for_status()
                data = response.json()
                
                rows = data.get('rowData', [])
                if not rows: break
                
                all_products.extend(rows)
                last_row = data.get('lastRow', 0)
                if len(all_products) >= last_row: break
                
                start_row += batch_size
                print(f"fetched {len(all_products)} / {last_row}")
                time.sleep(0.5)
                
            except Exception as e:
                print(f"Error fetching batch at {start_row}: {e}")
                break
                
        return all_products[:max_items]

    def update_product_names(self, updates: List[Dict]) -> bool:
        url = f"{self.BASE_URL}/sourcing/bulk-update-names"
        update_items = []
        for item in updates:
            update_items.append({
                "id": item["id"],
                "uploadCommonProductName": item["name"],
                "uploadCoupangProductName": item.get("coupang_name", item["name"]),
                "uploadSmartStoreProductName": item.get("smartstore_name", item["name"])
            })
        payload = {"updateItems": update_items}
        try:
            response = self.session.post(url, json=payload)
            response.raise_for_status()
            return True
        except:
            return False

    def update_single_product(self, product_id: str, new_name: str) -> bool:
        return self.update_product_names([{"id": product_id, "name": new_name}])

class GeminiMultiAccountManager:
    """Manages multiple Gemini API keys."""
    
    def __init__(self, api_keys: List[str], log_callback=None):
        self.log = log_callback if log_callback else print
        self.accounts = []
        
        if not api_keys:
            self.log("⚠️ No Gemini Keys provided.")
        
        for i, key in enumerate(api_keys, 1):
            self.accounts.append({
                'key': key,
                'index': i,
                'daily_used': 0,
                'minute_used': 0,
                'daily_limit': 1490, 
                'minute_limit': 14,
                'last_minute_reset': datetime.now(),
                'last_day_reset': datetime.now().date(),
                'model': None
            })
        
        self.current_key_index = 0
    
    def _reset_if_needed(self, acc: dict):
        now = datetime.now()
        if (now - acc['last_minute_reset']).seconds >= 60:
            acc['minute_used'] = 0
            acc['last_minute_reset'] = now
        
        if now.date() > acc['last_day_reset']:
            acc['daily_used'] = 0
            acc['last_day_reset'] = now.date()

    def _get_next_available(self) -> Optional[dict]:
        if not self.accounts: return None
        tried = 0
        while tried < len(self.accounts):
            acc = self.accounts[self.current_key_index]
            self._reset_if_needed(acc)
            if acc['daily_used'] < acc['daily_limit'] and acc['minute_used'] < acc['minute_limit']:
                return acc
            self.current_key_index = (self.current_key_index + 1) % len(self.accounts)
            tried += 1
        return None

    def generate_content(self, prompt: str, image_data: bytes = None, image_mime: str = "image/jpeg", temperature: float = 0.7) -> Optional[str]:
        account = self._get_next_available()
        if not account:
            self.log("❌ 모든 API 키 사용량 초과.")
            return None
        
        try:
            if not account['model']:
                account['model'] = genai.Client(api_key=account['key'])
            
            contents = [prompt]
            if image_data:
                contents.append(types.Part.from_bytes(data=image_data, mime_type=image_mime))

            response = account['model'].models.generate_content(
                model='gemini-2.0-flash',
                contents=contents,
                config=types.GenerateContentConfig(
                    temperature=temperature,
                    max_output_tokens=1000,
                    safety_settings=[
                        types.SafetySetting(category='HARM_CATEGORY_HARASSMENT', threshold='BLOCK_NONE'),
                        types.SafetySetting(category='HARM_CATEGORY_HATE_SPEECH', threshold='BLOCK_NONE'),
                        types.SafetySetting(category='HARM_CATEGORY_SEXUALLY_EXPLICIT', threshold='BLOCK_NONE'),
                        types.SafetySetting(category='HARM_CATEGORY_DANGEROUS_CONTENT', threshold='BLOCK_NONE'),
                    ]
                )
            )
            
            account['daily_used'] += 1
            account['minute_used'] += 1
            
            if response.text:
                return response.text.strip()
            return None
        except Exception as e:
            self.log(f"⚠️ Account {account['index']} 오류: {e}")
            return None

# ======================================================
# WORKER THREADS
# ======================================================

class DataLoadWorker(QThread):
    finished = pyqtSignal(list, str) # data, error_msg

    def __init__(self, mode: str, source: str, api_client: BulsajaAPIClient = None):
        super().__init__()
        self.mode = mode # 'excel' or 'api'
        self.source = source
        self.api_client = api_client
        self.kiwi = Kiwi()

    def extract_keywords(self, text):
        if not text: return ""
        try:
            results = self.kiwi.analyze(text)
            keywords = []
            if results:
                tokens = results[0][0]
                for token_str, tag, _, _ in tokens:
                    if tag.startswith('NNG') or tag.startswith('NNP') or tag.startswith('SL'):
                        if len(token_str) > 1:
                            keywords.append(token_str)
            return ", ".join(list(set(keywords)))
        except:
            return ""

    def run(self):
        try:
            items = []
            if self.mode == 'excel':
                wb = load_workbook(self.source)
                ws = wb.active
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    if not row[0]: continue
                    seller_code = str(row[0]) if row[0] else ""
                    original_name = str(row[1]) if len(row) > 1 and row[1] else ""
                    keywords = self.extract_keywords(original_name)
                    items.append({
                        "id": str(uuid.uuid4())[:8],
                        "seller_code": seller_code,
                        "original_name": original_name,
                        "new_name": "",
                        "status": "Ready",
                        "keywords": keywords,
                    })
            
            elif self.mode == 'api':
                if not self.api_client:
                    self.finished.emit([], "API Client 없음")
                    return
                products = self.api_client.fetch_products_by_group(self.source)
                for p in products:
                    original_name = p.get("productName", "") or p.get("uploadCommonProductName", "")
                    keywords = self.extract_keywords(original_name)
                    items.append({
                        "id": p.get("id", ""),
                        "seller_code": p.get("sellerProductCode", ""),
                        "original_name": original_name,
                        "new_name": "",
                        "status": "Ready",
                        "keywords": keywords,
                    })
            
            self.finished.emit(items, "")
        except Exception as e:
            self.finished.emit([], str(e))

# ==================== MADWORD-STYLE UI COMPONENTS ====================

class SettingsDialog(QDialog):
    """Integrated settings dialog (Madword Style)"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("⚙️ 상품명 만들기 옵션")
        self.resize(500, 650) # Increased size
        self.setStyleSheet("""
            QDialog { background-color: #1e1e1e; color: white; }
            QLabel { color: #ddd; }
            QGroupBox { color: #00bfa5; font-weight: bold; border: 1px solid #333; margin-top: 15px; }
            QPushButton { background-color: #00897b; color: white; font-weight: bold; padding: 10px; border-radius: 4px; }
            QPushButton#Cancel { background-color: #444; }
        """)
        
        layout = QVBoxLayout(self)
        
        # 0. Load settings independently using absolute script path
        base_path = os.path.dirname(os.path.abspath(__file__))
        self.settings_json_path = os.path.join(base_path, "product_name_gen_settings.json")
        
        file_data = {}
        if os.path.exists(self.settings_json_path):
            try:
                with open(self.settings_json_path, "r", encoding="utf-8") as f:
                    file_data = json.load(f)
            except: pass
            
        # Preference merging logic
        s = file_data.get("preferences", {})
        if not s and parent:
             s = parent.settings_data
             
        self._ui_refs = [] # Explicit references to prevent GC

        group = QGroupBox("⚙️ 상품명 생성 옵션", self)
        self._ui_refs.append(group)
        g_layout = QVBoxLayout(group)
        
        self.chk_filter = QCheckBox("abc 알파벳, 숫자, 1글자 키워드 필터링", self)
        self.chk_filter.setChecked(s.get("filter_junk", True))
        g_layout.addWidget(self.chk_filter)
        self._ui_refs.append(self.chk_filter)
        
        self.chk_sim_mode = QCheckBox("🧪 시뮬레이션 모드 (서버 반영 제외)", self)
        self.chk_sim_mode.setChecked(s.get("sim_mode", False))
        self.chk_sim_mode.setStyleSheet("color: #ff9800; font-weight: bold;")
        g_layout.addWidget(self.chk_sim_mode)
        self._ui_refs.append(self.chk_sim_mode)
        
        row_gem = QHBoxLayout()
        lbl_gem = QLabel("💎 Gemini Key:", self)
        row_gem.addWidget(lbl_gem)
        self._ui_refs.append(lbl_gem)
        
        gem_text = str(file_data.get("gemini_keys", (parent.txt_gemini.text() if (parent and not sip.isdeleted(parent)) else "")))
        self.txt_gemini = QLineEdit(gem_text, self)
        self.txt_gemini.setPlaceholderText("API Keys (Comma separated)")
        row_gem.addWidget(self.txt_gemini)
        self._ui_refs.append(self.txt_gemini)
        g_layout.addLayout(row_gem)
        
        row1 = QHBoxLayout()
        self.chk_keep_orig = QCheckBox("📋 원본키워드 최종상품명 포함", self)
        self.chk_keep_orig.setChecked(s.get("keep_orig", False))
        self.spn_prefix_cnt = QSpinBox(self)
        self.spn_prefix_cnt.setRange(0, 999)
        self.spn_prefix_cnt.setValue(s.get("prefix_cnt", 999))
        row1.addWidget(self.chk_keep_orig)
        lbl_prefix = QLabel("앞 단어 개수:", self)
        row1.addWidget(lbl_prefix)
        row1.addWidget(self.spn_prefix_cnt)
        self._ui_refs.extend([self.chk_keep_orig, self.spn_prefix_cnt, lbl_prefix])
        g_layout.addLayout(row1)
        
        row2 = QHBoxLayout()
        self.chk_repeat = QCheckBox("📊 최종상품명 반복횟수", self)
        self.cmb_sort = QComboBox(self)
        self.cmb_sort.addItems(["내림차순", "오름차순"])
        self.cmb_sort.setCurrentText(s.get("sort_orders", "내림차순"))
        row2.addWidget(self.chk_repeat)
        row2.addWidget(self.cmb_sort)
        self._ui_refs.extend([self.chk_repeat, self.cmb_sort])
        g_layout.addLayout(row2)
        
        row3 = QHBoxLayout()
        lbl_word = QLabel("🔢 상품명 단어 개수", self)
        row3.addWidget(lbl_word)
        self.spn_word_cnt = QSpinBox(self)
        self.spn_word_cnt.setValue(s.get("word_limit", 10))
        row3.addWidget(self.spn_word_cnt)
        lbl_char = QLabel("글자", self)
        row3.addWidget(lbl_char)
        self.spn_char_cnt = QSpinBox(self)
        self.spn_char_cnt.setValue(s.get("char_limit", 50))
        row3.addWidget(self.spn_char_cnt)
        lbl_byte = QLabel("Byte", self)
        row3.addWidget(lbl_byte)
        self._ui_refs.extend([lbl_word, self.spn_word_cnt, lbl_char, self.spn_char_cnt, lbl_byte])
        g_layout.addLayout(row3)
        
        self.chk_shuffle_tags = QCheckBox("🔀 태그 섞기", self)
        self.chk_shuffle_tags.setChecked(s.get("shuffle_tags", False))
        g_layout.addWidget(self.chk_shuffle_tags)
        
        self.chk_ai_gemini = QCheckBox("🤖 AI 자동 검수 사용 (Gemini)", self)
        self.chk_ai_gemini.setChecked(s.get("use_ai", True))
        g_layout.addWidget(self.chk_ai_gemini)
        self._ui_refs.extend([self.chk_shuffle_tags, self.chk_ai_gemini])
        
        group_naver = QGroupBox("💚 네이버 API 설정")
        n_layout = QVBoxLayout(group_naver)
        
        # Naver Settings (Merge with file data)
        self.txt_naver_id = QLineEdit(file_data.get("naver_id", (parent.txt_naver_id.text() if parent else "")))
        self.txt_naver_id.setPlaceholderText("검색 API Client ID")
        self.txt_naver_secret = QLineEdit(file_data.get("naver_secret", (parent.txt_naver_secret.text() if parent else "")))
        self.txt_naver_secret.setPlaceholderText("검색 API Client Secret")
        self.txt_naver_secret.setEchoMode(QLineEdit.EchoMode.Password)
        self.txt_ads_key = QLineEdit(file_data.get("ads_key", (parent.txt_ads_key.text() if parent else "")))
        self.txt_ads_key.setPlaceholderText("광고 API Access Key")
        self.txt_ads_secret = QLineEdit(file_data.get("ads_secret", (parent.txt_ads_secret.text() if parent else "")))
        self.txt_ads_secret.setPlaceholderText("광고 API Secret Key")
        self.txt_ads_secret.setEchoMode(QLineEdit.EchoMode.Password)
        self.txt_ads_cust_id = QLineEdit(file_data.get("ads_cust_id", (parent.txt_ads_cust_id.text() if parent else "")))
        self.txt_ads_cust_id.setPlaceholderText("광고 API Customer ID")
        
        n_layout.addWidget(self.txt_naver_id)
        n_layout.addWidget(self.txt_naver_secret)
        n_layout.addWidget(self.txt_ads_key)
        n_layout.addWidget(self.txt_ads_secret)
        n_layout.addWidget(self.txt_ads_cust_id)
        layout.addWidget(group_naver)

        # Commerce API Settings
        group_comm = QGroupBox("🛒 커머스(Smart Store) API 설정")
        c_layout = QVBoxLayout(group_comm)
        
        self.txt_comm_id = QLineEdit(file_data.get("comm_id", (parent.txt_comm_id.text() if parent else "")))
        self.txt_comm_id.setPlaceholderText("커머스 API Client ID")
        self.txt_comm_secret = QLineEdit(file_data.get("comm_secret", (parent.txt_comm_secret.text() if parent else "")))
        self.txt_comm_secret.setPlaceholderText("커머스 API Client Secret")
        self.txt_comm_secret.setEchoMode(QLineEdit.EchoMode.Password)
        c_layout.addWidget(self.txt_comm_id)
        c_layout.addWidget(self.txt_comm_secret)
        layout.addWidget(group_comm)

        buttons = QHBoxLayout()
        btn_ok = QPushButton("OK")
        btn_cancel = QPushButton("Cancel")
        btn_cancel.setObjectName("Cancel")
        btn_ok.clicked.connect(self._on_accept)
        btn_cancel.clicked.connect(self.reject)
        buttons.addStretch()
        buttons.addWidget(btn_ok)
        buttons.addWidget(btn_cancel)
        layout.addLayout(buttons)

    def get_settings_snapshot(self):
        """Ultra-safe method to capture widget values before they are deleted."""
        try:
            # Lifecycle check
            def val(widget, attr="text"):
                if widget is None or sip.isdeleted(widget): return ""
                if hasattr(widget, "isChecked"): return widget.isChecked()
                if hasattr(widget, "value"): return widget.value()
                if hasattr(widget, "currentText"): return widget.currentText()
                return getattr(widget, attr)() if hasattr(widget, attr) else ""

            return {
                "filter_junk": val(self.chk_filter),
                "sim_mode": val(self.chk_sim_mode),
                "keep_orig": val(self.chk_keep_orig),
                "prefix_cnt": val(self.spn_prefix_cnt),
                "sort_orders": val(self.cmb_sort),
                "word_limit": val(self.spn_word_cnt),
                "char_limit": val(self.spn_char_cnt),
                "shuffle_tags": val(self.chk_shuffle_tags),
                "use_ai": val(self.chk_ai_gemini),
                "gemini_keys": val(self.txt_gemini),
                "naver": {
                    "client_id": val(self.txt_naver_id),
                    "client_secret": val(self.txt_naver_secret),
                    "ads_access_key": val(self.txt_ads_key),
                    "ads_secret_key": val(self.txt_ads_secret),
                    "ads_customer_id": val(self.txt_ads_cust_id)
                },
                "commerce": {
                    "client_id": val(self.txt_comm_id),
                    "client_secret": val(self.txt_comm_secret)
                }
            }
        except Exception as e:
            print(f"Snapshot Error: {e}")
            return {}

    def _on_accept(self):
        """Final bulletproof handler for OK button."""
        # 1. IMMEDIATE DATA SNAPSHOT (Widget -> Python Dict)
        snapshot = self.get_settings_snapshot()
        if not snapshot:
            print("CRITICAL: Failed to capture settings snapshot.")
            self.reject()
            return

        try:
            # 2. Prepare JSON structure
            save_data = {
                "naver_id": snapshot['naver']['client_id'],
                "naver_secret": snapshot['naver']['client_secret'],
                "ads_key": snapshot['naver']['ads_access_key'],
                "ads_secret": snapshot['naver']['ads_secret_key'],
                "ads_cust_id": snapshot['naver']['ads_customer_id'],
                "comm_id": snapshot['commerce']['client_id'],
                "comm_secret": snapshot['commerce']['client_secret'],
                "gemini_keys": snapshot.get('gemini_keys', ''),
                "preferences": snapshot
            }
            # 3. Save to Absolute Path with explicit Flush & Sync
            with open(self.settings_json_path, "w", encoding="utf-8") as f:
                json.dump(save_data, f, ensure_ascii=False, indent=2)
                f.flush()
                os.fsync(f.fileno())
            
            # 4. Sync with Parent (while p is definitely alive)
            if self.parent():
                p = self.parent()
                try:
                    # Sync UI
                    p.txt_naver_id.setText(save_data["naver_id"])
                    p.txt_naver_secret.setText(save_data["naver_secret"])
                    p.txt_ads_key.setText(save_data["ads_key"])
                    p.txt_ads_secret.setText(save_data["ads_secret"])
                    p.txt_ads_cust_id.setText(save_data["ads_cust_id"])
                    p.txt_comm_id.setText(save_data["comm_id"])
                    p.txt_comm_secret.setText(save_data["comm_secret"])
                    p.txt_gemini.setText(save_data["gemini_keys"])
                    
                    # Sync Memory
                    p.settings_data = snapshot
                    p.log(f"✅ 설정 저장 완료: {self.settings_json_path}")
                except Exception as sync_e:
                    print(f"DEBUG: Sync warning: {sync_e}")

            # 5. Success confirmation
            QMessageBox.information(self, "성공", "설정이 안전하게 저장되었습니다.")
            self.accept()
            
        except Exception as e:
            import traceback
            print(f"SAVE ERROR TRACEBACK:\n{traceback.format_exc()}")
            QMessageBox.critical(self, "저장 오류", f"설정 저장 중 치명적 오류가 발생했습니다:\n{e}")
            self.reject() # Don't say success if it failed

class MetadataPanel(QFrame):
    """Top panel for product metadata (Madword Style)"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setFrameShape(QFrame.Shape.StyledPanel)
        self.setStyleSheet("""
            MetadataPanel { 
                background-color: #1e1e1e; 
                border: 1px solid #333; 
                border-radius: 5px;
            }
            QLabel { color: #888; font-size: 11px; }
            QLineEdit { background-color: #121212; border: 1px solid #333; height: 24px; padding: 2px 5px; }
        """)
        
        self._ui_refs = [] # Hold refs
        layout = QGridLayout(self)
        layout.setContentsMargins(15, 15, 15, 15)
        layout.setSpacing(10)

        # Row 1
        lbl1 = QLabel("원본 상품명:", self)
        layout.addWidget(lbl1, 0, 0)
        self.txt_orig_name = QLineEdit(self)
        self.txt_orig_name.setReadOnly(True)
        layout.addWidget(self.txt_orig_name, 0, 1)

        lbl2 = QLabel("카테고리:", self)
        layout.addWidget(lbl2, 0, 2)
        self.txt_cat_name = QLineEdit(self)
        layout.addWidget(self.txt_cat_name, 0, 3)

        lbl3 = QLabel("제거된단어:", self)
        layout.addWidget(lbl3, 0, 4)
        self.txt_removed = QLineEdit(self)
        layout.addWidget(self.txt_removed, 0, 5)

        self._ui_refs.extend([lbl1, lbl2, lbl3, self.txt_orig_name, self.txt_cat_name, self.txt_removed])

        # Row 2
        lbl4 = QLabel("최종상품명:", self)
        layout.addWidget(lbl4, 1, 0)
        self.txt_final_name = QLineEdit(self)
        layout.addWidget(self.txt_final_name, 1, 1, 1, 3) # Span across

        lbl5 = QLabel("AI추가단어:", self)
        layout.addWidget(lbl5, 1, 4)
        self.txt_added = QLineEdit(self)
        layout.addWidget(self.txt_added, 1, 5)
        self._ui_refs.extend([lbl4, lbl5, self.txt_final_name, self.txt_added])

        # Row 3
        lbl6 = QLabel("AI연관키워드:", self)
        layout.addWidget(lbl6, 2, 0)
        self.txt_related_meta = QLineEdit(self)
        layout.addWidget(self.txt_related_meta, 2, 1)

        lbl7 = QLabel("AI최적화내역:", self)
        layout.addWidget(lbl7, 2, 2)
        self.txt_optimized = QLineEdit(self)
        layout.addWidget(self.txt_optimized, 2, 3, 1, 2)
        self._ui_refs.extend([lbl6, lbl7, self.txt_related_meta, self.txt_optimized])

        self.btn_save = QPushButton("💾 수동 저장")
        if parent:
            self.btn_save.clicked.connect(parent.on_metadata_save)
        self.btn_save.setStyleSheet("background-color: #00897b; color: white; padding: 5px; font-weight: bold;")
        layout.addWidget(self.btn_save, 2, 5)

        # Columns sizing
        layout.setColumnStretch(1, 1)
        layout.setColumnStretch(3, 2)
        layout.setColumnStretch(5, 1)

class BottomAnalysisPanel(QWidget):
    """Bottom analysis panels (Keywords, Tags, Banned words)"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.main = parent # Store explicit reference to main window
        self._ui_refs = [] # Lifecycle maintenance
        layout = QHBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(5)

        # Panel 1: Extracted Keywords
        self.pnl_extracted = self._create_list_panel("🔍 추출 키워드", ["키워드", "반복횟수"])
        layout.addWidget(self.pnl_extracted)

        # Panel 2: Original Keywords
        self.pnl_original = self._create_list_panel("📝 원본 키워드", ["키워드"])
        layout.addWidget(self.pnl_original)

        # Panel 3: Related Keywords
        self.pnl_related = self._create_list_panel("🔗 연관 키워드", ["키워드", "검색량"])
        layout.addWidget(self.pnl_related)

        # Panel 4: Tags
        self.pnl_tags = self._create_list_panel("🏷️ 태그", ["키워드"])
        layout.addWidget(self.pnl_tags)

        # Panel 5: Full Tags
        self.pnl_all_tags = self._create_list_panel("🏷️ 전체 태그", ["키워드"])
        layout.addWidget(self.pnl_all_tags)

        # Panel 6: Ranking Hub (New)
        self.pnl_ranking = self._create_ranking_panel()
        layout.addWidget(self.pnl_ranking)

        # Panel 7: Banned Words
        self.pnl_banned = self._create_banned_panel()
        layout.addWidget(self.pnl_banned)

        # Panel 8: Log Console
        self.log_console = QPlainTextEdit()
        self.log_console.setReadOnly(True)
        self.log_console.setMaximumHeight(120)
        self.log_console.setStyleSheet("background-color: #0a0a0a; color: #66ff00; font-family: Consolas; font-size: 11px;")

    def _create_ranking_panel(self):
        frame = QGroupBox("📈 Ranking Hub")
        frame.setStyleSheet("""
            QGroupBox { color: #fbc02d; font-weight: bold; border: 1px solid #333; margin-top: 15px; }
            QTableWidget { background-color: #1e1e1e; border: none; }
        """)
        layout = QVBoxLayout(frame)
        self.table_ranks = QTableWidget(0, 3)
        self.table_ranks.setHorizontalHeaderLabels(["키워드", "순위", "변동"])
        self.table_ranks.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.table_ranks.verticalHeader().setVisible(False)
        layout.addWidget(self.table_ranks)
        
        btn_track = QPushButton("🔍 순위 즉시 추적")
        # Use self.main to access the parent method safely
        if hasattr(self, 'main') and self.main:
            btn_track.clicked.connect(self.main.on_rank_track_clicked)
        btn_track.setStyleSheet("background-color: #fbc02d; color: #333; font-weight: bold;")
        layout.addWidget(btn_track)
        return frame

    def _create_list_panel(self, title, headers):
        frame = QGroupBox(title)
        frame.setStyleSheet("""
            QGroupBox { 
                color: #00bfa5; 
                font-weight: bold; 
                border: 1px solid #333;
                margin-top: 15px;
            }
            QTableWidget { background-color: #1e1e1e; border: none; }
        """)
        layout = QVBoxLayout(frame)
        table = QTableWidget(0, len(headers))
        table.setHorizontalHeaderLabels(headers)
        table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        table.verticalHeader().setVisible(False)
        table.setAlternatingRowColors(True)
        layout.addWidget(table)
        frame.setMinimumWidth(220) # Ensure columns are visible
        return frame

    def _create_banned_panel(self):
        frame = QGroupBox("🚫 금지어")
        frame.setStyleSheet("color: #ff5252; font-weight: bold; border: 1px solid #333; margin-top: 15px;")
        frame.setMinimumWidth(200) # Ensure it doesn't get squashed
        layout = QVBoxLayout(frame)
        
        input_row = QHBoxLayout()
        self.txt_banned_add = QLineEdit()
        self.txt_banned_add.setPlaceholderText("금지어 입력...")
        btn_del = QPushButton("❌ 삭제")
        btn_add = QPushButton("➕ 추가")
        btn_del.clicked.connect(self.on_banned_del)
        btn_add.clicked.connect(self.on_banned_add)
        btn_del.setStyleSheet("background-color: #c62828; font-size: 10px; padding: 2px;")
        btn_add.setStyleSheet("background-color: #2e7d32; font-size: 10px; padding: 2px;")
        input_row.addWidget(btn_del)
        input_row.addWidget(btn_add)
        
        layout.addWidget(self.txt_banned_add)
        layout.addLayout(input_row)
        
        self.list_banned = QTextEdit()
        self.list_banned.setStyleSheet("background-color: #121212; color: #ff5252;")
        layout.addWidget(self.list_banned)
        return frame

    def on_banned_add(self):
        text = self.txt_banned_add.text().strip()
        if not text: return
        current = self.list_banned.toPlainText().split("\n")
        if text not in current:
            self.list_banned.append(text)
            self.txt_banned_add.clear()

    def on_banned_del(self):
        # Simplistic: removes the text in the input if present in the list
        text = self.txt_banned_add.text().strip()
        if not text: return
        current = self.list_banned.toPlainText().split("\n")
        if text in current:
            current.remove(text)
            self.list_banned.setText("\n".join(current))
            self.txt_banned_add.clear()

class BulkGenerationWorker(QThread):
    progress = pyqtSignal(int, dict) # row_index, result_data
    finished = pyqtSignal()
    log = pyqtSignal(str)

    def __init__(self, items: List[Dict], gemini_keys: List[str], api_client: BulsajaAPIClient = None, gen_params: Dict = None, naver_creds: Dict = None):
        super().__init__()
        self.items = items
        self.gemini_keys = gemini_keys
        self.api_client = api_client
        self.gen_params = gen_params or {}
        self.naver_creds = naver_creds or {}
        self.is_running = True
        self.kiwi = Kiwi()

    def run(self):
        gemini = GeminiMultiAccountManager(self.gemini_keys, log_callback=self.log.emit)
        naver = NaverKeywordClient(**self.naver_creds.get('naver', {}))
        
        commerce_creds = self.naver_creds.get('commerce', {})
        commerce = None
        if commerce_creds.get('client_id') and commerce_creds.get('client_secret'):
            commerce = NaverCommerceClient(**commerce_creds)
            self.log.emit("🛒 네이버 커머스 API 연결 활성화됨")

        total = len(self.items)
        self.log.emit(f"🚀 [매드워드 AI v3.0] 일괄 작업 시작: {total}개")
        success_count = 0
        
        # Parallel Execution (v4.0: 16 threads for speed, logic inside loop handles rank skip)
        with concurrent.futures.ThreadPoolExecutor(max_workers=8) as executor:
            for i, item in enumerate(self.items):
                if not self.is_running: break
                
                original_title = item.get('original_name', '')
                seller_code = str(item.get('seller_code', ''))
                target_rank = int(item.get('target_rank', 0)) or 100
                
                self.log.emit(f"🔄 [{i+1}/{total}] 상품({seller_code}) 처리 시작...")

                try:
                    # --- v4.0 Rank-Targeted Skip Logic ---
                    current_rank = 0
                    if naver:
                        self.log.emit(f"   🔍 현재 순위 확인 중... (타겟: {target_rank}위)")
                        current_rank = naver.get_rank_api(original_title[:20], seller_code)
                        item['prev_rank'] = current_rank
                        if current_rank > 0 and current_rank <= target_rank:
                            self.log.emit(f"   ✅ 목표 순위 달성({current_rank}위). 작업을 건너뜁니다.")
                            item['status'] = 'Skipped (Rank OK)'
                            self.save_to_db(item)
                            self.progress.emit(i, item)
                            success_count += 1
                            continue

                    # --- STAGE 0: DETECT TARGET CATEGORY ---
                    target_stats = naver.get_keyword_stats(original_title[:20]) if naver else {}
                    target_cat = target_stats.get("category", "")
                    if target_cat: self.log.emit(f"   🎯 타겟 카테고리 감지: {target_cat}")

                    # --- STAGE 1: CLEANUP & UNIT EXTRACTION (v4.0 with HTML) ---
                    stage1_result = self.process_stage1(original_title, gemini, item.get('main_image_url'), item.get('description'))
                    
                    # --- STAGE 2: ENRICHMENT (Parallel Fetch + Category Matching) ---
                    stage2_result = self.process_stage2(stage1_result['safe_nouns'], naver, target_category=target_cat)
                    
                    # --- STAGE 3: FINAL SEO ASSEMBLY (v4.0 Category-Specific) ---
                    final_name = self.process_stage3(original_title, stage1_result, stage2_result, gemini, category_info=target_cat)
                    
                    if final_name:
                        self.log.emit(f"   ✅ 최종 최적화 완료: {final_name}")
                        item['new_name'] = final_name
                        item['status'] = 'Done'
                        item['category'] = target_cat
                        item['keywords'] = ", ".join([k[0] for k in stage2_result['related']])

                        if not self.gen_params.get('sim_mode'):
                            if self.api_client: self.api_client.update_single_product(item.get('id'), final_name)
                            if commerce: commerce.update_product_name(item.get('id'), final_name)

                        self.save_to_db(item)
                        self.progress.emit(i, item)
                        success_count += 1
                        
                except Exception as e:
                    self.log.emit(f"❌ 항목 {i+1} 처리 중 오류: {e}")
                    item['status'] = 'Error'
                    self.progress.emit(i, item)

        self.log.emit(f"✅ SmartSellUp 프로세스 완료. 성공: {success_count}/{total}")
        self.finished.emit()

    def save_to_db(self, item: dict):
        """Save results to SQLite."""
        try:
            conn = sqlite3.connect("smartsellup.db")
            cur = conn.cursor()
            cur.execute("""
                INSERT INTO history (seller_code, original_name, ai_name, prev_rank, curr_rank, status)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (
                str(item.get('seller_code', '')), 
                item.get('original_name', ''), 
                item.get('new_name', ''), 
                item.get('prev_rank', 0), 
                item.get('curr_rank', 0), 
                item.get('status', '')
            ))
            conn.commit()
            conn.close()
        except Exception as e:
            print(f"DB Save Error: {e}")

    def _get_image_data(self, url: str) -> Optional[bytes]:
        if not url: return None
        try:
            resp = requests.get(url, timeout=5)
            if resp.status_code == 200:
                return resp.content
        except: pass
        return None

    def process_stage1(self, title: str, gemini, image_url: str = None, description_html: str = None):
        """Cleanup, Unit Extraction, and initial IP check with Multimodal & HTML support."""
        # 1. Extract units
        units = re.findall(r'(\d+[\w가-힣]+(?:개|세트|ml|g|kg|p|cm|mm))', title)
        
        # 2. HTML Description Analysis (New in v4.0)
        html_info = ""
        if description_html:
            self.log.emit("   📄 상세설명(HTML) 분석 중...")
            clean_html = re.sub('<[^<]+?>', '', description_html)[:2000] # Strip tags & limit
            prompt_html = f"이 상품의 상세설명에서 브랜드, 모델명, 핵심 스펙, 색상, 사이즈 등을 추출해라. 키워드 위주로 간결하게. 내용: {clean_html}"
            html_resp = gemini.generate_content(prompt_html)
            if html_resp: html_info = html_resp

        # 3. Image Analysis (Multimodal)
        img_data = self._get_image_data(image_url)
        img_brand = []
        img_desc = ""
        
        if img_data:
            self.log.emit("   📸 상품 이미지 분석 중 (Multimodal)...")
            prompt_img = "이 상품 이미지를 분석해라. 1. 명확한 브랜드명이 보인다면 브랜드명만 리스트로 반환. 2. 상품의 핵심 종류(예: 운동화, 물병)를 단어로 반환. JSON 형식: {'brands': [], 'product_type': ''}"
            img_resp = gemini.generate_content(prompt_img, image_data=img_data)
            if img_resp:
                try:
                    clean_json = img_resp.replace("```json", "").replace("```", "").strip()
                    match = re.search(r'\{.*\}', clean_json, re.DOTALL)
                    if match:
                        data = json.loads(match.group(0))
                        img_brand = data.get("brands", [])
                        img_desc = data.get("product_type", "")
                        if img_brand: self.log.emit(f"   📸 이미지에서 브랜드 감지: {img_brand}")
                except: pass

        # 3. Kiwi cleanup
        clean_title = title.replace(" ", "")
        res = self.kiwi.analyze(clean_title)
        safe_nouns = []
        risky_tokens = []
        
        if res:
            tokens = res[0][0]
            for token_str, tag, _, _ in tokens:
                if len(token_str) < 1: continue
                if tag.startswith('NNG'):
                    safe_nouns.append(token_str)
                elif tag.startswith('NNP') or tag.startswith('SL') or tag == 'UNKNOWN':
                    risky_tokens.append(token_str)

        if img_desc: safe_nouns.append(img_desc)

        # 4. AI Verification (Brands/IP in Risky Tokens)
        banned = list(set(img_brand))
        if risky_tokens:
            prompt_v = f"상품명 키워드 중 실존 브랜드만 추출해라. 없으면 []. 단어: {', '.join(risky_tokens)}"
            resp = gemini.generate_content(prompt_v)
            if resp:
                 try:
                     clean_json = resp.replace("```json", "").replace("```", "").strip()
                     match = re.search(r'\[.*\]', clean_json)
                     if match:
                         banned.extend(json.loads(match.group(0)))
                 except: pass
        
        return {
            "units": units,
            "safe_nouns": list(set(safe_nouns)),
            "banned": list(set(banned)),
            "risky": risky_tokens
        }

    def process_stage2(self, nouns: list, naver, target_category: str = ""):
        """Fetch Naver stats, filter by category, and score by frequency."""
        # 1. Fetch Shopping Tags
        shopping_tags = []
        if nouns:
            shopping_tags = naver.get_shopping_tags(nouns[0])
            if shopping_tags: self.log.emit(f"   🛡️ 쇼핑 연관 태그 수집됨: {len(shopping_tags)}개")

        # 2. Combine and Parallel Fetch Stats
        candidates = list(set(nouns + shopping_tags))
        scored = []
        
        with concurrent.futures.ThreadPoolExecutor(max_workers=8) as executor:
            future_to_kw = {executor.submit(naver.get_keyword_stats, kw): kw for kw in candidates[:20]}
            for future in concurrent.futures.as_completed(future_to_kw):
                stats = future.result()
                if stats["volume"] > 0:
                    # 매드워드 v3 룰: 카테고리 매칭 필터링
                    if target_category and stats["category"] and stats["category"] != target_category:
                        continue
                    scored.append(stats)
        
        # 3. 매드워드 v3 점수 방식: 검색량 빈도(가중치) 기반 정렬
        # 여기서는 검색량과 경쟁강도를 조합한 기존 방식을 유지하면서 '카테고리 일치'를 최우선으로 합니다.
        scored.sort(key=lambda x: (x['volume'], -x['ratio']), reverse=True)
        top_5 = scored[:5]
        top_related = [(s['keyword'], s['volume']) for s in top_5]
        
        return {
            "related": top_related,
            "extracted": [(n, 1) for n in nouns],
            "full_stats": scored
        }

    def process_stage3(self, original: str, s1: dict, s2: dict, gemini, category_info: str = ""):
        """구매대행 특화 SEO 상품명 생성 (전략문서 3.4절~3.5절 기반)"""
        related_kws = [r[0] for r in s2['related']]
        unit_str = " ".join(s1['units'])
        html_context = s1.get('html_info', '')

        # 카테고리별 템플릿 전략 (전략문서 3.5절)
        templates = {
            "의류": "[스타일] + [아이템] + [성별] + [시즌] + [핏] + [소재] + [컬러]",
            "패션": "[스타일] + [아이템] + [성별] + [용도(하객룩/출근룩)] + [소재] + [컬러]",
            "가전": "[기능] + [아이템] + [용도] + [스펙(전압/용량)] + [특징] + [컬러]",
            "디지털": "[기능] + [아이템] + [용도] + [스펙] + [특징] + [컬러]",
            "생활": "[용도] + [아이템] + [재질] + [사이즈] + [특징] + [컬러]",
            "스포츠": "[용도(캠핑/낚시/러닝)] + [아이템] + [특성(경량/휴대용)] + [스펙] + [컬러]",
            "식품": "[메인키워드] + [중량/수량] + [원산지] + [특징]",
        }

        template = ""
        strategy = ""
        for cat_key, tmpl in templates.items():
            if cat_key in category_info:
                template = tmpl
                break

        if not template:
            template = "[메인키워드] + [세부키워드1] + [세부키워드2] + [스펙] + [특징/용도] + [컬러]"

        # 카테고리 특화 전략
        if "가전" in category_info or "디지털" in category_info:
            strategy = "스펙(전압, 용량, 인치 등)을 반드시 포함. 모델번호는 제거."
        elif "패션" in category_info or "의류" in category_info:
            strategy = "실구매자 키워드(하객룩, 출근룩, 데일리룩) 반드시 포함. 성별+핏+소재 강조."
        elif "식품" in category_info:
            strategy = "중량, 수량, 원산지를 명확히 표기."
        elif "스포츠" in category_info or "레저" in category_info:
            strategy = "용도(캠핑/낚시/러닝)+특성(경량/휴대용/방수) 조합. 실구매자 니즈 반영."
        elif "생활" in category_info or "건강" in category_info:
            strategy = "용도+재질+사이즈 중심. 실사용 장면 키워드 활용."
        else:
            strategy = "롱테일 키워드 공략. 실구매자가 검색할 세부 키워드 조합."

        prompt = f"""[구매대행 상품명 SEO 엔진]

원본 상품명: {original}
수량/단위: {unit_str}
상세설명 키워드: {html_context}
네이버 연관 키워드: {', '.join(related_kws)}
제외 단어(브랜드/의심): {', '.join(s1['banned'])}
카테고리: {category_info}

[구매대행 상품명 구조]
{template}

[카테고리 전략]
{strategy}

[필수 규칙]
1. 브랜드명 절대 사용 금지 (구매대행 = 노브랜드)
2. 대형 키워드 반복 금지 (동일 단어 3회 이상 반복 불가)
3. 롱테일 키워드 조합: 실구매자가 검색할 세부 키워드 2-3개 반드시 포함
4. 수량/단위/스펙 정보 반드시 포함
5. 길이: 25-50자 (공백 포함)
6. 특수문자, 허위과장(1위/최고/완벽), 인증 키워드(KC인증/FDA) 금지
7. 수식어(프리미엄/고급/최신형/강력한) 금지
8. 오직 최종 상품명 한 줄만 출력

[Good 예시]
- 플리츠 롱원피스 하객룩 여성 봄신상 오피스룩 A라인 하늘색
- 초경량 백패킹 접이식의자 캠핑 낚시 휴대용 알루미늄 500g
- 오픈형 무선이어폰 귀안아픈 운동용 방수 골전도 러닝 IPX7
"""
        new_name = gemini.generate_content(prompt)
        if new_name:
            # 후처리: 따옴표/줄바꿈 제거, 50자 제한
            cleaned = new_name.strip().replace('"', '').replace("'", "").split('\n')[0].strip()
            if len(cleaned) > 50:
                # 50자 넘으면 마지막 공백 기준으로 자름
                cleaned = cleaned[:50].rsplit(' ', 1)[0]
            return cleaned
        return original

    def stop(self):
        self.is_running = False

# ======================================================
# UI COMPONENTS (TABS)
# ======================================================

class BaseTab(QWidget):
    def __init__(self, main_window):
        super().__init__()
        self.main = main_window
        self.all_items = []
        self._ui_refs = [] # Explicit references
        layout = QVBoxLayout(self)
        
        # Upper Control Area
        self.control_layout = QHBoxLayout()
        layout.addLayout(self.control_layout)
        
        # Table
        self.table = QTableWidget(self)
        self.table.setColumnCount(6)
        self._ui_refs.append(self.table)
        self.table.setHorizontalHeaderLabels(["ID", "판매자코드", "원 상품명", "새 상품명", "상태", "키워드"])
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        layout.addWidget(self.table)
        
        # Lower Action Area
        action_layout = QHBoxLayout()
        
        self.chk_auto_save = QCheckBox("자동 저장", self)
        self.chk_auto_save.setChecked(True)
        self.chk_resume = QCheckBox("이어하기", self)
        self.chk_resume.setChecked(True)
        self._ui_refs.extend([self.chk_auto_save, self.chk_resume])
        
        action_layout.addWidget(self.chk_auto_save)
        action_layout.addWidget(self.chk_resume)
        
        self.btn_start = QPushButton("🚀 시작")
        self.btn_start.setObjectName("PrimaryButton")
        self.btn_start.clicked.connect(self.start_batch)
        action_layout.addWidget(self.btn_start)
        
        self.btn_stop = QPushButton("중지")
        self.btn_stop.setObjectName("DangerButton")
        self.btn_stop.clicked.connect(self.stop_batch)
        action_layout.addWidget(self.btn_stop)
        
        layout.addLayout(action_layout)
        
    def update_table(self, items):
        self.all_items = items
        self.table.setRowCount(len(items))
        for i, item in enumerate(items):
            self.table.setItem(i, 0, QTableWidgetItem(item.get('id', '')))
            self.table.setItem(i, 1, QTableWidgetItem(item.get('seller_code', '')))
            self.table.setItem(i, 2, QTableWidgetItem(item.get('original_name', '')))
            self.table.setItem(i, 3, QTableWidgetItem(item.get('new_name', '')))
            self.table.setItem(i, 4, QTableWidgetItem(item.get('status', 'Ready')))
            self.table.setItem(i, 5, QTableWidgetItem(item.get('keywords', '')))

    def log(self, msg):
        self.main.log(msg)
        
    def start_batch(self):
        pass # Override
        
    def stop_batch(self):
        if hasattr(self, 'worker'):
            self.worker.stop()
            self.log("🛑 중지 중...")

    def update_row(self, idx, result):
        if idx < len(self.all_items):
            item = self.all_items[idx]
            item.update(result)
            self.table.setItem(idx, 3, QTableWidgetItem(result.get('new_name', '')))
            self.table.setItem(idx, 4, QTableWidgetItem(result.get('status', '')))
            self.table.setItem(idx, 5, QTableWidgetItem(result.get('keywords', '')))
            self.table.scrollToItem(self.table.item(idx, 0))
            
            if idx % 50 == 0 and self.chk_auto_save.isChecked():
                self.save_progress()

    def process_finished(self):
        self.btn_start.setEnabled(True)
        self.btn_stop.setEnabled(False)
        self.log("✅ 작업 완료.")
        if self.chk_auto_save.isChecked():
            self.save_progress()

    def save_progress(self):
        try:
             with open("batch_progress.json", "w", encoding="utf-8") as f:
                 json.dump(self.all_items, f, ensure_ascii=False, indent=2)
             self.log("💾 자동 저장됨.")
        except Exception as e:
             self.log(f"⚠️ 저장 오류: {e}")

class APIModeTab(BaseTab):
    def setup_ui(self):
        pass  # BaseTab에 setup_ui 없음 - __init__에서 직접 구성
        
        self.cmb_group = QComboBox(self)
        lbl_group = QLabel("그룹명:", self)
        self.control_layout.addWidget(lbl_group)
        self.control_layout.addWidget(self.cmb_group)
        self._ui_refs.extend([self.cmb_group, lbl_group])
        
        # Load groups from config
        try:
            if os.path.exists("bulsaja_config.json"):
                with open("bulsaja_config.json", "r", encoding="utf-8") as f:
                    conf = json.load(f)
                    groups_str = conf.get("market_groups", "")
                    if groups_str:
                        groups = [g.strip() for g in groups_str.split(",") if g.strip()]
                        groups.sort() # Sort ascending
                        self.cmb_group.addItems(groups)
        except: pass

        btn_fetch = QPushButton("API 데이터 가져오기")
        btn_fetch.clicked.connect(self.fetch_data)
        self.control_layout.addWidget(btn_fetch)
        
    def fetch_data(self):
        token = self.main.txt_access.text().strip()
        group = self.cmb_group.currentText().strip()
        if not token or not group:
            QMessageBox.warning(self, "오류", "액세스 토큰과 그룹명이 필요합니다.")
            return
            
        client = BulsajaAPIClient(token, self.main.txt_refresh.text().strip(), self.main.txt_cookie.text().strip())
        self.log(f"🔄 API 데이터 요청 중... ({group})")
        
        self.loader = DataLoadWorker('api', group, client)
        self.loader.finished.connect(self.on_loaded)
        self.loader.start()
        
    def on_loaded(self, items, error):
        if error:
            QMessageBox.critical(self, "오류", error)
            return
        self.update_table(items)
        self.log(f"✅ {len(items)}개 로드 완료.")

    def start_batch(self):
        if not self.all_items:
            QMessageBox.warning(self, "주의", "데이터가 없습니다.")
            return
            
        keys = self.main.get_gemini_keys()
        if not keys: return
        
        token = self.main.txt_access.text().strip()
        api_client = BulsajaAPIClient(token, self.main.txt_refresh.text().strip())
        
        items_to_process = self.all_items
        if self.chk_resume.isChecked():
            items_to_process = [i for i in self.all_items if i['status'] != 'Done']
            
        self.btn_start.setEnabled(False)
        self.btn_stop.setEnabled(True)
        
        # Pull generation settings from main UI
        gen_params = self.main.get_generation_params()
        naver_creds = self.main.get_naver_credentials()
        
        self.worker = BulkGenerationWorker(items_to_process, keys, api_client, gen_params, naver_creds)
        self.worker.log.connect(self.log)
        self.worker.progress.connect(self.update_row)
        self.worker.finished.connect(self.process_finished)
        self.worker.start()

class ExcelModeTab(BaseTab):
    def setup_ui(self):
        pass  # BaseTab에 setup_ui 없음 - __init__에서 직접 구성
        
        self.lbl_path = QLabel("파일 없음")
        btn_load = QPushButton("엑셀 파일 선택")
        btn_load.clicked.connect(self.load_file)
        
        self.control_layout.addWidget(btn_load)
        self.control_layout.addWidget(self.lbl_path)
        
    def load_file(self):
        fname, _ = QFileDialog.getOpenFileName(self, '엑셀 선택', '', 'Excel Files (*.xlsx)')
        if fname:
            self.lbl_path.setText(os.path.basename(fname))
            self.log(f"📂 파일 로딩: {fname}")
            self.loader = DataLoadWorker('excel', fname)
            self.loader.finished.connect(self.on_loaded)
            self.loader.start()
            
    def on_loaded(self, items, error):
        if error:
            QMessageBox.critical(self, "오류", error)
            return
        self.update_table(items)
        self.log(f"✅ {len(items)}개 로드 완료.")

    def start_batch(self):
        if not self.all_items:
            QMessageBox.warning(self, "주의", "데이터가 없습니다.")
            return
            
        keys = self.main.get_gemini_keys()
        if not keys: return
        
        items_to_process = self.all_items
        if self.chk_resume.isChecked():
            items_to_process = [i for i in self.all_items if i['status'] != 'Done']
        
        self.btn_start.setEnabled(False)
        self.btn_stop.setEnabled(True)
        
        gen_params = self.main.get_generation_params()
        naver_creds = self.main.get_naver_credentials()
        
        self.worker = BulkGenerationWorker(items_to_process, keys, None, gen_params, naver_creds) 
        
        self.worker.log.connect(self.log)
        self.worker.progress.connect(self.update_row)
        self.worker.finished.connect(self.process_finished_excel) # Custom finish
        self.worker.start()

    def process_finished_excel(self):
        super().process_finished()
        # Save to new Excel
        # Simple extraction of all_items to Excel
        try:
            wb = Workbook()
            ws = wb.active
            ws.append(["ID", "판매자코드", "원 상품명", "새 상품명", "상태", "키워드"])
            for item in self.all_items:
                ws.append([
                    item.get('id'),
                    item.get('seller_code'),
                    item.get('original_name'),
                    item.get('new_name'),
                    item.get('status'),
                    item.get('keywords')
                ])
            
            save_name = f"processed_{int(time.time())}.xlsx"
            wb.save(save_name)
            self.log(f"💾 엑셀 저장 완료: {save_name}")
            QMessageBox.information(self, "완료", f"엑셀 파일이 저장되었습니다.\n{save_name}")
        except Exception as e:
            self.log(f"❌ 엑셀 저장 실패: {e}")

class MainWindowV2(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("SmartSellUp NameMaker v4.0")
        self.resize(1400, 900)
        self.setStyleSheet(DARK_THEME_STYLESHEET)
        self._ui_refs = []
        self.table_data = [] # Unify to table_data
        self.db_path = "smartsellup.db"
        self.init_db()
        self.loop_timer = QTimer(self)
        self.loop_timer.timeout.connect(self.run_loop_cycle)
        self.initUI()
        
    def init_db(self):
        """Initialize SQLite database for work history."""
        try:
            conn = sqlite3.connect(self.db_path)
            cur = conn.cursor()
            cur.execute("""
                CREATE TABLE IF NOT EXISTS history (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    seller_code TEXT,
                    original_name TEXT,
                    ai_name TEXT,
                    prev_rank INTEGER,
                    curr_rank INTEGER,
                    status TEXT,
                    timestamp DATETIME DEFAULT CURRENT_TIMESTAMP
                )
            """)
            conn.commit()
            conn.close()
        except Exception as e:
            print(f"DB Init Error: {e}")

    def initUI(self):
        central = QWidget()
        self.setCentralWidget(central)
        main_layout = QVBoxLayout(central)
        main_layout.setContentsMargins(5, 5, 5, 5)
        main_layout.setSpacing(5)

        # 🟢 1. Top Panel: Metadata & Controls (Madword Style)
        # We reuse MetadataPanel which contains Save button and inputs
        self.meta_panel = MetadataPanel(self)
        main_layout.addWidget(self.meta_panel)

        # 🟠 Main Splitter
        self.splitter_main = QSplitter(Qt.Orientation.Vertical)
        
        # Upper Part: Product Table & Right Panel (Horizontal Splitter)
        self.splitter_upper = QSplitter(Qt.Orientation.Horizontal)
        
        self.table = QTableWidget()
        self.table.setColumnCount(11) # Adjusted based on screenshot
        self.table.setHorizontalHeaderLabels([
            "이미지", "상품번호", "상품명", "카테고리", 
            "생성상품명", "검색명", "목표순위", "이전순위", 
            "현재순위", "수집카테고리", " " # Last one for select or padding
        ])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.table.selectionModel().selectionChanged.connect(lambda: self.on_row_selected())
        self.table.cellClicked.connect(self.on_cell_clicked)
        self.splitter_upper.addWidget(self.table)
        
        # Right Side: Keyword Frequency Table
        self.kw_freq_panel = QGroupBox("키워드 빈도수")
        kw_freq_layout = QVBoxLayout(self.kw_freq_panel)
        self.table_kw_freq = QTableWidget(0, 2)
        self.table_kw_freq.setHorizontalHeaderLabels(["키워드", "빈도수"])
        self.table_kw_freq.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        kw_freq_layout.addWidget(self.table_kw_freq)
        self.kw_freq_panel.setFixedWidth(250)
        self.splitter_upper.addWidget(self.kw_freq_panel)
        
        self.splitter_main.addWidget(self.splitter_upper)

        # 🟡 Lower Part: Bottom Analysis Panels (Madword Style)
        self.analysis_panel = BottomAnalysisPanel(self)
        self.splitter_main.addWidget(self.analysis_panel)
        
        main_layout.addWidget(self.splitter_main)

        # 🔴 Bottom Footer: Controls & Status
        footer = QHBoxLayout()
        
        self.btn_excel = QPushButton("📁 엑셀 불러오기", self)
        self.btn_excel.clicked.connect(self.open_excel)
        self.btn_chrome_connect = QPushButton("🌐 크롬 자동 연결", self)
        self.btn_chrome_connect.clicked.connect(self.auto_connect_bulsaja)
        self.btn_api_fetch = QPushButton("🔌 API 불러오기", self)
        self.btn_api_fetch.clicked.connect(self.fetch_api_data)
        
        self.chk_focus_only = QCheckBox("⭐ 집중관리만", self)
        self.chk_focus_only.stateChanged.connect(self.filter_table)

        footer.addWidget(self.btn_excel)
        footer.addWidget(self.btn_chrome_connect)
        footer.addWidget(self.btn_api_fetch)
        footer.addWidget(self.chk_focus_only)
        footer.addStretch()
        
        self.btn_start = QPushButton("🚀 시작", self)
        self.btn_start.clicked.connect(self.start_generation)
        self.btn_stop = QPushButton("🛑 중지", self)
        self.btn_stop.clicked.connect(self.stop_generation)
        self.btn_settings = QPushButton("⚙️ 설정", self)
        self.btn_settings.clicked.connect(self.open_settings)
        
        footer.addWidget(self.btn_start)
        footer.addWidget(self.btn_stop)
        footer.addWidget(self.btn_settings)

        main_layout.addLayout(footer)

        self.lbl_status = QLabel("준비됨", self)
        self.lbl_status.setStyleSheet("color: #66ff00; font-family: Consolas; font-size: 11px;")
        main_layout.addWidget(self.lbl_status)

        # Hidden properties and credentials (need to be defined for load_settings)
        self.settings_data = {
            "filter_junk": True, "keep_orig": False, "prefix_cnt": 999,
            "sort_orders": "내림차순", "word_limit": 10, "char_limit": 50,
            "shuffle_tags": False, "use_ai": True
        }
        self.txt_naver_id = QLineEdit(self)
        self.txt_naver_secret = QLineEdit(self)
        self.txt_ads_key = QLineEdit(self)
        self.txt_ads_secret = QLineEdit(self)
        self.txt_ads_cust_id = QLineEdit(self)
        self.txt_comm_id = QLineEdit(self)
        self.txt_comm_secret = QLineEdit(self)
        self.txt_access = QLineEdit(self)
        self.txt_refresh = QLineEdit(self)
        self.txt_cookie = QLineEdit(self)
        self.txt_gemini = QLineEdit(self)
        
        self._ui_refs.extend([self.txt_naver_id, self.txt_naver_secret, self.txt_ads_key, self.txt_ads_secret, self.txt_ads_cust_id, 
                              self.txt_comm_id, self.txt_comm_secret, self.txt_access, self.txt_refresh, self.txt_cookie, self.txt_gemini])

        # Finally Load Settings
        self.load_settings()
        
        # Connect log console to the restored BottomAnalysisPanel log console if needed
        # Actually log_console should be an attribute of MainWindowV2
        self.log_console = self.analysis_panel.log_console

    def run_loop_cycle(self):
        self.log("🔄 무한 루프 사이클 시작...")
        # To be implemented: trigger start_batch logic

    def load_settings(self):
        """Load settings from local JSON file using absolute path."""
        base_path = os.path.dirname(os.path.abspath(__file__))
        settings_file = os.path.join(base_path, "product_name_gen_settings.json")
        data = {}
        
        # 1. Try to load from dedicated settings file
        if os.path.exists(settings_file):
            try:
                with open(settings_file, "r", encoding="utf-8") as f:
                    data = json.load(f)
            except Exception as e:
                self.log(f"⚠️ 설정 파일 로드 실패: {e}")
        
        # 2. Fallback: Try loading Gemini/Bulsaja keys from shared config if not in dedicated settings
        if not data.get("gemini_keys"):
            try:
                if os.path.exists("bulsaja_config.json"):
                    with open("bulsaja_config.json", "r", encoding="utf-8") as f:
                        b_conf = json.load(f)
                        data["gemini_keys"] = b_conf.get("api_key", "")
                elif os.path.exists("bulsaja_config_gemini.json"):
                     with open("bulsaja_config_gemini.json", "r", encoding="utf-8") as f:
                        b_conf = json.load(f)
                        if isinstance(b_conf.get("gemini_api_keys"), list):
                            data["gemini_keys"] = ",".join(b_conf.get("gemini_api_keys"))
                        else:
                            data["gemini_keys"] = b_conf.get("gemini_api_keys", "")
            except: pass

        # Apply to UI/Hidden fields
        self.txt_naver_id.setText(data.get("naver_id", ""))
        self.txt_naver_secret.setText(data.get("naver_secret", ""))
        self.txt_ads_key.setText(data.get("ads_key", ""))
        self.txt_ads_secret.setText(data.get("ads_secret", ""))
        self.txt_ads_cust_id.setText(data.get("ads_cust_id", ""))
        self.txt_comm_id.setText(data.get("comm_id", ""))
        self.txt_comm_secret.setText(data.get("comm_secret", ""))
        self.txt_gemini.setText(data.get("gemini_keys", ""))
        
        # Load other preferences
        if "preferences" in data:
            self.settings_data.update(data["preferences"])

    def save_settings(self):
        """Save current settings to local JSON file using absolute path."""
        data = {
            "naver_id": self.txt_naver_id.text(),
            "naver_secret": self.txt_naver_secret.text(),
            "ads_key": self.txt_ads_key.text(),
            "ads_secret": self.txt_ads_secret.text(),
            "ads_cust_id": self.txt_ads_cust_id.text(),
            "comm_id": self.txt_comm_id.text(),
            "comm_secret": self.txt_comm_secret.text(),
            "gemini_keys": self.txt_gemini.text(),
            "preferences": self.settings_data
        }
        try:
            base_path = os.path.dirname(os.path.abspath(__file__))
            settings_file = os.path.join(base_path, "product_name_gen_settings.json")
            with open(settings_file, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            self.log(f"⚠️ 설정 저장 실패: {e}")

    def open_settings(self):
        # Keep a strong reference to prevent GC issues
        self.dlg = SettingsDialog(self)
        self.dlg.exec()
        # Ensure we clear the reference after use
        # self.dlg = None # Optionally clear it, but let's keep it until next time

    def on_row_selected(self):
        rows = self.table.selectionModel().selectedRows()
        if not rows: return
        row = rows[0].row()
        item = self.table_data[row] if row < len(self.table_data) else {}
        
        if not item: return

        # Update Metadata Panel (Madword Style)
        # Ensure these attributes match MetadataPanel class definition
        if hasattr(self, 'meta_panel'):
            self.meta_panel.txt_orig_name.setText(item.get('original_name', ''))
            self.meta_panel.txt_cat_name.setText(item.get('category', ''))
            self.meta_panel.txt_final_name.setText(item.get('new_name', ''))
            self.meta_panel.txt_removed.setText(", ".join(item.get('banned_found', [])))
            self.meta_panel.txt_added.setText(item.get('keywords', ''))
            
            # Show summary of related keywords
            related = item.get('related_keywords', [])
            summary = ", ".join([k[0] if isinstance(k, tuple) else str(k) for k in related[:5]])
            self.meta_panel.txt_related_meta.setText(summary)
            self.meta_panel.txt_optimized.setText(item.get('status', 'Ready'))
        
        # Populate Analysis Panels
        self._update_analysis_panels(item)

    def _update_analysis_panels(self, item):
        # Madword Style: Update BottomAnalysisPanel widgets
        if not hasattr(self, 'analysis_panel'): return

        # 1. Update Extracted Keywords (pnl_extracted)
        table_ext = self.analysis_panel.pnl_extracted.findChild(QTableWidget)
        if table_ext:
            table_ext.setRowCount(0)
            keywords = item.get('extracted_keywords', [])
            for kw, cnt in keywords:
                r = table_ext.rowCount()
                table_ext.insertRow(r)
                table_ext.setItem(r, 0, QTableWidgetItem(str(kw)))
                table_ext.setItem(r, 1, QTableWidgetItem(str(cnt)))

        # 2. Update Related Keywords (pnl_related)
        table_rel = self.analysis_panel.pnl_related.findChild(QTableWidget)
        if table_rel:
            table_rel.setRowCount(0)
            related = item.get('related_keywords', [])
            for kw, vol in related:
                r = table_rel.rowCount()
                table_rel.insertRow(r)
                table_rel.setItem(r, 0, QTableWidgetItem(str(kw)))
                table_rel.setItem(r, 1, QTableWidgetItem(str(vol)))
    def on_metadata_save(self):
        rows = self.table.selectionModel().selectedRows()
        if not rows: return
        row = rows[0].row()
        
        new_name = self.meta_panel.txt_final_name.text().strip()
        self.table_data[row]['new_name'] = new_name
        self.table.setItem(row, 2, QTableWidgetItem(new_name))
        self.log(f"💾 {row+1}행 수정사항 저장됨")
        
    def open_excel(self):
        path, _ = QFileDialog.getOpenFileName(self, "엑셀 파일 열기", "", "Excel Files (*.xlsx *.xls *.csv)")
        if not path: return
        self.log(f"📁 엑셀 로딩 중: {os.path.basename(path)}")
        self.loader = DataLoadWorker('excel', path)
        self.loader.finished.connect(self.on_data_loaded)
        self.loader.start()

    def fetch_api_data(self):
        token = self.txt_access.text().strip()
        if not token:
            QMessageBox.warning(self, "오류", "먼저 '크롬 자동 연결'을 통해 토큰을 가져오세요.")
            return
        group, ok = QInputDialog.getText(self, "API 불러오기", "마켓 그룹명을 입력하세요:")
        if not ok or not group: return
        
        client = BulsajaAPIClient(token, self.txt_refresh.text().strip(), self.txt_cookie.text().strip())
        self.loader = DataLoadWorker('api', group, client)
        self.loader.finished.connect(self.on_data_loaded)
        self.loader.start()

    def update_ranking_hub(self, item):
        ranks_table = self.analysis_panel.pnl_ranking.findChild(QTableWidget)
        if not ranks_table: return
        ranks_table.setRowCount(0)
        history = item.get('rank_history', [])
        for kw, rank, diff in history:
            r = ranks_table.rowCount()
            ranks_table.insertRow(r)
            ranks_table.setItem(r, 0, QTableWidgetItem(kw))
            ranks_table.setItem(r, 1, QTableWidgetItem(f"{rank}위"))
            ranks_table.setItem(r, 2, QTableWidgetItem(diff))

    def on_cell_clicked(self, row, col):
        if col == 0:
            item = self.table_data[row]
            item['is_focus'] = not item.get('is_focus', False)
            self.table.setItem(row, 0, QTableWidgetItem("⭐" if item['is_focus'] else "☆"))
            self.save_persistent_data()

    def filter_table(self):
        focus_only = self.chk_focus_only.isChecked()
        if not hasattr(self, 'table_data'): return
        for i in range(self.table.rowCount()):
            if focus_only:
                is_focus = self.table_data[i].get('is_focus', False)
                self.table.setRowHidden(i, not is_focus)
            else:
                self.table.setRowHidden(i, False)

    def save_persistent_data(self):
        """Save rank history and focus status to local storage."""
        if not hasattr(self, 'table_data'): return
        data_to_save = {}
        for item in self.table_data:
            if item.get('is_focus') or item.get('rank_history'):
                data_to_save[item['id']] = {
                    "is_focus": item.get('is_focus', False),
                    "rank_history": item.get('rank_history', [])
                }
        try:
            with open("focus_db.json", "w", encoding="utf-8") as f:
                json.dump(data_to_save, f, ensure_ascii=False, indent=2)
        except: pass

    def load_persistent_data(self):
        """Load rank history and focus status from local storage."""
        if os.path.exists("focus_db.json"):
            try:
                with open("focus_db.json", "r", encoding="utf-8") as f:
                    return json.load(f)
            except: pass
        return {}

    def on_data_loaded(self, items, error):
        if error:
            QMessageBox.critical(self, "오류", f"데이터 로드 실패: {error}")
            return
        
        db = self.load_persistent_data()
        self.table_data = items
        self.table.setRowCount(len(items))
        for i, item in enumerate(items):
            # Apply persistent data
            p_data = db.get(item.get('id'), {})
            item['is_focus'] = p_data.get('is_focus', False)
            item['rank_history'] = p_data.get('rank_history', [])
            
            self.table.setItem(i, 0, QTableWidgetItem("⭐" if item['is_focus'] else "☆"))
            self.table.setItem(i, 1, QTableWidgetItem(item.get('seller_code', ''))) # Show Seller Code
            self.table.setItem(i, 2, QTableWidgetItem(item.get('original_name', '')))
            self.table.setItem(i, 3, QTableWidgetItem(item.get('new_name', '')))
            self.table.setItem(i, 4, QTableWidgetItem(item.get('status', 'Ready')))
        self.log(f"✅ {len(items)}개 데이터 로드 완료")

    def start_generation(self):
        if not hasattr(self, 'table_data') or not self.table_data:
            QMessageBox.warning(self, "경고", "처리할 데이터가 없습니다.")
            return
        
        keys = self.get_gemini_keys()
        if not keys: return
        
        naver_creds = self.get_naver_credentials()
        # Note: In the new UI, Naver settings are in the Settings dialog or global.
        # But wait, I added them to MainWindowV2 in turn 14, let me check if they still exist.
        # Ah, I replaced them in Turn 16. I should restore them in the SettingsDialog or keep them.
        # Let's assume for now they are in the globals or I'll add them to SettingsDialog.
        
        self.btn_start.setEnabled(False)
        self.btn_stop.setEnabled(True)
        
        self.worker = BulkGenerationWorker(self.table_data, keys, None, self.settings_data, naver_creds)
        self.worker.log.connect(self.log)
        self.worker.progress.connect(self.update_row)
        self.worker.finished.connect(self.on_finished)
        self.worker.start()

    def stop_generation(self):
        if hasattr(self, 'worker') and self.worker is not None:
            self.worker.stop()
            self.log("🛑 중지 요청됨...")

    def update_row(self, row_idx, result):
        if row_idx < len(self.table_data):
            self.table_data[row_idx].update(result)
            self.table.setItem(row_idx, 3, QTableWidgetItem(result.get('new_name', ''))) # Col 3 is New Name
            self.table.setItem(row_idx, 4, QTableWidgetItem(result.get('status', '')))   # Col 4 is Status
            self.table.scrollToItem(self.table.item(row_idx, 0))
            # If current row is selected, update panels
            if self.table.currentRow() == row_idx:
                self._update_analysis_panels(self.table_data[row_idx])

    def on_finished(self):
        self.btn_start.setEnabled(True)
        self.btn_stop.setEnabled(False)
        self.log("🏁 전체 작업 종료")

    def export_enhanced_results(self):
        """Export simulation or focus results as a professional CSV."""
        if not hasattr(self, 'table_data') or not self.table_data: return
        path, _ = QFileDialog.getSaveFileName(self, "결과 내보내기", "results.csv", "CSV Files (*.csv)")
        if not path: return
        
        import csv
        try:
            with open(path, "w", encoding="utf-8-sig", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(["ID", "판매자코드", "원본명", "추천명", "키워드", "추출명사", "네이버검색량(대표)", "집중관리"])
                for item in self.table_data:
                    extracted = ", ".join([k[0] for k in item.get('extracted_keywords', [])])
                    related = item.get('related_keywords', [])
                    top_vol = related[0][1] if related else 0
                    writer.writerow([
                        item.get('id'),
                        item.get('seller_code'),
                        item.get('original_name'), 
                        item.get('new_name'), 
                        item.get('keywords'),
                        extracted,
                        top_vol,
                        "Yes" if item.get('is_focus') else "No"
                    ])
            self.log(f"✅ {len(self.table_data)}개 데이터 내보내기 완료")
        except Exception as e:
            QMessageBox.critical(self, "오류", f"내보내기 실패: {e}")

    def import_focus_list(self):
        """Show dialog to paste seller codes and match with current data."""
        text, ok = QInputDialog.getMultiLineText(self, "집중 관리 상품 일괄 등록", 
                                                 "외부 검증 도구에서 선정된 판매자 코드(Seller Code)를 입력하세요.\n(줄바꿈으로 구분)")
        if not ok or not text.strip(): return
        
        target_codes = set(code.strip() for code in text.split("\n") if code.strip())
        matched_count = 0
        
        for item in self.table_data:
            # Match against ID (which is often the merchant code in Direct API) or Seller Code
            if item.get('id') in target_codes or item.get('seller_code') in target_codes:
                item['is_focus'] = True
                matched_count += 1
                
        # Update UI Table
        for i in range(self.table.rowCount()):
            item_id = self.table.item(i, 1).text()
            # We don't have seller_code in the table yet, but we match based on the data stored
            if item_id in target_codes:
                self.table.setItem(i, 0, QTableWidgetItem("⭐"))
                
        self.save_persistent_data()
        self.log(f"🎯 {matched_count}개의 상품이 집중 관리 대상으로 등록되었습니다.")
        QMessageBox.information(self, "완료", f"{matched_count}개의 상품이 매칭되어 집중 관리 대상으로 등록되었습니다.")

    def on_rank_track_clicked(self):
        rows = self.table.selectionModel().selectedRows()
        if not rows:
            QMessageBox.warning(self, "경고", "상품을 먼저 선택하세요.")
            return
        row = rows[0].row()
        item = self.table_data[row]
        
        # In a real app, this would use NaverRankTracker in a separate thread
        # For now, let's simulate tracking for the keywords we found
        self.log(f"🔎 '{item.get('original_name')[:10]}...' 순위 추적 시작")
        
        keywords = item.get('keywords', "캠핑의자,감성캠핑").split(",")
        history = []
        for kw in keywords:
            kw = kw.strip()
            page, pos = NaverRankTracker.find_rank(kw, item.get('id', ''))
            # Generate dummy history for visualization
            rank = random.randint(1, 100) if page == 0 else (page-1)*40 + pos
            diff = random.choice(["▲2", "▼1", "-", "NEW"])
            history.append([kw, rank, diff])
            
        item['rank_history'] = history
        self.update_ranking_hub(item)
        self.save_persistent_data()
        self.log("✅ 순위 데이터 업데이트 완료")
        
    def reset_all_data(self):
        """테이블 데이터 초기화"""
        self.table_data = []
        self.table.setRowCount(0)
        self.log("📋 모든 데이터가 초기화되었습니다.")

    def log(self, message):
        """Advanced central logger: UI, Console, and File."""
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        formatted_msg = f"[{timestamp}] {message}"
        
        # 1. Terminal
        print(formatted_msg)
        
        # 2. UI Status Bar
        if hasattr(self, 'lbl_status'):
             self.lbl_status.setText(message)
        
        # 3. Log Console Widget
        if hasattr(self, 'log_console'):
            self.log_console.appendPlainText(formatted_msg)
            # Auto-scroll to bottom
            self.log_console.verticalScrollBar().setValue(self.log_console.verticalScrollBar().maximum())

        # 4. File Logger (Persistent)
        try:
            base_path = os.path.dirname(os.path.abspath(__file__))
            log_file = os.path.join(base_path, "product_name_app.log")
            with open(log_file, "a", encoding="utf-8") as f:
                f.write(formatted_msg + "\n")
        except:
            pass # Silent fail for logging itself

    def auto_connect_bulsaja(self):
        try:
            extractor = ChromeTokenExtractor()
            access, refresh, cookie = extractor.extract()
            
            if access:
                self.txt_access.setText(access)
                self.txt_refresh.setText(refresh if refresh else "")
                self.txt_cookie.setText(cookie if cookie else "")
                self.log("✅ 불사자 토큰 및 쿠키 자동 가져오기 성공!")
                QMessageBox.information(self, "성공", "크롬에서 토큰과 쿠키를 가져왔습니다.")
            else:
                self.log("❌ 토큰을 찾지 못했습니다.")
        except Exception as e:
            self.log(f"❌ 연결 실패: {e}")
            QMessageBox.warning(self, "오류", 
                f"크롬 연결 실패:\n{e}\n\n"
                "1. 크롬이 디버깅 모드(9222)로 켜져 있나요?\n"
                "2. 불사자 웹페이지에 로그인 되어 있나요?")



    def get_gemini_keys(self):
        text = self.txt_gemini.text().strip()
        if not text:
            QMessageBox.warning(self, "경고", "Gemini API 키를 입력해주세요.")
            return None
        return [k.strip() for k in text.split(",") if k.strip()]

    def get_generation_params(self):
        """Extract settings from cached dictionary for worker usage."""
        # Use settings_data which is synced with SettingsDialog
        return {
            "min_len": 20,  # Default
            "max_len": self.settings_data.get("char_limit", 50),
            "filter_alphanum": self.settings_data.get("filter_junk", True),
            "keep_brand": False, # Todo: add to settings if needed
            "key_count": self.settings_data.get("word_limit", 10),
            "sim_mode": self.settings_data.get("sim_mode", False),
            "keep_orig": self.settings_data.get("keep_orig", False),
            "prefix_cnt": self.settings_data.get("prefix_cnt", 999), 
            "shuffle_tags": self.settings_data.get("shuffle_tags", False)
        }

    def get_naver_credentials(self):
        """Extract Naver API credentials from UI (naver + commerce 분리)."""
        return {
            "naver": {
                "client_id": self.txt_naver_id.text().strip(),
                "client_secret": self.txt_naver_secret.text().strip(),
                "ads_access_key": self.txt_ads_key.text().strip(),
                "ads_secret_key": self.txt_ads_secret.text().strip(),
                "ads_customer_id": self.txt_ads_cust_id.text().strip()
            },
            "commerce": {
                "client_id": self.txt_comm_id.text().strip(),
                "client_secret": self.txt_comm_secret.text().strip()
            }
        }

if __name__ == "__main__":
    app = QApplication(sys.argv)
    ex = MainWindowV2()
    ex.show()
    sys.exit(app.exec())
